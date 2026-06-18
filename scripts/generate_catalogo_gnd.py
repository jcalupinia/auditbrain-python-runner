"""Genera backend/app/ict/catalogo_gnd.py desde el archivo del cliente.

Fuente: gastos_no_deducibles_CMGND_final.xlsx (hoja "gastos no deducibles").
  Col A = No. Casillero (7xxx)
  Col B = Descripción breve del gasto no deducible
  Col C = Normativa aplicable (puede tener saltos de línea)

El módulo resultante mapea casillero -> (descripción, normativa) y lo consume
el A5 Cuadro A para autocompletar las columnas E (descripción) y G (normativa)
cuando traslada un casillero no deducible declarado en el F-101.

Uso:
    python scripts/generate_catalogo_gnd.py [ruta_xlsx]
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_SRC = Path(r"C:\Users\jcalu\Downloads\gastos_no_deducibles_CMGND_final.xlsx")
OUT = Path(__file__).resolve().parent.parent / "backend" / "app" / "ict" / "catalogo_gnd.py"


def main() -> int:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    if not src.exists():
        print(f"ERROR: no existe {src}")
        return 1

    wb = load_workbook(src, data_only=True)
    ws = wb["gastos no deducibles"]
    rows: list[tuple[str, str, str]] = []
    for r in range(4, ws.max_row + 1):
        cas = ws.cell(r, 1).value
        if cas in (None, ""):
            continue
        desc = (ws.cell(r, 2).value or "").strip()
        norm = (ws.cell(r, 3).value or "").strip()
        rows.append((str(cas).strip(), desc, norm))
    rows.sort(key=lambda x: int(x[0]))

    out: list[str] = []
    out.append('"""Catálogo de Gastos No Deducibles (GND) — descripción y normativa SRI.')
    out.append("")
    out.append("GENERADO automáticamente desde el archivo del cliente")
    out.append('`gastos_no_deducibles_CMGND_final.xlsx` (hoja "gastos no deducibles").')
    out.append("NO editar a mano: regenerar con scripts/generate_catalogo_gnd.py.")
    out.append("")
    out.append('Mapea cada casillero de gasto no deducible (rango 7001-7999, "VALOR NO')
    out.append('DEDUCIBLE" en el F-101) a una tupla (descripción, normativa). Lo usa el')
    out.append("A5 Cuadro A para autocompletar las columnas E (descripción del tipo de")
    out.append("gasto) y G (normativa aplicable) cuando traslada un casillero declarado.")
    out.append("")
    out.append(f"Cobertura: {len(rows)} casilleros (100% de los 7xxx no deducibles del F-101).")
    out.append('"""')
    out.append("from __future__ import annotations")
    out.append("")
    out.append("")
    out.append("# casillero -> (descripcion_tipo_gasto, normativa_aplicable)")
    out.append("GND_CASILLERO_INFO: dict[str, tuple[str, str]] = {")
    for cas, desc, norm in rows:
        out.append(f"    {cas!r}: ({desc!r}, {norm!r}),")
    out.append("}")
    out.append("")
    out.append("")
    out.append("def gnd_descripcion(casillero: str) -> str | None:")
    out.append('    """Descripción del tipo de gasto no deducible, o None si no está."""')
    out.append("    info = GND_CASILLERO_INFO.get(str(casillero).strip())")
    out.append("    return info[0] if info else None")
    out.append("")
    out.append("")
    out.append("def gnd_normativa(casillero: str) -> str | None:")
    out.append('    """Normativa aplicable al gasto no deducible, o None si no está."""')
    out.append("    info = GND_CASILLERO_INFO.get(str(casillero).strip())")
    out.append("    return info[1] if info else None")
    out.append("")

    OUT.write_text("\n".join(out), encoding="utf-8")
    print(f"Generado {OUT} con {len(rows)} casilleros")
    return 0


if __name__ == "__main__":
    sys.exit(main())
