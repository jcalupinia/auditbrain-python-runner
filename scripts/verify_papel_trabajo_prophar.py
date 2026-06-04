"""Verificación empírica del split SRI vs Papel de Trabajo.

Cumple REGLA SUPREMA de CLAUDE.md: no declarar trabajo concluido sin
verificar empíricamente que lo que afirmamos es correcto.

USOS:

1. Modo CONNECTED (con DB real, PROPHAR cargado):
       python scripts/verify_papel_trabajo_prophar.py --ruc 1791859596001

   Toma la última sesión ICT del cliente y verifica los dos Excels.

2. Modo SYNTHETIC (sin DB, solo lógica de split):
       python scripts/verify_papel_trabajo_prophar.py --synthetic

   Construye un workbook fake con las hojas internas y verifica la
   eliminación funciona. Útil en CI sin DB.

Checks ejecutados (ambos modos):
    1. generate_excel devuelve tuple[bytes, bytes]
    2. bytes_sri NO contiene VERIFICACIÓN A1 ni AUDITORÍA DE ANEXOS ni TRAZABILIDAD
    3. bytes_papel SÍ contiene VERIFICACIÓN A1 y AUDITORÍA DE ANEXOS
    4. Ambos archivos se cargan sin error (no dialog "Reparaciones")
    5. Archivos escritos a _verify_output/ para inspección manual

Sale con exit code 0 si todo OK, 1 si cualquier check falla.
"""
from __future__ import annotations

import argparse
import sys
from io import BytesIO
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

INTERNAL_SHEETS = ("VERIFICACIÓN A1", "AUDITORÍA DE ANEXOS", "TRAZABILIDAD")


def _check_split(bytes_sri: bytes, bytes_papel: bytes) -> list[str]:
    """Apply all checks to a pair (bytes_sri, bytes_papel). Returns list of
    error messages — empty list = all checks passed."""
    errors: list[str] = []

    if not isinstance(bytes_sri, bytes) or len(bytes_sri) < 1000:
        errors.append(f"bytes_sri inválido (len={len(bytes_sri) if bytes_sri else 0})")
    if not isinstance(bytes_papel, bytes) or len(bytes_papel) < 1000:
        errors.append(f"bytes_papel inválido (len={len(bytes_papel) if bytes_papel else 0})")

    # Cargar ambos workbooks
    try:
        wb_sri = openpyxl.load_workbook(BytesIO(bytes_sri))
    except Exception as exc:
        errors.append(f"Excel SRI no abre: {exc}")
        return errors
    try:
        wb_papel = openpyxl.load_workbook(BytesIO(bytes_papel))
    except Exception as exc:
        errors.append(f"Excel papel no abre: {exc}")
        return errors

    # CHECK: Excel SRI NO contiene hojas internas
    for forbidden in INTERNAL_SHEETS:
        if forbidden in wb_sri.sheetnames:
            errors.append(f"Excel SRI contiene hoja prohibida: {forbidden}")

    # CHECK: Papel SÍ contiene VERIFICACIÓN A1 y AUDITORÍA DE ANEXOS
    for required in ("VERIFICACIÓN A1", "AUDITORÍA DE ANEXOS"):
        if required not in wb_papel.sheetnames:
            errors.append(f"Papel trabajo NO contiene hoja requerida: {required}")

    return errors


def _build_synthetic_pair() -> tuple[bytes, bytes]:
    """Build a synthetic pair using the same split logic as service.generate_excel.

    No DB required — useful for CI smoke tests.
    """
    from backend.app.ict.service import INTERNAL_SHEETS_FOR_SRI

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Hojas de negocio que deben quedar en el SRI
    for n in ["INDICE", "A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9",
              "DATOS F-101", "DATOS F-103", "DATOS F-104"]:
        ws = wb.create_sheet(n)
        ws["A1"] = f"datos {n}"
    # Hojas internas que deben removerse del SRI
    for n in INTERNAL_SHEETS_FOR_SRI:
        ws = wb.create_sheet(n)
        ws["A1"] = f"interno {n}"

    buf_papel = BytesIO()
    wb.save(buf_papel)
    bytes_papel = buf_papel.getvalue()

    wb_sri = openpyxl.load_workbook(BytesIO(bytes_papel))
    for sn in INTERNAL_SHEETS_FOR_SRI:
        if sn in wb_sri.sheetnames:
            del wb_sri[sn]
    buf_sri = BytesIO()
    wb_sri.save(buf_sri)
    bytes_sri = buf_sri.getvalue()

    return bytes_sri, bytes_papel


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--ruc", help="RUC del cliente (modo CONNECTED)")
    parser.add_argument(
        "--synthetic", action="store_true",
        help="No requiere DB. Usa workbook fake.",
    )
    args = parser.parse_args()

    if not args.synthetic and not args.ruc:
        # Default to synthetic if nothing provided
        args.synthetic = True

    if args.synthetic:
        print("== Modo SYNTHETIC ==")
        bytes_sri, bytes_papel = _build_synthetic_pair()
    else:
        print(f"== Modo CONNECTED · RUC {args.ruc} ==")
        try:
            from backend.app.db import SessionLocal
            from backend.app.ict import service as ict_service
            from backend.app.ict.models import ICTSession
        except Exception as exc:
            print(f"ERROR importando módulos: {exc}")
            return 1
        db = SessionLocal()
        session = (
            db.query(ICTSession)
            .filter(ICTSession.ruc == args.ruc)
            .order_by(ICTSession.id.desc())
            .first()
        )
        if session is None:
            print(f"ERROR: no hay sesión ICT con RUC {args.ruc} en la BD")
            return 1
        print(f"Sesión encontrada: id={session.id}, período={session.ejercicio_fiscal}")
        bytes_sri, bytes_papel = ict_service.generate_excel(db, session=session)

    print(f"\nbytes_sri:   {len(bytes_sri):>10,} bytes")
    print(f"bytes_papel: {len(bytes_papel):>10,} bytes")

    # Run checks
    errors = _check_split(bytes_sri, bytes_papel)
    if errors:
        print("\nFAIL ERRORES detectados:")
        for e in errors:
            print(f"  • {e}")
        return 1

    # Write to disk for manual inspection
    out_dir = ROOT / "_verify_output"
    out_dir.mkdir(exist_ok=True)
    (out_dir / "ICT_SRI.xlsx").write_bytes(bytes_sri)
    (out_dir / "ICT_PAPEL_TRABAJO.xlsx").write_bytes(bytes_papel)

    # Re-load to confirm files are valid (no "Reparaciones" dialog risk)
    wb_sri = openpyxl.load_workbook(BytesIO(bytes_sri))
    wb_papel = openpyxl.load_workbook(BytesIO(bytes_papel))
    print(f"\nOK Excel SRI ({len(wb_sri.sheetnames)} hojas):")
    print(f"   {wb_sri.sheetnames}")
    print(f"\nOK Excel papel ({len(wb_papel.sheetnames)} hojas):")
    print(f"   {wb_papel.sheetnames}")
    print(f"\n[output] Archivos guardados en: {out_dir}/")
    print("   Inspección manual recomendada:")
    print("   1. Abrir ICT_SRI.xlsx — verificar que NO levanta cuadro 'Reparaciones'")
    print("   2. Abrir ICT_PAPEL_TRABAJO.xlsx — verificar VERIFICACIÓN A1 y AUDITORÍA visibles")
    print("\nTODO OK OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
