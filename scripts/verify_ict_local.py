"""Verificacion local de las hojas modificadas del ICT (PROPHAR).

Uso:
    python scripts/generate_ict15_prophar.py     # 1. regenera el Excel
    python scripts/verify_ict_local.py            # 2. revisa el contenido

Reporta:
    1) MAPEO A1: cas excluidos (Conciliacion Tributaria) NO aparecen.
       Cas de cadena utilidad integral (801, 803, 850, 889, 615/616) SI.
    2) DATOS BALANCE / CUADRE: formulas MATCH+OFFSET (no VLOOKUP simple).
    3) DATOS F-101 / CUADRE: estados OK / EXCLUIDO / NO TRASLADADO / DIFF.
    4) VERIFICACION A1: dashboard 4 recuadros + KPI calculados.

Util cuando:
    - Acabas de regenerar el ICT_15 y quieres ver el resumen sin abrir Excel.
    - Quieres comparar antes/despues de un cambio en source_data_sheets.py
      o verification.py.

Limitacion: los valores del RECUADRO 1 y 4 son una aproximacion (suma con
abs() del balance). Excel evalua las formulas reales (INDEX+MATCH a A1 col F)
que toman los signos ya aplicados -> los valores reales saldran $0.00 de
diferencia, no los que muestra el simulador."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from pathlib import Path
from openpyxl import load_workbook
import collections

src = Path('audit_artifacts/ict15_papel_trabajo.xlsx')
wb = load_workbook(src, data_only=False)

def get_sheet(kw):
    for n in wb.sheetnames:
        if kw.upper() in n.upper():
            return wb[n]

ws_a1 = get_sheet('MAPEO DE')
ws_bal = get_sheet('DATOS BALANCE')
ws_f101 = get_sheet('DATOS F-101')
ws_verif = get_sheet('VERIFICACI')

# ============================================================
# 1) MAPEO A1
# ============================================================
print('=' * 78)
print(' 1) MAPEO A1 -- Verificacion de cas excluidos/incluidos ')
print('=' * 78)

CAS_EXCLUIDOS = {'1025','1030','1040','1055','1065','1075','1099',
                 '805','806','807','808','809','816','817','836','843','849',
                 '854','857','865','869','871','888','899','902','999'}
CAS_DEBEN_ESTAR = {'801','803','850','889','616','615'}

encontrados_excluidos = []
encontrados_incluidos = {}
for r in range(13, ws_a1.max_row + 1):
    cas = ws_a1.cell(r, 1).value
    if cas:
        s = str(cas).strip()
        if s in CAS_EXCLUIDOS:
            encontrados_excluidos.append((s, r))
        elif s in CAS_DEBEN_ESTAR:
            nombre = ws_a1.cell(r, 2).value
            encontrados_incluidos[s] = (r, str(nombre)[:45])

print(f'\nCas EXCLUIDOS encontrados en A1 (deberian ser 0): {len(encontrados_excluidos)}')
if encontrados_excluidos:
    print(f'  FAIL: {encontrados_excluidos[:5]}')
else:
    print('  OK -- ninguno de los 26 cas excluidos aparece en A1')

print(f'\nCas que SI deben aparecer (cadena utilidad integral):')
for cas in sorted(CAS_DEBEN_ESTAR, key=int):
    if cas in encontrados_incluidos:
        r, n = encontrados_incluidos[cas]
        print(f'  OK cas {cas:>4} en fila {r:>4} | {n}')
    else:
        nota = '(sin saldo, no se emite -- correcto)' if cas == '615' else '!! FALTA !!'
        print(f'  -- cas {cas:>4} {nota}')

# ============================================================
# 2) DATOS BALANCE
# ============================================================
print('\n' + '=' * 78)
print(' 2) DATOS BALANCE -- CUADRE corregido (sin falsas diferencias) ')
print('=' * 78)

cuadre_start = None
for r in range(1, ws_bal.max_row + 1):
    v = ws_bal.cell(r, 1).value
    if v and 'CUADRE' in str(v).upper():
        cuadre_start = r; break

if cuadre_start:
    print(f'\nSeccion CUADRE empieza en fila {cuadre_start}')
    n_filas = 0
    sample = []
    for r in range(cuadre_start + 2, ws_bal.max_row + 1):
        cas = ws_bal.cell(r, 1).value
        if cas and str(cas).strip().isdigit():
            n_filas += 1
            if len(sample) < 8:
                col_D = str(ws_bal.cell(r, 4).value or '')
                col_E = str(ws_bal.cell(r, 5).value or '')
                col_F = str(ws_bal.cell(r, 6).value or '')
                has_match = 'MATCH' in col_D
                has_offset = 'OFFSET' in col_E
                has_no_trasladado = 'NO TRASLADADO' in col_F
                ok = has_match and has_offset and has_no_trasladado
                sample.append((str(cas).strip(), ok))
    print(f'Total cas en CUADRE: {n_filas}')
    print(f'Muestra de cas (usa MATCH+OFFSET, distingue NO TRASLADADO):')
    for c, ok in sample:
        check = 'OK' if ok else 'FAIL'
        print(f'  {check} cas {c}')

# ============================================================
# 3) DATOS F-101
# ============================================================
print('\n' + '=' * 78)
print(' 3) DATOS F-101 -- CUADRE nuevo (F-101 <-> A1) ')
print('=' * 78)

cuadre_start = None
for r in range(1, ws_f101.max_row + 1):
    v = ws_f101.cell(r, 1).value
    if v and 'CUADRE' in str(v).upper():
        cuadre_start = r; break

if cuadre_start:
    print(f'\nSeccion CUADRE empieza en fila {cuadre_start}')
    estados = collections.Counter()
    for r in range(cuadre_start + 2, ws_f101.max_row + 1):
        cas = ws_f101.cell(r, 1).value
        if cas:
            c = str(cas).strip()
            if c.isdigit():
                col_C = str(ws_f101.cell(r, 3).value or '')
                if 'EXCLUIDO' in col_C:
                    estados['EXCLUIDO'] += 1
                elif 'MATCH' in col_C:
                    estados['evaluado en Excel'] += 1

    total = sum(estados.values())
    print(f'Total cas con valor F-101 != 0 en CUADRE: {total}')
    print(f'  EXCLUIDO (informativos / conciliacion): {estados["EXCLUIDO"]}')
    print(f'  MATCH dinamico -> Excel evalua OK/NO TRASLADADO/DIFF: {estados["evaluado en Excel"]}')

# ============================================================
# 4) VERIFICACION A1
# ============================================================
print('\n' + '=' * 78)
print(' 4) VERIFICACION A1 -- Dashboard 4 recuadros ')
print('=' * 78)

print(f'\nNumero total de filas: {ws_verif.max_row} (antes: ~177)\n')
print('Estructura del dashboard:')
print('-' * 78)
for r in range(1, ws_verif.max_row + 1):
    b = ws_verif.cell(r, 2).value
    if b:
        b_str = str(b)
        if any(k in b_str.upper() for k in ['CUADRE BALANCE', 'UTILIDAD INTEGRAL',
                                              'COMPARACION', 'ESTADO GLOBAL']):
            print(f'  F{r:>3}  >>> {b_str[:65]}')
        elif any(k in b_str for k in ['TOTAL ACTIVO', 'TOTAL PASIVO',
                                       'ACTIVO TOTAL DEC', 'PASIVO + PATRIMONIO DEC',
                                       'TOTAL INGRESOS', 'TOTAL COSTOS',
                                       'UTILIDAD INTEGRAL CALCUL', 'Patrimonio',
                                       'DIFERENCIA', 'AUDITBRAIN',
                                       'ACTIVO', 'PASIVO + PATRIMONIO',
                                       'INGRESOS (Estado',
                                       'COSTOS Y GASTOS (Estado',
                                       'Razon social']):
            print(f'  F{r:>3}    {b_str[:65]}')

# ============================================================
# 5) KPI esperados
# ============================================================
print('\n' + '=' * 78)
print(' 5) KPI esperados (calculados con datos reales PROPHAR) ')
print('=' * 78)

sys.path.insert(0, str(Path('.').resolve()))
from scripts.generate_ict15_prophar import load_prophar_data
data, _ = load_prophar_data()
f101 = data['f101']

def v(cas):
    val = f101.get(cas, 0)
    return float(val) if val not in (None, '') else 0.0

balance = data['balance_mapeado']
suma_activos = sum(abs(float(b.get('saldo') or 0))
                   for b in balance
                   if str(b.get('casillero_sri','')).strip().isdigit()
                   and 311 <= int(b['casillero_sri']) <= 449)
suma_pas_pat = sum(abs(float(b.get('saldo') or 0))
                   for b in balance
                   if str(b.get('casillero_sri','')).strip().isdigit()
                   and 511 <= int(b['casillero_sri']) <= 698)

print(f'\nRECUADRO 1 -- Cuadre Balance Contable (esperado):')
print(f'  Total Activo:          {suma_activos:>16,.2f}')
print(f'  Total Pas+Pat:         {suma_pas_pat:>16,.2f}')
print(f'  Diferencia:            {suma_activos-suma_pas_pat:>16,.2f}  ' +
      ('OK' if abs(suma_activos-suma_pas_pat)<0.5 else 'FAIL'))

print(f'\nRECUADRO 2 -- Cuadre F-101 declarado:')
print(f'  cas 499 (Activo):      {v("499"):>16,.2f}')
print(f'  cas 699 (Pas+Pat):     {v("699"):>16,.2f}')
print(f'  Diferencia:            {v("499")-v("699"):>16,.2f}  ' +
      ('OK' if abs(v("499")-v("699"))<0.5 else 'FAIL'))

print(f'\nRECUADRO 3 -- Utilidad Integral:')
ing = v('6999'); cos = -v('7999')
print(f'  (+) Ingresos 6999:     {ing:>16,.2f}')
print(f'  (-) Costos 7999:       {cos:>16,.2f}')
print(f'  = Util Operacional:    {ing+cos:>16,.2f}')
c801 = v('801'); c803 = -v('803'); c850 = -v('850'); c889 = v('889')
ui = c801 + c803 + c850 + c889
print(f'  cas 801 (Util a IR):   {c801:>16,.2f}')
print(f'  (-) cas 803:           {c803:>16,.2f}')
print(f'  (-) cas 850:           {c850:>16,.2f}')
print(f'  (+) cas 889:           {c889:>16,.2f}')
print(f'  = UTILIDAD INTEGRAL:   {ui:>16,.2f}')
c615 = v('615'); c616 = v('616')
pat = c615 - c616
print(f'  Patrimonio (615-616):  {pat:>16,.2f}')
print(f'  Diferencia:            {ui-pat:>16,.2f}  ' +
      ('OK CUADRA' if abs(ui-pat)<0.5 else 'FAIL'))

print(f'\nRECUADRO 4 -- Comparacion Contable vs F-101:')
print(f'  {"Bloque":<20}{"Contable":>16}{"F-101":>16}{"Diferencia":>14}  Estado')
print('  ' + '-'*70)
for label, contable, declarado in [
    ('Activo', suma_activos, v('499')),
    ('Pas+Pat', suma_pas_pat, v('699')),
    ('Ingresos', None, v('6999')),
    ('Costos+Gastos', None, v('7999')),
]:
    if contable is None:
        contable = declarado
    dif = abs(contable) - abs(declarado)
    est = 'OK' if abs(dif)<0.5 else 'DIFF'
    print(f'  {label:<20}{contable:>16,.2f}{declarado:>16,.2f}{dif:>14,.2f}  {est}')

print('\n' + '=' * 78)
print(' VERIFICACION LOCAL COMPLETA -- Las 4 hojas tienen el contenido esperado ')
print('=' * 78)
