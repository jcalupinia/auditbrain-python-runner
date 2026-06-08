"""Fase A — Extracción del Golden Master ICT_14.

Lee `ICT_14_PAPEL_TRABAJO.xlsx` (corregido por el cliente) y produce:

  - audit_artifacts/golden_ict14.json         (estructurado por hoja)
  - audit_artifacts/golden_ict14_redcells.csv (subset filtrable)
  - audit_artifacts/golden_ict14_inventory.md (conteos por hoja)
  - audit_artifacts/ict14_source.sha256       (hash de control)

Diseño: lectura NO destructiva (openpyxl data_only=False, read_only=False
para preservar info de fill/font).

Aplica skill `finance:reconciliation` (lado A = golden master).
"""
from __future__ import annotations

import csv
import hashlib
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\jcalu\Downloads\ICT_14_PAPEL_TRABAJO.xlsx")
OUT_DIR = Path("audit_artifacts")

# Variantes de rojo que el cliente puede haber usado para marcar atención.
# Pueden estar en fill (fondo) o en font (color de texto).
RED_FILLS = {
    "FFFF0000",  # rojo puro
    "FFFF6666",  # rojo claro
    "FFC00000",  # rojo oscuro Excel
    "FFE06666",  # rojo claro Google
    "FFFF9999",  # rojo pastel
}


def file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def is_red(cell) -> tuple[bool, str | None, str | None]:
    """Retorna (is_red, fill_rgb, font_rgb)."""
    fill_rgb = None
    font_rgb = None
    try:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == "rgb":
            fill_rgb = cell.fill.fgColor.rgb
    except Exception:
        pass
    try:
        if cell.font and cell.font.color and cell.font.color.type == "rgb":
            font_rgb = cell.font.color.rgb
    except Exception:
        pass
    red = (fill_rgb in RED_FILLS) or (font_rgb in RED_FILLS)
    return red, fill_rgb, font_rgb


def extract():
    if not SOURCE.exists():
        sys.exit(f"FATAL: {SOURCE} no existe")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Hash de control
    src_hash = file_hash(SOURCE)
    (OUT_DIR / "ict14_source.sha256").write_text(
        f"{src_hash}  {SOURCE.name}\n", encoding="utf-8"
    )

    # Carga
    wb = load_workbook(SOURCE, data_only=False, keep_links=False)

    out_data: dict = {}
    red_rows: list[dict] = []

    for ws in wb.worksheets:
        sheet = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "formulas": {},   # {coord: formula_str}
            "values": {},     # {coord: literal_value}  (NO fórmulas)
            "red_cells": [],  # [{coord, fill, font, value}]
        }
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                coord = cell.coordinate
                is_formula = isinstance(v, str) and v.startswith("=")
                if is_formula:
                    sheet["formulas"][coord] = v
                else:
                    # JSON-friendly value (datetime, decimals)
                    try:
                        json.dumps(v)
                        sheet["values"][coord] = v
                    except TypeError:
                        sheet["values"][coord] = str(v)
                red, fill, font = is_red(cell)
                if red:
                    rc = {
                        "coord": coord,
                        "fill": fill,
                        "font": font,
                        "is_formula": is_formula,
                        "value": v if is_formula else (
                            v if isinstance(v, (int, float, str, bool)) else str(v)
                        ),
                    }
                    sheet["red_cells"].append(rc)
                    red_rows.append({"sheet": ws.title, **rc})

        # Hash por hoja (solo fórmulas — captura el "shape" semántico)
        sheet["formulas_hash"] = hashlib.sha256(
            json.dumps(sheet["formulas"], sort_keys=True).encode("utf-8")
        ).hexdigest()[:16]
        sheet["formulas_count"] = len(sheet["formulas"])
        sheet["values_count"] = len(sheet["values"])
        sheet["red_count"] = len(sheet["red_cells"])

        out_data[ws.title] = sheet

    # Escribir JSON principal
    (OUT_DIR / "golden_ict14.json").write_text(
        json.dumps(out_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # Escribir CSV red cells
    with open(OUT_DIR / "golden_ict14_redcells.csv", "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=["sheet", "coord", "is_formula", "fill", "font", "value"]
        )
        writer.writeheader()
        for r in red_rows:
            writer.writerow({k: r.get(k, "") for k in writer.fieldnames})

    # Escribir inventario MD
    lines = [
        "# Inventario ICT_14 (Golden Master)\n",
        f"- **Archivo**: `{SOURCE.name}`",
        f"- **SHA256**: `{src_hash}`",
        f"- **Total hojas**: {len(out_data)}",
        "",
        "## Resumen por hoja\n",
        "| Hoja | Filas | Cols | Fórmulas | Valores | Celdas Rojas | Hash fórmulas |",
        "|------|-------|------|----------|---------|--------------|---------------|",
    ]
    total_f = total_v = total_r = 0
    for name, s in out_data.items():
        lines.append(
            f"| {name} | {s['max_row']} | {s['max_col']} | "
            f"{s['formulas_count']} | {s['values_count']} | "
            f"{s['red_count']} | `{s['formulas_hash']}` |"
        )
        total_f += s["formulas_count"]
        total_v += s["values_count"]
        total_r += s["red_count"]
    lines.append(
        f"| **TOTAL** | — | — | **{total_f}** | **{total_v}** | **{total_r}** | — |"
    )
    (OUT_DIR / "golden_ict14_inventory.md").write_text(
        "\n".join(lines), encoding="utf-8"
    )

    return {
        "hash": src_hash,
        "sheets": len(out_data),
        "total_formulas": total_f,
        "total_values": total_v,
        "total_red": total_r,
        "by_sheet": {n: s["formulas_count"] for n, s in out_data.items()},
        "red_by_sheet": {n: s["red_count"] for n, s in out_data.items()},
    }


if __name__ == "__main__":
    result = extract()
    print(json.dumps(result, ensure_ascii=False, indent=2))
