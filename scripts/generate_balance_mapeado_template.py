"""Genera el Excel modelo "Balance Mapeado ICT 2025" para que el cliente
lo descargue, lo llene con sus cuentas y lo suba al portal.

Estructura exacta que espera el parser
(backend/app/ict/parsers/balance_mapeado_excel.py):

  - Filas 1-10: encabezado descriptivo / instrucciones (ignorado por el parser)
  - Fila 11: encabezados de columnas (el parser detecta "Códigos SRI" + "Saldo")
  - Filas 12+: datos
      Col A — Cod.Cuenta.Contable (opcional)
      Col B — Descripción Cuenta Contable (opcional pero recomendado)
      Col C — (espacio entre Descripción y SRI)
      Col D — Códigos SRI         (OBLIGATORIO)
      Col E — Saldos 31 DIC        (OBLIGATORIO numérico)

Uso:
    python scripts/generate_balance_mapeado_template.py
    → genera frontend-client/public/plantillas/balance_mapeado_ICT_2025.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ----- Estilos -----
THIN = Side(border_style="thin", color="A0A0A0")
THICK = Side(border_style="medium", color="2D5F8B")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_THICK = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

FILL_TITLE = PatternFill("solid", fgColor="1F3A5F")
FILL_HEADER = PatternFill("solid", fgColor="4A7BA8")
FILL_INFO = PatternFill("solid", fgColor="F4F7FB")
FILL_OBLIGATORIO = PatternFill("solid", fgColor="FFF3CD")  # amarillo claro
FILL_OPCIONAL = PatternFill("solid", fgColor="E8F5E9")     # verde claro
FILL_EJEMPLO = PatternFill("solid", fgColor="FAFAFC")
FILL_SECCION = PatternFill("solid", fgColor="D6E4F0")
FILL_TOTAL = PatternFill("solid", fgColor="E8F1F8")

FONT_TITLE = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Calibri", size=11, bold=True, color="1F3A5F")
FONT_INFO = Font(name="Calibri", size=9, color="5A6575")
FONT_INFO_BOLD = Font(name="Calibri", size=9, bold=True, color="2D5F8B")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_DATA = Font(name="Calibri", size=10)
FONT_EJEMPLO_ITALIC = Font(name="Calibri", size=9, italic=True, color="9E9E9E")
FONT_SECCION = Font(name="Calibri", size=11, bold=True, color="1F3A5F")
FONT_TOTAL = Font(name="Calibri", size=11, bold=True, color="1B5E20")


# ----- Filas de ejemplo: cubren TODOS los bloques principales del F-101 -----
# Estructura: (codigo, descripcion, casillero_sri, saldo, es_seccion?)
EJEMPLOS: list[tuple[str, str, str, float | None, bool]] = [
    # ─── ACTIVOS CORRIENTES ──────────────────────────────────────────────────
    ("", "ACTIVOS CORRIENTES", "", None, True),
    ("1.1.01.001", "Caja General",                                "311",  5000.00, False),
    ("1.1.01.002", "Caja Chica",                                  "311",  500.00, False),
    ("1.1.02.001", "Banco Pichincha Cta. Cte. 1234567",           "311",  85000.00, False),
    ("1.1.02.002", "Banco Produbanco Cta. Ahorros 9876543",       "311",  12500.00, False),
    ("1.1.03.001", "Inversiones Temporales — Pólizas",            "312",  50000.00, False),
    ("1.1.04.001", "Clientes Locales Relacionados",               "313",  35000.00, False),
    ("1.1.04.002", "Clientes Locales NO Relacionados",            "315",  120000.00, False),
    ("1.1.04.099", "(-) Provisión Cuentas Incobrables",           "317",  -3500.00, False),
    ("1.1.05.001", "Otras Cuentas por Cobrar Empleados",          "325",  2800.00, False),
    ("1.1.06.001", "Anticipo IR Retenido",                        "336",  4200.00, False),
    ("1.1.06.002", "Crédito Tributario IVA",                      "338",  8900.00, False),
    ("1.1.07.001", "Inventario Productos Terminados",             "340",  180000.00, False),
    ("1.1.07.002", "Inventario Materias Primas",                  "341",  95000.00, False),
    ("1.1.07.099", "(-) Provisión Obsolescencia Inventario",      "347",  -2500.00, False),

    # ─── ACTIVOS NO CORRIENTES ───────────────────────────────────────────────
    ("", "ACTIVOS NO CORRIENTES", "", None, True),
    ("1.2.01.001", "Terrenos",                                    "381",  150000.00, False),
    ("1.2.01.002", "Edificios",                                   "382",  400000.00, False),
    ("1.2.01.003", "Maquinaria y Equipo",                         "383",  220000.00, False),
    ("1.2.01.004", "Muebles y Enseres",                           "383",  35000.00, False),
    ("1.2.01.005", "Equipo de Cómputo",                           "383",  18000.00, False),
    ("1.2.01.006", "Vehículos",                                   "383",  62000.00, False),
    ("1.2.01.099", "(-) Depreciación Acumulada Edificios",        "384",  -45000.00, False),
    ("1.2.01.098", "(-) Depreciación Acumulada Maq. y Eq.",       "385",  -68000.00, False),
    ("1.2.01.097", "(-) Depreciación Acumulada Vehículos",        "386",  -22000.00, False),
    ("1.2.02.001", "Software y Licencias",                        "391",  12000.00, False),
    ("1.2.02.099", "(-) Amortización Acumulada Intangibles",      "392",  -4800.00, False),
    ("1.2.03.001", "Inversiones Largo Plazo",                     "432",  25000.00, False),

    # ─── PASIVOS CORRIENTES ──────────────────────────────────────────────────
    ("", "PASIVOS CORRIENTES", "", None, True),
    ("2.1.01.001", "Proveedores Locales",                         "513",  85000.00, False),
    ("2.1.01.002", "Proveedores del Exterior",                    "514",  18500.00, False),
    ("2.1.02.001", "Cuentas por Pagar Empleados",                 "519",  6800.00, False),
    ("2.1.03.001", "Préstamo Bancario Corto Plazo",               "525",  120000.00, False),
    ("2.1.04.001", "Impuesto a la Renta por Pagar",               "532",  15500.00, False),
    ("2.1.04.002", "Participación Trabajadores por Pagar",        "533",  6200.00, False),
    ("2.1.04.003", "IESS por Pagar",                              "534",  4800.00, False),
    ("2.1.04.004", "Jubilación Patronal Corriente",               "535",  2500.00, False),
    ("2.1.04.005", "Beneficios Sociales por Pagar",               "536",  12800.00, False),
    ("2.1.05.001", "Provisiones Corrientes — Otras",              "544",  8500.00, False),

    # ─── PASIVOS NO CORRIENTES ───────────────────────────────────────────────
    ("", "PASIVOS NO CORRIENTES", "", None, True),
    ("2.2.01.001", "Préstamo Bancario Largo Plazo",               "555",  280000.00, False),
    ("2.2.02.001", "Pasivo por Impuesto a la Renta Diferido",     "572",  15800.00, False),
    ("2.2.02.002", "Jubilación Patronal Largo Plazo",             "573",  68500.00, False),
    ("2.2.02.003", "Desahucio",                                   "574",  32500.00, False),

    # ─── PATRIMONIO ──────────────────────────────────────────────────────────
    ("", "PATRIMONIO", "", None, True),
    ("3.1.01.001", "Capital Social",                              "601",  500000.00, False),
    ("3.1.02.001", "Reserva Legal",                               "604",  45000.00, False),
    ("3.1.03.001", "Utilidades Acumuladas Ejercicios Anteriores", "611",  185000.00, False),
    ("3.1.03.002", "(-) Pérdidas Acumuladas",                     "612",  -22000.00, False),
    ("3.1.03.003", "Utilidad del Ejercicio",                      "614",  85000.00, False),

    # ─── INGRESOS ────────────────────────────────────────────────────────────
    ("", "INGRESOS DE ACTIVIDADES ORDINARIAS", "", None, True),
    ("4.1.01.001", "Ventas Locales Tarifa 12% IVA",               "6005", 850000.00, False),
    ("4.1.01.002", "Ventas Locales Tarifa 0% IVA",                "6001", 125000.00, False),
    ("4.1.02.001", "Exportaciones de Bienes",                     "6011", 0.00, False),
    ("4.1.03.001", "Ingresos por Servicios",                      "6021", 35000.00, False),
    ("4.1.04.001", "Otros Ingresos Operacionales",                "6051", 8500.00, False),

    # ─── COSTOS Y GASTOS ─────────────────────────────────────────────────────
    ("", "COSTOS Y GASTOS", "", None, True),
    ("5.1.01.001", "Inventario Inicial Bienes Reventa",           "7001", 165000.00, False),
    ("5.1.02.001", "Compras Netas Locales Bienes",                "7002", 425000.00, False),
    ("5.1.03.001", "(-) Inventario Final Bienes Reventa",         "7010", -180000.00, False),
    ("5.2.01.001", "Sueldos y Salarios",                          "7041", 145000.00, False),
    ("5.2.01.002", "Beneficios Sociales",                         "7053", 28500.00, False),
    ("5.2.02.001", "Honorarios Profesionales",                    "7065", 22000.00, False),
    ("5.2.03.001", "Mantenimiento y Reparaciones",                "7113", 8500.00, False),
    ("5.2.04.001", "Combustibles y Lubricantes",                  "7117", 6800.00, False),
    ("5.2.05.001", "Promoción y Publicidad",                      "7173", 12500.00, False),
    ("5.2.06.001", "Gastos de Viaje",                             "7182", 4200.00, False),
    ("5.2.07.001", "Gastos de Gestión",                           "7185", 3800.00, False),
    ("5.2.08.001", "Servicios Básicos",                           "7197", 9500.00, False),
    ("5.2.09.001", "Arriendos",                                   "7201", 36000.00, False),
    ("5.2.10.001", "Depreciaciones del Ejercicio",                "7225", 24500.00, False),
    ("5.2.11.001", "Impuestos, Contribuciones y Otros",           "7257", 8200.00, False),
    ("5.2.12.001", "Otros Gastos Operacionales",                  "7299", 5500.00, False),
]


def _write_intro_section(ws) -> None:
    """Filas 1-10: título + instrucciones — el parser las ignora."""
    # Título principal
    ws.merge_cells("A1:E2")
    c = ws.cell(1, 1, value="BALANCE MAPEADO — PLANTILLA ICT 2025")
    c.font = FONT_TITLE
    c.fill = FILL_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    # Subtítulo + sello
    ws.merge_cells("A3:E3")
    c = ws.cell(3, 1, value="AuditBrain · Audit Consulting Group · www.auditconsulting.ec")
    c.font = FONT_INFO
    c.alignment = Alignment(horizontal="center")

    # Instrucciones
    ws.merge_cells("A5:E5")
    c = ws.cell(5, 1, value="📋 INSTRUCCIONES")
    c.font = FONT_SUBTITLE
    c.alignment = Alignment(horizontal="left", indent=1)

    # 5 instrucciones en filas 6-10 (fila 11 reservada para headers del parser)
    instrucciones = [
        "1. Una fila por cada cuenta contable del balance al 31 de diciembre.",
        "2. Columnas OBLIGATORIAS (amarillo): 'Códigos SRI' (D) y 'Saldos 31 DIC' (E).",
        "3. 'Códigos SRI' = casillero del F-101 al que mapea la cuenta (ej. 311 = Efectivo).",
        "4. Saldos negativos son normales para pasivos/patrimonio — el sistema los normaliza.",
        "5. Sube en el portal: ICT 2025 → Balance Mapeado → ▶ Procesar.",
    ]
    for i, txt in enumerate(instrucciones, start=6):
        ws.merge_cells(f"A{i}:E{i}")
        c = ws.cell(i, 1, value=txt)
        c.font = FONT_INFO
        c.fill = FILL_INFO
        c.alignment = Alignment(horizontal="left", indent=2, wrap_text=True)
        c.border = BORDER


def _write_headers(ws, row: int = 11) -> None:
    """Fila 11 con los headers EXACTOS que el parser detecta."""
    headers = [
        ("Cod.Cuenta.Contable",        FILL_OPCIONAL,     "Opcional pero recomendado"),
        ("Descripción Cuenta Contable", FILL_OPCIONAL,    "Opcional pero recomendado"),
        ("",                            FILL_HEADER,       ""),  # col C separadora
        ("Códigos SRI",                FILL_OBLIGATORIO,  "OBLIGATORIO: nº casillero F-101"),
        ("Saldos 31 DIC",              FILL_OBLIGATORIO,  "OBLIGATORIO: numérico"),
    ]
    for i, (text, fill, _) in enumerate(headers, start=1):
        c = ws.cell(row, i, value=text)
        c.font = FONT_HEADER if fill == FILL_HEADER else Font(
            name="Calibri", size=10, bold=True, color="1F3A5F"
        )
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_THICK
    ws.row_dimensions[row].height = 32


def _write_data_rows(ws, start_row: int = 12) -> int:
    """Filas 12+ con ejemplos. Devuelve la última fila usada."""
    r = start_row
    for codigo, desc, cas, saldo, es_seccion in EJEMPLOS:
        if es_seccion:
            # Fila separadora de sección — el parser la ignora (sin SRI ni saldo)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            c = ws.cell(r, 1, value=f"━━━━━━━━━━  {desc}  ━━━━━━━━━━")
            c.font = FONT_SECCION
            c.fill = FILL_SECCION
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
            ws.row_dimensions[r].height = 22
            r += 1
            continue

        # Fila de dato real
        ws.cell(r, 1, value=codigo).font = FONT_DATA
        ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(r, 1).border = BORDER

        ws.cell(r, 2, value=desc).font = FONT_DATA
        ws.cell(r, 2).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(r, 2).border = BORDER

        # Col C vacía (separadora)
        ws.cell(r, 3, value="").border = BORDER

        # Col D: Códigos SRI
        cd = ws.cell(r, 4, value=cas)
        cd.font = Font(name="Calibri", size=10, bold=True, color="1F3A5F")
        cd.alignment = Alignment(horizontal="center")
        cd.border = BORDER
        cd.fill = FILL_OBLIGATORIO

        # Col E: Saldo
        ce = ws.cell(r, 5, value=saldo)
        ce.font = FONT_DATA
        ce.number_format = '#,##0.00;-#,##0.00;0.00'
        ce.alignment = Alignment(horizontal="right")
        ce.border = BORDER
        ce.fill = FILL_OBLIGATORIO

        r += 1

    # Agregar 20 filas vacías al final para que el usuario las llene
    for _ in range(20):
        for c in range(1, 6):
            cell = ws.cell(r, c)
            cell.border = BORDER
            if c == 4:
                cell.fill = FILL_OBLIGATORIO
            elif c == 5:
                cell.fill = FILL_OBLIGATORIO
                cell.number_format = '#,##0.00;-#,##0.00;0.00'
                cell.alignment = Alignment(horizontal="right")
        r += 1

    return r - 1


def _set_column_widths(ws) -> None:
    widths = {"A": 22, "B": 55, "C": 3, "D": 16, "E": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _add_freeze_panes(ws) -> None:
    """Mantiene visible el header al hacer scroll."""
    ws.freeze_panes = "A12"


def _add_summary_footer(ws, last_data_row: int) -> None:
    """Después de los datos: nota + estadísticas con fórmulas."""
    r = last_data_row + 2
    # Línea separadora
    for c in range(1, 6):
        cell = ws.cell(r, c)
        cell.fill = FILL_TOTAL
        cell.border = Border(top=THICK)
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(r, 1, value="📊 RESUMEN AUTOMÁTICO (no editar)")
    c.font = FONT_TOTAL
    c.alignment = Alignment(horizontal="left", indent=1)

    r += 1
    metricas = [
        ("Total filas con datos", f"=COUNTA(D12:D{last_data_row})"),
        ("Total filas con saldo", f"=COUNT(E12:E{last_data_row})"),
        ("Suma de saldos (debe ser ≈ 0 si balance cuadra)", f"=SUM(E12:E{last_data_row})"),
    ]
    for label, formula in metricas:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(r, 1, value=label)
        c.font = FONT_INFO_BOLD
        c.alignment = Alignment(horizontal="left", indent=2)
        c.fill = FILL_TOTAL

        cv = ws.cell(r, 5, value=formula)
        cv.font = FONT_TOTAL
        cv.number_format = '#,##0.00;-#,##0.00;0.00'
        cv.alignment = Alignment(horizontal="right")
        cv.fill = FILL_TOTAL
        cv.border = BORDER
        r += 1


def build_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Mapeado"

    _set_column_widths(ws)
    _write_intro_section(ws)
    _write_headers(ws, row=11)
    last = _write_data_rows(ws, start_row=12)
    _add_summary_footer(ws, last_data_row=last)
    _add_freeze_panes(ws)
    return wb


def main() -> None:
    wb = build_template()
    out = (Path(__file__).parent.parent / "frontend-client" / "public"
           / "plantillas" / "balance_mapeado_ICT_2025.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"OK — plantilla generada en: {out}")
    print(f"     tamaño: {out.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
