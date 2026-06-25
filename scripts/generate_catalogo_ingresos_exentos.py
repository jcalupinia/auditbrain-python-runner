"""Genera backend/app/ict/catalogo_ingresos_exentos.py desde el archivo del cliente.

Fuente: INGRESOS EXCCENTOS.xlsx (hoja "Hoja1").
  Col A = No. Casillero (6xxx)
  Col B = Descripción breve del ingreso exento / no objeto
  Col C = Normativa de respaldo

El módulo resultante mapea casillero -> (descripción, normativa) y lo consume
el A4 Cuadro 1 para autocompletar las columnas E (descripción del tipo de
ingreso exento) y F (normativa de respaldo) cuando traslada un casillero
exento declarado en el F-101.

Uso:
    python scripts/generate_catalogo_ingresos_exentos.py [ruta_xlsx]
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_SRC = Path(r"C:\Users\jcalu\Downloads\INGRESOS EXCCENTOS.xlsx")
OUT = (Path(__file__).resolve().parent.parent
       / "backend" / "app" / "ict" / "catalogo_ingresos_exentos.py")


def main() -> int:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    if not src.exists():
        print(f"ERROR: no existe {src}")
        return 1

    wb = load_workbook(src, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[tuple[str, str, str]] = []
    for r in range(4, ws.max_row + 1):
        cas = ws.cell(r, 1).value
        if cas in (None, ""):
            continue
        desc = (ws.cell(r, 2).value or "").strip()
        norm = (ws.cell(r, 3).value or "").strip()
        rows.append((str(cas).strip(), desc, norm))
    rows.sort(key=lambda x: int(x[0]))

    out: list[str] = []
    out.append('"""Catálogo de Ingresos Exentos / No Objeto (CMIE) — descripción y normativa.')
    out.append("")
    out.append("GENERADO automáticamente desde el archivo del cliente")
    out.append('`INGRESOS EXCCENTOS.xlsx`.')
    out.append("NO editar a mano: regenerar con scripts/generate_catalogo_ingresos_exentos.py.")
    out.append("")
    out.append("Mapea cada casillero de ingreso exento / no objeto (rango 6001-6999) a una")
    out.append("tupla (descripción, normativa). Lo usa el A4 Cuadro 1 para autocompletar las")
    out.append("columnas E (descripción del tipo de ingreso exento) y F (normativa de")
    out.append("respaldo) cuando traslada un casillero exento declarado en el F-101.")
    out.append("")
    out.append(f"Cobertura: {len(rows)} casilleros definidos por el cliente.")
    out.append('"""')
    out.append("from __future__ import annotations")
    out.append("")
    out.append("")
    out.append("# casillero -> (descripcion_tipo_ingreso, normativa_respaldo)")
    out.append("IE_CASILLERO_INFO: dict[str, tuple[str, str]] = {")
    for cas, desc, norm in rows:
        out.append(f"    {cas!r}: ({desc!r}, {norm!r}),")
    out.append("}")
    out.append("")
    out.append("")
    out.append("def ie_descripcion(casillero: str) -> str | None:")
    out.append('    """Descripción del tipo de ingreso exento / no objeto, o None."""')
    out.append("    info = IE_CASILLERO_INFO.get(str(casillero).strip())")
    out.append("    return info[0] if info else None")
    out.append("")
    out.append("")
    out.append("def ie_normativa(casillero: str) -> str | None:")
    out.append('    """Normativa de respaldo del ingreso exento, o None."""')
    out.append("    info = IE_CASILLERO_INFO.get(str(casillero).strip())")
    out.append("    return info[1] if info else None")
    out.append("")

    OUT.write_text("\n".join(out), encoding="utf-8")
    print(f"Generado {OUT} con {len(rows)} casilleros")
    return 0


if __name__ == "__main__":
    sys.exit(main())
