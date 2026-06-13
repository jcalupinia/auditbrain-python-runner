"""Genera ICT_15 con datos PROPHAR usando los fillers directamente.

Reproduce el flujo de `service.generate_excel` sin requerir DB.
Diseñado para auditoría comparativa contra ICT_14 (golden master).

Output:
  audit_artifacts/ict15_papel_trabajo.xlsx
  audit_artifacts/ict15_sri.xlsx
  audit_artifacts/ict15_generation_log.txt
"""
from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

# Asegurar import de backend.* (script ejecutado desde scripts/)
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

PROPHAR_DIR = Path(r"C:\Users\jcalu\Downloads\Información PROPHAR\Información PROPHAR")
BALANCE_PATH = Path(r"C:\Users\jcalu\Downloads\BALANCE MAPEADO.xlsx")
OUT_DIR = Path("audit_artifacts")


def load_prophar_data():
    """Parsea todos los archivos PROPHAR."""
    log = []
    from backend.app.ict.parsers.f101_pdf import parse_f101
    from backend.app.ict.parsers.balance_mapeado_excel import parse_balance_mapeado
    from backend.app.ict.parsers.f103_pdf import parse_f103
    from backend.app.ict.parsers.f104_pdf import parse_f104

    # F-101
    f101_bytes = (PROPHAR_DIR / "F101 2025.pdf").read_bytes()
    f101_res = parse_f101(f101_bytes)
    f101_casilleros = f101_res.get("casilleros", {})
    log.append(f"F-101: {len(f101_casilleros)} casilleros")

    # Balance Mapeado
    bal_res = parse_balance_mapeado(BALANCE_PATH.read_bytes())
    balance = bal_res.get("cuentas", [])
    bal_sin_saldo = bal_res.get("cuentas_sin_saldo", [])
    log.append(f"Balance: {len(balance)} cuentas ({len(bal_sin_saldo)} sin saldo)")

    # F-103 (12 meses)
    f103_monthly = {}
    f103_dir = PROPHAR_DIR / "103"
    if f103_dir.exists():
        for pdf in sorted(f103_dir.glob("*.pdf")):
            month = int(pdf.name[:2])
            try:
                r = parse_f103(pdf.read_bytes())
                f103_monthly[month] = r.get("casilleros", {})
            except Exception as e:
                log.append(f"  F-103 {pdf.name}: ERROR {e}")
        log.append(f"F-103: {len(f103_monthly)} meses parseados")

    # F-104 (12 meses)
    f104_monthly = {}
    f104_dir = PROPHAR_DIR / "104"
    if f104_dir.exists():
        for pdf in sorted(f104_dir.glob("*.pdf")):
            month = int(pdf.name[:2])
            try:
                r = parse_f104(pdf.read_bytes())
                f104_monthly[month] = r.get("casilleros", {})
            except Exception as e:
                log.append(f"  F-104 {pdf.name}: ERROR {e}")
        log.append(f"F-104: {len(f104_monthly)} meses parseados")

    return {
        "f101": f101_casilleros,
        "balance_mapeado": balance,
        "balance_mapeado_cuentas_sin_saldo": bal_sin_saldo,
        "f103_monthly": f103_monthly,
        "f104_monthly": f104_monthly,
    }, log


def make_mock_session(shared_context: dict):
    """Crea un mock que generate_excel pueda consumir sin DB."""
    # Crear anexos mock con anexo_code y extracted_data
    # Distribuir el shared_context según el anexo "lógico" que lo contiene.
    anexos = []
    distributions = {
        "A1": {"f101": shared_context["f101"],
               "balance_mapeado": shared_context["balance_mapeado"],
               "balance_mapeado_cuentas_sin_saldo":
                   shared_context["balance_mapeado_cuentas_sin_saldo"]},
        "A2": {"f104_monthly": shared_context["f104_monthly"]},
        "A5": {"f103_monthly": shared_context["f103_monthly"]},
    }
    for code in ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9"]:
        anexos.append(SimpleNamespace(
            anexo_code=code,
            status="ready",
            extracted_data=distributions.get(code, {}),
        ))

    return SimpleNamespace(
        id=999,
        ruc="1791859596001",
        razon_social="PROPHAR S.A.",
        ejercicio_fiscal=2025,
        numero_adhesivo="",
        anexos=anexos,
    )


def patch_generate_excel_no_db():
    """Monkey-patch para hacer generate_excel funcionar sin Session real."""
    # Las funciones internas que llaman a `db` realmente no lo usan
    # (en realidad usan `session` solo para lectura). Si pasamos
    # `db=None` y el código no lo dereferencia, funciona.
    pass


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print("=== Generando ICT_15 con datos PROPHAR ===\n")

    # 1. Parsear todos los datos
    data, log = load_prophar_data()
    for l in log:
        print(f"  {l}")

    # 2. Construir mock session
    session = make_mock_session(data)
    print(f"\nMock session: RUC={session.ruc}, anexos={len(session.anexos)}")

    # 3. Generar Excel
    print("\nGenerando Excel (este paso puede tardar)...")
    try:
        from backend.app.ict.service import generate_excel
        bytes_sri, bytes_papel = generate_excel(db=None, session=session)
        print(f"  Bytes SRI:    {len(bytes_sri):,}")
        print(f"  Bytes Papel:  {len(bytes_papel):,}")
    except Exception as e:
        import traceback
        print(f"\nERROR: {e}")
        traceback.print_exc()
        return 1

    # 4. Guardar
    (OUT_DIR / "ict15_sri.xlsx").write_bytes(bytes_sri)
    (OUT_DIR / "ict15_papel_trabajo.xlsx").write_bytes(bytes_papel)
    (OUT_DIR / "ict15_generation_log.txt").write_text(
        "\n".join(log) + "\n", encoding="utf-8"
    )

    # 5. Validar que abre sin reparaciones
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(bytes_papel))
    print(f"\nHojas generadas: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
