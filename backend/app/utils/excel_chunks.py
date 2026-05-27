"""Lectura de Excel en chunks para archivos grandes.

Problema: cargar un Excel de 50k+ filas con pandas.read_excel() en una
sola operación puede superar el rlimit de memoria (180 MB en starter).

Solución: leer por bloques (chunks) y aplicar la función agregadora en
cada bloque. Mantiene el consumo de memoria proporcional al chunk_size,
no al tamaño total del archivo.

Limitaciones conocidas:
- openpyxl/pandas no soportan chunksize nativo para .xlsx (sí para CSV).
- Implementamos chunking manual leyendo el archivo en read-only mode.
- Si el archivo tiene fórmulas complejas, pueden no recalcularse en
  modo read-only — devolvemos el último valor cacheado.

Uso típico:

    from backend.app.utils.excel_chunks import iter_excel_chunks, sum_column

    # Sumar una columna sin cargar todo en memoria
    total = sum_column(path="ventas.xlsx", sheet="Detalle", column="monto")

    # Procesar cualquier lógica custom
    for chunk_df in iter_excel_chunks("mayor.xlsx", chunk_size=5000):
        # chunk_df es un pandas DataFrame de hasta 5000 filas
        process(chunk_df)
"""

from __future__ import annotations

from pathlib import Path
from typing import Callable, Iterator

import pandas as pd
from openpyxl import load_workbook


def iter_excel_chunks(
    path: str | Path,
    sheet: str | int = 0,
    chunk_size: int = 5000,
    header_row: int = 1,
) -> Iterator[pd.DataFrame]:
    """Itera sobre un Excel devolviendo DataFrames de `chunk_size` filas.

    Args:
        path: ruta al archivo .xlsx (no .xls).
        sheet: nombre o índice de la hoja.
        chunk_size: filas por chunk. Default 5000.
        header_row: fila 1-indexed donde están los encabezados.

    Yields:
        pandas.DataFrame con las filas del chunk, columnas tipadas
        según pandas autodetect.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]

        rows_iter = ws.iter_rows(values_only=True)

        # Saltar filas hasta header_row
        headers = None
        for idx, row in enumerate(rows_iter, start=1):
            if idx == header_row:
                headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(row)]
                break
        if headers is None:
            return

        buffer: list[tuple] = []
        for row in rows_iter:
            buffer.append(row)
            if len(buffer) >= chunk_size:
                yield pd.DataFrame(buffer, columns=headers)
                buffer = []
        if buffer:
            yield pd.DataFrame(buffer, columns=headers)
    finally:
        wb.close()


def count_rows(path: str | Path, sheet: str | int = 0) -> int:
    """Cuenta filas de una hoja sin cargar el contenido."""
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
        return int(ws.max_row or 0)
    finally:
        wb.close()


def sum_column(
    path: str | Path,
    sheet: str | int = 0,
    column: str | int = 0,
    chunk_size: int = 10000,
) -> float:
    """Suma una columna numérica sin cargar todo el archivo.

    Útil para totales de mayor contable, ventas, etc., sin OOM.
    """
    total = 0.0
    for chunk in iter_excel_chunks(path, sheet=sheet, chunk_size=chunk_size):
        if isinstance(column, int):
            col_name = chunk.columns[column]
        else:
            col_name = column
        series = pd.to_numeric(chunk[col_name], errors="coerce")
        total += float(series.fillna(0).sum())
    return total


def groupby_sum(
    path: str | Path,
    sheet: str | int = 0,
    group_col: str | int = 0,
    sum_col: str | int = 1,
    chunk_size: int = 10000,
) -> dict[str, float]:
    """GroupBy con suma sin cargar el archivo entero.

    Devuelve un dict {clave_grupo: suma}. Compatible con archivos
    arbitrariamente grandes mientras el número de grupos únicos
    quepa en memoria.
    """
    accum: dict[str, float] = {}
    for chunk in iter_excel_chunks(path, sheet=sheet, chunk_size=chunk_size):
        group_name = (
            chunk.columns[group_col] if isinstance(group_col, int) else group_col
        )
        sum_name = (
            chunk.columns[sum_col] if isinstance(sum_col, int) else sum_col
        )
        chunk[sum_name] = pd.to_numeric(chunk[sum_name], errors="coerce").fillna(0)
        partial = chunk.groupby(group_name)[sum_name].sum()
        for k, v in partial.items():
            key = str(k)
            accum[key] = accum.get(key, 0.0) + float(v)
    return accum


def apply_to_chunks(
    path: str | Path,
    func: Callable[[pd.DataFrame], dict],
    sheet: str | int = 0,
    chunk_size: int = 5000,
) -> list[dict]:
    """Aplica una función a cada chunk y devuelve lista de resultados.

    Útil para validaciones, detección de duplicados parciales,
    transformaciones por bloque, etc.
    """
    return [func(chunk) for chunk in iter_excel_chunks(path, sheet=sheet, chunk_size=chunk_size)]


# ---------------------------------------------------------------------------
# Helpers de seguridad y diagnóstico
# ---------------------------------------------------------------------------

def estimate_memory_mb(path: str | Path, sheet: str | int = 0) -> float:
    """Estima cuánta RAM consumiría leer el Excel completo con pandas.

    Aproximación: 8 bytes por celda numérica × overhead 5x. Útil para
    decidir si vale la pena usar chunks o cargar todo de una vez.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        cells = rows * cols
        # Estimación conservadora: 40 bytes/celda con overhead pandas.
        return cells * 40 / (1024 * 1024)
    finally:
        wb.close()


def should_use_chunks(path: str | Path, sheet: str | int = 0, threshold_mb: float = 100.0) -> bool:
    """Decide si conviene usar chunks según tamaño estimado.

    Default threshold: 100 MB (deja margen sobre el rlimit de 180 MB).
    """
    return estimate_memory_mb(path, sheet) > threshold_mb
