"""Lee el 'balance resumido' (plantilla .xlsx del informe de auditoría externa).

La plantilla la genera este mismo backend (ver exporter.build_plantilla). El
parser no depende de coordenadas fijas: localiza la fila de encabezado por la
columna 'clave' y las columnas de año por su valor numérico de encabezado, de
modo que tolera filas/columnas insertadas por el usuario.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from backend.app.tax.planificacion_utilidades import schema
from backend.app.tax.planificacion_utilidades.mapping import (
    PLANTILLA_KEY_HEADER,
    PLANTILLA_SHEET,
)


def extract_balance_resumido(xlsx_bytes: bytes) -> dict:
    """Devuelve {data, params, warnings, source} desde la plantilla resumida."""
    try:
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    except Exception:
        return _empty(["No se pudo abrir el .xlsx (¿archivo corrupto o protegido?)."])

    ws = wb[PLANTILLA_SHEET] if PLANTILLA_SHEET in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return _empty(["La hoja está vacía."])

    header_idx, key_col, year_cols = _locate_header(rows)
    if header_idx is None:
        return _empty([
            f"No se encontró el encabezado '{PLANTILLA_KEY_HEADER}'. "
            "Use la plantilla generada por el sistema (botón 'Plantilla')."
        ])

    warnings: list[str] = []
    data: dict[str, list] = {k: [None, None, None] for k in schema.INPUT_KEYS}
    # year_cols: lista de (col_index, anio) ordenada por anio.
    col_for_pos = _align_years([a for _, a in year_cols])

    valid_keys = set(schema.INPUT_KEYS)
    leidos = 0
    for r in rows[header_idx + 1:]:
        key = (str(r[key_col]).strip() if key_col < len(r) and r[key_col] is not None
               else "")
        if key not in valid_keys:
            continue
        for (col, anio) in year_cols:
            pos = col_for_pos.get(anio)
            if pos is None or col >= len(r):
                continue
            val = _num(r[col])
            if val is not None:
                data[key][pos] = val
                leidos += 1

    if leidos == 0:
        warnings.append("No se leyeron valores numéricos; revise la plantilla.")

    params: dict[str, str] = {}
    razon = _scan_param(rows, "razón social") or _scan_param(rows, "razon social")
    if razon:
        params["empresa"] = razon
    ruc = _scan_param(rows, "ruc")
    if ruc:
        params["ruc"] = ruc

    return {"data": data, "params": params, "warnings": warnings,
            "source": "resumido"}


def _locate_header(rows):
    """Encuentra (idx_fila, col_clave, [(col, anio)]) escaneando las primeras filas."""
    for idx, r in enumerate(rows[:15]):
        key_col = None
        year_cols: list[tuple[int, int]] = []
        for ci, cell in enumerate(r):
            if cell is None:
                continue
            s = str(cell).strip().lower()
            if s == PLANTILLA_KEY_HEADER:
                key_col = ci
            year = _as_year(cell)
            if year is not None:
                year_cols.append((ci, year))
        if key_col is not None and year_cols:
            return idx, key_col, sorted(year_cols, key=lambda x: x[1])
    return None, None, []


def _align_years(found_years: list[int]) -> dict[int, int]:
    """Mapea cada año hallado a su posición 0/1/2 según ANIOS.

    Si el año coincide con ANIOS usa ese índice; si no, asigna por orden.
    """
    out: dict[int, int] = {}
    leftover = [a for a in found_years if a not in schema.ANIOS]
    free_positions = [i for i, a in enumerate(schema.ANIOS) if a not in found_years]
    for a in found_years:
        if a in schema.ANIOS:
            out[a] = schema.ANIOS.index(a)
    for a in leftover:
        if free_positions:
            out[a] = free_positions.pop(0)
    return out


def _as_year(cell) -> int | None:
    try:
        v = int(float(cell))
    except (TypeError, ValueError):
        return None
    return v if 2000 <= v <= 2099 else None


def _num(cell) -> float | None:
    if cell is None or isinstance(cell, str):
        try:
            return float(str(cell).replace(",", "")) if cell not in (None, "") else None
        except ValueError:
            return None
    try:
        return float(cell)
    except (TypeError, ValueError):
        return None


def _scan_param(rows, label: str) -> str | None:
    """Busca 'label' en col A y devuelve el primer valor no vacío a su derecha."""
    for r in rows[:15]:
        for ci, cell in enumerate(r):
            if cell and str(cell).strip().lower().startswith(label):
                for nxt in r[ci + 1:]:
                    if nxt not in (None, ""):
                        return str(nxt).strip()
    return None


def _empty(warnings: list[str]) -> dict:
    return {"data": {k: [None, None, None] for k in schema.INPUT_KEYS},
            "params": {}, "warnings": warnings, "source": "resumido"}
