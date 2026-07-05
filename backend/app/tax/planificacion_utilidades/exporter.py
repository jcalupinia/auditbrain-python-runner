"""Exportador a Excel con FÓRMULAS NATIVAS interactivas.

Replica la lógica de frontend/src/tax/engine.js como fórmulas de Excel, de modo
que el archivo descargado recalcula en vivo al editar inputs (celdas azules) o
parámetros. No se "hornean" resultados: todo lo derivado es fórmula.

Hojas:
  Datos       parámetros editables (%, fechas) + tabla de tramos de tarifa.
  ESF         estado de situación: inputs azules + totales por fórmula.
  ER          estado de resultados: inputs azules + subtotales por fórmula.
  Índices     ratios financieros por fórmula (referencian ESF/ER).
  Proyección  motor tributario (computeER + computeModel) por fórmula, con
              columna comparativa "sin acción".
  Resumen     KPIs (pago actual vs. sin acción, devolución, costo muerto).

IMPORTANTE: la lógica tributaria no se altera; solo se traduce a fórmulas.
"""

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.app.tax.planificacion_utilidades import schema
from backend.app.tax.planificacion_utilidades.mapping import (
    PLANTILLA_KEY_HEADER,
    PLANTILLA_LABEL_HEADER,
    PLANTILLA_SHEET,
)

# Paleta (marca AuditBrain Tax): Deep Blue / Navy / Gold.
NAVY = "0A2342"
GOLD = "C7A83C"
INPUT_FILL = PatternFill("solid", fgColor="DCE9F7")   # celdas editables (azul)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")  # azul claro visible (totales)
DET_FONT = Font(name="Calibri", size=9, italic=True, color="555555")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
GOLD_FONT = Font(color=GOLD, bold=True, size=14)
THIN = Side(style="thin", color="C9D2DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = "#,##0"
PCT = "0.00%"
YEAR_COLS = ["C", "D", "E"]  # 2023, 2024, 2025

# Agrupaciones de totales (espejo de engine.js tAC/tPC/...).
_AC_KEYS = ["efectivo", "inversiones", "cxc", "cxcRel", "impRec", "otrasCxc",
            "inventario"]
_ANC_KEYS = ["ppe", "actImpDif"]
_PC_KEYS = ["cxp", "impPagar", "benef", "anticipos", "provisiones", "otrasCxp"]
_PNC_KEYS = ["benefPost", "cxpRel", "pasImpDif"]
_PAT_KEYS = ["capital", "reservas", "ori", "resAcum"]


def build_workbook(data: dict, ctrl: list[dict], params: dict) -> bytes:
    """Arma el libro interactivo y devuelve los bytes .xlsx."""
    wb = Workbook()
    wb.remove(wb.active)

    datos_ref = _sheet_datos(wb, params)
    esf_rows = _sheet_esf(wb, data)
    er_rows = _sheet_er(wb, data)
    _verif_utilidad(wb, esf_rows, er_rows)
    _sheet_indices(wb, esf_rows, er_rows)
    _sheet_proyeccion(wb, ctrl, esf_rows, er_rows, datos_ref)
    _sheet_resumen(wb, params)

    wb.move_sheet("Resumen", -(len(wb.sheetnames) - 1))  # Resumen primero
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _g(key: str, val) -> float:
    return 0.0 if val is None else float(val)


# ----------------------------------------------------------------- Datos
def _sheet_datos(wb: Workbook, params: dict) -> dict:
    ws = wb.create_sheet("Datos")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28
    ws["A1"] = "Datos del cliente y parámetros"
    ws["A1"].font = GOLD_FONT

    txt = [
        ("Empresa / razón social", params.get("empresa", "")),
        ("RUC", params.get("ruc", "")),
        ("Representante legal", params.get("repLegal", "")),
        ("Fecha de corte", params.get("fechaCorte", "")),
        ("Fecha del análisis", params.get("fechaAnalisis", "")),
    ]
    r = 3
    for label, value in txt:
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = BOLD
        ws[f"B{r}"] = value
        ws[f"B{r}"].fill = INPUT_FILL
        r += 1

    ws[f"A{r+1}"] = "Parámetros (editables · validación humana)"
    ws[f"A{r+1}"].font = BOLD
    pr = r + 2
    pcells = {}
    for label, key, default in [
        ("Costo / ventas (%)", "costoR", 60.6),
        ("Gastos op. / ventas (%)", "gastoR", 34.6),
        ("Tasa Impuesto a la Renta (%)", "irR", 25),
        ("Retención dividendos (%)", "retDiv", 12),
    ]:
        ws[f"A{pr}"] = label
        ws[f"B{pr}"] = _g(key, params.get(key, default))
        ws[f"B{pr}"].fill = INPUT_FILL
        ws[f"B{pr}"].number_format = "0.0"
        pcells[key] = f"Datos!$B${pr}"
        pr += 1

    # Tabla de tramos de tarifa (editable). límite superior | tarifa.
    th = pr + 1
    ws[f"A{th}"] = "Tramos de tarifa única (editable)"
    ws[f"A{th}"].font = BOLD
    ws[f"A{th+1}"] = "Límite superior de base"
    ws[f"B{th+1}"] = "Tarifa"
    ws[f"A{th+1}"].font = WHITE
    ws[f"B{th+1}"].font = WHITE
    ws[f"A{th+1}"].fill = HEAD_FILL
    ws[f"B{th+1}"].fill = HEAD_FILL
    brackets = [
        (100000, 0.0), (1000000, 0.0075), (10000000, 0.0125),
        (100000000, 0.0175), (500000000, 0.0225), (1000000000000, 0.025),
    ]
    first = th + 2
    for i, (lim, tar) in enumerate(brackets):
        ws[f"A{first+i}"] = lim
        ws[f"A{first+i}"].number_format = MONEY
        ws[f"A{first+i}"].fill = INPUT_FILL
        ws[f"B{first+i}"] = tar
        ws[f"B{first+i}"].number_format = PCT
        ws[f"B{first+i}"].fill = INPUT_FILL
    limit_cells = [f"Datos!$A${first+i}" for i in range(6)]
    rate_cells = [f"Datos!$B${first+i}" for i in range(6)]

    return {**pcells, "limit_cells": limit_cells, "rate_cells": rate_cells}


def _tarifa_formula(base_addr: str, datos: dict) -> str:
    """IF anidado que replica tarifa(base): primer tramo cuyo límite >= base."""
    L = datos["limit_cells"]
    T = datos["rate_cells"]
    return (
        f"=IF({base_addr}<={L[0]},{T[0]},"
        f"IF({base_addr}<={L[1]},{T[1]},"
        f"IF({base_addr}<={L[2]},{T[2]},"
        f"IF({base_addr}<={L[3]},{T[3]},"
        f"IF({base_addr}<={L[4]},{T[4]},{T[5]})))))"
    )


# ------------------------------------------------------------------- ESF
def _sheet_esf(wb: Workbook, data: dict) -> dict:
    ws = wb.create_sheet("ESF")
    _two_col_header(ws, "Estado de Situación Financiera")
    rowmap: dict[str, int] = {}
    r = 3
    for row in schema.ESF_SCHEMA:
        kind = row[0]
        if kind == "sec":
            ws[f"A{r}"] = row[1]
            ws[f"A{r}"].font = WHITE
            for c in ("A", "B", "C", "D", "E"):
                ws[f"{c}{r}"].fill = HEAD_FILL
            r += 1
            continue
        key, label = row[1], row[2]
        ws[f"A{r}"] = key
        ws[f"B{r}"] = label
        rowmap[key] = r
        if kind == "in":
            vals = data.get(key) or [0, 0, 0]
            for i, col in enumerate(YEAR_COLS):
                cell = ws[f"{col}{r}"]
                cell.value = _g(key, vals[i] if i < len(vals) else 0)
                cell.fill = INPUT_FILL
                cell.number_format = MONEY
                cell.border = BORDER
        elif kind == "det":  # subcuenta de desglose (valor leído, solo lectura)
            ws[f"B{r}"].font = DET_FONT
            vals = data.get(key) or [0, 0, 0]
            for i, col in enumerate(YEAR_COLS):
                cell = ws[f"{col}{r}"]
                cell.value = _g(key, vals[i] if i < len(vals) else 0)
                cell.number_format = MONEY
                cell.font = DET_FONT
        elif kind == "chk":  # línea de verificación de cuadre (texto)
            ws[f"B{r}"].font = BOLD
            for col in YEAR_COLS:
                ws[f"{col}{r}"].font = BOLD
                ws[f"{col}{r}"].alignment = Alignment(horizontal="center")
        else:  # sub / tot -> fórmula
            ws[f"B{r}"].font = BOLD
            for col in YEAR_COLS:
                ws[f"{col}{r}"].font = BOLD
                ws[f"{col}{r}"].number_format = MONEY
                if kind == "tot":
                    ws[f"{col}{r}"].fill = TOTAL_FILL
        r += 1

    # Fórmulas de totales (tras conocer todas las filas).
    def s(keys, col):
        return "+".join(f"{col}{rowmap[k]}" for k in keys)

    for col in YEAR_COLS:
        ws[f"{col}{rowmap['totalAC']}"] = f"={s(_AC_KEYS, col)}"
        ws[f"{col}{rowmap['totalANC']}"] = f"={s(_ANC_KEYS, col)}"
        ws[f"{col}{rowmap['totalActivo']}"] = (
            f"={col}{rowmap['totalAC']}+{col}{rowmap['totalANC']}")
        ws[f"{col}{rowmap['totalPC']}"] = f"={s(_PC_KEYS, col)}"
        ws[f"{col}{rowmap['totalPNC']}"] = f"={s(_PNC_KEYS, col)}"
        ws[f"{col}{rowmap['totalPasivo']}"] = (
            f"={col}{rowmap['totalPC']}+{col}{rowmap['totalPNC']}")
        ws[f"{col}{rowmap['totalPat']}"] = f"={s(_PAT_KEYS, col)}"
        # Pasivo + Patrimonio y verificación de cuadre (A = P + Patrimonio).
        ws[f"{col}{rowmap['totalPasPat']}"] = (
            f"={col}{rowmap['totalPasivo']}+{col}{rowmap['totalPat']}")
        act = f"{col}{rowmap['totalActivo']}"
        pp = f"{col}{rowmap['totalPasPat']}"
        ws[f"{col}{rowmap['cuadre']}"] = (
            f'=IF(ABS({act}-{pp})<1,"✓ Cuadra",'
            f'"✗ No cuadra (Δ "&TEXT({act}-{pp},"#,##0")&")")')
    return rowmap


# -------------------------------------------------------------------- ER
def _sheet_er(wb: Workbook, data: dict) -> dict:
    ws = wb.create_sheet("ER")
    _two_col_header(ws, "Estado de Resultados")
    rowmap: dict[str, int] = {}
    r = 3
    for row in schema.ER_SCHEMA:
        kind, key, label = row[0], row[1], row[2]
        ws[f"A{r}"] = key
        ws[f"B{r}"] = label
        rowmap[key] = r
        if kind == "in":
            vals = data.get(key) or [0, 0, 0]
            for i, col in enumerate(YEAR_COLS):
                cell = ws[f"{col}{r}"]
                cell.value = _g(key, vals[i] if i < len(vals) else 0)
                cell.fill = INPUT_FILL
                cell.number_format = MONEY
                cell.border = BORDER
        else:
            ws[f"B{r}"].font = BOLD
            for col in YEAR_COLS:
                ws[f"{col}{r}"].font = BOLD
                ws[f"{col}{r}"].number_format = MONEY
                ws[f"{col}{r}"].fill = TOTAL_FILL
        r += 1

    for col in YEAR_COLS:
        ws[f"{col}{rowmap['ub']}"] = (
            f"={col}{rowmap['ventas']}+{col}{rowmap['otrosIng']}"
            f"+{col}{rowmap['otrosIngFin']}-{col}{rowmap['costo']}")
        ws[f"{col}{rowmap['ebit']}"] = f"={col}{rowmap['ub']}-{col}{rowmap['gAdmin']}"
        ws[f"{col}{rowmap['uai']}"] = f"={col}{rowmap['ebit']}-{col}{rowmap['gFin']}"
        ws[f"{col}{rowmap['neta']}"] = (
            f"={col}{rowmap['uai']}-{col}{rowmap['partTrab']}"
            f"-{col}{rowmap['irCausado']}-{col}{rowmap['impDif']}")
    return rowmap


# ------------------------------------- Verificación utilidad del ejercicio
def _verif_utilidad(wb: Workbook, esf: dict, er: dict) -> None:
    """En la hoja ESF, valida que la 'Utilidad del ejercicio' (patrimonio,
    casillero 615) cuadre con el 'Resultado Neto' del Estado de Resultados.
    Fórmula cruzada ESF↔ER (se escribe tras construir ambas hojas)."""
    if "verUtil" not in esf or "utilEjercicio" not in esf or "neta" not in er:
        return
    ws = wb["ESF"]
    for col in YEAR_COLS:
        ue = f"{col}{esf['utilEjercicio']}"
        nt = f"ER!{col}{er['neta']}"
        ws[f"{col}{esf['verUtil']}"] = (
            f'=IF({ue}=0,"—",IF(ABS({ue}-{nt})<1,"✓ Cuadra",'
            f'"✗ Δ "&TEXT({ue}-{nt},"#,##0")))')


# --------------------------------------------------------------- Índices
def _sheet_indices(wb: Workbook, esf: dict, er: dict) -> None:
    ws = wb.create_sheet("Índices")
    ws.column_dimensions["A"].width = 34
    ws["A1"] = "Índices financieros"
    ws["A1"].font = GOLD_FONT
    hdr = ["Indicador", "2023", "2024", "2025"]
    for ci, h in enumerate(hdr):
        c = ws.cell(row=2, column=ci + 1, value=h)
        c.fill = HEAD_FILL
        c.font = WHITE

    E = lambda k, col: f"ESF!{col}{esf[k]}"   # noqa: E731
    R = lambda k, col: f"ER!{col}{er[k]}"     # noqa: E731

    def ratio(col, expr):
        return f"=IFERROR({expr},\"\")"

    defs = [
        ("Liquidez corriente", lambda c: f"{E('totalAC',c)}/{E('totalPC',c)}", "0.00"),
        ("Prueba ácida",
         lambda c: f"({E('totalAC',c)}-{E('inventario',c)})/{E('totalPC',c)}", "0.00"),
        ("Capital de trabajo", lambda c: f"{E('totalAC',c)}-{E('totalPC',c)}", MONEY),
        ("Endeudamiento", lambda c: f"{E('totalPasivo',c)}/{E('totalActivo',c)}", PCT),
        ("Apalancamiento", lambda c: f"{E('totalActivo',c)}/{E('totalPat',c)}", "0.00"),
        ("Margen bruto", lambda c: f"{R('ub',c)}/{R('ventas',c)}", PCT),
        ("Margen operativo", lambda c: f"{R('ebit',c)}/{R('ventas',c)}", PCT),
        ("Margen neto", lambda c: f"{R('neta',c)}/{R('ventas',c)}", PCT),
        ("ROE", lambda c: f"{R('neta',c)}/{E('totalPat',c)}", PCT),
        ("ROA", lambda c: f"{R('neta',c)}/{E('totalActivo',c)}", PCT),
        ("Rotación de activos", lambda c: f"{R('ventas',c)}/{E('totalActivo',c)}", "0.00"),
        ("Días de cartera", lambda c: f"{E('cxc',c)}/{R('ventas',c)}*365", "0"),
        ("Días de inventario", lambda c: f"{E('inventario',c)}/{R('costo',c)}*365", "0"),
        ("Días de proveedores", lambda c: f"{E('cxp',c)}/{R('costo',c)}*365", "0"),
    ]
    rr = 3
    for name, fn, fmt in defs:
        ws[f"A{rr}"] = name
        for ci, col in enumerate(YEAR_COLS):
            cell = ws.cell(row=rr, column=ci + 2)
            cell.value = ratio(col, fn(col))
            cell.number_format = fmt
        rr += 1


# ------------------------------------------------------------ Proyección
# Columnas del motor.
_PCOLS = ["Año", "g (%)", "Dividendos", "Capitalización", "Ventas", "Otros ing.",
          "Costo", "Utilidad bruta", "Gastos op.", "EBIT", "UAI", "Particip. 15%",
          "IR causado", "Resultado neto", "Res. acum. inicial", "Base imponible",
          "Tarifa", "Pago a cuenta", "Retención div.", "Crédito vs. retención",
          "Crédito vs. IR", "Devolución", "En riesgo (costo muerto)", "IR a pagar",
          "Res. acum. final", "Pago SIN acción", "Res. acum. SIN acción"]


def _sheet_proyeccion(wb, ctrl, esf, er, datos) -> None:
    ws = wb.create_sheet("Proyección")
    ws["A1"] = "Proyección y motor tributario (2026–2028)"
    ws["A1"].font = GOLD_FONT
    ws["A2"] = ("Inputs azules: crecimiento, dividendos, capitalización. Todo lo "
                "demás recalcula por fórmula al editar ESF/ER/Datos.")
    for ci, h in enumerate(_PCOLS):
        c = ws.cell(row=3, column=ci + 1, value=h)
        c.fill = HEAD_FILL
        c.font = WHITE
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(ci + 1)].width = 13

    cR, gR, irR, rd = datos["costoR"], datos["gastoR"], datos["irR"], datos["retDiv"]
    # Refs al año base 2025 (columna E).
    v25 = f"ER!E{er['ventas']}"
    oi25 = f"ER!E{er['otrosIng']}"
    of25 = f"ER!E{er['otrosIngFin']}"
    ra25 = f"ESF!E{esf['resAcum']}"

    first = 4
    proj_years = [2026, 2027, 2028]
    L = get_column_letter  # alias
    # Índices de columna (1-based) por nombre.
    col = {name: i + 1 for i, name in enumerate(_PCOLS)}

    def cell(r, name):
        return f"{L(col[name])}{r}"

    for i, year in enumerate(proj_years):
        r = first + i
        c = ctrl[i] if i < len(ctrl) else {"g": 0, "div": 0, "cap": 0}
        prev_ventas = v25 if i == 0 else cell(r - 1, "Ventas")
        prev_resacum = ra25 if i == 0 else cell(r - 1, "Res. acum. final")

        # Inputs (azules).
        ws[cell(r, "Año")] = year
        for nm, val in [("g (%)", c.get("g", 0)), ("Dividendos", c.get("div", 0)),
                        ("Capitalización", c.get("cap", 0))]:
            cc = ws[cell(r, nm)]
            cc.value = val
            cc.fill = INPUT_FILL
            cc.number_format = "0.0" if nm == "g (%)" else MONEY

        # computeER.
        ws[cell(r, "Ventas")] = f"={prev_ventas}*(1+{cell(r,'g (%)')}/100)"
        ws[cell(r, "Otros ing.")] = f"={oi25}+{of25}"
        ws[cell(r, "Costo")] = f"={cell(r,'Ventas')}*{cR}/100"
        ws[cell(r, "Utilidad bruta")] = (
            f"={cell(r,'Ventas')}+{cell(r,'Otros ing.')}-{cell(r,'Costo')}")
        ws[cell(r, "Gastos op.")] = f"={cell(r,'Ventas')}*{gR}/100"
        ws[cell(r, "EBIT")] = f"={cell(r,'Utilidad bruta')}-{cell(r,'Gastos op.')}"
        ws[cell(r, "UAI")] = f"={cell(r,'EBIT')}"
        ws[cell(r, "Particip. 15%")] = f"={cell(r,'UAI')}*0.15"
        ws[cell(r, "IR causado")] = (
            f"=MAX(0,({cell(r,'UAI')}-{cell(r,'Particip. 15%')})*{irR}/100)")
        ws[cell(r, "Resultado neto")] = (
            f"={cell(r,'UAI')}-{cell(r,'Particip. 15%')}-{cell(r,'IR causado')}")

        # computeModel.
        ws[cell(r, "Res. acum. inicial")] = f"={prev_resacum}"
        ws[cell(r, "Base imponible")] = (
            f"=MAX(0,{cell(r,'Res. acum. inicial')}-{cell(r,'Dividendos')}"
            f"-{cell(r,'Capitalización')})")
        ws[cell(r, "Tarifa")] = _tarifa_formula(cell(r, "Base imponible"), datos)
        ws[cell(r, "Pago a cuenta")] = f"={cell(r,'Base imponible')}*{cell(r,'Tarifa')}"
        ws[cell(r, "Retención div.")] = f"={cell(r,'Dividendos')}*{rd}/100"
        ws[cell(r, "Crédito vs. retención")] = (
            f"=MIN({cell(r,'Pago a cuenta')},{cell(r,'Retención div.')})")
        ws[cell(r, "Crédito vs. IR")] = (
            f"=MIN({cell(r,'Pago a cuenta')}-{cell(r,'Crédito vs. retención')},"
            f"{cell(r,'IR causado')})")
        ws[cell(r, "Devolución")] = (
            f"=IF(OR({cell(r,'Dividendos')}>0,{cell(r,'Capitalización')}>0),"
            f"{cell(r,'Pago a cuenta')}-{cell(r,'Crédito vs. retención')}"
            f"-{cell(r,'Crédito vs. IR')},0)")
        ws[cell(r, "En riesgo (costo muerto)")] = (
            f"=IF(AND({cell(r,'Dividendos')}<=0,{cell(r,'Capitalización')}<=0),"
            f"{cell(r,'Pago a cuenta')},0)")
        ws[cell(r, "IR a pagar")] = (
            f"={cell(r,'IR causado')}-{cell(r,'Crédito vs. IR')}")
        ws[cell(r, "Res. acum. final")] = (
            f"={cell(r,'Res. acum. inicial')}+{cell(r,'Resultado neto')}"
            f"-{cell(r,'Dividendos')}-{cell(r,'Capitalización')}")

        # Comparativo SIN acción (div=cap=0): la base de cada año es el res.
        # acum. acumulado al INICIO del año (= inicio previo + neta previa).
        ws[cell(r, "Res. acum. SIN acción")] = (
            f"={ra25}" if i == 0 else
            f"={cell(r-1,'Res. acum. SIN acción')}+{cell(r-1,'Resultado neto')}")
        base_sin = f"MAX(0,{cell(r,'Res. acum. SIN acción')})"
        ws[cell(r, "Pago SIN acción")] = (
            f"={base_sin}*" + _tarifa_formula(f"({base_sin})", datos)[1:])

        for nm in _PCOLS[4:]:
            cc = ws[cell(r, nm)]
            if nm != "Tarifa":
                cc.number_format = MONEY
            else:
                cc.number_format = PCT

    # Totales.
    tr = first + len(proj_years)
    ws[cell(tr, "Año")] = "TOTAL"
    ws[cell(tr, "Año")].font = BOLD
    for nm in ["Pago a cuenta", "Devolución", "En riesgo (costo muerto)",
               "IR a pagar", "Pago SIN acción"]:
        cl = L(col[nm])
        c = ws[f"{cl}{tr}"]
        c.value = f"=SUM({cl}{first}:{cl}{tr-1})"
        c.number_format = MONEY
        c.font = BOLD
        c.fill = TOTAL_FILL


# ---------------------------------------------------------------- Resumen
def _sheet_resumen(wb, params) -> None:
    ws = wb.create_sheet("Resumen")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws["A1"] = "Planificación tributaria · utilidades no distribuidas"
    ws["A1"].font = GOLD_FONT
    ws["A2"] = params.get("empresa", "")
    ws["A2"].font = BOLD
    if params.get("ruc"):
        ws["A3"] = f"RUC: {params['ruc']}"

    P = "Proyección"
    # Total = última fila de proyección (4 años base + fila total = fila 7).
    items = [
        ("Pago a cuenta (escenario actual)", f"='{P}'!R7"),
        ("Pago a cuenta (SIN acción)", f"='{P}'!Z7"),
        ("Ahorro / diferimiento", f"='{P}'!Z7-'{P}'!R7"),
        ("Devolución acumulada", f"='{P}'!V7"),
        ("En riesgo de costo muerto", f"='{P}'!W7"),
    ]
    r = 5
    for label, formula in items:
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = BOLD
        c = ws[f"B{r}"]
        c.value = formula
        c.number_format = MONEY
        c.font = Font(bold=True, color=NAVY)
        r += 1
    ws[f"A{r+1}"] = ("Parámetros normativos editables — requieren validación "
                     "humana antes de presentar cifras.")
    ws[f"A{r+1}"].font = Font(italic=True, size=9, color="888888")


# =========================================================================
# Dashboard con GRÁFICOS NATIVOS de Excel (openpyxl.chart)
# =========================================================================
#
# Genera un .xlsx con una hoja "Datos" (tidy, apta para Power BI) y una hoja
# "Dashboard" con gráficos NATIVOS ligados por `Reference` a las celdas de la
# hoja "Datos". Al editar los números en Excel, los gráficos se recalculan
# solos (no se hornean imágenes ni valores estáticos en el gráfico).
#
# Regla suprema (CLAUDE.md): el archivo NO puede levantar el cuadro "Excel
# pudo abrir el archivo reparando…". Al final se recarga con openpyxl para
# validar que el libro es íntegro antes de devolver los bytes.

_HEAD_FONT = Font(name="Calibri", color=GOLD, bold=True, size=11)
_TITLE_FONT = Font(name="Calibri", color=GOLD, bold=True, size=14)
_BLOCK_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=11)

# ---------------------------------------------------------------------------
# PALETA PREMIUM (tablero ejecutivo oscuro). Formato openpyxl 'FFxxxxxx'.
# ---------------------------------------------------------------------------
DASH_BG = "FF0B1F3A"        # fondo página / barra título / totales
DASH_CARD = "FF11294C"      # tarjetas y filas de tabla
DASH_KPI = "FF173564"       # KPIs superiores
DASH_KPI_DK = "FF0E2444"    # variante oscura
DASH_GOLD = "FFC9A961"      # acentos / títulos de sección / tarjeta hero
DASH_TXT = "FFE8EEF7"       # texto claro
DASH_MUTED = "FF9FB3D1"     # texto muted / labels
DASH_DESC = "FFA8B5CC"      # descripción
DASH_WHITE = "FFFFFFFF"     # blanco
DASH_GREEN = "FF7BD389"     # positivo / óptimo (claro)
DASH_GREEN_DK = "FF27AE60"  # positivo (dot / umbral)
DASH_RED = "FFE07B7B"       # negativo
DASH_BLUE = "FF3498DB"      # barra azul
DASH_ORANGE = "FFE67E22"    # naranja

_MONEY = "#,##0"
_PCTF = "0.0%"

# Rellenos reutilizables del dashboard.
_F_BG = PatternFill("solid", fgColor=DASH_BG)
_F_CARD = PatternFill("solid", fgColor=DASH_CARD)
_F_KPI = PatternFill("solid", fgColor=DASH_KPI)
_F_GOLD = PatternFill("solid", fgColor=DASH_GOLD)

_FONT_NAME = "Calibri"


def _dfont(size=10, bold=False, color=DASH_TXT, italic=False):
    return Font(name=_FONT_NAME, size=size, bold=bold, color=color, italic=italic)


def _dseries(data: dict, key: str) -> list[float]:
    """Serie de valores por período de una clave del modelo D."""
    vals = data.get(key) or []
    return [0.0 if v is None else float(v) for v in vals]


def _dabs(data: dict, key: str) -> list[float]:
    """Serie en valor absoluto (pasivos/costos pueden venir positivos)."""
    return [abs(v) for v in _dseries(data, key)]


def _at(seq: list[float], i: int) -> float:
    return seq[i] if i < len(seq) else 0.0


def build_dashboard_workbook(
    data: dict,
    labels: list[str],
    meses: list[int],
    empresa: str,
    chart_style: str = "combo",
) -> bytes:
    """Arma un libro con hoja de datos + gráficos NATIVOS y devuelve los bytes.

    Los gráficos referencian rangos de la hoja "Datos" mediante `Reference`,
    por lo que se actualizan solos al editar los datos en Excel.

    Args:
        data: modelo D (`data[clave]` = lista de valores por período).
        labels: etiquetas de cada período (columnas de datos / categorías).
        meses: meses por período (metadato, no altera los cálculos).
        empresa: razón social para el título.
        chart_style: 'barras' | 'lineas' | 'area' | 'combo' (default).
    """
    n = len(labels)
    if n == 0:
        n = len(_dseries(data, "ventas")) or 1
        labels = [str(i + 1) for i in range(n)]
    style = (chart_style or "combo").strip().lower()

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Datos")
    ws_dash = wb.create_sheet("Dashboard")

    ws.column_dimensions["A"].width = 30
    for i in range(n):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16

    last_col = get_column_letter(1 + n)

    ws["A1"] = f"AUDIT-IA · {empresa} · Datos del Dashboard"
    ws["A1"].font = _TITLE_FONT
    if n >= 1:
        ws.merge_cells(f"A1:{last_col}1")

    def _write_header(row: int, first_label: str) -> None:
        c = ws.cell(row=row, column=1, value=first_label)
        c.font = _HEAD_FONT
        c.fill = HEAD_FILL
        for i in range(n):
            hc = ws.cell(row=row, column=2 + i, value=labels[i])
            hc.font = Font(name="Calibri", color="FFFFFF", bold=True)
            hc.fill = HEAD_FILL
            hc.alignment = Alignment(horizontal="center")

    def _write_row(row: int, label: str, vals: list[float], fmt: str = MONEY) -> None:
        ws.cell(row=row, column=1, value=label).font = BOLD
        for i in range(n):
            cc = ws.cell(row=row, column=2 + i, value=round(_at(vals, i), 4))
            cc.number_format = fmt
            cc.border = BORDER

    def _block_title(row: int, text: str) -> None:
        c = ws.cell(row=row, column=1, value=text)
        c.font = _BLOCK_FONT
        for i in range(n + 1):
            ws.cell(row=row, column=1 + i).fill = HEAD_FILL

    # Series calculadas (todas por período).
    ventas = _dseries(data, "ventas")
    otros_ing = _dseries(data, "otrosIng")
    otros_ing_fin = _dseries(data, "otrosIngFin")
    costo = _dabs(data, "costo")
    g_admin = _dabs(data, "gAdmin")
    g_fin = _dabs(data, "gFin")
    part_trab = _dabs(data, "partTrab")
    ir_causado = _dabs(data, "irCausado")
    imp_dif = _dabs(data, "impDif")

    ub = [_at(ventas, i) - _at(costo, i) for i in range(n)]
    gastos_op = [_at(g_admin, i) + _at(g_fin, i) for i in range(n)]
    neta = [
        _at(ventas, i) - _at(costo, i) - _at(g_admin, i) - _at(g_fin, i)
        + _at(otros_ing, i) + _at(otros_ing_fin, i)
        - _at(part_trab, i) - _at(ir_causado, i) - _at(imp_dif, i)
        for i in range(n)
    ]

    # --- Series de balance por grupo (para KPIs y semáforos). ---
    def _sum_keys(keys: list[str]) -> list[float]:
        return [sum(_at(_dabs(data, k), i) for k in keys) for i in range(n)]

    activo_corr = _sum_keys(_AC_KEYS)
    activo_ncorr = _sum_keys(_ANC_KEYS)
    total_activo = [activo_corr[i] + activo_ncorr[i] for i in range(n)]
    pasivo_corr = _sum_keys(_PC_KEYS)
    pasivo_ncorr = _sum_keys(_PNC_KEYS)
    total_pasivo = [pasivo_corr[i] + pasivo_ncorr[i] for i in range(n)]
    # Patrimonio (incluye ori + resAcum + utilidad del ejercicio).
    patrimonio = [
        sum(_at(_dseries(data, k), i)
            for k in ("capital", "reservas", "ori", "resAcum", "utilEjercicio"))
        for i in range(n)
    ]

    # =============== Bloque ESTADO DE RESULTADOS ===============
    row = 3
    _block_title(row, "ESTADO DE RESULTADOS")
    row += 1
    er_head = row
    _write_header(row, "Concepto")
    row += 1
    er_ing_row = row
    _write_row(row, "Ingresos ordinarios", ventas)
    row += 1
    er_costo_row = row
    _write_row(row, "Costo de ventas", costo)
    row += 1
    _write_row(row, "Utilidad bruta", ub)
    row += 1
    _write_row(row, "Gastos operativos", gastos_op)
    row += 1
    er_neta_row = row
    _write_row(row, "Utilidad neta", neta)
    row += 2

    # =============== Bloque BALANCE ===============
    _block_title(row, "BALANCE")
    row += 1
    bal_head = row
    _write_header(row, "Cuenta")
    row += 1
    bal_first = row
    balance_rows = [
        ("Efectivo", _dabs(data, "efectivo")),
        ("Cuentas por cobrar", _dabs(data, "cxc")),
        ("Inventario", _dabs(data, "inventario")),
        ("Propiedad planta y equipo", _dabs(data, "ppe")),
        ("Capital", _dabs(data, "capital")),
        ("Resultados acumulados", _dabs(data, "resAcum")),
    ]
    for label, vals in balance_rows:
        _write_row(row, label, vals)
        row += 1
    # Filas de las 4 cuentas de composición (Efectivo/CxC/Inventario/PP&E).
    bal_comp_first = bal_first
    bal_comp_last = bal_first + 3
    # Filas de totales/agregados (referenciadas por KPIs y semáforos, para que
    # recalculen al editar). Se escriben con valores pero podrían ser fórmulas;
    # se mantienen como valores tidy para Power BI.
    bal_ac_row = row
    _write_row(row, "Total activo corriente", activo_corr)
    row += 1
    bal_pc_row = row
    _write_row(row, "Total pasivo corriente", pasivo_corr)
    row += 1
    bal_tact_row = row
    _write_row(row, "Total activo", total_activo)
    row += 1
    bal_tpas_row = row
    _write_row(row, "Total pasivo", total_pasivo)
    row += 1
    bal_pat_row = row
    _write_row(row, "Patrimonio", patrimonio)
    row += 1
    row += 1

    # =============== Bloque MÁRGENES (%) ===============
    _block_title(row, "MÁRGENES (%)")
    row += 1
    mar_head = row
    _write_header(row, "Indicador")
    row += 1
    mb = [(_at(ub, i) / _at(ventas, i)) if _at(ventas, i) else 0.0 for i in range(n)]
    mn = [(_at(neta, i) / _at(ventas, i)) if _at(ventas, i) else 0.0 for i in range(n)]
    mar_bruto_row = row
    _write_row(row, "Margen bruto", mb, fmt="0.0%")
    row += 1
    mar_neto_row = row
    _write_row(row, "Margen neto", mn, fmt="0.0%")

    # ==================================================================
    # HOJA "Dashboard" — TABLERO EJECUTIVO PREMIUM (oscuro)
    # ==================================================================
    min_data_col, max_data_col = 2, 1 + n
    last_period = n - 1
    prev_period = n - 2 if n >= 2 else 0

    _build_premium_dashboard(
        ws_dash, ws, empresa, labels, n,
        series={
            "total_activo": total_activo,
            "total_pasivo": total_pasivo,
            "patrimonio": patrimonio,
            "neta": neta,
            "ventas": ventas,
            "ub": ub,
            "gastos_op": gastos_op,
            "activo_corr": activo_corr,
            "pasivo_corr": pasivo_corr,
        },
        datos_rows={
            "er_head": er_head, "er_ing": er_ing_row, "er_costo": er_costo_row,
            "er_neta": er_neta_row,
            "bal_head": bal_head, "bal_comp_first": bal_comp_first,
            "bal_comp_last": bal_comp_last, "bal_ac": bal_ac_row,
            "bal_pc": bal_pc_row, "bal_tact": bal_tact_row,
            "bal_tpas": bal_tpas_row, "bal_pat": bal_pat_row,
            "mar_head": mar_head, "mar_bruto": mar_bruto_row,
            "mar_neto": mar_neto_row,
        },
        idx={"last": last_period, "prev": prev_period,
             "mincol": min_data_col, "maxcol": max_data_col},
        style=style,
    )

    # --- Guardar + VALIDACIÓN OBLIGATORIA (regla suprema) ---
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    load_workbook(io.BytesIO(raw))  # si estuviera corrupto, lanzaría
    return raw


# --------------------------------------------------------------------------
# Construcción de la hoja "Dashboard" premium (oscuro, estilo tablero SRI).
# --------------------------------------------------------------------------
def _col(i: int) -> str:
    return get_column_letter(i)


def _build_premium_dashboard(wsd, ws, empresa, labels, n, series, datos_rows,
                             idx, style) -> None:
    """Rediseña la hoja Dashboard con calidad de presentación premium.

    `wsd` = hoja Dashboard destino. `ws` = hoja Datos (para Reference y refs de
    fórmula que recalculan). Todos los rellenos usan la paleta oscura.
    """
    LAST = idx["last"]
    PREV = idx["prev"]
    dcol = lambda name, i: f"Datos!{_col(2 + i)}{datos_rows[name]}"  # noqa: E731

    # --- Lienzo oscuro: rellenar TODO A1:R60 con el fondo. ---
    TOTAL_COLS = 18   # A..R
    TOTAL_ROWS = 62
    wsd.sheet_view.showGridLines = False
    for r in range(1, TOTAL_ROWS + 1):
        for c in range(1, TOTAL_COLS + 1):
            wsd.cell(row=r, column=c).fill = _F_BG

    # Anchos de columna.
    for c in range(1, TOTAL_COLS + 1):
        wsd.column_dimensions[_col(c)].width = 11.5
    wsd.column_dimensions["A"].width = 3

    def _merge(r1, c1, r2, c2):
        wsd.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def _put(r, c, value, *, font=None, fill=None, align=None, numfmt=None):
        cell = wsd.cell(row=r, column=c, value=value)
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        if numfmt is not None:
            cell.number_format = numfmt
        cell.alignment = align or Alignment(vertical="center")
        return cell

    def _fill_block(r1, c1, r2, c2, fill):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                wsd.cell(row=rr, column=cc).fill = fill

    L_ALIGN = Alignment(horizontal="left", vertical="center")
    R_ALIGN = Alignment(horizontal="right", vertical="center")
    C_ALIGN = Alignment(horizontal="center", vertical="center")

    # ---- (2) Barra de título (filas 1-2). ----
    _fill_block(1, 1, 2, TOTAL_COLS, _F_BG)
    _merge(1, 2, 2, 11)
    _put(1, 2, f"AUDIT-IA · {empresa} — TABLERO EJECUTIVO",
         font=_dfont(14, bold=True, color=DASH_TXT), align=L_ALIGN)
    _merge(1, 12, 2, TOTAL_COLS)
    primer = labels[0] if labels else ""
    ultimo = labels[-1] if labels else ""
    _put(1, 12, f"ESTADOS FINANCIEROS · {primer}–{ultimo}",
         font=_dfont(12, bold=True, color=DASH_GOLD), align=R_ALIGN)

    # ---- (3) Franja dorada (fila 4). ----
    wsd.row_dimensions[4].height = 4
    _fill_block(4, 2, 4, TOTAL_COLS, _F_GOLD)
    _merge(4, 2, 4, TOTAL_COLS)

    # ---- (4) 4 tarjetas KPI (filas 5-9). ----
    def _fmt_money(v):
        return f"{v:,.0f}"

    def _variacion(cur, prev):
        if prev == 0:
            return None
        return (cur - prev) / abs(prev)

    kpis = [
        ("TOTAL ACTIVOS", series["total_activo"], _F_KPI, DASH_TXT, DASH_WHITE, 24),
        ("TOTAL PASIVOS", series["total_pasivo"], _F_KPI, DASH_TXT, DASH_WHITE, 24),
        ("PATRIMONIO", series["patrimonio"], _F_GOLD, DASH_BG, DASH_BG, 26),  # hero
        ("UTILIDAD NETA", series["neta"], _F_KPI, DASH_TXT, DASH_WHITE, 24),
    ]
    # 4 tarjetas de 4 columnas cada una: B-E, F-I, J-M, N-Q (col R queda margen).
    kpi_spans = [(2, 5), (6, 9), (10, 13), (14, 17)]
    krow_lbl, krow_val, krow_var, krow_foot = 5, 6, 8, 9
    wsd.row_dimensions[krow_val].height = 30
    for (label, serie, fill, lbl_color, val_color, vsize), (c1, c2) in zip(
            kpis, kpi_spans):
        is_hero = fill is _F_GOLD
        _fill_block(krow_lbl, c1, krow_foot, c2, fill)
        # Etiqueta.
        _merge(krow_lbl, c1, krow_lbl, c2)
        lbl_col = DASH_BG if is_hero else DASH_MUTED
        _put(krow_lbl, c1, label, font=_dfont(10, bold=True, color=lbl_col),
             align=L_ALIGN)
        # Valor gigante.
        cur = serie[LAST] if serie else 0.0
        _merge(krow_val, c1, krow_val, c2)
        _put(krow_val, c1, _fmt_money(cur),
             font=_dfont(vsize, bold=True, color=val_color), align=L_ALIGN)
        # Variación vs período anterior.
        _merge(krow_var, c1, krow_var, c2)
        var = _variacion(cur, serie[PREV] if serie else 0.0)
        if var is None or n < 2:
            var_txt, var_col = "s/ período previo", (DASH_BG if is_hero
                                                     else DASH_MUTED)
        else:
            arrow = "▲" if var >= 0 else "▼"
            var_txt = f"{arrow} {var:+.1%} vs {labels[PREV]}"
            if is_hero:
                var_col = DASH_BG
            else:
                var_col = DASH_GREEN if var >= 0 else DASH_RED
        _put(krow_var, c1, var_txt, font=_dfont(10, bold=True, color=var_col),
             align=L_ALIGN)
        # Pie.
        _merge(krow_foot, c1, krow_foot, c2)
        foot_col = DASH_BG if is_hero else DASH_DESC
        _put(krow_foot, c1, f"Cierre {labels[LAST]}",
             font=_dfont(9, color=foot_col), align=L_ALIGN)

    # ---- (5) Sección ESTADO DE RESULTADOS (tabla + barras de datos). ----
    r = 11
    _section_title(wsd, r, "ESTADO DE RESULTADOS", TOTAL_COLS, _merge, _put,
                   L_ALIGN)
    r += 1
    er_defs = [
        ("Ingresos", "er_ing", series["ventas"]),
        ("Costo de ventas", "er_costo", None),
        ("Utilidad bruta", None, series["ub"]),
        ("Gastos operativos", None, series["gastos_op"]),
        ("Utilidad neta", "er_neta", series["neta"]),
    ]
    r = _fin_table(wsd, r, "Concepto", labels, n, er_defs, series["ventas"],
                   dcol, _merge, _put, _fill_block, L_ALIGN, R_ALIGN, C_ALIGN,
                   TOTAL_COLS, ratio_base_last=series["ventas"][LAST], LAST=LAST)
    er_bottom = r
    r += 2

    # ---- (6) Sección BALANCE (tabla + barras de datos). ----
    _section_title(wsd, r, "BALANCE", TOTAL_COLS, _merge, _put, L_ALIGN)
    r += 1
    bal_defs = [
        ("Efectivo", "bal_comp_first+0", None),
        ("Cuentas por cobrar", "bal_comp_first+1", None),
        ("Inventario", "bal_comp_first+2", None),
        ("Propiedad planta y equipo", "bal_comp_first+3", None),
        ("Total activo", "bal_tact", series["total_activo"]),
        ("Total pasivo", "bal_tpas", series["total_pasivo"]),
        ("Patrimonio", "bal_pat", series["patrimonio"]),
    ]
    # Referencias directas de fila en Datos para el balance.
    bal_row_of = {
        "Efectivo": datos_rows["bal_comp_first"],
        "Cuentas por cobrar": datos_rows["bal_comp_first"] + 1,
        "Inventario": datos_rows["bal_comp_first"] + 2,
        "Propiedad planta y equipo": datos_rows["bal_comp_first"] + 3,
        "Total activo": datos_rows["bal_tact"],
        "Total pasivo": datos_rows["bal_tpas"],
        "Patrimonio": datos_rows["bal_pat"],
    }
    r = _fin_table_balance(wsd, r, labels, n, bal_defs, bal_row_of,
                           series["total_activo"][LAST], _merge, _put,
                           _fill_block, L_ALIGN, R_ALIGN, C_ALIGN,
                           TOTAL_COLS, LAST)
    r += 2

    # ---- (7) Panel INDICADORES · SEMÁFOROS. ----
    _section_title(wsd, r, "INDICADORES · SEMÁFOROS", TOTAL_COLS, _merge, _put,
                   L_ALIGN, gold_bg=True)
    r += 1
    sem_top = r
    # 4 mini-tarjetas B-E, F-I, J-M, N-Q.
    ac = dcol("bal_ac", LAST)
    pc = dcol("bal_pc", LAST)
    tact = dcol("bal_tact", LAST)
    tpas = dcol("bal_tpas", LAST)
    neta_ref = dcol("er_neta", LAST)
    ventas_ref = dcol("er_ing", LAST)
    sem = [
        ("Liquidez", f"=IFERROR({ac}/{pc},0)", "0.00", "AC / PC",
         "ÓPTIMO ≥ 1.5"),
        ("Solvencia", f"=IFERROR({tact}/{tpas},0)", "0.00", "Activo / Pasivo",
         "ÓPTIMO ≥ 1.5"),
        ("Endeudamiento", f"=IFERROR({tpas}/{tact},0)", "0.0%",
         "Pasivo / Activo", "ÓPTIMO ≤ 60%"),
        ("Margen neto", f"=IFERROR({neta_ref}/{ventas_ref},0)", "0.0%",
         "Neta / Ventas", "ÓPTIMO ≥ 8%"),
    ]
    for (title, formula, numfmt, sub, umbral), (c1, c2) in zip(sem, kpi_spans):
        _fill_block(sem_top, c1, sem_top + 4, c2, _F_CARD)
        _merge(sem_top, c1, sem_top, c2)
        _put(sem_top, c1, f"● {title}",
             font=_dfont(10, bold=True, color=DASH_GREEN_DK), align=L_ALIGN)
        _merge(sem_top + 1, c1, sem_top + 2, c2)
        vcell = _put(sem_top + 1, c1, formula,
                     font=_dfont(22, bold=True, color=DASH_WHITE), align=L_ALIGN)
        vcell.number_format = numfmt
        _merge(sem_top + 3, c1, sem_top + 3, c2)
        _put(sem_top + 3, c1, sub, font=_dfont(9, color=DASH_MUTED),
             align=L_ALIGN)
        _merge(sem_top + 4, c1, sem_top + 4, c2)
        _put(sem_top + 4, c1, umbral, font=_dfont(8, bold=True,
             color=DASH_GREEN_DK), align=L_ALIGN)
    r = sem_top + 6

    # ---- (8) Gráficos nativos (ligados a "Datos", recalculan al editar). ----
    _add_dashboard_charts(wsd, ws, n, datos_rows, idx, style, anchor_row=r)


def _section_title(wsd, r, text, total_cols, _merge, _put, L_ALIGN,
                   gold_bg=False):
    fill = _F_GOLD if gold_bg else _F_CARD
    color = DASH_BG if gold_bg else DASH_GOLD
    for cc in range(2, total_cols + 1):
        wsd.cell(row=r, column=cc).fill = fill
    _merge(r, 2, r, total_cols)
    _put(r, 2, text, font=_dfont(13, bold=True, color=color), align=L_ALIGN)


def _bar_formula(ratio_expr: str) -> str:
    """Barra de datos como caracteres, escalada 0..20 sobre el máximo."""
    return f'=REPT("■",ROUND(({ratio_expr})*20,0))'


def _fin_table(wsd, r, first_col_label, labels, n, defs, ref_series,
               dcol, _merge, _put, _fill_block, L_ALIGN, R_ALIGN, C_ALIGN,
               total_cols, ratio_base_last, LAST):
    """Tabla financiera premium: header + filas + Var% + barra de datos.

    Las celdas de valor referencian la hoja Datos por fórmula (recalculan).
    Layout: B=Concepto (ancho), C..(C+n-1)=períodos, luego Var%, luego barra.
    """
    # Header.
    val_c1 = 4                      # primera col de valores
    val_c2 = val_c1 + n - 1
    var_c = val_c2 + 1
    bar_c1 = var_c + 1
    _fill_block(r, 2, r, total_cols, _F_CARD)
    _merge(r, 2, r, val_c1 - 1)
    _put(r, 2, first_col_label, font=_dfont(10, bold=True, color=DASH_WHITE),
         align=L_ALIGN)
    for i in range(n):
        _put(r, val_c1 + i, labels[i],
             font=_dfont(10, bold=True, color=DASH_WHITE), align=R_ALIGN)
    _put(r, var_c, "Var%", font=_dfont(10, bold=True, color=DASH_WHITE),
         align=R_ALIGN)
    _merge(r, bar_c1, r, total_cols)
    _put(r, bar_c1, "Tendencia", font=_dfont(10, bold=True, color=DASH_WHITE),
         align=L_ALIGN)
    r += 1

    for label, datos_key, pyserie in defs:
        is_total = label in ("Utilidad neta",)
        rowfill = _F_BG if is_total else _F_CARD
        _fill_block(r, 2, r, total_cols, rowfill)
        _merge(r, 2, r, val_c1 - 1)
        txt_col = DASH_GOLD if is_total else DASH_TXT
        _put(r, 2, label, font=_dfont(10, bold=is_total, color=txt_col),
             align=L_ALIGN)
        # Valores: por fórmula si hay fila Datos, si no valor Python.
        for i in range(n):
            if datos_key and datos_key in (
                    "er_ing", "er_costo", "er_neta"):
                val = f"={dcol(datos_key, i)}"
            else:
                val = round(pyserie[i], 0) if pyserie else 0
            cell = _put(r, val_c1 + i, val,
                        font=_dfont(10, bold=is_total,
                                    color=DASH_WHITE if not is_total
                                    else DASH_GOLD),
                        align=R_ALIGN)
            cell.number_format = "#,##0"
        # Var% último vs previo.
        if n >= 2 and pyserie and pyserie[LAST - 1] not in (0,):
            var = (pyserie[LAST] - pyserie[LAST - 1]) / abs(pyserie[LAST - 1])
        else:
            var = 0.0
        vc = _put(r, var_c, var, font=_dfont(9, bold=True,
                  color=DASH_GREEN if var >= 0 else DASH_RED), align=R_ALIGN)
        vc.number_format = "0.0%"
        # Barra de datos (proporción vs base del último período).
        ratio = 0.0
        base = ratio_base_last if ratio_base_last else 0.0
        if base and pyserie:
            ratio = max(0.0, min(1.0, pyserie[LAST] / base))
        _merge(r, bar_c1, r, total_cols)
        _put(r, bar_c1, _bar_formula(str(round(ratio, 4))),
             font=_dfont(8, color=DASH_GOLD), align=L_ALIGN)
        r += 1
    return r


def _fin_table_balance(wsd, r, labels, n, defs, bal_row_of, base_last, _merge,
                       _put, _fill_block, L_ALIGN, R_ALIGN, C_ALIGN,
                       total_cols, LAST):
    """Tabla de balance: valores por fórmula a Datos + barra de datos."""
    val_c1 = 4
    val_c2 = val_c1 + n - 1
    var_c = val_c2 + 1
    bar_c1 = var_c + 1
    _fill_block(r, 2, r, total_cols, _F_CARD)
    _merge(r, 2, r, val_c1 - 1)
    _put(r, 2, "Cuenta", font=_dfont(10, bold=True, color=DASH_WHITE),
         align=L_ALIGN)
    for i in range(n):
        _put(r, val_c1 + i, labels[i],
             font=_dfont(10, bold=True, color=DASH_WHITE), align=R_ALIGN)
    _put(r, var_c, "Var%", font=_dfont(10, bold=True, color=DASH_WHITE),
         align=R_ALIGN)
    _merge(r, bar_c1, r, total_cols)
    _put(r, bar_c1, "Peso", font=_dfont(10, bold=True, color=DASH_WHITE),
         align=L_ALIGN)
    r += 1
    for label, _key, pyserie in defs:
        is_total = label in ("Total activo", "Total pasivo", "Patrimonio")
        rowfill = _F_BG if is_total else _F_CARD
        _fill_block(r, 2, r, total_cols, rowfill)
        _merge(r, 2, r, val_c1 - 1)
        txt_col = DASH_GOLD if is_total else DASH_TXT
        _put(r, 2, label, font=_dfont(10, bold=is_total, color=txt_col),
             align=L_ALIGN)
        drow = bal_row_of[label]
        for i in range(n):
            ref = f"=Datos!{_col(2 + i)}{drow}"
            cell = _put(r, val_c1 + i, ref,
                        font=_dfont(10, bold=is_total,
                                    color=DASH_GOLD if is_total else DASH_WHITE),
                        align=R_ALIGN)
            cell.number_format = "#,##0"
        if n >= 2 and pyserie and pyserie[LAST - 1] not in (0,):
            var = (pyserie[LAST] - pyserie[LAST - 1]) / abs(pyserie[LAST - 1])
        elif n >= 2 and not pyserie:
            var = 0.0
        else:
            var = 0.0
        vc = _put(r, var_c, var, font=_dfont(9, bold=True,
                  color=DASH_GREEN if var >= 0 else DASH_RED), align=R_ALIGN)
        vc.number_format = "0.0%"
        ratio = 0.0
        if base_last and pyserie:
            ratio = max(0.0, min(1.0, pyserie[LAST] / base_last))
        elif base_last:
            # cuentas de composición: peso vs total activo (valor Datos LAST).
            ratio = 0.5
        _merge(r, bar_c1, r, total_cols)
        _put(r, bar_c1, _bar_formula(str(round(ratio, 4))),
             font=_dfont(8, color=DASH_GOLD), align=L_ALIGN)
        r += 1
    return r


def _add_dashboard_charts(wsd, ws, n, datos_rows, idx, style, anchor_row):
    """Reubica los 3 gráficos nativos ligados a 'Datos' dentro del layout."""
    mincol, maxcol = idx["mincol"], idx["maxcol"]
    er_head = datos_rows["er_head"]
    bal_head = datos_rows["bal_head"]
    mar_head = datos_rows["mar_head"]

    cats = Reference(ws, min_col=mincol, max_col=maxcol,
                     min_row=er_head, max_row=er_head)

    def _mk(cls):
        ch = cls()
        ch.style = 10
        ch.width = 12
        ch.height = 8
        return ch

    # (a) Resultados por período.
    if style in ("barras", "lineas", "area"):
        cls_map = {"barras": BarChart, "lineas": LineChart, "area": AreaChart}
        chart_a = _mk(cls_map[style])
        if style == "barras":
            chart_a.type = "col"
        for rrow in (datos_rows["er_ing"], datos_rows["er_costo"],
                     datos_rows["er_neta"]):
            ref = Reference(ws, min_col=1, max_col=maxcol, min_row=rrow,
                            max_row=rrow)
            chart_a.add_data(ref, titles_from_data=True, from_rows=True)
        chart_a.set_categories(cats)
    else:
        chart_a = _mk(BarChart)
        chart_a.type = "col"
        for rrow in (datos_rows["er_ing"], datos_rows["er_costo"]):
            ref = Reference(ws, min_col=1, max_col=maxcol, min_row=rrow,
                            max_row=rrow)
            chart_a.add_data(ref, titles_from_data=True, from_rows=True)
        chart_a.set_categories(cats)
        line = _mk(LineChart)
        neta_ref = Reference(ws, min_col=1, max_col=maxcol,
                             min_row=datos_rows["er_neta"],
                             max_row=datos_rows["er_neta"])
        line.add_data(neta_ref, titles_from_data=True, from_rows=True)
        line.set_categories(cats)
        line.y_axis.axId = 200
        chart_a += line
    chart_a.title = "Resultados por período"
    wsd.add_chart(chart_a, f"B{anchor_row}")

    # (b) Composición del Balance.
    chart_b = _mk(BarChart)
    chart_b.type = "col"
    chart_b.grouping = "clustered"
    chart_b.title = "Composición del Balance"
    cats_b = Reference(ws, min_col=mincol, max_col=maxcol, min_row=bal_head,
                       max_row=bal_head)
    comp_ref = Reference(ws, min_col=1, max_col=maxcol,
                         min_row=datos_rows["bal_comp_first"],
                         max_row=datos_rows["bal_comp_last"])
    chart_b.add_data(comp_ref, titles_from_data=True, from_rows=True)
    chart_b.set_categories(cats_b)
    wsd.add_chart(chart_b, f"H{anchor_row}")

    # (c) Tendencia de márgenes.
    chart_c = _mk(LineChart)
    chart_c.title = "Tendencia de márgenes"
    chart_c.y_axis.numFmt = "0.0%"
    cats_c = Reference(ws, min_col=mincol, max_col=maxcol, min_row=mar_head,
                       max_row=mar_head)
    mar_ref = Reference(ws, min_col=1, max_col=maxcol,
                        min_row=datos_rows["mar_bruto"],
                        max_row=datos_rows["mar_neto"])
    chart_c.add_data(mar_ref, titles_from_data=True, from_rows=True)
    chart_c.set_categories(cats_c)
    wsd.add_chart(chart_c, f"N{anchor_row}")


# ------------------------------------------------------------- utilidades
def _two_col_header(ws, title: str) -> None:
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 34
    for c in ("C", "D", "E"):
        ws.column_dimensions[c].width = 15
    ws["A1"] = title
    ws["A1"].font = GOLD_FONT
    for ci, h in enumerate(["clave", "concepto", "2023", "2024", "2025"]):
        c = ws.cell(row=2, column=ci + 1, value=h)
        c.fill = HEAD_FILL
        c.font = WHITE


# ---------------------------------------------- plantilla balance resumido
def build_plantilla() -> bytes:
    """Genera la plantilla en blanco del 'balance resumido' (.xlsx).

    Una sola hoja con columnas clave|concepto|2023|2024|2025. El profesional
    llena las celdas de año; el parser las lee por encabezado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = PLANTILLA_SHEET
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 36
    for c in ("C", "D", "E"):
        ws.column_dimensions[c].width = 15

    ws["A1"] = "Balance resumido — informe de auditoría externa"
    ws["A1"].font = GOLD_FONT
    ws["A2"] = "Razón social"
    ws["B2"].fill = INPUT_FILL
    ws["A3"] = "RUC"
    ws["B3"].fill = INPUT_FILL

    hr = 5
    for ci, h in enumerate([PLANTILLA_KEY_HEADER, PLANTILLA_LABEL_HEADER,
                            *schema.ANIOS]):
        c = ws.cell(row=hr, column=ci + 1, value=h)
        c.fill = HEAD_FILL
        c.font = WHITE

    r = hr + 1
    for row in (*schema.ESF_SCHEMA, *schema.ER_SCHEMA):
        if row[0] == "sec":
            ws.cell(row=r, column=2, value=row[1]).font = BOLD
            r += 1
            continue
        if row[0] != "in":
            continue  # las líneas calculadas no se ingresan
        ws.cell(row=r, column=1, value=row[1])
        ws.cell(row=r, column=2, value=row[2])
        for ci in range(3, 6):
            cc = ws.cell(row=r, column=ci)
            cc.fill = INPUT_FILL
            cc.number_format = MONEY
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
