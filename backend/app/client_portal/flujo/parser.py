# backend/app/client_portal/flujo/parser.py
"""Parser de la balanza homologada (hoja tipo "Mapeo") que sube el cliente.

Detecta las columnas por encabezado (cuenta contable, código Super Cías, código
SRI, saldo) para tolerar variaciones de formato entre archivos de clientes.
Acepta números en formato regional (`.` o `,` como decimal)."""
from __future__ import annotations
import io
import re
from datetime import datetime, date

from openpyxl import load_workbook

# palabras clave para detectar cada columna (normalizadas a minúsculas sin tildes)
_CLAVES = {
    "cuenta": ("cuenta", "cta", "cod.cuenta", "codigo cuenta", "cod cuenta"),
    "nombre": ("descrip", "nombre", "detalle", "concepto"),
    "super_cias": ("super", "supercias", "super cias"),
    "sri": ("sri",),
    "saldo": ("saldo", "31 dic", "valor"),
}


def _norm(s) -> str:
    s = str(s or "").strip().lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "a"), ("ñ", "n")):
        s = s.replace(a, b)
    return s


def _parse_saldo(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "")
    if not s:
        return None
    # heurística formato regional: el separador decimal es el que aparece al final
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):      # europeo: 1.234,56
            s = s.replace(".", "").replace(",", ".")
        else:                                 # us: 2,000.00
            s = s.replace(",", "")
    elif "," in s:
        # coma sola: decimal si 1-2 dígitos tras la última coma, si no miles
        ent, _, dec = s.rpartition(",")
        s = (ent.replace(",", "") + "." + dec) if len(dec) in (1, 2) else s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


_MESES = ["", "ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]


def _etiqueta_periodo(cell) -> str | None:
    """Etiqueta de período si la celda es fecha (objeto o texto) o un año.
    Fecha -> '31-may-2026'; año (int/float/texto 4 díg.) -> '2025'. Si no, None."""
    if isinstance(cell, (datetime, date)):
        return f"{cell.day:02d}-{_MESES[cell.month]}-{cell.year}"
    s = str(cell or "").strip()
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})$", s)              # ISO yyyy-mm-dd
    if m:
        y, mo, da = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12:
            return f"{da:02d}-{_MESES[mo]}-{y}"
    m = re.match(r"(\d{2})[/-](\d{2})[/-](\d{4})$", s)        # dd/mm/yyyy o dd-mm-yyyy
    if m:
        da, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12:
            return f"{da:02d}-{_MESES[mo]}-{y}"
    m = re.match(r"(19|20)\d{2}", s)                           # año solo (permite '2024.0')
    if m and (len(s) == 4 or s[4:] in (".0", "")):
        return s[:4]
    return None


def parse_balanza_multiperiodo(contenido: bytes) -> dict:
    """Lee un balance CRUDO ``Código | Cuenta | período1..N`` (sin columnas de
    homologación) y devuelve ``{"periodos": [labels], "estado": "esf"|"eri",
    "filas": [{cuenta, nombre, saldos:[...]}]}``. `saldos` va alineado a `periodos`.
    Clasifica el estado por el dígito dominante del código (1/2/3 = esf, 4/5/6 = eri).
    """
    wb = load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    filas_raw = list(ws.iter_rows(values_only=True))
    if not filas_raw:
        return {"periodos": [], "estado": "esf", "filas": []}
    def _tiene_label_cuenta(fila) -> bool:
        for v in fila:
            n = _norm(v)
            if n and any(k in n for k in ("codigo", "cuenta", "cod.cuenta", "cta")):
                return True
        return False

    hr = None
    per_cols: list[int] = []
    labels: list[str] = []
    fallback = None  # (i, cols, labs) primera fila con períodos aunque no tenga label
    for i, fila in enumerate(filas_raw[:15]):
        cols, labs = [], []
        for j, v in enumerate(fila):
            lab = _etiqueta_periodo(v)
            if lab:
                cols.append(j); labs.append(lab)
        if not cols:
            continue
        if fallback is None:
            fallback = (i, cols, labs)
        if _tiene_label_cuenta(fila):
            hr, per_cols, labels = i, cols, labs
            break
    if hr is None and fallback is not None:
        hr, per_cols, labels = fallback
    if hr is None:
        return {"periodos": [], "estado": "esf", "filas": []}

    filas: list[dict] = []
    digitos: dict[str, int] = {}
    for fila in filas_raw[hr + 1:]:
        cuenta = str(fila[0]).strip() if fila and fila[0] is not None else ""
        if not cuenta:
            continue
        nombre = str(fila[1]).strip() if len(fila) > 1 and fila[1] is not None else ""
        saldos = [_parse_saldo(fila[c]) if c < len(fila) else None for c in per_cols]
        saldos = [s if s is not None else 0.0 for s in saldos]
        filas.append({"cuenta": cuenta, "nombre": nombre, "saldos": saldos})
        d = cuenta[:1]
        if d.isdigit():
            digitos[d] = digitos.get(d, 0) + 1
    estado = "eri" if sum(digitos.get(d, 0) for d in "456") > sum(digitos.get(d, 0) for d in "123") else "esf"
    return {"periodos": labels, "estado": estado, "filas": filas}


def _detectar_columnas(fila_encabezado) -> dict[str, int]:
    """Devuelve {campo: indice_columna} detectando por encabezado."""
    col: dict[str, int] = {}
    for idx, celda in enumerate(fila_encabezado):
        h = _norm(celda)
        if not h:
            continue
        for campo, claves in _CLAVES.items():
            if campo in col:
                continue
            if any(k in h for k in claves):
                col[campo] = idx
                break
    return col


def parse_balanza(contenido: bytes) -> list[dict]:
    """Lee la primera hoja con encabezados reconocibles y devuelve filas
    ``{"cuenta", "nombre", "super_cias", "sri", "saldo"}``. Ignora filas sin
    código Super Cías o sin saldo numérico."""
    wb = load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    for ws in wb.worksheets:
        # busca la fila de encabezados en las primeras 15 filas
        for r_idx, fila in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
            col = _detectar_columnas(fila)
            if "super_cias" in col and "saldo" in col:
                return _leer_filas(ws, col, desde=r_idx + 2)
    return []


def _leer_filas(ws, col: dict[str, int], desde: int) -> list[dict]:
    out: list[dict] = []
    imax = max(col.values())
    for fila in ws.iter_rows(min_row=desde, values_only=True):
        if fila is None or len(fila) <= imax:
            continue
        sc = str(fila[col["super_cias"]] or "").strip().replace(".", "")
        if not sc.isdigit():
            continue
        saldo = _parse_saldo(fila[col["saldo"]])
        if saldo is None:
            continue
        out.append({
            "cuenta": str(fila[col["cuenta"]]).strip() if "cuenta" in col and fila[col["cuenta"]] is not None else "",
            "nombre": str(fila[col["nombre"]]).strip() if "nombre" in col and fila[col["nombre"]] is not None else "",
            "super_cias": sc,
            "sri": str(fila[col["sri"]]).strip() if "sri" in col and fila[col["sri"]] is not None else "",
            "saldo": saldo,
        })
    return out
