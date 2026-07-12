# backend/app/client_portal/flujo/generador.py
"""Generador del Excel de la Herramienta Flujo de Efectivo.

Corre los 8 motores (homologación, ESF, ERI, flujo, patrimonio, F-101,
indicadores, no-efectivo) sobre dos balanzas (año anterior / año actual) y
produce un ``.xlsx`` (openpyxl) con 9 hojas presentadas profesionalmente
(bordes, fuentes Calibri, formato numérico, identidad visual AUDIT-IA).

El archivo es un papel de trabajo auditable: los valores son numéricos ya
calculados (no fórmulas vivas), pero la hoja ``Homologación`` deja el rastro
de cada saldo y la hoja ``RESUMEN`` muestra las cuadraturas de forma explícita.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import (
    catalogos,
    motor,
    motor_er,
    motor_f101,
    motor_flujo,
    motor_indicadores,
    motor_no_efectivo,
    motor_notas,
    motor_patrimonio,
    motor_resumen,
    patrimonio_matriz,
)

# ----------------------------------------------------------------------------
# Identidad visual (CLAUDE.md): Gold #C7A83C / Navy #0A2342.
# ----------------------------------------------------------------------------
GOLD = "C7A83C"
NAVY = "0A2342"
AZUL_CLARO = "DCE6F1"   # fondo filas TOTAL
VERDE = "C6EFCE"        # semáforo cuadra
VERDE_TXT = "006100"
ROJO = "FFC7CE"         # semáforo no cuadra
ROJO_TXT = "9C0006"
BLANCO = "FFFFFF"
GRIS = "6B7280"

NUM_FMT = "#,##0.00"

_thin = Side(style="thin", color="BFBFBF")
_double = Side(style="double", color="000000")
BORDE_THIN = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDE_TOTAL = Border(left=_thin, right=_thin, top=_double, bottom=_double)

DEEP_BLUE = "071B2F"    # banda de encabezado ejecutiva
ZEBRA = "F3F6FA"        # fila alterna sutil

FONT_DATA = Font(name="Calibri", size=9, color="1F2937")
FONT_TOTAL = Font(name="Calibri", size=10, bold=True, color=NAVY)
FONT_BLOQUE = Font(name="Calibri", size=10, bold=True, color=BLANCO)
FONT_HEADER = Font(name="Calibri", size=9, bold=True, color=BLANCO)
FONT_TITULO = Font(name="Calibri", size=16, bold=True, color=NAVY)
FONT_MARCA = Font(name="Calibri", size=9, italic=True, color=GRIS)
# banda ejecutiva premium
FONT_BAND_T = Font(name="Calibri", size=17, bold=True, color=GOLD)
FONT_BAND_S = Font(name="Calibri", size=9, bold=True, color="D7E0EA")
FONT_BAND_R = Font(name="Calibri", size=9, bold=True, color=GOLD)

FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_TOTAL = PatternFill("solid", fgColor=AZUL_CLARO)
FILL_GOLD = PatternFill("solid", fgColor=GOLD)
FILL_BAND = PatternFill("solid", fgColor=DEEP_BLUE)
FILL_ZEBRA = PatternFill("solid", fgColor=ZEBRA)
FILL_BLOQUE = PatternFill("solid", fgColor=NAVY)

# borde gold inferior para acentos
_gold_side = Side(style="thin", color=GOLD)
BORDE_HEADER = Border(left=_thin, right=_thin, top=_thin, bottom=Side(style="medium", color=GOLD))

AL_IZQ = Alignment(horizontal="left", vertical="center", wrap_text=False)
AL_DER = Alignment(horizontal="right", vertical="center")
AL_CEN = Alignment(horizontal="center", vertical="center")

MARCA = "AuditConsulting Auditores Cía. Ltda."
PLATAFORMA = "AUDIT-IA"


# ----------------------------------------------------------------------------
# Helpers de escritura segura
# ----------------------------------------------------------------------------
def _safe_text(v) -> str:
    """Evita que Excel interprete un texto como fórmula (cuadro "Reparaciones").

    Antepone un apóstrofo lógico prefijando con espacio de ancho cero NO — mejor:
    si el texto empieza con =, +, - o @, se prefija un apóstrofo real que
    openpyxl trata como texto literal sin mostrarlo. Para máxima compatibilidad
    simplemente anteponemos un espacio fino solo cuando haga falta.
    """
    s = "" if v is None else str(v)
    if s[:1] in ("=", "+", "-", "@"):
        return "'" + s
    return s


def _celda_texto(ws: Worksheet, row: int, col: int, valor, *, font=FONT_DATA,
                 al=AL_IZQ, fill=None, borde=BORDE_THIN):
    c = ws.cell(row=row, column=col, value=_safe_text(valor))
    c.font = font
    c.alignment = al
    c.border = borde
    if fill is not None:
        c.fill = fill
    return c


def _celda_num(ws: Worksheet, row: int, col: int, valor, *, font=FONT_DATA,
               fill=None, borde=BORDE_THIN):
    try:
        num = round(float(valor), 2)
    except (TypeError, ValueError):
        num = 0.0
    c = ws.cell(row=row, column=col, value=num)
    c.number_format = NUM_FMT
    c.font = font
    c.alignment = AL_DER
    c.border = borde
    if fill is not None:
        c.fill = fill
    return c


def _encabezados(ws: Worksheet, row: int, titulos: list[str], anchos: list[int]):
    for i, (t, w) in enumerate(zip(titulos, anchos), start=1):
        c = ws.cell(row=row, column=i, value=_safe_text(t))
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDE_HEADER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22


def _bloque(ws: Worksheet, row: int, texto: str, ncols: int):
    c = ws.cell(row=row, column=1, value=_safe_text("  " + texto))
    c.font = FONT_BLOQUE
    c.alignment = AL_IZQ
    for i in range(1, ncols + 1):
        cell = ws.cell(row=row, column=i)
        cell.fill = FILL_BLOQUE
        cell.border = Border(bottom=_gold_side)
    ws.row_dimensions[row].height = 19


def _zebra(ws: Worksheet, row: int, ncols: int):
    """Sombreado alterno sutil para legibilidad (fila par)."""
    for i in range(1, ncols + 1):
        cell = ws.cell(row=row, column=i)
        if cell.fill is None or cell.fill.fgColor.rgb in (None, "00000000"):
            cell.fill = FILL_ZEBRA


def _titulo_hoja(ws: Worksheet, titulo: str, ncols: int) -> int:
    """Banda de encabezado ejecutiva premium (Deep Blue + acento gold).
    Devuelve la fila donde empieza el contenido."""
    n = max(ncols, 3)
    # banda de dos filas
    for r in (1, 2):
        for i in range(1, n + 1):
            ws.cell(row=r, column=i).fill = FILL_BAND
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    t = ws.cell(row=1, column=1, value=_safe_text("  " + titulo))
    t.font = FONT_BAND_T
    t.alignment = Alignment(horizontal="left", vertical="center")
    s = ws.cell(row=2, column=1,
                value=_safe_text(f"  {MARCA} · {PLATAFORMA}"))
    s.font = FONT_BAND_S
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 16
    # línea de acento gold
    for i in range(1, n + 1):
        ws.cell(row=3, column=i).fill = FILL_GOLD
    ws.row_dimensions[3].height = 3
    ws.freeze_panes = "A5"
    return 5


# ----------------------------------------------------------------------------
# Hojas
# ----------------------------------------------------------------------------
def _hoja_resumen(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("RESUMEN")
    ws.sheet_view.showGridLines = False
    _titulo_hoja(ws, "Estado de Flujo de Efectivo — Resumen ejecutivo", 5)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30

    # Semáforos de cuadratura
    _encabezados_resumen(ws, 6)
    fila = 7
    cuad_esf = ctx["cuadre_esf"]
    cuad_af = ctx["flujo"]
    filas_sem = [
        ("Cuadre ESF (Activo = Pasivo + Patrimonio)", cuad_esf["diferencia"],
         cuad_esf["cuadra"], "Diferencia debe ser 0.00"),
        ("Cuadre Flujo de Efectivo (AF)", cuad_af["cuadre_af"],
         cuad_af["cuadra"], "Efectivo final calc. − real"),
    ]
    for nombre, dif, cuadra, nota in filas_sem:
        _celda_texto(ws, fila, 2, nombre, font=FONT_DATA)
        _celda_num(ws, fila, 3, dif)
        estado = ws.cell(row=fila, column=4,
                         value=_safe_text("CUADRA" if cuadra else "NO CUADRA"))
        estado.font = Font(name="Calibri", size=9, bold=True,
                           color=VERDE_TXT if cuadra else ROJO_TXT)
        estado.fill = PatternFill("solid", fgColor=VERDE if cuadra else ROJO)
        estado.alignment = AL_CEN
        estado.border = BORDE_THIN
        _celda_texto(ws, fila, 2, nombre)
        ws.cell(row=fila, column=2).border = BORDE_THIN
        ws.cell(row=fila, column=3).border = BORDE_THIN
        # nota
        n = ws.cell(row=fila, column=5, value=_safe_text(nota))
        n.font = FONT_MARCA
        fila += 1

    # Utilidad neta (informativa, no semáforo)
    _celda_texto(ws, fila, 2, "Utilidad neta del ejercicio")
    _celda_num(ws, fila, 3, ctx["cascada"]["utilidad_neta"])
    ws.cell(row=fila, column=4, value="").border = BORDE_THIN
    fila += 1
    _celda_texto(ws, fila, 2, "Resultado integral del ejercicio")
    _celda_num(ws, fila, 3, ctx["resultado_integral"])
    ws.cell(row=fila, column=4, value="").border = BORDE_THIN
    fila += 2

    # Nota al pie
    pie = ws.cell(row=fila + 1, column=2,
                  value=_safe_text("Semáforo verde = cuadra (diferencia ≤ 1.00). "
                                   "Rojo = requiere revisión del auditor."))
    pie.font = FONT_MARCA


def _encabezados_resumen(ws: Worksheet, row: int):
    for col, t in ((2, "Verificación"), (3, "Diferencia"), (4, "Estado")):
        c = ws.cell(row=row, column=col, value=_safe_text(t))
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = AL_CEN
        c.border = BORDE_THIN


def _hoja_homologacion(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("Homologación")
    ws.sheet_view.showGridLines = False
    inicio = _titulo_hoja(ws, "Homologación — rastro de auditoría", 4)
    titulos = ["Cuenta", "Cód. Super Cías", "Cód. SRI", "Saldo"]
    anchos = [46, 18, 14, 18]

    fila = inicio
    for etiqueta, balanza in (("AÑO ANTERIOR", ctx["bal_ant"]),
                              ("AÑO ACTUAL", ctx["bal_act"])):
        _bloque(ws, fila, etiqueta, 4)
        fila += 1
        _encabezados(ws, fila, titulos, anchos)
        fila += 1
        total = 0.0
        for j, f in enumerate(balanza):
            z = FILL_ZEBRA if j % 2 else None
            _celda_texto(ws, fila, 1, f.get("cuenta", ""), fill=z)
            _celda_texto(ws, fila, 2, f.get("super_cias", ""), al=AL_CEN, fill=z)
            _celda_texto(ws, fila, 3, f.get("sri", ""), al=AL_CEN, fill=z)
            _celda_num(ws, fila, 4, f.get("saldo", 0.0), fill=z)
            total += float(f.get("saldo", 0.0) or 0.0)
            fila += 1
        _celda_texto(ws, fila, 1, "TOTAL saldos", font=FONT_TOTAL,
                     fill=FILL_TOTAL, borde=BORDE_TOTAL)
        for col in (2, 3):
            _celda_texto(ws, fila, col, "", fill=FILL_TOTAL, borde=BORDE_TOTAL)
        _celda_num(ws, fila, 4, total, font=FONT_TOTAL, fill=FILL_TOTAL,
                   borde=BORDE_TOTAL)
        fila += 2


def _hoja_estructura(wb: Workbook, nombre: str, titulo: str, estructura,
                     tot_ant: dict, tot_act: dict, subtotales: list | None = None):
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False
    inicio = _titulo_hoja(ws, titulo, 5)
    titulos = ["Código", "Etiqueta", "Saldo Anterior", "Saldo Actual", "Variación"]
    anchos = [16, 50, 18, 18, 16]
    _encabezados(ws, inicio, titulos, anchos)
    fila = inicio + 1
    idx = 0
    for n in estructura:
        ant = round(float(tot_ant.get(n.codigo, 0.0)), 2)
        act = round(float(tot_act.get(n.codigo, 0.0)), 2)
        es_seccion = len(n.codigo) <= 1
        font = FONT_TOTAL if es_seccion else FONT_DATA
        fill = FILL_TOTAL if es_seccion else (FILL_ZEBRA if idx % 2 else None)
        idx += 1
        _celda_texto(ws, fila, 1, n.codigo, al=AL_CEN, font=font, fill=fill)
        _celda_texto(ws, fila, 2, n.etiqueta, font=font, fill=fill)
        _celda_num(ws, fila, 3, ant, font=font, fill=fill)
        _celda_num(ws, fila, 4, act, font=font, fill=fill)
        _celda_num(ws, fila, 5, round(act - ant, 2), font=font, fill=fill)
        fila += 1

    if subtotales:
        fila += 1
        _bloque(ws, fila, "SUBTOTALES DE LA CASCADA (año actual)", 5)
        fila += 1
        for etiqueta, valor in subtotales:
            _celda_texto(ws, fila, 1, "", font=FONT_TOTAL, fill=FILL_TOTAL,
                         borde=BORDE_TOTAL)
            _celda_texto(ws, fila, 2, etiqueta, font=FONT_TOTAL, fill=FILL_TOTAL,
                         borde=BORDE_TOTAL)
            _celda_texto(ws, fila, 3, "", font=FONT_TOTAL, fill=FILL_TOTAL,
                         borde=BORDE_TOTAL)
            _celda_num(ws, fila, 4, valor, font=FONT_TOTAL, fill=FILL_TOTAL,
                       borde=BORDE_TOTAL)
            _celda_texto(ws, fila, 5, "", font=FONT_TOTAL, fill=FILL_TOTAL,
                         borde=BORDE_TOTAL)
            fila += 1


def _hoja_flujo(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("Flujo de Efectivo")
    ws.sheet_view.showGridLines = False
    inicio = _titulo_hoja(ws, "Estado de Flujo de Efectivo — método indirecto", 2)
    _encabezados(ws, inicio, ["Concepto", "Valor"], [46, 20])
    fila = inicio + 1
    f = ctx["flujo"]
    filas = [
        ("Flujo de actividades de OPERACIÓN", f["operacion"], False),
        ("Flujo de actividades de INVERSIÓN", f["inversion"], False),
        ("Flujo de actividades de FINANCIAMIENTO", f["financiamiento"], False),
        ("Incremento neto de efectivo", f["incremento_neto"], True),
        ("Efectivo al inicio del período", f["efectivo_inicial"], False),
        ("Efectivo al final (calculado)", f["efectivo_final_calculado"], True),
        ("Efectivo al final (real balance)", f["efectivo_final_real"], False),
        ("AF — cuadre (calc. − real, debe ser 0.00)", f["cuadre_af"], True),
    ]
    for etiqueta, valor, total in filas:
        font = FONT_TOTAL if total else FONT_DATA
        fill = FILL_TOTAL if total else None
        borde = BORDE_TOTAL if total else BORDE_THIN
        _celda_texto(ws, fila, 1, etiqueta, font=font, fill=fill, borde=borde)
        _celda_num(ws, fila, 2, valor, font=font, fill=fill, borde=borde)
        fila += 1


def _hoja_patrimonio(wb: Workbook, ctx: dict):
    """Estado de Evolución del Patrimonio — matriz oficial 99xx (16 movimientos ×
    18 componentes + TOTAL), con encabezados agrupados. Reproduce la hoja
    ``Estado de Evolucion del Patrimo`` del modelo (validada 288/288 celdas)."""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Evolución del Patrimonio")
    ws.sheet_view.showGridLines = False
    mat = ctx["patrimonio_matriz"]
    cols = mat["columnas"]       # 18 componentes [{codigo,nombre}]
    grupos = mat["grupos"]       # [{nombre, cols:[codigo...]}]
    filas = mat["filas"]         # 16 movimientos
    ncomp = len(cols)
    col_total = 3 + ncomp        # columna del TOTAL

    hr1 = _titulo_hoja(ws, "Estado de Evolución del Patrimonio (matriz oficial 99xx)",
                       col_total)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 34
    for i in range(ncomp):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13
    ws.column_dimensions[get_column_letter(col_total)].width = 16

    # Encabezado fila 1: Cód. | Movimiento | [grupos] | TOTAL PATRIMONIO
    _celda_texto(ws, hr1, 1, "Cód.", al=AL_CEN, font=FONT_TOTAL, fill=FILL_TOTAL)
    _celda_texto(ws, hr1, 2, "Movimiento", font=FONT_TOTAL, fill=FILL_TOTAL)
    c = 3
    for g in grupos:
        n = len(g["cols"])
        if n > 1:
            ws.merge_cells(start_row=hr1, start_column=c, end_row=hr1, end_column=c + n - 1)
        _celda_texto(ws, hr1, c, g.get("nombre") or " ", al=AL_CEN,
                     font=FONT_TOTAL, fill=FILL_TOTAL)
        for cc in range(c, c + n):
            celda = ws.cell(row=hr1, column=cc)
            celda.fill = FILL_TOTAL
            celda.border = BORDE_THIN
        c += n
    _celda_texto(ws, hr1, col_total, "TOTAL PATRIMONIO", al=AL_CEN,
                 font=FONT_TOTAL, fill=FILL_TOTAL)

    # Encabezado fila 2: códigos de componente
    hr2 = hr1 + 1
    _celda_texto(ws, hr2, 1, " ", fill=FILL_TOTAL)
    _celda_texto(ws, hr2, 2, " ", fill=FILL_TOTAL)
    for i, comp in enumerate(cols):
        _celda_texto(ws, hr2, 3 + i, comp["codigo"], al=AL_CEN, font=FONT_DATA,
                     fill=FILL_TOTAL)
    _celda_texto(ws, hr2, col_total, " ", fill=FILL_TOTAL)

    # Filas de movimiento
    secciones = {"99", "9901", "9902"}
    fila = hr2 + 1
    for f in filas:
        es_sec = f["codigo"] in secciones
        font = FONT_TOTAL if es_sec else FONT_DATA
        fill = FILL_TOTAL if es_sec else None
        _celda_texto(ws, fila, 1, f["codigo"], al=AL_CEN, font=font, fill=fill)
        _celda_texto(ws, fila, 2, f["nombre"], font=font, fill=fill)
        for i, comp in enumerate(cols):
            _celda_num(ws, fila, 3 + i, f["celdas"].get(comp["codigo"], 0.0),
                       font=font, fill=fill)
        _celda_num(ws, fila, col_total, f["celdas"].get("total", 0.0),
                   font=FONT_TOTAL, fill=(fill or FILL_TOTAL))
        fila += 1


def _hoja_no_efectivo(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("Movimiento no Efectivo")
    ws.sheet_view.showGridLines = False
    inicio = _titulo_hoja(ws, "Movimientos que no son efectivo (add-backs)", 2)
    ne = ctx["no_efectivo"]
    _encabezados(ws, inicio, ["Categoría", "Monto"], [40, 20])
    fila = inicio + 1
    for cat in motor_no_efectivo.CATEGORIAS:
        _celda_texto(ws, fila, 1, cat)
        _celda_num(ws, fila, 2, ne.get(cat, 0.0))
        fila += 1
    # otras categorías fuera del set canónico
    for k, v in ne.items():
        if k in motor_no_efectivo.CATEGORIAS or k in ("total", "detalle"):
            continue
        _celda_texto(ws, fila, 1, str(k))
        _celda_num(ws, fila, 2, v)
        fila += 1
    _celda_texto(ws, fila, 1, "TOTAL no efectivo", font=FONT_TOTAL,
                 fill=FILL_TOTAL, borde=BORDE_TOTAL)
    _celda_num(ws, fila, 2, ne.get("total", 0.0), font=FONT_TOTAL,
               fill=FILL_TOTAL, borde=BORDE_TOTAL)
    fila += 2

    detalle = ne.get("detalle", {})
    if detalle:
        _bloque(ws, fila, "DETALLE POR CÓDIGO ERI", 2)
        fila += 1
        _encabezados(ws, fila, ["Código ERI", "Monto"], [40, 20])
        fila += 1
        for cod in sorted(detalle):
            _celda_texto(ws, fila, 1, cod, al=AL_CEN)
            _celda_num(ws, fila, 2, detalle[cod])
            fila += 1


def _hoja_f101(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("Formulario 101")
    ws.sheet_view.showGridLines = False
    inicio = _titulo_hoja(ws, "Formulario 101 — casilleros (año actual)", 2)
    _encabezados(ws, inicio, ["Casillero", "Valor"], [18, 22])
    fila = inicio + 1
    cas = ctx["f101"]
    for cod in sorted(cas, key=lambda c: int(c) if str(c).isdigit() else 0):
        valor = cas[cod]
        if round(float(valor), 2) == 0.0:
            continue
        _celda_texto(ws, fila, 1, str(cod), al=AL_CEN)
        _celda_num(ws, fila, 2, valor)
        fila += 1


# Cuadro de mando de indicadores: (categoría, [(clave, etiqueta, fórmula, tipo,
# graficable)...]). Los graficables van primero en cada categoría (bloque
# contiguo) para que el gráfico nativo tome un rango de igual escala.
_IND_DASHBOARD = [
    ("LIQUIDEZ", [
        ("razon_corriente", "Razón corriente", "Act. corriente / Pas. corriente", "ratio", True),
        ("prueba_acida", "Prueba ácida", "(Act. corriente − Inventarios) / Pas. corriente", "ratio", True),
        ("capital_trabajo", "Capital de trabajo", "Act. corriente − Pas. corriente", "monto", False),
    ]),
    ("ACTIVIDAD", [
        ("dias_cartera", "Días cartera", "Cuentas por cobrar / Ventas × 365", "dias", True),
        ("dias_inventario", "Días inventario", "Inventario prom. / Costo × 365", "dias", True),
        ("dias_proveedores", "Días proveedores", "Cuentas por pagar / Costo × 365", "dias", True),
        ("ciclo_efectivo", "Ciclo de efectivo", "Cartera + Inventario − Proveedores", "dias", True),
        ("eficiencia_activos", "Eficiencia de activos", "Ventas netas / Activos", "ratio", False),
    ]),
    ("ENDEUDAMIENTO", [
        ("endeudamiento_total", "Endeudamiento total", "Pasivos / Activos", "pct", True),
        ("endeudamiento_lp", "Endeudamiento a largo plazo", "Pasivo no corriente / Activos", "pct", True),
        ("endeudamiento_patrimonial", "Rel. endeudamiento patrimonial", "Pasivo / Patrimonio", "ratio", True),
        ("apalancamiento", "Apalancamiento", "Activos / Patrimonio", "ratio", True),
        ("endeudamiento_financiero", "Endeudamiento financiero", "Pasivo financiero / Patrimonio", "ratio", False),
    ]),
    ("RENTABILIDAD", [
        ("roi", "ROI", "Utilidad operativa / Activos", "pct", True),
        ("margen_operativo", "Margen operativo", "Utilidad operativa / Ventas", "pct", True),
        ("roe", "ROE", "Utilidad neta / Patrimonio", "pct", True),
        ("margen_neto", "Margen neto", "Utilidad neta / Ventas", "pct", True),
        ("roa", "ROA", "Utilidad neta / Activos", "pct", True),
    ]),
    ("ADICIONALES", [
        ("ebit", "EBIT", "Utilidad operativa", "monto", True),
        ("ebitda", "EBITDA", "EBIT + Depreciación + Amortización", "monto", True),
    ]),
]
_IND_FMT = {"monto": NUM_FMT, "ratio": "0.0000", "pct": "0.00%", "dias": "0.0"}


def _hoja_indicadores(wb: Workbook, ctx: dict):
    """Cuadro de mando de indicadores financieros — tabla por categoría con
    columnas del año actual y anterior + fórmula, y un gráfico de barras nativo
    (openpyxl) por categoría comparando ambos años. Los gráficos son nativos de
    Excel: el archivo abre sin reparación al descargarlo."""
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("Indicadores")
    ws.sheet_view.showGridLines = False
    fila = _titulo_hoja(ws, "Cuadro de mando — Indicadores financieros", 4)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    ind = ctx["indicadores"]
    ind_ant = ctx.get("indicadores_ant", {})

    for categoria, filas_cat in _IND_DASHBOARD:
        cat_start = fila
        _bloque(ws, fila, categoria, 4)
        fila += 1
        # Encabezado de columnas
        _celda_texto(ws, fila, 1, "Indicador", font=FONT_TOTAL, fill=FILL_TOTAL)
        _celda_texto(ws, fila, 2, "Fórmula", font=FONT_TOTAL, fill=FILL_TOTAL)
        _celda_texto(ws, fila, 3, "Año actual", al=AL_DER, font=FONT_TOTAL, fill=FILL_TOTAL)
        _celda_texto(ws, fila, 4, "Año anterior", al=AL_DER, font=FONT_TOTAL, fill=FILL_TOTAL)
        hdr = fila
        fila += 1
        primera_dato = fila
        n_graf = 0
        contando = True
        for clave, etiqueta, formula, tipo, graf in filas_cat:
            _celda_texto(ws, fila, 1, etiqueta)
            _celda_texto(ws, fila, 2, formula)
            for col, dic in ((3, ind), (4, ind_ant)):
                c = ws.cell(row=fila, column=col, value=round(float(dic.get(clave, 0.0)), 4))
                c.number_format = _IND_FMT[tipo]
                c.font = FONT_DATA
                c.alignment = AL_DER
                c.border = BORDE_THIN
            if contando and graf:
                n_graf += 1
            else:
                contando = False
            fila += 1

        # Gráfico nativo de la categoría (bloque graficable contiguo)
        if n_graf >= 1:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = categoria
            chart.height = 6.2
            chart.width = 13
            chart.gapWidth = 60
            data = Reference(ws, min_col=3, max_col=4, min_row=hdr,
                             max_row=primera_dato + n_graf - 1)
            cats = Reference(ws, min_col=1, max_col=1, min_row=primera_dato,
                             max_row=primera_dato + n_graf - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.legend.position = "b"
            ws.add_chart(chart, "F%d" % cat_start)

        fila += 1  # separación entre categorías


def _hoja_er_esf(wb: Workbook, ctx: dict):
    """Balance resumido — Estado de Resultados + ESF condensados (año actual y
    anterior), replicando la hoja "ER y ESF" del modelo."""
    ws = wb.create_sheet("Balance resumido")
    ws.sheet_view.showGridLines = False
    fila = _titulo_hoja(ws, "Estado de Resultados y Balance — resumido", 3)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    res = ctx["resumen"]

    for titulo, seccion in (("ESTADO DE RESULTADOS", res["er"]),
                            ("ESTADO DE SITUACIÓN FINANCIERA", res["esf"])):
        fila += 1
        _bloque(ws, fila, titulo, 3)
        fila += 1
        _celda_texto(ws, fila, 1, "", fill=FILL_TOTAL)
        _celda_texto(ws, fila, 2, "Año actual", al=AL_DER, font=FONT_TOTAL, fill=FILL_TOTAL)
        _celda_texto(ws, fila, 3, "Año anterior", al=AL_DER, font=FONT_TOTAL, fill=FILL_TOTAL)
        fila += 1
        for f in seccion:
            es_tot = f["es_total"]
            font = FONT_TOTAL if es_tot else FONT_DATA
            fill = FILL_TOTAL if es_tot else None
            borde = BORDE_TOTAL if es_tot else BORDE_THIN
            _celda_texto(ws, fila, 1, f["concepto"], font=font, fill=fill, borde=borde)
            _celda_num(ws, fila, 2, f["act"], font=font, fill=fill, borde=borde)
            _celda_num(ws, fila, 3, f["ant"], font=font, fill=fill, borde=borde)
            fila += 1
        fila += 1


def _hoja_notas(wb: Workbook, ctx: dict):
    """Notas a los Estados Financieros — desglose por rubro (ESF y ERI).

    Por cada rubro con saldo: encabezado (código + nombre + años), las cuentas
    de detalle con saldo anterior/actual, y una fila de subtotal del rubro.
    """
    ws = wb.create_sheet("Notas")
    ws.sheet_view.showGridLines = False
    fila = _titulo_hoja(ws, "Notas a los Estados Financieros — desglose por rubro", 4)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    notas = ctx["notas"]
    anio_ant = ctx.get("anio_anterior", "Año anterior")
    anio_act = ctx.get("anio_actual", "Año actual")

    for titulo_bloque, grupo in (("ESTADO DE SITUACIÓN FINANCIERA", notas["esf"]),
                                 ("ESTADO DE RESULTADOS INTEGRAL", notas["eri"])):
        if not grupo:
            continue
        fila += 1
        _bloque(ws, fila, titulo_bloque, 4)
        fila += 2
        for nt in grupo:
            # Encabezado del rubro: código + nombre + años
            _celda_texto(ws, fila, 1, nt["codigo"], al=AL_CEN, font=FONT_TOTAL, fill=FILL_TOTAL)
            _celda_texto(ws, fila, 2, nt["nombre"], font=FONT_TOTAL, fill=FILL_TOTAL)
            _celda_texto(ws, fila, 3, str(anio_ant), al=AL_CEN, font=FONT_TOTAL, fill=FILL_TOTAL)
            _celda_texto(ws, fila, 4, str(anio_act), al=AL_CEN, font=FONT_TOTAL, fill=FILL_TOTAL)
            fila += 1
            # Cuentas de detalle
            for f in nt["filas"]:
                _celda_texto(ws, fila, 1, f["codigo"], al=AL_CEN)
                _celda_texto(ws, fila, 2, f["nombre"])
                _celda_num(ws, fila, 3, f["ant"])
                _celda_num(ws, fila, 4, f["act"])
                fila += 1
            # Subtotal del rubro
            _celda_texto(ws, fila, 1, "", font=FONT_TOTAL, fill=FILL_TOTAL, borde=BORDE_TOTAL)
            _celda_texto(ws, fila, 2, "Subtotal " + nt["nombre"], font=FONT_TOTAL,
                         fill=FILL_TOTAL, borde=BORDE_TOTAL)
            _celda_num(ws, fila, 3, nt["total_ant"], font=FONT_TOTAL, fill=FILL_TOTAL,
                       borde=BORDE_TOTAL)
            _celda_num(ws, fila, 4, nt["total_act"], font=FONT_TOTAL, fill=FILL_TOTAL,
                       borde=BORDE_TOTAL)
            fila += 2  # fila en blanco entre rubros


# ----------------------------------------------------------------------------
# API pública
# ----------------------------------------------------------------------------
def generar_excel(balanza_anterior: list[dict], balanza_actual: list[dict]) -> bytes:
    """Corre los 8 motores y devuelve los bytes de un ``.xlsx`` con 9 hojas.

    Args:
        balanza_anterior: filas ``{"cuenta","super_cias","sri","saldo"}`` del
            año anterior (salida de ``parser.parse_balanza``).
        balanza_actual: idem para el año actual.

    Returns:
        Los bytes del libro Excel (openpyxl), listo para escribir a disco o
        enviar al cliente.
    """
    # --- Catálogos (se cargan una vez) ---
    est_esf = catalogos.cargar_estructura("esf")
    est_eri = catalogos.cargar_estructura("eri")
    clasificacion = catalogos.cargar_clasificacion_flujo()
    agregados = catalogos.cargar_agregados_f101()
    cat_no_efectivo = catalogos.cargar_no_efectivo()

    # --- Homologación (SUMIF por Super Cías) ---
    saldos_ant, _sin_ant = motor.homologar_balanza(balanza_anterior)
    saldos_act, _sin_act = motor.homologar_balanza(balanza_actual)

    # --- Totales por código con rollup ---
    tot_esf_ant = motor.totales_por_codigo(est_esf, saldos_ant)
    tot_esf_act = motor.totales_por_codigo(est_esf, saldos_act)
    tot_eri_ant = motor.totales_por_codigo(est_eri, saldos_ant)
    tot_eri_act = motor.totales_por_codigo(est_eri, saldos_act)

    # --- Motores ---
    cuadre_esf = motor.cuadre(tot_esf_act)
    flujo = motor_flujo.flujo_efectivo(tot_esf_ant, tot_esf_act, clasificacion)
    cascada = motor_er.cascada_resultados(tot_eri_act)
    patrimonio = motor_patrimonio.evolucion(tot_esf_ant, tot_esf_act)
    no_efectivo = motor_no_efectivo.gastos_no_efectivo(tot_eri_act, cat_no_efectivo)
    notas = motor_notas.notas_estados(est_esf, est_eri, tot_esf_ant, tot_esf_act,
                                      tot_eri_ant, tot_eri_act)
    ori885 = motor_f101.ori_del_periodo(balanza_anterior, balanza_actual)
    f101 = motor_f101.casilleros_completos(balanza_actual, agregados,
                                           extras={"885": ori885})
    resumen = motor_resumen.balance_resumido(
        tot_esf_ant, tot_esf_act, tot_eri_ant, tot_eri_act)
    indicadores = motor_indicadores.indicadores(
        tot_esf_act, cascada, resumen=resumen, no_efectivo=no_efectivo,
        tot_esf_ant=tot_esf_ant, anio="act")
    cascada_ant = motor_er.cascada_resultados(tot_eri_ant)
    no_efectivo_ant = motor_no_efectivo.gastos_no_efectivo(tot_eri_ant, cat_no_efectivo)
    indicadores_ant = motor_indicadores.indicadores(
        tot_esf_ant, cascada_ant, resumen=resumen, no_efectivo=no_efectivo_ant,
        anio="ant")

    # Resultado integral REAL = utilidad neta + ORI del período. El ORI viene de
    # la reclasificación actuarial (motor_f101.ori_del_periodo), NO del código 800
    # del ERI (que suele venir en 0). Sin este ajuste el resumen mostraba el
    # resultado integral igual a la utilidad neta (bug detectado 2026-07-11).
    resultado_integral = round(cascada["utilidad_neta"] + ori885, 2)

    ctx = {
        "bal_ant": balanza_anterior,
        "bal_act": balanza_actual,
        "cuadre_esf": cuadre_esf,
        "flujo": flujo,
        "cascada": cascada,
        "ori": ori885,
        "resultado_integral": resultado_integral,
        "patrimonio": patrimonio,
        "patrimonio_matriz": patrimonio_matriz.matriz_patrimonio(
            balanza_anterior, balanza_actual),
        "no_efectivo": no_efectivo,
        "notas": notas,
        "resumen": resumen,
        "f101": f101,
        "indicadores": indicadores,
        "indicadores_ant": indicadores_ant,
    }

    # --- Construcción del libro (orden de hojas exigido) ---
    wb = Workbook()
    wb.remove(wb.active)  # elimina la hoja por defecto

    _hoja_resumen(wb, ctx)
    _hoja_homologacion(wb, ctx)
    _hoja_estructura(wb, "ESF", "Estado de Situación Financiera", est_esf,
                     tot_esf_ant, tot_esf_act)
    subtotales_eri = [
        ("Ganancia bruta", cascada["ganancia_bruta"]),
        ("Utilidad operativa", cascada["utilidad_operativa"]),
        ("Utilidad antes de impuestos", cascada["utilidad_antes_ir"]),
        ("Utilidad neta", cascada["utilidad_neta"]),
        ("Resultado integral", resultado_integral),
    ]
    _hoja_estructura(wb, "ERI", "Estado de Resultados Integral", est_eri,
                     tot_eri_ant, tot_eri_act, subtotales=subtotales_eri)
    _hoja_flujo(wb, ctx)
    _hoja_patrimonio(wb, ctx)
    _hoja_no_efectivo(wb, ctx)
    _hoja_er_esf(wb, ctx)
    _hoja_f101(wb, ctx)
    _hoja_notas(wb, ctx)
    _hoja_indicadores(wb, ctx)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
