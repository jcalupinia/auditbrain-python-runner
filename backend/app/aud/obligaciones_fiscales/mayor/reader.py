"""Lectura de un Mayor General en Excel, agnóstica del ERP de origen.

El formato varía por cliente, así que el encabezado y las columnas se
autodetectan por sinónimos. Si no se logra el mapeo mínimo, el resultado lo
reporta en lugar de fallar: el auditor mapea las columnas a mano.
"""

from __future__ import annotations

import datetime
import unicodedata
from io import BytesIO

from openpyxl import load_workbook

from backend.app.aud.obligaciones_fiscales.cedulas.base import _parse_amount_sri
from backend.app.aud.obligaciones_fiscales.mayor.tipos import (
    COLUMNAS_MINIMAS,
    LecturaMayor,
    Movimiento,
)

# Sinónimos por campo. Se compara por igualdad exacta sobre el encabezado
# normalizado; el "contiene" sólo se usa como respaldo y NUNCA para 'cuenta'
# (porque 'Persona Cruce Cuenta' lo capturaría por error).
SINONIMOS: dict[str, tuple[str, ...]] = {
    # 'cuenta' va AL FINAL: es un sinónimo débil de código (solo aparece en
    # encabezados de dos columnas tipo 'Cuenta | Nombre', donde 'Cuenta' es
    # el código). Cuando el encabezado real ya tiene una columna 'Cuenta'
    # propia (el nombre de la cuenta), esa se resuelve primero por su
    # propio código exacto ('cuenta contable', 'cta', ...) antes de llegar
    # a esta columna, así que el orden no la roba.
    "codigo": ("codigo", "cod", "cod cuenta", "codigo cuenta", "cuenta contable",
               "nro cuenta", "numero de cuenta", "cta", "cuenta"),
    "cuenta": ("cuenta", "nombre", "nombre cuenta", "nombre de cuenta",
               "descripcion cuenta", "detalle cuenta"),
    "fecha": ("fecha", "fecha asiento", "fecha movimiento", "f asiento"),
    "asiento": ("asiento", "comprobante", "nro asiento", "numero asiento",
                "no asiento", "diario", "nro comprobante"),
    "documento": ("documento", "doc", "nro documento", "comprobante venta",
                  "factura"),
    "identificacion": ("identificacion", "ruc", "cedula", "ruc cedula", "nit",
                       "identificacion tercero"),
    "persona": ("persona", "razon social", "tercero", "proveedor", "cliente",
                "beneficiario", "nombre tercero"),
    "descripcion": ("descripcion", "glosa", "detalle", "concepto",
                    "observacion", "observaciones"),
    "debe": ("debe", "debito", "cargo", "debitos"),
    "haber": ("haber", "credito", "abono", "creditos"),
    "saldo": ("saldo", "saldo final", "saldo acumulado"),
}

# Campos donde el respaldo por "contiene" sería peligroso.
SOLO_EXACTO = frozenset({"cuenta", "saldo"})

MAX_FILAS_BUSQUEDA_ENCABEZADO = 30

# Prefijos (normalizados) que delatan una fila de acumulado por cuenta
# ("TOTAL CUENTA", "SALDO ANTERIOR", ...) en vez de un movimiento real.
_PREFIJOS_FILA_ACUMULADO = (
    "total", "suma", "subtotal", "saldo anterior", "saldo inicial",
)

# Cuantas entradas de "importe no parseable" se listan en errores como
# máximo, para no inundar el reporte de lectura.
MAX_ERRORES_IMPORTES_NO_PARSEABLES = 10


def _norm(valor) -> str:
    """minúsculas, sin tildes, sin puntuación, espacios colapsados."""
    if valor is None:
        return ""
    s = unicodedata.normalize("NFKD", str(valor))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = "".join(c if c.isalnum() else " " for c in s.lower())
    return " ".join(s.split())


def _mapear_encabezado(celdas: list) -> dict[str, int]:
    """Devuelve {campo: índice de columna} para una fila candidata."""
    normalizadas = [_norm(c) for c in celdas]
    mapeo: dict[str, int] = {}
    usadas: set[int] = set()

    for campo, opciones in SINONIMOS.items():
        for i, texto in enumerate(normalizadas):
            if i in usadas or not texto:
                continue
            if texto in opciones:
                mapeo[campo] = i
                usadas.add(i)
                break

    for campo, opciones in SINONIMOS.items():
        if campo in mapeo or campo in SOLO_EXACTO:
            continue
        for i, texto in enumerate(normalizadas):
            if i in usadas or not texto:
                continue
            if any(texto.startswith(o) for o in opciones):
                mapeo[campo] = i
                usadas.add(i)
                break

    # Último recurso (defecto 6): si 'cuenta' (nombre de la cuenta) no se
    # resolvió por NINGÚN sinónimo propio pero sí hay una columna
    # 'descripcion' mapeada, esa columna es probablemente el nombre de la
    # cuenta (ERP sin columna 'Nombre' separada, ej.
    # 'Cta | Descripción | Debe | Haber'). Solo cede cuando 'cuenta' no se
    # resolvió de ninguna otra forma, para no confundir la glosa del
    # movimiento del encabezado real de 12 columnas (donde 'Cuenta' ya se
    # mapea por su propio sinónimo antes de llegar a 'Descripción').
    if "cuenta" not in mapeo and "descripcion" in mapeo:
        mapeo["cuenta"] = mapeo.pop("descripcion")

    return mapeo


def _fecha(valor) -> datetime.date | None:
    if isinstance(valor, datetime.datetime):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor
    if valor is None:
        return None
    texto = str(valor).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def _texto(valor) -> str:
    return "" if valor is None else str(valor).strip()


def _limpiar_importe(texto: str) -> str:
    """Normaliza formatos habituales de exportaciones de ERP antes de
    delegar el parseo a `_parse_amount_sri`: quita símbolos de moneda
    ('$', 'USD') y espacios, convierte el negativo contable entre
    paréntesis ('(150.00)') en '-150.00', y trata un guion solo ('-') o
    una celda vacía como cero.
    """
    t = (texto or "").strip()
    if not t:
        return t
    for simbolo in ("$", "USD", "usd"):
        t = t.replace(simbolo, "")
    t = t.strip()
    negativo = t.startswith("(") and t.endswith(")")
    if negativo:
        t = t[1:-1].strip()
    t = t.replace(" ", "")
    if t in ("-", ""):
        return "0"
    if negativo and not t.startswith("-"):
        t = "-" + t
    return t


def _importe(valor, *, lectura: LecturaMayor, fila_num: int, campo: str) -> float:
    """Convierte una celda de importe a float.

    Si tras limpiar el texto sigue sin poder parsearse, NO desaparece en
    silencio como 0.00: se cuenta en `lectura.importes_no_parseables` y se
    deja rastro (fila y texto original) en `lectura.errores`, hasta un
    máximo de entradas para no inundar el reporte.
    """
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = _texto(valor)
    limpio = _limpiar_importe(texto)
    if not limpio:
        return 0.0
    parsed = _parse_amount_sri(limpio)
    if parsed is not None:
        return parsed
    lectura.importes_no_parseables += 1
    if len(lectura.errores) < MAX_ERRORES_IMPORTES_NO_PARSEABLES:
        lectura.errores.append(
            f"Fila {fila_num}: importe no parseable en '{campo}': {texto!r}"
        )
    return 0.0


def _es_fila_acumulado(cuenta: str, descripcion: str) -> bool:
    """Filas de TOTAL/SUBTOTAL/SALDO que un ERP emite al cierre de cada
    cuenta: llevan código pero no son un movimiento (defecto 2). El texto
    delator puede venir en el nombre de la cuenta o en la glosa/descripción.
    """
    for texto in (cuenta, descripcion):
        n = _norm(texto)
        if any(n.startswith(p) for p in _PREFIJOS_FILA_ACUMULADO):
            return True
    return False


def _leer_hoja(ws, mapeo: dict[str, int], fila_encabezado: int, lectura: LecturaMayor) -> None:
    """Lee los movimientos de UNA hoja ya mapeada y los agrega a `lectura`."""
    col = mapeo

    def celda(fila, campo):
        i = col.get(campo)
        return fila[i] if i is not None and i < len(fila) else None

    for n, fila in enumerate(
        ws.iter_rows(min_row=fila_encabezado + 1, values_only=True),
        start=fila_encabezado + 1,
    ):
        codigo = _texto(celda(fila, "codigo"))
        if not codigo:
            lectura.filas_descartadas += 1
            continue
        if _norm(codigo) in SINONIMOS["codigo"]:
            # Encabezado repetido a mitad del listado (paginación del ERP).
            lectura.filas_descartadas += 1
            continue

        fecha = _fecha(celda(fila, "fecha"))
        asiento = _texto(celda(fila, "asiento"))
        cuenta = _texto(celda(fila, "cuenta"))
        descripcion = _texto(celda(fila, "descripcion"))

        if not fecha and not asiento and _es_fila_acumulado(cuenta, descripcion):
            # Fila de TOTAL/SUBTOTAL/SALDO ANTERIOR/INICIAL: son los
            # acumulados de la cuenta, no un movimiento; si se cargara
            # duplicaría exactamente el debe y el haber de la cuenta.
            lectura.filas_descartadas += 1
            continue

        lectura.movimientos.append(
            Movimiento(
                codigo=codigo,
                cuenta=cuenta,
                fecha=fecha,
                asiento=asiento,
                documento=_texto(celda(fila, "documento")),
                identificacion=_texto(celda(fila, "identificacion")),
                persona=_texto(celda(fila, "persona")),
                descripcion=descripcion,
                debe=_importe(celda(fila, "debe"), lectura=lectura, fila_num=n, campo="debe"),
                haber=_importe(celda(fila, "haber"), lectura=lectura, fila_num=n, campo="haber"),
                saldo=_importe(celda(fila, "saldo"), lectura=lectura, fila_num=n, campo="saldo"),
                fila=n,
            )
        )


def leer_mayor(contenido: bytes) -> LecturaMayor:
    """Lee un mayor en .xlsx/.xlsm y devuelve sus movimientos normalizados.

    Si el mayor viene repartido en varias hojas con el mismo encabezado
    (habitual cuando el ERP exporta una hoja por mes o por tipo de
    comprobante), se leen TODAS las que alcancen el mapeo mínimo: elegir
    solo la de mejor puntaje perdería los movimientos de las demás sin
    avisar.
    """
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        return LecturaMayor(errores=[f"No se pudo abrir el Excel: {e}"])

    # Defecto 4: cualquier excepción durante el barrido de hojas o la
    # lectura de filas debe dejar el workbook (y su ZipFile/handles)
    # cerrado igual, para no acumular archivos abiertos en un backend de
    # larga vida. try/finally garantiza el cierre en TODOS los caminos,
    # incluida una excepción que se propaga hacia arriba.
    try:
        candidatos = []  # [(hoja, fila_encabezado, mapeo, puntaje), ...]
        for nombre in wb.sheetnames:
            ws = wb[nombre]
            mejor_de_hoja = None  # (fila, mapeo, puntaje)
            for fila_idx, fila in enumerate(
                ws.iter_rows(max_row=MAX_FILAS_BUSQUEDA_ENCABEZADO, values_only=True),
                start=1,
            ):
                mapeo = _mapear_encabezado(list(fila))
                puntaje = len(mapeo)
                if mejor_de_hoja is None or puntaje > mejor_de_hoja[2]:
                    mejor_de_hoja = (fila_idx, mapeo, puntaje)
            if mejor_de_hoja is not None:
                fila_idx, mapeo, puntaje = mejor_de_hoja
                candidatos.append((nombre, fila_idx, mapeo, puntaje))

        if not candidatos or max(c[3] for c in candidatos) == 0:
            return LecturaMayor(
                errores=["No se detectó una fila de encabezado reconocible."],
                columnas_faltantes=list(COLUMNAS_MINIMAS),
            )

        # Hoja de referencia para reportar columnas detectadas/faltantes
        # cuando NINGUNA hoja alcanza el mapeo mínimo (el auditor mapea a
        # mano).
        mejor = max(candidatos, key=lambda c: c[3])
        mapeo_mejor = mejor[2]
        columnas_faltantes = [c for c in COLUMNAS_MINIMAS if c not in mapeo_mejor]
        if columnas_faltantes:
            return LecturaMayor(
                columnas_detectadas=mapeo_mejor,
                columnas_faltantes=columnas_faltantes,
                hoja=mejor[0],
                fila_encabezado=mejor[1],
            )

        lectura = LecturaMayor()
        for nombre, fila_encabezado, mapeo, _puntaje in candidatos:
            if any(c not in mapeo for c in COLUMNAS_MINIMAS):
                continue  # esta hoja concreta no alcanza el mapeo mínimo
            if not lectura.hojas_leidas:
                lectura.hoja = nombre
                lectura.fila_encabezado = fila_encabezado
                lectura.columnas_detectadas = mapeo
            lectura.hojas_leidas.append(nombre)
            _leer_hoja(wb[nombre], mapeo, fila_encabezado, lectura)

        return lectura
    finally:
        wb.close()
