"""Arma el libro DM de Obligaciones Fiscales.

Orden de las hojas: primero el resumen del motor y su detalle, después los
casilleros declarados. Las cédulas DM3..DM7 leen de estas hojas POR FÓRMULA
usando los mapas de direcciones que este módulo recolecta.

Orden de CONSTRUCCIÓN obligatorio (no es el orden de las pestañas): DM4 →
DM5 → DM6 → DM7 → DM3, porque cada una referencia a la anterior por fórmula
(DM6 lee de DM4 y DM5; DM7 de DM5; DM3 de DM7). Construir en otro orden
significaría referenciar un mapa de direcciones que todavía no existe.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm3_saldos import build_dm3
from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm4_compras import build_dm4
from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm5_ventas import build_dm5
from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm6_iva import build_dm6
from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm7_retenciones import build_dm7
from backend.app.aud.obligaciones_fiscales.libro.fuentes import (
    construir_hojas_de_casilleros,
)
from backend.app.aud.obligaciones_fiscales.libro.hoja_detalle import build_hoja_detalle
from backend.app.aud.obligaciones_fiscales.libro.hoja_mayores import build_hoja_mayores

ORDEN_HOJAS = [
    "Mayores homologados", "Detalle mayor",
    "DM3 Revisión de saldos", "DM4 Compras", "DM5 Ventas", "DM6 IVA",
    "DM7 Retenciones x pagar",
    "DATOS F-104", "DATOS F-103",
]


def _periodos_del_ejercicio(f104_monthly: dict, f103_monthly: dict, periodo: str) -> list[str]:
    """Los 12 períodos "YYYY-MM" del ejercicio.

    Se prefieren las claves reales de las declaraciones subidas (así se
    respetan los meses efectivamente presentes); si no hay ninguna, se
    derivan del año del ejercicio para que las cédulas igual se construyan
    con la matriz completa (en cero) en vez de quedar vacías.
    """
    claves = set(f104_monthly or {}) | set(f103_monthly or {})
    if claves:
        return sorted(claves)
    anio = str(periodo or "").strip()
    if len(anio) == 4 and anio.isdigit():
        return [f"{anio}-{m:02d}" for m in range(1, 13)]
    return []


def armar_libro(
    *,
    clasificacion,
    movimientos,
    f104_monthly: dict,
    f103_monthly: dict,
    cliente: str = "",
    periodo: str = "",
    preparado_por: str | None = None,
    revisado_por: str | None = None,
) -> bytes:
    """Devuelve los bytes del libro DM con sus hojas de datos y sus cédulas."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    dir_mayores = build_hoja_mayores(wb, clasificacion)
    build_hoja_detalle(
        wb, movimientos, {f.codigo_cuenta: f.categoria_final for f in clasificacion}
    )
    hojas_datos = construir_hojas_de_casilleros(
        wb, f104_monthly=f104_monthly, f103_monthly=f103_monthly
    )
    dir_f104 = hojas_datos["f104"]
    dir_f103 = hojas_datos["f103"]

    periodos = _periodos_del_ejercicio(f104_monthly, f103_monthly, periodo)
    nombres_cuenta = {f.codigo_cuenta: f.nombre_cuenta for f in clasificacion}

    kwargs_comunes = dict(cliente=cliente, periodo=periodo,
                          preparado_por=preparado_por, revisado_por=revisado_por)

    dir_dm4 = build_dm4(
        wb, dir_mayores=dir_mayores, dir_f104=dir_f104, periodos=periodos,
        nombres_cuenta=nombres_cuenta, **kwargs_comunes,
    )
    dir_dm5 = build_dm5(
        wb, dir_mayores=dir_mayores, dir_f104=dir_f104, periodos=periodos,
        nombres_cuenta=nombres_cuenta, **kwargs_comunes,
    )
    build_dm6(
        wb, dir_dm5=dir_dm5, dir_dm4=dir_dm4, dir_mayores=dir_mayores,
        dir_f104=dir_f104, periodos=periodos, **kwargs_comunes,
    )
    dir_dm7 = build_dm7(
        wb, dir_mayores=dir_mayores, dir_f104=dir_f104, dir_f103=dir_f103,
        periodos=periodos, nombres_cuenta=nombres_cuenta, **kwargs_comunes,
    )
    build_dm3(
        wb, dir_mayores=dir_mayores, dir_f104=dir_f104, dir_dm7=dir_dm7,
        periodos=periodos, **kwargs_comunes,
    )

    orden = [h for h in ORDEN_HOJAS if h in wb.sheetnames]
    orden += [h for h in wb.sheetnames if h not in orden]
    wb._sheets = [wb[h] for h in orden]

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
