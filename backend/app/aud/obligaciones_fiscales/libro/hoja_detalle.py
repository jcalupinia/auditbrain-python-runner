"""Hoja 'Detalle mayor': cada movimiento con la categoría que se le asignó.

Una sola hoja con autofiltro, en vez de una hoja por categoría: así el
auditor rastrea cualquier subtotal del resumen hasta la factura que lo
originó sin saltar entre pestañas.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

SHEET_DETALLE = "Detalle mayor"
SIN_CLASIFICAR = "SIN_CLASIFICAR"

ENCABEZADO = [
    "Categoría", "Código", "Cuenta", "Fecha", "Asiento", "Documento",
    "Identificación", "Persona", "Descripción", "Debe", "Haber", "Saldo", "Mes",
]
ANCHOS = [16, 14, 34, 12, 20, 24, 16, 32, 34, 14, 14, 14, 8]

FONT_TITULO = Font(name="Calibri", size=11, bold=True)
FONT_ENCABEZADO = Font(name="Calibri", size=10, bold=True)
FONT_DATA = Font(name="Calibri", size=9)
FORMATO_NUM = "#,##0.00"
FILA_ENCABEZADO = 3


def build_hoja_detalle(wb: Workbook, movimientos, categorias: dict[str, str]) -> None:
    """Escribe todos los movimientos clasificados en una sola hoja filtrable."""
    if SHEET_DETALLE in wb.sheetnames:
        del wb[SHEET_DETALLE]
    ws = wb.create_sheet(SHEET_DETALLE)

    ws.cell(1, 1, "DETALLE DEL MAYOR · movimientos clasificados").font = FONT_TITULO

    for i, texto in enumerate(ENCABEZADO, start=1):
        c = ws.cell(FILA_ENCABEZADO, i, texto)
        c.font = FONT_ENCABEZADO
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    fila = FILA_ENCABEZADO + 1
    for m in movimientos:
        valores = [
            categorias.get(m.codigo, SIN_CLASIFICAR),
            m.codigo, m.cuenta, m.fecha, m.asiento, m.documento,
            m.identificacion, m.persona, m.descripcion,
            m.debe, m.haber, m.saldo, m.mes or "",
        ]
        for i, v in enumerate(valores, start=1):
            c = ws.cell(fila, i, v)
            c.font = FONT_DATA
            if i in (10, 11, 12):
                c.number_format = FORMATO_NUM
            if i == 4 and m.fecha:
                c.number_format = "yyyy-mm-dd"
        fila += 1

    ultima = max(fila - 1, FILA_ENCABEZADO + 1)
    ws.auto_filter.ref = (
        f"A{FILA_ENCABEZADO}:{get_column_letter(len(ENCABEZADO))}{ultima}"
    )
    ws.freeze_panes = f"A{FILA_ENCABEZADO + 1}"
    for i, ancho in enumerate(ANCHOS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
