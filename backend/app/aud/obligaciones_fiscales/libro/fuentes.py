"""Hojas de datos fuente: los casilleros declarados al SRI.

Se apoyan en los builders del ICT, que ya generan la matriz completa de
casilleros por mes y devuelven el mapa de direcciones que las cédulas usan
para referenciarlas POR FÓRMULA.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from backend.app.aud.obligaciones_fiscales.libro.ats import ResumenATS
from backend.app.aud.obligaciones_fiscales.libro.estilos import (
    BORDE, FONT_DATA, FONT_ENCABEZADO_TABLA, FORMATO_NUM,
)
from backend.app.ict.fillers.source_data_sheets import (
    build_f103_sheet,
    build_f104_sheet,
)

SHEET_ATS = "DATOS ATS"

# Porcentajes fijos de retención de IVA que reconoce el SRI, con su
# casillero de control del F-104 (721..731). "NC" (nota de crédito) no tiene
# casillero propio: es informativo.
_BUCKETS_RET_IVA = ("10", "20", "30", "50", "70", "100", "NC")

_CAMPOS_FIJOS = (
    ("compras_bi_0", "Compras · BI tarifa 0%"),
    ("compras_bi_gravada", "Compras · BI tarifa ≠0%"),
    ("compras_iva", "Compras · Valor IVA"),
    ("ventas_bi_0", "Ventas · BI tarifa 0%"),
    ("ventas_bi_gravada", "Ventas · BI tarifa ≠0%"),
    ("ventas_iva", "Ventas · Valor IVA"),
    ("anulados", "Comprobantes anulados (informativo)"),
    ("iva_le_retuvieron", "IVA que le retuvieron (cas. 609)"),
    ("renta_le_retuvieron", "Renta que le retuvieron (informativo)"),
    ("ret_renta_total", "Retenciones de renta · total según ATS"),
    ("ret_iva_total", "Retenciones de IVA · total según ATS"),
)


def _valor_campo_ats(resumen: ResumenATS | None, campo: str) -> float:
    """Resuelve el valor literal de un campo fijo, código de renta o
    porcentaje de retención de IVA para un resumen de un mes."""
    if resumen is None:
        return 0.0
    directos = {
        "compras_bi_0": resumen.compras.bi_0,
        "compras_bi_gravada": resumen.compras.bi_gravada,
        "compras_iva": resumen.compras.iva,
        "ventas_bi_0": resumen.ventas.bi_0,
        "ventas_bi_gravada": resumen.ventas.bi_gravada,
        "ventas_iva": resumen.ventas.iva,
        "anulados": float(resumen.comprobantes_anulados or 0),
        "iva_le_retuvieron": resumen.iva_que_le_retuvieron,
        "renta_le_retuvieron": resumen.renta_que_le_retuvieron,
        "ret_renta_total": resumen.retenciones_renta_valor_total,
        "ret_iva_total": resumen.retenciones_iva_total,
    }
    if campo in directos:
        return directos[campo]
    if campo.startswith("renta_codigo:"):
        codigo = campo.split(":", 1)[1]
        return sum(f.valor_retenido for f in resumen.retenciones_renta if f.codigo == codigo)
    if campo.startswith("iva_pct:"):
        pct = campo.split(":", 1)[1]
        if pct == "NC":
            return sum(f.valor_retenido for f in resumen.retenciones_iva if f.porcentaje is None)
        objetivo = float(pct)
        return sum(
            f.valor_retenido for f in resumen.retenciones_iva if f.porcentaje == objetivo
        )
    return 0.0


def construir_hoja_ats(
    wb: Workbook, resumenes: dict[str, ResumenATS]
) -> dict[tuple[str, str], str]:
    """Crea 'DATOS ATS': valores literales por mes que DM8 referencia por
    fórmula. Devuelve {(campo, "01".."12") → addr}.

    Los campos fijos (compras/ventas/anulados/le-efectuaron/totales) están
    siempre presentes. Los códigos de retención de renta son un catálogo
    abierto (varían por cliente): se listan los que aparezcan en cualquiera
    de los meses recibidos. Los porcentajes de retención de IVA son fijos
    (10/20/30/50/70/100/NC, catálogo del SRI).
    """
    if SHEET_ATS in wb.sheetnames:
        del wb[SHEET_ATS]
    ws = wb.create_sheet(SHEET_ATS)

    ws.cell(1, 1, "DATOS ATS · Talón Resumen del Anexo Transaccional").font = Font(
        name="Calibri", size=11, bold=True
    )

    meses = sorted(resumenes.keys()) if resumenes else [f"2025-{m:02d}" for m in range(1, 13)]

    codigos_renta = sorted({
        f.codigo for r in resumenes.values() for f in r.retenciones_renta
    })

    filas_campo: list[tuple[str, str]] = (
        list(_CAMPOS_FIJOS)
        + [(f"renta_codigo:{c}", f"Ret. renta {c}") for c in codigos_renta]
        + [
            (f"iva_pct:{pct}", f"Ret. IVA {pct}%" if pct != "NC" else "Ret. IVA NC")
            for pct in _BUCKETS_RET_IVA
        ]
    )

    encabezado = ["Campo"] + list(meses) + ["Total"]
    for i, texto in enumerate(encabezado, start=1):
        c = ws.cell(3, i, texto)
        c.font = FONT_ENCABEZADO_TABLA
        c.border = BORDE

    lookup: dict[tuple[str, str], str] = {}
    fila = 4
    for campo, etiqueta in filas_campo:
        ws.cell(fila, 1, etiqueta).font = FONT_DATA
        for j, periodo in enumerate(meses):
            col = 2 + j
            valor = _valor_campo_ats(resumenes.get(periodo), campo)
            c = ws.cell(fila, col, valor)
            c.font = FONT_DATA
            c.number_format = FORMATO_NUM
            c.border = BORDE
            mes = periodo.split("-")[-1]
            lookup[(campo, mes)] = f"'{SHEET_ATS}'!{get_column_letter(col)}{fila}"
        col_total = 2 + len(meses)
        if meses:
            ini, fin = get_column_letter(2), get_column_letter(col_total - 1)
            t = ws.cell(fila, col_total, f"=SUM({ini}{fila}:{fin}{fila})")
            t.font = FONT_DATA
            t.number_format = FORMATO_NUM
            t.border = BORDE
        fila += 1

    ws.column_dimensions["A"].width = 34
    for j in range(len(meses) + 1):
        ws.column_dimensions[get_column_letter(2 + j)].width = 14

    return lookup


def a_periodos_anuales(month_data: dict) -> dict:
    """{"01": {"periodo": "01/2025", ...}} → {"2025-01": {"casilleros": {...}}}.

    Los builders del ICT indexan por período completo; el extractor de F-104
    de esta herramienta indexa por mes. Los meses sin período detectado se
    descartan: sin año no se puede ubicar la columna.
    """
    salida: dict[str, dict] = {}
    for datos in (month_data or {}).values():
        periodo = (datos or {}).get("periodo")
        if not periodo or "/" not in str(periodo):
            continue
        mes, anio = str(periodo).split("/", 1)
        salida[f"{anio}-{int(mes):02d}"] = {"casilleros": datos.get("casilleros", {})}
    return salida


def construir_hojas_de_casilleros(
    wb: Workbook, *, f104_monthly: dict, f103_monthly: dict
) -> dict[str, dict]:
    """Crea DATOS F-104 y DATOS F-103. Devuelve {"f104": lookup, "f103": lookup}."""
    return {
        "f104": build_f104_sheet(wb, f104_monthly or {}),
        "f103": build_f103_sheet(wb, f103_monthly or {}),
    }


def leer_declaraciones(job_dir: Path) -> tuple[dict, dict]:
    """Lee los PDFs subidos del job y los deja en formato de períodos anuales."""
    from backend.app.aud.obligaciones_fiscales import file_storage
    from backend.app.aud.obligaciones_fiscales.cedulas.f104_extractor import (
        extract_all_f104,
    )
    from backend.app.ict.parsers.f103_pdf import parse_all_f103

    f104_mes, _ = extract_all_f104(file_storage.list_inputs(job_dir, "f104"))
    f103_monthly, _ = parse_all_f103(file_storage.list_inputs(job_dir, "f103"))
    return a_periodos_anuales(f104_mes), (f103_monthly or {})


def leer_ats(job_dir: Path) -> dict[str, ResumenATS]:
    """Lee los ATS (PDF o XML) subidos del job y los agrupa por período."""
    from backend.app.aud.obligaciones_fiscales import file_storage
    from backend.app.aud.obligaciones_fiscales.libro.ats import parse_all_ats

    por_periodo, _errores = parse_all_ats(file_storage.list_inputs(job_dir, "ats"))
    return por_periodo
