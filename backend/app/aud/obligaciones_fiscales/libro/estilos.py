"""Formato compartido por las cédulas del libro DM."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

FUENTE_MARCAS = "Arial"

FONT_TITULO_FIRMA = Font(name="Arial", size=12, bold=True)
FONT_TITULO_CEDULA = Font(name="Arial", size=11, bold=True)
FONT_ETIQUETA = Font(name="Arial", size=10)
FONT_DATO = Font(name="Arial", size=10, bold=True)
FONT_ENCABEZADO_TABLA = Font(name="Calibri", size=10, bold=True)
FONT_DATA = Font(name="Calibri", size=9)
FONT_TOTAL = Font(name="Calibri", size=10, bold=True)

BORDE = Border(*[Side(style="thin")] * 4)
BORDE_TOTAL = Border(
    top=Side(style="double"), bottom=Side(style="double"),
    left=Side(style="thin"), right=Side(style="thin"),
)
RELLENO_TOTAL = PatternFill("solid", fgColor="DCE6F1")
FORMATO_NUM = "#,##0.00"

# Marcas de auditoría. Son fijas de la cédula: NO se calculan contra ninguna
# tolerancia. En el modelo del auditor la misma marca aparecía en Arial y en
# Wingdings; aquí se unifican.
MARCAS = {
    "verificado": ("ü", "Cotejado según libros"),
    "declarado": ("£", "Cotejado según formularios"),
    "diferencia": ("‡", "Diferencia determinada"),
    "sumado": ("Σ", "Sumado, totalizado"),
}


def escribir_encabezado_cedula(
    ws,
    *,
    titulo: str,
    referencia: str,
    cliente: str,
    periodo: str,
    preparado_por: str | None = None,
    revisado_por: str | None = None,
    fecha_corte=None,
) -> None:
    """Encabezado estándar (filas 1 a 10) de una cédula DM."""
    ws.cell(1, 1, "OBLIGACIONES FISCALES").font = FONT_TITULO_FIRMA
    ws.cell(3, 1, titulo).font = FONT_TITULO_CEDULA

    ws.cell(5, 1, "Nombre del cliente").font = FONT_ETIQUETA
    ws.cell(6, 1, cliente or "").font = FONT_DATO
    ws.cell(5, 4, "Periodo terminado").font = FONT_ETIQUETA
    ws.cell(6, 4, periodo or "").font = FONT_DATO

    ws.cell(7, 1, "Preparado por:").font = FONT_ETIQUETA
    ws.cell(8, 1, preparado_por or "").font = FONT_DATO
    ws.cell(7, 3, "Fecha:").font = FONT_ETIQUETA
    ws.cell(8, 3, fecha_corte or "").font = FONT_DATO
    ws.cell(7, 4, "Referencia").font = FONT_ETIQUETA
    ws.cell(8, 4, referencia).font = FONT_DATO

    ws.cell(9, 1, "Revisado por:").font = FONT_ETIQUETA
    ws.cell(10, 1, revisado_por or "").font = FONT_DATO


def escribir_leyenda_marcas(ws, *, fila: int) -> int:
    """Leyenda de las marcas al pie de la cédula. Devuelve la fila siguiente."""
    for i, (simbolo, texto) in enumerate(MARCAS.values()):
        c = ws.cell(fila + i, 1, simbolo)
        c.font = Font(name=FUENTE_MARCAS, size=10)
        ws.cell(fila + i, 2, texto).font = Font(name=FUENTE_MARCAS, size=9)
    return fila + len(MARCAS)
