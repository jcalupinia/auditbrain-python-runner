"""Hoja 'Mayores homologados': el resumen que produce el motor.

Es la ÚNICA fuente del "según libros" de todas las cédulas. Cada cédula lee
el subtotal de su categoría por fórmula, usando el mapa de direcciones que
esta función devuelve.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from backend.app.aud.obligaciones_fiscales.mayor.catalogo import CATEGORIAS
from backend.app.aud.obligaciones_fiscales.mayor.ventas_tarifa import (
    separar_ventas_por_tarifa,
)

SHEET_MAYORES = "Mayores homologados"
MESES = [f"{m:02d}" for m in range(1, 13)]
NOMBRES_MES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
               "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
SIN_CLASIFICAR = "SIN_CLASIFICAR"

FONT_TITULO = Font(name="Calibri", size=11, bold=True)
FONT_DATA = Font(name="Calibri", size=9)
FONT_TOTAL = Font(name="Calibri", size=10, bold=True)
BORDE = Border(*[Side(style="thin")] * 4)
RELLENO_TOTAL = PatternFill("solid", fgColor="DCE6F1")
FORMATO_NUM = "#,##0.00"

COL_CATEGORIA, COL_CODIGO, COL_NOMBRE = 1, 2, 3
COL_PRIMER_MES = 4
COL_TOTAL = COL_PRIMER_MES + 12

CAT_VENTAS = "VENTAS"
# Los tres tramos del desglose de ventas por tarifa, en el orden en que se
# escriben, con el título de su sub-bloque.
TRAMOS_VENTAS = (
    ("gravada", "Ventas gravadas (≠ 0%)"),
    ("cero", "Ventas 0%"),
    ("por_asignar", "Por asignar (asientos con tarifas mezcladas)"),
)


def _orden(codigo: str | None) -> int:
    cat = CATEGORIAS.get(codigo or "")
    return cat.orden if cat else 99


def _addr(letra: str, fila: int) -> str:
    """Dirección CALIFICADA con el nombre de esta hoja.

    Las cédulas publican estas direcciones tal cual (`f"={addr}"`), así que
    sin el prefijo Excel las resolvería contra la propia hoja de la cédula:
    ése fue el bug de las referencias circulares de DM5/DM7. El nombre lleva
    espacios, por eso va siempre entre comillas simples.
    """
    return f"'{SHEET_MAYORES}'!{letra}{fila}"


def _bloque_desglose_ventas(ws, *, fila: int, cuentas, desglose, nombres,
                            fila_subtotal_ventas: int,
                            lookup: dict) -> int:
    """Bloque aparte con las ventas separadas en gravadas / 0% / por asignar.

    Va DEBAJO de la tabla principal y no dentro de ella: si el desglose
    fuese tres columnas más de la fila de cada cuenta, el subtotal de la
    categoría (un SUM sobre el rango de filas) contaría cada venta cuatro
    veces. Como bloque separado, la tabla de arriba sigue siendo la misma
    que el auditor ya conoce y el desglose se cuadra contra ella con una
    fila de control explícita.

    Cada tramo agrupa TODAS las cuentas de venta, para que su subtotal sea
    un SUM sobre un rango contiguo (es lo que DM5 necesita) y para que el
    auditor lea de corrido "lo gravado", "lo 0%" y "lo que falta clasificar".
    """
    ws.cell(fila, COL_CATEGORIA,
            "DESGLOSE DE VENTAS POR TARIFA · separación por asiento contable"
            ).font = FONT_TITULO
    fila += 1
    ws.cell(fila, COL_CATEGORIA,
            "Cada asiento se reparte según su contrapartida de IVA en ventas. "
            "Lo que no se puede separar sin ambigüedad queda 'por asignar': "
            "no se prorratea, se revisa.").font = FONT_DATA
    fila += 2

    encabezado = ["Tramo", "Código", "Cuenta"] + NOMBRES_MES + ["Total"]
    for i, texto in enumerate(encabezado, start=1):
        c = ws.cell(fila, i, texto)
        c.font = FONT_TOTAL
        c.border = BORDE
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    fila += 1

    filas_subtotal: list[int] = []
    for tramo, titulo in TRAMOS_VENTAS:
        primera = fila
        for codigo in cuentas:
            por_mes = desglose.get(codigo, {}).get(tramo, {})
            ws.cell(fila, COL_CATEGORIA, titulo).font = FONT_DATA
            ws.cell(fila, COL_CODIGO, codigo).font = FONT_DATA
            ws.cell(fila, COL_NOMBRE, nombres.get(codigo, "")).font = FONT_DATA
            for j, mes in enumerate(MESES):
                c = ws.cell(fila, COL_PRIMER_MES + j, float(por_mes.get(mes, 0.0)))
                c.font = FONT_DATA
                c.number_format = FORMATO_NUM
                c.border = BORDE
                lookup[(f"cuenta:{codigo}:{tramo}", mes)] = _addr(
                    get_column_letter(COL_PRIMER_MES + j), fila
                )
            ini = get_column_letter(COL_PRIMER_MES)
            fin = get_column_letter(COL_PRIMER_MES + 11)
            t = ws.cell(fila, COL_TOTAL, f"=SUM({ini}{fila}:{fin}{fila})")
            t.font = FONT_DATA
            t.number_format = FORMATO_NUM
            t.border = BORDE
            lookup[(f"cuenta:{codigo}:{tramo}", "TOTAL")] = _addr(
                get_column_letter(COL_TOTAL), fila
            )
            fila += 1

        ultima = fila - 1
        etiqueta = ws.cell(fila, COL_NOMBRE, f"Subtotal {titulo}")
        etiqueta.font = FONT_TOTAL
        etiqueta.fill = RELLENO_TOTAL
        for j in range(13):
            col = COL_PRIMER_MES + j
            letra = get_column_letter(col)
            c = ws.cell(fila, col, f"=SUM({letra}{primera}:{letra}{ultima})")
            c.font = FONT_TOTAL
            c.fill = RELLENO_TOTAL
            c.number_format = FORMATO_NUM
            c.border = BORDE
            clave = MESES[j] if j < 12 else "TOTAL"
            lookup[(f"{CAT_VENTAS}:{tramo}", clave)] = _addr(letra, fila)
        filas_subtotal.append(fila)
        fila += 2

    etiqueta = ws.cell(fila, COL_NOMBRE,
                       "Control: desglose − Subtotal VENTAS (debe ser cero)")
    etiqueta.font = FONT_TOTAL
    for j in range(13):
        col = COL_PRIMER_MES + j
        letra = get_column_letter(col)
        suma = ",".join(f"{letra}{f}" for f in filas_subtotal)
        c = ws.cell(fila, col,
                    f"=ROUND(SUM({suma})-{letra}{fila_subtotal_ventas},2)")
        c.font = FONT_TOTAL
        c.number_format = FORMATO_NUM
        c.border = BORDE
    return fila + 2


def build_hoja_mayores(wb: Workbook, filas, movimientos=None) -> dict[tuple[str, str], str]:
    """Crea la hoja resumen. Devuelve {(categoria, "01".."12"|"TOTAL") → addr}.

    Todas las direcciones devueltas van calificadas con el nombre de la hoja
    (`'Mayores homologados'!D30`). La única excepción es la clave
    `("orden:<categoria>", "cuentas")`, que guarda una LISTA de códigos de
    cuenta, no una dirección.

    Con `movimientos` se añade además el bloque de desglose de las ventas por
    tarifa y sus direcciones `("cuenta:<codigo>:gravada"|"cero"|"por_asignar",
    mes)` y `("VENTAS:<tramo>", mes)`. Sin movimientos no hay forma de separar
    por asiento y la hoja queda exactamente como estaba.
    """
    if SHEET_MAYORES in wb.sheetnames:
        del wb[SHEET_MAYORES]
    ws = wb.create_sheet(SHEET_MAYORES)

    ws.cell(1, 1, "MAYORES HOMOLOGADOS · resumen por cuenta y mes").font = FONT_TITULO

    encabezado = ["Categoría", "Código", "Cuenta"] + NOMBRES_MES + ["Total"]
    for i, texto in enumerate(encabezado, start=1):
        c = ws.cell(3, i, texto)
        c.font = FONT_TOTAL
        c.border = BORDE
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    por_categoria: dict[str, list] = {}
    for f in filas:
        por_categoria.setdefault(f.categoria_final or SIN_CLASIFICAR, []).append(f)

    lookup: dict[tuple[str, str], str] = {}
    fila = 4
    fila_subtotal_ventas = 0
    for categoria in sorted(por_categoria, key=lambda c: (_orden(c), c)):
        cuentas = sorted(por_categoria[categoria], key=lambda f: f.codigo_cuenta)
        primera = fila
        for f in cuentas:
            ws.cell(fila, COL_CATEGORIA, categoria).font = FONT_DATA
            ws.cell(fila, COL_CODIGO, f.codigo_cuenta).font = FONT_DATA
            ws.cell(fila, COL_NOMBRE, f.nombre_cuenta).font = FONT_DATA
            por_mes = f.por_mes_json or {}
            for j, mes in enumerate(MESES):
                c = ws.cell(fila, COL_PRIMER_MES + j, float(por_mes.get(mes, 0.0)))
                c.font = FONT_DATA
                c.number_format = FORMATO_NUM
                c.border = BORDE
                lookup[(f"cuenta:{f.codigo_cuenta}", MESES[j])] = _addr(
                    get_column_letter(COL_PRIMER_MES + j), fila
                )
            ini = get_column_letter(COL_PRIMER_MES)
            fin = get_column_letter(COL_PRIMER_MES + 11)
            t = ws.cell(fila, COL_TOTAL, f"=SUM({ini}{fila}:{fin}{fila})")
            t.font = FONT_DATA
            t.number_format = FORMATO_NUM
            t.border = BORDE
            lookup[(f"cuenta:{f.codigo_cuenta}", "TOTAL")] = _addr(
                get_column_letter(COL_TOTAL), fila
            )
            fila += 1

        lookup[(f"orden:{categoria}", "cuentas")] = [f.codigo_cuenta for f in cuentas]
        ultima = fila - 1
        etiqueta = ws.cell(fila, COL_NOMBRE, f"Subtotal {categoria}")
        etiqueta.font = FONT_TOTAL
        etiqueta.fill = RELLENO_TOTAL
        for j in range(13):
            col = COL_PRIMER_MES + j
            letra = get_column_letter(col)
            c = ws.cell(fila, col, f"=SUM({letra}{primera}:{letra}{ultima})")
            c.font = FONT_TOTAL
            c.fill = RELLENO_TOTAL
            c.number_format = FORMATO_NUM
            c.border = BORDE
            clave = MESES[j] if j < 12 else "TOTAL"
            lookup[(categoria, clave)] = _addr(letra, fila)
        if categoria == CAT_VENTAS:
            fila_subtotal_ventas = fila
        fila += 2  # una fila en blanco entre bloques

    cuentas_ventas = lookup.get((f"orden:{CAT_VENTAS}", "cuentas")) or []
    if movimientos and cuentas_ventas:
        desglose = separar_ventas_por_tarifa(
            movimientos,
            {f.codigo_cuenta: f.categoria_final for f in filas},
        )
        fila = _bloque_desglose_ventas(
            ws, fila=fila + 1, cuentas=cuentas_ventas, desglose=desglose,
            nombres={f.codigo_cuenta: f.nombre_cuenta for f in filas},
            fila_subtotal_ventas=fila_subtotal_ventas, lookup=lookup,
        )

    ws.freeze_panes = "D4"
    for col, ancho in ((1, 16), (2, 14), (3, 46)):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    for j in range(13):
        ws.column_dimensions[get_column_letter(COL_PRIMER_MES + j)].width = 14
    return lookup
