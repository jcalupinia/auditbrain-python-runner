"""DM5 Ventas — ventas gravadas, ventas 0% e IVA en ventas, libros vs declaraciones.

Tres bloques "libros vs declarado vs diferencia". El número de cuentas de
venta varía por cliente, así que las posiciones de las filas de subtotal se
calculan en cada bloque, nunca se hardcodean.

Las cuentas de VENTAS ≠0% y VENTAS 0% son las MISMAS cuentas del mayor, pero
NO las mismas cifras: la hoja de mayores publica, además del total por cuenta
y mes, el desglose por tarifa que se obtiene asiento por asiento (ver
`mayor/ventas_tarifa.py`). El bloque ≠0% lee el tramo gravado y el bloque 0%
lee el tramo 0%. Antes ambos leían la misma celda del resumen y contrastaban
esa única cifra contra casilleros distintos del F-104.

Lo que la separación no puede resolver sin adivinar (asientos con tarifas
mezcladas donde varias combinaciones de líneas cuadran con el IVA) se muestra
en una fila propia dentro del bloque ≠0%, FUERA del "Según libros". Va en ese
bloque porque el remanente sólo aparece en asientos que SÍ tienen IVA —los
asientos sin IVA se resuelven como 0% de una— así que es base gravada
pendiente de identificar, no venta 0% no declarada.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.bloques import (
    COL_ETIQUETA, COL_PRIMER_MES, MESES, escribir_encabezado_meses, fila_diferencia,
    fila_referencias, fila_suma_rango,
)
from backend.app.aud.obligaciones_fiscales.libro.estilos import (
    BORDE, FONT_TOTAL, FORMATO_NUM, RELLENO_TOTAL, escribir_encabezado_cedula,
    escribir_leyenda_marcas,
)

SHEET_DM5 = "DM5 Ventas"
CASILLEROS_VENTAS = ["411", "412", "444"]
CASILLEROS_VENTAS_0 = ["412", "413", "414", "415", "417", "418", "444"]
CASILLEROS_IVA_VENTAS = ["421", "422", "423", "424", "454"]

ETIQUETA_POR_ASIGNAR = "Por asignar (revisar asientos con tarifas mezcladas)"


def _dirs_cuenta(dir_mayores: dict, codigo: str, tramo: str | None = None) -> dict[str, str]:
    """Direcciones mensuales de una cuenta, opcionalmente de un tramo.

    Si se pide un tramo (`gravada` / `cero`) y la hoja de mayores no lo
    publicó —se armó sin movimientos, así que no hubo cómo separar— se cae al
    total de la cuenta: es preferible mostrar el total en los dos bloques,
    como antes, que dejar la cédula en cero.
    """
    if tramo:
        del_tramo = {
            mes: dir_mayores[(f"cuenta:{codigo}:{tramo}", mes)]
            for mes in MESES
            if (f"cuenta:{codigo}:{tramo}", mes) in dir_mayores
        }
        if del_tramo:
            return del_tramo
    return {
        mes: dir_mayores[(f"cuenta:{codigo}", mes)]
        for mes in MESES
        if (f"cuenta:{codigo}", mes) in dir_mayores
    }


def _dirs_casillero(dir_f104: dict, periodos: list[str], cas: str) -> dict[str, str]:
    salida = {}
    for periodo in periodos:
        mes = periodo.split("-")[-1]
        addr = dir_f104.get((periodo, cas))
        if addr:
            salida[mes] = addr
    return salida


def _bloque_cuentas(ws, *, fila: int, titulo: str, cuentas: list[str],
                    dir_mayores: dict, nombres_cuenta: dict,
                    tramo: str | None = None,
                    extra: tuple[str, dict[str, str]] | None = None) -> tuple[int, int]:
    """Encabezado + una fila por cuenta + 'Según libros'. Devuelve (fila_siguiente, fila_libros).

    `extra` es una fila informativa que se escribe DESPUÉS del 'Según libros'
    y por tanto queda fuera de su suma.
    """
    escribir_encabezado_meses(ws, fila=fila, titulo=titulo)
    fila += 1
    primera = fila
    for codigo in cuentas:
        fila_referencias(ws, fila=fila, etiqueta=nombres_cuenta.get(codigo, codigo),
                         direcciones=_dirs_cuenta(dir_mayores, codigo, tramo))
        fila += 1
    ultima = fila - 1
    fila_libros = fila
    fila_suma_rango(ws, fila=fila_libros, etiqueta="Según libros", desde=primera, hasta=ultima)
    fila += 1
    if extra:
        etiqueta, direcciones = extra
        fila_referencias(ws, fila=fila, etiqueta=etiqueta, direcciones=direcciones)
        fila += 1
    fila += 1
    return fila, fila_libros


def _bloque_declarado(ws, *, fila: int, casilleros: list[str], dir_f104: dict,
                      periodos: list[str]) -> tuple[int, int]:
    """Una fila por casillero + 'Según declaraciones'. Devuelve (fila_siguiente, fila_declarado)."""
    primero = fila
    for cas in casilleros:
        fila_referencias(ws, fila=fila, etiqueta=f"Casillero {cas}",
                         direcciones=_dirs_casillero(dir_f104, periodos, cas))
        fila += 1
    fila_declarado = fila
    fila_suma_rango(ws, fila=fila_declarado, etiqueta="Según declaraciones",
                    desde=primero, hasta=fila - 1)
    fila += 1
    return fila, fila_declarado


def _fila_suma_de_filas(ws, *, fila: int, etiqueta: str, filas: list[int]) -> None:
    """Fila que suma celdas puntuales (no un rango contiguo) de la misma hoja.

    Es el caso de 'Total ventas declaradas': la suma de dos filas 'Según
    declaraciones' que quedan separadas por el bloque de casilleros
    intermedio, así que no se puede usar SUM sobre un RANGO; sí sobre la
    lista de celdas. Se escribe `=SUM(C20,C35)` y no `=C20+C35` a propósito:
    en las cédulas DM una fórmula que son puras referencias unidas por '+'
    significa siempre "referencia a otra hoja" y debe ir calificada (ver
    tests/test_of_libro_direcciones_calificadas.py). Esta es la única suma de
    celdas sueltas de la MISMA hoja, y como SUM no rompe esa regla.
    """
    e = ws.cell(fila, COL_ETIQUETA, etiqueta)
    e.font = FONT_TOTAL
    e.fill = RELLENO_TOTAL
    for j in range(13):
        col = get_column_letter(COL_PRIMER_MES + j)
        formula = "=SUM(" + ",".join(f"{col}{f}" for f in filas) + ")"
        c = ws.cell(fila, COL_PRIMER_MES + j, formula)
        c.font = FONT_TOTAL
        c.fill = RELLENO_TOTAL
        c.number_format = FORMATO_NUM
        c.border = BORDE


def build_dm5(
    wb: Workbook,
    *,
    dir_mayores: dict,
    dir_f104: dict,
    periodos: list[str],
    nombres_cuenta: dict[str, str],
    cliente: str,
    periodo: str,
    preparado_por: str | None = None,
    revisado_por: str | None = None,
) -> dict[tuple[str, str], str]:
    """Construye DM5.

    Devuelve {("ventas_libros"|"ventas_0_libros"|"iva_ventas_libros"|
    "total_declarado", mes) → addr}, el mapa que DM6 consume.
    """
    if SHEET_DM5 in wb.sheetnames:
        del wb[SHEET_DM5]
    ws = wb.create_sheet(SHEET_DM5)

    escribir_encabezado_cedula(
        ws, titulo="Ventas e IVA en ventas", referencia="DM5",
        cliente=cliente, periodo=periodo,
        preparado_por=preparado_por, revisado_por=revisado_por,
    )

    cuentas_ventas = dir_mayores.get(("orden:VENTAS", "cuentas"), [])
    cuentas_iva_ventas = dir_mayores.get(("orden:IVA_VENTAS", "cuentas"), [])

    fila = 13

    # --- Bloque 1: Ventas ≠ 0% ---
    dirs_por_asignar = {
        mes: dir_mayores[("VENTAS:por_asignar", mes)]
        for mes in MESES
        if ("VENTAS:por_asignar", mes) in dir_mayores
    }
    fila, fila_libros_1 = _bloque_cuentas(
        ws, fila=fila, titulo="VENTAS ≠ 0%", cuentas=cuentas_ventas,
        dir_mayores=dir_mayores, nombres_cuenta=nombres_cuenta, tramo="gravada",
        extra=(ETIQUETA_POR_ASIGNAR, dirs_por_asignar) if dirs_por_asignar else None,
    )
    fila, fila_decl_1 = _bloque_declarado(
        ws, fila=fila, casilleros=CASILLEROS_VENTAS, dir_f104=dir_f104, periodos=periodos,
    )
    fila_diferencia(ws, fila=fila, etiqueta="Diferencia",
                    fila_libros=fila_libros_1, fila_declarado=fila_decl_1)
    fila += 3

    # --- Bloque 2: Ventas 0% (mismas cuentas de libros, otros casilleros) ---
    fila, fila_libros_2 = _bloque_cuentas(
        ws, fila=fila, titulo="VENTAS 0%", cuentas=cuentas_ventas,
        dir_mayores=dir_mayores, nombres_cuenta=nombres_cuenta, tramo="cero",
    )
    fila, fila_decl_2 = _bloque_declarado(
        ws, fila=fila, casilleros=CASILLEROS_VENTAS_0, dir_f104=dir_f104, periodos=periodos,
    )
    fila_diferencia(ws, fila=fila, etiqueta="Diferencia",
                    fila_libros=fila_libros_2, fila_declarado=fila_decl_2)
    fila += 1

    fila_total_declarado = fila
    _fila_suma_de_filas(ws, fila=fila_total_declarado, etiqueta="Total ventas declaradas",
                        filas=[fila_decl_1, fila_decl_2])
    fila += 3

    # --- Bloque 3: IVA en ventas ---
    fila, fila_libros_3 = _bloque_cuentas(
        ws, fila=fila, titulo="IVA EN VENTAS", cuentas=cuentas_iva_ventas,
        dir_mayores=dir_mayores, nombres_cuenta=nombres_cuenta,
    )
    fila, fila_decl_3 = _bloque_declarado(
        ws, fila=fila, casilleros=CASILLEROS_IVA_VENTAS, dir_f104=dir_f104, periodos=periodos,
    )
    fila_diferencia(ws, fila=fila, etiqueta="Diferencia",
                    fila_libros=fila_libros_3, fila_declarado=fila_decl_3)
    fila += 3

    escribir_leyenda_marcas(ws, fila=fila)

    for col, ancho in ((1, 24), (2, 30)):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    for j in range(13):
        ws.column_dimensions[get_column_letter(COL_PRIMER_MES + j)].width = 15

    salida: dict[tuple[str, str], str] = {}
    for clave, fila_origen in (
        ("ventas_libros", fila_libros_1),
        ("ventas_0_libros", fila_libros_2),
        ("iva_ventas_libros", fila_libros_3),
        ("total_declarado", fila_total_declarado),
    ):
        salida.update({
            (clave, m): f"'{SHEET_DM5}'!{get_column_letter(COL_PRIMER_MES + j)}{fila_origen}"
            for j, m in enumerate(MESES)
        })
    return salida
