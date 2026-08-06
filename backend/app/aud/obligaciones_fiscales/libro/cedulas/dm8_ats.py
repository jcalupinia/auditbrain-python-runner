"""DM8 ATS — cruza el Anexo Transaccional (ATS) contra lo declarado en los
formularios F-104 (IVA/retenciones de IVA) y F-103 (retenciones de renta).

Mapeo verificado empíricamente contra un anexo real de diciembre 2025 (ver
"Mapeo fila de DM8 → campo del Talón Resumen" en el spec de diseño del
proyecto):

| Bloque de DM8                | Se compara contra                     |
|-------------------------------|----------------------------------------|
| Ventas 0%                     | casilleros 413 + 415                   |
| Ventas gravadas               | casillero 411                          |
| IVA en ventas                 | casillero 421                          |
| Compras 0%                    | casillero 519 − importaciones (*)      |
| Compras gravadas              | casillero 519                          |
| IVA en compras                | casillero 520                          |
| Retenciones de IVA (por %)    | casilleros 721/723/725/727/729/731     |
| Retenciones de renta (código) | casillero 499 del F-103                |
| IVA que le retuvieron         | casillero 609                          |
| Renta que le retuvieron       | informativo (anticipo de IR, fuera de
                                   alcance de un cruce mensual)             |
| Comprobantes anulados         | informativo                            |

(*) "Compras 0% vs 519 − importaciones" es la única fila de este mapeo que
NO forma parte de la lista de cruces verificados empíricamente con
diciembre 2025 (esa lista no incluye compras 0%): se implementa tal como la
documenta el spec, pero conviene revalidarla con un cliente que declare
importaciones.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.bloques import (
    COL_PRIMER_MES, MESES, escribir_encabezado_meses, fila_diferencia,
    fila_referencias, fila_suma_rango,
)
from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm7_retenciones import (
    CASILLERO_RET_IVA_CONTROL, CASILLERO_RET_RENTA, CASILLEROS_RET_IVA,
)
from backend.app.aud.obligaciones_fiscales.libro.estilos import (
    escribir_encabezado_cedula, escribir_leyenda_marcas,
)

SHEET_DM8 = "DM8 ATS"

CASILLERO_VENTAS_GRAVADAS = "411"
CASILLEROS_VENTAS_0 = ["413", "415"]
CASILLERO_IVA_VENTAS = "421"
CASILLERO_COMPRAS_TOTAL = "519"
CASILLERO_IVA_COMPRAS = "520"
CASILLERO_IVA_RETENIDO_RECIBIDO = "609"

# Casilleros de importaciones "Valor Neto" del F-104 (compras). Ver nota (*)
# en el docstring del módulo.
CASILLEROS_IMPORTACIONES = ["513", "514", "515", "516"]

# 721..731 en el mismo orden que los porcentajes fijos del catálogo de
# retención de IVA que reconoce el SRI.
_PORCENTAJES_CON_CASILLERO = list(
    zip(("10", "20", "30", "50", "70", "100"), CASILLEROS_RET_IVA)
)


def _dirs_campo(dir_ats: dict, campo: str) -> dict[str, str]:
    return {mes: dir_ats[(campo, mes)] for mes in MESES if (campo, mes) in dir_ats}


def _dirs_casillero(dir_datos: dict, periodos: list[str], cas: str) -> dict[str, str]:
    salida = {}
    for periodo in periodos:
        mes = periodo.split("-")[-1]
        addr = dir_datos.get((periodo, cas))
        if addr:
            salida[mes] = addr
    return salida


def _bloque_simple(
    ws, *, fila: int, titulo: str, etiqueta_ats: str, campo_ats: str, dir_ats: dict,
    etiqueta_declarado: str, casilleros_declarado: list[str], dir_f104: dict,
    periodos: list[str],
) -> int:
    """Encabezado + 'Según ATS' + N casilleros + 'Según declaración' + Diferencia."""
    escribir_encabezado_meses(ws, fila=fila, titulo=titulo)
    fila += 1
    fila_ats = fila
    fila_referencias(ws, fila=fila, etiqueta=etiqueta_ats, direcciones=_dirs_campo(dir_ats, campo_ats))
    fila += 2

    primero = fila
    for cas in casilleros_declarado:
        fila_referencias(ws, fila=fila, etiqueta=f"Casillero {cas}",
                         direcciones=_dirs_casillero(dir_f104, periodos, cas))
        fila += 1
    fila_declarado = fila
    fila_suma_rango(ws, fila=fila_declarado, etiqueta=etiqueta_declarado,
                    desde=primero, hasta=fila - 1)
    fila += 1
    fila_diferencia(ws, fila=fila, etiqueta="Diferencia",
                    fila_libros=fila_ats, fila_declarado=fila_declarado)
    fila += 3
    return fila


def build_dm8(
    wb: Workbook,
    *,
    dir_ats: dict,
    dir_f104: dict,
    dir_f103: dict,
    periodos: list[str],
    cliente: str,
    periodo: str,
    preparado_por: str | None = None,
    revisado_por: str | None = None,
) -> dict[tuple[str, str], str]:
    """Construye DM8. No publica direcciones: nada la consume por fórmula."""
    if SHEET_DM8 in wb.sheetnames:
        del wb[SHEET_DM8]
    ws = wb.create_sheet(SHEET_DM8)

    escribir_encabezado_cedula(
        ws, titulo="Anexo Transaccional (ATS) vs. formularios declarados",
        referencia="DM8", cliente=cliente, periodo=periodo,
        preparado_por=preparado_por, revisado_por=revisado_por,
    )

    fila = 13

    # --- Ventas 0% ---
    fila = _bloque_simple(
        ws, fila=fila, titulo="VENTAS 0% SEGÚN ATS",
        etiqueta_ats="Ventas 0% según ATS", campo_ats="ventas_bi_0", dir_ats=dir_ats,
        etiqueta_declarado="Según declaración (413+415)",
        casilleros_declarado=CASILLEROS_VENTAS_0, dir_f104=dir_f104, periodos=periodos,
    )

    # --- Ventas gravadas ---
    fila = _bloque_simple(
        ws, fila=fila, titulo="VENTAS GRAVADAS SEGÚN ATS",
        etiqueta_ats="Ventas gravadas según ATS", campo_ats="ventas_bi_gravada", dir_ats=dir_ats,
        etiqueta_declarado=f"Según declaración (casillero {CASILLERO_VENTAS_GRAVADAS})",
        casilleros_declarado=[CASILLERO_VENTAS_GRAVADAS], dir_f104=dir_f104, periodos=periodos,
    )

    # --- IVA en ventas ---
    fila = _bloque_simple(
        ws, fila=fila, titulo="IVA EN VENTAS SEGÚN ATS",
        etiqueta_ats="IVA ventas según ATS", campo_ats="ventas_iva", dir_ats=dir_ats,
        etiqueta_declarado=f"Según declaración (casillero {CASILLERO_IVA_VENTAS})",
        casilleros_declarado=[CASILLERO_IVA_VENTAS], dir_f104=dir_f104, periodos=periodos,
    )

    # --- Compras 0% = ATS vs (519 − importaciones) ---
    escribir_encabezado_meses(ws, fila=fila, titulo="COMPRAS 0% SEGÚN ATS")
    fila += 1
    fila_ats_compras0 = fila
    fila_referencias(ws, fila=fila, etiqueta="Compras 0% según ATS",
                     direcciones=_dirs_campo(dir_ats, "compras_bi_0"))
    fila += 2

    fila_519 = fila
    fila_referencias(ws, fila=fila, etiqueta=f"Casillero {CASILLERO_COMPRAS_TOTAL}",
                     direcciones=_dirs_casillero(dir_f104, periodos, CASILLERO_COMPRAS_TOTAL))
    fila += 1
    primer_import = fila
    for cas in CASILLEROS_IMPORTACIONES:
        fila_referencias(ws, fila=fila, etiqueta=f"Casillero {cas} (importaciones)",
                         direcciones=_dirs_casillero(dir_f104, periodos, cas))
        fila += 1
    fila_import_total = fila
    fila_suma_rango(ws, fila=fila_import_total, etiqueta="Total importaciones",
                    desde=primer_import, hasta=fila - 1)
    fila += 1
    fila_declarado_compras0 = fila
    fila_diferencia(ws, fila=fila, etiqueta=f"Según declaración ({CASILLERO_COMPRAS_TOTAL} − importaciones)",
                    fila_libros=fila_519, fila_declarado=fila_import_total)
    fila += 1
    fila_diferencia(ws, fila=fila, etiqueta="Diferencia",
                    fila_libros=fila_ats_compras0, fila_declarado=fila_declarado_compras0)
    fila += 3

    # --- Compras gravadas ---
    fila = _bloque_simple(
        ws, fila=fila, titulo="COMPRAS GRAVADAS SEGÚN ATS",
        etiqueta_ats="Compras gravadas según ATS", campo_ats="compras_bi_gravada", dir_ats=dir_ats,
        etiqueta_declarado=f"Según declaración (casillero {CASILLERO_COMPRAS_TOTAL})",
        casilleros_declarado=[CASILLERO_COMPRAS_TOTAL], dir_f104=dir_f104, periodos=periodos,
    )

    # --- IVA en compras ---
    fila = _bloque_simple(
        ws, fila=fila, titulo="IVA EN COMPRAS SEGÚN ATS",
        etiqueta_ats="IVA compras según ATS", campo_ats="compras_iva", dir_ats=dir_ats,
        etiqueta_declarado=f"Según declaración (casillero {CASILLERO_IVA_COMPRAS})",
        casilleros_declarado=[CASILLERO_IVA_COMPRAS], dir_f104=dir_f104, periodos=periodos,
    )

    # --- Comprobantes anulados (informativo, no tiene contraparte declarada) ---
    escribir_encabezado_meses(ws, fila=fila, titulo="COMPROBANTES ANULADOS (INFORMATIVO)")
    fila += 1
    fila_referencias(ws, fila=fila, etiqueta="Anulados según ATS",
                     direcciones=_dirs_campo(dir_ats, "anulados"))
    fila += 3

    # --- Retenciones de IVA por porcentaje ---
    escribir_encabezado_meses(ws, fila=fila, titulo="RETENCIONES DE IVA SEGÚN ATS (POR PORCENTAJE)")
    fila += 1
    for pct, cas in _PORCENTAJES_CON_CASILLERO:
        fila_ats_pct = fila
        fila_referencias(ws, fila=fila, etiqueta=f"Ret. IVA {pct}% según ATS",
                         direcciones=_dirs_campo(dir_ats, f"iva_pct:{pct}"))
        fila += 1
        fila_referencias(ws, fila=fila, etiqueta=f"Casillero {cas}",
                         direcciones=_dirs_casillero(dir_f104, periodos, cas))
        fila_cas_pct = fila
        fila += 1
        fila_diferencia(ws, fila=fila, etiqueta=f"Diferencia {pct}%",
                        fila_libros=fila_ats_pct, fila_declarado=fila_cas_pct)
        fila += 2

    # NC: nota de crédito, no tiene casillero de contraparte.
    fila_referencias(ws, fila=fila, etiqueta="Ret. IVA NC según ATS (informativo)",
                     direcciones=_dirs_campo(dir_ats, "iva_pct:NC"))
    fila += 2

    fila_ats_iva_total = fila
    fila_referencias(ws, fila=fila, etiqueta="Total retenciones de IVA según ATS",
                     direcciones=_dirs_campo(dir_ats, "ret_iva_total"))
    fila += 1
    fila_referencias(
        ws, fila=fila, etiqueta=f"Casillero {CASILLERO_RET_IVA_CONTROL} (control del total)",
        direcciones=_dirs_casillero(dir_f104, periodos, CASILLERO_RET_IVA_CONTROL),
    )
    fila_cas_control = fila
    fila += 1
    fila_diferencia(ws, fila=fila, etiqueta="Diferencia",
                    fila_libros=fila_ats_iva_total, fila_declarado=fila_cas_control)
    fila += 3

    # --- Retenciones de renta por código (catálogo abierto: varía por cliente) ---
    codigos_renta = sorted({
        clave[0].split(":", 1)[1] for clave in dir_ats if clave[0].startswith("renta_codigo:")
    })
    escribir_encabezado_meses(ws, fila=fila, titulo="RETENCIONES DE RENTA SEGÚN ATS (POR CÓDIGO)")
    fila += 1
    primera_renta = fila
    for codigo in codigos_renta:
        fila_referencias(ws, fila=fila, etiqueta=f"Ret. renta {codigo} según ATS",
                         direcciones=_dirs_campo(dir_ats, f"renta_codigo:{codigo}"))
        fila += 1
    fila_total_ats_renta = fila
    if codigos_renta:
        fila_suma_rango(ws, fila=fila_total_ats_renta, etiqueta="Total ATS",
                        desde=primera_renta, hasta=fila - 1)
    else:
        fila_referencias(ws, fila=fila, etiqueta="Total ATS",
                         direcciones=_dirs_campo(dir_ats, "ret_renta_total"))
    fila += 1
    fila_referencias(ws, fila=fila, etiqueta=f"Casillero {CASILLERO_RET_RENTA} (F-103)",
                     direcciones=_dirs_casillero(dir_f103, periodos, CASILLERO_RET_RENTA))
    fila_cas_renta = fila
    fila += 1
    fila_diferencia(ws, fila=fila, etiqueta="Diferencia",
                    fila_libros=fila_total_ats_renta, fila_declarado=fila_cas_renta)
    fila += 3

    # --- Retenciones que le efectuaron ---
    escribir_encabezado_meses(ws, fila=fila, titulo="RETENCIONES QUE LE EFECTUARON")
    fila += 1
    fila_ats_iva_recibido = fila
    fila_referencias(ws, fila=fila, etiqueta="IVA que le retuvieron según ATS",
                     direcciones=_dirs_campo(dir_ats, "iva_le_retuvieron"))
    fila += 1
    fila_referencias(
        ws, fila=fila, etiqueta=f"Casillero {CASILLERO_IVA_RETENIDO_RECIBIDO}",
        direcciones=_dirs_casillero(dir_f104, periodos, CASILLERO_IVA_RETENIDO_RECIBIDO),
    )
    fila_cas_609 = fila
    fila += 1
    fila_diferencia(ws, fila=fila, etiqueta="Diferencia",
                    fila_libros=fila_ats_iva_recibido, fila_declarado=fila_cas_609)
    fila += 2
    fila_referencias(
        ws, fila=fila,
        etiqueta="Renta que le retuvieron según ATS (informativo, cruza con el anticipo de IR)",
        direcciones=_dirs_campo(dir_ats, "renta_le_retuvieron"),
    )
    fila += 3

    escribir_leyenda_marcas(ws, fila=fila)

    for col, ancho in ((1, 24), (2, 48)):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    for j in range(13):
        ws.column_dimensions[get_column_letter(COL_PRIMER_MES + j)].width = 15

    return {}
