"""Anatomía común de las cédulas DM.

Cada bloque es una tabla de 12 meses + total con: filas de detalle (una por
cuenta o por casillero), una fila "Según libros", una fila "Según
declaraciones" y una fila "Diferencia". Todas las cifras son fórmulas.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter

from backend.app.aud.obligaciones_fiscales.libro.estilos import (
    BORDE, FONT_DATA, FONT_ENCABEZADO_TABLA, FONT_TOTAL, FORMATO_NUM, RELLENO_TOTAL,
)

MESES = [f"{m:02d}" for m in range(1, 13)]
NOMBRES_MES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
               "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

COL_TITULO, COL_ETIQUETA, COL_PRIMER_MES = 1, 2, 3
COL_TOTAL = COL_PRIMER_MES + 12


def _col(mes_idx: int) -> str:
    return get_column_letter(COL_PRIMER_MES + mes_idx)


def escribir_encabezado_meses(ws, *, fila: int, titulo: str, etiqueta: str = "Cuenta") -> None:
    ws.cell(fila, COL_TITULO, titulo).font = FONT_ENCABEZADO_TABLA
    ws.cell(fila, COL_ETIQUETA, etiqueta).font = FONT_ENCABEZADO_TABLA
    for j, nombre in enumerate(NOMBRES_MES):
        c = ws.cell(fila, COL_PRIMER_MES + j, nombre)
        c.font = FONT_ENCABEZADO_TABLA
        c.border = BORDE
    c = ws.cell(fila, COL_TOTAL, "Total")
    c.font = FONT_ENCABEZADO_TABLA
    c.border = BORDE


def fila_referencias(ws, *, fila: int, etiqueta: str, direcciones: dict[str, str]) -> None:
    """Una fila cuyos 12 meses apuntan por fórmula a otra hoja.

    Un mes sin dirección se escribe como 0: una celda vacía en un papel de
    trabajo se lee como dato faltante, no como ausencia de movimiento.
    """
    ws.cell(fila, COL_ETIQUETA, etiqueta).font = FONT_DATA
    for j, mes in enumerate(MESES):
        addr = direcciones.get(mes)
        c = ws.cell(fila, COL_PRIMER_MES + j, f"={addr}" if addr else 0)
        c.font = FONT_DATA
        c.number_format = FORMATO_NUM
        c.border = BORDE
    t = ws.cell(fila, COL_TOTAL, f"=SUM({_col(0)}{fila}:{_col(11)}{fila})")
    t.font = FONT_DATA
    t.number_format = FORMATO_NUM
    t.border = BORDE


def fila_suma_direcciones(ws, *, fila: int, etiqueta: str,
                          direcciones_por_mes: dict[str, list[str]]) -> None:
    """Una fila donde cada mes es la SUMA de varias celdas de otra hoja.

    Es el caso de 'Según declaraciones': varios casilleros del mismo mes.
    """
    ws.cell(fila, COL_ETIQUETA, etiqueta).font = FONT_DATA
    for j, mes in enumerate(MESES):
        addrs = direcciones_por_mes.get(mes) or []
        valor = ("=" + "+".join(addrs)) if addrs else 0
        c = ws.cell(fila, COL_PRIMER_MES + j, valor)
        c.font = FONT_DATA
        c.number_format = FORMATO_NUM
        c.border = BORDE
    t = ws.cell(fila, COL_TOTAL, f"=SUM({_col(0)}{fila}:{_col(11)}{fila})")
    t.font = FONT_DATA
    t.number_format = FORMATO_NUM
    t.border = BORDE


def fila_suma_rango(ws, *, fila: int, etiqueta: str, desde: int, hasta: int) -> None:
    """Fila de subtotal sobre un rango vertical de filas de la misma hoja."""
    e = ws.cell(fila, COL_ETIQUETA, etiqueta)
    e.font = FONT_TOTAL
    e.fill = RELLENO_TOTAL
    for j in range(13):
        col = get_column_letter(COL_PRIMER_MES + j)
        c = ws.cell(fila, COL_PRIMER_MES + j, f"=SUM({col}{desde}:{col}{hasta})")
        c.font = FONT_TOTAL
        c.fill = RELLENO_TOTAL
        c.number_format = FORMATO_NUM
        c.border = BORDE


def fila_diferencia(ws, *, fila: int, etiqueta: str, fila_libros: int,
                    fila_declarado: int) -> None:
    """Libros menos declarado, redondeado a 2 decimales.

    El ROUND evita el ruido de coma flotante que en el archivo modelo del
    auditor producía diferencias como -3,6e-12.
    """
    e = ws.cell(fila, COL_ETIQUETA, etiqueta)
    e.font = FONT_TOTAL
    for j in range(13):
        col = get_column_letter(COL_PRIMER_MES + j)
        c = ws.cell(fila, COL_PRIMER_MES + j,
                    f"=ROUND({col}{fila_libros}-{col}{fila_declarado},2)")
        c.font = FONT_TOTAL
        c.number_format = FORMATO_NUM
        c.border = BORDE
