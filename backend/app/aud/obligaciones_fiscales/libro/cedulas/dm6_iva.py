"""DM6 IVA — la conciliacion grande de IVA en ventas y compras.

A diferencia de DM3..DM5, aqui las filas son los MESES (una por mes, doce en
total) y las columnas son las 29 variables {1}..{29} de la conciliacion, mas
una columna adicional AF fuera de la numeracion (retenciones de IVA
declaradas). No lleva filas de cuentas.

Columna B corresponde a {1}, C a {2}, ..., AD a {29} (offset de una columna
porque la columna A guarda el nombre del mes). `letra_columna(n)` calcula esa
letra para no hardcodearla en once sitios distintos.

Encadenamiento entre meses (la parte critica): J{9} y T{19} de enero salen de
un casillero declarado; de febrero en adelante salen de L{11} y W{22} DE LA
FILA DEL MES ANTERIOR en esta misma hoja. Nunca se referencia la propia fila:
eso crearia un ciclo, porque L depende de K y M depende de J.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.bloques import MESES, NOMBRES_MES
from backend.app.aud.obligaciones_fiscales.libro.estilos import (
    BORDE, FONT_DATA, FONT_ENCABEZADO_TABLA, FORMATO_NUM, escribir_encabezado_cedula,
    escribir_leyenda_marcas,
)

SHEET_DM6 = "DM6 IVA"
TARIFA_POR_DEFECTO = 0.15

COL_MES = 1
FILA_ENCABEZADO_TABLA = 12
FILA_PRIMER_MES = 13

# Columna AF: fuera de la numeracion {1}..{29}, casillero 609 (retenciones de
# IVA declaradas). Ver tabla "Mapeo de casilleros confirmado empiricamente"
# del plan; no forma parte de la conciliacion numerada, es informativa.
COL_AF = "AF"

# Etiquetas descriptivas de cada columna. Las que dicen "(manual)" quedan en 0
# porque en el modelo real del auditor son cero por diseno (no hay percepcion
# de IVA ni ajustes manuales en el caso validado): no es que falte definirlas,
# es que su valor esperado es 0. D y E SI tienen fuente (casilleros 413 y
# 417+418): dejarlas en 0 literal producia un descuadre falso en AB {27}.
_ETIQUETAS = {
    1: "Ventas <> 0% (libros, DM5)",
    2: "Ventas 0% (libros, DM5)",
    3: "Ventas 0% sin derecho a credito tributario (cas. 413)",
    4: "Exportaciones de bienes y servicios (cas. 417+418)",
    5: "Transferencias a contado (cas. 480)",
    6: "Tarifa de IVA del mes",
    7: "IVA en ventas (B x G)",
    8: "IVA en la diferencia (cas. 424)",
    9: "Impuesto a liquidar mes anterior",
    10: "Transferencias a contado x tarifa + diferencia",
    11: "Impuesto causado del mes",
    12: "Total impuesto a liquidar",
    13: "Base imponible de compras (DM4)",
    14: "IVA en compras (N x G)",
    15: "Otro credito tributario (cero por diseno)",
    16: "Factor de proporcionalidad",
    17: "Credito tributario aplicable",
    18: "Retenciones que le efectuaron (cero por diseno)",
    19: "Credito del mes anterior",
    20: "IVA retenido por clientes (libros)",
    21: "Ajuste manual (cero por diseno)",
    22: "Saldo a favor del contribuyente",
    23: "Impuesto a pagar",
    24: "Total ventas declaradas (DM5)",
    25: "Saldo credito tributario prox. mes (cas. 615+617)",
    26: "Impuesto por percepcion (cas. 699)",
    27: "Diferencia ventas (Y-B-C-D-E)",
    28: "Diferencia saldo a favor (Z-W)",
    29: "Diferencia impuesto a pagar (AA-X)",
}
_ETIQUETA_AF = "Retenciones de IVA declaradas (cas. 609)"


def letra_columna(n: int) -> str:
    """La letra de columna de la variable {n} (1..29). {1}->B, {29}->AD."""
    return get_column_letter(n + 1)


def _mapa_mes_a_periodo(periodos: list[str]) -> dict[str, str]:
    return {p.split("-")[-1]: p for p in periodos}


def _dir_casillero(dir_f104: dict, mapa_periodo: dict, mes: str, cas: str) -> str | None:
    periodo = mapa_periodo.get(mes)
    if not periodo:
        return None
    return dir_f104.get((periodo, cas))


def _formula_o_cero(addr: str | None):
    return f"={addr}" if addr else 0


def _formula_suma_o_cero(addr1: str | None, addr2: str | None):
    if addr1 and addr2:
        return f"={addr1}+{addr2}"
    if addr1:
        return f"={addr1}"
    if addr2:
        return f"={addr2}"
    return 0


def _escribir(ws, col: str, fila: int, valor, *, es_tarifa: bool = False) -> None:
    c = ws[f"{col}{fila}"]
    c.value = valor
    c.font = FONT_DATA
    c.border = BORDE
    c.number_format = "0.00%" if es_tarifa else FORMATO_NUM


def build_dm6(
    wb: Workbook,
    *,
    dir_dm5: dict,
    dir_dm4: dict,
    dir_mayores: dict,
    dir_f104: dict,
    periodos: list[str],
    cliente: str,
    periodo: str,
    tarifas: dict[str, float] | None = None,
    preparado_por: str | None = None,
    revisado_por: str | None = None,
) -> dict[tuple[str, str], str]:
    """Construye DM6. Devuelve {(f"col_{n}", mes) → addr} para las 29
    columnas numeradas y {("col_AF", mes) → addr} para la columna adicional.
    """
    if SHEET_DM6 in wb.sheetnames:
        del wb[SHEET_DM6]
    ws = wb.create_sheet(SHEET_DM6)
    tarifas = tarifas or {}
    mapa_periodo = _mapa_mes_a_periodo(periodos)

    escribir_encabezado_cedula(
        ws, titulo="Conciliación de IVA", referencia="DM6",
        cliente=cliente, periodo=periodo,
        preparado_por=preparado_por, revisado_por=revisado_por,
    )

    # --- encabezado de la tabla ---
    ws.cell(FILA_ENCABEZADO_TABLA, COL_MES, "Mes").font = FONT_ENCABEZADO_TABLA
    for n, etiqueta in _ETIQUETAS.items():
        col = letra_columna(n)
        c = ws[f"{col}{FILA_ENCABEZADO_TABLA}"]
        c.value = etiqueta
        c.font = FONT_ENCABEZADO_TABLA
        c.border = BORDE
    c = ws[f"{COL_AF}{FILA_ENCABEZADO_TABLA}"]
    c.value = _ETIQUETA_AF
    c.font = FONT_ENCABEZADO_TABLA
    c.border = BORDE

    b, c_, d, e, f, g, h, i_, j, k, l, m, n, o, p, q, r, s, t, u, v, w, x, y, z, aa, ab, ac, ad = (
        letra_columna(idx) for idx in range(1, 30)
    )

    fila_anterior: int | None = None
    for idx, mes in enumerate(MESES):
        fila = FILA_PRIMER_MES + idx
        es_enero = idx == 0

        ws.cell(fila, COL_MES, NOMBRES_MES[idx]).font = FONT_DATA
        ws.cell(fila, COL_MES).border = BORDE

        # --- columnas que vienen de fuera ---
        _escribir(ws, b, fila, _formula_o_cero(dir_dm5.get(("ventas_libros", mes))))
        _escribir(ws, c_, fila, _formula_o_cero(dir_dm5.get(("ventas_0_libros", mes))))
        # {3} D: ventas 0% que NO dan derecho a credito tributario, casillero 413.
        _escribir(ws, d, fila, _formula_o_cero(_dir_casillero(dir_f104, mapa_periodo, mes, "413")))
        # {4} E: exportaciones de bienes y servicios, casilleros 417 + 418.
        _escribir(ws, e, fila, _formula_suma_o_cero(
            _dir_casillero(dir_f104, mapa_periodo, mes, "417"),
            _dir_casillero(dir_f104, mapa_periodo, mes, "418"),
        ))
        _escribir(ws, f, fila, _formula_o_cero(_dir_casillero(dir_f104, mapa_periodo, mes, "480")))
        _escribir(ws, g, fila, tarifas.get(mes, TARIFA_POR_DEFECTO), es_tarifa=True)

        # --- {7} H = B*G ---
        _escribir(ws, h, fila, f"={b}{fila}*{g}{fila}")

        _escribir(ws, i_, fila, _formula_o_cero(_dir_casillero(dir_f104, mapa_periodo, mes, "424")))

        # --- {9} J: enero = casillero 483; resto = L del mes anterior ---
        if es_enero:
            _escribir(ws, j, fila, _formula_o_cero(
                _dir_casillero(dir_f104, mapa_periodo, mes, "483")))
        else:
            _escribir(ws, j, fila, f"={l}{fila_anterior}")

        # --- {10} K = F*G + I ---
        _escribir(ws, k, fila, f"={f}{fila}*{g}{fila}+{i_}{fila}")
        # --- {11} L = (H+I) - K ---
        _escribir(ws, l, fila, f"=({h}{fila}+{i_}{fila})-{k}{fila}")
        # --- {12} M = J + K ---
        _escribir(ws, m, fila, f"={j}{fila}+{k}{fila}")

        _escribir(ws, n, fila, _formula_o_cero(dir_dm4.get(("base", mes))))
        # --- {14} O = N*G ---
        _escribir(ws, o, fila, f"={n}{fila}*{g}{fila}")
        _escribir(ws, p, fila, 0)  # {15} cero por diseno en el modelo real del auditor

        # --- {16} Q = IF(SUM(B:E)=0, 0, (B+C+E)/(B+C+D+E)) ---
        _escribir(ws, q, fila,
                 f"=IF(SUM({b}{fila}:{e}{fila})=0,0,"
                 f"({b}{fila}+{c_}{fila}+{e}{fila})/"
                 f"({b}{fila}+{c_}{fila}+{d}{fila}+{e}{fila}))")
        # --- {17} R = (O+P)*Q ---
        _escribir(ws, r, fila, f"=({o}{fila}+{p}{fila})*{q}{fila}")

        _escribir(ws, s, fila, 0)  # {18} cero por diseno en el modelo real del auditor

        # --- {19} T: enero = casillero 605+606; resto = W del mes anterior ---
        if es_enero:
            _escribir(ws, t, fila, _formula_suma_o_cero(
                _dir_casillero(dir_f104, mapa_periodo, mes, "605"),
                _dir_casillero(dir_f104, mapa_periodo, mes, "606"),
            ))
        else:
            _escribir(ws, t, fila, f"={w}{fila_anterior}")

        _escribir(ws, u, fila, _formula_o_cero(dir_mayores.get(("IVA_RETENIDO", mes))))
        _escribir(ws, v, fila, 0)  # {21} cero por diseno en el modelo real del auditor

        expr = f"({m}{fila}-{r}{fila}-{s}{fila}-{t}{fila}-{u}{fila}+{v}{fila})"
        # --- {22} W = ABS(IF((M-R-S-T-U+V)<0, (M-R-S-T-U+V), 0)) ---
        _escribir(ws, w, fila, f"=ABS(IF({expr}<0,{expr},0))")
        # --- {23} X = IF((M-R-S-T-U+V)>0, (M-R-S-T-U+V), 0) ---
        _escribir(ws, x, fila, f"=IF({expr}>0,{expr},0)")

        _escribir(ws, y, fila, _formula_o_cero(dir_dm5.get(("total_declarado", mes))))
        _escribir(ws, z, fila, _formula_suma_o_cero(
            _dir_casillero(dir_f104, mapa_periodo, mes, "615"),
            _dir_casillero(dir_f104, mapa_periodo, mes, "617"),
        ))
        _escribir(ws, aa, fila, _formula_o_cero(_dir_casillero(dir_f104, mapa_periodo, mes, "699")))

        # --- {27} AB = Y-B-C-D-E ---
        _escribir(ws, ab, fila, f"={y}{fila}-{b}{fila}-{c_}{fila}-{d}{fila}-{e}{fila}")
        # --- {28} AC = Z-W ---
        _escribir(ws, ac, fila, f"={z}{fila}-{w}{fila}")
        # --- {29} AD = AA-X ---
        _escribir(ws, ad, fila, f"={aa}{fila}-{x}{fila}")

        _escribir(ws, COL_AF, fila, _formula_o_cero(
            _dir_casillero(dir_f104, mapa_periodo, mes, "609")))

        fila_anterior = fila

    fila_leyenda = FILA_PRIMER_MES + len(MESES) + 2
    escribir_leyenda_marcas(ws, fila=fila_leyenda)

    ws.column_dimensions[get_column_letter(COL_MES)].width = 14
    for n in range(1, 30):
        ws.column_dimensions[letra_columna(n)].width = 16
    ws.column_dimensions[COL_AF].width = 16

    salida: dict[tuple[str, str], str] = {}
    for idx, mes in enumerate(MESES):
        fila = FILA_PRIMER_MES + idx
        for n in range(1, 30):
            salida[(f"col_{n}", mes)] = f"'{SHEET_DM6}'!{letra_columna(n)}{fila}"
        salida[("col_AF", mes)] = f"'{SHEET_DM6}'!{COL_AF}{fila}"
    return salida
