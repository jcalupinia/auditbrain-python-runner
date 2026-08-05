"""DM7 Retenciones por pagar — retenciones de IVA (F-104) y de renta (F-103).

Dos bloques independientes, cada uno con el mismo patrón que DM4/DM5:

Bloque 1 — Retenciones de IVA: una fila por cuenta de la categoría RET_IVA
(libros) → *Según libros* → una fila por casillero declarado (721, 723, 725,
727, 729, 731) → *Según declaraciones* → el casillero 799 como CONTROL del
total (no se suma al rango de declaraciones: es la cifra que el propio F-104
reporta como total retenido, y sirve para que el auditor verifique que la
suma de los seis casilleros cuadra con el total que declaró el SRI) →
*Diferencia* (libros − declarado).

Bloque 2 — Retenciones de renta: una fila por cuenta de la categoría
RET_RENTA (libros) → *Según libros* → el casillero 499 del F-103 (único,
no hay varios casilleros que sumar) → *Según declaraciones* → *Diferencia*.

Publica ("ret_iva_declarado", mes) y ("ret_renta_declarado", mes), que DM3
consume para el pasivo de SRI por pagar al cierre.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.bloques import (
    COL_PRIMER_MES, MESES, escribir_encabezado_meses, fila_diferencia,
    fila_referencias, fila_suma_rango,
)
from backend.app.aud.obligaciones_fiscales.libro.estilos import (
    escribir_encabezado_cedula, escribir_leyenda_marcas,
)

SHEET_DM7 = "DM7 Retenciones x pagar"
CASILLEROS_RET_IVA = ["721", "723", "725", "727", "729", "731"]
CASILLERO_RET_IVA_CONTROL = "799"
CASILLERO_RET_RENTA = "499"


def _dirs_cuenta(dir_mayores: dict, codigo: str) -> dict[str, str]:
    return {
        mes: dir_mayores[(f"cuenta:{codigo}", mes)]
        for mes in MESES
        if (f"cuenta:{codigo}", mes) in dir_mayores
    }


def _dirs_casillero(dir_datos: dict, periodos: list[str], cas: str) -> dict[str, str]:
    salida = {}
    for periodo in periodos:
        mes = periodo.split("-")[-1]
        addr = dir_datos.get((periodo, cas))
        if addr:
            salida[mes] = addr
    return salida


def _bloque_cuentas(ws, *, fila: int, titulo: str, cuentas: list[str],
                    dir_mayores: dict, nombres_cuenta: dict) -> tuple[int, int]:
    """Encabezado + una fila por cuenta + 'Según libros'. Devuelve (fila_siguiente, fila_libros)."""
    escribir_encabezado_meses(ws, fila=fila, titulo=titulo)
    fila += 1
    primera = fila
    for codigo in cuentas:
        fila_referencias(ws, fila=fila, etiqueta=nombres_cuenta.get(codigo, codigo),
                         direcciones=_dirs_cuenta(dir_mayores, codigo))
        fila += 1
    ultima = fila - 1
    fila_libros = fila
    fila_suma_rango(ws, fila=fila_libros, etiqueta="Según libros", desde=primera, hasta=ultima)
    fila += 2
    return fila, fila_libros


def build_dm7(
    wb: Workbook,
    *,
    dir_mayores: dict,
    dir_f104: dict,
    dir_f103: dict,
    periodos: list[str],
    nombres_cuenta: dict[str, str],
    cliente: str,
    periodo: str,
    preparado_por: str | None = None,
    revisado_por: str | None = None,
) -> dict[tuple[str, str], str]:
    """Construye DM7. Devuelve {("ret_iva_declarado"|"ret_renta_declarado", mes) → addr}."""
    if SHEET_DM7 in wb.sheetnames:
        del wb[SHEET_DM7]
    ws = wb.create_sheet(SHEET_DM7)

    escribir_encabezado_cedula(
        ws, titulo="Retenciones por pagar de IVA y de la fuente", referencia="DM7",
        cliente=cliente, periodo=periodo,
        preparado_por=preparado_por, revisado_por=revisado_por,
    )

    cuentas_ret_iva = dir_mayores.get(("orden:RET_IVA", "cuentas"), [])
    cuentas_ret_renta = dir_mayores.get(("orden:RET_RENTA", "cuentas"), [])

    fila = 13

    # --- Bloque 1: Retenciones de IVA ---
    fila, fila_libros_iva = _bloque_cuentas(
        ws, fila=fila, titulo="RETENCIONES DE IVA", cuentas=cuentas_ret_iva,
        dir_mayores=dir_mayores, nombres_cuenta=nombres_cuenta,
    )
    primer_cas = fila
    for cas in CASILLEROS_RET_IVA:
        fila_referencias(ws, fila=fila, etiqueta=f"Casillero {cas}",
                         direcciones=_dirs_casillero(dir_f104, periodos, cas))
        fila += 1
    fila_declarado_iva = fila
    fila_suma_rango(ws, fila=fila_declarado_iva, etiqueta="Según declaraciones",
                    desde=primer_cas, hasta=fila - 1)
    fila += 1

    fila_referencias(
        ws, fila=fila, etiqueta=f"Casillero {CASILLERO_RET_IVA_CONTROL} (control del total)",
        direcciones=_dirs_casillero(dir_f104, periodos, CASILLERO_RET_IVA_CONTROL),
    )
    fila += 1

    fila_diferencia(ws, fila=fila, etiqueta="Diferencia",
                    fila_libros=fila_libros_iva, fila_declarado=fila_declarado_iva)
    fila += 3

    # --- Bloque 2: Retenciones de renta ---
    fila, fila_libros_renta = _bloque_cuentas(
        ws, fila=fila, titulo="RETENCIONES DE RENTA", cuentas=cuentas_ret_renta,
        dir_mayores=dir_mayores, nombres_cuenta=nombres_cuenta,
    )
    primer_cas_renta = fila
    fila_referencias(ws, fila=fila, etiqueta=f"Casillero {CASILLERO_RET_RENTA}",
                     direcciones=_dirs_casillero(dir_f103, periodos, CASILLERO_RET_RENTA))
    fila += 1
    fila_declarado_renta = fila
    fila_suma_rango(ws, fila=fila_declarado_renta, etiqueta="Según declaraciones",
                    desde=primer_cas_renta, hasta=fila - 1)
    fila += 1

    fila_diferencia(ws, fila=fila, etiqueta="Diferencia",
                    fila_libros=fila_libros_renta, fila_declarado=fila_declarado_renta)
    fila += 3

    escribir_leyenda_marcas(ws, fila=fila)

    for col, ancho in ((1, 24), (2, 30)):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    for j in range(13):
        ws.column_dimensions[get_column_letter(COL_PRIMER_MES + j)].width = 15

    salida: dict[tuple[str, str], str] = {}
    for clave, fila_origen in (
        ("ret_iva_declarado", fila_declarado_iva),
        ("ret_renta_declarado", fila_declarado_renta),
    ):
        salida.update({
            (clave, m): f"'{SHEET_DM7}'!{get_column_letter(COL_PRIMER_MES + j)}{fila_origen}"
            for j, m in enumerate(MESES)
        })
    return salida
