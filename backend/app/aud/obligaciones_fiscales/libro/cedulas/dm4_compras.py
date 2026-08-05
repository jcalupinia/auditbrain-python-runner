"""DM4 Compras — IVA en compras y base imponible, libros vs declaraciones."""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.bloques import (
    COL_PRIMER_MES, COL_TOTAL, MESES, escribir_encabezado_meses, fila_diferencia,
    fila_referencias, fila_suma_direcciones, fila_suma_rango,
)
from backend.app.aud.obligaciones_fiscales.libro.estilos import (
    FONT_DATA, FORMATO_NUM, escribir_encabezado_cedula, escribir_leyenda_marcas,
)

SHEET_DM4 = "DM4 Compras"
CASILLEROS_IVA = ["520", "521", "522", "523", "524", "525", "526", "555", "560"]
CASILLEROS_BASE = ["510", "511", "512"]
TARIFA_POR_DEFECTO = 0.15


def _dirs_cuenta(dir_mayores: dict, codigo: str) -> dict[str, str]:
    return {
        mes: dir_mayores[(f"cuenta:{codigo}", mes)]
        for mes in MESES
        if (f"cuenta:{codigo}", mes) in dir_mayores
    }


def _dirs_casillero(dir_f104: dict, periodos: list[str], cas: str) -> dict[str, str]:
    salida = {}
    for periodo in periodos:
        mes = periodo.split("-")[-1]
        addr = dir_f104.get((periodo, cas))
        if addr:
            salida[mes] = addr
    return salida


def build_dm4(
    wb: Workbook,
    *,
    dir_mayores: dict,
    dir_f104: dict,
    periodos: list[str],
    nombres_cuenta: dict[str, str],
    cliente: str,
    periodo: str,
    tarifas: dict[str, float] | None = None,
    preparado_por: str | None = None,
    revisado_por: str | None = None,
) -> dict[tuple[str, str], str]:
    """Construye DM4. Devuelve {("libros"|"declarado"|"base", mes) → addr}."""
    if SHEET_DM4 in wb.sheetnames:
        del wb[SHEET_DM4]
    ws = wb.create_sheet(SHEET_DM4)
    tarifas = tarifas or {}

    escribir_encabezado_cedula(
        ws, titulo="Compras e IVA en compras", referencia="DM4",
        cliente=cliente, periodo=periodo,
        preparado_por=preparado_por, revisado_por=revisado_por,
    )

    # --- Bloque 1: IVA en compras ---
    escribir_encabezado_meses(ws, fila=13, titulo="IVA EN COMPRAS")
    fila = 14
    primera_cuenta = fila
    for codigo in dir_mayores.get(("orden:IVA_COMPRAS", "cuentas"), []):
        fila_referencias(ws, fila=fila, etiqueta=nombres_cuenta.get(codigo, codigo),
                         direcciones=_dirs_cuenta(dir_mayores, codigo))
        fila += 1
    ultima_cuenta = fila - 1

    fila_libros = fila
    fila_suma_rango(ws, fila=fila_libros, etiqueta="Según libros",
                    desde=primera_cuenta, hasta=ultima_cuenta)
    fila += 2

    primer_cas = fila
    for cas in CASILLEROS_IVA:
        fila_referencias(ws, fila=fila, etiqueta=f"Casillero {cas}",
                         direcciones=_dirs_casillero(dir_f104, periodos, cas))
        fila += 1
    fila_declarado = fila
    fila_suma_rango(ws, fila=fila_declarado, etiqueta="Según declaraciones",
                    desde=primer_cas, hasta=fila - 1)
    fila += 1

    fila_dif = fila
    fila_diferencia(ws, fila=fila_dif, etiqueta="Diferencia",
                    fila_libros=fila_libros, fila_declarado=fila_declarado)
    fila += 3

    # --- Bloque 2: base imponible = IVA de libros / tarifa del mes ---
    escribir_encabezado_meses(ws, fila=fila, titulo="BASE IMPONIBLE DE COMPRAS")
    fila += 1
    fila_base = fila
    ws.cell(fila, 2, "Compras gravadas (IVA ÷ tarifa)").font = FONT_DATA
    for j, mes in enumerate(MESES):
        col = get_column_letter(COL_PRIMER_MES + j)
        tarifa = tarifas.get(mes, TARIFA_POR_DEFECTO)
        c = ws.cell(fila, COL_PRIMER_MES + j, f"={col}{fila_libros}/{tarifa}")
        c.font = FONT_DATA
        c.number_format = FORMATO_NUM
    ini = get_column_letter(COL_PRIMER_MES)
    fin = get_column_letter(COL_PRIMER_MES + 11)
    ws.cell(fila, COL_TOTAL, f"=SUM({ini}{fila}:{fin}{fila})").number_format = FORMATO_NUM
    fila += 2

    primer_cas_base = fila
    for cas in CASILLEROS_BASE:
        fila_referencias(ws, fila=fila, etiqueta=f"Casillero {cas}",
                         direcciones=_dirs_casillero(dir_f104, periodos, cas))
        fila += 1
    fila_declarado_base = fila
    fila_suma_rango(ws, fila=fila_declarado_base, etiqueta="Según declaraciones",
                    desde=primer_cas_base, hasta=fila - 1)
    fila += 1
    fila_diferencia(ws, fila=fila, etiqueta="Diferencia",
                    fila_libros=fila_base, fila_declarado=fila_declarado_base)
    fila += 3

    escribir_leyenda_marcas(ws, fila=fila)

    for col, ancho in ((1, 24), (2, 30)):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    for j in range(13):
        ws.column_dimensions[get_column_letter(COL_PRIMER_MES + j)].width = 15

    return {
        **{("libros", m): f"'{SHEET_DM4}'!{get_column_letter(COL_PRIMER_MES + j)}{fila_libros}"
           for j, m in enumerate(MESES)},
        **{("base", m): f"'{SHEET_DM4}'!{get_column_letter(COL_PRIMER_MES + j)}{fila_base}"
           for j, m in enumerate(MESES)},
    }
