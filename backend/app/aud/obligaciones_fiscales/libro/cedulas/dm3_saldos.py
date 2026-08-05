"""DM3 Revisión de saldos — tres cifras anuales, libros vs F-104.

A diferencia de DM4..DM7, aquí no hay tabla de 12 meses: cada bloque es UNA
sola cifra anual. El "según libros" es el movimiento ACUMULADO DEL AÑO de la
cuenta (la columna Total del resumen "Mayores homologados"), no el saldo al
cierre.

Tres bloques, cada uno con su propia cuenta y su propio casillero:

| Bloque              | Cuenta (parametrizable)  | Casillero declarado          |
|----------------------|--------------------------|-------------------------------|
| Crédito tributario    | 1.1.5.1.2 (por defecto)  | 615 + 617 de diciembre        |
| IVA Diferido           | 2.1.7.4.2 (por defecto)  | 485 de diciembre              |
| SRI por Pagar           | 2.1.7.5.6 (por defecto)  | 859 de diciembre + retenciones
                                                        de renta de diciembre (DM7)  |

Si el cliente no tiene alguna de esas cuentas en su mayor, el bloque igual
se escribe con 0: no se puede referenciar por fórmula una celda que no
existe, así que se deja el valor en 0 literal y una nota junto al bloque
para que el auditor sepa qué cuenta se esperaba encontrar.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter

from backend.app.aud.obligaciones_fiscales.libro.estilos import (
    BORDE, FONT_DATA, FONT_ENCABEZADO_TABLA, FONT_TITULO_CEDULA, FONT_TOTAL,
    FORMATO_NUM, RELLENO_TOTAL, escribir_encabezado_cedula, escribir_leyenda_marcas,
)

SHEET_DM3 = "DM3 Revisión de saldos"

COL_CODIGO, COL_ETIQUETA, COL_VALOR = 1, 2, 3

DICIEMBRE = "12"


def _periodo_de(periodos: list[str], mes: str) -> str | None:
    for p in periodos:
        if p.split("-")[-1] == mes:
            return p
    return None


def _addr_casillero(dir_f104: dict, periodos: list[str], cas: str, mes: str = DICIEMBRE):
    periodo = _periodo_de(periodos, mes)
    if not periodo:
        return None
    return dir_f104.get((periodo, cas))


def _bloque_saldo(
    ws,
    *,
    fila: int,
    titulo: str,
    codigo_cuenta: str,
    nombre_cuenta: str,
    dir_mayores: dict,
    formula_declarado: str,
    etiqueta_declarado: str,
) -> int:
    """Escribe un bloque de una sola cifra anual. Devuelve la fila siguiente."""
    ws.cell(fila, COL_CODIGO, titulo).font = FONT_TITULO_CEDULA
    fila += 1

    for i, texto in enumerate(("Código", "Cuenta", "Valor US$")):
        c = ws.cell(fila, COL_CODIGO + i, texto)
        c.font = FONT_ENCABEZADO_TABLA
        c.border = BORDE
    fila += 1

    fila_codigo = fila
    ws.cell(fila, COL_CODIGO, codigo_cuenta).font = FONT_DATA
    ws.cell(fila, COL_ETIQUETA, nombre_cuenta).font = FONT_DATA
    fila += 1

    addr_total = dir_mayores.get((f"cuenta:{codigo_cuenta}", "TOTAL"))
    if addr_total:
        ws.cell(fila, COL_ETIQUETA,
               f"Nota: cuenta {codigo_cuenta} presente en el mayor").font = FONT_DATA
    else:
        c = ws.cell(fila, COL_ETIQUETA,
                   f"⚠ Nota: la cuenta {codigo_cuenta} no aparece en el mayor del "
                   "cliente; el auditor debe verificar si debía existir.")
        c.font = FONT_DATA
    fila += 1

    fila_libros = fila
    e = ws.cell(fila, COL_ETIQUETA, "Según libros")
    e.font = FONT_TOTAL
    e.fill = RELLENO_TOTAL
    valor_libros = f"={addr_total}" if addr_total else 0
    v = ws.cell(fila, COL_VALOR, valor_libros)
    v.font = FONT_TOTAL
    v.fill = RELLENO_TOTAL
    v.number_format = FORMATO_NUM
    v.border = BORDE
    fila += 1

    fila_declarado = fila
    e = ws.cell(fila, COL_ETIQUETA, etiqueta_declarado)
    e.font = FONT_DATA
    v = ws.cell(fila, COL_VALOR, formula_declarado)
    v.font = FONT_DATA
    v.number_format = FORMATO_NUM
    v.border = BORDE
    fila += 1

    e = ws.cell(fila, COL_ETIQUETA, "Diferencia")
    e.font = FONT_TOTAL
    v = ws.cell(fila, COL_VALOR, f"=ROUND(C{fila_libros}-C{fila_declarado},2)")
    v.font = FONT_TOTAL
    v.number_format = FORMATO_NUM
    v.border = BORDE
    fila += 3

    return fila


def build_dm3(
    wb,
    *,
    dir_mayores: dict,
    dir_f104: dict,
    dir_dm7: dict,
    periodos: list[str],
    cliente: str,
    periodo: str,
    cuenta_credito_tributario: str = "1.1.5.1.2",
    nombre_credito_tributario: str = "Crédito Tributario IVA",
    cuenta_iva_diferido: str = "2.1.7.4.2",
    nombre_iva_diferido: str = "IVA Diferido",
    cuenta_sri_por_pagar: str = "2.1.7.5.6",
    nombre_sri_por_pagar: str = "SRI por Pagar",
    preparado_por: str | None = None,
    revisado_por: str | None = None,
) -> dict[str, str]:
    """Construye DM3. No publica direcciones: nada la consume por fórmula."""
    if SHEET_DM3 in wb.sheetnames:
        del wb[SHEET_DM3]
    ws = wb.create_sheet(SHEET_DM3)

    escribir_encabezado_cedula(
        ws, titulo="Revisión de saldos", referencia="DM3",
        cliente=cliente, periodo=periodo,
        preparado_por=preparado_por, revisado_por=revisado_por,
    )

    fila = 12

    # --- Bloque 1: Crédito tributario = 615 + 617 de diciembre ---
    addr615 = _addr_casillero(dir_f104, periodos, "615")
    addr617 = _addr_casillero(dir_f104, periodos, "617")
    partes = [a for a in (addr615, addr617) if a]
    formula_credito = ("=" + "+".join(partes)) if partes else 0
    fila = _bloque_saldo(
        ws, fila=fila, titulo="CREDITO TRIBUTARIO",
        codigo_cuenta=cuenta_credito_tributario, nombre_cuenta=nombre_credito_tributario,
        dir_mayores=dir_mayores, formula_declarado=formula_credito,
        etiqueta_declarado="Según F-104 casillero 615+617 (diciembre)",
    )

    # --- Bloque 2: IVA Diferido = 485 de diciembre ---
    addr485 = _addr_casillero(dir_f104, periodos, "485")
    formula_diferido = f"={addr485}" if addr485 else 0
    fila = _bloque_saldo(
        ws, fila=fila, titulo="IVA DIFERIDO",
        codigo_cuenta=cuenta_iva_diferido, nombre_cuenta=nombre_iva_diferido,
        dir_mayores=dir_mayores, formula_declarado=formula_diferido,
        etiqueta_declarado="Según F-104 casillero 485 (diciembre)",
    )

    # --- Bloque 3: SRI por Pagar = 859 de diciembre + retenciones de renta
    # de diciembre (DM7) ---
    addr859 = _addr_casillero(dir_f104, periodos, "859")
    addr_ret_renta_dic = dir_dm7.get(("ret_renta_declarado", DICIEMBRE))
    partes_sri = [a for a in (addr859, addr_ret_renta_dic) if a]
    formula_sri = ("=" + "+".join(partes_sri)) if partes_sri else 0
    fila = _bloque_saldo(
        ws, fila=fila, titulo="PASIVO: SRI POR PAGAR",
        codigo_cuenta=cuenta_sri_por_pagar, nombre_cuenta=nombre_sri_por_pagar,
        dir_mayores=dir_mayores, formula_declarado=formula_sri,
        etiqueta_declarado="Según F-104 casillero 859 (diciembre) + retenciones "
                            "de renta de diciembre (DM7)",
    )

    escribir_leyenda_marcas(ws, fila=fila)

    ws.column_dimensions[get_column_letter(COL_CODIGO)].width = 16
    ws.column_dimensions[get_column_letter(COL_ETIQUETA)].width = 60
    ws.column_dimensions[get_column_letter(COL_VALOR)].width = 16

    return {}
