"""Carga la plantilla baked-in DM Obligaciones Fiscales y la puebla con datos.

Para M1:
- DM6 IVA: columnas C, D, E (casilleros F-104 415, 413, 417) por mes
- DM7 Retenciones: columnas H, I, J, K, L, M (casilleros 721, 723, 725,
  729, 731, 727) por mes
- Encabezado: cliente, periodo en todas las pestañas relevantes

El resto de pestañas (DM, DM1, DM2, DM3, DM4, DM5, DM8, DM9, DM10) quedan
con su contenido original de plantilla. Se completan en M2.
"""

from __future__ import annotations

import datetime
import io
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

TEMPLATE_PATH = Path(__file__).parent / "templates" / "dm_obligaciones_fiscales.xlsx"
ASSETS_DIR = Path(__file__).parent / "assets"

LOGO_PATHS = {
    "audit_consulting": ASSETS_DIR / "logo_audit_consulting.png",
    "partner_auditing": ASSETS_DIR / "logo_partner_auditing.png",
}

# Celda ancla del logo (esquina superior derecha de cada cédula).
# Dimensiones aproximadas en píxeles para que se vea balanceado.
LOGO_ANCHOR = "F1"
LOGO_WIDTH_PX = 140
LOGO_HEIGHT_PX = 60

# Hojas donde se inserta el logo. Todas las cédulas DM* (DM10 tiene layout
# distinto, se evalúa después si el ancla F1 no funciona).
LOGO_SHEETS = [
    "DM  Programa de Auditoria",
    "DM1 Cuestionario de Auditoria ",
    "DM2 Cédula Sumaria",
    "DM3 Revisión de saldos",
    "DM4 Compras ",
    "DM5 Ventas ",
    "DM6 IVA",
    "DM7 Retenciones x pagar",
    "DM8 ATS",
    "DM9 Límite costos y gastos",
    "DM10 Hoja de hallazgos",
]

DM6_SHEET = "DM6 IVA"
DM7_SHEET = "DM7 Retenciones x pagar"
DM6_FIRST_ROW = 20  # Enero
DM7_FIRST_ROW = 21  # Enero

# Mapeo columna del Excel DM6 -> clave del row_data
DM6_COL_MAP = {
    3: "c415",   # C: Ventas tarifa 0% (c/ derecho)
    4: "c413",   # D: Ventas tarifa 0% (s/ derecho)
    5: "c417",   # E: Exportaciones
}

# Mapeo columna del Excel DM7 -> clave del row_data
# Orden según plantilla: H=10%, I=20%, J=30%, K=70%, L=100%, M=50%
DM7_COL_MAP = {
    8: "c721",   # H: 10%
    9: "c723",   # I: 20%
    10: "c725",  # J: 30%
    11: "c729",  # K: 70%
    12: "c731",  # L: 100%
    13: "c727",  # M: 50%
}

# Mapeo de escritura del encabezado por campo del formulario.
# Solo escribimos en SOURCE cells. Las demás hojas reciben el valor via
# fórmulas cross-sheet ya existentes en la plantilla (ej. DM6!A5 lee de
# DM3!A5 que lee de DM2!A5 ... que lee de DM Programa!A5).
HEADER_WRITES = {
    "cliente_name": [
        ("DM  Programa de Auditoria",     "A5"),  # propaga a DM1..DM7 via formula
        ("DM9 Límite costos y gastos",    "A5"),  # source propio
        ("DM10 Hoja de hallazgos",        "B4"),  # source propio
    ],
    "period_end": [
        ("DM  Programa de Auditoria",     "D5"),
        ("DM9 Límite costos y gastos",    "D5"),
        # DM10!D4 es formula -> DM8 ATS!D16:E16; no tocar
    ],
    "prepared_by_name": [
        ("DM  Programa de Auditoria",     "A7"),  # propaga via formula chain
        ("DM9 Límite costos y gastos",    "A7"),
        ("DM10 Hoja de hallazgos",        "B5"),
    ],
    "reviewed_by_name": [
        # Cada hoja tiene su SOURCE (cada cédula podría ser revisada por persona
        # distinta, pero nuestro form recibe UN solo revisor, así que ponemos
        # el mismo en todas las que tienen source).
        ("DM  Programa de Auditoria",     "A9"),
        ("DM1 Cuestionario de Auditoria ", "A9"),
        ("DM2 Cédula Sumaria",            "A9"),
        ("DM3 Revisión de saldos",        "A9"),
        ("DM9 Límite costos y gastos",    "A9"),
        ("DM10 Hoja de hallazgos",        "D5"),
    ],
}


def assemble(
    *,
    cliente_name: str,
    period_label: str,
    period_end: datetime.date | None,
    prepared_by_name: str | None,
    reviewed_by_name: str | None,
    firma_auditora: str | None = None,
    dm6_data: dict,
    dm7_data: dict,
) -> bytes:
    """Carga plantilla, escribe encabezados + DM6 + DM7 + logo, devuelve bytes."""
    wb = load_workbook(TEMPLATE_PATH)

    _write_headers(
        wb,
        cliente_name=cliente_name,
        period_end=period_end,
        prepared_by_name=prepared_by_name,
        reviewed_by_name=reviewed_by_name,
    )

    _insert_logo(wb, firma_auditora)

    if DM7_SHEET in wb.sheetnames:
        _populate_monthly_grid(
            wb[DM7_SHEET], DM7_FIRST_ROW, DM7_COL_MAP, dm7_data.get("rows", [])
        )
    if DM6_SHEET in wb.sheetnames:
        _populate_monthly_grid(
            wb[DM6_SHEET], DM6_FIRST_ROW, DM6_COL_MAP, dm6_data.get("rows", [])
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _insert_logo(wb, firma_auditora: str | None) -> None:
    """Inserta el PNG del logo correspondiente en F1 de cada cédula.

    Si firma_auditora es None o desconocida, o el archivo no existe,
    no inserta nada (silencioso — el resto del Excel se genera igual).
    """
    if not firma_auditora:
        return
    path = LOGO_PATHS.get(firma_auditora)
    if not path or not path.exists():
        return
    for sheet_name in LOGO_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        try:
            # Nueva Image por hoja: openpyxl no permite la misma instancia
            # en múltiples hojas.
            img = XLImage(str(path))
            img.width = LOGO_WIDTH_PX
            img.height = LOGO_HEIGHT_PX
            img.anchor = LOGO_ANCHOR
            wb[sheet_name].add_image(img)
        except Exception:
            # Si openpyxl rechaza la imagen en una hoja específica, seguimos
            # con las demás. No queremos que un logo malo tumbe el job.
            pass


def _write_headers(
    wb,
    *,
    cliente_name: str,
    period_end: datetime.date | None,
    prepared_by_name: str | None,
    reviewed_by_name: str | None,
) -> None:
    """Escribe los campos del form en las SOURCE cells de cada hoja.

    Las hojas que tienen fórmulas hacia las SOURCE reciben el valor
    automáticamente al abrir el Excel — Excel recalcula las fórmulas.
    """
    field_values = {
        "cliente_name": cliente_name,
        "period_end": period_end,
        "prepared_by_name": prepared_by_name,
        "reviewed_by_name": reviewed_by_name,
    }
    for field, targets in HEADER_WRITES.items():
        value = field_values.get(field)
        if value is None or value == "":
            continue
        for sheet_name, coord in targets:
            if sheet_name not in wb.sheetnames:
                continue
            _try_write(wb[sheet_name], coord, value)


def _populate_monthly_grid(
    ws,
    first_row: int,
    col_map: dict[int, str],
    rows: list[dict],
) -> None:
    """Escribe cada fila mensual en su row del Excel.

    rows: lista de 12 dicts (Enero..Diciembre).
    col_map: {col_index: key_in_row_dict}.
    Solo escribe celdas cuando el valor NO es None (para no romper formulas
    o sobreescribir hardcoded zeros).
    """
    for i, row_data in enumerate(rows):
        excel_row = first_row + i
        for col, key in col_map.items():
            v = row_data.get(key)
            if v is not None:
                ws.cell(row=excel_row, column=col, value=v)


def _try_write(ws, coord: str, value) -> None:
    try:
        ws[coord] = value
    except Exception:
        pass
