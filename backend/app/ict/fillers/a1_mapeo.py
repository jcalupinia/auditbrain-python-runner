"""Filler for MAPEO A1 sheet usando Balance Mapeado (casillero pre-asignado).

Metodología REPLICADA del anexo oficial ICT 2024 del SRI:

  - Cada casillero F-101 ocupa una o más filas (una por cada cuenta
    contable que mapea a él).
  - PRIMERA fila del casillero:
      A: número de casillero       B: nombre del casillero
      C: valor declarado F-101     D/E/F: primera cuenta del balance
      G: fórmula DIFERENCIA:
         · 1 sola cuenta  → "=F<row>-C<row>"
         · N cuentas      → "=SUM(F<row>:F<row+N-1>)-C<row>"
      H: observaciones (vacío)
  - FILAS SIGUIENTES del mismo casillero:
      Solo D/E/F (código, descripción, saldo de cada cuenta adicional).
      Cols A, B, C, G, H quedan en BLANCO — la fórmula G de la primera
      fila YA suma todo el rango.
  - Casilleros SIN cuentas en el balance: G = "=F<row>-C<row>" (apunta
    a su propia fila, donde F<row> está vacío → fórmula evaluará a
    -C<row>, evidenciando que falta el saldo contable).
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell

from backend.app.ict.cell_maps.a1 import (
    A1_CASILLEROS_ORDERED,
    A1_FIRST_DATA_ROW,
    A1_HEADER_MAP,
    A1_SHEET,
)


def _safe_set(ws, cell_addr: str, value) -> bool:
    """Wrapper local: delega al central que protege MergedCells + fórmulas
    y registra la escritura en el trace log para la hoja Trazabilidad."""
    from backend.app.ict.fillers.base import safe_set
    return safe_set(ws, cell_addr, value, anexo="A1",
                    origen="A1 Mapeo (F-101 + Balance Mapeado)")


def _safe_set_formula(ws, cell_addr: str, formula: str, *, casillero: str | None = None) -> bool:
    """Wrapper local para fórmulas DELIBERADAS (sobreescribe fórmulas
    viejas del template cuando es necesario, p. ej. =SUM(F13:F25)-C13).
    El parámetro casillero opcional permite registrar el casillero referenciado
    para la sección de COBERTURA del dashboard de VERIFICACIÓN."""
    from backend.app.ict.fillers.base import safe_set_formula
    return safe_set_formula(ws, cell_addr, formula, anexo="A1",
                            casillero=casillero,
                            origen="A1 Mapeo (fórmula referencial)")


class A1Filler:
    anexo_code = "A1"

    # Casilleros que marcan FIN de bloque mayor — después de ellos
    # se inserta una fila en blanco como separador visual.
    BLOQUE_BREAKS = {"361", "499", "550", "599", "698", "699", "1005", "6999", "7999"}
    # Casilleros que son TOTAL (formato negrita + borde doble)
    TOTAL_CASILLEROS = {"361", "449", "499", "550", "589", "599", "698", "699",
                        "1005", "1045", "6999", "7991", "7992", "7999"}

    # Casilleros que F-101 declara como (-) NEGATIVOS (deterioros, depreciaciones,
    # pérdidas, inventarios finales). Sus saldos en el balance del cliente vienen
    # con signo CONTABLE NEGATIVO mientras F-101 los expresa positivos. Para que
    # la fórmula de diferencia dé 0 al cuadrar usamos ABS() sobre el sumatorio.
    #
    # _CORE: casilleros conocidos antes del fix 2026-06-04. Calibrados manualmente.
    # NEGATIVE_CASILLEROS: union de _CORE + auto-detección por prefijo "(-)" en el
    # catálogo OFICIAL F-101. La auto-detección garantiza que cualquier cas del
    # balance que en el SRI sea de naturaleza (-) (deterioro, amortización,
    # depreciación) quede correctamente clasificado, incluso si no estaba en _CORE.
    _NEGATIVE_CORE = {
        "314", "317", "324", "327", "329",  # deterioros cuentas por cobrar
        "347",                                 # deterioro inventarios
        "384", "385", "386",                  # depreciación acumulada PPE
        "392", "393",                          # amortización acumulada intangibles
        "602",                                 # capital no pagado
        "612", "616",                          # pérdidas acumuladas / del ejercicio
        "7010", "7022", "7028", "7034",       # inventarios finales (restan en CoGS)
    }

    @classmethod
    def _build_negative_set(cls) -> set[str]:
        """Combina _NEGATIVE_CORE con cas del catálogo cuyos nombres empiezan con
        '(-)'. Defensivo contra el bug de 'saldos de línea': si SRI agrega un
        nuevo cas negativo, queda clasificado correctamente sin tocar este file."""
        from backend.app.ict.catalogo_f101 import F101_CASILLERO_NAMES
        auto_detected = {
            cas for cas, nombre in F101_CASILLERO_NAMES.items()
            if nombre.strip().startswith("(-)")
            and cas.isdigit() and 311 <= int(cas) <= 699
        }
        return cls._NEGATIVE_CORE | auto_detected

    # Se construye una vez al cargar la clase. Es read-only durante la sesión.
    NEGATIVE_CASILLEROS: set[str] = set()  # se completa abajo

    # Mapping de TOTAL → identificador de bloque para que F<row_total> tenga
    # fórmula SUM del rango del bloque que totaliza.
    # PRIMARIOS suman cuentas individuales; COMPUESTOS suman sub-totales.
    PRIMARY_TOTAL_BLOCKS = {
        "361": "ACT_CORR",      "449": "ACT_NO_CORR",
        "550": "PAS_CORR",      "589": "PAS_NO_CORR",
        "698": "PATRIMONIO",
        "1005": "ING_ORD",      "1045": "ING_NO_OP",
        "7991": "COSTOS_OP",    "7992": "GASTOS",
    }
    COMPOSITE_TOTALS = {
        # cas_total → (sub_total_1, sub_total_2)
        "499": ("361", "449"),
        "599": ("550", "589"),
        "699": ("599", "698"),
        "6999": ("1005", "1045"),    # TOTAL INGRESOS = ord + no operacionales
        "7999": ("7991", "7992"),
    }
    # En qué bloque empieza cada casillero PRIMARIO. Lookup directo por casillero
    # del primer casillero del bloque para resetear block_start_row.
    # IMPORTANTE: estos valores deben coincidir EXACTAMENTE con los primeros
    # casilleros de cada bloque en A1_CASILLEROS_ORDERED.
    BLOCK_FIRST_CAS = {
        "ACT_CORR":    "311",   "ACT_NO_CORR": "362",
        "PAS_CORR":    "511",   "PAS_NO_CORR": "555",   # 555 es el primero en cell_map
        "PATRIMONIO":  "601",
        "ING_ORD":     "6001",  "ING_NO_OP":   "6033",
        "COSTOS_OP":   "7001",  "GASTOS":      "7173",
    }

    # Rangos de casilleros que en el F-101 figuran en POSITIVO pero el sistema
    # contable del cliente PUEDE traerlos en NEGATIVO (convención contable:
    # créditos = negativos). Para que la columna F del A1 sea visualmente
    # comparable con el F-101, normalizamos esos saldos a POSITIVO via ABS().
    # Esto cubre: pasivos corrientes/no corrientes/patrimonio + todas las
    # cuentas "(-)" del activo (deterioros, depreciaciones, amortizaciones).
    @classmethod
    def _needs_abs_normalization(cls, casillero: str) -> bool:
        """[LEGACY] True si el saldo del balance debe normalizarse a positivo
        para coincidir con el signo del F-101. Reemplazado por la regla
        unificada de signos pero se mantiene por compatibilidad."""
        if casillero in cls.NEGATIVE_CASILLEROS:
            return True
        if not casillero.isdigit():
            return False
        num = int(casillero)
        if 511 <= num <= 599:
            return True
        if 601 <= num <= 698:
            return True
        return False

    @classmethod
    def _is_pasivo_o_patrimonio(cls, casillero: str) -> bool:
        """True si el casillero pertenece al bloque PASIVOS (511-599) o
        PATRIMONIO (601-698). REGLA: estos casilleros requieren INVERTIR
        el signo del balance contable (créditos negativos → positivos en
        el A1 para coincidir con la convención SRI del F-101)."""
        if not casillero.isdigit():
            return False
        num = int(casillero)
        if 511 <= num <= 599:
            return True
        if 601 <= num <= 698:
            return True
        return False

    @classmethod
    def _is_ingreso_estado_resultados(cls, casillero: str) -> bool:
        """True si el casillero pertenece a INGRESOS (6001-6999).

        REGLA cliente (2026-06-05): los ingresos en el balance vienen como
        CRÉDITOS (saldos negativos por convención contable). El F-101 los
        declara en POSITIVO. Para que la columna F del A1 coincida con la
        columna C (F-101 declarado), hay que INVERTIR el signo del balance.

        Ejemplo: cas 6001 (VENTAS NETAS LOCALES)
          - Balance contable: -5,069,641.37 (crédito)
          - F-101 declarado: 5,069,641.37 (positivo)
          - A1 col F debe mostrar: 5,069,641.37 → fórmula =-'DATOS BALANCE'!D{row}

        Excepción: cas con (-) en el nombre (ej. cas 6147 DESCUENTOS) ya
        están en NEGATIVE_CASILLEROS y mantienen el flujo is_negative
        (=-ABS) — esos representan devoluciones/descuentos que conceptual-
        mente RESTAN al ingreso, deben quedar negativos en A1.
        """
        if not casillero.isdigit():
            return False
        num = int(casillero)
        return 6001 <= num <= 6999

    @classmethod
    def _build_signed_sum_formula(
        cls,
        col: str,
        componentes: list[str],
        casillero_to_row: dict[str, int],
    ) -> str | None:
        """Construye una fórmula tipo `=+C13+C22-C25-C29+...` que SUMA los
        componentes del bloque, RESTANDO los casilleros que están en
        NEGATIVE_CASILLEROS (cuentas (-) acumuladas que conceptualmente
        restan: deterioros, depreciaciones, amortizaciones, pérdidas).

        Args:
            col: 'C' (declarado) o 'F' (contable).
            componentes: lista de casilleros del bloque en orden.
            casillero_to_row: lookup cas → fila A1.

        Returns:
            Fórmula `=+Cxx+Cyy-Czz...` o None si no hay componentes con fila.
        """
        if not componentes:
            return None
        parts: list[str] = []
        for cas in componentes:
            row = casillero_to_row.get(cas)
            if row is None:
                continue
            sign = "-" if cas in cls.NEGATIVE_CASILLEROS else "+"
            parts.append(f"{sign}{col}{row}")
        if not parts:
            return None
        # Resulta en p.ej. "=+C13+C22-C25+C26..." — el signo + inicial es válido
        return "=" + "".join(parts)

    def fill(
        self,
        workbook: Workbook,
        session_data: dict,
        anexo_data: dict,
    ) -> dict:
        ws = workbook[A1_SHEET]
        filled = 0
        warnings: list[str] = []

        # ⚠ FIX BUG MERGED CELLS:
        # El template del SRI tiene merged cells PRE-DEFINIDAS en el rango de
        # datos (filas 13+) para algunos casilleros. Cuando el filler hace
        # `ws.insert_rows()` para cuentas extra de un casillero con múltiples
        # cuentas (ej. cas 311 con 7 bancos), todos los merges del template
        # se DESPLAZAN hacia abajo. Esto causa que safe_set() omita silenciosa-
        # mente la escritura de A/B/C en otros casilleros (cas 321, 322, 325,
        # 327, etc.) porque la celda destino quedó dentro de un rango merged
        # heredado del template. RESULTADO visible: el A1 muestra el número
        # del casillero (A) pero el nombre (B) y el valor (C) vacíos.
        #
        # SOLUCIÓN: un-merge TODOS los rangos por debajo del header. El filler
        # tendrá libertad total para escribir; el formatter al final crea las
        # merges DESEADAS (A:A, C:C por grupo de casillero).
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row >= A1_FIRST_DATA_ROW:
                ws.unmerge_cells(str(mr))

        # ⚠ FIX bug "notas SRI residuales del template" (2026-06-05):
        # El template del SRI trae filas con texto explicativo
        # (ej. "a) Corresponde al número, nombre y valor del casillero...")
        # entre los bloques de cas. Cuando el filler escribe la lista de cas,
        # esas filas quedan en medio del A1. El cliente PROPHAR reportó esto
        # en ICT_9_PAPEL_TRABAJO.xlsx fila 52.
        # SOLUCIÓN: limpiar TODAS las celdas A-H desde A1_FIRST_DATA_ROW hasta
        # el final del template antes de empezar a escribir cas. El filler
        # tendrá un canvas limpio y el formato se aplica al final.
        last_row = ws.max_row
        for r in range(A1_FIRST_DATA_ROW, last_row + 1):
            for c in range(1, 9):  # A-H
                cell = ws.cell(r, c)
                if cell.value is not None:
                    cell.value = None

        # Header
        for cell_addr, key in A1_HEADER_MAP.items():
            if _safe_set(ws, cell_addr, session_data.get(key, "")):
                filled += 1

        f101 = anexo_data.get("f101", {})
        # balance_mapeado is a list of {casillero_sri, codigo, descripcion, saldo}
        balance_mapeado = anexo_data.get("balance_mapeado", [])

        # Lookups a las hojas DATOS — permiten que C y F sean FÓRMULAS
        # referenciales en vez de valores literales.
        f101_lookup: dict[str, int] = anexo_data.get("_f101_lookup", {}) or {}
        # balance_lookup[i] = row donde se escribió la cuenta i en DATOS BALANCE
        balance_lookup: list[int] = anexo_data.get("_balance_lookup", []) or []

        # Group balance items by casillero_sri + recordar su índice original
        # para poder generar la fórmula =DATOS BALANCE!D<row> correcta.
        by_casillero: dict[str, list[dict]] = {}
        for idx, item in enumerate(balance_mapeado):
            cas = str(item.get("casillero_sri", "")).strip()
            if cas:
                item_with_idx = dict(item)
                item_with_idx["_source_row"] = (
                    balance_lookup[idx] if idx < len(balance_lookup) else None
                )
                by_casillero.setdefault(cas, []).append(item_with_idx)

        # Tracking de grupos para formatting posterior
        casillero_groups: list[dict] = []
        current_row = A1_FIRST_DATA_ROW

        # Tracking de bloques para que los TOTALES tengan fórmulas =SUM correctas
        # en la columna F (saldos contables), igualando lo que pasa en C (declarado).
        # block_start_rows[bloque_id] = fila donde empieza el bloque
        # total_rows[casillero_total] = fila donde se escribió ese TOTAL
        # casillero_to_row[cas] = fila donde se escribió ese cas (para SUM con signos)
        # bloque_to_casilleros[bloque_id] = lista ordenada de casilleros del bloque
        block_start_rows: dict[str, int] = {}
        total_rows: dict[str, int] = {}
        casillero_to_row: dict[str, int] = {}
        bloque_to_casilleros: dict[str, list[str]] = {}
        current_bloque_id: str | None = None

        # REGLA cliente (2026-06-05): A1 solo debe contener cas con saldo
        # (en F-101 o balance contable) + ocultar los informativos. Los
        # TOTALES siempre se muestran (cuadratura).
        # Razón: el SRI necesita identificar los cas del balance vs F-101,
        # no los meramente informativos.
        from backend.app.ict.fillers.source_data_sheets import (
            _es_informativo, F101_TOTALES,
        )

        def _cas_es_relevante_a1(cas: str, nombre: str) -> bool:
            """¿Este cas debe aparecer en A1?"""
            if cas in self.TOTAL_CASILLEROS or cas in F101_TOTALES:
                return True  # TOTAL siempre
            if _es_informativo(nombre):
                return False  # Nunca informativos
            # Tiene saldo en F-101?
            v = f101.get(cas)
            try:
                f101_val = float(v) if v not in (None, "") else 0
            except (TypeError, ValueError):
                f101_val = 0
            if f101_val != 0:
                return True
            # Tiene saldo contable?
            for item in by_casillero.get(cas, []):
                try:
                    if float(item.get("saldo") or 0) != 0:
                        return True
                except (TypeError, ValueError):
                    pass
            return False

        # Pre-pass: determinar el conjunto de cas que SÍ se van a emitir y
        # cuál es el PRIMER cas emitido de cada bloque. Esto permite que
        # block_start_rows se registre correctamente aunque el cas
        # hardcoded como first_cas se haya filtrado.
        cas_a_emitir: set[str] = set()
        for cas, nombre in A1_CASILLEROS_ORDERED:
            if _cas_es_relevante_a1(cas, nombre):
                cas_a_emitir.add(cas)
        # first_emitted_per_block[bloque_id] = primer cas EMITIDO de ese bloque
        first_emitted_per_block: dict[str, str] = {}
        for cas, nombre in A1_CASILLEROS_ORDERED:
            if cas not in cas_a_emitir:
                continue
            if not cas.isdigit():
                continue
            n = int(cas)
            # Map cas → bloque por rango
            bloque = None
            if 311 <= n <= 360: bloque = "ACT_CORR"
            elif 362 <= n <= 448: bloque = "ACT_NO_CORR"
            elif 511 <= n <= 549 or n == 593: bloque = "PAS_CORR"
            elif 551 <= n <= 588 or (590 <= n <= 598 and n != 593): bloque = "PAS_NO_CORR"
            elif 601 <= n <= 697: bloque = "PATRIMONIO"
            elif 6001 <= n <= 6018: bloque = "ING_ORD"
            elif 6019 <= n <= 6998: bloque = "ING_NO_OP"
            elif 7001 <= n <= 7172: bloque = "COSTOS_OP"
            elif 7173 <= n <= 7990: bloque = "GASTOS"
            if bloque and bloque not in first_emitted_per_block:
                first_emitted_per_block[bloque] = cas

        for casillero, casillero_nombre in A1_CASILLEROS_ORDERED:
            # FILTRO REGLA cliente: solo cas relevantes en A1.
            if casillero not in cas_a_emitir:
                continue

            valor_declarado = f101.get(casillero)
            matching = by_casillero.get(casillero, [])
            n_accounts = len(matching)
            is_total = casillero in self.TOTAL_CASILLEROS
            is_negative = casillero in self.NEGATIVE_CASILLEROS
            row_start = current_row

            # Detectar inicio de bloque usando el PRIMER CAS EMITIDO de cada
            # bloque (no el first_cas hardcoded, que podría haberse filtrado).
            for bloque_id, primer_cas_emitido in first_emitted_per_block.items():
                if casillero == primer_cas_emitido and bloque_id not in block_start_rows:
                    block_start_rows[bloque_id] = current_row
                    current_bloque_id = bloque_id
                    bloque_to_casilleros.setdefault(bloque_id, [])

            # Registrar el cas en el bloque actual (excepto el propio TOTAL)
            if current_bloque_id and not is_total:
                bloque_to_casilleros[current_bloque_id].append(casillero)

            # Track posición fila para este cas — usado en la fórmula del TOTAL
            casillero_to_row[casillero] = current_row

            # === Casillero + nombre (cols A, B) ===
            if _safe_set(ws, f"A{current_row}", casillero):
                filled += 1
            if _safe_set(ws, f"B{current_row}", casillero_nombre):
                filled += 1

            # === REGLA UNIFICADA DE SIGNOS (CLAUDE.md / regla del usuario) ===
            # "Tienen que tener el mismo criterio entre los saldos de la
            # declaración de impuesto a la renta y lo contable."
            #
            # Col C — Valor declarado (F-101):
            #   · Cas normal             → '=DATOS F-101'!Cxxx  (positivo del PDF)
            #   · Cas (-) NEGATIVO       → =-DATOS F-101!Cxxx   (invierte signo)
            #     (el PDF SRI trae el deterioro/depreciación/pérdida en POSITIVO,
            #     pero conceptualmente RESTA → mostrarlo en NEGATIVO en A1)
            #
            # Col F — Saldo contable (Balance del cliente):
            #   · Activo normal          → '=DATOS BALANCE'!Dxxx (signo balance)
            #   · Pasivo/Patrimonio      → =-DATOS BALANCE!Dxxx  (invierte signo)
            #     (balance trae créditos negativos; F-101 los muestra positivos)
            #   · Cas (-) del activo     → =-ABS(DATOS BALANCE!Dxxx) (siempre neg)
            #     (deterioro/depreciación, ambos signos posibles en balance,
            #     siempre debe quedar NEGATIVO para coincidir con C)
            #   · Cas (-) del patrimonio (612 Pérdidas) → mismo: =-ABS(...)
            #
            # Resultado: C y F siempre con el MISMO SIGNO para cada casillero.
            # Por eso los TOTALES son SUMA SIMPLE: =SUM(C13:C69) y =SUM(F13:F69).

            # === C — Valor declarado del F-101 (con signo según nombre) ===
            if is_total and casillero in self.COMPOSITE_TOTALS:
                sub1, sub2 = self.COMPOSITE_TOTALS[casillero]
                if sub1 in total_rows and sub2 in total_rows:
                    _safe_set_formula(ws, f"C{current_row}",
                                      f"=C{total_rows[sub1]}+C{total_rows[sub2]}")
                    filled += 1
                elif casillero in f101_lookup:
                    _safe_set_formula(
                        ws, f"C{current_row}",
                        f"='DATOS F-101'!C{f101_lookup[casillero]}",
                    )
                    filled += 1
            elif is_total and casillero in self.PRIMARY_TOTAL_BLOCKS:
                # TOTAL primario: SUMA simple del bloque (los signos ya están
                # aplicados en cada cas individual).
                bloque_id = self.PRIMARY_TOTAL_BLOCKS[casillero]
                first_row = block_start_rows.get(bloque_id)
                if first_row:
                    _safe_set_formula(
                        ws, f"C{current_row}",
                        f"=SUM(C{first_row}:C{current_row-1})",
                        casillero=casillero,
                    )
                    filled += 1
                elif casillero in f101_lookup:
                    # Fallback 1: bloque sin detalles emitidos (filtro
                    # de relevancia descartó todo). Referencia directa F-101.
                    _safe_set_formula(
                        ws, f"C{current_row}",
                        f"='DATOS F-101'!C{f101_lookup[casillero]}",
                        casillero=casillero,
                    )
                    filled += 1
                elif valor_declarado is not None:
                    # Fallback 2: sin lookup (test sin DATOS F-101). Valor
                    # literal — preserva la cuadratura aunque sin fórmula.
                    _safe_set(ws, f"C{current_row}", float(valor_declarado))
                    filled += 1
                current_bloque_id = None  # cerrar bloque
            elif casillero in f101_lookup:
                # Casillero normal: referencia con signo según nombre.
                # is_negative → multiplicar por -1 (cuenta acumulada que resta).
                base_ref = f"'DATOS F-101'!C{f101_lookup[casillero]}"
                formula_c = f"=-{base_ref}" if is_negative else f"={base_ref}"
                _safe_set_formula(
                    ws, f"C{current_row}", formula_c,
                    casillero=casillero,
                )
                filled += 1
            elif valor_declarado is not None:
                # Fallback: si el casillero no está en F-101 lookup
                val_with_sign = -abs(float(valor_declarado)) if is_negative else valor_declarado
                if _safe_set(ws, f"C{current_row}", val_with_sign): filled += 1
            else:
                if not is_total:
                    warnings.append(f"Casillero {casillero} no encontrado en F-101")

            # === F — Saldo contable del Balance (con signo unificado) ===
            # Para TOTALES: SUMA simple, ya que los componentes tienen signo correcto.
            if is_total:
                f_formula_for_total = None
                if casillero in self.PRIMARY_TOTAL_BLOCKS:
                    bloque_id = self.PRIMARY_TOTAL_BLOCKS[casillero]
                    first_row = block_start_rows.get(bloque_id)
                    if first_row:
                        f_formula_for_total = f"=SUM(F{first_row}:F{current_row-1})"
                    elif casillero in f101_lookup:
                        # Fallback 1: bloque sin detalles emitidos → ref F-101.
                        f_formula_for_total = f"='DATOS F-101'!C{f101_lookup[casillero]}"
                    elif valor_declarado is not None:
                        # Fallback 2: sin lookup ni detalles → fórmula refiriendo
                        # a la celda C de la misma fila (asume balance = declarado).
                        f_formula_for_total = f"=C{current_row}"
                    if bloque_id in block_start_rows:
                        del block_start_rows[bloque_id]
                elif casillero in self.COMPOSITE_TOTALS:
                    sub1, sub2 = self.COMPOSITE_TOTALS[casillero]
                    if sub1 in total_rows and sub2 in total_rows:
                        f_formula_for_total = (
                            f"=F{total_rows[sub1]}+F{total_rows[sub2]}"
                        )

                if f_formula_for_total:
                    _safe_set_formula(ws, f"F{current_row}", f_formula_for_total)
                    filled += 1
                total_rows[casillero] = current_row

            if n_accounts == 0:
                # Sin cuentas contables.
                # Regla unificada: G = F - C (mismo signo en ambas columnas).
                _safe_set_formula(ws, f"G{current_row}", f"=F{current_row}-C{current_row}")

                if (not is_total and valor_declarado is not None and
                    valor_declarado != 0):
                    warnings.append(
                        f"Casillero {casillero} ({casillero_nombre}): declarado "
                        f"{valor_declarado:,.2f} pero el Balance Mapeado no aporta "
                        f"cuentas contables — diferencia será -{valor_declarado:,.2f}"
                    )
                # Trackear grupo (1 fila)
                casillero_groups.append({
                    "casillero": casillero, "row_start": row_start,
                    "row_end": current_row, "is_total": is_total,
                })
                current_row += 1
                # Separador después de bloques mayores
                if casillero in self.BLOQUE_BREAKS:
                    current_row += 1  # fila en blanco
                continue

            # === Primera fila: primera cuenta en D/E/F + fórmula G ===
            first = matching[0]
            if _safe_set(ws, f"D{current_row}", first.get("codigo", "")): filled += 1
            if _safe_set(ws, f"E{current_row}", first.get("descripcion", "")): filled += 1

            # REGLA UNIFICADA DE SIGNOS — col F debe coincidir con col C:
            #   · is_negative (cas (-) acumulado) → SIEMPRE NEGATIVO: =-ABS(...)
            #     (el balance puede traerlo + o -, fijamos NEGATIVO para
            #     coincidir con el signo de C que también es NEGATIVO).
            #     Cubre: deterioros, depreciaciones, amortizaciones,
            #     pérdidas acumuladas, devoluciones/descuentos en ventas
            #     (cas 6147), reversiones (cas 7299/7300).
            #   · Pasivo o Patrimonio normal      → INVERTIR signo balance: =-(...)
            #     (balance trae créditos negativos; col F los muestra positivos
            #     igual que el F-101)
            #   · Ingreso 6001-6999 (excepto los is_negative) → INVERTIR
            #     signo balance: =-(...) (balance trae créditos negativos,
            #     F-101 los declara positivos)
            #     [regla agregada 2026-06-05 a pedido del cliente]
            #   · Activo normal y Costo/Gasto normal → mantener signo: =(...)
            is_pas_pat = self._is_pasivo_o_patrimonio(casillero)
            is_ingreso = self._is_ingreso_estado_resultados(casillero)
            src_row = first.get("_source_row")

            def _formula_f(ref_or_value: str, is_literal: bool = False) -> str | float:
                """Devuelve fórmula F con el signo aplicado según las reglas.

                FIX 2026-06-05 (reporte cliente): antes usábamos `=-{ref}` para
                pasivo/patrimonio/ingreso asumiendo que el balance los traía
                NEGATIVOS (crédito). Pero PROPHAR los carga con signo VARIABLE
                (algunos negativos, otros ya normalizados a positivo). Esto
                generaba signo invertido erróneamente cuando el saldo venía
                positivo (resultado: doble del valor en col G de diferencia).

                Solución: usar `=ABS(...)` en vez de `=-{ref}` para
                NORMALIZAR a positivo sin importar el signo de entrada. La
                cuadratura sigue siendo correcta porque ABS(-X) = X = -(-X)
                y ABS(+X) = X.
                """
                if is_literal:
                    val = float(ref_or_value) if ref_or_value else 0
                    if is_negative:
                        return -abs(val)
                    if is_pas_pat or is_ingreso:
                        return abs(val)
                    return val
                if is_negative:
                    return f"=-ABS({ref_or_value})"
                if is_pas_pat or is_ingreso:
                    return f"=ABS({ref_or_value})"
                return f"={ref_or_value}"

            if src_row:
                base_ref = f"'DATOS BALANCE'!D{src_row}"
                _safe_set_formula(ws, f"F{current_row}", _formula_f(base_ref))
                filled += 1
            else:
                saldo_f = _formula_f(first.get("saldo", 0), is_literal=True)
                if _safe_set(ws, f"F{current_row}", saldo_f):
                    filled += 1

            # Fórmula G (Diferencia) — ahora SIMPLE: F-C porque ambos tienen
            # el mismo signo (regla unificada). Sin ABS porque F ya viene
            # con signo correcto.
            if n_accounts > 1:
                end_row = current_row + n_accounts - 1
                formula = f"=SUM(F{current_row}:F{end_row})-C{current_row}"
            else:
                formula = f"=F{current_row}-C{current_row}"
            _safe_set_formula(ws, f"G{current_row}", formula)

            # === Filas adicionales: cuentas extra del mismo casillero ===
            for offset, item in enumerate(matching[1:], start=1):
                row_n = current_row + offset
                ws.insert_rows(row_n)
                if _safe_set(ws, f"D{row_n}", item.get("codigo", "")): filled += 1
                if _safe_set(ws, f"E{row_n}", item.get("descripcion", "")): filled += 1
                src_row = item.get("_source_row")
                if src_row:
                    base_ref = f"'DATOS BALANCE'!D{src_row}"
                    _safe_set_formula(ws, f"F{row_n}", _formula_f(base_ref))
                    filled += 1
                else:
                    saldo_f = _formula_f(item.get("saldo", 0), is_literal=True)
                    if _safe_set(ws, f"F{row_n}", saldo_f):
                        filled += 1

            # Trackear grupo (N filas)
            casillero_groups.append({
                "casillero": casillero,
                "row_start": row_start,
                "row_end": current_row + n_accounts - 1,
                "is_total": is_total,
            })

            current_row += n_accounts
            # Separador después de bloques mayores
            if casillero in self.BLOQUE_BREAKS:
                current_row += 1  # fila en blanco entre bloques

        # === Reporte de cobertura para warnings ===
        casilleros_a1_set = {c for c, _ in A1_CASILLEROS_ORDERED}

        # Casilleros del balance que NO aparecen en el A1 (quedan disponibles
        # para otros anexos via shared_context, pero son útiles de listar):
        extra_casilleros = set(by_casillero.keys()) - casilleros_a1_set
        if extra_casilleros:
            cuentas_huerfanas = sum(len(by_casillero[c]) for c in extra_casilleros)
            warnings.append(
                f"Balance Mapeado: {len(extra_casilleros)} casilleros con "
                f"{cuentas_huerfanas} cuentas NO mapean al A1 (se trasladan a "
                f"A2-A9 vía shared_context): {sorted(extra_casilleros)[:10]}"
                f"{'...' if len(extra_casilleros) > 10 else ''}"
            )

        # === REGLA RUNTIME: cada TOTAL DEBE tener fórmula en F y valor en C ===
        # Si esta validación falla, significa que en algún punto el filler
        # NO escribió la fórmula SUM en F<row_total> o no recibió el valor
        # declarado en C, y la VERIFICACIÓN A1 saldrá con columnas vacías
        # (el bug que reportó el usuario). Esto NO falla el job — sólo
        # agrega warnings explícitos al output para que el auditor sepa.
        # El test en tests/test_ict_a1_totales_regla.py es la guardia
        # estática que evita la regresión en CI.
        totales_sin_F = []
        totales_sin_C = []
        for grp in casillero_groups:
            cas = grp["casillero"]
            if not grp.get("is_total"):
                continue
            r = grp["row_start"]
            f_val = ws[f"F{r}"].value
            if not (isinstance(f_val, str) and f_val.startswith("=")):
                totales_sin_F.append(f"cas {cas} (fila {r})")
            c_val = ws[f"C{r}"].value
            if c_val is None or c_val == "":
                totales_sin_C.append(f"cas {cas} (fila {r})")

        if totales_sin_F:
            warnings.append(
                f"⚠ REGLA TOTAL: {len(totales_sin_F)} casilleros TOTAL sin "
                f"fórmula SUM en F (saldo contable): {totales_sin_F}. "
                f"Revisar PRIMARY_TOTAL_BLOCKS y COMPOSITE_TOTALS en "
                f"a1_mapeo.py."
            )
        if totales_sin_C:
            warnings.append(
                f"⚠ REGLA TOTAL: {len(totales_sin_C)} casilleros TOTAL sin "
                f"valor en C (declarado F-101): {totales_sin_C}. "
                f"El F-101 del cliente no declaró estos totales o el parser "
                f"f101_pdf.py los está perdiendo."
            )

        # === Aplicar formato profesional al A1 ===
        # REGLA (CLAUDE.md): los anexos del ICT deben verse correctamente
        # presentados, equivalentes al formato oficial del SRI Ecuador.
        try:
            from backend.app.ict.fillers.formatting import format_a1_sheet
            format_a1_sheet(ws, casillero_groups=casillero_groups,
                            first_data_row=A1_FIRST_DATA_ROW)
        except Exception:
            import logging
            logging.exception("format_a1_sheet falló")

        return {"filled_cells": filled, "warnings": warnings}


# === Inicialización estática de NEGATIVE_CASILLEROS ===
# Se ejecuta UNA VEZ al importar el módulo. Combina _NEGATIVE_CORE
# (calibrado manual) con los cas (-) auto-detectados del catálogo F-101.
# Resultado: any cas del balance con nombre que empieza "(-)" queda
# clasificado como negativo, sin requerir tocar este archivo cada vez
# que SRI agrega un nuevo cas de naturaleza (-).
A1Filler.NEGATIVE_CASILLEROS = A1Filler._build_negative_set()
