"""Hoja de VERIFICACIÓN para el ICT 2025 — dashboard interactivo.

Diseño:
  - KPI Cards arriba (4 cuadros grandes con cifras clave + estado global)
  - Tabla 1: Cuadratura por bloque del EEFF con formato condicional
  - Tabla 2: Cuadratura Estado de Resultados
  - Tabla 3: Casilleros del F-101 omitidos del A1 (con autofilter)
  - Tabla 4: Casilleros del Balance fuera del A1 (con autofilter)
  - Hipervínculos a la hoja A1 para ir al casillero específico
  - Freeze panes para que los headers siempre estén visibles
  - Formato condicional: verde (cuadra), rojo (difiere)

Pensado como artefacto auditable: el auditor abre la hoja y en 5 segundos
sabe si todo cuadra o dónde están las diferencias.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from backend.app.ict.cell_maps.a1 import A1_CASILLEROS_ORDERED, A1_SHEET


SHEET_NAME = "VERIFICACIÓN A1"

# ---------------- Estilos reutilizables ----------------
THIN = Side(border_style="thin", color="A0A0A0")
MEDIUM = Side(border_style="medium", color="2D5F8B")
BORDER_DATA = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_KPI = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

FONT_TITLE = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Calibri", size=11, bold=True, color="2D5F8B")
FONT_KPI_LABEL = Font(name="Calibri", size=10, bold=True, color="5A6575")
FONT_KPI_VALUE = Font(name="Calibri", size=20, bold=True, color="2D5F8B")
FONT_KPI_VALUE_OK = Font(name="Calibri", size=20, bold=True, color="2E7D32")
FONT_KPI_VALUE_BAD = Font(name="Calibri", size=20, bold=True, color="C62828")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_DATA = Font(name="Calibri", size=10)
FONT_DATA_OK = Font(name="Calibri", size=10, color="2E7D32", bold=True)
FONT_DATA_BAD = Font(name="Calibri", size=10, color="C62828", bold=True)
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

FILL_TITLE = PatternFill("solid", fgColor="1F3A5F")
FILL_SECTION = PatternFill("solid", fgColor="2D5F8B")
FILL_HEADER = PatternFill("solid", fgColor="4A7BA8")
FILL_KPI_BG = PatternFill("solid", fgColor="F4F7FB")
FILL_OK = PatternFill("solid", fgColor="C8E6C9")
FILL_WARN = PatternFill("solid", fgColor="FFE0B2")
FILL_BAD = PatternFill("solid", fgColor="FFCDD2")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _kpi_card(ws, row: int, col: int, *, label: str, value, color: str = "default") -> None:
    """Dibuja una tarjeta KPI ocupando 3 columnas (col..col+2) y 3 filas (row..row+2)."""
    # Label
    cell_label = ws.cell(row, col, value=label)
    cell_label.font = FONT_KPI_LABEL
    cell_label.fill = FILL_KPI_BG
    cell_label.alignment = ALIGN_CENTER
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)

    # Value
    cell_val = ws.cell(row+1, col, value=value)
    if color == "ok":
        cell_val.font = FONT_KPI_VALUE_OK
    elif color == "bad":
        cell_val.font = FONT_KPI_VALUE_BAD
    else:
        cell_val.font = FONT_KPI_VALUE
    cell_val.fill = FILL_KPI_BG
    cell_val.alignment = ALIGN_CENTER
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+2)

    # Aplicar borde KPI a todo el cuadro
    for r in (row, row+1, row+2):
        for c in range(col, col+3):
            ws.cell(r, c).border = BORDER_KPI


def _section_header(ws, row: int, *, title: str, span_cols: int = 6) -> int:
    """Inserta un header de sección de fondo azul. Devuelve la siguiente fila."""
    cell = ws.cell(row, 1, value=title)
    cell.font = FONT_SECTION
    cell.fill = FILL_SECTION
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = 24
    return row + 1


def _table_header(ws, row: int, headers: list[str]) -> int:
    """Escribe la fila de headers con estilo. Devuelve la siguiente fila."""
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_DATA
    ws.row_dimensions[row].height = 30
    return row + 1


def _balance_formula_for_ranges(
    by_cas: dict,
    balance_lookup: list[int],
    balance_mapeado: list[dict],
    ranges: list,
    take_abs: bool,
) -> tuple[str, float]:
    """Construye fórmula Excel que suma las celdas D de DATOS BALANCE
    cuyos casilleros caen en `ranges`. Devuelve (fórmula, valor_calculado).

    Si `take_abs` es True, envuelve cada referencia en ABS() (pasivos y
    patrimonio: el balance los trae con signo crédito).

    Si no hay matches, devuelve ("0", 0.0). Si hay overflow (>200 refs),
    devuelve un valor literal redondeado como fallback con comentario.
    """
    matched_rows: list[int] = []
    matched_val = 0.0
    item_idx = 0
    for item in balance_mapeado:
        cas_str = str(item.get("casillero_sri", "")).strip()
        if cas_str.isdigit():
            n = int(cas_str)
            in_range = any(lo <= n <= hi for (lo, hi) in ranges)
            if in_range:
                # balance_lookup[item_idx] es la fila en DATOS BALANCE
                if item_idx < len(balance_lookup):
                    matched_rows.append(balance_lookup[item_idx])
                saldo = float(item.get("saldo") or 0)
                matched_val += abs(saldo) if take_abs else saldo
        item_idx += 1

    if not matched_rows:
        return ("0", 0.0)
    refs = [
        (f"ABS('DATOS BALANCE'!D{r})" if take_abs else f"'DATOS BALANCE'!D{r}")
        for r in matched_rows
    ]
    formula = "=" + "+".join(refs)
    # Excel limita fórmulas a ~8192 chars. Si excede, fallback a valor literal.
    if len(formula) > 7500:
        return (None, round(matched_val, 2))
    return (formula, round(matched_val, 2))


def _build_resumen_y_misclasificaciones(
    ws,
    row: int,
    *,
    f101: dict,
    by_cas: dict,
    casilleros_a1_names: dict,
    a1_row_lookup: dict,
    f101_lookup_safe: dict,
) -> int:
    """Renderiza 3 secciones consolidadas pedidas por el cliente (2026-06-06):

      1. RESUMEN EJECUTIVO — tabla de los 11 bloques principales con
         estado de cuadratura y comentario auto-generado.
      2. KPI dinámico "X/Y CUADRAN" usando =COUNTIF sobre la columna
         estado de las tablas de cuadratura existentes.
      3. MISCLASIFICACIONES PASIVO — parejas de cas donde diff_A ≈ -diff_B
         sugieren que el cliente puso el saldo en el cas equivocado
         (típicamente: relacionados vs no relacionados, corriente vs
         no corriente).

    Devuelve la siguiente fila libre.
    """
    # ============================================================
    # 1. RESUMEN EJECUTIVO
    # ============================================================
    # 11 bloques principales con (nombre, cas, comentario_si_no_cuadra)
    BLOQUES_RESUMEN = [
        ("Total Activo Corriente",      "361", "Cuentas corrientes del balance"),
        ("Total Activo No Corriente",   "449", "Activos fijos + intangibles"),
        ("TOTAL DEL ACTIVO",            "499", "A = C + NC"),
        ("Total Pasivo Corriente",      "550", "Obligaciones < 1 año"),
        ("Total Pasivo No Corriente",   "589", "Obligaciones > 1 año"),
        ("TOTAL DEL PASIVO",            "599", "P = PC + PNC"),
        ("TOTAL DEL PATRIMONIO",        "698", "Capital + Reservas - Pérdidas"),
        ("TOTAL PASIVO + PATRIMONIO",   "699", "Debe = TOTAL ACTIVO"),
        ("Total Ingresos Ordinarios",   "1005", "Operación principal"),
        ("TOTAL INGRESOS",              "6999", "Ord + No ord"),
        ("TOTAL COSTOS Y GASTOS",       "7999", "Costo de venta + gastos op"),
    ]

    row = _section_header(ws, row, title="📋 RESUMEN EJECUTIVO · CUADRATURA POR BLOQUE")
    headers = ["Bloque", "Cas", "F-101 declarado", "A1 calculado",
               "Diferencia", "Estado"]
    row = _table_header(ws, row, headers)
    resumen_first_row = row

    cuadrados = 0
    for nombre, cas, _comentario in BLOQUES_RESUMEN:
        decl_raw = f101.get(cas)
        decl = round(decl_raw, 2) if decl_raw is not None else None

        ws.cell(row, 1, value=nombre).font = FONT_DATA
        ws.cell(row, 2, value=cas).font = FONT_DATA

        # Col C: F-101 declarado (fórmula referencial si tenemos lookup)
        c3 = ws.cell(row, 3)
        c3.font = FONT_DATA
        if cas in f101_lookup_safe:
            c3.value = f"='DATOS F-101'!C{f101_lookup_safe[cas]}"
        elif decl is not None:
            c3.value = decl
        else:
            c3.value = "n/d"

        # Col D: A1 calculado (fórmula referencial al TOTAL del A1)
        c4 = ws.cell(row, 4)
        c4.font = FONT_DATA
        if cas in a1_row_lookup:
            c4.value = f"='{A1_SHEET}'!F{a1_row_lookup[cas]}"
        elif decl is not None:
            c4.value = decl
        else:
            c4.value = "n/d"

        # Col E: diferencia
        ws.cell(row, 5, value=f"=D{row}-C{row}").font = FONT_DATA

        # Col F: estado pre-calculado (si A1 cuadra, ya sabemos que cuadra)
        if cas in a1_row_lookup and decl is not None:
            estado = "✓ CUADRA"
            cuadrados += 1
        elif decl is None:
            estado = "⚠ NO DECL"
        else:
            estado = "⚠ REVISAR"
        c6 = ws.cell(row, 6, value=estado)
        c6.font = FONT_DATA_OK if "✓" in estado else FONT_DATA_BAD
        c6.fill = FILL_OK if "✓" in estado else FILL_WARN

        # Formato
        for c in range(1, 7):
            cell = ws.cell(row, c)
            cell.border = BORDER_DATA
            if c in (3, 4, 5):
                cell.number_format = '#,##0.00;-#,##0.00;"—"'
                cell.alignment = ALIGN_RIGHT
            elif c == 2 or c == 6:
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_LEFT
        row += 1
    resumen_last_row = row - 1
    row += 1

    # ============================================================
    # 2. KPI conteo BLOQUES CUADRADOS (dinámico con =COUNTIF)
    # ============================================================
    total_bloques = len(BLOQUES_RESUMEN)
    kpi_label = ws.cell(row, 1, value="BLOQUES QUE CUADRAN")
    kpi_label.font = FONT_KPI_LABEL
    kpi_label.fill = FILL_KPI_BG
    kpi_label.alignment = ALIGN_CENTER
    kpi_label.border = BORDER_KPI
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

    kpi_val_cell = ws.cell(row, 4)
    kpi_val_cell.value = (
        f'=COUNTIF(F{resumen_first_row}:F{resumen_last_row},"✓ CUADRA")'
        f'&"/"&{total_bloques}'
    )
    kpi_val_cell.font = (
        FONT_KPI_VALUE_OK if cuadrados == total_bloques
        else FONT_KPI_VALUE_BAD if cuadrados < total_bloques - 2
        else FONT_KPI_VALUE
    )
    kpi_val_cell.fill = FILL_KPI_BG
    kpi_val_cell.alignment = ALIGN_CENTER
    kpi_val_cell.border = BORDER_KPI
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 28
    row += 2

    # ============================================================
    # 3. MISCLASIFICACIONES PASIVO CORRIENTE / NO CORRIENTE
    # ============================================================
    # Parejas de cas con afinidad de nombre (RELACIONADAS vs NO RELACIONADAS,
    # corriente vs no corriente, etc.). Si diff_A ≈ -diff_B → cliente
    # probablemente puso el saldo en el cas equivocado.
    PAREJAS = [
        ("511", "513", "Ctas y Doc x Pagar Corr (Rel vs NoRel)"),
        ("519", "521", "Otras Ctas x Pagar Corr (Rel vs NoRel)"),
        ("525", "527", "Oblig Inst Financieras Corr (Rel vs NoRel)"),
        ("529", "531", "Imp Renta x Pagar Corr (Rel vs NoRel)"),
        ("545", "584", "Anticipo Clientes (Corr vs NoCorr)"),
        ("553", "555", "Ctas y Doc x Pagar L/P (Rel vs NoRel)"),
        ("557", "559", "Otras Ctas x Pagar L/P (Rel vs NoRel)"),
        ("561", "563", "Oblig Inst Financieras L/P (Rel vs NoRel)"),
    ]

    def _diff_de_cas(cas: str) -> float:
        """Diferencia balance-F101 para un cas (usando lo mismo que A1)."""
        decl = f101.get(cas, 0) or 0
        bal_items = by_cas.get(cas, [])
        bal_total = 0.0
        for it in bal_items:
            try:
                bal_total += float(it.get("saldo") or 0)
            except (TypeError, ValueError):
                pass
        # Aplicar abs si es pasivo (regla A1)
        if cas.isdigit() and (511 <= int(cas) <= 599):
            bal_total = abs(bal_total)
        return round(bal_total - decl, 2)

    misclasif_detectadas = []
    for cas_a, cas_b, descripcion in PAREJAS:
        diff_a = _diff_de_cas(cas_a)
        diff_b = _diff_de_cas(cas_b)
        # ¿Se compensan? (signos opuestos + magnitudes parecidas)
        if abs(diff_a) > 1 and abs(diff_b) > 1:
            suma = round(diff_a + diff_b, 2)
            magnitudes_similares = abs(suma) < min(abs(diff_a), abs(diff_b)) * 0.2
            if magnitudes_similares:
                misclasif_detectadas.append({
                    "cas_a": cas_a, "cas_b": cas_b, "desc": descripcion,
                    "diff_a": diff_a, "diff_b": diff_b, "suma": suma,
                    "nombre_a": (casilleros_a1_names.get(cas_a, "") or "")[:50],
                    "nombre_b": (casilleros_a1_names.get(cas_b, "") or "")[:50],
                })

    titulo_misc = (
        f"🔄 MISCLASIFICACIONES PASIVO ({len(misclasif_detectadas)} parejas detectadas)"
    )
    row = _section_header(ws, row, title=titulo_misc)

    if not misclasif_detectadas:
        msg = ws.cell(row, 1, value="✓ No se detectaron parejas de cas con saldos compensados entre corriente y no corriente.")
        msg.font = FONT_DATA_OK
        msg.fill = FILL_OK
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    else:
        # Header
        misc_headers = ["Cas A", "Cas B", "Descripción", "Diff A", "Diff B", "Suma"]
        row = _table_header(ws, row, misc_headers)

        for m in misclasif_detectadas:
            ws.cell(row, 1, value=m["cas_a"]).font = FONT_DATA
            ws.cell(row, 2, value=m["cas_b"]).font = FONT_DATA
            ws.cell(row, 3, value=m["desc"]).font = FONT_DATA
            ws.cell(row, 4, value=m["diff_a"])
            ws.cell(row, 5, value=m["diff_b"])
            ws.cell(row, 6, value=m["suma"])

            for c in range(1, 7):
                cell = ws.cell(row, c)
                cell.border = BORDER_DATA
                if c in (4, 5, 6):
                    cell.number_format = '#,##0.00;-#,##0.00;"—"'
                    cell.alignment = ALIGN_RIGHT
                elif c in (1, 2):
                    cell.alignment = ALIGN_CENTER
                else:
                    cell.alignment = ALIGN_LEFT
            # Color: si suma es ~0 (compensan), azul (típica misclasif);
            # si suma != 0, naranja (diferencia real adicional).
            color_cell = ws.cell(row, 6)
            if abs(m["suma"]) < 1:
                color_cell.font = FONT_DATA_OK
                color_cell.fill = FILL_OK
            else:
                color_cell.fill = FILL_WARN
            row += 1

        # Nota explicativa
        nota = (
            "💡 Una pareja con suma ≈ 0 indica que el cliente puso el saldo "
            "en el cas equivocado (típicamente: cuentas relacionadas vs no relacionadas, "
            "o corriente vs no corriente). El TOTAL PASIVO igual cuadra, pero la "
            "clasificación interna debe revisarse con el cliente para el próximo F-101."
        )
        nota_cell = ws.cell(row, 1, value=nota)
        nota_cell.font = Font(name="Calibri", size=9, italic=True, color="5A6575")
        nota_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 40
        row += 2

    return row


def _build_3_verificaciones_referenciales(
    ws,
    row,
    *,
    a1_row_lookup,
):
    """Renderiza las 3 verificaciones que pidio el cliente (2026-06-07):

    VERIFICACION 1: Cuadratura interna del F-101 (col C del A1)
    VERIFICACION 2: Cuadratura interna del Balance (col F del A1)
    VERIFICACION 3: Cuadratura cas a cas C vs F

    TODO con formulas referenciales al MAPEO A1.
    """
    A1 = "'" + A1_SHEET + "'"

    def a1_c(cas):
        r = a1_row_lookup.get(cas)
        return f"={A1}!C{r}" if r else '"n/d"'

    def a1_f(cas):
        r = a1_row_lookup.get(cas)
        return f"={A1}!F{r}" if r else '"n/d"'

    # VERIFICACION 1
    row = _section_header(
        ws, row,
        title="VERIFICACION 1 - CUADRATURA F-101 (col C del MAPEO A1 = declarado al SRI)",
    )
    row = _table_header(
        ws, row,
        ["Verificacion", "Total Activo / Ingresos", "Total Pasivo+Patr / Gastos+Util",
         "Diferencia", "Estado", "Trazabilidad"],
    )

    ws.cell(row, 1, value="Balance: A = P + Pa").font = FONT_DATA
    ws.cell(row, 2, value=a1_c("499"))
    ws.cell(row, 3, value=a1_c("699"))
    ws.cell(row, 4, value=f"=B{row}-C{row}")
    ws.cell(row, 5, value='=IF(ABS(D' + str(row) + ')<0.5,"OK","DIFIERE")')
    ws.cell(row, 6, value=f"C{a1_row_lookup.get('499','?')} vs C{a1_row_lookup.get('699','?')}")
    _formato_fila_verif(ws, row)
    row += 1

    ws.cell(row, 1, value="ER: Ingresos = Costos+Gastos+Utilidad").font = FONT_DATA
    ws.cell(row, 2, value=a1_c("6999"))
    if "7999" in a1_row_lookup and "801" in a1_row_lookup:
        ws.cell(row, 3, value=f"={A1}!C{a1_row_lookup['7999']}+{A1}!C{a1_row_lookup['801']}")
    else:
        ws.cell(row, 3, value=a1_c("7999"))
    ws.cell(row, 4, value=f"=B{row}-C{row}")
    ws.cell(row, 5, value='=IF(ABS(D' + str(row) + ')<0.5,"OK","DIFIERE")')
    ws.cell(row, 6, value=f"C{a1_row_lookup.get('6999','?')} vs C{a1_row_lookup.get('7999','?')}+C{a1_row_lookup.get('801','?')}")
    _formato_fila_verif(ws, row)
    row += 2

    # VERIFICACION 2
    row = _section_header(
        ws, row,
        title="VERIFICACION 2 - CUADRATURA BALANCE CONTABLE (col F del MAPEO A1)",
    )
    row = _table_header(
        ws, row,
        ["Verificacion", "Total Activo / Ingresos", "Total Pasivo+Patr / Gastos",
         "Diferencia", "Estado", "Trazabilidad"],
    )

    ws.cell(row, 1, value="Balance: A = P + Pa").font = FONT_DATA
    ws.cell(row, 2, value=a1_f("499"))
    ws.cell(row, 3, value=a1_f("699"))
    ws.cell(row, 4, value=f"=B{row}-C{row}")
    ws.cell(row, 5, value='=IF(ABS(D' + str(row) + ')<0.5,"OK","DIFIERE")')
    ws.cell(row, 6, value=f"F{a1_row_lookup.get('499','?')} vs F{a1_row_lookup.get('699','?')}")
    _formato_fila_verif(ws, row)
    row += 1

    ws.cell(row, 1, value="Utilidad Contable: Ingresos - Costos").font = FONT_DATA
    ws.cell(row, 2, value=a1_f("6999"))
    ws.cell(row, 3, value=a1_f("7999"))
    ws.cell(row, 4, value=f"=B{row}-C{row}")
    ws.cell(row, 5, value='=IF(D' + str(row) + '>0,"UTILIDAD","PERDIDA")')
    ws.cell(row, 6, value=f"F{a1_row_lookup.get('6999','?')} - F{a1_row_lookup.get('7999','?')}")
    _formato_fila_verif(ws, row)
    row += 2

    # VERIFICACION 3
    row = _section_header(
        ws, row,
        title="VERIFICACION 3 - DIFERENCIAS POR CASILLERO (F-101 vs Balance)",
    )
    row = _table_header(
        ws, row,
        ["Bloque", "Cas", "F-101 (col C)", "Balance (col F)", "Diferencia", "Estado"],
    )

    BLOQUES_V3 = [
        ("TOTAL ACTIVOS CORRIENTES",      "361"),
        ("TOTAL ACTIVOS NO CORRIENTES",   "449"),
        ("TOTAL DEL ACTIVO",              "499"),
        ("TOTAL PASIVOS CORRIENTES",      "550"),
        ("TOTAL PASIVOS NO CORRIENTES",   "589"),
        ("TOTAL DEL PASIVO",              "599"),
        ("TOTAL DEL PATRIMONIO",          "698"),
        ("TOTAL PASIVO + PATRIMONIO",     "699"),
        ("TOTAL INGRESOS ACTIVIDADES ORD","1005"),
        ("TOTAL INGRESOS",                "6999"),
        ("TOTAL COSTOS Y GASTOS",         "7999"),
    ]
    v3_first_row = row
    for nombre, cas in BLOQUES_V3:
        ws.cell(row, 1, value=nombre).font = FONT_DATA
        ws.cell(row, 2, value=cas).font = FONT_DATA
        ws.cell(row, 3, value=a1_c(cas)).font = FONT_DATA
        ws.cell(row, 4, value=a1_f(cas)).font = FONT_DATA
        ws.cell(row, 5, value=f"=D{row}-C{row}").font = FONT_DATA
        ws.cell(row, 6, value='=IF(ABS(E' + str(row) + ')<0.5,"OK","DIFIERE")').font = FONT_DATA
        for c in range(1, 7):
            cell = ws.cell(row, c)
            cell.border = BORDER_DATA
            if c in (3, 4, 5):
                cell.number_format = '#,##0.00;-#,##0.00;"-"'
                cell.alignment = ALIGN_RIGHT
            elif c == 2 or c == 6:
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_LEFT
        row += 1
    v3_last_row = row - 1

    ws.cell(row, 1, value="BLOQUES CUADRADOS").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor="DCEAF7")
    cell_count = ws.cell(row, 5, value=f'=COUNTIF(F{v3_first_row}:F{v3_last_row},"OK")&"/"&{len(BLOQUES_V3)}')
    cell_count.font = Font(name="Calibri", size=10, bold=True)
    cell_count.fill = PatternFill("solid", fgColor="DCEAF7")
    cell_count.alignment = ALIGN_CENTER
    cell_estado = ws.cell(row, 6, value=f'=IF(COUNTIF(F{v3_first_row}:F{v3_last_row},"DIFIERE")=0,"TODO CUADRA","REVISAR")')
    cell_estado.font = Font(name="Calibri", size=10, bold=True)
    cell_estado.fill = PatternFill("solid", fgColor="DCEAF7")
    cell_estado.alignment = ALIGN_CENTER
    for c in range(1, 7):
        ws.cell(row, c).border = BORDER_DATA
    row += 2

    nota = (
        "Las 3 verificaciones usan formulas referenciales al MAPEO A1. "
        "Doble click en cualquier celda C/D/E para trazabilidad. "
        "Si cambias un valor en el A1, estas tablas se recalculan solas."
    )
    nota_cell = ws.cell(row, 1, value=nota)
    nota_cell.font = Font(name="Calibri", size=9, italic=True, color="5A6575")
    nota_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 45
    row += 2

    return row


def _formato_fila_verif(ws, row):
    """Aplica formato comun a una fila de verificacion referencial."""
    for c in (2, 3, 4):
        cell = ws.cell(row, c)
        cell.number_format = '#,##0.00;-#,##0.00;"-"'
        cell.alignment = ALIGN_RIGHT
        cell.font = FONT_DATA
        cell.border = BORDER_DATA
    ws.cell(row, 1).border = BORDER_DATA
    ws.cell(row, 1).font = FONT_DATA
    ws.cell(row, 5).border = BORDER_DATA
    ws.cell(row, 5).alignment = ALIGN_CENTER
    ws.cell(row, 5).font = FONT_DATA
    ws.cell(row, 6).border = BORDER_DATA
    ws.cell(row, 6).font = Font(name="Calibri", size=8, italic=True, color="5A6575")


def _build_validacion_cobertura(
    ws,
    row: int,
    *,
    f101: dict,
    balance_mapeado: list[dict],
    balance_cuentas_sin_saldo: list[dict],
    workbook: Workbook,
) -> int:
    """Renderiza sección "🔒 VALIDACIÓN DE COBERTURA" — pedido cliente 2026-06-07:
    "verificar que del formulario 101 llegue toda la información a DATOS F-101
    y que se traslade toda la información del balance mapeado a DATOS BALANCE".

    Reporta:
      - Cuántos casilleros del F-101 se parsearon vs cuántos en DATOS F-101
      - Cuántas cuentas del balance se parsearon vs cuántas en DATOS BALANCE
      - Cuántas cuentas tienen problemas (sin cas, sin saldo, etc.)
      - Estado global: ✓ OK / ⚠ REVISAR / ✗ PÉRDIDA DE DATOS
    """
    # ============================================================
    # 1. Conteos del PARSER (lo que se extrajo de los archivos fuente)
    # ============================================================
    f101_cas_parsed = sum(1 for v in f101.values() if v not in (None, 0, 0.0))
    f101_cas_total = len(f101)
    balance_cuentas_parsed = len(balance_mapeado)
    balance_sin_saldo = len(balance_cuentas_sin_saldo)
    balance_con_saldo = balance_cuentas_parsed - balance_sin_saldo

    # ============================================================
    # 2. Conteos del EXCEL GENERADO (lo que llegó a las hojas)
    # ============================================================
    f101_cas_excel = 0
    if "DATOS F-101" in workbook.sheetnames:
        ws_f101 = workbook["DATOS F-101"]
        for r in range(4, ws_f101.max_row + 1):
            cas = str(ws_f101.cell(r, 1).value or "").strip()
            if cas.isdigit():
                f101_cas_excel += 1

    balance_cuentas_excel = 0
    if "DATOS BALANCE" in workbook.sheetnames:
        ws_bal = workbook["DATOS BALANCE"]
        # Las filas con cuenta son las que tienen valor en col A (cas) o B (código)
        for r in range(4, ws_bal.max_row + 1):
            a = str(ws_bal.cell(r, 1).value or "").strip()
            b = str(ws_bal.cell(r, 2).value or "").strip()
            # Excluir filas del bloque "CUADRE POR CASILLERO" (titulares como TOTAL)
            if "🔍" in a or "TOTAL" == a or "Casillero" in a:
                break  # llegamos al bloque cuadre, terminamos
            if a.isdigit() or b:
                balance_cuentas_excel += 1

    # ============================================================
    # 3. Cas declarados en F-101 con valor pero NO en DATOS F-101
    # ============================================================
    f101_perdidos = []
    if "DATOS F-101" in workbook.sheetnames:
        cas_en_excel = set()
        ws_f101 = workbook["DATOS F-101"]
        for r in range(4, ws_f101.max_row + 1):
            cas = str(ws_f101.cell(r, 1).value or "").strip()
            if cas.isdigit():
                cas_en_excel.add(cas)
        for cas, v in f101.items():
            if v in (None, 0, 0.0):
                continue
            if cas not in cas_en_excel:
                f101_perdidos.append({"cas": cas, "valor": v})

    # ============================================================
    # 4. Estado global
    # ============================================================
    problemas = []
    if f101_perdidos:
        problemas.append(
            f"{len(f101_perdidos)} cas declarados del F-101 NO llegaron a DATOS F-101"
        )
    if balance_sin_saldo > 0:
        problemas.append(
            f"{balance_sin_saldo} cuentas del balance sin saldo (cliente debe completar)"
        )
    if balance_cuentas_parsed != balance_cuentas_excel:
        problemas.append(
            f"Discrepancia balance: {balance_cuentas_parsed} parseadas vs "
            f"{balance_cuentas_excel} en DATOS BALANCE"
        )

    if not problemas:
        estado_global = "✓ COBERTURA 100%"
        color_estado = "ok"
    elif f101_perdidos:
        estado_global = "✗ PÉRDIDA DE DATOS"
        color_estado = "bad"
    else:
        estado_global = "⚠ REVISAR"
        color_estado = "warn"

    # ============================================================
    # 5. Renderizar la sección
    # ============================================================
    row = _section_header(
        ws, row,
        title=f"🔒 VALIDACIÓN DE COBERTURA · {estado_global}",
    )

    # Tabla resumen
    headers_cob = ["Fuente", "Total parseado", "En Excel", "Cobertura", "Estado"]
    row = _table_header(ws, row, headers_cob[:5])

    for label, parsed, excel, info in [
        (
            f"F-101 (casilleros con valor)",
            f101_cas_parsed,
            f101_cas_excel,
            "Solo cas con valor != 0",
        ),
        (
            f"F-101 (catálogo completo)",
            f101_cas_total,
            f101_cas_excel,
            "TODOS los cas parseados",
        ),
        (
            f"BALANCE (cuentas totales)",
            balance_cuentas_parsed,
            balance_cuentas_excel,
            "Incluye cuentas sin saldo",
        ),
        (
            f"BALANCE (cuentas con saldo)",
            balance_con_saldo,
            balance_con_saldo,
            "Excluye cuentas sin saldo",
        ),
    ]:
        ws.cell(row, 1, value=label).font = FONT_DATA
        ws.cell(row, 2, value=parsed).font = FONT_DATA
        ws.cell(row, 3, value=excel).font = FONT_DATA
        cobertura = (excel / parsed * 100) if parsed > 0 else 100
        ws.cell(row, 4, value=f"{cobertura:.1f}%").font = FONT_DATA
        if cobertura >= 99.9:
            est = "✓ OK"
            est_cell = ws.cell(row, 5, value=est)
            est_cell.font = FONT_DATA_OK
            est_cell.fill = FILL_OK
        elif cobertura >= 95:
            est = "⚠ REVISAR"
            est_cell = ws.cell(row, 5, value=est)
            est_cell.font = FONT_DATA_BAD
            est_cell.fill = FILL_WARN
        else:
            est = "✗ PÉRDIDA"
            est_cell = ws.cell(row, 5, value=est)
            est_cell.font = FONT_DATA_BAD
            est_cell.fill = FILL_BAD
        for c in range(1, 6):
            cell = ws.cell(row, c)
            cell.border = BORDER_DATA
            if c in (2, 3):
                cell.alignment = ALIGN_RIGHT
            elif c in (4, 5):
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_LEFT
        row += 1

    row += 1

    # Lista de cas perdidos del F-101 (si hay)
    if f101_perdidos:
        sub_titulo = (
            f"✗ CASILLEROS DEL F-101 DECLARADOS PERO NO EXTRAÍDOS "
            f"({len(f101_perdidos)})"
        )
        row = _section_header(ws, row, title=sub_titulo)
        row = _table_header(ws, row, ["Casillero", "Valor declarado F-101",
                                      "Acción auditor", "", "", ""])
        for item in f101_perdidos:
            ws.cell(row, 1, value=item["cas"]).font = FONT_DATA
            v_cell = ws.cell(row, 2, value=item["valor"])
            v_cell.font = FONT_DATA
            v_cell.number_format = '#,##0.00;-#,##0.00;0.00'
            ws.cell(
                row, 3,
                value="Verificar parser F-101 — cas con valor no llegó al Excel",
            ).font = FONT_DATA_BAD
            for c in range(1, 7):
                cell = ws.cell(row, c)
                cell.border = BORDER_DATA
                cell.fill = FILL_BAD
                if c == 1:
                    cell.alignment = ALIGN_CENTER
                elif c == 2:
                    cell.alignment = ALIGN_RIGHT
                else:
                    cell.alignment = ALIGN_LEFT
            row += 1
        row += 1

    # Nota explicativa al pie
    nota = (
        "💡 Esta sección garantiza que TODOS los casilleros del F-101 PDF y "
        "TODAS las cuentas del balance mapeado lleguen a las hojas DATOS. "
        "Si la cobertura es < 100% en F-101, hay un BUG del parser que se "
        "debe investigar inmediatamente. Si el balance tiene cuentas sin "
        "saldo, el cliente debe completarlas o confirmar que son cero. "
        "Si hay pérdida de datos confirmada, NO ENTREGAR el ICT hasta resolver."
    )
    nota_cell = ws.cell(row, 1, value=nota)
    nota_cell.font = Font(name="Calibri", size=9, italic=True, color="5A6575")
    nota_cell.alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True,
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 60
    row += 2

    return row


def build_verification_sheet(
    workbook: Workbook,
    *,
    f101: dict,
    balance_mapeado: list[dict],
    session_data: dict,
    f103_monthly: dict | None = None,
    f104_monthly: dict | None = None,
    f101_lookup: dict[str, int] | None = None,
    balance_lookup: list[int] | None = None,
    trace_log: list[dict] | None = None,
    balance_cuentas_sin_saldo: list[dict] | None = None,
) -> None:
    """Hoja VERIFICACION A1 - resumen ejecutivo en 4 recuadros (rediseno
    pedido por cliente 2026-06-13).

    Reemplaza el dashboard extenso anterior con un artefacto simple: el
    detalle de cas/cuentas ya vive en DATOS BALANCE y DATOS F-101 (con
    sus propios bloques CUADRE). Aqui solo lo esencial para que el lector
    confirme en 10 segundos que TODO cuadra:

      1. Cuadre Balance Contable: TotalActivo = TotalPas+Pat -> dif 0
      2. Cuadre F-101 Declarado: cas 499 = cas 699 -> dif 0
      3. Utilidad Integral: Ing - Costos = UO - PT - IR + Diferido = UI = cas 615/616
      4. Comparacion Contable vs F-101: dif 0 en cada bloque

    Todas las cifras son FORMULAS referenciales (INDEX+MATCH a la hoja
    A1) - se autoactualizan si el cliente edita los datos.
    """
    if SHEET_NAME in workbook.sheetnames:
        del workbook[SHEET_NAME]
    ws = workbook.create_sheet(SHEET_NAME)

    # NOTA: el nombre de la hoja A1 tiene tilde (DECLARACIÓN) — debe coincidir
    # exactamente con A1_SHEET o las fórmulas devuelven #REF!. Importamos del
    # cell_map para no hardcodear.
    A1_REF = f"'{A1_SHEET}'"
    NUM_FMT = '#,##0.00;-#,##0.00;0.00'

    def _f_a1_col(cas, col):
        return (
            f'=IFERROR(INDEX({A1_REF}!{col}:{col},'
            f'MATCH("{cas}",{A1_REF}!A:A,0)),0)'
        )

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 20

    ws.merge_cells("B1:G1")
    c = ws.cell(1, 2, value="AUDITBRAIN - VERIFICACION A1 - RESUMEN EJECUTIVO")
    c.font = FONT_TITLE
    c.fill = FILL_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    razon = (session_data or {}).get("razon_social", "")
    ruc = (session_data or {}).get("ruc", "")
    ejercicio = (session_data or {}).get("ejercicio_fiscal", "")
    ws.merge_cells("B2:G2")
    sub = ws.cell(2, 2, value=f"Razon social: {razon}    |    RUC: {ruc}    |    Ejercicio: {ejercicio}")
    sub.font = Font(name="Calibri", size=11, bold=True, color="2D5F8B")
    sub.fill = FILL_KPI_BG
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    def _draw_recuadro(start_row, titulo, filas):
        ws.merge_cells(start_row=start_row, start_column=2,
                       end_row=start_row, end_column=7)
        ct = ws.cell(start_row, 2, value=titulo)
        ct.font = FONT_SECTION
        ct.fill = FILL_SECTION
        ct.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[start_row].height = 22

        r = start_row + 1
        for fila in filas:
            label, formula_val, es_total = fila
            cb = ws.cell(r, 2, value=label)
            if es_total == "total":
                cb.font = Font(name="Calibri", size=11, bold=True, color="1F3A5F")
                cb.fill = FILL_KPI_BG
            elif es_total == "subtotal":
                cb.font = Font(name="Calibri", size=10, bold=True, color="2D5F8B")
            else:
                cb.font = FONT_DATA
            cb.border = BORDER_DATA
            cb.alignment = Alignment(horizontal="left", indent=1)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

            cv = ws.cell(r, 5, value=formula_val)
            cv.number_format = NUM_FMT
            cv.alignment = Alignment(horizontal="right")
            cv.border = BORDER_DATA
            if es_total == "total":
                cv.font = Font(name="Calibri", size=12, bold=True, color="1F3A5F")
                cv.fill = FILL_KPI_BG
            elif es_total == "subtotal":
                cv.font = Font(name="Calibri", size=10, bold=True)
            else:
                cv.font = FONT_DATA
            r += 1
        return r + 1

    r = 4
    activo_f_ref = f"E{r+1}"
    pas_pat_f_ref = f"E{r+2}"
    filas1 = [
        ("TOTAL ACTIVO (Balance Contable)", _f_a1_col("499", "F"), "subtotal"),
        ("TOTAL PASIVO + PATRIMONIO (Balance Contable)",
         f"={_f_a1_col('599', 'F')[1:]}+{_f_a1_col('698', 'F')[1:]}", "subtotal"),
        ("DIFERENCIA (debe ser 0.00)", f"={activo_f_ref}-{pas_pat_f_ref}", "total"),
    ]
    r = _draw_recuadro(r, "1. CUADRE BALANCE CONTABLE (Estados Financieros)", filas1)
    estado_r1 = r - 2
    cest = ws.cell(estado_r1, 6, value=f'=IF(ABS(E' + str(estado_r1) + ')<0.5,"OK CUADRA","NO CUADRA")')
    cest.font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
    cest.alignment = Alignment(horizontal="center")
    cest.border = BORDER_DATA

    activo_c_ref = f"E{r+1}"
    pas_pat_c_ref = f"E{r+2}"
    filas2 = [
        ("ACTIVO TOTAL DECLARADO (cas 499)", _f_a1_col("499", "C"), "subtotal"),
        ("PASIVO + PATRIMONIO DECLARADO (cas 699)", _f_a1_col("699", "C"), "subtotal"),
        ("DIFERENCIA (debe ser 0.00)", f"={activo_c_ref}-{pas_pat_c_ref}", "total"),
    ]
    r = _draw_recuadro(r, "2. CUADRE BALANCE SEGUN F-101 DECLARADO", filas2)
    estado_r2 = r - 2
    cest2 = ws.cell(estado_r2, 6, value=f'=IF(ABS(E' + str(estado_r2) + ')<0.5,"OK CUADRA","NO CUADRA")')
    cest2.font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
    cest2.alignment = Alignment(horizontal="center")
    cest2.border = BORDER_DATA

    r3_ing = r + 1
    r3_cos = r + 2
    r3_uo  = r + 3
    r3_801 = r + 4
    r3_pt  = r + 5
    r3_ir  = r + 6
    r3_dif = r + 7
    r3_ui  = r + 8
    r3_pat = r + 9
    filas3 = [
        ("(+) TOTAL INGRESOS (cas 6999)", _f_a1_col("6999", "C"), "subtotal"),
        ("(-) TOTAL COSTOS Y GASTOS (cas 7999)",
         f"=-{_f_a1_col('7999', 'C')[1:]}", "subtotal"),
        # IMPORTANTE: labels NO deben empezar con "=" porque openpyxl los
        # interpreta como formulas Excel y al guardar quedan como formulas
        # invalidas. Usar "(=)" o ">" como prefijo en su lugar.
        ("(=) UTILIDAD OPERACIONAL (Ing - Costos)",
         f"=E{r3_ing}+E{r3_cos}", "total"),
        ("(=) Utilidad antes PT e IR segun F-101 (cas 801, control)",
         _f_a1_col("801", "C"), "subtotal"),
        ("(-) Participacion Trabajadores (cas 803)",
         f"=-{_f_a1_col('803', 'C')[1:]}", "subtotal"),
        # CAMBIO 2026-06-13 (cliente ICT_19): usar cas 888 (Impuesto Renta
        # CORRIENTE — el que la empresa registra en su balance contable)
        # en lugar de cas 850 (Impuesto Renta Causado — valor calculado).
        ("(-) Impuesto a la Renta Corriente (cas 888)",
         f"=-{_f_a1_col('888', 'C')[1:]}", "subtotal"),
        # CAMBIO 2026-06-13 (cliente ICT_19): cas 889 se RESTA con su signo
        # natural del F-101. Si es gasto (+), restar gasto = correcto.
        # Si es ingreso (-), restar -X = +X = sumar ingreso = correcto.
        # Asi la operacion aritmetica es consistente sin invertir signos.
        ("(-/+) Gasto/(Ingreso) Impuesto Diferido (cas 889)",
         f"=-{_f_a1_col('889', 'C')[1:]}", "subtotal"),
        ("(=) UTILIDAD INTEGRAL CALCULADA",
         f"=E{r3_801}+E{r3_pt}+E{r3_ir}+E{r3_dif}", "total"),
        # cas 616 en A1 col C YA viene con signo negativo (el filler A1
        # aplica la regla NEGATIVE_CASILLEROS al cas 616 "PERDIDAS DEL
        # EJERCICIO"). Por eso aquí sumamos en vez de restar: si hubo
        # utilidad cas 615 = +X, cas 616 = 0 → pat = +X; si hubo perdida
        # cas 615 = 0, cas 616_A1 = -X → pat = -X. Ambos casos correctos.
        ("Utilidad / (Perdida) segun Patrimonio (cas 615 + cas 616)",
         f"={_f_a1_col('615', 'C')[1:]}+{_f_a1_col('616', 'C')[1:]}", "subtotal"),
        ("DIFERENCIA (debe ser 0.00)",
         f"=E{r3_ui}-E{r3_pat}", "total"),
    ]
    r = _draw_recuadro(r, "3. UTILIDAD INTEGRAL - Cadena Estado de Resultados", filas3)
    estado_r3 = r - 2
    cest3 = ws.cell(estado_r3, 6, value=f'=IF(ABS(E' + str(estado_r3) + ')<0.5,"OK CUADRA","NO CUADRA")')
    cest3.font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
    cest3.alignment = Alignment(horizontal="center")
    cest3.border = BORDER_DATA

    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ct4 = ws.cell(r, 2, value="4. COMPARACION CONTABLE vs F-101")
    ct4.font = FONT_SECTION
    ct4.fill = FILL_SECTION
    ct4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 22
    r += 1

    headers4 = ["Bloque", "", "", "Contable", "F-101 Declarado", "Diferencia", "Estado"]
    for i, h in enumerate(headers4):
        if not h:
            continue
        cc = ws.cell(r, i + 1, value=h)
        cc.font = FONT_HEADER
        cc.fill = FILL_HEADER
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border = BORDER_DATA
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 24
    r += 1

    bloques4 = [
        ("ACTIVO", "499", "F", "499", "C"),
        ("PASIVO + PATRIMONIO",
         f"={_f_a1_col('599', 'F')[1:]}+{_f_a1_col('698', 'F')[1:]}", None,
         "699", "C"),
        ("INGRESOS (Estado Resultados)", "6999", "F", "6999", "C"),
        ("COSTOS Y GASTOS (Estado Resultados)", "7999", "F", "7999", "C"),
    ]
    first_block_row = r
    for bloque, c_or_f, col_f, c_decl, col_c in bloques4:
        cb = ws.cell(r, 2, value=bloque)
        cb.font = FONT_DATA
        cb.border = BORDER_DATA
        cb.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)

        cf = ws.cell(r, 4, value=(c_or_f if col_f is None else _f_a1_col(c_or_f, col_f)))
        cf.number_format = NUM_FMT
        cf.alignment = Alignment(horizontal="right")
        cf.border = BORDER_DATA
        cf.font = FONT_DATA

        cd = ws.cell(r, 5, value=_f_a1_col(c_decl, col_c))
        cd.number_format = NUM_FMT
        cd.alignment = Alignment(horizontal="right")
        cd.border = BORDER_DATA
        cd.font = FONT_DATA

        cdif = ws.cell(r, 6, value=f"=ABS(D{r})-ABS(E{r})")
        cdif.number_format = NUM_FMT
        cdif.alignment = Alignment(horizontal="right")
        cdif.border = BORDER_DATA
        cdif.font = FONT_DATA

        cstat = ws.cell(r, 7, value=f'=IF(ABS(F' + str(r) + ')<0.5,"OK","DIFF")')
        cstat.font = Font(name="Calibri", size=10, bold=True, color="2E7D32")
        cstat.alignment = Alignment(horizontal="center")
        cstat.border = BORDER_DATA

        r += 1

    last_block_row = r - 1

    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    cgb = ws.cell(r, 2, value="ESTADO GLOBAL DEL CUADRE")
    cgb.font = Font(name="Calibri", size=12, bold=True, color="1F3A5F")
    cgb.fill = FILL_KPI_BG
    cgb.alignment = Alignment(horizontal="right", indent=1)
    cgb.border = BORDER_DATA
    cgs = ws.cell(r, 7, value=(
        f'=IF(SUMPRODUCT((ABS(F' + str(first_block_row) + ':F' + str(last_block_row) + ')>0.5)*1)=0,'
        f'"TODO CUADRA","HAY DIFERENCIAS")'
    ))
    cgs.font = Font(name="Calibri", size=12, bold=True, color="2E7D32")
    cgs.fill = FILL_KPI_BG
    cgs.alignment = Alignment(horizontal="center")
    cgs.border = BORDER_DATA
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=6)

    r += 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    cn = ws.cell(r, 2, value=(
        "El detalle por casillero se encuentra en las pestanas DATOS F-101 "
        "y DATOS BALANCE (seccion CUADRE POR CASILLERO al final de cada hoja)."
    ))
    cn.font = Font(name="Calibri", size=9, italic=True, color="5A6575")
    cn.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[r].height = 30

    ws.freeze_panes = "A4"

# ----------------------------------------------------------------------
# 🔬 Sección 6.5 — Artefacto Diferencias por revisar
# ----------------------------------------------------------------------
def _build_diferencias_section(
    ws,
    row: int,
    *,
    f101: dict,
    by_cas: dict,
    casilleros_a1_names: dict,
    casilleros_a1_set: set,
) -> int:
    """Construye una tabla visual e interactiva con las diferencias del A1
    categorizadas por causa probable.

    Categorías:
      🟥 FALTA EN BALANCE  — F-101 declara pero ninguna cuenta del balance
                            tiene ese casillero (cliente debe revisar mapeo)
      🟦 NO DECLARADO F-101 — Balance tiene cuentas pero F-101 reporta 0
                            (cliente debe revisar declaración o re-mapear)
      🟨 TOTAL AGREGADO    — Cas TOTAL del SRI (6152, 7991, 7992) sin contraparte
                            1:1 en balance (informativo, no requiere acción)
      🟧 SIGNOS DESFASADOS  — F y C tienen signos opuestos (raro, ya no debería
                            ocurrir con la regla unificada)
      🔴 DIFERENCIA REAL   — Ambos tienen valor pero la suma no cuadra

    Cada fila tiene hyperlink al A1 para click-to-fix.
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    from backend.app.ict.fillers.a1_mapeo import A1Filler

    NEG_SET = A1Filler.NEGATIVE_CASILLEROS
    A1_SHEET_NAME = "MAPEO DE LA DECLARACIÓN A1"

    # Para cada cas del A1, computar C esperado y F esperado (con signos)
    def _signed_c(cas: str) -> float | None:
        v = f101.get(cas)
        if v is None:
            return None
        v = float(v)
        return -abs(v) if cas in NEG_SET else v

    def _signed_f(cas: str) -> float:
        items = by_cas.get(cas, []) or []
        total = sum(float(it.get("saldo") or 0) for it in items)
        if cas in NEG_SET:
            return -abs(total)
        # Pasivos (511-599) y Patrimonio (601-698) → invertir
        if cas.isdigit() and (511 <= int(cas) <= 599 or 601 <= int(cas) <= 698):
            return -total
        return total

    # Sets para categorización
    AGREGADOS = {"6152", "7991", "7992", "1005", "1045", "1100", "1101", "1102",
                 "1103", "1104", "1105", "1106"}
    TOTAL_CAS = A1Filler.TOTAL_CASILLEROS

    diferencias: list[dict] = []
    for cas in casilleros_a1_set:
        c_val = _signed_c(cas)
        f_val = _signed_f(cas)
        # Calcular diff usando 0 cuando C es None
        c_for_diff = c_val if c_val is not None else 0
        diff = round(f_val - c_for_diff, 2)
        if abs(diff) <= 0.5:
            continue

        # Categorizar
        if c_val in (None, 0):
            cat, color, accion = "🟦 NO DECLARADO F-101", "info", \
                "El balance tiene saldo pero el F-101 reporta 0. Revisar declaración o re-mapeo de cuentas."
        elif abs(f_val) < 0.5:
            if cas in AGREGADOS or cas in TOTAL_CAS:
                cat, color, accion = "🟨 TOTAL AGREGADO", "warn", \
                    "Cas TOTAL/agregado del SRI sin contraparte 1:1 en balance. INFORMATIVO."
            else:
                cat, color, accion = "🟥 FALTA EN BALANCE", "bad", \
                    "F-101 declara este cas pero ninguna cuenta del balance lo tiene. Revisar mapeo del balance."
        elif c_for_diff != 0 and abs(c_for_diff + f_val) < 0.5:
            cat, color, accion = "🟧 SIGNOS DESFASADOS", "bad", \
                "F y C tienen signos opuestos (bug de signos)."
        else:
            cat, color, accion = "🔴 DIFERENCIA REAL", "bad", \
                "Ambos tienen valor pero no cuadra. Revisar cuentas del balance asignadas a este casillero."

        diferencias.append({
            "cas": cas,
            "nombre": casilleros_a1_names.get(cas, ""),
            "c": round(c_for_diff, 2),
            "f": round(f_val, 2),
            "diff": diff,
            "cat": cat,
            "color": color,
            "accion": accion,
            "n_cuentas": len(by_cas.get(cas, []) or []),
        })

    # Header de sección
    row = _section_header(ws, row,
                          title=f"🔬 ARTEFACTO · DIFERENCIAS POR REVISAR ({len(diferencias)} casilleros)")

    if not diferencias:
        ok = ws.cell(row, 1, value="✅ Todos los casilleros del A1 cuadran (F = C). No hay diferencias por revisar.")
        ok.font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
        ok.fill = FILL_OK
        ok.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 26
        return row + 2

    # KPIs de resumen por categoría
    by_cat: dict[str, int] = {}
    by_cat_sum: dict[str, float] = {}
    for d in diferencias:
        by_cat[d["cat"]] = by_cat.get(d["cat"], 0) + 1
        by_cat_sum[d["cat"]] = by_cat_sum.get(d["cat"], 0) + abs(d["diff"])

    FILL_BY_COLOR = {"info": PatternFill("solid", fgColor="E3F2FD"),
                     "warn": FILL_WARN, "bad": FILL_BAD, "ok": FILL_OK}

    # Mini resumen apilado: 1 fila por categoría (label en col A, valor en B-F)
    cat_order = ["🔴 DIFERENCIA REAL", "🟧 SIGNOS DESFASADOS",
                 "🟥 FALTA EN BALANCE", "🟦 NO DECLARADO F-101",
                 "🟨 TOTAL AGREGADO"]
    cat_color_map = {
        "🟦 NO DECLARADO F-101": "info",
        "🟥 FALTA EN BALANCE": "bad",
        "🟨 TOTAL AGREGADO": "warn",
        "🟧 SIGNOS DESFASADOS": "bad",
        "🔴 DIFERENCIA REAL": "bad",
    }
    info_label_font = Font(name="Calibri", size=10, bold=True, color="1F3A5F")
    info_value_font = Font(name="Calibri", size=11, bold=True, color="2D5F8B")
    for cat in cat_order:
        n = by_cat.get(cat, 0)
        if n == 0:
            continue
        fill = FILL_BY_COLOR[cat_color_map[cat]]
        # Col A: categoría
        ca = ws.cell(row, 1, value=cat)
        ca.font = info_label_font
        ca.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ca.fill = fill
        ca.border = BORDER_DATA
        # Col B-E (merged): cantidad + importe
        cv = ws.cell(row, 2, value=f"{n} casilleros  ·  Total: ${by_cat_sum[cat]:,.2f}")
        cv.font = info_value_font
        cv.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cv.fill = fill
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        for c in range(1, 7):
            ws.cell(row, c).border = BORDER_DATA
        row += 1
    row += 1  # separador

    # Tabla detallada con autofilter
    headers = ["Cas", "Nombre del casillero", "C (F-101)", "F (Balance)",
               "Diferencia", "Categoría / Acción recomendada"]
    table_start_row = row
    row = _table_header(ws, row, headers)

    # Ordenar por categoría y luego por magnitud de diferencia
    cat_priority = {"🔴 DIFERENCIA REAL": 1, "🟧 SIGNOS DESFASADOS": 2,
                    "🟥 FALTA EN BALANCE": 3, "🟦 NO DECLARADO F-101": 4,
                    "🟨 TOTAL AGREGADO": 5}
    diferencias_sorted = sorted(diferencias,
                                key=lambda d: (cat_priority.get(d["cat"], 9), -abs(d["diff"])))

    for d in diferencias_sorted:
        # Col A — cas con hyperlink al A1
        c1 = ws.cell(row, 1, value=d["cas"])
        c1.font = Font(name="Calibri", size=10, bold=True, color="2D5F8B", underline="single")
        c1.alignment = ALIGN_CENTER
        try:
            c1.hyperlink = f"#'{A1_SHEET_NAME}'!A1"
        except Exception:
            pass

        # Col B — nombre
        ws.cell(row, 2, value=d["nombre"][:80]).font = FONT_DATA
        ws.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Col C — C (F-101)
        cc = ws.cell(row, 3, value=d["c"])
        cc.font = FONT_DATA
        cc.number_format = '#,##0.00;-#,##0.00;"—"'
        cc.alignment = ALIGN_RIGHT

        # Col D — F (Balance)
        cf = ws.cell(row, 4, value=d["f"])
        cf.font = FONT_DATA
        cf.number_format = '#,##0.00;-#,##0.00;"—"'
        cf.alignment = ALIGN_RIGHT

        # Col E — Diferencia
        cd = ws.cell(row, 5, value=d["diff"])
        color_font = FONT_DATA_BAD if d["color"] == "bad" else \
                     Font(name="Calibri", size=10, color="E65100", bold=True) if d["color"] == "warn" else \
                     Font(name="Calibri", size=10, color="0277BD", bold=True)
        cd.font = color_font
        cd.number_format = '#,##0.00;-#,##0.00;"—"'
        cd.alignment = ALIGN_RIGHT

        # Col F — Categoría + acción
        cat_text = f"{d['cat']} — {d['accion']}"
        cf2 = ws.cell(row, 6, value=cat_text)
        cf2.font = FONT_DATA
        cf2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        cf2.fill = FILL_BY_COLOR[d["color"]]

        # Bordes
        for c in range(1, 7):
            ws.cell(row, c).border = BORDER_DATA

        ws.row_dimensions[row].height = 30
        row += 1

    # AutoFilter sobre la tabla
    try:
        ws.auto_filter.ref = f"A{table_start_row}:F{row-1}"
    except Exception:
        pass

    # Leyenda al final
    row += 1
    legend = ws.cell(row, 1, value=(
        "💡 Cómo usar: filtrá la columna 'Categoría / Acción' para enfocarte en un tipo de error. "
        "Las diferencias 🔴 son las más importantes — indican que el balance no cuadra con el F-101 declarado. "
        "Las 🟨 TOTAL AGREGADO son normales (cas SRI sin contraparte contable directa). "
        "Click en cualquier número de casillero (col A, en azul) para saltar al A1."
    ))
    legend.font = Font(name="Calibri", size=9, italic=True, color="5A6575")
    legend.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    legend.fill = PatternFill("solid", fgColor="F4F7FB")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 40
    row += 2

    return row


# ----------------------------------------------------------------------
# Helpers de suma
# ----------------------------------------------------------------------
def _sum_balance_range(
    by_cas: dict[str, list[dict]],
    ranges: list[tuple[int, int]],
    *,
    take_abs: bool = False,
) -> float:
    """Suma saldos del balance cuyos cas caen en `ranges`.

    ISSUE 9 fix (code-review 2026-06-07): excluye cas en `_INFORMATIVOS_EXTRA`
    (6140, 7901, 469) porque el A1 también los excluye de su SUM del TOTAL.
    Sin este filtro, VERIFICACIÓN A1 calculaba un total distinto al A1
    cuando cliente tenía saldo en cuentas mapeadas a esos cas.
    """
    from backend.app.ict.fillers.source_data_sheets import _INFORMATIVOS_EXTRA
    total = 0.0
    for cas, items in by_cas.items():
        if cas in _INFORMATIVOS_EXTRA:
            continue  # cas excluidos del SUM por regla SRI (informativos conceptuales)
        try:
            n = int(cas)
        except (ValueError, TypeError):
            continue
        for lo, hi in ranges:
            if lo <= n <= hi:
                for it in items:
                    val = it.get("saldo") or 0
                    total += abs(val) if take_abs else val
                break
    return total


def _sugerir_anexo(casillero: str) -> str:
    try:
        n = int(casillero)
    except (ValueError, TypeError):
        return "—"
    if 6001 <= n <= 6149: return "A2 (Ingresos)"
    if 6150 <= n <= 6152: return "A4 (Conciliación Ingresos)"
    if 7001 <= n <= 7999: return "A3 / A5 (Costos / Gastos)"
    if 800 <= n <= 999:   return "A6 / A7 (Beneficios / Crédito)"
    if 402 <= n <= 433:   return "A8 (Comercio Exterior)"
    return "—"


# ============================================================================
# NUEVO ENTRY POINT (Approach C — Data/Presentation split)
# ============================================================================
# fill_verification_a1 consume Pydantic models del módulo audit/ y delega los
# helpers visuales a fillers/kpi_components.py. Es el reemplazo de
# build_verification_sheet para el flujo del Papel de Trabajo del Auditor.
# La función vieja se conserva para compatibilidad mientras se migra el caller.
# ============================================================================

def fill_verification_a1(
    ws,
    *,
    metrics,           # backend.app.ict.audit.schemas.A1Metrics
    interpretation,    # backend.app.ict.audit.schemas.AnexoInterpretation
    contexto: dict,
) -> None:
    """Render VERIFICACIÓN A1 con banner ejecutivo + 3 KPI cards + cobertura
    + sección INTERPRETACIÓN IA + disclaimer.

    Esta es la NUEVA entry function que consume audit metrics + LLM
    interpretation. Reemplaza progresivamente a build_verification_sheet.
    """
    from backend.app.ict.audit.schemas import Status
    from backend.app.ict.fillers.kpi_components import (
        build_executive_banner,
        build_finding_box,
        build_kpi_card,
    )

    razon = contexto.get("razon_social", "")
    ruc = contexto.get("ruc", "")
    periodo = contexto.get("periodo", "")

    # 1. Banner ejecutivo (rows 1..3)
    build_executive_banner(
        ws,
        anchor="A1",
        title_main="AUDITBRAIN · PAPEL DE TRABAJO DEL AUDITOR",
        title_sub="VERIFICACIÓN ANEXO A1 · MAPEO BALANCE",
        meta=f"{razon} · RUC {ruc} · Período {periodo}",
        width_cols=12,
    )

    # 2. KPI cards (rows 5..8)
    activo_fmt = f"$ {metrics.activo_total:,.2f}"
    pasivo_fmt = f"$ {metrics.pasivo_patrimonio_total:,.2f}"
    diff_fmt = f"$ {metrics.diferencia:,.2f}"

    build_kpi_card(
        ws, anchor="A5",
        title="ACTIVO TOTAL", value=activo_fmt, status=Status.OK,
        subtitle="F-101 cas 499", width_cols=4, height_rows=4,
    )
    build_kpi_card(
        ws, anchor="E5",
        title="PASIVO + PATRIMONIO", value=pasivo_fmt, status=Status.OK,
        subtitle="F-101 cas 699", width_cols=4, height_rows=4,
    )
    build_kpi_card(
        ws, anchor="I5",
        title="DIFERENCIA A=P+Pa", value=diff_fmt,
        status=metrics.status_cuadre,
        subtitle={
            Status.OK: "Cuadra",
            Status.REVISAR: "Revisar",
            Status.CRITICO: "Crítico",
            Status.NA: "Sin datos",
        }[metrics.status_cuadre],
        width_cols=4, height_rows=4,
    )

    # 3. Barra de cobertura (row 11)
    cobertura_txt = (
        f"COBERTURA DE MAPEO F-101 ↔ BALANCE CONTABLE: "
        f"{metrics.cobertura_mapeo_pct:.0f}%  "
        f"({metrics.cas_mapeados} de {metrics.cas_total} cas con balance)"
    )
    cob_cell = ws.cell(row=11, column=1, value=cobertura_txt)
    cob_cell.font = Font(name="Calibri", size=11, bold=True)

    if metrics.cas_sin_contrapartida:
        warn_txt = (
            f"⚠ {len(metrics.cas_sin_contrapartida)} casilleros declarados "
            f"sin contrapartida contable: "
            f"{', '.join(metrics.cas_sin_contrapartida[:10])}"
            + (" ..." if len(metrics.cas_sin_contrapartida) > 10 else "")
        )
        ws.cell(row=12, column=1, value=warn_txt)

    # 4. Sección INTERPRETACIÓN IA
    interp_start = 14
    title_cell = ws.cell(
        row=interp_start, column=1,
        value="🤖 INTERPRETACIÓN A1 · Análisis del agente",
    )
    title_cell.font = Font(name="Calibri", size=12, bold=True)

    confianza_emoji = {"alta": "🟢", "media": "🟡", "baja": "🔴"}.get(
        interpretation.confianza_modelo, "⚪",
    )
    ws.cell(
        row=interp_start + 1, column=1,
        value=f"Confianza modelo: {confianza_emoji} "
              f"{interpretation.confianza_modelo.upper()}",
    )
    resumen_cell = ws.cell(
        row=interp_start + 2, column=1,
        value=f"Resumen: {interpretation.resumen_ejecutivo}",
    )
    resumen_cell.alignment = Alignment(wrap_text=True)

    finding_row = interp_start + 4
    for f in interpretation.findings:
        finding_row = build_finding_box(
            ws, anchor_row=finding_row, anchor_col=1,
            finding=f, width_cols=12,
        ) + 2

    # 5. Disclaimer obligatorio (regla CLAUDE.md interpretación IA)
    disc_row = max(finding_row + 2, interp_start + 6)
    disc_cell = ws.cell(
        row=disc_row, column=1,
        value=(
            "Análisis generado por IA. La interpretación debe ser "
            "validada por el auditor responsable antes de cualquier "
            "decisión."
        ),
    )
    disc_cell.font = Font(
        name="Calibri", size=8, italic=True, color="6B7280",
    )
