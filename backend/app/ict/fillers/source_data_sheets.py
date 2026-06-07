"""Hojas de DATOS FUENTE para el ICT 2025.

Genera al final del workbook 4 hojas con TODOS los datos parseados de
los formularios SRI + Balance Mapeado del cliente. Cada anexo (A1..A9)
escribe FÓRMULAS que referencian estas hojas en lugar de valores
literales, de modo que:

  1. El auditor puede hacer doble-click en cualquier valor del anexo y
     ver desde QUÉ casillero/cuenta proviene.
  2. Si se actualiza un valor en la hoja DATOS (cambio manual del
     auditor), todos los anexos se recalculan automáticamente.
  3. Es trivial verificar qué casilleros del F-101/F-103/F-104 quedaron
     sin usar (la hoja DATOS los muestra TODOS, y la hoja
     VERIFICACIÓN reporta cuáles no se referenciaron).

Hojas generadas (al final, antes de VERIFICACIÓN y TRAZABILIDAD):
  · DATOS F-101            — un casillero por fila (anual)
  · DATOS F-103            — pivot mes×casillero (12 meses retenciones)
  · DATOS F-104            — pivot mes×casillero (12 meses IVA)
  · DATOS BALANCE MAPEADO — todas las cuentas con su casillero+saldo
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_F101 = "DATOS F-101"
SHEET_F103 = "DATOS F-103"
SHEET_F104 = "DATOS F-104"
SHEET_BALANCE = "DATOS BALANCE"


# ============================================================================
# REGLA del proyecto (CLAUDE.md / pedido cliente 2026-06-05):
# "Para A1 y DATOS F-101 solo extraer la información que tenga saldo y que NO
#  diga informativo".
# ============================================================================
# TOTALES del F-101 — SIEMPRE se muestran (aunque tengan saldo 0) porque
# son indicadores de cuadratura del balance.
F101_TOTALES = {
    "361", "449", "499",         # totales activos
    "550", "589", "599",          # totales pasivos
    "698", "699",                  # totales patrimonio + p+pa
    "1005", "1025", "1030", "1040", "1045", "1055", "1065", "1075", "1099",
    "6999",                        # total ingresos
    "7991", "7992", "7999",       # totales costos/gastos
    "899", "999",                  # totales impuesto
}


def _es_total_f101(cas: str) -> bool:
    """¿Es un casillero TOTAL del F-101?"""
    return cas in F101_TOTALES


# Cas que CONCEPTUALMENTE son informativos del subbloque revaluaciones
# pero el catálogo SRI 2025 NO los marca con "(INFORMATIVO)" en el nombre.
# Verificado empíricamente (PROPHAR 2025): incluirlos infla la cuadratura
# del cas 449 (TOTAL ACT NO CORR) porque son subtotales del subbloque
# revaluaciones — sus componentes ya están contabilizados en cas 385.
_INFORMATIVOS_EXTRA: frozenset[str] = frozenset({
    "469",  # (-) TOTAL DEPRECIACION ACUMULADA DEL AJUSTE ACUMULADO POR
            # REVALUACIONES Y OTROS AJUSTES NEGATIVOS — falta "(INFORMATIVO)"
            # en el catálogo SRI 2025 pero es subtotal del subbloque.
    "6140", # VENTAS NETAS DE PROPIEDADES PLANTA Y EQUIPO — verificado
            # empíricamente PROPHAR 2025: el F-101 SRI NO suma este cas
            # al TOTAL INGRESOS (6999). El ingreso real va en cas 6035
            # (UTILIDAD EN VENTA DE PPE). Sin este filtro, el A1 inflaba
            # cas 6999 en $176.52 (mismo valor que cas 6035, duplicado).
    "7901", # BAJA DE INVENTARIOS — verificado empíricamente PROPHAR 2025:
            # el F-101 SRI NO suma este cas al TOTAL COSTOS Y GASTOS
            # (7999). Es un cas con tratamiento tributario especial.
            # Sin este filtro, el A1 inflaba cas 7999 en $328,047.09.
})


def _es_informativo(nombre: str, cas: str | None = None) -> bool:
    """¿El cas es meramente informativo (no parte de la cuadratura)?

    SRI marca varios cas con "(INFORMATIVO)" o "(CASILLERO INFORMATIVO)" en
    el nombre — son cas opcionales que el auditor no necesita ver en A1
    o en DATOS F-101 si no tienen saldo.

    Adicional (2026-06-06): cas en _INFORMATIVOS_EXTRA también se tratan
    como informativos aunque el catálogo SRI no los marque (omisión SRI).
    """
    # Cas conceptualmente informativo aunque el nombre no lo diga.
    if cas and cas in _INFORMATIVOS_EXTRA:
        return True
    if not nombre:
        return False
    upper = nombre.upper()
    return (
        "INFORMATIVO" in upper
        or "CASILLERO INFORMATIVO" in upper
    )


def _es_excluido_estado_resultados(cas: str, nombre: str) -> bool:
    """¿Este cas del estado de resultados (6001-7999) debe ocultarse del A1?

    Regla cliente (2026-06-06): "QUITA DE A1 DEL ESTADO DE RESULTADOS LOS
    CASILLEROS INFORMATIVOS, INGRESOS EXENTOS Y GASTOS NO DEDUCIBLES.
    SOLO DEBE QUEDAR LAS CUENTAS QUE SE COMPARAN CON SALDOS CONTABLES."

    Excluye 4 familias del estado de resultados (rango 6001-7999):
      - VALOR EXENTO ... (~70 cas, son ingresos exentos del 25% IR)
      - VALOR NO DEDUCIBLE ... (~99 cas, ajuste tributario sin saldo contable)
      - INGRESOS NO OBJETO DE IMPUESTO A LA RENTA (cas 6150, 7906)
      - Cualquier cas marcado (INFORMATIVO) en el estado de resultados

    Estos cas son ajustes tributarios o información declarativa que NO
    tienen contraparte en el libro mayor del cliente — el auditor no
    puede cuadrarlos. Mostrarlos en A1 genera ruido y confunde.

    Los TOTALES (6999, 7999, etc.) NUNCA se excluyen: la cuadratura es
    responsabilidad de la fila TOTAL, no del detalle.

    NO se aplica al balance (311-699): ahí no hay este patrón.
    """
    if not cas or not cas.isdigit():
        return False
    n = int(cas)
    # Solo aplica al estado de resultados.
    if not (6001 <= n <= 7999):
        return False
    if not nombre:
        return False
    upper = nombre.upper().strip()
    # Patrones de exclusión validados contra catálogo F-101 OFICIAL SRI 2025.
    if upper.startswith("VALOR EXENTO"):
        return True
    if upper.startswith("VALOR NO DEDUCIBLE"):
        return True
    if "NO OBJETO DE IMPUESTO" in upper:
        return True
    if _es_informativo(nombre, cas):
        return True
    return False


def es_cas_relevante_f101(cas: str, valor, nombre: str) -> bool:
    """REGLA: un cas del F-101 es relevante (se muestra en A1 / DATOS F-101) si:
        1) Es un TOTAL (cuadratura), o
        2) Tiene saldo != 0 Y no es informativo.

    Cas con saldo 0 que no son TOTAL → ocultar.
    Cas marcados (INFORMATIVO) con saldo 0 → ocultar.
    Cas con saldo != 0 marcados (INFORMATIVO) → ocultar (regla del cliente).

    Devuelve True si el cas debe aparecer.
    """
    if _es_total_f101(cas):
        return True
    if _es_informativo(nombre, cas):
        return False
    try:
        v = float(valor) if valor is not None else 0
    except (TypeError, ValueError):
        v = 0
    return v != 0


def _safe_text(s) -> str:
    """Escapa el texto para que Excel NO lo interprete como fórmula.

    Excel trata cualquier celda cuyo valor empiece con `=`, `+`, `-`, `@`
    como una fórmula. Si el texto contiene paréntesis u operadores
    desbalanceados (caso: nombre "(=) Rebaja del saldo..." que terminó
    como "=) Rebaja..." después de limpieza), Excel lanza error de
    parseo y al abrir el archivo aplica "reparación" eliminando la celda.

    REGLA SUPREMA (CLAUDE.md): si genero un Excel, tiene que abrir SIN
    el cuadro "Excel pudo abrir el archivo reparando o quitando contenido
    que no se podía leer". Esta función evita ese caso.

    Devuelve el texto prefijado con apóstrofo si empieza con un carácter
    que Excel interpreta como fórmula. El apóstrofo es el escape estándar
    de Excel para forzar texto literal (no se ve en la celda).
    """
    if s is None:
        return ""
    text = str(s)
    if text and text[0] in ("=", "+", "-", "@"):
        return "'" + text
    return text


# ---- Estilos compartidos ----
THIN = Side(border_style="thin", color="A0A0A0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_DATA = Font(name="Calibri", size=9)
FILL_TITLE = PatternFill("solid", fgColor="1F3A5F")
FILL_HEADER = PatternFill("solid", fgColor="4A7BA8")


def _write_title(ws, title: str, span_cols: int = 4) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span_cols)
    c = ws.cell(1, 1, value=title)
    c.font = FONT_TITLE
    c.fill = FILL_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 26


def _write_header(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 24


# ---------------- F-101 ----------------
def build_f101_sheet(
    wb: Workbook,
    f101: dict,
    casillero_names: dict[str, str] | None = None,
) -> dict[str, int]:
    """Crea hoja DATOS F-101 con TODOS los casilleros canónicos del F-101 SRI.

    REGLA del proyecto (CLAUDE.md / pedido del usuario):
        "verificar que se trasladen TODOS los casilleros con sus códigos,
        nombres, valores — no importa que estén en cero, con la finalidad
        de que no haya saldos de líneas [vacías]"

    Estrategia:
      1. La lista CANÓNICA viene de backend/app/ict/catalogo_f101.py
         (F101_CASILLERO_NAMES). El test estático garantiza que cubre
         todos los casilleros del parser.
      2. Para cada casillero de la lista canónica se escribe SIEMPRE:
            · Columna A — número del casillero
            · Columna B — nombre oficial SRI (NUNCA vacío)
            · Columna C — valor declarado (0.00 si el PDF no lo trae)
            · Columna D — observación (vacío para edición manual)
      3. Si el F-101 trae casilleros EXTRAS no contemplados en el catálogo
         (caso edge: PDF con formato no estándar), también se escriben al
         final con el nombre del legacy lookup como fallback.
      4. Runtime check: si algún casillero del PDF no tiene nombre en el
         catálogo, se emite un warning vía logging (regla runtime).

    Args:
        f101: dict casillero_str → valor (extraído por parse_f101).
        casillero_names: legacy, opcional. Ya no se usa como fuente
            primaria — solo fallback para casilleros extras del PDF.

    Returns:
        casillero_to_row: {"311": 4, "315": 5, ...} para usar en fórmulas.
    """
    from backend.app.ict.catalogo_f101 import F101_CASILLERO_NAMES

    if SHEET_F101 in wb.sheetnames:
        del wb[SHEET_F101]
    ws = wb.create_sheet(SHEET_F101)

    _write_title(ws, "📄 DATOS F-101 · Declaración Anual del Impuesto a la Renta")
    _write_header(ws, 3, ["Casillero", "Nombre del Casillero", "Valor Declarado", "Observación"])

    # === REGLA: la lista canónica define qué casilleros aparecen ===
    canonical_cas = sorted(
        F101_CASILLERO_NAMES.keys(),
        key=lambda x: int(x) if x.isdigit() else 99999,
    )
    # Detectar casilleros del PDF que NO están en el catálogo (caso edge).
    legacy_names = casillero_names or {}
    extras_in_pdf = sorted(
        set(f101.keys()) - set(F101_CASILLERO_NAMES.keys()),
        key=lambda x: int(x) if x.isdigit() else 99999,
    )
    if extras_in_pdf:
        import logging
        logging.warning(
            "DATOS F-101: %d casilleros del PDF NO están en el catálogo canónico: %s. "
            "Agregar a backend/app/ict/catalogo_f101.py para que tengan nombre.",
            len(extras_in_pdf), extras_in_pdf[:20],
        )

    casillero_to_row: dict[str, int] = {}
    row = 4

    # DATOS F-101 conserva TODOS los 888 cas del catálogo (vista completa
    # del declarante / referencia documental del SRI). El filtro de
    # relevancia (solo con saldo + sin informativos) aplica SOLO al A1
    # (regla cliente 2026-06-05: "datos f101 puede estar completo si se
    # quiere, pero a1 no debe tener informativos").
    for cas in canonical_cas:
        val = f101.get(cas, 0)
        if val is None:
            val = 0
        nombre = F101_CASILLERO_NAMES[cas]
        ws.cell(row, 1, value=cas).font = FONT_DATA
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.cell(row, 1).border = BORDER
        ws.cell(row, 2, value=_safe_text(nombre)).font = FONT_DATA
        ws.cell(row, 2).border = BORDER
        c_val = ws.cell(row, 3, value=val)
        c_val.font = FONT_DATA
        c_val.number_format = '#,##0.00;-#,##0.00;0.00'
        c_val.alignment = Alignment(horizontal="right")
        c_val.border = BORDER
        ws.cell(row, 4, value="").border = BORDER
        casillero_to_row[cas] = row
        row += 1

    # Paso 2: extras del PDF que no estaban en el catálogo, al final
    for cas in extras_in_pdf:
        val = f101.get(cas) or 0
        nombre = legacy_names.get(cas, "(no catalogado — actualizar catalogo_f101.py)")
        ws.cell(row, 1, value=cas).font = FONT_DATA
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.cell(row, 1).border = BORDER
        ws.cell(row, 2, value=_safe_text(nombre)).font = FONT_DATA
        ws.cell(row, 2).border = BORDER
        c_val = ws.cell(row, 3, value=val)
        c_val.font = FONT_DATA
        c_val.number_format = '#,##0.00;-#,##0.00;0.00'
        c_val.alignment = Alignment(horizontal="right")
        c_val.border = BORDER
        ws.cell(row, 4, value="⚠ no catalogado").border = BORDER
        casillero_to_row[cas] = row
        row += 1

    # === REGLA runtime: verificar que NO haya filas con nombre vacío ===
    # Si algún row.B quedó vacío, es bug — emitir warning para alertar.
    for r in range(4, row):
        b_val = ws.cell(r, 2).value
        if b_val is None or str(b_val).strip() == "":
            cas_r = ws.cell(r, 1).value
            import logging
            logging.warning(
                "DATOS F-101 fila %d (cas %s): NOMBRE VACÍO. Regresión de regla.",
                r, cas_r,
            )

    # Anchos
    widths = {"A": 14, "B": 60, "C": 18, "D": 32}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:D{row-1}"
    return casillero_to_row


# ---------------- F-103 ----------------
def build_f103_sheet(wb: Workbook, f103_monthly: dict) -> dict[tuple[str, str], str]:
    """Crea hoja DATOS F-103 con TODOS los casilleros canónicos del F-103 SRI.

    REGLA del proyecto (pedido del usuario):
        "se trasladen TODOS los casilleros con sus códigos, nombres, valores,
        no importa que estén en cero, con la finalidad de que no haya saldos
        de líneas [vacías]"

    Estrategia (igual que F-101):
      · Lista canónica viene de catalogo_f103.py (F103_CASILLERO_NAMES).
      · SIEMPRE escribe todos los casilleros canónicos, aunque tengan 0.
      · Si el cliente no subió F-103, igual se muestra la matriz vacía
        para que el auditor sepa qué iba a esperar.
      · Casilleros extras del PDF que no están en el catálogo se agregan
        al final con observación "⚠ no catalogado".

    Returns:
        lookup: {(periodo, casillero) → "addr"} ej. ("2025-01", "302") → "C5"
                + lookup ("ANUAL", cas) → addr de la columna TOTAL ANUAL.
    """
    from backend.app.ict.catalogo_f103 import F103_CASILLERO_NAMES

    if SHEET_F103 in wb.sheetnames:
        del wb[SHEET_F103]
    ws = wb.create_sheet(SHEET_F103)

    _write_title(ws, "📋 DATOS F-103 · Declaraciones Mensuales de Retenciones IR")

    # Meses presentes (si no hay datos, generar matriz vacía con 12 meses
    # placeholder para visualizar la estructura completa).
    meses = sorted(f103_monthly.keys()) if f103_monthly else [
        f"2025-{m:02d}" for m in range(1, 13)
    ]

    # Casilleros: todos los canónicos + extras del PDF
    canonical_cas = list(F103_CASILLERO_NAMES.keys())
    extras_pdf: set[str] = set()
    for periodo in meses:
        if periodo not in f103_monthly:
            continue
        pdf_cas = set((f103_monthly[periodo].get("casilleros") or {}).keys())
        extras_pdf.update(pdf_cas - set(F103_CASILLERO_NAMES.keys()))
    extras_pdf_sorted = sorted(extras_pdf, key=lambda x: int(x) if x.isdigit() else 99999)
    sorted_cas = sorted(canonical_cas, key=lambda x: int(x) if x.isdigit() else 99999) + extras_pdf_sorted

    if extras_pdf:
        import logging
        logging.warning(
            "DATOS F-103: %d casilleros del PDF NO están en catalogo_f103.py: %s",
            len(extras_pdf), extras_pdf_sorted[:20],
        )

    # Header: Casillero | Nombre | mes1..mesN | TOTAL ANUAL
    headers = ["Casillero", "Nombre del Casillero"] + list(meses) + ["TOTAL ANUAL"]
    _write_header(ws, 3, headers)

    n_meses = len(meses)
    first_data_col = 3                              # mes 1 = col C
    last_data_col_idx = 2 + n_meses                 # último mes
    total_col_idx = 3 + n_meses                     # columna TOTAL ANUAL

    lookup: dict[tuple[str, str], str] = {}
    row = 4

    for cas in sorted_cas:
        # Columna A — casillero
        ws.cell(row, 1, value=cas).font = FONT_DATA
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.cell(row, 1).border = BORDER

        # Columna B — nombre del catálogo (NUNCA vacío si está en canónicos)
        nombre = F103_CASILLERO_NAMES.get(cas, "(no catalogado — actualizar catalogo_f103.py)")
        ws.cell(row, 2, value=_safe_text(nombre)).font = FONT_DATA
        ws.cell(row, 2).border = BORDER
        ws.cell(row, 2).alignment = Alignment(horizontal="left", wrap_text=True)

        # Columnas de cada mes — valor o 0 si no se declaró
        for j, periodo in enumerate(meses, start=first_data_col):
            mes_data = f103_monthly.get(periodo) or {}
            val = (mes_data.get("casilleros") or {}).get(cas, 0)
            if val is None:
                val = 0
            c = ws.cell(row, j, value=val)
            c.font = FONT_DATA
            c.number_format = '#,##0.00;-#,##0.00;0.00'
            c.alignment = Alignment(horizontal="right")
            c.border = BORDER
            lookup[(periodo, cas)] = f"{get_column_letter(j)}{row}"

        # Columna TOTAL ANUAL = SUM de los meses
        first_col_letter = get_column_letter(first_data_col)
        last_col_letter = get_column_letter(last_data_col_idx)
        total_col_letter = get_column_letter(total_col_idx)
        c_total = ws.cell(row, total_col_idx,
                          value=f"=SUM({first_col_letter}{row}:{last_col_letter}{row})")
        c_total.font = Font(name="Calibri", size=9, bold=True)
        c_total.number_format = '#,##0.00;-#,##0.00;0.00'
        c_total.alignment = Alignment(horizontal="right")
        c_total.border = BORDER
        lookup[("ANUAL", cas)] = f"{total_col_letter}{row}"
        row += 1

    # REGLA runtime: ninguna fila debe quedar sin nombre
    for r in range(4, row):
        b_val = ws.cell(r, 2).value
        if b_val is None or str(b_val).strip() == "":
            cas_r = ws.cell(r, 1).value
            import logging
            logging.warning(
                "DATOS F-103 fila %d (cas %s): NOMBRE VACÍO. Regresión de regla.",
                r, cas_r,
            )

    # Anchos: A=14 (cas), B=55 (nombre largo), meses=13, total=15
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    for j in range(first_data_col, last_data_col_idx + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13
    ws.column_dimensions[get_column_letter(total_col_idx)].width = 15

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(total_col_idx)}{row-1}"
    return lookup


# ---------------- F-104 ----------------
def build_f104_sheet(wb: Workbook, f104_monthly: dict) -> dict[tuple[str, str], str]:
    """Crea hoja DATOS F-104 con TODOS los casilleros canónicos del F-104 SRI.

    REGLA del proyecto (pedido del usuario):
        "se trasladen TODOS los casilleros con sus códigos, nombres, valores,
        no importa que estén en cero, con la finalidad de que no haya saldos
        de líneas [vacías]"

    Estrategia idéntica a F-103: catálogo canónico
    (backend/app/ict/catalogo_f104.py) define las 22+ filas. Si el cliente
    sube F-104 con valores 0 para ciertos casilleros, igual aparecen.
    Si no sube F-104, se muestra matriz vacía con la estructura completa.

    Returns:
        lookup: {(periodo, casillero) → "addr"} + ("ANUAL", cas) → total.
    """
    from backend.app.ict.catalogo_f104 import F104_CASILLERO_NAMES

    if SHEET_F104 in wb.sheetnames:
        del wb[SHEET_F104]
    ws = wb.create_sheet(SHEET_F104)

    _write_title(ws, "📑 DATOS F-104 · Declaraciones Mensuales de IVA")

    # Meses presentes (si no hay datos, 12 meses placeholder)
    meses = sorted(f104_monthly.keys()) if f104_monthly else [
        f"2025-{m:02d}" for m in range(1, 13)
    ]

    # Casilleros: todos los canónicos + extras del PDF
    canonical_cas = list(F104_CASILLERO_NAMES.keys())
    extras_pdf: set[str] = set()
    for periodo in meses:
        if periodo not in f104_monthly:
            continue
        d = f104_monthly.get(periodo) or {}
        cas_dict = d.get("casilleros") if isinstance(d, dict) else None
        if cas_dict:
            extras_pdf.update(set(cas_dict.keys()) - set(F104_CASILLERO_NAMES.keys()))
    extras_pdf_sorted = sorted(extras_pdf, key=lambda x: int(x) if x.isdigit() else 99999)
    sorted_cas = sorted(canonical_cas, key=lambda x: int(x) if x.isdigit() else 99999) + extras_pdf_sorted

    if extras_pdf:
        import logging
        logging.warning(
            "DATOS F-104: %d casilleros del PDF NO están en catalogo_f104.py: %s",
            len(extras_pdf), extras_pdf_sorted[:20],
        )

    # Header: Casillero | Nombre | mes1..mesN | TOTAL ANUAL
    headers = ["Casillero", "Nombre del Casillero"] + list(meses) + ["TOTAL ANUAL"]
    _write_header(ws, 3, headers)

    n_meses = len(meses)
    first_data_col = 3
    last_data_col_idx = 2 + n_meses
    total_col_idx = 3 + n_meses

    lookup: dict[tuple[str, str], str] = {}
    row = 4

    for cas in sorted_cas:
        ws.cell(row, 1, value=cas).font = FONT_DATA
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.cell(row, 1).border = BORDER

        nombre = F104_CASILLERO_NAMES.get(cas, "(no catalogado — actualizar catalogo_f104.py)")
        ws.cell(row, 2, value=_safe_text(nombre)).font = FONT_DATA
        ws.cell(row, 2).border = BORDER
        ws.cell(row, 2).alignment = Alignment(horizontal="left", wrap_text=True)

        for j, periodo in enumerate(meses, start=first_data_col):
            d = f104_monthly.get(periodo) or {}
            cas_dict = d.get("casilleros") if isinstance(d, dict) else None
            val = (cas_dict or {}).get(cas, 0)
            if val is None:
                val = 0
            c = ws.cell(row, j, value=val)
            c.font = FONT_DATA
            c.number_format = '#,##0.00;-#,##0.00;0.00'
            c.alignment = Alignment(horizontal="right")
            c.border = BORDER
            lookup[(periodo, cas)] = f"{get_column_letter(j)}{row}"

        first_col_letter = get_column_letter(first_data_col)
        last_col_letter = get_column_letter(last_data_col_idx)
        total_col_letter = get_column_letter(total_col_idx)
        c_total = ws.cell(row, total_col_idx,
                          value=f"=SUM({first_col_letter}{row}:{last_col_letter}{row})")
        c_total.font = Font(name="Calibri", size=9, bold=True)
        c_total.number_format = '#,##0.00;-#,##0.00;0.00'
        c_total.alignment = Alignment(horizontal="right")
        c_total.border = BORDER
        lookup[("ANUAL", cas)] = f"{total_col_letter}{row}"
        row += 1

    # REGLA runtime: ninguna fila sin nombre
    for r in range(4, row):
        b_val = ws.cell(r, 2).value
        if b_val is None or str(b_val).strip() == "":
            cas_r = ws.cell(r, 1).value
            import logging
            logging.warning(
                "DATOS F-104 fila %d (cas %s): NOMBRE VACÍO. Regresión de regla.",
                r, cas_r,
            )

    # Anchos
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    for j in range(first_data_col, last_data_col_idx + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13
    ws.column_dimensions[get_column_letter(total_col_idx)].width = 15

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(total_col_idx)}{row-1}"
    return lookup


# ---------------- BALANCE MAPEADO ----------------
def build_balance_sheet(wb: Workbook, balance: list[dict]) -> list[int]:
    """Crea hoja DATOS BALANCE con TODAS las cuentas del balance mapeado.

    Returns:
        item_to_row: lista del mismo largo que balance, donde item_to_row[i]
            es la fila en DATOS BALANCE donde se escribió la cuenta i de la lista
            original. Permite a A1 generar fórmulas tipo
            ='DATOS BALANCE'!D<row_idx>.
    """
    if SHEET_BALANCE in wb.sheetnames:
        del wb[SHEET_BALANCE]
    ws = wb.create_sheet(SHEET_BALANCE)

    _write_title(ws, "📊 DATOS BALANCE MAPEADO · Cuentas y saldos del cliente",
                 span_cols=5)
    _write_header(ws, 3, ["Casillero SRI", "Código Contable", "Nombre Cuenta",
                          "Saldo 31-dic", "Origen"])

    item_to_row: list[int] = []
    row = 4
    for item in balance:
        cas = str(item.get("casillero_sri", "")).strip()
        codigo = item.get("codigo", "")
        desc = item.get("descripcion", "")
        saldo = item.get("saldo", 0)

        ws.cell(row, 1, value=cas).font = FONT_DATA
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.cell(row, 1).border = BORDER
        ws.cell(row, 2, value=codigo).font = FONT_DATA
        ws.cell(row, 2).border = BORDER
        ws.cell(row, 3, value=desc).font = FONT_DATA
        ws.cell(row, 3).border = BORDER
        c_saldo = ws.cell(row, 4, value=saldo)
        c_saldo.font = FONT_DATA
        c_saldo.number_format = '#,##0.00;-#,##0.00;0.00'
        c_saldo.alignment = Alignment(horizontal="right")
        c_saldo.border = BORDER
        ws.cell(row, 5, value="Balance Mapeado del cliente").font = FONT_DATA
        ws.cell(row, 5).border = BORDER

        item_to_row.append(row)
        row += 1

    detail_last_row = row - 1
    detail_first_row = 4

    # ================================================================
    # BLOQUE CUADRE CON MAPEO A1 (pedido cliente 2026-06-06)
    # ================================================================
    # Para cada casillero SRI único, comparamos:
    #   - Suma balance contable (SUMIF sobre col D del detalle de arriba)
    #   - Saldo final que el A1 muestra en col F del cas (VLOOKUP a A1)
    # La diferencia en valores absolutos debe ser 0 (A1 toma el balance
    # con signo aplicado, pero el monto bruto es el mismo).
    casilleros_unicos = sorted(
        {str(it.get("casillero_sri", "")).strip() for it in balance
         if str(it.get("casillero_sri", "")).strip().isdigit()},
        key=lambda x: int(x),
    )

    if casilleros_unicos:
        row += 2  # separación visual
        # Título
        title_cell = ws.cell(row, 1,
                             value="🔍 CUADRE POR CASILLERO · Balance ↔ MAPEO A1")
        title_cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="2D5F8B")
        title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 24
        row += 1

        # Headers
        headers = [
            "Casillero",
            "# Cuentas",
            "Suma Balance",
            "Saldo A1 col F",
            "Diferencia |F|-|Bal|",
            "Cuadre",
        ]
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row, i, value=h)
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="4A7BA8")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[row].height = 28
        row += 1
        resumen_first = row

        # Referencia a la hoja A1 — usamos comillas porque tiene espacios
        A1_REF = "'MAPEO DE LA DECLARACIÓN A1'"
        # Rango col A:F del A1 (debería cubrir todos los cas)
        # Usamos VLOOKUP con 4° argumento FALSE para coincidencia exacta.
        # Si el cas no está en A1 (raro), la cifra será #N/A — convertimos
        # a 0 con IFERROR.

        for cas in casilleros_unicos:
            ws.cell(row, 1, value=cas).font = FONT_DATA
            ws.cell(row, 1).alignment = Alignment(horizontal="center")
            ws.cell(row, 1).border = BORDER

            # Col B: cuántas cuentas hay con ese cas
            f_count = f'=COUNTIF(A{detail_first_row}:A{detail_last_row},"{cas}")'
            cb = ws.cell(row, 2, value=f_count)
            cb.font = FONT_DATA
            cb.alignment = Alignment(horizontal="center")
            cb.border = BORDER

            # Col C: suma balance del cas (SUMIF)
            f_sumbal = f'=SUMIF(A{detail_first_row}:A{detail_last_row},"{cas}",D{detail_first_row}:D{detail_last_row})'
            cc = ws.cell(row, 3, value=f_sumbal)
            cc.font = FONT_DATA
            cc.number_format = '#,##0.00;-#,##0.00;0.00'
            cc.alignment = Alignment(horizontal="right")
            cc.border = BORDER

            # Col D: saldo del A1 col F (VLOOKUP)
            # Buscar cas en col A de A1, devolver col F (índice 6)
            # IFERROR para evitar #N/A si no encuentra
            f_a1 = (
                f'=IFERROR(VLOOKUP("{cas}",{A1_REF}!$A:$F,6,FALSE),0)'
            )
            cd = ws.cell(row, 4, value=f_a1)
            cd.font = FONT_DATA
            cd.number_format = '#,##0.00;-#,##0.00;0.00'
            cd.alignment = Alignment(horizontal="right")
            cd.border = BORDER

            # Col E: diferencia en valor absoluto (debe ser 0)
            # ABS(D{row}) - ABS(C{row})
            f_diff = f'=ABS(D{row})-ABS(C{row})'
            ce = ws.cell(row, 5, value=f_diff)
            ce.font = FONT_DATA
            ce.number_format = '#,##0.00;-#,##0.00;0.00'
            ce.alignment = Alignment(horizontal="right")
            ce.border = BORDER

            # Col F: estado (texto IF basado en col E)
            # Si |diff| < 0.5 → "✓ OK", sino "✗ DIFF"
            f_estado = f'=IF(ABS(E{row})<0.5,"✓ OK","✗ DIFF")'
            cf = ws.cell(row, 6, value=f_estado)
            cf.font = FONT_DATA
            cf.alignment = Alignment(horizontal="center")
            cf.border = BORDER

            row += 1

        resumen_last = row - 1

        # Fila TOTAL al final del bloque
        ws.cell(row, 1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="DCEAF7")
        ws.cell(row, 1).border = BORDER

        # # cas únicos
        c_cnt = ws.cell(row, 2, value=len(casilleros_unicos))
        c_cnt.font = Font(name="Calibri", size=10, bold=True)
        c_cnt.alignment = Alignment(horizontal="center")
        c_cnt.fill = PatternFill("solid", fgColor="DCEAF7")
        c_cnt.border = BORDER

        # Suma total balance
        c_sum_bal = ws.cell(row, 3, value=f'=SUM(C{resumen_first}:C{resumen_last})')
        c_sum_bal.font = Font(name="Calibri", size=10, bold=True)
        c_sum_bal.number_format = '#,##0.00;-#,##0.00;0.00'
        c_sum_bal.alignment = Alignment(horizontal="right")
        c_sum_bal.fill = PatternFill("solid", fgColor="DCEAF7")
        c_sum_bal.border = BORDER

        # Suma total A1 col F
        c_sum_a1 = ws.cell(row, 4, value=f'=SUM(D{resumen_first}:D{resumen_last})')
        c_sum_a1.font = Font(name="Calibri", size=10, bold=True)
        c_sum_a1.number_format = '#,##0.00;-#,##0.00;0.00'
        c_sum_a1.alignment = Alignment(horizontal="right")
        c_sum_a1.fill = PatternFill("solid", fgColor="DCEAF7")
        c_sum_a1.border = BORDER

        # Diferencias absolutas (cuántas filas con diff > 0.5)
        c_cnt_diff = ws.cell(
            row, 5,
            value=f'=COUNTIF(F{resumen_first}:F{resumen_last},"✗ DIFF")&" con diff"',
        )
        c_cnt_diff.font = Font(name="Calibri", size=10, bold=True, color="C62828")
        c_cnt_diff.alignment = Alignment(horizontal="center")
        c_cnt_diff.fill = PatternFill("solid", fgColor="DCEAF7")
        c_cnt_diff.border = BORDER

        # Estado global
        c_glob = ws.cell(
            row, 6,
            value=(
                f'=IF(COUNTIF(F{resumen_first}:F{resumen_last},"✗ DIFF")=0,'
                f'"✓ TODO CUADRA","⚠ REVISAR")'
            ),
        )
        c_glob.font = Font(name="Calibri", size=10, bold=True)
        c_glob.alignment = Alignment(horizontal="center")
        c_glob.fill = PatternFill("solid", fgColor="DCEAF7")
        c_glob.border = BORDER
        row += 1

    widths = {"A": 14, "B": 22, "C": 50, "D": 18, "E": 30, "F": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:E{detail_last_row}"
    return item_to_row
