"""Filler protocol + helpers for ICT 2025 Excel template manipulation.

Este módulo centraliza:
  - load_template(): carga el template oficial preservando fórmulas y links.
  - safe_set(): escribe en una celda de forma segura. NUNCA sobreescribe
    MergedCells ni fórmulas; opcionalmente registra cada escritura en el
    trace log para que generate_excel pueda producir la hoja
    "📋 Trazabilidad" con el linaje completo origen → destino.
  - reset_trace() / get_trace(): API para gestionar el log entre
    invocaciones de generate_excel.
"""

from __future__ import annotations

from contextvars import ContextVar
from pathlib import Path
from typing import Protocol

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.merge import MergedCell

TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "ict_2025_template.xlsx"


def load_template() -> Workbook:
    """Load the official SRI ICT 2025 template preserving formulas."""
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"ICT template not found at {TEMPLATE_PATH}. "
            "Copy the official SRI template here before running ICT generation."
        )
    return load_workbook(TEMPLATE_PATH, data_only=False, keep_links=True)


# ---------------------------------------------------------------------------
# Trace log — registro de origen→destino para la hoja "📋 Trazabilidad"
# Usamos ContextVar para que sea seguro en concurrencia (cada request tiene
# su propio trace) y para que los fillers no tengan que recibir un parámetro.
# ---------------------------------------------------------------------------

_TRACE: ContextVar[list[dict] | None] = ContextVar("_ict_trace", default=None)


def reset_trace() -> None:
    """Inicia un nuevo trace log para la generación actual."""
    _TRACE.set([])


def get_trace() -> list[dict]:
    """Devuelve el trace acumulado (lista de dicts) o vacío si no hay sesión."""
    return _TRACE.get() or []


def _record(anexo: str | None, casillero: str | None, sheet: str,
            cell_addr: str, value, origen: str | None, status: str) -> None:
    """Append a trace entry — silently no-op if no active trace."""
    trace = _TRACE.get()
    if trace is None:
        return
    trace.append({
        "anexo": anexo or "",
        "casillero": str(casillero or ""),
        "sheet": sheet,
        "cell": cell_addr,
        "origen": origen or "",
        "valor": value,
        "status": status,  # "written" | "skipped_formula" | "skipped_merged"
    })


def safe_set_formula(
    ws,
    cell_addr: str,
    formula: str,
    *,
    anexo: str | None = None,
    casillero: str | None = None,
    origen: str | None = None,
) -> bool:
    """Escribe una FÓRMULA en una celda, sobreescribiendo cualquier valor
    o fórmula anterior. Úsalo cuando el filler intencionalmente está
    actualizando una fórmula calculada (p. ej. ``=SUM(F13:F25)-C13`` que
    reemplaza la fórmula vieja del template ``=F15-C13``).

    NO usa el guard de protección de fórmulas porque la intención es
    explícita. Sigue respetando MergedCells y registra la operación en
    el trace log.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        raise ValueError(f"safe_set_formula esperaba '=...' pero recibió: {formula!r}")
    try:
        cell = ws[cell_addr]
    except Exception:
        _record(anexo, casillero, ws.title, cell_addr, formula, origen, "error")
        return False
    if isinstance(cell, MergedCell):
        _record(anexo, casillero, ws.title, cell_addr, formula, origen, "skipped_merged")
        return False
    cell.value = formula
    _record(anexo, casillero, ws.title, cell_addr, formula, origen, "written")
    return True


def safe_set(
    ws,
    cell_addr: str,
    value,
    *,
    anexo: str | None = None,
    casillero: str | None = None,
    origen: str | None = None,
) -> bool:
    """Escribe ``value`` en ``ws[cell_addr]`` de forma defensiva.

    Devuelve True si se escribió, False si se omitió por alguna razón
    (MergedCell, fórmula, o error). En todos los casos registra la
    operación en el trace log si está activo, para auditoría.

    Reglas de protección:
      1. Si la celda es MergedCell → skip (escribir lanzaría AttributeError).
      2. Si la celda tiene una fórmula (cell.data_type == "f" o value que
         empieza con "=") → skip. Esto preserva 100% las fórmulas del
         template oficial del SRI, incluso si el cell_map de un filler
         apunta accidentalmente a una celda calculada.

    Args:
        ws: Worksheet de openpyxl.
        cell_addr: dirección estilo "A1", "C13".
        value: valor a escribir (str, int, float).
        anexo: código del anexo (A1..A9, INDICE) para trazabilidad.
        casillero: número del casillero SRI de origen (opcional).
        origen: descripción humana del origen ("F-101 página 1",
            "Balance Mapeado fila 39", etc).
    """
    try:
        cell = ws[cell_addr]
    except Exception:
        _record(anexo, casillero, ws.title, cell_addr, value, origen, "error")
        return False

    # Guard 1: MergedCell
    if isinstance(cell, MergedCell):
        _record(anexo, casillero, ws.title, cell_addr, value, origen, "skipped_merged")
        return False

    # Guard 2: fórmula. data_type "f" o str que empieza con "=".
    existing = cell.value
    is_formula = (
        getattr(cell, "data_type", None) == "f"
        or (isinstance(existing, str) and existing.startswith("="))
    )
    if is_formula:
        _record(anexo, casillero, ws.title, cell_addr, value, origen, "skipped_formula")
        return False

    cell.value = value
    _record(anexo, casillero, ws.title, cell_addr, value, origen, "written")
    return True


def write_trace_sheet(workbook: Workbook) -> None:
    """Genera la hoja TRAZABILIDAD como dashboard interactivo.

    Diseño:
      - Header con título y stats arriba
      - KPI Cards con cifras clave (total escrituras, por anexo)
      - Tabla principal con autofilter en cada columna
      - Formato condicional: color de fondo distinto por anexo
      - Iconos en columna Estado: ✓ escritura ok, ⚠ omitida por fórmula, ⛌ merged
      - Freeze panes para mantener headers visibles
      - Hipervínculos a la celda destino real para navegar rápido

    Pensado para que el auditor pueda filtrar por anexo, por casillero,
    por sheet o por estado y verificar el linaje origen→destino al instante.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from collections import Counter

    trace = get_trace()

    SHEET_NAME = "TRAZABILIDAD"
    if SHEET_NAME in workbook.sheetnames:
        del workbook[SHEET_NAME]
    ws = workbook.create_sheet(SHEET_NAME)

    # ---- Stats globales ----
    written = [t for t in trace if t.get("status") == "written"]
    skipped_f = [t for t in trace if t.get("status") == "skipped_formula"]
    skipped_m = [t for t in trace if t.get("status") == "skipped_merged"]
    by_anexo = Counter(t.get("anexo", "—") for t in written)

    # ---- Estilos ----
    THIN = Side(border_style="thin", color="A0A0A0")
    MEDIUM = Side(border_style="medium", color="2D5F8B")
    BORDER_DATA = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BORDER_KPI = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

    FONT_TITLE = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    FONT_KPI_LABEL = Font(name="Calibri", size=9, bold=True, color="5A6575")
    FONT_KPI_VALUE = Font(name="Calibri", size=18, bold=True, color="2D5F8B")
    FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    FONT_DATA = Font(name="Calibri", size=9)
    FONT_LEGEND = Font(name="Calibri", size=9, italic=True, color="5A6575")
    FILL_TITLE = PatternFill("solid", fgColor="1F3A5F")
    FILL_HEADER = PatternFill("solid", fgColor="4A7BA8")
    FILL_KPI_BG = PatternFill("solid", fgColor="F4F7FB")

    # Color por anexo (paleta tenue)
    ANEXO_COLORS = {
        "INDICE": "FFF9C4",  # amarillo claro
        "A1":     "E1F5FE",  # azul claro
        "A2":     "E8F5E9",  # verde claro
        "A3":     "FCE4EC",  # rosa claro
        "A4":     "F3E5F5",  # morado claro
        "A5":     "FFF3E0",  # naranja claro
        "A6":     "E0F2F1",  # turquesa claro
        "A7":     "E8EAF6",  # índigo claro
        "A8":     "FFEBEE",  # rojo claro
        "A9":     "F1F8E9",  # lima claro
    }

    # ---- Título ----
    ws.merge_cells("A1:G2")
    tcell = ws.cell(1, 1, value="🔗 TRAZABILIDAD · Linaje origen → destino")
    tcell.font = FONT_TITLE
    tcell.fill = FILL_TITLE
    tcell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    # ---- KPI Cards (fila 4-6) ----
    def kpi(row, col, label, value, color="default"):
        ws.cell(row, col, value=label).font = FONT_KPI_LABEL
        ws.cell(row, col).fill = FILL_KPI_BG
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)

        v = ws.cell(row+1, col, value=value)
        v.font = FONT_KPI_VALUE
        v.fill = FILL_KPI_BG
        v.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+1)

        for r in (row, row+1, row+2):
            for c in (col, col+1):
                ws.cell(r, c).border = BORDER_KPI

    kpi(4, 1, "ESCRITURAS EXITOSAS", f"{len(written):,}")
    kpi(4, 3, "FÓRMULAS PROTEGIDAS", f"{len(skipped_f):,}")
    kpi(4, 5, "CELDAS MERGED", f"{len(skipped_m):,}")

    # ---- Stats por anexo (fila 8) ----
    ws.cell(8, 1, value="Por anexo:").font = Font(name="Calibri", size=10, bold=True)
    col_offset = 2
    for anexo, count in sorted(by_anexo.items()):
        c1 = ws.cell(8, col_offset, value=anexo)
        c1.font = Font(name="Calibri", size=9, bold=True)
        c1.alignment = Alignment(horizontal="center")
        c1.fill = PatternFill("solid", fgColor=ANEXO_COLORS.get(anexo, "EEEEEE"))
        c1.border = BORDER_DATA
        c2 = ws.cell(9, col_offset, value=count)
        c2.font = Font(name="Calibri", size=10, bold=True, color="2D5F8B")
        c2.alignment = Alignment(horizontal="center")
        c2.fill = PatternFill("solid", fgColor=ANEXO_COLORS.get(anexo, "EEEEEE"))
        c2.border = BORDER_DATA
        col_offset += 1

    # ---- Leyenda ----
    legend_row = 11
    leg = ws.cell(legend_row, 1, value=(
        "Leyenda: ✓ Escritura exitosa · ⚠ Omitida (era fórmula del template, se respetó) · "
        "⛌ Omitida (era celda combinada) · Filtra por columna usando las flechitas del encabezado."
    ))
    leg.font = FONT_LEGEND
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=7)

    # ---- Tabla principal de escrituras (con autofilter) ----
    table_start = 13
    headers = ["Anexo", "Casillero", "Hoja", "Celda", "Valor escrito", "Origen", "Estado"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(table_start, i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_DATA
    ws.row_dimensions[table_start].height = 26

    # Combinar TODAS las entradas (written + skipped) para auditoría completa
    all_entries = sorted(trace, key=lambda x: (x.get("anexo", ""), x.get("sheet", ""), x.get("cell", "")))

    STATUS_ICON = {
        "written": "✓ OK",
        "skipped_formula": "⚠ Fórmula respetada",
        "skipped_merged": "⛌ Celda combinada",
        "error": "✗ Error",
    }

    for i, entry in enumerate(all_entries, start=table_start + 1):
        anexo = entry.get("anexo", "")
        casillero = entry.get("casillero", "")
        sheet = entry.get("sheet", "")
        cell_addr = entry.get("cell", "")
        valor = entry.get("valor", "")
        origen = entry.get("origen", "")
        status = entry.get("status", "")
        status_disp = STATUS_ICON.get(status, status)

        # Fila completa con fondo del color del anexo
        fill_row = PatternFill("solid", fgColor=ANEXO_COLORS.get(anexo, "FFFFFF"))

        # Anexo
        c1 = ws.cell(i, 1, value=anexo)
        c1.font = Font(name="Calibri", size=9, bold=True, color="2D5F8B")
        c1.alignment = Alignment(horizontal="center", vertical="center")
        # Casillero
        c2 = ws.cell(i, 2, value=casillero)
        c2.font = FONT_DATA
        c2.alignment = Alignment(horizontal="center", vertical="center")
        # Hoja
        c3 = ws.cell(i, 3, value=sheet)
        c3.font = FONT_DATA
        c3.alignment = Alignment(horizontal="left", vertical="center")
        # Celda destino con hipervínculo
        c4 = ws.cell(i, 4, value=cell_addr)
        c4.font = Font(name="Calibri", size=9, color="2D5F8B", underline="single")
        c4.alignment = Alignment(horizontal="center", vertical="center")
        try:
            # Hipervínculo a la hoja+celda real del anexo
            safe_sheet_ref = f"'{sheet}'!{cell_addr}" if " " in sheet else f"{sheet}!{cell_addr}"
            c4.hyperlink = f"#{safe_sheet_ref}"
        except Exception:
            pass
        # Valor
        c5 = ws.cell(i, 5, value=valor if isinstance(valor, (int, float, str)) else str(valor))
        c5.font = FONT_DATA
        if isinstance(valor, (int, float)) and not isinstance(valor, bool):
            c5.number_format = '#,##0.00;-#,##0.00;"—"'
            c5.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c5.alignment = Alignment(horizontal="left", vertical="center")
        # Origen
        c6 = ws.cell(i, 6, value=origen)
        c6.font = FONT_DATA
        c6.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Estado
        c7 = ws.cell(i, 7, value=status_disp)
        if status == "written":
            c7.font = Font(name="Calibri", size=9, color="2E7D32", bold=True)
        elif status == "skipped_formula":
            c7.font = Font(name="Calibri", size=9, color="EF6C00")
        elif status == "skipped_merged":
            c7.font = Font(name="Calibri", size=9, color="757575")
        else:
            c7.font = Font(name="Calibri", size=9, color="C62828")
        c7.alignment = Alignment(horizontal="center", vertical="center")

        for c in range(1, 8):
            ws.cell(i, c).fill = fill_row
            ws.cell(i, c).border = BORDER_DATA

    last_row = table_start + len(all_entries)
    # AutoFilter
    ws.auto_filter.ref = f"A{table_start}:G{last_row}"

    # Anchos de columna
    widths = {"A": 10, "B": 14, "C": 32, "D": 12, "E": 22, "F": 40, "G": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Freeze panes: mantiene título + KPIs + headers visibles
    ws.freeze_panes = f"A{table_start + 1}"

    # Si no hay nada, mensaje informativo
    if not all_entries:
        ws.cell(row=table_start + 1, column=1,
                value="(sin entradas — los fillers no registraron trazabilidad)")


class Filler(Protocol):
    """Protocol every anexo filler implements."""

    anexo_code: str

    def fill(
        self,
        workbook: Workbook,
        session_data: dict,
        anexo_data: dict,
    ) -> dict:
        """Fill cells in the workbook for this anexo's sheet.

        Returns:
            {filled_cells: int, warnings: list[str]}

        IMPORTANT:
        - Only write cells listed in this anexo's cell_map
        - NEVER overwrite cells with formulas (preserve template)
        """
        ...
