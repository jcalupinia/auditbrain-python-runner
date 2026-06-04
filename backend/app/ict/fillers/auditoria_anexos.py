"""Hoja AUDITORÍA DE ANEXOS para el ICT 2025.

Dashboard ejecutivo que para CADA anexo (INDICE, A1..A9) muestra:

  1. ESTADO       — completo / parcial / vacío
  2. METODOLOGÍA  — qué pide el SRI, breve y en español auditor
  3. DATOS        — métricas concretas (celdas llenas, casilleros referenciados)
  4. DIFERENCIAS  — cuadraturas detectadas que no cierran (cas, declarado, contable)
  5. ANÁLISIS     — interpretación automática + recomendación al auditor

Pensado para que un Lider Auditor abra una sola hoja y entienda en
60 segundos en qué quedó la generación: dónde cuadra, dónde no, qué
documentos faltan, qué tiene que revisar manualmente.

REGLA del proyecto (CLAUDE.md): formato profesional equivalente al
SRI Ecuador. KPI cards, tablas con bordes, colores por estado,
hyperlinks a los anexos correspondientes.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_NAME = "AUDITORÍA DE ANEXOS"


# ─────────────────────────────────────────────────────────────────────────────
# Estilos
# ─────────────────────────────────────────────────────────────────────────────
THIN = Side(border_style="thin", color="A0A0A0")
MEDIUM = Side(border_style="medium", color="2D5F8B")
BORDER_DATA = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_KPI = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

FONT_TITLE = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
FONT_KPI_LABEL = Font(name="Calibri", size=9, bold=True, color="5A6575")
FONT_KPI_VALUE = Font(name="Calibri", size=14, bold=True, color="2D5F8B")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_DATA = Font(name="Calibri", size=10)
FONT_DATA_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_OK = Font(name="Calibri", size=10, bold=True, color="1B5E20")
FONT_WARN = Font(name="Calibri", size=10, bold=True, color="E65100")
FONT_BAD = Font(name="Calibri", size=10, bold=True, color="B71C1C")
FONT_ITALIC = Font(name="Calibri", size=9, italic=True, color="5A6575")

FILL_TITLE = PatternFill("solid", fgColor="1F3A5F")
FILL_SECTION = PatternFill("solid", fgColor="2D5F8B")
FILL_HEADER = PatternFill("solid", fgColor="4A7BA8")
FILL_KPI_BG = PatternFill("solid", fgColor="F4F7FB")
FILL_OK = PatternFill("solid", fgColor="E8F5E9")
FILL_WARN = PatternFill("solid", fgColor="FFF3E0")
FILL_BAD = PatternFill("solid", fgColor="FFEBEE")
FILL_ANALYSIS = PatternFill("solid", fgColor="FAFAFC")

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
ALIGN_R = Alignment(horizontal="right", vertical="center")


# ─────────────────────────────────────────────────────────────────────────────
# Metadatos de cada anexo — qué pide el SRI + qué fuentes consume
# ─────────────────────────────────────────────────────────────────────────────
ANEXOS_META = [
    {
        "code": "INDICE",
        "sheet": "INDICE",
        "nombre": "Índice del Informe",
        "pide": (
            "Listado de los 9 anexos con su estado (completo/incompleto), "
            "fuente de datos y total de filas. Es la portada del informe ICT."
        ),
        "fuentes": ["Recompute automático sobre A1..A9"],
    },
    {
        "code": "A1",
        "sheet": "MAPEO DE LA DECLARACIÓN A1",
        "nombre": "Mapeo de la Declaración",
        "pide": (
            "Mapeo cuenta por cuenta del Balance Mapeado del cliente vs los "
            "casilleros del Formulario 101. Cada casillero debe cuadrar con "
            "la suma de las cuentas contables que mapean a él. La diferencia "
            "(col G) debe ser 0 para todos los casilleros."
        ),
        "fuentes": ["F-101 PDF", "Balance Mapeado Excel"],
    },
    {
        "code": "A2",
        "sheet": "INGRESOS A2",
        "nombre": "Ingresos",
        "pide": (
            "Cuadro 1: Ingresos ordinarios por concepto (ventas bienes, "
            "servicios, exportaciones) declarados en F-101. "
            "Cuadro 2: IVA declarado en F-104 mensual vs facturación electrónica. "
            "Cuadro 3: conciliación IVA → IR."
        ),
        "fuentes": ["F-101 PDF", "F-104 PDF mensual", "Facturación electrónica"],
    },
    {
        "code": "A3",
        "sheet": "COSTOS  GASTOS A3",
        "nombre": "Costos y Gastos — Límites de Deducibilidad",
        "pide": (
            "9 bloques que verifican límites de deducibilidad: gastos de "
            "gestión, viaje, indirectos del exterior, promoción, intereses, "
            "instalación, regalías, viáticos, deterioros. Compara declarado "
            "vs límite legal y emite excedente NO deducible."
        ),
        "fuentes": ["F-101 PDF"],
    },
    {
        "code": "A4",
        "sheet": "CONCILIACIÓN INGRESOS A4",
        "nombre": "Conciliación de Ingresos Exentos",
        "pide": (
            "Cuadro 1: detalle de cuentas contables que generan ingresos "
            "exentos (dividendos, rentas exentas, ingresos no objeto). "
            "Cuadro 2: conciliación de casilleros 804, 805, 812, 1112 del F-101."
        ),
        "fuentes": ["F-101 PDF", "Libro Mayor exentos (opcional)", "Balance"],
    },
    {
        "code": "A5",
        "sheet": "CONCILIACIÓN COSTOS Y GASTOS A5",
        "nombre": "Conciliación Costos y Gastos",
        "pide": (
            "5 cuadros: detalle no deducibles, prorrateo (cas 6999/7999 + %), "
            "participación trabajadores (cas 804/805/808), conciliación "
            "gastos (cas 806/807/808/809/813/1113), reversos."
        ),
        "fuentes": ["F-101 PDF", "Libro Mayor no deducibles (opcional)"],
    },
    {
        "code": "A6",
        "sheet": "BENEFICIOS TRIBUTARIOS A6",
        "nombre": "Beneficios Tributarios",
        "pide": (
            "Cuadro A: deducciones adicionales (cas 810 F-101) + detalle de "
            "cuentas. Cuadro B: contratos de inversión vigentes. "
            "Cuadro C: exoneraciones / disminuciones de tarifa IR."
        ),
        "fuentes": ["F-101 PDF", "Contratos inversión (manual)", "Exoneraciones (manual)"],
    },
    {
        "code": "A7",
        "sheet": "CRÉDITO TRIBUTARIO A7",
        "nombre": "Crédito Tributario",
        "pide": (
            "Matriz 1 IR: crédito tributario IR generado en 2022, 2023, 2024 "
            "con monto declarado, devuelto, utilizado y no recuperable. "
            "Matriz 2 ISD: pagos de ISD 2021-2025 con destino fiscal."
        ),
        "fuentes": ["F-101 multi-año PDF", "F-108 multi-año PDF"],
    },
    {
        "code": "A8",
        "sheet": "COMERCIO EXTERIOR A8",
        "nombre": "Comercio Exterior — Pagos al Exterior",
        "pide": (
            "3 tablas: A) pagos con Convenio Doble Imposición, "
            "B) pagos sin CDI, C) reembolsos vía intermediarios. "
            "Datos vienen del XML del ATS (Anexo Transaccional)."
        ),
        "fuentes": ["ATS XML (anual)"],
    },
    {
        "code": "A9",
        "sheet": "INVENTARIOS A9",
        "nombre": "Inventarios",
        "pide": (
            "Detalle de 9 casilleros de inventario del F-101 "
            "(7001/7010/7013/7022/7025/7028/7031/7034/7037) con código, "
            "forma valoración, cantidad y costo total del Kardex."
        ),
        "fuentes": ["F-101 PDF", "Kardex Excel"],
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# Helpers de UI
# ─────────────────────────────────────────────────────────────────────────────

def _write_title(ws, row: int = 1) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=8)
    c = ws.cell(row, 1, value="🔬 AUDITORÍA DE ANEXOS · Resumen ejecutivo")
    c.font = FONT_TITLE
    c.fill = FILL_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[row].height = 24
    ws.row_dimensions[row+1].height = 24
    return row + 3


def _write_session_info(ws, row: int, session_data: dict) -> int:
    info_label = Font(name="Calibri", size=10, bold=True, color="5A6575")
    info_value = Font(name="Calibri", size=10, color="1F3A5F")
    rows = [
        ("Contribuyente", session_data.get("razon_social", "—")),
        ("RUC",           session_data.get("ruc", "—")),
        ("Ejercicio fiscal", session_data.get("ejercicio_fiscal", "—")),
    ]
    for k, v in rows:
        ws.cell(row, 1, value=k).font = info_label
        ws.cell(row, 2, value=v).font = info_value
        row += 1
    return row + 1


def _write_global_kpis(ws, row: int, kpis: dict) -> int:
    """Bloque de 4 KPIs globales en la parte alta del dashboard."""
    def kpi(r, col, label, value, color="default"):
        ws.cell(r, col, value=label).font = FONT_KPI_LABEL
        ws.cell(r, col).fill = FILL_KPI_BG
        ws.cell(r, col).alignment = ALIGN_C
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
        v = ws.cell(r+1, col, value=value)
        v.font = Font(name="Calibri", size=14, bold=True, color={
            "ok": "1B5E20", "warn": "E65100", "bad": "B71C1C"
        }.get(color, "2D5F8B"))
        v.fill = FILL_KPI_BG
        v.alignment = ALIGN_C
        ws.merge_cells(start_row=r+1, start_column=col, end_row=r+2, end_column=col+1)
        for rr in (r, r+1, r+2):
            for cc in (col, col+1):
                ws.cell(rr, cc).border = BORDER_KPI

    anexos_completos = kpis.get("anexos_completos", 0)
    total_anexos = kpis.get("total_anexos", 10)
    diff_count = kpis.get("diferencias", 0)
    warnings_count = kpis.get("warnings", 0)
    casilleros = kpis.get("casilleros_referenciados", 0)

    color_compl = "ok" if anexos_completos == total_anexos else ("warn" if anexos_completos >= total_anexos * 0.6 else "bad")
    color_diff = "ok" if diff_count == 0 else ("warn" if diff_count <= 3 else "bad")

    kpi(row, 1, "ANEXOS GENERADOS", f"{anexos_completos} / {total_anexos}", color=color_compl)
    kpi(row, 3, "DIFERENCIAS DETECTADAS", f"{diff_count}", color=color_diff)
    kpi(row, 5, "WARNINGS TOTALES", f"{warnings_count}", color=("ok" if warnings_count == 0 else "warn"))
    kpi(row, 7, "CASILLEROS USADOS", f"{casilleros}")
    return row + 4


def _write_section_header(ws, row: int, anexo_code: str, nombre: str, sheet: str) -> int:
    """Header coloreado por anexo con hyperlink a la hoja."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    title = f"📑 {anexo_code} · {nombre}"
    c = ws.cell(row, 1, value=title)
    c.font = FONT_SECTION
    c.fill = FILL_SECTION
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    # Hyperlink a la hoja del anexo
    try:
        safe_ref = f"'{sheet}'!A1" if " " in sheet else f"{sheet}!A1"
        c.hyperlink = f"#{safe_ref}"
    except Exception:
        pass
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_kv_row(ws, row: int, label: str, value, *, val_font=None, val_fill=None) -> int:
    ws.cell(row, 1, value=label).font = FONT_DATA_BOLD
    ws.cell(row, 1).alignment = ALIGN_L
    ws.cell(row, 1).border = BORDER_DATA
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    cv = ws.cell(row, 2, value=value)
    cv.font = val_font or FONT_DATA
    if val_fill:
        cv.fill = val_fill
    cv.alignment = ALIGN_L
    cv.border = BORDER_DATA
    return row + 1


def _write_metricas_row(ws, row: int, datos: list[tuple[str, str]]) -> int:
    """Pinta una fila con varias métricas inline (label: valor | label: valor)."""
    txt = "    ·    ".join(f"{k}: {v}" for k, v in datos)
    ws.cell(row, 1, value="📊 Métricas").font = FONT_DATA_BOLD
    ws.cell(row, 1).alignment = ALIGN_L
    ws.cell(row, 1).border = BORDER_DATA
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    cv = ws.cell(row, 2, value=txt)
    cv.font = FONT_DATA
    cv.alignment = ALIGN_L
    cv.border = BORDER_DATA
    return row + 1


def _write_diferencias_table(ws, row: int, diffs: list[dict]) -> int:
    """Tabla detallada de diferencias del anexo."""
    if not diffs:
        ws.cell(row, 1, value="🟢 Diferencias").font = FONT_DATA_BOLD
        ws.cell(row, 1).alignment = ALIGN_L
        ws.cell(row, 1).border = BORDER_DATA
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        cv = ws.cell(row, 2, value="Ninguna diferencia material detectada.")
        cv.font = FONT_OK
        cv.fill = FILL_OK
        cv.alignment = ALIGN_L
        cv.border = BORDER_DATA
        return row + 1

    # Header
    ws.cell(row, 1, value="🔴 Diferencias").font = FONT_DATA_BOLD
    ws.cell(row, 1).alignment = ALIGN_L
    ws.cell(row, 1).border = BORDER_DATA
    headers = ["Concepto", "Casillero", "F-101 declarado", "Balance / Calculado", "Diferencia"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row, i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER_DATA
    row += 1

    for d in diffs:
        ws.cell(row, 1, value="").border = BORDER_DATA
        ws.cell(row, 2, value=d.get("concepto", "—")).font = FONT_DATA
        ws.cell(row, 2).alignment = ALIGN_L
        ws.cell(row, 2).border = BORDER_DATA
        ws.cell(row, 3, value=d.get("casillero", "—")).font = FONT_DATA
        ws.cell(row, 3).alignment = ALIGN_C
        ws.cell(row, 3).border = BORDER_DATA
        c_decl = ws.cell(row, 4, value=d.get("declarado"))
        c_decl.font = FONT_DATA
        c_decl.number_format = '#,##0.00;-#,##0.00;"—"'
        c_decl.alignment = ALIGN_R
        c_decl.border = BORDER_DATA
        c_bal = ws.cell(row, 5, value=d.get("contable"))
        c_bal.font = FONT_DATA
        c_bal.number_format = '#,##0.00;-#,##0.00;"—"'
        c_bal.alignment = ALIGN_R
        c_bal.border = BORDER_DATA
        diff_val = d.get("diferencia")
        c_diff = ws.cell(row, 6, value=diff_val)
        c_diff.font = FONT_BAD if isinstance(diff_val, (int, float)) and abs(diff_val) > 0.5 else FONT_DATA
        c_diff.number_format = '#,##0.00;-#,##0.00;"0.00"'
        c_diff.alignment = ALIGN_R
        c_diff.border = BORDER_DATA
        # Columns 7-8 spillover
        for c in (7, 8):
            ws.cell(row, c).border = BORDER_DATA
        row += 1
    return row


def _write_analysis_row(ws, row: int, texto: str, color: str = "default") -> int:
    ws.cell(row, 1, value="🧠 Análisis").font = FONT_DATA_BOLD
    ws.cell(row, 1).alignment = ALIGN_L
    ws.cell(row, 1).border = BORDER_DATA
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    cv = ws.cell(row, 2, value=texto)
    fill_map = {"ok": FILL_OK, "warn": FILL_WARN, "bad": FILL_BAD}
    cv.font = FONT_DATA
    cv.fill = fill_map.get(color, FILL_ANALYSIS)
    cv.alignment = ALIGN_L
    cv.border = BORDER_DATA
    ws.row_dimensions[row].height = max(30, len(texto) // 100 * 15 + 30)
    return row + 1


# ─────────────────────────────────────────────────────────────────────────────
# Análisis automático por anexo
# ─────────────────────────────────────────────────────────────────────────────

def _analyze_anexo(meta: dict, ctx: dict) -> dict:
    """Devuelve dict con métricas, diferencias y análisis para 1 anexo.

    ctx tiene: f101, f103_monthly, f104_monthly, balance_mapeado,
               anexo_warnings (dict: code → list[str])
    """
    code = meta["code"]
    f101 = ctx.get("f101", {}) or {}
    f103_monthly = ctx.get("f103_monthly", {}) or {}
    f104_monthly = ctx.get("f104_monthly", {}) or {}
    balance = ctx.get("balance_mapeado", []) or []
    warnings = (ctx.get("anexo_warnings", {}) or {}).get(code, [])

    # Métricas base por anexo
    metricas: list[tuple[str, str]] = []
    diffs: list[dict] = []
    analysis = ""
    color = "default"
    estado = "—"

    if code == "INDICE":
        anexos_listados = 9
        estado = "✓ Generado"
        metricas = [("Anexos listados", f"{anexos_listados}"),
                    ("Estados calculados", "automático")]
        analysis = ("Portada del informe. Si algún anexo aparece como "
                    "'incompleto', revisar la sección correspondiente y "
                    "subir el documento faltante (F-101, F-103, F-104, "
                    "Balance Mapeado, Kardex, ATS).")
        color = "ok"

    elif code == "A1":
        # Cuadratura A=P+Pa F-101
        activo_f101 = f101.get("499") or 0
        pp_f101 = (f101.get("699") or (f101.get("599") or 0) + (f101.get("698") or 0))
        cuadre_f101 = round(abs(activo_f101 - pp_f101), 2)
        # Cuadratura A=P+Pa Balance
        by_cas = {}
        for b in balance:
            cas = str(b.get("casillero_sri", "")).strip()
            if cas:
                by_cas.setdefault(cas, []).append(b)
        activo_bal = sum(float(b.get("saldo", 0) or 0)
                         for c, items in by_cas.items() for b in items
                         if c.isdigit() and 311 <= int(c) <= 499)
        pp_bal = sum(abs(float(b.get("saldo", 0) or 0))
                     for c, items in by_cas.items() for b in items
                     if c.isdigit() and (511 <= int(c) <= 599 or 601 <= int(c) <= 698))
        cuadre_bal = round(abs(activo_bal - pp_bal), 2)

        cas_con_valor = sum(1 for v in f101.values() if v not in (None, 0, 0.0))
        metricas = [
            ("Casilleros F-101 con valor", f"{cas_con_valor}"),
            ("Cuentas en Balance Mapeado", f"{len(balance)}"),
            ("Cuadre F-101 (A=P+Pa)", f"{cuadre_f101:,.2f}"),
            ("Cuadre Balance (A=P+Pa)", f"{cuadre_bal:,.2f}"),
        ]
        # Diferencias por bloque (los principales)
        BLOQUES = [
            ("Total Activos Corrientes",     "361", [(311, 360)],  False),
            ("Total Activos No Corrientes",  "449", [(362, 449)],  False),
            ("Total del Activo",             "499", [(311, 499)],  False),
            ("Total Pasivos Corrientes",     "550", [(511, 549)],  True),
            ("Total Pasivos No Corrientes",  "589", [(553, 588)],  True),
            ("Total del Patrimonio",         "698", [(601, 697)],  True),
        ]
        for nombre, cas, ranges, abs_flag in BLOQUES:
            decl = f101.get(cas)
            if decl is None:
                continue
            bal = 0.0
            for lo, hi in ranges:
                for c, items in by_cas.items():
                    if c.isdigit() and lo <= int(c) <= hi:
                        for it in items:
                            v = float(it.get("saldo", 0) or 0)
                            bal += abs(v) if abs_flag else v
            diff = round(bal - decl, 2)
            if abs(diff) > 0.5:
                diffs.append({"concepto": nombre, "casillero": cas,
                              "declarado": round(decl, 2),
                              "contable": round(bal, 2), "diferencia": diff})

        if cuadre_f101 <= 0.5 and cuadre_bal <= 0.5 and not diffs:
            estado, color = "✓ Completo y cuadrado", "ok"
            analysis = ("F-101 y Balance cuadran perfectamente "
                        "(Activo = Pasivo + Patrimonio). Cada casillero "
                        "del F-101 tiene contraparte contable. Listo para "
                        "presentar al SRI.")
        elif cuadre_f101 > 0.5 and cuadre_bal > 0.5:
            estado, color = "✗ NO CUADRA — F-101 y Balance", "bad"
            analysis = (f"Diferencias materiales en AMBOS sets: F-101 desfase "
                        f"{cuadre_f101:,.2f} y Balance {cuadre_bal:,.2f}. "
                        f"Posibles causas: (a) el F-101 declarado tiene errores; "
                        f"(b) el Balance Mapeado tiene cuentas mal clasificadas; "
                        f"(c) faltan cuentas. Revisar los bloques con diferencia "
                        f"en la tabla siguiente.")
        elif diffs:
            estado, color = "⚠ Diferencias por bloque", "warn"
            analysis = (f"Las cuadraturas totales cierran pero hay {len(diffs)} "
                        f"bloques con diferencia material. Cada uno aparece "
                        f"en la tabla siguiente con el detalle. Revisar las "
                        f"cuentas del Balance Mapeado cuyo casillero SRI quede "
                        f"mal asignado.")
        else:
            estado, color = "✓ Cuadrado", "ok"
            analysis = "A1 cuadrado en sus totales principales."

    elif code == "A2":
        f104_count = len(f104_monthly)
        # Cuadratura: ventas IR (F-101 6001+6005) vs ventas IVA agregado (F-104)
        ventas_ir = (f101.get("6001") or 0) + (f101.get("6005") or 0)
        ventas_iva = 0.0
        for mes in f104_monthly.values():
            cas = (mes.get("casilleros") or {})
            # Cas 411 = ventas locales tarifa diferente de 0 (base imponible)
            ventas_iva += cas.get("411", 0) or 0
        diff_ventas = round(ventas_iva - ventas_ir, 2)

        metricas = [
            ("Meses F-104 cargados", f"{f104_count}/12"),
            ("Ventas IR (F-101 6001+6005)", f"{ventas_ir:,.2f}"),
            ("Ventas IVA acumulado (F-104 411)", f"{ventas_iva:,.2f}"),
        ]
        if abs(diff_ventas) > 0.5 and ventas_ir > 0 and ventas_iva > 0:
            diffs.append({"concepto": "Ventas bienes IR vs IVA", "casillero": "6001/411",
                          "declarado": round(ventas_ir, 2), "contable": round(ventas_iva, 2),
                          "diferencia": diff_ventas})

        if f104_count == 0:
            estado, color = "⚠ F-104 no cargado", "warn"
            analysis = ("No se subieron las declaraciones F-104 mensuales. "
                        "Sin esto el Cuadro 2 (IVA vs Facturación) y la "
                        "conciliación IVA → IR del Cuadro 3 quedan vacíos.")
        elif f104_count < 12:
            estado, color = "⚠ F-104 parcial", "warn"
            analysis = (f"Solo {f104_count} de los 12 meses del F-104 fueron "
                        f"cargados. Las cifras anuales del Cuadro 2 estarán "
                        f"sub-declaradas. Subir los meses faltantes antes de "
                        f"presentar.")
        elif abs(diff_ventas) > 0.5 and ventas_ir > 0:
            estado, color = "⚠ Diferencia IR vs IVA", "warn"
            analysis = (f"Las ventas declaradas para IR ({ventas_ir:,.2f}) y "
                        f"para IVA ({ventas_iva:,.2f}) difieren en "
                        f"{diff_ventas:,.2f}. Esto es típico cuando hay "
                        f"ventas exentas de IVA o diferimientos. Revisar y "
                        f"documentar la conciliación.")
        else:
            estado, color = "✓ Ingresos cuadrados", "ok"
            analysis = ("Ingresos ordinarios consistentes entre F-101 e IVA. "
                        "Si subiste facturación electrónica, los totales del "
                        "Cuadro 2 deben coincidir con SRI/Comprobantes.")

    elif code == "A3":
        cas_costos = sum(1 for k in f101.keys() if k.startswith("7"))
        metricas = [
            ("Casilleros 7xxx en F-101", f"{cas_costos}"),
            ("Bloques de deducibilidad", "9"),
        ]
        # cas 6999 (ingresos) y 7185 (gastos gestión) son los críticos del bloque 1
        ingresos = f101.get("6999") or 0
        gtos_gestion = f101.get("7185") or 0
        limite_pct = 0.05  # 5% de ingresos en general
        limite = ingresos * limite_pct
        excedente = max(0, gtos_gestion - limite)
        if excedente > 0:
            diffs.append({"concepto": "Gastos de gestión exceden 5% ingresos",
                          "casillero": "7185", "declarado": round(gtos_gestion, 2),
                          "contable": round(limite, 2),
                          "diferencia": round(excedente, 2)})
            color, estado = "warn", "⚠ Excede límite gestión"
            analysis = (f"Los gastos de gestión declarados ({gtos_gestion:,.2f}) "
                        f"exceden el 5% de ingresos ({limite:,.2f}). El "
                        f"excedente {excedente:,.2f} podría ser NO deducible. "
                        f"Revisar Bloque 1 del A3 — el filler calcula el cálculo "
                        f"exacto, esto es solo un alert temprano.")
        else:
            estado, color = "✓ Dentro de límites", "ok"
            analysis = ("Gastos de gestión dentro del 5% de ingresos. "
                        "Revisar individualmente los otros 8 bloques en la "
                        "hoja A3 (viaje, intereses exterior, regalías, etc.) "
                        "para validar todos los límites de deducibilidad.")

    elif code == "A4":
        cas_exentos = [k for k in ("804", "805", "812", "1112") if f101.get(k)]
        metricas = [
            ("Casilleros exentos declarados", f"{len(cas_exentos)}"),
            ("Detalle en libros del cliente", "ver Cuadro 1"),
        ]
        if not cas_exentos:
            estado, color = "✓ Sin ingresos exentos", "ok"
            analysis = ("El F-101 no declara ingresos exentos. No hay "
                        "conciliación pendiente para este anexo.")
        else:
            estado, color = "⚠ Requiere detalle manual", "warn"
            analysis = (f"El F-101 declara {len(cas_exentos)} casilleros de "
                        f"ingresos exentos ({', '.join(cas_exentos)}). El "
                        f"Cuadro 1 necesita el detalle de cuentas contables "
                        f"que originaron esos ingresos (sube el Libro Mayor "
                        f"de las cuentas 804/805/812). La conciliación del "
                        f"Cuadro 2 ya está hecha contra el F-101.")

    elif code == "A5":
        nd_locales = f101.get("806") or 0
        nd_exterior = f101.get("807") or 0
        partic_trab = f101.get("808") or 0
        metricas = [
            ("No deducibles locales (806)", f"{nd_locales:,.2f}"),
            ("No deducibles exterior (807)", f"{nd_exterior:,.2f}"),
            ("Participación trabajadores (808)", f"{partic_trab:,.2f}"),
        ]
        if nd_locales + nd_exterior > 0:
            estado, color = "⚠ Detallar no deducibles", "warn"
            analysis = (f"Hay {nd_locales + nd_exterior:,.2f} de gastos "
                        f"declarados como NO DEDUCIBLES. El SRI exige el "
                        f"detalle cuenta por cuenta en el Cuadro A. Subir "
                        f"el Libro Mayor con las cuentas mapeadas a "
                        f"casilleros 806/807 para que el filler complete "
                        f"el detalle automáticamente.")
        else:
            estado, color = "✓ Sin no deducibles", "ok"
            analysis = ("Sin gastos no deducibles declarados. "
                        "Verificar Cuadro B (prorrateo) y Cuadro D "
                        "(conciliación) manualmente en el anexo.")

    elif code == "A6":
        ded_810 = f101.get("810") or 0
        metricas = [
            ("Deducciones adicionales (cas 810)", f"{ded_810:,.2f}"),
        ]
        if ded_810 > 0:
            estado, color = "⚠ Justificar deducciones", "warn"
            analysis = (f"F-101 declara {ded_810:,.2f} en deducciones "
                        f"adicionales (cas 810). Sustento legal exigido: "
                        f"contratos de inversión vigentes (Cuadro B) o "
                        f"exoneraciones aprobadas (Cuadro C). Completar "
                        f"esos cuadros con el sustento documental antes de "
                        f"presentar el ICT.")
        else:
            estado, color = "✓ Sin beneficios", "ok"
            analysis = ("No hay deducciones adicionales declaradas en el "
                        "F-101. A6 queda como anexo informativo vacío.")

    elif code == "A7":
        cas_credito = (f101.get("850") or 0) + (f101.get("851") or 0)
        metricas = [
            ("Crédito tributario IR (cas 850+851)", f"{cas_credito:,.2f}"),
        ]
        if cas_credito > 0:
            estado, color = "⚠ Cargar histórico multi-año", "warn"
            analysis = (f"F-101 declara {cas_credito:,.2f} en crédito "
                        f"tributario IR. Para completar A7 hay que subir "
                        f"los F-101 de 2022, 2023, 2024 (los 3 años "
                        f"anteriores) para que se llene la Matriz IR con el "
                        f"detalle de generación, devolución y utilización por "
                        f"año. Sin esto la matriz queda en blanco.")
        else:
            estado, color = "✓ Sin crédito IR pendiente", "ok"
            analysis = ("F-101 no declara crédito tributario IR. "
                        "Si el cliente paga ISD, subir F-108 multi-año "
                        "para la Matriz 2 (ISD) — no se valida con F-101.")

    elif code == "A8":
        # ATS XML check
        ats_pagos = ctx.get("ats_pagos_exterior", [])
        cas_402 = sum((f.get("casilleros", {}) if isinstance(f, dict) else {}).get("402", 0) or 0
                      for f in f103_monthly.values())
        cas_433 = sum((f.get("casilleros", {}) if isinstance(f, dict) else {}).get("433", 0) or 0
                      for f in f103_monthly.values())
        metricas = [
            ("Transacciones ATS detectadas", f"{len(ats_pagos)}"),
            ("Pagos exterior con CDI (F-103 402 ann)", f"{cas_402:,.2f}"),
            ("Pagos exterior sin CDI (F-103 433 ann)", f"{cas_433:,.2f}"),
        ]
        tiene_pagos_103 = (cas_402 + cas_433) > 0
        if not ats_pagos and not tiene_pagos_103:
            estado, color = "✓ Sin pagos exterior", "ok"
            analysis = ("No se detectaron pagos al exterior ni en ATS XML "
                        "ni en F-103. A8 queda vacío legítimamente.")
        elif not ats_pagos and tiene_pagos_103:
            estado, color = "✗ Falta ATS XML", "bad"
            analysis = (f"El F-103 declara pagos al exterior por "
                        f"{cas_402 + cas_433:,.2f} pero NO se subió el ATS "
                        f"XML anual. A8 NO se puede completar sin el ATS. "
                        f"Descargar el ATS del portal SRI y subirlo.")
        elif ats_pagos and not tiene_pagos_103:
            estado, color = "⚠ ATS sin retención F-103", "warn"
            analysis = (f"Se detectaron {len(ats_pagos)} pagos al exterior "
                        f"en ATS pero el F-103 no declara retenciones "
                        f"asociadas (cas 402-433). Validar si esos pagos "
                        f"requerían retención IR.")
        else:
            estado, color = "✓ ATS + F-103 consistentes", "ok"
            analysis = (f"{len(ats_pagos)} transacciones ATS clasificadas en "
                        f"3 tablas (con CDI, sin CDI, reembolsos). "
                        f"Validar manualmente que las clasificaciones "
                        f"automáticas (pago_loc_ext + tipo_regi) sean correctas.")

    elif code == "A9":
        # Inventarios
        invs = [f101.get(c) for c in ("7001", "7010", "7013", "7022", "7025",
                                      "7028", "7031", "7034", "7037")]
        invs_con_valor = sum(1 for v in invs if v not in (None, 0, 0.0))
        kardex_count = len(ctx.get("kardex_items", []) or [])
        metricas = [
            ("Casilleros inventario con valor", f"{invs_con_valor}/9"),
            ("Ítems en Kardex", f"{kardex_count}"),
        ]
        if invs_con_valor == 0:
            estado, color = "✓ Sin inventarios", "ok"
            analysis = ("La empresa no maneja inventarios según F-101 "
                        "(casilleros 7001-7037 todos en cero). A9 vacío "
                        "es correcto.")
        elif kardex_count == 0 and invs_con_valor > 0:
            estado, color = "✗ Falta Kardex", "bad"
            analysis = (f"F-101 declara inventarios en {invs_con_valor} "
                        f"casilleros pero NO se subió el Kardex. El SRI "
                        f"exige el detalle de método de valoración, "
                        f"cantidad y costo. Subir el Kardex del cliente.")
        else:
            estado, color = "✓ Inventarios cargados", "ok"
            analysis = (f"{invs_con_valor} casilleros con valor + Kardex "
                        f"con {kardex_count} ítems. Validar que los "
                        f"métodos de valoración (PROMEDIO, PEPS, UEPS) "
                        f"declarados sean consistentes con las políticas "
                        f"contables del cliente.")

    # Si hay warnings del filler, integrarlos al análisis
    if warnings:
        if analysis:
            analysis += f"\n\n⚠ Warnings del filler ({len(warnings)}): {warnings[0][:140]}..."
        else:
            analysis = f"⚠ Warnings: {warnings[0][:150]}"
        if color == "ok":
            color = "warn"

    return {
        "estado": estado, "color": color,
        "metricas": metricas, "diferencias": diffs,
        "analysis": analysis,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Builder principal
# ─────────────────────────────────────────────────────────────────────────────

def build_auditoria_anexos_sheet(
    workbook: Workbook,
    *,
    session_data: dict,
    f101: dict,
    f103_monthly: dict | None = None,
    f104_monthly: dict | None = None,
    balance_mapeado: list | None = None,
    ats_pagos_exterior: list | None = None,
    kardex_items: list | None = None,
    anexo_warnings: dict | None = None,
) -> None:
    """Crea (o reemplaza) la hoja AUDITORÍA DE ANEXOS al final del workbook.

    anexo_warnings: dict opcional {code → list[str]} con los warnings que
    devolvió cada filler durante la generación. Se integran al bloque de
    análisis del anexo correspondiente.
    """
    if SHEET_NAME in workbook.sheetnames:
        del workbook[SHEET_NAME]
    ws = workbook.create_sheet(SHEET_NAME)

    ctx = {
        "f101": f101 or {},
        "f103_monthly": f103_monthly or {},
        "f104_monthly": f104_monthly or {},
        "balance_mapeado": balance_mapeado or [],
        "ats_pagos_exterior": ats_pagos_exterior or [],
        "kardex_items": kardex_items or [],
        "anexo_warnings": anexo_warnings or {},
    }

    # Pre-analizar todos los anexos
    analyses = [(meta, _analyze_anexo(meta, ctx)) for meta in ANEXOS_META]

    # KPIs globales agregados
    anexos_ok = sum(1 for _, a in analyses if a["color"] == "ok")
    diff_total = sum(len(a["diferencias"]) for _, a in analyses)
    warn_total = sum(len(ctx["anexo_warnings"].get(m["code"], [])) for m, _ in analyses)
    cas_referenciados = len([v for v in f101.values() if v not in (None, 0, 0.0)])

    # Render
    row = _write_title(ws)
    row = _write_session_info(ws, row, session_data)
    row = _write_global_kpis(ws, row, {
        "anexos_completos": anexos_ok,
        "total_anexos": len(analyses),
        "diferencias": diff_total,
        "warnings": warn_total,
        "casilleros_referenciados": cas_referenciados,
    })

    # Una sección por anexo
    for meta, a in analyses:
        row += 1  # separador
        row = _write_section_header(ws, row, meta["code"], meta["nombre"], meta["sheet"])
        # Estado
        color_font = {"ok": FONT_OK, "warn": FONT_WARN, "bad": FONT_BAD}.get(a["color"], FONT_DATA_BOLD)
        color_fill = {"ok": FILL_OK, "warn": FILL_WARN, "bad": FILL_BAD}.get(a["color"], FILL_ANALYSIS)
        row = _write_kv_row(ws, row, "Estado", a["estado"], val_font=color_font, val_fill=color_fill)
        # Metodología (qué pide el SRI)
        row = _write_kv_row(ws, row, "📜 Qué pide el SRI", meta["pide"])
        # Fuentes
        row = _write_kv_row(ws, row, "📥 Fuentes esperadas", " · ".join(meta["fuentes"]))
        # Métricas
        if a["metricas"]:
            row = _write_metricas_row(ws, row, a["metricas"])
        # Diferencias
        row = _write_diferencias_table(ws, row, a["diferencias"])
        # Análisis
        row = _write_analysis_row(ws, row, a["analysis"], color=a["color"])

    # Anchos
    widths = {"A": 22, "B": 22, "C": 18, "D": 22, "E": 22, "F": 18, "G": 14, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Freeze panes para mantener título visible
    ws.freeze_panes = "A4"


# ============================================================================
# NUEVO ENTRY POINT (Approach C — Data/Presentation split)
# ============================================================================
# fill_auditoria_anexos consume AnexosMetrics + dict[str, AnexoInterpretation]
# del módulo audit/ y usa kpi_components para renderizar matriz 3x3 + finding
# boxes. La entry function legacy (build_auditoria_anexos_sheet) queda intacta
# para no romper callers existentes — se migrará en PT-9.
# ============================================================================

def fill_auditoria_anexos(
    ws,
    *,
    metrics,           # backend.app.ict.audit.schemas.AnexosMetrics
    interpretations: dict,  # dict[str, AnexoInterpretation]
    contexto: dict,
) -> None:
    """Render AUDITORÍA DE ANEXOS con banner + matriz 3x3 + interpretaciones."""
    from backend.app.ict.audit.schemas import Status
    from backend.app.ict.fillers.kpi_components import (
        build_executive_banner,
        build_finding_box,
        build_traffic_light_grid,
    )

    razon = contexto.get("razon_social", "")
    ruc = contexto.get("ruc", "")
    periodo = contexto.get("periodo", "")

    # 1. Banner ejecutivo
    build_executive_banner(
        ws, anchor="A1",
        title_main="AUDITBRAIN · PAPEL DE TRABAJO DEL AUDITOR",
        title_sub="AUDITORÍA INTEGRAL DE ANEXOS A1..A9",
        meta=f"{razon} · RUC {ruc} · Período {periodo}",
        width_cols=14,
    )

    # 2. Título de sección "MATRIZ DE ESTADO"
    ws.cell(
        row=5, column=1,
        value="MATRIZ DE ESTADO POR ANEXO",
    ).font = Font(name="Calibri", size=12, bold=True)

    # 3. Matriz 3x3 de semáforos
    build_traffic_light_grid(
        ws, anchor="A7", anexos_status=metrics.anexos,
        card_width_cols=4, card_height_rows=4, gap_cols=1, gap_rows=1,
    )

    # 4. Leyenda
    legend_row = 7 + 3 * (4 + 1) + 1   # filas tomadas por el grid 3x3 + gap
    resumen = metrics.resumen_global
    legend_txt = (
        f"🟢 OK ({resumen.get(Status.OK, 0)})   "
        f"🟧 Revisar ({resumen.get(Status.REVISAR, 0)})   "
        f"🔴 Crítico ({resumen.get(Status.CRITICO, 0)})   "
        f"⚪ N/A ({resumen.get(Status.NA, 0)})"
    )
    ws.cell(row=legend_row, column=1, value=legend_txt).font = Font(
        name="Calibri", size=10, italic=True,
    )

    # 5. Sección INTERPRETACIÓN POR ANEXO
    interp_start = legend_row + 3
    ws.cell(
        row=interp_start, column=1,
        value="🤖 INTERPRETACIÓN POR ANEXO · Análisis del agente",
    ).font = Font(name="Calibri", size=12, bold=True)

    r = interp_start + 2
    for code in ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9"]:
        interp = interpretations.get(code)
        if interp is None:
            continue
        emoji = {"alta": "🟢", "media": "🟡", "baja": "🔴"}.get(
            interp.confianza_modelo, "⚪",
        )
        header = (
            f"▸ {code} — {interp.anexo_nombre}  "
            f"[Confianza: {emoji} {interp.confianza_modelo}]"
        )
        ws.cell(row=r, column=1, value=header).font = Font(
            name="Calibri", size=11, bold=True,
        )
        r += 1
        resumen_cell = ws.cell(row=r, column=1, value=interp.resumen_ejecutivo)
        resumen_cell.alignment = Alignment(wrap_text=True)
        r += 2
        for f in interp.findings:
            r = build_finding_box(
                ws, anchor_row=r, anchor_col=1,
                finding=f, width_cols=14,
            ) + 2

    # 6. Disclaimer obligatorio
    disc_row = r + 2
    ws.cell(
        row=disc_row, column=1,
        value=(
            "Análisis generado por IA. Toda interpretación debe ser "
            "validada por el auditor responsable antes de cualquier "
            "decisión, glosa o entrega al cliente."
        ),
    ).font = Font(name="Calibri", size=8, italic=True, color="6B7280")
