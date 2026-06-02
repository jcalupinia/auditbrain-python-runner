"""Helpers de formato visual para los anexos del ICT 2025.

REGLA GENERAL (también en CLAUDE.md): los anexos del ICT generados DEBEN
verse profesionalmente presentados, equivalentes al formato oficial del
SRI Ecuador (referencia: ``1791240154001_Anexo ICT_2024_07.xlsx`` en
``docs/referencias/``). Esto significa:

  - Tamaños de fuente consistentes (8 para datos, 9 para headers).
  - Bordes thin en cada fila de datos.
  - Filas en blanco como separadores visuales entre bloques (Activos
    Corrientes / No Corrientes / Pasivos / Patrimonio / Resultados).
  - Merged cells para que las columnas "casillero", "nombre", "valor
    declarado" y "diferencia" cubran TODA la altura del grupo de
    cuentas contables que mapean a ese casillero.
  - Anchos de columna fijos (A=14, B=36, C=14, D=14, E=34, F=14, G=14,
    H=22 para Observaciones, I=14).
  - Filas TOTAL en NEGRITA con borde superior doble.

Cada filler de anexo (a1..a9) debe llamar al helper apropiado de este
módulo al final de su ``fill()`` para aplicar el formato.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ----- Estilos base reutilizables -----
THIN = Side(border_style="thin", color="000000")
THICK = Side(border_style="thick", color="000000")
DOUBLE = Side(border_style="double", color="000000")

DATA_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOTAL_BORDER = Border(left=THIN, right=THIN, top=DOUBLE, bottom=DOUBLE)
NO_BORDER = Border()

DATA_FONT = Font(name="Calibri", size=9, bold=False)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
SECTION_HEADER_FONT = Font(name="Calibri", size=11, bold=True)

HEADER_FILL = PatternFill("solid", fgColor="2D5F8B")  # azul corporativo SRI
TOTAL_FILL = PatternFill("solid", fgColor="E8F1F8")   # azul muy claro
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0") # azul medio claro

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


# ----- Anchos de columna estándar para A1 -----
A1_COLUMN_WIDTHS = {
    "A": 12.0,   # Casillero
    "B": 40.0,   # Nombre Casillero
    "C": 16.0,   # Valor declarado
    "D": 18.0,   # Código cuenta contable
    "E": 38.0,   # Nombre cuenta contable
    "F": 16.0,   # Saldo
    "G": 16.0,   # Diferencia
    "H": 24.0,   # Observaciones
}


def apply_column_widths(ws, widths: dict[str, float]) -> None:
    """Aplica anchos de columna definidos en el dict."""
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def _is_writable(ws, row: int, col: int) -> bool:
    """True si la celda (row,col) no es MergedCell (puede ser modificada)."""
    cell = ws.cell(row, col)
    return not isinstance(cell, MergedCell)


def safe_apply_style(ws, row: int, col: int, *,
                     font=None, fill=None, border=None, alignment=None,
                     number_format=None) -> None:
    """Aplica estilos a una celda saltando si es MergedCell. Algunos
    atributos no se pueden setear en MergedCells y romperían el filler."""
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    if font is not None: cell.font = font
    if fill is not None: cell.fill = fill
    if border is not None: cell.border = border
    if alignment is not None: cell.alignment = alignment
    if number_format is not None: cell.number_format = number_format


def safe_merge(ws, range_str: str) -> None:
    """Hace merge de un rango, manejando excepciones (ya merged, etc.)."""
    try:
        ws.merge_cells(range_str)
    except Exception:
        pass


def format_a1_sheet(
    ws,
    *,
    casillero_groups: list[dict],
    first_data_row: int = 13,
    total_cols: int = 9,
) -> None:
    """Aplica formato profesional a la hoja A1 después del llenado.

    Args:
        ws: worksheet del A1.
        casillero_groups: lista de dicts {casillero, row_start, row_end,
            is_total} para saber dónde aplicar merges y separadores.
        first_data_row: fila donde empiezan los datos (default 13).
        total_cols: cantidad de columnas a estilizar (1..total_cols).

    Acciones:
        1. Aplica anchos de columna estándar A1_COLUMN_WIDTHS.
        2. Para cada grupo de casillero con N>1 cuentas:
           merge A, B, C, G a lo alto del grupo.
        3. Aplica bordes thin a cada celda con datos.
        4. Marca filas TOTAL en negrita y borde doble.
        5. Aplica número con 2 decimales a las columnas C, F, G.
        6. Wrap text en B y E (nombres largos).
        7. Centra verticalmente cada fila merged.
    """
    apply_column_widths(ws, A1_COLUMN_WIDTHS)

    for grp in casillero_groups:
        cas = grp["casillero"]
        r_start = grp["row_start"]
        r_end = grp["row_end"]
        is_total = grp.get("is_total", False)

        # Merge A, B, C, G a lo alto del grupo si hay más de una fila
        if r_end > r_start:
            for col_letter in ("A", "B", "C", "G"):
                safe_merge(ws, f"{col_letter}{r_start}:{col_letter}{r_end}")

        # Estilo de filas
        is_data_row = True
        for r in range(r_start, r_end + 1):
            font_to_use = TOTAL_FONT if is_total else DATA_FONT
            fill_to_use = TOTAL_FILL if is_total else None
            border_to_use = TOTAL_BORDER if is_total else DATA_BORDER

            for c in range(1, total_cols + 1):
                # Alineación según columna
                if c == 1:  # A — Casillero
                    align = ALIGN_CENTER
                elif c == 2:  # B — Nombre Casillero
                    align = ALIGN_LEFT
                elif c in (3, 6, 7):  # C, F, G — Numéricos
                    align = ALIGN_RIGHT
                elif c == 4:  # D — Código cuenta
                    align = ALIGN_LEFT
                elif c == 5:  # E — Nombre cuenta
                    align = ALIGN_LEFT
                else:
                    align = ALIGN_LEFT

                safe_apply_style(
                    ws, r, c,
                    font=font_to_use,
                    fill=fill_to_use,
                    border=border_to_use,
                    alignment=align,
                )

                # Formato número para cols C, F, G
                if c in (3, 6, 7):
                    safe_apply_style(ws, r, c, number_format='#,##0.00;-#,##0.00;""')

        # Altura mínima de fila para que el wrap se vea bien
        for r in range(r_start, r_end + 1):
            if ws.row_dimensions[r].height is None or ws.row_dimensions[r].height < 16:
                ws.row_dimensions[r].height = 16

        # Separador visual: fila en blanco después del grupo (sólo si NO es TOTAL final)
        # Esto lo deja la lógica del filler (insertando fila en blanco entre bloques).
