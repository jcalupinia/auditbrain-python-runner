"""Reusable visual components for ICT audit artifacts (Excel).

These helpers render KPI cards, traffic-light grids, executive banners
and finding boxes following the SRI Ecuador + Big 4 hybrid aesthetic.

Each helper takes a worksheet + anchor + payload (typed via audit schemas)
and applies styles. They never decide the layout: the caller chooses where.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from backend.app.ict.audit.schemas import (
    AnexoFinding,
    AnexoStatus,
    Status,
)

# Color palette: SRI institutional + Big 4 accents
STATUS_COLORS: dict[Status, dict[str, str]] = {
    Status.OK:       {"fill": "DCFCE7", "border": "16A34A", "text": "166534"},
    Status.REVISAR:  {"fill": "FEF3C7", "border": "F59E0B", "text": "92400E"},
    Status.CRITICO:  {"fill": "FEE2E2", "border": "C0392B", "text": "991B1B"},
    Status.NA:       {"fill": "F1F5F9", "border": "94A3B8", "text": "475569"},
}

SEVERITY_BORDERS = {
    "critico":     {"color": "C0392B", "weight": "medium"},
    "material":    {"color": "E67E22", "weight": "medium"},
    "leve":        {"color": "F1C40F", "weight": "thin"},
    "informativo": {"color": "3498DB", "weight": "thin"},
}

# SRI brand colors (from oficial 2024 ARCOLANDS template)
SRI_BLUE = "1E3A8A"
SRI_LIGHT = "DBEAFE"


def _thin_border(color: str = "94A3B8") -> Border:
    side = Side(border_style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _medium_border(color: str) -> Border:
    side = Side(border_style="medium", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _parse_anchor(anchor: str) -> tuple[int, int]:
    """Convert 'B2' → (row=2, col=2)."""
    col_letters = "".join(c for c in anchor if c.isalpha())
    row = int("".join(c for c in anchor if c.isdigit()))
    col = 0
    for ch in col_letters.upper():
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return row, col


def build_kpi_card(
    ws: Worksheet,
    *,
    anchor: str,
    title: str,
    value: str,
    status: Status,
    subtitle: str = "",
    width_cols: int = 3,
    height_rows: int = 4,
) -> None:
    """Render a KPI card at `anchor` (e.g. 'B2'), spanning width_cols × height_rows.

    Layout:
      Row N:     [TITLE bold]
      Row N+1:   [empty separator]
      Row N+2:   [LARGE VALUE + emoji status]
      Row N+3:   [subtitle small]
    """
    row, col = _parse_anchor(anchor)
    colors = STATUS_COLORS[status]
    fill = PatternFill("solid", fgColor=colors["fill"])
    border = _medium_border(colors["border"])

    title_cell = ws.cell(row=row, column=col, value=title)
    title_cell.font = Font(name="Calibri", size=10, bold=True,
                           color=colors["text"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    if width_cols > 1:
        ws.merge_cells(start_row=row, end_row=row,
                       start_column=col, end_column=col + width_cols - 1)

    value_row = row + 2
    value_cell = ws.cell(row=value_row, column=col, value=value)
    value_cell.font = Font(name="Calibri", size=18, bold=True,
                           color=colors["text"])
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    if width_cols > 1:
        ws.merge_cells(start_row=value_row, end_row=value_row,
                       start_column=col, end_column=col + width_cols - 1)

    if subtitle:
        sub_row = row + 3
        sub_cell = ws.cell(row=sub_row, column=col, value=subtitle)
        sub_cell.font = Font(name="Calibri", size=8, italic=True,
                             color=colors["text"])
        sub_cell.alignment = Alignment(horizontal="center")
        if width_cols > 1:
            ws.merge_cells(start_row=sub_row, end_row=sub_row,
                           start_column=col, end_column=col + width_cols - 1)

    # Fill background and border across all cells of the card
    for r in range(row, row + height_rows):
        for c in range(col, col + width_cols):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = border


def build_traffic_light_grid(
    ws: Worksheet,
    *,
    anchor: str,
    anexos_status: list[AnexoStatus],
    card_width_cols: int = 3,
    card_height_rows: int = 4,
    gap_cols: int = 1,
    gap_rows: int = 1,
) -> None:
    """Render a 3×3 grid of mini KPI cards, one per anexo A1..A9."""
    row, col = _parse_anchor(anchor)
    assert len(anexos_status) == 9, "traffic light grid expects exactly 9 anexos"

    for idx, st in enumerate(anexos_status):
        grid_row = idx // 3
        grid_col = idx % 3
        cell_row = row + grid_row * (card_height_rows + gap_rows)
        cell_col = col + grid_col * (card_width_cols + gap_cols)
        anchor_str = f"{get_column_letter(cell_col)}{cell_row}"
        emoji = {"ok": "🟢", "revisar": "🟧",
                 "critico": "🔴", "na": "⚪"}[st.status.value]
        # Title = codigo + emoji in subtitle row
        build_kpi_card(
            ws, anchor=anchor_str, title=st.codigo,
            value=emoji, status=st.status,
            subtitle=st.observacion_corta,
            width_cols=card_width_cols, height_rows=card_height_rows,
        )


def build_executive_banner(
    ws: Worksheet,
    *,
    anchor: str,
    title_main: str,
    title_sub: str,
    meta: str = "",
    width_cols: int = 10,
) -> None:
    """Render the executive banner at top of the sheet."""
    row, col = _parse_anchor(anchor)

    main_cell = ws.cell(row=row, column=col, value=title_main)
    main_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    main_cell.fill = PatternFill("solid", fgColor=SRI_BLUE)
    main_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, end_row=row,
                   start_column=col, end_column=col + width_cols - 1)

    sub_cell = ws.cell(row=row + 1, column=col, value=title_sub)
    sub_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    sub_cell.fill = PatternFill("solid", fgColor=SRI_BLUE)
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row + 1, end_row=row + 1,
                   start_column=col, end_column=col + width_cols - 1)

    if meta:
        meta_cell = ws.cell(row=row + 2, column=col, value=meta)
        meta_cell.font = Font(name="Calibri", size=9, italic=True,
                              color="FFFFFF")
        meta_cell.fill = PatternFill("solid", fgColor=SRI_BLUE)
        meta_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row + 2, end_row=row + 2,
                       start_column=col, end_column=col + width_cols - 1)


def build_finding_box(
    ws: Worksheet,
    *,
    anchor_row: int,
    anchor_col: int,
    finding: AnexoFinding,
    width_cols: int = 8,
) -> int:
    """Render an AnexoFinding as a bordered box. Returns the last row written."""
    sev = finding.severity
    border_spec = SEVERITY_BORDERS.get(sev, SEVERITY_BORDERS["informativo"])
    border = Border(
        left=Side(border_style=border_spec["weight"], color=border_spec["color"]),
        right=Side(border_style=border_spec["weight"], color=border_spec["color"]),
        top=Side(border_style=border_spec["weight"], color=border_spec["color"]),
        bottom=Side(border_style=border_spec["weight"], color=border_spec["color"]),
    )
    emoji = {"critico": "🔴", "material": "🟧",
             "leve": "🟡", "informativo": "🔵"}[sev]

    r = anchor_row
    # Row 1: Severity header
    title_text = f"{emoji} {sev.upper()} · {finding.titulo}"
    ws.cell(row=r, column=anchor_col, value=title_text).font = Font(
        name="Calibri", size=11, bold=True, color=border_spec["color"]
    )
    ws.merge_cells(start_row=r, end_row=r,
                   start_column=anchor_col,
                   end_column=anchor_col + width_cols - 1)
    r += 1

    sections = [
        ("Descripción técnica", finding.descripcion_tecnica),
        ("Implicación tributaria", finding.implicacion_tributaria),
        ("Recomendación", finding.recomendacion),
    ]
    for label, value in sections:
        ws.cell(row=r, column=anchor_col, value=label).font = Font(
            name="Calibri", size=9, bold=True
        )
        r += 1
        ws.cell(row=r, column=anchor_col, value=value).font = Font(
            name="Calibri", size=9
        )
        ws.cell(row=r, column=anchor_col).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, end_row=r,
                       start_column=anchor_col,
                       end_column=anchor_col + width_cols - 1)
        r += 1

    # Footer: casilleros + monto
    monto_str = (f"${finding.monto_disputa:,.2f}"
                 if finding.monto_disputa is not None else "—")
    footer = (
        f"Casilleros: {', '.join(finding.casilleros_afectados) or '—'}  |  "
        f"Monto disputa: {monto_str}"
    )
    ws.cell(row=r, column=anchor_col, value=footer).font = Font(
        name="Calibri", size=8, italic=True
    )
    ws.merge_cells(start_row=r, end_row=r,
                   start_column=anchor_col,
                   end_column=anchor_col + width_cols - 1)
    last_row = r

    # Apply borders to all cells in the box
    for rr in range(anchor_row, last_row + 1):
        for cc in range(anchor_col, anchor_col + width_cols):
            ws.cell(row=rr, column=cc).border = border

    return last_row
