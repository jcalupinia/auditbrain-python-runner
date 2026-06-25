"""Helpers para EXPANDIR bloques tabulares en los anexos del ICT.

Cuando un cliente declara más datos que filas disponibles en un cuadro de
plantilla (ej. A5 Cuadro A: 5 filas pero 15 casilleros no deducibles), hay
que insertar filas en runtime SIN perder ningún dato (REGLA SUPREMA del
proyecto: nunca truncar información del cliente).

`openpyxl.insert_rows` tiene 3 limitaciones conocidas (verificadas en 3.1.5):
  1. NO reajusta las referencias de las fórmulas desplazadas.
  2. NO desplaza los merged cells (quedan en su posición vieja).
  3. NO copia el estilo/alto a las filas nuevas (quedan en blanco).

Este módulo aporta:
  - `shift_formula_rows`: reajusta refs de fila >= umbral en una fórmula.
  - `expand_tabular_block`: inserta N filas en un cuadro y deja la hoja
    consistente (fórmulas reajustadas, merges recreados, estilo copiado).
"""
from __future__ import annotations

import re
from copy import copy

from openpyxl.utils import get_column_letter


# Referencia de celda LOCAL (misma hoja): columna(s) + fila, con $ opcional.
# - lookbehind: no debe venir precedida de letra/dígito/_/$/! (evita capturar
#   parte de un nombre o una ref a OTRA hoja escrita tras '!').
# - lookahead: no debe seguir '(' (evita confundir una función con una ref).
_CELL_REF = re.compile(
    r"(?<![A-Za-z0-9_$!])(\$?)([A-Z]{1,3})(\$?)(\d+)(?![A-Za-z0-9_(])"
)

# Una ref externa completa tras un nombre de hoja: !C500 ó !$C$500:$C$520
_EXTERNAL_REF = re.compile(r"!\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?")


def _shift_segment(seg: str, threshold: int, amount: int) -> str:
    def repl(m: re.Match) -> str:
        col_abs, col, row_abs, row = m.group(1), m.group(2), m.group(3), m.group(4)
        r = int(row)
        if r >= threshold:
            r += amount
        return f"{col_abs}{col}{row_abs}{r}"

    return _CELL_REF.sub(repl, seg)


def shift_formula_rows(formula, threshold: int, amount: int):
    """Suma ``amount`` a cada referencia de fila >= ``threshold`` en una
    fórmula de Excel. Respeta:
      - refs absolutas ($B$30), rangos (G28:G32)
      - literales de texto entre comillas dobles ("0,00%")
      - nombres de hoja entre comillas simples y sus refs externas
        ('DATOS F-101'!C500 NO se desplaza — vive en otra hoja)

    Si ``formula`` no es una cadena que empiece con '=' o ``amount`` es 0,
    se devuelve sin cambios.
    """
    if amount == 0 or not isinstance(formula, str) or not formula.startswith("="):
        return formula

    out: list[str] = []
    i, n = 0, len(formula)
    while i < n:
        ch = formula[i]
        if ch == '"':
            # Literal de texto: copiar intacto hasta el cierre.
            j = formula.find('"', i + 1)
            if j == -1:
                j = n - 1
            out.append(formula[i:j + 1])
            i = j + 1
        elif ch == "'":
            # Nombre de hoja entre comillas: copiar intacto.
            j = formula.find("'", i + 1)
            if j == -1:
                j = n - 1
            out.append(formula[i:j + 1])
            i = j + 1
            # Si sigue '!<ref>', es una ref a OTRA hoja → copiar sin desplazar.
            if i < n and formula[i] == "!":
                m = _EXTERNAL_REF.match(formula[i:])
                if m:
                    out.append(m.group(0))
                    i += m.end()
                else:
                    out.append("!")
                    i += 1
        else:
            # Segmento normal (fórmula local) hasta la próxima comilla.
            j = i
            while j < n and formula[j] not in "\"'":
                j += 1
            out.append(_shift_segment(formula[i:j], threshold, amount))
            i = j
    return "".join(out)


def _copy_cell_style(src_cell, dst_cell) -> None:
    """Copia el estilo (no el valor) de una celda a otra."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def expand_tabular_block(
    ws,
    *,
    insert_at: int,
    amount: int,
    style_row: int,
    inner_merges: list[tuple[int, int]] | None = None,
    last_col: int = 13,
) -> None:
    """Inserta ``amount`` filas en ``insert_at`` y deja la hoja consistente.

    `openpyxl.insert_rows` es poco fiable al desplazar: pierde valores, bordes
    y merges de algunas celdas (verificado en 3.1.5 — caso A5 con 5 filas
    insertadas perdía los casilleros 807/808 del Cuadro D y sus bordes). En
    vez de confiar en él, se toma un SNAPSHOT completo (valor + estilo) de
    todas las celdas a desplazar ANTES de insertar y se RE-APLICA después en
    la posición + amount. Así el resultado es determinista e idéntico al
    original, solo desplazado.

    Pasos:
      1. Snapshot (valor + estilo) de cada celda en filas >= insert_at.
      2. Capturar merges (>= insert_at) y alturas de fila; eliminar esos
         merges para que el desplazamiento no los corrompa.
      3. insert_rows(insert_at, amount).
      4. Re-aplicar el snapshot a la posición + amount (valor con fórmulas
         reajustadas + estilo completo). Re-aplicar alturas.
      5. Recrear los merges desplazados (+amount).
      6. Formatear las filas nuevas desde ``style_row`` + merges internos.

    Args:
      insert_at:   fila donde se insertan las nuevas (las que estaban aquí
                   y debajo bajan +amount).
      amount:      número de filas a insertar (>0).
      style_row:   fila plantilla cuyo estilo/alto se copia a las nuevas.
      inner_merges: lista de (col_ini, col_fin) a fusionar dentro de cada
                   fila nueva (ej. [(5,6),(7,8),(9,10)] para E:F,G:H,I:J).
      last_col:    última columna (1-based) a snapshotear/formatear.
    """
    if amount <= 0:
        return

    max_row = ws.max_row

    # 1) Snapshot de valor + estilo de TODAS las celdas en filas >= insert_at
    #    (incluye celdas vacías con borde — las que openpyxl suele perder).
    snapshot: dict[tuple[int, int], dict] = {}
    for r in range(insert_at, max_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(r, c)
            snapshot[(r, c)] = {
                "value": cell.value,
                "font": copy(cell.font),
                "border": copy(cell.border),
                "fill": copy(cell.fill),
                "number_format": cell.number_format,
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            }

    # 2) Capturar merges (>= insert_at) y alturas; eliminar esos merges para
    #    que el insert_rows no los deje en estado inconsistente.
    old_merges = [
        (mc.min_col, mc.min_row, mc.max_col, mc.max_row)
        for mc in ws.merged_cells.ranges
        if mc.min_row >= insert_at
    ]
    for col_ini, row_ini, col_fin, row_fin in old_merges:
        try:
            ws.unmerge_cells(start_row=row_ini, start_column=col_ini,
                             end_row=row_fin, end_column=col_fin)
        except (KeyError, ValueError):
            pass
    old_heights = {
        r: ws.row_dimensions[r].height
        for r in range(insert_at, max_row + 1)
        if ws.row_dimensions[r].height is not None
    }

    # 3) Insertar las filas.
    ws.insert_rows(insert_at, amount)

    # 4) Re-aplicar snapshot (valor + estilo) a la posición + amount. Las
    #    fórmulas se reajustan (refs de fila >= insert_at → +amount).
    for (r, c), snap in snapshot.items():
        dst = ws.cell(r + amount, c)
        val = snap["value"]
        if isinstance(val, str) and val.startswith("="):
            val = shift_formula_rows(val, threshold=insert_at, amount=amount)
        dst.value = val
        dst.font = snap["font"]
        dst.border = snap["border"]
        dst.fill = snap["fill"]
        dst.number_format = snap["number_format"]
        dst.alignment = snap["alignment"]
        dst.protection = snap["protection"]
    for r, h in old_heights.items():
        ws.row_dimensions[r + amount].height = h

    # 5) Recrear los merges desplazados (+amount).
    for col_ini, row_ini, col_fin, row_fin in old_merges:
        ws.merge_cells(
            start_row=row_ini + amount, start_column=col_ini,
            end_row=row_fin + amount, end_column=col_fin,
        )

    # 6) Formatear las filas nuevas desde style_row + merges internos.
    src_height = ws.row_dimensions[style_row].height
    for new_row in range(insert_at, insert_at + amount):
        if src_height is not None:
            ws.row_dimensions[new_row].height = src_height
        for col in range(1, last_col + 1):
            _copy_cell_style(ws.cell(style_row, col), ws.cell(new_row, col))
        for col_ini, col_fin in (inner_merges or []):
            ws.merge_cells(
                start_row=new_row, start_column=col_ini,
                end_row=new_row, end_column=col_fin,
            )
