"""Pure functions to compute quantitative audit metrics from a workbook.

These functions are SIDE-EFFECT-FREE: they only read openpyxl workbooks
and return Pydantic models. No Excel writes, no I/O, no LLM calls.
"""
from __future__ import annotations

from decimal import Decimal
from typing import Optional

from openpyxl.workbook import Workbook

from backend.app.ict.audit.classifiers import semaforo_from_diff
from backend.app.ict.audit.schemas import (
    A1Metrics,
    AnexoStatus,
    AnexosMetrics,
    Status,
)

ANEXO_CODES = ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9"]
ANEXO_NOMBRES = {
    "A1": "Mapeo del balance general",
    "A2": "Conciliación de ingresos",
    "A3": "Costos y gastos",
    "A4": "Conciliación de ingresos (ajustes)",
    "A5": "Conciliación de costos (ajustes)",
    "A6": "Beneficios tributarios",
    "A7": "Crédito tributario IVA",
    "A8": "Comercio exterior",
    "A9": "Inventarios",
}


def _read_cas_value(sheet, cas: str) -> Optional[Decimal]:
    """Find a row in DATOS F-101 where col A == cas and return col C as Decimal."""
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row and row[0].value and str(row[0].value).strip() == cas:
            val = row[2].value if len(row) > 2 else None
            if val is None:
                return None
            return Decimal(str(val))
    return None


def compute_a1_metrics(wb: Workbook) -> A1Metrics:
    """Compute A1 mapeo metrics from a workbook with DATOS F-101 and A1 sheets."""
    f101 = wb["DATOS F-101"] if "DATOS F-101" in wb.sheetnames else None
    a1 = wb["A1"] if "A1" in wb.sheetnames else None

    activo_total = _read_cas_value(f101, "499") if f101 else None
    pasivo_pat_total = _read_cas_value(f101, "699") if f101 else None
    activo_total = activo_total or Decimal("0")
    pasivo_pat_total = pasivo_pat_total or Decimal("0")
    diferencia = activo_total - pasivo_pat_total

    cas_total = 0
    cas_mapeados = 0
    cas_sin_contrapartida: list[str] = []
    if a1 is not None:
        for row in a1.iter_rows(min_row=2, values_only=False):
            cas_cell = row[0].value if len(row) > 0 else None
            if cas_cell is None:
                continue
            cas = str(cas_cell).strip()
            if not cas or not cas[0].isdigit():
                continue
            cas_total += 1
            contable = row[5].value if len(row) > 5 else None
            if contable not in (None, "", 0):
                cas_mapeados += 1
            else:
                cas_sin_contrapartida.append(cas)

    cobertura = (cas_mapeados / cas_total * 100.0) if cas_total > 0 else 0.0
    status = semaforo_from_diff(diferencia, activo_total)

    return A1Metrics(
        activo_total=activo_total,
        pasivo_patrimonio_total=pasivo_pat_total,
        diferencia=diferencia,
        status_cuadre=status,
        cobertura_mapeo_pct=round(cobertura, 2),
        cas_mapeados=cas_mapeados,
        cas_total=cas_total,
        cas_sin_contrapartida=cas_sin_contrapartida,
    )


def _status_for_anexo(wb: Workbook, code: str) -> AnexoStatus:
    """Compute status for a single anexo by inspecting its sheet."""
    nombre = ANEXO_NOMBRES.get(code, code)
    if code not in wb.sheetnames:
        return AnexoStatus(
            codigo=code, nombre=nombre, status=Status.NA,
            observacion_corta="Hoja no generada",
        )
    sheet = wb[code]
    has_data = any(
        cell.value not in (None, "")
        for row in sheet.iter_rows(min_row=2, max_row=10, values_only=False)
        for cell in row
    )
    if not has_data:
        return AnexoStatus(
            codigo=code, nombre=nombre, status=Status.NA,
            observacion_corta="Sin datos",
        )
    return AnexoStatus(
        codigo=code, nombre=nombre, status=Status.OK,
        observacion_corta="Generado",
    )


def compute_anexos_metrics(wb: Workbook) -> AnexosMetrics:
    """Compute status snapshot for all 9 anexos."""
    statuses = [_status_for_anexo(wb, code) for code in ANEXO_CODES]
    resumen: dict[Status, int] = {s: 0 for s in Status}
    for st in statuses:
        resumen[st.status] += 1
    return AnexosMetrics(anexos=statuses, resumen_global=resumen)
