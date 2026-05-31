"""Parser for SRI Facturación Electrónica monthly report.

The client exports from "SRI en Línea" → "Comprobantes" → resumen mensual
as an Excel file. Different SRI export formats are handled tolerantly.

Returns aggregated monthly totals for use in A2 Cuadro 2 (columnas G-H).
"""

from __future__ import annotations

import re
from io import BytesIO

from openpyxl import load_workbook

from backend.app.ict.parsers.balance_excel import _norm


def parse_facturacion(excel_bytes: bytes) -> dict:
    """Parse SRI facturación electrónica report.

    Returns:
        {
            'meses': {
                'MM': {'emitidas': float, 'anuladas': float, 'neto': float},
                ...
            },
            'totales': {'emitidas': float, 'anuladas': float, 'neto': float},
            'errores': []
        }
    """
    try:
        wb = load_workbook(BytesIO(excel_bytes), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        return {
            "meses": {},
            "totales": {"emitidas": 0.0, "anuladas": 0.0, "neto": 0.0},
            "errores": [f"Excel inválido: {e}"],
        }

    ws = wb.active

    # Find header row: look for "mes", "periodo", or "fecha" alongside
    # "emitidas", "total facturas", "facturado", etc.
    header_row: int | None = None
    col_map: dict[str, int] = {}

    for r in range(1, min(20, ws.max_row) + 1):
        row_cells = list(ws[r])
        row_vals = [_norm(c.value) for c in row_cells]

        for i, h in enumerate(row_vals):
            if ("mes" in h or "periodo" in h or "fecha" in h) and "mes" not in col_map:
                col_map["mes"] = i
            if (
                "emitidas" in h
                or "total facturas" in h
                or "facturado" in h
                or "total_emitidas" in h
                or "numero" in h
            ) and "emitidas" not in col_map:
                col_map["emitidas"] = i
            if (
                "anuladas" in h
                or "notas de credito" in h
                or "nota de credito" in h
                or " nc" in h
                or h.startswith("nc")
                or "canceladas" in h
            ) and "anuladas" not in col_map:
                col_map["anuladas"] = i

        if "mes" in col_map and "emitidas" in col_map:
            header_row = r
            break

    if header_row is None:
        return {
            "meses": {},
            "totales": {"emitidas": 0.0, "anuladas": 0.0, "neto": 0.0},
            "errores": [
                "No se detectó encabezado con columnas 'Mes/Periodo' y "
                "'Emitidas/Total'. Asegúrese de exportar el reporte de "
                "facturación electrónica desde SRI en Línea."
            ],
        }

    meses: dict[str, dict] = {}
    total_emit = 0.0
    total_anul = 0.0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue

        mes_raw = row[col_map["mes"]] if col_map["mes"] < len(row) else None
        if mes_raw is None or str(mes_raw).strip() == "":
            continue

        mes_str = str(mes_raw).strip()

        # Normalize to MM: accept "01", "1", "enero", "2025/01", "01/2025", etc.
        m = re.search(r"\b(\d{1,2})\b", mes_str)
        if not m:
            continue
        mes_num = int(m.group(1))
        if not 1 <= mes_num <= 12:
            continue
        mes_key = f"{mes_num:02d}"

        def _to_float(raw) -> float:
            try:
                return float(raw) if raw is not None else 0.0
            except (ValueError, TypeError):
                return 0.0

        emit = _to_float(
            row[col_map["emitidas"]] if col_map["emitidas"] < len(row) else None
        )
        anul = _to_float(
            row[col_map["anuladas"]]
            if "anuladas" in col_map and col_map["anuladas"] < len(row)
            else None
        )

        meses[mes_key] = {
            "emitidas": emit,
            "anuladas": anul,
            "neto": emit - anul,
        }
        total_emit += emit
        total_anul += anul

    return {
        "meses": meses,
        "totales": {
            "emitidas": total_emit,
            "anuladas": total_anul,
            "neto": total_emit - total_anul,
        },
        "errores": [],
    }
