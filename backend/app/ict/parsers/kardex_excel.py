"""Parser para Kardex de inventarios Excel.

Lee items con: código de cuenta, descripción, forma valoración,
cantidad existencia, costo total. Tolerante a formatos de ERPs ecuatorianos.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook


def parse_kardex(excel_bytes: bytes) -> dict:
    """Returns {items: [{codigo_cuenta, descripcion, forma_valoracion, cantidad, costo_total}], errores}."""
    try:
        wb = load_workbook(BytesIO(excel_bytes), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        return {"items": [], "errores": [f"Excel inválido: {e}"]}

    ws = wb.active
    from backend.app.ict.parsers.balance_excel import _norm

    header_row = None
    col_map: dict[str, int] = {}
    for r in range(1, min(20, ws.max_row) + 1):
        row = [_norm(c.value) for c in ws[r]]
        for i, h in enumerate(row):
            if ("codigo" in h or "código" in h) and "codigo_cuenta" not in col_map:
                col_map["codigo_cuenta"] = i
            elif ("descripcion" in h or "descripción" in h or "nombre" in h) and "descripcion" not in col_map:
                col_map["descripcion"] = i
            elif ("forma" in h or "valoracion" in h or "valoración" in h or "metodo" in h or "método" in h) and "forma" not in col_map:
                col_map["forma"] = i
            elif ("cantidad" in h or "existencia" in h or "stock" in h) and "cantidad" not in col_map:
                col_map["cantidad"] = i
            elif ("costo" in h or "valor" in h) and "costo_total" not in col_map:
                col_map["costo_total"] = i
        if "codigo_cuenta" in col_map and "costo_total" in col_map:
            header_row = r
            break

    if header_row is None:
        return {"items": [], "errores": ["No se detectó encabezado del Kardex"]}

    items: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        codigo = row[col_map["codigo_cuenta"]] if col_map["codigo_cuenta"] < len(row) else None
        if codigo is None or str(codigo).strip() == "":
            continue
        items.append({
            "codigo_cuenta": str(codigo).strip(),
            "descripcion": str(row[col_map["descripcion"]]).strip() if "descripcion" in col_map and col_map["descripcion"] < len(row) and row[col_map["descripcion"]] is not None else "",
            "forma_valoracion": str(row[col_map["forma"]]).strip() if "forma" in col_map and col_map["forma"] < len(row) and row[col_map["forma"]] is not None else "PROMEDIO",
            "cantidad": str(row[col_map["cantidad"]]).strip() if "cantidad" in col_map and col_map["cantidad"] < len(row) and row[col_map["cantidad"]] is not None else "",
            "costo_total": float(row[col_map["costo_total"]]) if col_map["costo_total"] < len(row) and row[col_map["costo_total"]] is not None else 0.0,
        })

    return {"items": items, "errores": []}
