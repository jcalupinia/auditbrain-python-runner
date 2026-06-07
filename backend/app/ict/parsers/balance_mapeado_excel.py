"""Parser for Balance Mapeado Excel (cliente ya mapeó cada cuenta a su casillero SRI).

Estructura esperada:
- Headers en fila 11 (tolerante: escanea hasta fila 20)
- Datos desde fila 13 (o la siguiente a los headers)
- A = Cod.Cuenta.Contable (opcional, algunas filas son subcuentas sin código propio)
- B = Descripción
- D = Códigos SRI (el casillero F-101, OBLIGATORIO para mapeo)
- E = Saldos 31 DIC (numérico, OBLIGATORIO)

Las filas sin D o sin E (numérico) se ignoran (son agrupadores).
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

_EXPECTED_HEADER_KEYWORDS = {
    "codigo": ["cod.cuenta.contable", "cod cuenta contable", "código cuenta", "cuenta"],
    "descripcion": ["descripción cuenta contable", "descripcion cuenta", "descripción"],
    "casillero_sri": ["códigos sri", "codigos sri", "código sri", "sri"],
    "saldo": ["saldos 31 dic", "saldo final", "saldo"],
}


def _norm(s) -> str:
    return str(s or "").strip().lower()


def _find_header(ws, max_scan: int = 20) -> tuple[int, dict[str, int]] | tuple[None, None]:
    """Find the header row + column indices.

    Returns (header_row_idx, {field: col_idx_0_based}) or (None, None).
    Field names: codigo, descripcion, casillero_sri, saldo.
    Codigo column is optional.
    """
    for r in range(1, min(max_scan, ws.max_row) + 1):
        row = [_norm(c.value) for c in ws[r]]
        col_map: dict[str, int] = {}
        for ci, cell in enumerate(row):
            for field, keywords in _EXPECTED_HEADER_KEYWORDS.items():
                if field in col_map:
                    continue
                if any(kw in cell for kw in keywords):
                    col_map[field] = ci
                    break
        # casillero_sri AND saldo are minimum required
        if "casillero_sri" in col_map and "saldo" in col_map:
            return r, col_map
    return None, None


def parse_balance_mapeado(excel_bytes: bytes) -> dict:
    """Parse Balance Mapeado Excel.

    Returns {
        'cuentas': [
            {'casillero_sri': '311', 'codigo': '5BS.11101.002',
             'descripcion': 'Caja Chica', 'saldo': 8500.0,
             'saldo_vacio': False},
            ...
        ],
        'errores': [],
        'advertencias': [],
        'cuentas_sin_saldo': [...],  # cuentas con cas pero saldo vacío
    }

    REGLA OBLIGATORIA (2026-06-07, pedido cliente):
    Cuando una fila del plan tiene CASILLERO asignado pero SALDO vacío
    (None / blank), NO se debe omitir silenciosamente. Se incluye en la
    salida con `saldo=0.0` y `saldo_vacio=True`, además de agregar una
    advertencia explícita a `advertencias` para que el auditor sepa que
    el cliente puede haber olvidado completar el saldo.

    Histórico: bug 2026-06-07 (cas 545 PROPHAR) — la cuenta 2011001
    ANTICIPOS DE CLIENTES con cas 545 tenía saldo vacío y el parser la
    descartaba silenciosamente. El A1 no la mostraba. El auditor no se
    enteraba de la omisión. La regla nueva detecta esto.
    """
    try:
        wb = load_workbook(BytesIO(excel_bytes), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        return {
            "cuentas": [], "errores": [f"Excel inválido: {e}"],
            "advertencias": [], "cuentas_sin_saldo": [],
        }

    ws = wb.active
    header_row, col_map = _find_header(ws)
    if header_row is None:
        return {
            "cuentas": [],
            "errores": [
                "No se detectó fila de encabezado. Esperado: 'Códigos SRI' y 'Saldo' como mínimo. "
                "Verifica que el archivo sea un Balance Mapeado con esas columnas."
            ],
            "advertencias": [], "cuentas_sin_saldo": [],
        }

    col_codigo = col_map.get("codigo")
    col_descripcion = col_map.get("descripcion")
    col_casillero = col_map["casillero_sri"]
    col_saldo = col_map["saldo"]

    cuentas: list[dict] = []
    errores: list[str] = []
    advertencias: list[str] = []
    cuentas_sin_saldo: list[dict] = []
    last_codigo = ""  # for sub-rows that inherit parent's code

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        # Skip totally empty rows
        if all(v is None for v in row):
            continue

        casillero_raw = row[col_casillero] if col_casillero < len(row) else None
        saldo_raw = row[col_saldo] if col_saldo < len(row) else None

        # === FILAS SIN CASILLERO ===
        # Si no hay casillero, es una fila agrupadora o vacía. Recordar codigo
        # para sub-filas posteriores y continuar.
        if casillero_raw is None:
            if col_codigo is not None and col_codigo < len(row) and row[col_codigo] is not None:
                last_codigo = str(row[col_codigo]).strip()
            continue

        # === REGLA NUEVA: fila CON cas pero SIN saldo ===
        # Antes (bug 2026-06-07): se ignoraba silenciosamente.
        # Ahora: se incluye con saldo=0.0 y se registra advertencia.
        if saldo_raw is None:
            casillero = str(casillero_raw).strip()
            if casillero.endswith(".0"):
                casillero = casillero[:-2]

            codigo = ""
            if col_codigo is not None and col_codigo < len(row) and row[col_codigo] is not None:
                codigo = str(row[col_codigo]).strip()
                last_codigo = codigo
            else:
                codigo = last_codigo

            descripcion = ""
            if (
                col_descripcion is not None
                and col_descripcion < len(row)
                and row[col_descripcion] is not None
            ):
                descripcion = str(row[col_descripcion]).strip()

            # Solo registrar si tiene código contable o descripción (ignorar
            # filas totalmente vacías excepto el cas, que no aportan).
            if codigo or descripcion:
                cuenta_vacia = {
                    "casillero_sri": casillero,
                    "codigo": codigo,
                    "descripcion": descripcion,
                    "saldo": 0.0,
                    "saldo_vacio": True,
                    "_source_excel_row": row_idx,
                }
                cuentas.append(cuenta_vacia)
                cuentas_sin_saldo.append(cuenta_vacia)
                advertencias.append(
                    f"Fila {row_idx}: cuenta {codigo or '?'} mapeada al casillero "
                    f"{casillero} NO tiene saldo (vacío). Cliente posiblemente "
                    f"omitió este saldo. Cuenta: '{descripcion[:60]}'"
                )
            continue

        try:
            saldo = float(saldo_raw)
        except (ValueError, TypeError):
            errores.append(f"Saldo inválido para casillero {casillero_raw}: {saldo_raw!r}")
            continue

        casillero = str(casillero_raw).strip()
        # Strip ".0" if Excel stored as float (e.g. 311.0 -> "311")
        if casillero.endswith(".0"):
            casillero = casillero[:-2]

        codigo = ""
        if col_codigo is not None and col_codigo < len(row) and row[col_codigo] is not None:
            codigo = str(row[col_codigo]).strip()
            last_codigo = codigo  # update parent for subsequent sub-rows
        else:
            # Sub-row: inherit last parent codigo
            codigo = last_codigo

        descripcion = ""
        if col_descripcion is not None and col_descripcion < len(row) and row[col_descripcion] is not None:
            descripcion = str(row[col_descripcion]).strip()

        cuentas.append({
            "casillero_sri": casillero,
            "codigo": codigo,
            "descripcion": descripcion,
            "saldo": saldo,
            "saldo_vacio": False,
        })

    return {
        "cuentas": cuentas,
        "errores": errores,
        "advertencias": advertencias,
        "cuentas_sin_saldo": cuentas_sin_saldo,
    }
