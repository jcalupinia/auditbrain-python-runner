"""Parser para Libro Mayor analítico Excel.

Similar al Balance pero con detalle de movimientos.
Para A4 (cuentas de ingresos exentos) y A5 (cuentas de gastos no deducibles).
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook


def parse_mayor(excel_bytes: bytes, cuentas_filtro: list[str] | None = None) -> dict:
    """Parse libro mayor. Optionally filter by account code prefixes.

    Returns {movimientos: [{codigo, nombre, debe, haber, saldo, tipo}], errores: []}.
    """
    try:
        wb = load_workbook(BytesIO(excel_bytes), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        return {"movimientos": [], "errores": [f"Excel inválido: {e}"]}

    ws = wb.active
    from backend.app.ict.parsers.balance_excel import _find_header_row, _norm
    header_row, col_map = _find_header_row(ws)
    if header_row is None:
        return {"movimientos": [], "errores": ["Encabezado no detectado"]}

    header = [_norm(c.value) for c in ws[header_row]]
    extra: dict[str, int] = {}
    for i, h in enumerate(header):
        if "debe" in h and "debe" not in extra:
            extra["debe"] = i
        elif "haber" in h and "haber" not in extra:
            extra["haber"] = i
        elif "tipo" in h and "tipo" not in extra:
            extra["tipo"] = i

    movimientos: list[dict] = []
    col_codigo = col_map["codigo"]
    col_nombre = col_map.get("nombre")
    col_saldo = col_map["saldo"]

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        codigo_raw = row[col_codigo] if col_codigo < len(row) else None
        if codigo_raw is None or str(codigo_raw).strip() == "":
            continue
        codigo = str(codigo_raw).strip()
        if cuentas_filtro and not any(codigo.startswith(p) for p in cuentas_filtro):
            continue
        movimientos.append({
            "codigo": codigo,
            "nombre": str(row[col_nombre]).strip() if col_nombre is not None and row[col_nombre] is not None else "",
            "saldo": float(row[col_saldo]) if row[col_saldo] is not None else 0.0,
            "debe": float(row[extra["debe"]]) if "debe" in extra and row[extra["debe"]] is not None else 0.0,
            "haber": float(row[extra["haber"]]) if "haber" in extra and row[extra["haber"]] is not None else 0.0,
            "tipo": str(row[extra["tipo"]]) if "tipo" in extra and row[extra["tipo"]] is not None else "",
        })

    return {"movimientos": movimientos, "errores": []}
