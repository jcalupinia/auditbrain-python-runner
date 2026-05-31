"""Parser for Balance de Comprobación Excel with tolerant header detection."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

_CODIGO_KEYWORDS = ("codigo", "código", "cuenta", "code")
_NOMBRE_KEYWORDS = ("nombre", "descripcion", "descripción", "name", "concepto")
_SALDO_KEYWORDS = ("saldo final", "saldo", "valor", "monto")


def _norm(text) -> str:
    return str(text or "").strip().lower()


def _find_header_row(ws, max_scan: int = 20):
    for row_idx in range(1, min(max_scan, ws.max_row) + 1):
        row = [_norm(c.value) for c in ws[row_idx]]
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if any(k in cell for k in _CODIGO_KEYWORDS) and "codigo" not in col_map:
                col_map["codigo"] = col_idx
            elif any(k in cell for k in _NOMBRE_KEYWORDS) and "nombre" not in col_map:
                col_map["nombre"] = col_idx
            elif any(k in cell for k in _SALDO_KEYWORDS) and "saldo" not in col_map:
                col_map["saldo"] = col_idx
        if "codigo" in col_map and "saldo" in col_map:
            return row_idx, col_map
    return None, None


def parse_balance(excel_bytes: bytes) -> dict:
    """Returns {'cuentas': {codigo: {nombre, saldo}}, 'errores': []}."""
    try:
        wb = load_workbook(BytesIO(excel_bytes), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        return {"cuentas": {}, "errores": [f"Excel inválido: {e}"]}

    ws = wb.active
    header_row, col_map = _find_header_row(ws)
    if header_row is None:
        return {
            "cuentas": {},
            "errores": [
                "No se detectó fila de encabezado con 'Código' y 'Saldo'. "
                "Asegúrate de que tu Excel tenga columnas claramente etiquetadas."
            ],
        }

    cuentas: dict[str, dict] = {}
    errores: list[str] = []
    col_codigo = col_map["codigo"]
    col_nombre = col_map.get("nombre")
    col_saldo = col_map["saldo"]

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        codigo_raw = row[col_codigo] if col_codigo < len(row) else None
        if codigo_raw is None or str(codigo_raw).strip() == "":
            continue
        codigo = str(codigo_raw).strip()
        nombre = (
            str(row[col_nombre]).strip()
            if col_nombre is not None and col_nombre < len(row) and row[col_nombre] is not None
            else ""
        )
        saldo_raw = row[col_saldo] if col_saldo < len(row) else None
        try:
            saldo = float(saldo_raw) if saldo_raw is not None else 0.0
        except (ValueError, TypeError):
            errores.append(f"Saldo inválido para código {codigo}")
            continue
        cuentas[codigo] = {"nombre": nombre, "saldo": saldo}

    return {"cuentas": cuentas, "errores": errores}
