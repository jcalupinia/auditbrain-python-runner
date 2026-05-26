"""Tests del excel_assembler que puebla la plantilla baked-in DM."""

import datetime
import io

from openpyxl import load_workbook

from backend.app.aud.obligaciones_fiscales import excel_assembler


def _empty_dm6():
    return {"rows": [
        {"mes": m, "c411": None, "c413": None, "c415": None, "c417": None,
         "c419": None, "c421": None, "c429": None, "c480": None, "c499": None,
         "c529": None, "has_data": False}
        for m in ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                  "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    ]}


def _empty_dm7():
    return {"rows": [
        {"mes": m, "c721": None, "c723": None, "c725": None, "c727": None,
         "c729": None, "c731": None, "c799": None, "has_data": False}
        for m in ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                  "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    ]}


def test_assemble_returns_valid_xlsx_bytes():
    out = excel_assembler.assemble(
        cliente_name="EMPRESA TEST S.A.",
        period_label="Ejercicio 2025",
        period_end=datetime.date(2025, 12, 31),
        prepared_by_name="JC",
        reviewed_by_name="AU",
        dm6_data=_empty_dm6(),
        dm7_data=_empty_dm7(),
    )
    assert isinstance(out, bytes)
    wb = load_workbook(io.BytesIO(out))
    assert "DM6 IVA" in wb.sheetnames
    assert "DM7 Retenciones x pagar" in wb.sheetnames


def test_assemble_writes_dm7_january_casilleros():
    dm7 = _empty_dm7()
    dm7["rows"][0] = {
        "mes": "Enero", "c721": 8594.74, "c723": 25.41, "c725": 304.62,
        "c727": 0.00, "c729": 997.30, "c731": 37.50, "c799": 9959.57,
        "has_data": True,
    }
    out = excel_assembler.assemble(
        cliente_name="X", period_label="2025", period_end=None,
        prepared_by_name=None, reviewed_by_name=None,
        dm6_data=_empty_dm6(), dm7_data=dm7,
    )
    wb = load_workbook(io.BytesIO(out))
    dm7_ws = wb["DM7 Retenciones x pagar"]
    # Enero está en fila 21 (DM7_FIRST_ROW)
    assert dm7_ws["H21"].value == 8594.74   # casillero 721 (10%)
    assert dm7_ws["I21"].value == 25.41     # casillero 723 (20%)
    assert dm7_ws["J21"].value == 304.62    # casillero 725 (30%)
    assert dm7_ws["K21"].value == 997.30    # casillero 729 (70%)
    assert dm7_ws["L21"].value == 37.50     # casillero 731 (100%)
    assert dm7_ws["M21"].value == 0.00      # casillero 727 (50%)


def test_assemble_writes_dm6_january_casilleros():
    dm6 = _empty_dm6()
    dm6["rows"][0] = {
        "mes": "Enero", "c411": 704667.18, "c413": 100.0, "c415": 200.0,
        "c417": 300.0, "c419": 717710.66, "c421": 105700.08,
        "c429": 107656.60, "c480": 717710.66, "c499": 107656.60, "c529": 87403.98,
        "c412": None, "c414": None, "c416": None, "c418": None,
        "has_data": True,
    }
    out = excel_assembler.assemble(
        cliente_name="X", period_label="2025", period_end=None,
        prepared_by_name=None, reviewed_by_name=None,
        dm6_data=dm6, dm7_data=_empty_dm7(),
    )
    wb = load_workbook(io.BytesIO(out))
    dm6_ws = wb["DM6 IVA"]
    # Enero en fila 20 (DM6_FIRST_ROW)
    assert dm6_ws["C20"].value == 200.0    # c415 -> col C
    assert dm6_ws["D20"].value == 100.0    # c413 -> col D
    assert dm6_ws["E20"].value == 300.0    # c417 -> col E


def test_assemble_preserves_template_formulas_on_other_columns():
    """No debemos pisar formulas como B20=DM5 Ventas!C19."""
    out = excel_assembler.assemble(
        cliente_name="X", period_label="2025", period_end=None,
        prepared_by_name=None, reviewed_by_name=None,
        dm6_data=_empty_dm6(), dm7_data=_empty_dm7(),
    )
    # Reabrir con keep_links/keep_formulas (load_workbook default)
    wb = load_workbook(io.BytesIO(out))
    dm6_ws = wb["DM6 IVA"]
    # B20 debe seguir siendo la formula original (no sobreescrita por None)
    v = dm6_ws["B20"].value
    assert v is not None
    assert isinstance(v, str) and v.startswith("=")
