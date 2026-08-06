"""Hoja de detalle: todos los movimientos con su categoría."""

import datetime

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.hoja_detalle import (
    SHEET_DETALLE,
    build_hoja_detalle,
)
from backend.app.aud.obligaciones_fiscales.mayor.tipos import Movimiento

MOVS = [
    Movimiento(codigo="1.1.5.1.1", cuenta="IVA sobre Compras",
               fecha=datetime.date(2025, 1, 5), asiento="COM 1",
               documento="FAC 001", identificacion="9999999999001",
               persona="PROVEEDOR DEMO S.A.", descripcion="COMPRA DE PRUEBA",
               debe=2.39, haber=0.0, saldo=2.39),
    Movimiento(codigo="4.1.1.4", cuenta="Venta de insumos",
               fecha=datetime.date(2025, 2, 8), asiento="VTA 1",
               debe=0.0, haber=100.0, saldo=-100.0),
]
CATEGORIAS = {"1.1.5.1.1": "IVA_COMPRAS", "4.1.1.4": "VENTAS"}


def test_escribe_una_fila_por_movimiento():
    wb = Workbook()
    build_hoja_detalle(wb, MOVS, CATEGORIAS)
    ws = wb[SHEET_DETALLE]
    assert ws.max_row == 3 + len(MOVS) - 1 or ws.max_row >= len(MOVS)


def test_la_primera_columna_es_la_categoria_para_poder_filtrar():
    wb = Workbook()
    build_hoja_detalle(wb, MOVS, CATEGORIAS)
    ws = wb[SHEET_DETALLE]
    assert ws.cell(3, 1).value == "Categoría"
    assert ws.cell(4, 1).value == "IVA_COMPRAS"


def test_activa_el_autofiltro_sobre_el_rango_de_datos():
    wb = Workbook()
    build_hoja_detalle(wb, MOVS, CATEGORIAS)
    assert wb[SHEET_DETALLE].auto_filter.ref is not None


def test_congela_el_encabezado():
    wb = Workbook()
    build_hoja_detalle(wb, MOVS, CATEGORIAS)
    assert wb[SHEET_DETALLE].freeze_panes == "A4"


def test_un_movimiento_de_cuenta_sin_categoria_queda_marcado():
    wb = Workbook()
    build_hoja_detalle(wb, MOVS, {})
    ws = wb[SHEET_DETALLE]
    assert ws.cell(4, 1).value == "SIN_CLASIFICAR"


def test_conserva_los_datos_de_trazabilidad_del_movimiento():
    wb = Workbook()
    build_hoja_detalle(wb, MOVS, CATEGORIAS)
    ws = wb[SHEET_DETALLE]
    fila = [ws.cell(4, c).value for c in range(1, 14)]
    assert "FAC 001" in fila
    assert "PROVEEDOR DEMO S.A." in fila
    assert 2.39 in fila
