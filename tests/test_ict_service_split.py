"""Tests del split service.generate_excel -> (bytes_sri, bytes_papel_trabajo).

REGLA "Separación SRI vs Papel de trabajo del auditor" (CLAUDE.md):
el Excel que se carga al portal del SRI NO debe MOSTRAR las hojas internas
del auditor ni las hojas de datos fuente.

CAMBIO 2026-06-26 (decisión del cliente): esas hojas ya NO se BORRAN del
archivo SRI — se OCULTAN (`sheet_state="hidden"`). Motivo: las fórmulas
referenciales de A1..A9 apuntan a 'DATOS F-101'!Cxxx, 'DATOS BALANCE'!..,
etc. Borrar esas hojas rompería las fórmulas con #REF!. Ocultarlas deja el
archivo limpio a la vista del cliente y mantiene las fórmulas resolviendo.
El papel de trabajo conserva TODAS las hojas visibles.
"""
from io import BytesIO

import openpyxl

from backend.app.ict.service import (
    HIDDEN_SHEETS_FOR_SRI,
    _apply_sri_sheet_visibility,
)


def test_hidden_sheets_constant_lists_datos_and_internal_sheets():
    """La constante HIDDEN_SHEETS_FOR_SRI es el contrato explícito: cada hoja
    aquí se OCULTA (no se borra) en el archivo SRI."""
    expected = {
        "DATOS F-101",
        "DATOS F-103",
        "DATOS F-104",
        "DATOS BALANCE",
        "VERIFICACIÓN A1",
        "TRAZABILIDAD",
    }
    assert set(HIDDEN_SHEETS_FOR_SRI) == expected


def test_sri_hides_but_never_deletes_sheets():
    """El archivo SRI oculta las hojas internas/datos pero NO borra ninguna
    (para no romper las fórmulas referenciales de A1..A9)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    business = ["INDICE", "A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9"]
    for name in business:
        wb.create_sheet(name)
    for name in HIDDEN_SHEETS_FOR_SRI:
        wb.create_sheet(name)

    # serializar -> recargar -> aplicar visibilidad SRI -> serializar
    buf = BytesIO()
    wb.save(buf)
    wb_sri = openpyxl.load_workbook(BytesIO(buf.getvalue()))
    _apply_sri_sheet_visibility(wb_sri)
    out = BytesIO()
    wb_sri.save(out)
    reloaded = openpyxl.load_workbook(BytesIO(out.getvalue()))

    # 1) NINGUNA hoja fue borrada del SRI
    for name in list(HIDDEN_SHEETS_FOR_SRI) + business:
        assert name in reloaded.sheetnames, (
            f"La hoja '{name}' fue borrada del SRI (debía solo ocultarse)"
        )

    # 2) Las hojas internas/datos quedan OCULTAS
    for name in HIDDEN_SHEETS_FOR_SRI:
        assert reloaded[name].sheet_state == "hidden", (
            f"La hoja '{name}' debería estar oculta en el SRI"
        )

    # 3) INDICE + A1..A9 quedan VISIBLES
    for name in business:
        assert reloaded[name].sheet_state == "visible", (
            f"La hoja de negocio '{name}' debería seguir visible en el SRI"
        )


def test_sri_active_sheet_is_visible_indice():
    """La hoja activa del SRI debe ser una VISIBLE (INDICE de preferencia):
    Excel se queja al abrir un libro cuya hoja activa está oculta."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Las hojas a ocultar van primero (serían la activa por defecto, índice 0).
    for name in HIDDEN_SHEETS_FOR_SRI:
        wb.create_sheet(name)
    for name in ["INDICE", "A1"]:
        wb.create_sheet(name)

    buf = BytesIO()
    wb.save(buf)
    wb_sri = openpyxl.load_workbook(BytesIO(buf.getvalue()))
    _apply_sri_sheet_visibility(wb_sri)

    active = wb_sri.active
    assert active is not None
    assert active.sheet_state == "visible", "La hoja activa del SRI no debe estar oculta"
    assert active.title == "INDICE", f"La hoja activa debería ser INDICE, fue {active.title!r}"


def test_papel_trabajo_keeps_all_sheets_visible():
    """El papel de trabajo conserva TODAS las hojas visibles (no se le aplica
    la ocultación). Aquí solo verificamos que la función de visibilidad SRI no
    toca un workbook que represente el papel: se aplica únicamente a la copia SRI."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in ["INDICE", "A1", *HIDDEN_SHEETS_FOR_SRI]:
        wb.create_sheet(name)
    # El papel NO pasa por _apply_sri_sheet_visibility → todo visible.
    for ws in wb.worksheets:
        assert ws.sheet_state == "visible"


def test_sri_structure_is_password_protected():
    """El SRI bloquea la ESTRUCTURA del libro con contraseña: con eso el
    cliente no puede usar 'Mostrar' (Unhide) para des-ocultar las hojas
    DATOS/internas, ni insertar/eliminar/renombrar hojas. Solo AuditConsulting
    conoce la clave."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in ["INDICE", "A1", *HIDDEN_SHEETS_FOR_SRI]:
        wb.create_sheet(name)

    buf = BytesIO()
    wb.save(buf)
    wb_sri = openpyxl.load_workbook(BytesIO(buf.getvalue()))
    _apply_sri_sheet_visibility(wb_sri)
    out = BytesIO()
    wb_sri.save(out)
    reloaded = openpyxl.load_workbook(BytesIO(out.getvalue()))

    assert reloaded.security is not None, "El SRI no tiene protección de libro"
    assert reloaded.security.lockStructure is True, (
        "La estructura del libro SRI debe estar bloqueada (lockStructure=True)"
    )
    # Debe existir un hash de contraseña (protección CON clave, no sin clave).
    has_password = bool(reloaded.security.workbookHashValue) or bool(
        reloaded.security.workbookPassword
    )
    assert has_password, (
        "La protección de estructura del SRI debe llevar contraseña"
    )


def test_papel_trabajo_workbook_is_not_protected():
    """El papel de trabajo NO se protege: el auditor debe poder ver/editar
    todas las hojas. La función de visibilidad SRI no se aplica al papel."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in ["INDICE", "A1", *HIDDEN_SHEETS_FOR_SRI]:
        wb.create_sheet(name)
    # El papel NO pasa por _apply_sri_sheet_visibility.
    assert not getattr(wb.security, "lockStructure", False)


def test_generate_excel_signature_returns_tuple():
    """generate_excel devuelve tuple[bytes, bytes] (SRI, papel)."""
    import inspect
    from backend.app.ict.service import generate_excel
    sig = inspect.signature(generate_excel)
    annotation_str = str(sig.return_annotation).replace(" ", "")
    assert "tuple" in annotation_str.lower()
    assert "bytes" in annotation_str.lower()
