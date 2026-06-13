"""Regression test: cas 801 UTILIDAD DEL EJERCICIO debe ser FORMULA
=C{row_6999}-C{row_7999} en col C del A1 (cliente ICT_19, 2026-06-13).

Antes: cas 801 col C era ='DATOS F-101'!Cxxx (traslado directo del valor
declarado por el contribuyente).

Despues: cas 801 col C es =C{row_6999}-C{row_7999} (calculo automatico
basado en los totales de Ingresos y Costos+Gastos del propio A1). Si el
F-101 declarado no cuadra con esta operacion, el auditor ve la diferencia
en col G — no se enmascara.

Razon: el cliente reporta que la utilidad declarada en F-101 puede tener
ajustes que el calculo simple no captura. El papel de trabajo debe
mostrar el CALCULO REAL, no el declarado, para que el auditor identifique
discrepancias.
"""
from __future__ import annotations

import openpyxl
import pytest


@pytest.fixture
def wb_con_a1_completo():
    """Genera un workbook minimo con A1 completo (cas 6999, 7999, 801)."""
    from openpyxl import Workbook
    from backend.app.ict.cell_maps.a1 import A1_SHEET
    from backend.app.ict.fillers.a1_mapeo import A1Filler
    from backend.app.ict.fillers.source_data_sheets import build_f101_sheet

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    # Crear hoja A1 (el filler la encuentra por A1_SHEET)
    wb.create_sheet(A1_SHEET)
    # Crear hoja DATOS F-101 (el filler usa f101_lookup)
    f101_data = {
        # Ingresos: cas 6001 (operacionales), 6999 (TOTAL)
        "6001": 5000.0,
        # Costos: cas 7001, 7999 (TOTAL)
        "7001": 2000.0,
        # Utilidad: cas 801 declarado distinto del calculo simple
        "801": 4000.0,  # F-101 declarado (3000 segun el calculo)
        # Otros cas de la cadena
        "803": 300.0,
        "888": 700.0,
        "889": 100.0,
    }
    f101_lookup = build_f101_sheet(wb, f101_data, {})

    # Llamar A1Filler
    filler = A1Filler()
    session_data = {"razon_social": "TEST", "ruc": "T",
                    "ejercicio_fiscal": "2025", "numero_adhesivo": ""}
    anexo_data = {
        "f101": f101_data,
        "balance_mapeado": [],
        "_f101_lookup": f101_lookup,
        "_balance_lookup": [],
    }
    filler.fill(wb, session_data, anexo_data)
    return wb


def _find_cas_row(ws, cas_buscado: str) -> int | None:
    for r in range(13, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).strip() == cas_buscado:
            return r
    return None


def test_cas_801_col_C_es_formula_ingresos_menos_costos(wb_con_a1_completo):
    """Verifica que cas 801 en A1 col C sea una formula =C{6999}-C{7999}
    en lugar de =DATOS F-101!Cxxx."""
    from backend.app.ict.cell_maps.a1 import A1_SHEET
    ws = wb_con_a1_completo[A1_SHEET]

    row_801 = _find_cas_row(ws, "801")
    assert row_801 is not None, "cas 801 debe aparecer en A1 (no excluido)"

    formula_c = str(ws.cell(row_801, 3).value or "")

    assert formula_c.startswith("="), (
        f"cas 801 col C debe ser formula, encontrado: {formula_c!r}"
    )

    # NO debe ser un traslado directo del F-101
    assert "DATOS F-101" not in formula_c, (
        f"cas 801 col C NO debe trasladar directo del F-101. Formula: {formula_c}"
    )

    # DEBE incluir una resta entre 2 celdas col C (ingresos - costos)
    import re
    refs = re.findall(r"C(\d+)", formula_c)
    assert len(refs) >= 2, (
        f"cas 801 col C debe tener 2 referencias (ing - costos), encontrado: {formula_c}"
    )
    assert "-" in formula_c, (
        f"cas 801 col C debe tener resta, encontrado: {formula_c}"
    )


def test_cas_801_apunta_a_6999_y_7999(wb_con_a1_completo):
    """La formula de cas 801 debe apuntar EXACTAMENTE a las filas donde
    estan cas 6999 (TOTAL INGRESOS) y cas 7999 (TOTAL COSTOS Y GASTOS)."""
    from backend.app.ict.cell_maps.a1 import A1_SHEET
    ws = wb_con_a1_completo[A1_SHEET]

    row_801 = _find_cas_row(ws, "801")
    row_6999 = _find_cas_row(ws, "6999")
    row_7999 = _find_cas_row(ws, "7999")
    assert row_801 and row_6999 and row_7999, (
        f"Faltan cas: 801={row_801}, 6999={row_6999}, 7999={row_7999}"
    )

    formula_c = str(ws.cell(row_801, 3).value or "")
    # Verificar que las refs son a las filas correctas
    assert f"C{row_6999}" in formula_c, (
        f"cas 801 debe referenciar C{row_6999} (TOTAL INGRESOS). Formula: {formula_c}"
    )
    assert f"C{row_7999}" in formula_c, (
        f"cas 801 debe referenciar C{row_7999} (TOTAL COSTOS Y GASTOS). Formula: {formula_c}"
    )
