"""Regression test: VERIFICACION A1 no debe tener celdas de texto que
empiecen con "=" (openpyxl las interpreta como formulas Excel).

REPORTE CLIENTE (2026-06-13, ICT_19): al abrir el Excel descargado de
produccion, Excel muestra dialogo "Reparaciones en ICT_19_PAPEL_TRABAJO.xlsx
- Registros quitados: Formula de /xl/worksheets/sheet15.xml parte".

Causa raiz: labels como "= UTILIDAD OPERACIONAL (Ing - Costos)" y
"= UTILIDAD INTEGRAL CALCULADA" estaban en el codigo como strings. openpyxl
detecta el "=" inicial y los marca como type='formula' en el XLSX, pero
"UTILIDAD OPERACIONAL" no es sintaxis Excel valida -> Excel los elimina.

Fix: cambiar el prefijo "=" por "(=)" en labels de filas TOTAL.
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from backend.app.ict.cell_maps.a1 import A1_SHEET
from backend.app.ict.fillers.verification import build_verification_sheet


def test_no_celdas_texto_empiezan_con_igual():
    """Recorre todas las celdas de VERIFICACION A1 y verifica que ninguna
    fila tenga un VALOR DE TEXTO (no formula real) que empiece con "=".

    Si fallara: el XLSX al guardarse marcaria esa celda como type='f'
    (formula) pero al ser texto invalido, Excel la quitaria al abrir.
    """
    wb = Workbook()
    wb.create_sheet(A1_SHEET)
    build_verification_sheet(
        wb,
        f101={"499": 1000, "699": 1000},
        balance_mapeado=[],
        session_data={
            "razon_social": "TEST", "ruc": "T", "ejercicio_fiscal": "2025"
        },
    )
    ws = wb["VERIFICACIÓN A1"]

    invalid_formulas = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            # Si empieza con "=" debe ser una FORMULA Excel valida.
            # Heuristica: si despues del "=" hay solo texto sin parentesis,
            # operadores ni funciones, NO es formula.
            body = v[1:].strip()
            if not body:
                continue
            # Una fórmula válida tiene al menos uno de: ( + - * / IFERROR ABS SUM
            # IF MATCH INDEX OFFSET COUNTIF SUMIF SUMPRODUCT
            tokens_validos = ['(', '+', '-', '*', '/', 'IFERROR', 'ABS', 'SUM',
                              'IF', 'MATCH', 'INDEX', 'OFFSET', 'COUNTIF',
                              'SUMIF', 'SUMPRODUCT', 'INDIRECT', 'VLOOKUP',
                              'HLOOKUP', '!']
            tiene_token_valido = any(t in body.upper() for t in tokens_validos)
            # Tambien aceptamos refs simples a celdas (E5, F28, etc.)
            import re
            es_ref_simple = bool(re.match(r'^[A-Z]+\d+', body))
            if not tiene_token_valido and not es_ref_simple:
                invalid_formulas.append((cell.coordinate, v))

    assert not invalid_formulas, (
        f"VERIFICACION A1 tiene {len(invalid_formulas)} celdas con texto que "
        f"empieza con '=' pero no es formula Excel valida -> Excel las quita "
        f"al abrir y muestra dialogo de reparacion:\n"
        + "\n".join(f"  {c}: {v[:60]}" for c, v in invalid_formulas[:10])
    )
