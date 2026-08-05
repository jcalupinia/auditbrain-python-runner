"""DM4 Compras: IVA en compras y base imponible, libros vs declaraciones."""

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm4_compras import (
    CASILLEROS_BASE, CASILLEROS_IVA, SHEET_DM4, build_dm4,
)

PERIODOS = [f"2025-{m:02d}" for m in range(1, 13)]

DIR_MAYORES = {
    ("cuenta:1.1.5.1.1", "01"): "'Mayores homologados'!D4",
    ("cuenta:1.1.5.1.3", "01"): "'Mayores homologados'!D5",
    ("orden:IVA_COMPRAS", "cuentas"): ["1.1.5.1.1", "1.1.5.1.3"],
    ("IVA_COMPRAS", "01"): "'Mayores homologados'!D6",
}
DIR_F104 = {("2025-01", cas): f"'DATOS F-104'!C{i}" for i, cas in enumerate(
    CASILLEROS_IVA + CASILLEROS_BASE, start=20)}

NOMBRES = {"1.1.5.1.1": "IVA sobre Compras", "1.1.5.1.3": "IVA en Importaciones"}


def _cedula(**kw):
    wb = Workbook()
    datos = dict(dir_mayores=DIR_MAYORES, dir_f104=DIR_F104, periodos=PERIODOS,
                 nombres_cuenta=NOMBRES, cliente="C", periodo="2025", tarifas={})
    datos.update(kw)
    build_dm4(wb, **datos)
    return wb[SHEET_DM4]


def test_lista_una_fila_por_cuenta_de_iva_en_compras():
    ws = _cedula()
    etiquetas = [ws.cell(r, 2).value for r in range(1, ws.max_row + 1)]
    assert "IVA sobre Compras" in etiquetas
    assert "IVA en Importaciones" in etiquetas


def test_el_valor_de_una_cuenta_es_una_formula_al_resumen():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(r, 2).value == "IVA sobre Compras")
    assert ws.cell(fila, 3).value == "='Mayores homologados'!D4"


def test_los_casilleros_son_formulas_a_la_hoja_de_datos_no_valores():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if str(ws.cell(r, 2).value or "").startswith("Casillero 520"))
    valor = ws.cell(fila, 3).value
    assert isinstance(valor, str) and valor.startswith("='DATOS F-104'!")


def test_estan_los_nueve_casilleros_de_iva_en_compras():
    ws = _cedula()
    etiquetas = [str(ws.cell(r, 2).value or "") for r in range(1, ws.max_row + 1)]
    for cas in CASILLEROS_IVA:
        assert any(e.startswith(f"Casillero {cas}") for e in etiquetas), cas


def test_la_diferencia_resta_libros_menos_declarado_y_redondea():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(r, 2).value == "Diferencia")
    assert ws.cell(fila, 3).value.startswith("=ROUND(")


def test_la_base_imponible_divide_el_iva_para_la_tarifa_del_mes():
    ws = _cedula(tarifas={"01": 0.15})
    fila = next(r for r in range(1, ws.max_row + 1)
                if str(ws.cell(r, 2).value or "").startswith("Compras gravadas"))
    assert "/0.15" in ws.cell(fila, 3).value


def test_la_tarifa_por_defecto_es_quince_por_ciento():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if str(ws.cell(r, 2).value or "").startswith("Compras gravadas"))
    assert "/0.15" in ws.cell(fila, 3).value


def test_una_tarifa_distinta_en_un_mes_se_respeta():
    """El ejercicio puede cruzar el cambio de 12% a 15%."""
    ws = _cedula(tarifas={"01": 0.12})
    fila = next(r for r in range(1, ws.max_row + 1)
                if str(ws.cell(r, 2).value or "").startswith("Compras gravadas"))
    assert "/0.12" in ws.cell(fila, 3).value


def test_lleva_el_encabezado_de_cedula_con_su_referencia():
    ws = _cedula()
    valores = [ws.cell(r, c).value for r in range(1, 11) for c in range(1, 6)]
    assert "OBLIGACIONES FISCALES" in valores
    assert "DM4" in valores
