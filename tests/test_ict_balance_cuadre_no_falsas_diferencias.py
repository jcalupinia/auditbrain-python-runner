"""Regression test: sección CUADRE de DATOS BALANCE no debe dar falsas diferencias.

REPORTE CLIENTE (2026-06-13, ICT_17): la tabla "CUADRE POR CASILLERO ·
Balance ↔ MAPEO A1" estaba marcando "✗ DIFF" en casi todos los cas con
múltiples cuentas (cas 311 con 5 cuentas, cas 325 con 9, etc.) — pero
en realidad SÍ estaban trasladados.

Causa raíz: VLOOKUP solo devuelve la PRIMERA ocurrencia. Como en A1 los
cas con N cuentas ocupan N filas pero solo la primera tiene el número
de cas en col A, VLOOKUP traía el saldo de UNA cuenta mientras SUMIF
del balance sumaba TODAS → diferencia falsa.

Efecto del bug: las falsas diferencias enmascaran las verdaderas. Si
un cas REAL no se trasladó (caso cas 618 en ICT_17 — bug catálogo), el
auditor no podía distinguirlo del ruido.

Solución: la sección CUADRE debe:
  1. NO usar VLOOKUP (solo trae primera fila).
  2. Detectar PRESENCIA en A1 con COUNTIF de col A.
  3. Sumar el rango completo del cas en A1 con OFFSET + MATCH + COUNTIF.
  4. Distinguir estados: OK / NO TRASLADADO / INCOMPLETO / DIFF / EXCLUIDO.

Tests:
  1. Las fórmulas generadas NO contienen VLOOKUP simple.
  2. Las fórmulas usan MATCH + OFFSET para sumar el rango del cas.
  3. El estado distingue "NO TRASLADADO" cuando el cas está ausente del A1.
"""
from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pytest


@pytest.fixture(scope="module")
def regenerated_ict15():
    """Genera ICT_15 con datos PROPHAR reales y lo carga read-only."""
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from openpyxl import load_workbook
    from scripts.generate_ict15_prophar import (
        load_prophar_data, make_mock_session,
    )
    from backend.app.ict.service import generate_excel

    data, _log = load_prophar_data()
    mock_session = make_mock_session(data)

    # Generamos en memoria via service.generate_excel
    sri_bytes, papel_bytes = generate_excel(db=None, session=mock_session)

    # Cargamos para inspeccionar fórmulas
    from io import BytesIO
    wb = load_workbook(BytesIO(papel_bytes), data_only=False)
    return wb


def _get_balance_sheet(wb):
    for n in wb.sheetnames:
        if "BALANCE" in n.upper() and "DATOS" in n.upper():
            return wb[n]
    raise AssertionError("DATOS BALANCE sheet not found")


def _find_cuadre_start(ws) -> int:
    """Busca la fila donde empieza la sección CUADRE."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and "CUADRE" in str(v).upper():
            return r
    raise AssertionError("Sección CUADRE no encontrada en DATOS BALANCE")


class TestCuadreNoFalsasDiferencias:
    """La sección CUADRE debe usar la fórmula corregida (sin VLOOKUP simple)."""

    def test_no_vlookup_simple_en_cuadre(self, regenerated_ict15):
        """Las fórmulas del CUADRE NO deben usar VLOOKUP simple (solo
        trae primera fila). Si lo usan, es el bug viejo."""
        ws = _get_balance_sheet(regenerated_ict15)
        cuadre = _find_cuadre_start(ws)

        # Buscar primera fila de datos del cuadre (saltando header)
        first_data_row = cuadre + 2  # título + headers

        # Inspeccionar fórmulas de las primeras 10 filas de datos
        vlookup_simple_found = []
        for r in range(first_data_row, min(first_data_row + 15, ws.max_row + 1)):
            cas = ws.cell(r, 1).value
            if not cas or not str(cas).strip().isdigit():
                continue
            formula_a1 = str(ws.cell(r, 4).value or "")
            # El bug viejo era VLOOKUP("cas", ..., 6, FALSE) — solo primera
            # fila. La solución correcta NO debe tener ese patrón.
            if "VLOOKUP(" in formula_a1.upper() and "FALSE" in formula_a1.upper():
                vlookup_simple_found.append((cas, formula_a1))

        assert not vlookup_simple_found, (
            f"Fórmulas con VLOOKUP simple detectadas — bug regresa: "
            f"{vlookup_simple_found[:3]}"
        )

    def test_cuadre_usa_match_y_count(self, regenerated_ict15):
        """La fórmula corregida debe usar MATCH (encuentra el cas en A1) +
        COUNTIF (cuenta cuentas balance) para sumar el rango completo."""
        ws = _get_balance_sheet(regenerated_ict15)
        cuadre = _find_cuadre_start(ws)
        first_data_row = cuadre + 2

        # Buscar al menos una fila con la fórmula correcta
        match_count_found = False
        for r in range(first_data_row, min(first_data_row + 30, ws.max_row + 1)):
            cas = ws.cell(r, 1).value
            if not cas or not str(cas).strip().isdigit():
                continue
            # Inspeccionar col E (suma A1) o col D (count A1)
            for col in (4, 5):
                f = str(ws.cell(r, col).value or "").upper()
                if "MATCH(" in f or "COUNTIF(" in f:
                    match_count_found = True
                    break
            if match_count_found:
                break

        assert match_count_found, (
            "Ninguna fórmula del CUADRE usa MATCH/COUNTIF. La solución "
            "correcta debe usarlos para sumar el rango completo del cas."
        )

    def test_estado_distingue_no_trasladado(self, regenerated_ict15):
        """La columna Estado debe poder mostrar 'NO TRASLADADO' cuando
        el cas tiene saldo en balance pero está ausente del A1.

        Verificación: la fórmula del estado debe mencionar 'TRASLADADO' o
        'AUSENTE' en alguna rama IF.
        """
        ws = _get_balance_sheet(regenerated_ict15)
        cuadre = _find_cuadre_start(ws)
        first_data_row = cuadre + 2

        estado_has_trasladado = False
        for r in range(first_data_row, min(first_data_row + 30, ws.max_row + 1)):
            cas = ws.cell(r, 1).value
            if not cas or not str(cas).strip().isdigit():
                continue
            f_estado = str(ws.cell(r, 6).value or "").upper()
            if "TRASLAD" in f_estado or "AUSENTE" in f_estado:
                estado_has_trasladado = True
                break

        assert estado_has_trasladado, (
            "La columna Estado debe mostrar 'NO TRASLADADO' cuando un cas "
            "está ausente del A1 — para detectar bugs como cas 618/6094."
        )

    def test_estado_no_solo_ok_o_diff(self, regenerated_ict15):
        """La columna Estado debe distinguir más de 2 estados: el bug
        viejo solo daba OK/DIFF. La solución debe distinguir al menos:
        OK, NO TRASLADADO, y opcionalmente DIFF / INCOMPLETO."""
        ws = _get_balance_sheet(regenerated_ict15)
        cuadre = _find_cuadre_start(ws)
        first_data_row = cuadre + 2

        estados_distintos = set()
        for r in range(first_data_row, min(first_data_row + 30, ws.max_row + 1)):
            cas = ws.cell(r, 1).value
            if not cas or not str(cas).strip().isdigit():
                continue
            f_estado = str(ws.cell(r, 6).value or "")
            # Contar palabras clave únicas en la fórmula
            for token in ["NO TRASLADADO", "INCOMPLETO", "OK", "DIFF", "EXCLUIDO"]:
                if token in f_estado.upper():
                    estados_distintos.add(token)

        assert len(estados_distintos) >= 3, (
            f"La fórmula de Estado debe distinguir ≥3 estados. "
            f"Encontrados: {estados_distintos}. Bug viejo: solo OK + DIFF."
        )
