"""DM7 Retenciones por pagar: retenciones de IVA (F-104) y de renta (F-103)."""

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm7_retenciones import (
    CASILLEROS_RET_IVA,
    CASILLERO_RET_IVA_CONTROL,
    CASILLERO_RET_RENTA,
    SHEET_DM7,
    build_dm7,
)

PERIODOS = [f"2025-{m:02d}" for m in range(1, 13)]

DIR_MAYORES = {
    ("cuenta:2.1.7.3.1", "01"): "'Mayores homologados'!D10",
    ("cuenta:2.1.7.3.2", "01"): "'Mayores homologados'!D11",
    ("orden:RET_IVA", "cuentas"): ["2.1.7.3.1", "2.1.7.3.2"],
    ("cuenta:2.1.7.2.1", "01"): "'Mayores homologados'!D14",
    ("orden:RET_RENTA", "cuentas"): ["2.1.7.2.1"],
}

_CAS_F104 = CASILLEROS_RET_IVA + [CASILLERO_RET_IVA_CONTROL]
DIR_F104 = {("2025-01", cas): f"'DATOS F-104'!C{i}" for i, cas in enumerate(_CAS_F104, start=20)}
DIR_F103 = {("2025-01", CASILLERO_RET_RENTA): "'DATOS F-103'!C40"}

NOMBRES = {
    "2.1.7.3.1": "Ret. 30% Bienes",
    "2.1.7.3.2": "Ret. 70% Servicios",
    "2.1.7.2.1": "Ret. 1% Bienes",
}


def _cedula(**kw):
    wb = Workbook()
    datos = dict(dir_mayores=DIR_MAYORES, dir_f104=DIR_F104, dir_f103=DIR_F103,
                 periodos=PERIODOS, nombres_cuenta=NOMBRES, cliente="C", periodo="2025")
    datos.update(kw)
    build_dm7(wb, **datos)
    return wb[SHEET_DM7]


def _etiquetas(ws):
    return [ws.cell(r, 2).value for r in range(1, ws.max_row + 1)]


def test_lista_una_fila_por_cuenta_de_retenciones_de_iva():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    assert "Ret. 30% Bienes" in etiquetas
    assert "Ret. 70% Servicios" in etiquetas


def test_el_valor_de_una_cuenta_es_una_formula_al_resumen():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(r, 2).value == "Ret. 30% Bienes")
    assert ws.cell(fila, 3).value == "='Mayores homologados'!D10"


def test_los_casilleros_de_iva_son_formulas_a_la_hoja_de_datos_no_valores():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if str(ws.cell(r, 2).value or "").startswith("Casillero 721"))
    valor = ws.cell(fila, 3).value
    assert isinstance(valor, str) and valor.startswith("='DATOS F-104'!")


def test_estan_los_seis_casilleros_de_retenciones_de_iva():
    ws = _cedula()
    etiquetas = [str(ws.cell(r, 2).value or "") for r in range(1, ws.max_row + 1)]
    for cas in CASILLEROS_RET_IVA:
        assert any(e.startswith(f"Casillero {cas}") for e in etiquetas), cas


def test_el_casillero_799_aparece_como_control_no_sumado():
    """799 es control del total: aparece como referencia aparte, no se
    mete al rango que suma 'Según declaraciones'."""
    ws = _cedula()
    etiquetas = [str(ws.cell(r, 2).value or "") for r in range(1, ws.max_row + 1)]
    fila_799 = next(r for r in range(1, ws.max_row + 1)
                     if str(ws.cell(r, 2).value or "").startswith("Casillero 799"))
    assert "control" in str(ws.cell(fila_799, 2).value).lower()
    valor = ws.cell(fila_799, 3).value
    assert isinstance(valor, str) and valor.startswith("='DATOS F-104'!")


def test_la_diferencia_de_iva_resta_libros_menos_declarado_y_redondea():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    filas_diff = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "Diferencia"]
    assert len(filas_diff) >= 1
    assert ws.cell(filas_diff[0], 3).value.startswith("=ROUND(")


def test_lista_una_fila_por_cuenta_de_retenciones_de_renta():
    ws = _cedula()
    assert "Ret. 1% Bienes" in _etiquetas(ws)


def test_el_casillero_499_es_formula_a_datos_f103():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if str(ws.cell(r, 2).value or "").startswith("Casillero 499"))
    valor = ws.cell(fila, 3).value
    assert valor == "='DATOS F-103'!C40"


def test_hay_dos_filas_de_diferencia_una_por_bloque():
    ws = _cedula()
    assert _etiquetas(ws).count("Diferencia") == 2


def test_la_diferencia_resta_libros_menos_declarado_y_redondea_en_ambos_bloques():
    ws = _cedula()
    filas = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "Diferencia"]
    for fila in filas:
        assert ws.cell(fila, 3).value.startswith("=ROUND(")


def test_lleva_el_encabezado_de_cedula_con_su_referencia():
    ws = _cedula()
    valores = [ws.cell(r, c).value for r in range(1, 11) for c in range(1, 6)]
    assert "OBLIGACIONES FISCALES" in valores
    assert "DM7" in valores


def test_publica_las_direcciones_que_dm3_consume():
    wb = Workbook()
    lookup = build_dm7(wb, dir_mayores=DIR_MAYORES, dir_f104=DIR_F104, dir_f103=DIR_F103,
                        periodos=PERIODOS, nombres_cuenta=NOMBRES, cliente="C", periodo="2025")
    for clave in ("ret_iva_declarado", "ret_renta_declarado"):
        assert (clave, "01") in lookup, clave
        assert lookup[(clave, "01")].startswith(f"'{SHEET_DM7}'!")


def test_la_direccion_publicada_de_ret_iva_declarado_apunta_a_la_fila_segun_declaraciones():
    wb = Workbook()
    lookup = build_dm7(wb, dir_mayores=DIR_MAYORES, dir_f104=DIR_F104, dir_f103=DIR_F103,
                        periodos=PERIODOS, nombres_cuenta=NOMBRES, cliente="C", periodo="2025")
    ws = wb[SHEET_DM7]
    addr = lookup[("ret_iva_declarado", "01")]
    fila = int(addr.split("!")[1].lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
    assert ws.cell(fila, 2).value == "Según declaraciones"


def test_el_bloque_de_iva_crece_con_mas_cuentas_de_retencion_y_el_subtotal_se_desplaza():
    """El numero de cuentas de retencion varia por cliente."""
    dir_mayores_grande = {
        **{(f"cuenta:2.1.7.3.{i}", "01"): f"'Mayores homologados'!D{i}" for i in range(1, 6)},
        ("orden:RET_IVA", "cuentas"): [f"2.1.7.3.{i}" for i in range(1, 6)],
        ("orden:RET_RENTA", "cuentas"): [],
    }
    nombres_grande = {f"2.1.7.3.{i}": f"Ret {i}" for i in range(1, 6)}
    wb = Workbook()
    build_dm7(wb, dir_mayores=dir_mayores_grande, dir_f104=DIR_F104, dir_f103=DIR_F103,
              periodos=PERIODOS, nombres_cuenta=nombres_grande, cliente="C", periodo="2025")
    ws = wb[SHEET_DM7]
    primera_fila_libros = next(r for r in range(1, ws.max_row + 1)
                                if ws.cell(r, 2).value == "Según libros")
    formula = ws.cell(primera_fila_libros, 3).value
    assert formula == f"=SUM(C14:C{primera_fila_libros - 1})"
