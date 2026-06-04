"""Tests for the refactored verification.py that uses audit + kpi_components."""
from datetime import datetime
from decimal import Decimal

import openpyxl
import pytest

from backend.app.ict.audit.schemas import (
    A1Metrics, AnexoFinding, AnexoInterpretation, Status,
)


@pytest.fixture
def fake_a1_metrics():
    return A1Metrics(
        activo_total=Decimal("21671880.68"),
        pasivo_patrimonio_total=Decimal("21671880.68"),
        diferencia=Decimal("0.00"),
        status_cuadre=Status.OK,
        cobertura_mapeo_pct=87.0,
        cas_mapeados=47,
        cas_total=54,
        cas_sin_contrapartida=["302", "305"],
    )


@pytest.fixture
def fake_a1_interpretation():
    return AnexoInterpretation(
        anexo_codigo="A1",
        anexo_nombre="Mapeo del balance",
        resumen_ejecutivo="Balance cuadra. 2 cas sin contrapartida.",
        findings=[
            AnexoFinding(
                severity="leve",
                categoria="conciliacion_inconsistente",
                titulo="2 cas sin contrapartida contable",
                descripcion_tecnica="cas 302, 305 declarados pero sin balance",
                implicacion_tributaria="Verificar art. 19 LORTI",
                recomendacion="Revisar mapeo manual",
                casilleros_afectados=["302", "305"],
            ),
        ],
        confianza_modelo="alta",
        requiere_revision_humana=False,
        timestamp_analisis=datetime(2026, 6, 4),
        modelo_usado="claude-sonnet-4-7-20260101",
        tokens_consumidos=1500,
    )


def test_fill_verification_writes_executive_banner_and_kpis(
    fake_a1_metrics, fake_a1_interpretation,
):
    """The new fill_verification_a1 entry point renders banner + KPI cards
    using kpi_components, then writes the interpretation IA section."""
    from backend.app.ict.fillers.verification import fill_verification_a1
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("VERIFICACIÓN A1")
    contexto = {"razon_social": "PROPHAR S.A.",
                "ruc": "1791859596001", "periodo": "2025"}
    fill_verification_a1(
        ws,
        metrics=fake_a1_metrics,
        interpretation=fake_a1_interpretation,
        contexto=contexto,
    )

    # Banner present
    found_banner = False
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for v in row:
            if v and "AUDITBRAIN" in str(v):
                found_banner = True
                break
    assert found_banner, "Banner ejecutivo no encontrado"

    # At least one cell shows the activo total
    found_activo = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v and "21,671,880" in str(v):
                found_activo = True
    assert found_activo, "Activo total no renderizado"

    # Sección INTERPRETACIÓN IA debe aparecer
    found_interpretation = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v and "INTERPRETACIÓN" in str(v).upper():
                found_interpretation = True
    assert found_interpretation, "Sección INTERPRETACIÓN IA no renderizada"

    # Disclaimer obligatorio (regla CLAUDE.md interpretación IA)
    found_disclaimer = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v and "validada por el auditor" in str(v).lower():
                found_disclaimer = True
    assert found_disclaimer, "Disclaimer IA obligatorio no presente"
