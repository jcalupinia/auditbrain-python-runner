"""Toda dirección que una cédula DM referencia debe llevar el nombre de hoja.

Regresión del bug de las 1.275 fórmulas rotas: `bloques.fila_referencias`
publica literalmente `f"={addr}"`, así que si `addr` viene sin prefijo de
hoja (`D30` en vez de `'Mayores homologados'!D30`) Excel la resuelve contra
la PROPIA hoja de la cédula. Eso produjo referencias circulares reales en
DM5 y DM7 y el libro se abría con ceros y aviso de circularidad.

El contrato queda fijado aquí: los mapas de direcciones que consumen las
cédulas (`build_hoja_mayores`, `construir_hojas_de_casilleros`) devuelven
SIEMPRE direcciones calificadas.
"""

import datetime
import re
from io import BytesIO

from openpyxl import Workbook, load_workbook

from backend.app.aud.obligaciones_fiscales.libro.ensamblador import armar_libro
from backend.app.aud.obligaciones_fiscales.libro.fuentes import (
    construir_hojas_de_casilleros,
)
from backend.app.aud.obligaciones_fiscales.libro.hoja_mayores import (
    SHEET_MAYORES,
    build_hoja_mayores,
)
from backend.app.aud.obligaciones_fiscales.mayor.tipos import Movimiento
from backend.app.ict.fillers.source_data_sheets import (
    SHEET_F103,
    SHEET_F104,
    build_f104_sheet,
)


class _Fila:
    """Doble de MayorClasificacionJob: solo lo que la hoja necesita."""

    def __init__(self, codigo, nombre, categoria, por_mes):
        self.codigo_cuenta = codigo
        self.nombre_cuenta = nombre
        self.categoria_final = categoria
        self.por_mes_json = por_mes
        self.n_movimientos = 1
        self.debe = 0.0
        self.haber = 0.0


FILAS = [
    _Fila("1.1.5.1.1", "IVA sobre Compras", "IVA_COMPRAS", {"01": 659.57}),
    _Fila("4.1.1.1", "Venta de mercadería", "VENTAS", {"01": 5000.0}),
    _Fila("2.1.7.4.1", "IVA sobre Ventas", "IVA_VENTAS", {"01": 750.0}),
    _Fila("2.1.7.3.1", "Ret. 30% Bienes", "RET_IVA", {"01": 12.0}),
    _Fila("2.1.7.2.1", "Ret. 1% Bienes", "RET_RENTA", {"01": 20.0}),
]


# ---------------------------------------------------------------- unidad ---

def test_todas_las_direcciones_del_mayor_llevan_el_nombre_de_su_hoja():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS)
    sin_calificar = {
        clave: addr
        for clave, addr in lookup.items()
        if not clave[0].startswith("orden:")
        and not str(addr).startswith(f"'{SHEET_MAYORES}'!")
    }
    assert sin_calificar == {}


def test_la_lista_de_cuentas_por_categoria_no_se_califica():
    """`orden:*` guarda una LISTA de códigos, no una dirección."""
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS)
    assert lookup[("orden:IVA_COMPRAS", "cuentas")] == ["1.1.5.1.1"]


def test_la_direccion_calificada_del_mayor_sigue_apuntando_a_su_celda():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS)
    addr = lookup[("cuenta:1.1.5.1.1", "01")]
    celda = addr.split("!", 1)[1]
    assert wb[SHEET_MAYORES][celda].value == 659.57


def test_todas_las_direcciones_de_los_casilleros_llevan_su_hoja():
    wb = Workbook()
    lookups = construir_hojas_de_casilleros(
        wb,
        f104_monthly={"2025-01": {"casilleros": {"429": 4341.16}}},
        f103_monthly={"2025-01": {"casilleros": {"499": 915.70}}},
    )
    assert lookups["f104"] and lookups["f103"]
    for addr in lookups["f104"].values():
        assert addr.startswith(f"'{SHEET_F104}'!"), addr
    for addr in lookups["f103"].values():
        assert addr.startswith(f"'{SHEET_F103}'!"), addr


def test_la_direccion_calificada_del_casillero_sigue_apuntando_a_su_celda():
    wb = Workbook()
    lookups = construir_hojas_de_casilleros(
        wb,
        f104_monthly={"2025-01": {"casilleros": {"429": 4341.16}}},
        f103_monthly={},
    )
    addr = lookups["f104"][("2025-01", "429")]
    assert wb[SHEET_F104][addr.split("!", 1)[1]].value == 4341.16


def test_el_builder_del_ict_sigue_devolviendo_direcciones_sin_prefijo():
    """El ICT califica del lado del consumidor: calificar en el builder
    compartido produciría doble prefijo y rompería A1..A9."""
    wb = Workbook()
    lookup = build_f104_sheet(wb, {"2025-01": {"casilleros": {"429": 4341.16}}})
    for addr in lookup.values():
        assert "!" not in addr, addr


# ------------------------------------------------------ regresión de libro ---

# Una "publicación de direcciones" es una fórmula compuesta EXCLUSIVAMENTE
# por referencias de celda unidas por '+': es la forma exacta que emiten
# `bloques.fila_referencias` (`=addr`), `bloques.fila_suma_direcciones`
# (`=addr1+addr2`) y `dm3._bloque_saldo` (`=addr615+addr617`). Toda
# referencia dentro de ellas apunta a OTRA hoja, así que todas deben ir
# calificadas. Mirar token por token —y no el `!` de la fórmula completa—
# atrapa además las MIXTAS como `=N139+'DM7 Retenciones x pagar'!N43`, donde
# el tramo bueno aportaba el `!` y escondía al desnudo.
#
# Excepción documentada: la aritmética interna de la matriz de DM6, que
# referencia su PROPIA fila (`=J13+K13`) o la del mes anterior (`=L13` en la
# fila 14). Se reconoce por eso: una referencia desnuda solo es legítima si
# cae en la misma fila o en la inmediata anterior. Las rotas reales apuntan
# lejos (`DM5!C14 = =D30`, `DM7!C19 = =C129`).
TOKEN_REF = re.compile(r"^(?:'[^']+'!|[A-Za-z_][A-Za-z0-9_.]*!)?(\$?[A-Z]{1,3}\$?)(\d+)$")


def _referencias_desnudas(formula: str, fila_celda: int) -> list[str]:
    """Los tokens sin prefijo de hoja de una publicación de direcciones."""
    tokens = formula[1:].split("+")
    partes = [TOKEN_REF.match(t) for t in tokens]
    if not all(partes):
        return []          # no es una publicación de direcciones: lleva
                           # operadores o funciones (aritmética de la cédula)
    return [
        t for t, m in zip(tokens, partes)
        if "!" not in t and int(m.group(2)) not in (fila_celda, fila_celda - 1)
    ]

HOJAS_DM = [
    "DM3 Revisión de saldos", "DM4 Compras", "DM5 Ventas", "DM6 IVA",
    "DM7 Retenciones x pagar", "DM8 ATS",
]

MOVS = [Movimiento(codigo="1.1.5.1.1", cuenta="IVA sobre Compras",
                   fecha=datetime.date(2025, 1, 5), asiento="COM 1", debe=659.57)]


def _libro():
    return load_workbook(BytesIO(armar_libro(
        clasificacion=FILAS,
        movimientos=MOVS,
        f104_monthly={
            f"2025-{m:02d}": {"casilleros": {"429": 4341.16, "411": 1000.0}}
            for m in range(1, 13)
        },
        f103_monthly={
            f"2025-{m:02d}": {"casilleros": {"499": 915.70, "721": 33.0}}
            for m in range(1, 13)
        },
        cliente="MI CLIENTE S.A.", periodo="2025",
    )))


def test_ninguna_cedula_dm_tiene_referencias_sin_nombre_de_hoja():
    wb = _libro()
    rotas = []
    for hoja in HOJAS_DM:
        for fila in wb[hoja].iter_rows():
            for celda in fila:
                valor = celda.value
                if not isinstance(valor, str) or not valor.startswith("="):
                    continue
                for token in _referencias_desnudas(valor, celda.row):
                    rotas.append(f"{hoja}!{celda.coordinate} = {valor} → {token}")
    assert rotas == [], f"{len(rotas)} fórmulas con referencia desnuda: {rotas[:10]}"


def test_el_discriminador_reconoce_una_formula_mixta_como_rota():
    """Guarda del propio test: una fórmula con un tramo bueno y otro desnudo
    (`=N139+'DM7…'!N43`) tiene que contar como rota."""
    assert _referencias_desnudas("=N139+'DM7 Retenciones x pagar'!N43", 35) == ["N139"]
    assert _referencias_desnudas("=D30", 14) == ["D30"]
    assert _referencias_desnudas("=C18+C19", 28) == ["C18", "C19"]
    assert _referencias_desnudas("='Mayores homologados'!D30", 14) == []
    assert _referencias_desnudas("='DATOS F-104'!C18+'DATOS F-104'!C19", 28) == []


def test_el_discriminador_no_marca_la_aritmetica_interna_de_las_cedulas():
    assert _referencias_desnudas("=SUM(C14:C25)", 26) == []
    assert _referencias_desnudas("=ROUND(C26-C31,2)", 32) == []
    assert _referencias_desnudas("=J13+K13", 13) == []      # DM6, misma fila
    assert _referencias_desnudas("=L13", 14) == []          # DM6, mes anterior
    assert _referencias_desnudas("=C15/0.15", 31) == []     # DM4, base desde IVA
