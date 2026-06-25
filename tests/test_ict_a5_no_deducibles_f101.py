"""Regression test: cuando F-101 declara valores en casilleros de gastos
NO DEDUCIBLES (rango 7001-7999, "VALOR NO DEDUCIBLE"), el A5 Cuadro A debe
trasladarlos automáticamente:
  - Col B (No. Casillero) ← código del casillero
  - Col L (Valor declarado) ← ='DATOS F-101'!C<row>

Instrucción cliente 2026-06-18 (equivalente al A4 de ingresos exentos):
fuente ÚNICA = pestaña DATOS F-101.

Caso crítico (PROPHAR): el Cuadro A solo trae 5 filas (17-21) pero el cliente
puede declarar muchos más (PROPHAR: 15). El A5 debe INSERTAR filas para que
entren TODOS sin truncar (REGLA SUPREMA del proyecto), reajustando la fórmula
del total y desplazando los Cuadros B/C/D/E hacia abajo.
"""
from __future__ import annotations

from openpyxl import Workbook

from backend.app.ict.catalogo_gnd import gnd_descripcion, gnd_normativa
from backend.app.ict.cell_maps.a5 import A5_SHEET
from backend.app.ict.fillers.a5_conciliacion_costos import A5Filler
from backend.app.ict.fillers.source_data_sheets import build_f101_sheet
from backend.app.ict.fillers.base import load_template


def _session() -> dict:
    return {"razon_social": "TEST A5 ND", "ruc": "T",
            "ejercicio_fiscal": "2025", "numero_adhesivo": ""}


def _build_a5(valores_f101: dict) -> tuple:
    """Construye un ICT con DATOS F-101 y rellena el A5. Devuelve (ws_a5, lookup)."""
    wb = load_template()
    f101 = dict(valores_f101)
    lookup = build_f101_sheet(wb, f101, {})
    A5Filler().fill(wb, _session(), {
        "f101": f101,
        "balance_mapeado": [],
        "_f101_lookup": lookup,
        "_balance_lookup": [],
    })
    return wb[A5_SHEET], lookup


# ── 1. Traslado básico (≤5 cas, sin inserción) ───────────────────────────────

def test_a5_traslada_no_deducibles_a_B_y_L():
    ws, _ = _build_a5({"7042": 9_207.26, "7048": 63_279.36, "7249": 758_785.38})
    # Orden creciente por casillero
    assert ws["B17"].value == "7042"
    assert ws["B18"].value == "7048"
    assert ws["B19"].value == "7249"
    # Col L = referencia al valor declarado en F-101
    for r in (17, 18, 19):
        l = ws.cell(r, 12).value  # col L
        assert isinstance(l, str) and "DATOS F-101" in l, f"L{r}={l!r}"


def test_a5_no_deducible_valor_cero_no_aparece():
    ws, _ = _build_a5({"7042": 0.0, "7048": 100.0})
    presentes = [ws.cell(r, 2).value for r in range(17, 22) if ws.cell(r, 2).value]
    assert "7048" in presentes
    assert "7042" not in presentes


# ── 2. Inserción de filas cuando hay > 5 cas no deducibles ───────────────────

def test_a5_inserta_filas_para_8_no_deducibles():
    f101 = {str(c): float(i + 1) * 100
            for i, c in enumerate([7042, 7048, 7057, 7060, 7063, 7069, 7177, 7183])}
    ws, _ = _build_a5(f101)
    # Los 8 casilleros deben estar en B17:B24 (5 originales + 3 insertadas)
    casilleros_en_b = [ws.cell(r, 2).value for r in range(17, 25)]
    for cas in f101:
        assert cas in casilleros_en_b, f"falta {cas} en B17:B24 = {casilleros_en_b}"


def test_a5_15_no_deducibles_todos_presentes_sin_truncar():
    cas_list = [7042, 7048, 7057, 7060, 7063, 7069, 7177, 7183,
                7186, 7192, 7198, 7204, 7210, 7243, 7249]
    f101 = {str(c): float(i + 1) * 1000 for i, c in enumerate(cas_list)}
    ws, _ = _build_a5(f101)
    # 15 cas → 10 filas insertadas → datos en B17:B31
    presentes = [ws.cell(r, 2).value for r in range(17, 32) if ws.cell(r, 2).value]
    for c in cas_list:
        assert str(c) in presentes, f"falta {c}; presentes={presentes}"
    assert len(presentes) == 15


# ── 3. Total K ajustado tras inserción ───────────────────────────────────────

def test_a5_total_K_cubre_todas_las_filas_tras_insercion():
    f101 = {str(c): 100.0 for c in [7042, 7048, 7057, 7060, 7063, 7069, 7177]}
    # 7 cas → 2 filas insertadas → datos 17..23, total en fila 24
    ws, _ = _build_a5(f101)
    total = ws.cell(24, 11).value  # K24 (col K = 11)
    assert isinstance(total, str) and total.startswith("=SUM(K17:K23)"), f"K24={total!r}"


# ── 4. Cuadros B/C/D se desplazan correctamente tras inserción ───────────────

def test_a5_cuadros_bcd_se_desplazan_tras_insercion():
    # 7 no deducibles (+2 filas) + casilleros de los cuadros B/C/D
    f101 = {
        "6999": 18_000_000.0, "7999": 12_000_000.0,
        "804": 5_000.0, "805": 1_200.0, "808": 0.0,
        "806": 50_000.0, "807": 0.0, "809": 750.0, "813": 0.0, "1113": 0.0,
    }
    for c in [7042, 7048, 7057, 7060, 7063, 7069, 7177]:
        f101[str(c)] = 100.0
    ws, _ = _build_a5(f101)
    extra = 2  # 7 - 5
    # Cuadro B: 6999 estaba en G34 → ahora G(34+2)=G36
    g36 = ws.cell(34 + extra, 7).value
    assert isinstance(g36, str) and "DATOS F-101" in g36, f"G36={g36!r}"
    # Cuadro C: cas 804 estaba en H58 → ahora H60
    h60 = ws.cell(58 + extra, 8).value
    assert isinstance(h60, str) and "DATOS F-101" in h60, f"H60={h60!r}"
    # Cuadro D: cas 806 estaba en H66 → ahora H68
    h68 = ws.cell(66 + extra, 8).value
    assert isinstance(h68, str) and "DATOS F-101" in h68, f"H68={h68!r}"


# ── 5. Autocompletado descripción (E) y normativa (G) desde catálogo GND ─────

def test_a5_autocompleta_descripcion_y_normativa_gnd():
    """Al trasladar un casillero no deducible, el A5 escribe:
      - Col E (descripción del tipo de gasto) desde el catálogo GND
      - Col G (normativa aplicable) desde el catálogo GND"""
    ws, _ = _build_a5({"7042": 9_207.26})
    assert ws["B17"].value == "7042"
    assert ws["E17"].value == gnd_descripcion("7042")
    assert ws["G17"].value == gnd_normativa("7042")
    assert ws["E17"].value  # no vacío
    assert ws["G17"].value  # no vacío


def test_a5_autocompleta_gnd_en_filas_insertadas():
    """La descripción/normativa también se autocompletan en las filas que se
    insertan dinámicamente (no solo en las 5 base)."""
    cas_list = [7042, 7048, 7057, 7060, 7063, 7069, 7177, 7183]  # 8 → +3 filas
    f101 = {str(c): float(i + 1) * 100 for i, c in enumerate(cas_list)}
    ws, _ = _build_a5(f101)
    # La fila 24 (insertada) corresponde al 8º casillero (7183)
    assert ws["B24"].value == "7183"
    assert ws["E24"].value == gnd_descripcion("7183")
    assert ws["G24"].value == gnd_normativa("7183")


# ── Preservación de formato del Cuadro D tras inserción (bug ICT_35) ──────────

def test_a5_insercion_preserva_casilleros_y_bordes_cuadro_d():
    """Con >5 no deducibles se insertan filas; los casilleros y bordes del
    Cuadro D (806/807/808/809/813/1113) NO deben perderse en el desplazamiento.
    Reproduce el bug del ICT_35: 807 y 808 desaparecían y perdían bordes."""
    cas = ["7177", "7180", "7183", "7186", "7192", "7198", "7210", "7234",
           "7243", "7249"]  # 10 no deducibles → +5 filas
    f101 = {c: 100.0 * (i + 1) for i, c in enumerate(cas)}
    f101.update({"6999": 1e7, "7999": 1e7, "806": 3000.0})
    ws, _ = _build_a5(f101)
    # Cuadro D desplazado +5: filas 71-76, casilleros en col G (7)
    esperado = {71: 806, 72: 807, 73: 808, 74: 809, 75: 813, 76: 1113}
    for row, cas_esp in esperado.items():
        val = ws.cell(row, 7).value
        assert val == cas_esp, f"G{row} esperado {cas_esp}, encontrado {val!r}"
        b = ws.cell(row, 7).border
        assert b.top.style and b.bottom.style, f"G{row}: bordes perdidos"
    # Merges A:F de las filas del Cuadro D deben existir
    merges = {str(m) for m in ws.merged_cells.ranges}
    for row in (71, 72, 73, 74, 75, 76):
        assert f"A{row}:F{row}" in merges, f"merge A{row}:F{row} perdido. {sorted(merges)}"


# ── REGLA estática: cobertura catálogo GND ⊇ casilleros 7xxx del F-101 ────────

def test_gnd_cubre_todos_los_no_deducibles_del_f101():
    """Todo casillero de gasto no deducible (7001-7999, 'VALOR NO DEDUCIBLE')
    del catálogo F-101 DEBE tener descripción y normativa en el catálogo GND.
    Si el F-101 incorpora un cas nuevo, este test obliga a actualizar el GND."""
    from backend.app.ict.catalogo_f101 import F101_CASILLERO_NAMES
    from backend.app.ict.catalogo_gnd import GND_CASILLERO_INFO

    f101_nd = {
        c for c in F101_CASILLERO_NAMES
        if c.isdigit() and 7001 <= int(c) <= 7999
        and F101_CASILLERO_NAMES[c].upper().startswith("VALOR NO DEDUCIBLE")
    }
    faltantes = sorted(f101_nd - set(GND_CASILLERO_INFO), key=int)
    assert not faltantes, f"Casilleros no deducibles sin entrada en GND: {faltantes}"
    # Toda entrada debe tener descripción y normativa no vacías
    for cas, (desc, norm) in GND_CASILLERO_INFO.items():
        assert desc, f"GND {cas}: descripción vacía"
        assert norm, f"GND {cas}: normativa vacía"


# ── 6. Sin no deducibles → no inserta, cuadros en posición original ───────────

def test_a5_sin_no_deducibles_cuadros_en_posicion_original():
    f101 = {"6999": 18_000_000.0, "7999": 12_000_000.0, "806": 50_000.0}
    ws, _ = _build_a5(f101)
    # Sin inserción: 6999 en G34, 806 en H66 (posiciones originales)
    assert isinstance(ws.cell(34, 7).value, str) and "DATOS F-101" in ws.cell(34, 7).value
    assert isinstance(ws.cell(66, 8).value, str) and "DATOS F-101" in ws.cell(66, 8).value
    # B17 vacío (no hay no deducibles)
    assert ws["B17"].value in (None, "")
