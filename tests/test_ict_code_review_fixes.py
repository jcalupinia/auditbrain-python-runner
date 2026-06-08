"""Tests de regresión para los issues detectados por code-reviewer
(2026-06-07, agente feature-dev:code-reviewer ID afeca977258ebd688).

Cubre los issues 1, 4, 5, 6, 7, 8, 9, 10, 11 — fijando comportamiento
esperado para prevenir regresiones futuras cuando cualquier developer
modifique las reglas de signo, sort_key o parsers.

Issues NO testeados aquí (por falta de fix):
  - ISSUE 2: regla de signo ingresos con saldo negativo (decisión
    pendiente — PROPHAR funciona con ABS, otros clientes pueden romper).
  - ISSUE 3: cas 613, 617 (cooperativas, sin caso real para verificar).
"""
from __future__ import annotations

from io import BytesIO
import pytest
from openpyxl import Workbook

from backend.app.ict.cell_maps.a1 import _a1_sort_key, A1_CASILLEROS_ORDERED
from backend.app.ict.fillers.a1_mapeo import A1Filler
from backend.app.ict.fillers.source_data_sheets import (
    _INFORMATIVOS_EXTRA,
    _es_informativo,
)
from backend.app.ict.catalogo_f101 import F101_CASILLERO_NAMES
from backend.app.ict.parsers.balance_mapeado_excel import (
    _normalize_casillero,
    parse_balance_mapeado,
)


# Asegurar que NEGATIVE_CASILLEROS está poblado.
A1Filler.NEGATIVE_CASILLEROS = A1Filler._build_negative_set()


# ============================================================
# ISSUE 1: cas de depreciación/deterioro activos no estándar
# ============================================================

class TestIssue1NegativeCasilleros:
    """Cas 398, 399, 401, 402, 405, 406, 410, 411 son depreciación/deterioro
    de propiedades de inversión, plantas vivas, animales vivos y activos de
    exploración. Sus nombres NO empiezan con "(-)" pero conceptualmente
    RESTAN — deben estar en NEGATIVE_CASILLEROS para que col F del A1 los
    muestre con `=-ABS(...)`.
    """

    @pytest.mark.parametrize("cas", ["398", "399", "401", "402", "405", "406", "410", "411"])
    def test_cas_dep_act_no_estandar_en_negative_casilleros(self, cas):
        assert cas in A1Filler.NEGATIVE_CASILLEROS, (
            f"cas {cas} debe estar en NEGATIVE_CASILLEROS para clientes "
            f"con activos biológicos, propiedades de inversión o exploración"
        )

    def test_cas_clasicos_siguen_en_negative_casilleros(self):
        """Regresión: el fix de issue 1 NO debe sacar cas que ya estaban."""
        for cas in ["314", "317", "324", "327", "329", "347", "384", "385",
                    "386", "392", "393", "602", "612", "616",
                    "7010", "7022", "7028", "7034"]:
            assert cas in A1Filler.NEGATIVE_CASILLEROS, (
                f"cas {cas} (clásico) ya debía estar en NEGATIVE_CASILLEROS"
            )


# ============================================================
# ISSUE 4: _a1_sort_key retorna tupla de 3 elementos siempre
# ============================================================

class TestIssue4SortKeyConsistente:
    """`_a1_sort_key` antes retornaba tuplas de 2 elementos en tier 8/9.
    Ahora todos los returns son de 3 elementos para consistencia.
    """

    @pytest.mark.parametrize("cas", [
        "311", "361", "449", "499", "550", "589", "599", "698", "699",
        "1005", "6999", "7999",
        "801", "850", "899", "999",  # tier 8 (antes 2 elementos)
        "1100", "5001", "9999",      # tier 9 (antes 2 elementos)
    ])
    def test_sort_key_retorna_3_elementos(self, cas):
        result = _a1_sort_key(cas)
        assert isinstance(result, tuple), f"Esperaba tupla, recibí {type(result)}"
        assert len(result) == 3, (
            f"cas {cas} retorna tupla de {len(result)} elementos: {result}. "
            f"Esperado: tupla de 3 elementos (tier, sub_tier, n)"
        )

    def test_sort_key_cas_no_numerico_tambien_3_elementos(self):
        result = _a1_sort_key("abc")
        assert len(result) == 3
        assert result[0] == 99  # tier para no-numéricos


# ============================================================
# ISSUE 7: parser balance normaliza casilleros con cualquier formato
# ============================================================

class TestIssue7NormalizeCasillero:
    """`_normalize_casillero` debe manejar todos los formatos que Excel
    puede emitir según el tipo de la celda y formato regional.
    """

    @pytest.mark.parametrize("inp,esperado", [
        # Strings limpios
        ("311", "311"),
        ("6001", "6001"),
        # Enteros
        (311, "311"),
        (6001, "6001"),
        # Floats con .0
        (311.0, "311"),
        (6001.0, "6001"),
        # Strings con .0
        ("311.0", "311"),
        ("6001.0", "6001"),
        # Strings con .00 (el bug histórico)
        ("311.00", "311"),
        ("6001.00", "6001"),
        # Strings con espacios
        ("311 ", "311"),
        (" 311", "311"),
        # Casos vacíos
        ("", ""),
        (None, ""),
        # Strings no numéricos (preservar)
        ("abc", "abc"),
    ])
    def test_normalize_casillero(self, inp, esperado):
        assert _normalize_casillero(inp) == esperado, (
            f"_normalize_casillero({inp!r}) → esperado {esperado!r}"
        )

    def test_parser_acepta_casilleros_con_dos_decimales(self):
        """E2E: cliente con Excel formateado como 'Número 2 decimales'
        en la columna cas SRI debe parsear correctamente."""
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1, value="Cod.Cuenta.Contable")
        ws.cell(1, 2, value="Descripción")
        ws.cell(1, 3, value="Códigos SRI")
        ws.cell(1, 4, value="Saldos 31 DIC")
        # Excel exporta como "311.00" — antes el parser fallaba
        ws.cell(2, 1, value="1101"); ws.cell(2, 2, value="Caja")
        ws.cell(2, 3, value="311.00"); ws.cell(2, 4, value=1000.0)
        ws.cell(3, 3, value="6001.00"); ws.cell(3, 4, value=5000.0)
        bio = BytesIO(); wb.save(bio)
        res = parse_balance_mapeado(bio.getvalue())
        cas_extraidos = {c["casillero_sri"] for c in res["cuentas"]}
        assert "311" in cas_extraidos, "cas 311.00 debe normalizarse a 311"
        assert "6001" in cas_extraidos, "cas 6001.00 debe normalizarse a 6001"
        assert "311.00" not in cas_extraidos, "no debe quedar el sufijo .00"


# ============================================================
# ISSUE 8: cas 6147 catálogo con paréntesis balanceado
# ============================================================

class TestIssue8Catalogo6147:
    def test_cas_6147_tiene_parentesis_balanceado(self):
        nombre = F101_CASILLERO_NAMES.get("6147", "")
        assert nombre, "cas 6147 debe existir en catálogo"
        # Debe terminar con (INFORMATIVO) — no INFORMATIVO)
        assert nombre.endswith("(INFORMATIVO)"), (
            f"cas 6147 termina con: ...{nombre[-30:]!r}. "
            f"Esperado: ...(INFORMATIVO)"
        )

    def test_cas_6147_paréntesis_balanceados(self):
        """Validar balance de paréntesis en todo el nombre."""
        nombre = F101_CASILLERO_NAMES.get("6147", "")
        assert nombre.count("(") == nombre.count(")"), (
            f"cas 6147 tiene paréntesis desbalanceados: "
            f"{nombre.count('(')} aperturas vs {nombre.count(')')} cierres"
        )

    def test_cas_6147_NO_está_en_negative_casilleros(self):
        """Documenta que cas 6147 NO está auto-negativizado (el nombre
        no empieza con "(-)" sino con "VALOR TOTAL CORRESPONDIENTE...").
        Si en el futuro se quiere tratar como negativo, agregar a
        _NEGATIVE_CORE manualmente. Este test fija la decisión actual.
        """
        assert "6147" not in A1Filler.NEGATIVE_CASILLEROS, (
            "cas 6147 NO debe estar en NEGATIVE_CASILLEROS — el nombre "
            "del catálogo no empieza con '(-)'. Si querés que sea "
            "tratado como negativo, agregalo a _NEGATIVE_CORE explícitamente."
        )

    def test_cas_6147_es_informativo(self):
        """cas 6147 ahora tiene (INFORMATIVO) en el nombre →
        _es_informativo lo detecta correctamente."""
        nombre = F101_CASILLERO_NAMES.get("6147", "")
        assert _es_informativo(nombre, "6147") is True


# ============================================================
# ISSUE 9: _INFORMATIVOS_EXTRA excluidos de _sum_balance_range
# ============================================================

class TestIssue9InformativosExtraExcluidosDeSum:
    def test_informativos_extra_contiene_cas_esperados(self):
        assert "469" in _INFORMATIVOS_EXTRA
        assert "6140" in _INFORMATIVOS_EXTRA
        assert "7901" in _INFORMATIVOS_EXTRA

    def test_sum_balance_range_excluye_informativos(self):
        """Si una cuenta del balance está mapeada a cas en _INFORMATIVOS_EXTRA,
        NO debe sumarse al total del rango. Esto coincide con la regla del A1
        que también excluye estos cas."""
        from backend.app.ict.fillers.verification import _sum_balance_range
        by_cas = {
            "6001": [{"saldo": 1000.0}],   # ingreso normal
            "6003": [{"saldo": 5000.0}],   # ingreso normal
            "6140": [{"saldo": 999.99}],   # informativo extra — debe excluirse
            "7901": [{"saldo": 333.33}],   # informativo extra — debe excluirse
        }
        # Rango ingresos: 6001-6999 — incluiría 6001, 6003, 6140
        total_ingresos = _sum_balance_range(by_cas, [(6001, 6999)])
        assert total_ingresos == 6000.0, (
            f"Esperaba 6000 (solo 6001+6003), recibí {total_ingresos}. "
            f"Cas 6140 ($999.99) NO debe sumarse."
        )
        # Rango costos: 7001-7999 — incluiría 7901 si no se excluyera
        total_costos = _sum_balance_range(by_cas, [(7001, 7999)])
        assert total_costos == 0.0, (
            f"Esperaba 0 (cas 7901 excluido), recibí {total_costos}"
        )


# ============================================================
# Tests adicionales sugeridos por el reviewer
# ============================================================

class TestVerificationE2E:
    """Tests E2E para asegurar que generate_excel y build_verification_sheet
    no rompen ante datos vacíos o inesperados."""

    def test_verification_sheet_sin_balance_no_rompe(self):
        """build_verification_sheet con balance_mapeado=[] no debe lanzar."""
        from backend.app.ict.fillers.verification import build_verification_sheet
        from backend.app.ict.cell_maps.a1 import A1_SHEET

        wb = Workbook()
        wb.create_sheet(A1_SHEET)
        # No debe lanzar excepción
        build_verification_sheet(
            wb,
            f101={"499": 1000},
            balance_mapeado=[],
            session_data={
                "ruc": "TEST",
                "razon_social": "TEST",
                "ejercicio_fiscal": "2025",
            },
        )
        assert "VERIFICACIÓN A1" in wb.sheetnames

    def test_verification_con_cuentas_sin_saldo_no_rompe(self):
        """Pasar balance_cuentas_sin_saldo no rompe la generación."""
        from backend.app.ict.fillers.verification import build_verification_sheet
        from backend.app.ict.cell_maps.a1 import A1_SHEET

        wb = Workbook()
        wb.create_sheet(A1_SHEET)
        build_verification_sheet(
            wb,
            f101={"499": 1000},
            balance_mapeado=[],
            session_data={
                "ruc": "T", "razon_social": "T", "ejercicio_fiscal": "2025",
            },
            balance_cuentas_sin_saldo=[
                {
                    "_source_excel_row": 100,
                    "codigo": "X",
                    "descripcion": "Y",
                    "casillero_sri": "545",
                }
            ],
        )
        ws = wb["VERIFICACIÓN A1"]
        # Verificar que la sección se renderizó
        found = False
        for r in range(1, ws.max_row + 1):
            v = str(ws.cell(r, 1).value or "")
            if "CUENTAS CON CASILLERO PERO SIN SALDO" in v:
                found = True
                break
        assert found, "Sección de cuentas sin saldo debe renderizarse"


class TestA1OrderRegression:
    """Regresión del fix 0af7207 — cas 490, 491 ANTES de cas 449 (TOTAL)."""

    def test_cas_490_491_antes_que_449(self):
        pos = {cas: i for i, (cas, _) in enumerate(A1_CASILLEROS_ORDERED)}
        assert pos["490"] < pos["449"], (
            f"cas 490 (pos {pos['490']}) debe ir antes que cas 449 "
            f"(pos {pos['449']}) para que SUM(...) lo incluya"
        )
        assert pos["491"] < pos["449"], (
            f"cas 491 (pos {pos['491']}) debe ir antes que cas 449 "
            f"(pos {pos['449']}) para que SUM(...) lo incluya"
        )

    def test_cas_449_antes_que_499(self):
        pos = {cas: i for i, (cas, _) in enumerate(A1_CASILLEROS_ORDERED)}
        assert pos["449"] < pos["499"], (
            "cas 449 (TOTAL ACT NO CORR) debe ir antes que 499 (TOTAL ACTIVO)"
        )

    def test_subtotales_informativos_despues_de_449(self):
        """Cas 468, 469 (subtotales informativos REVALUACIONES) deben ir
        DESPUÉS del cas 449 para que el SUM del TOTAL NO los incluya."""
        pos = {cas: i for i, (cas, _) in enumerate(A1_CASILLEROS_ORDERED)}
        # Si están en A1 (cas 469 ahora está filtrado por _INFORMATIVOS_EXTRA)
        if "468" in pos:
            assert pos["468"] > pos["449"], (
                "cas 468 (subtotal informativo) debe ir DESPUÉS del cas 449"
            )
