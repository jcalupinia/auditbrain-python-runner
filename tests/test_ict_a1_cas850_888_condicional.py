"""Regression test: cas 850 / cas 888 son condicionales en A1 (cliente
ICT_21, 2026-06-13).

Regla (CAMBIO ICT_21):
  - Si la empresa DECLARO cas 888 (IR Corriente) en el F-101 con valor != 0
    → mostrar cas 888 en A1, ocultar cas 850.
  - Si NO declaró cas 888 (valor 0 en F-101) → mostrar cas 850 (IR Causado)
    como fallback, ocultar cas 888.

Razon (CAMBIO ICT_21): la regla previa chequeaba cuenta balance, pero el
cliente puede mapear una cuenta a cas 888 SIN declarar valor en cas 888
del F-101 (el contador declara bajo cas 850). En ese caso el A1 mostraba
cas 888 con col C = 0 y col F = saldo, generando descuadre falso. La
regla correcta usa el F-101 declarado como criterio.
"""
from __future__ import annotations

import openpyxl
import pytest


def _find_cas_row(ws, cas_buscado: str) -> int | None:
    for r in range(13, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).strip() == cas_buscado:
            return r
    return None


def _build_wb_con_f101_888(f101_888_valor: float, balance_888_saldo: float = 0):
    """Genera ICT con cas 888 declarado en F-101 con el valor indicado.

    El criterio actual (CAMBIO ICT_21) es el VALOR F-101, no el balance.
    balance_888_saldo solo afecta a la col F (no a la decision de mostrar).
    """
    from openpyxl import Workbook
    from backend.app.ict.cell_maps.a1 import A1_SHEET
    from backend.app.ict.fillers.a1_mapeo import A1Filler
    from backend.app.ict.fillers.source_data_sheets import build_f101_sheet

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.create_sheet(A1_SHEET)
    f101_data = {
        "6001": 5000.0, "7001": 2000.0, "801": 3000.0, "803": 300.0,
        "850": 700.0,  # IR causado declarado
        "888": f101_888_valor,  # CLAVE: este es el criterio
        "889": 100.0,
    }
    f101_lookup = build_f101_sheet(wb, f101_data, {})

    balance = []
    if balance_888_saldo != 0:
        balance.append({
            "casillero_sri": "888", "codigo": "BAL888",
            "descripcion": "GASTO IR CORRIENTE", "saldo": balance_888_saldo,
        })

    filler = A1Filler()
    session_data = {"razon_social": "TEST", "ruc": "T",
                    "ejercicio_fiscal": "2025", "numero_adhesivo": ""}
    anexo_data = {
        "f101": f101_data,
        "balance_mapeado": balance,
        "_f101_lookup": f101_lookup,
        "_balance_lookup": [],
    }
    filler.fill(wb, session_data, anexo_data)
    return wb


def test_cas_888_aparece_si_tiene_valor_f101():
    """Cuando F-101 declara cas 888 != 0 → cas 888 aparece, 850 NO."""
    from backend.app.ict.cell_maps.a1 import A1_SHEET
    wb = _build_wb_con_f101_888(f101_888_valor=500.0)
    ws = wb[A1_SHEET]

    row_888 = _find_cas_row(ws, "888")
    row_850 = _find_cas_row(ws, "850")

    assert row_888 is not None, "cas 888 debe aparecer cuando F-101 lo declara"
    assert row_850 is None, "cas 850 NO debe aparecer cuando cas 888 esta declarado"


def test_cas_850_aparece_si_cas_888_no_esta_declarado_f101():
    """Cuando F-101 NO declara cas 888 (valor 0) → cas 850 aparece como fallback."""
    from backend.app.ict.cell_maps.a1 import A1_SHEET
    wb = _build_wb_con_f101_888(f101_888_valor=0)
    ws = wb[A1_SHEET]

    row_888 = _find_cas_row(ws, "888")
    row_850 = _find_cas_row(ws, "850")

    assert row_850 is not None, (
        "cas 850 debe aparecer como fallback cuando cas 888 F-101 = 0"
    )
    assert row_888 is None, (
        "cas 888 NO debe aparecer si F-101 no lo declara"
    )


def test_cas_850_aparece_aunque_balance_tenga_cuenta_en_888():
    """REGRESION ICT_21: caso reportado por cliente. La empresa mapea una
    cuenta del balance al cas 888 (con saldo != 0) pero NO declara cas 888
    en su F-101 (declara bajo cas 850). Antes el A1 mostraba cas 888 con
    col C=0 y col F=saldo, generando descuadre. Ahora debe mostrar cas 850.
    """
    from backend.app.ict.cell_maps.a1 import A1_SHEET
    # F-101: cas 888 = 0, cas 850 = 700 (declarado)
    # Balance: cas 888 = 990 (mapeado por el contador pero no declarado)
    wb = _build_wb_con_f101_888(f101_888_valor=0, balance_888_saldo=990.0)
    ws = wb[A1_SHEET]

    row_888 = _find_cas_row(ws, "888")
    row_850 = _find_cas_row(ws, "850")

    assert row_850 is not None, (
        "cas 850 debe aparecer cuando F-101 cas 888 = 0 (aunque balance tenga 888)"
    )
    assert row_888 is None, (
        "cas 888 NO debe aparecer si F-101 lo declara en 0 (era el bug del cliente)"
    )


def test_nunca_aparecen_ambos_simultaneamente():
    """Cas 850 y cas 888 son MUTUAMENTE EXCLUYENTES en A1 — nunca aparecen ambos."""
    from backend.app.ict.cell_maps.a1 import A1_SHEET

    for f101_valor, balance in [(500.0, 0), (0, 0), (500.0, 990.0), (0, 990.0)]:
        wb = _build_wb_con_f101_888(f101_888_valor=f101_valor,
                                     balance_888_saldo=balance)
        ws = wb[A1_SHEET]
        row_888 = _find_cas_row(ws, "888")
        row_850 = _find_cas_row(ws, "850")
        assert not (row_888 and row_850), (
            f"Con f101={f101_valor}, balance={balance}: cas 888 y 850 NUNCA "
            f"deben aparecer juntos (encontrado: 888={row_888}, 850={row_850})"
        )
