"""Regression test: cuando F-101 declara valores en cas de ingresos
exentos / no objeto, A4 Cuadro 1 debe trasladarlos automaticamente.

Reportado por cliente 2026-06-17: si el F-101 declara valor en cas
6081/6083/6085/6094/6150 (RENTAS EXENTAS / NO OBJETO DE IMPUESTO), el
A4 debe poblar:
  - Col B (casillero) desde B16 con el numero del casillero
  - Col G (valor) desde G16 con referencia al valor F-101

Antes: B16:B25 quedaba vacio cuando no habia `mayor_exentos` ni cuentas
del balance mapeadas a cas 804/805/812/1112. La col G era SUMIF reactivo
al balance pero como B estaba vacio, evaluaba a "" → el detalle de los
ingresos exentos declarados en F-101 nunca aparecia.

Ahora: cas 6081, 6083, 6085, 6094, 6150 con valor != 0 en F-101 se
trasladan a las primeras filas libres de B16:B25 y la formula G
correspondiente sigue funcionando reactivamente.
"""
from __future__ import annotations

import pytest


# Cas oficiales SRI de ingresos exentos / no objeto (rango 6001-6999)
# que NO empiezan con "VALOR EXENTO" (esos son subitems informativos).
CAS_EXENTOS_NO_OBJETO = ["6081", "6083", "6085", "6094", "6150"]


def _cas_exentos_disponibles(n: int) -> list[str]:
    """Primeros n casilleros que el A4 traslada (patrón F-101 ∩ librería CMIE)."""
    from backend.app.ict.catalogo_f101 import F101_CASILLERO_NAMES
    from backend.app.ict.catalogo_ingresos_exentos import IE_CASILLERO_INFO
    out: list[str] = []
    for cas in sorted(F101_CASILLERO_NAMES, key=lambda x: int(x) if x.isdigit() else 99999):
        if not cas.isdigit() or not (6001 <= int(cas) <= 6999):
            continue
        nom = F101_CASILLERO_NAMES[cas].upper()
        es_exento = (nom.startswith("VALOR EXENTO") or "RENTAS EXENTAS" in nom
                     or "NO OBJETO DE IMPUESTO" in nom)
        if es_exento and cas in IE_CASILLERO_INFO:
            out.append(cas)
        if len(out) >= n:
            break
    return out


def _build_wb_a4_con_exentos(valores_f101: dict[str, float]):
    """Genera workbook ICT con valores F-101 en cas exentos y devuelve
    la hoja A4 ya rellenada."""
    from openpyxl import Workbook
    from backend.app.ict.cell_maps.a4 import A4_SHEET
    from backend.app.ict.fillers.a4_conciliacion_ingresos import A4Filler
    from backend.app.ict.fillers.source_data_sheets import build_f101_sheet
    from backend.app.ict.fillers.base import load_template

    wb = load_template()  # incluye la hoja A4 con sus fórmulas de template
    # Necesitamos también DATOS F-101 para que las referencias funcionen.
    f101_data = dict(valores_f101)
    f101_lookup = build_f101_sheet(wb, f101_data, {})

    filler = A4Filler()
    session_data = {"razon_social": "TEST", "ruc": "T",
                    "ejercicio_fiscal": "2025", "numero_adhesivo": ""}
    anexo_data = {
        "f101": f101_data,
        "balance_mapeado": [],
        "_f101_lookup": f101_lookup,
        "_balance_lookup": [],
    }
    filler.fill(wb, session_data, anexo_data)
    return wb[A4_SHEET], f101_lookup


def test_a4_traslada_cas_6094_otras_rentas_exentas_con_valor():
    """Cliente declara cas 6094 (OTRAS RENTAS EXENTAS) = 800.00 en F-101.
    El A4 debe poblar B16='6094' y G16 con la formula referencial."""
    ws, f101_lookup = _build_wb_a4_con_exentos({"6094": 800.0})
    assert ws["B16"].value == "6094", (
        f"B16 debe ser '6094' (cas exento con valor en F-101), "
        f"encontrado: {ws['B16'].value!r}"
    )
    g16 = ws["G16"].value
    # G16 puede ser la formula SUMIF reactiva existente O una formula que
    # apunta al valor F-101 del cas. Lo importante: no debe estar vacia y
    # debe ser una fórmula.
    assert isinstance(g16, str) and g16.startswith("="), (
        f"G16 debe ser fórmula, encontrado: {g16!r}"
    )


def test_a4_traslada_multiples_cas_exentos():
    """Si hay 3 cas exentos con valor (todos en la librería CMIE), llena 3
    filas en orden creciente (6081 < 6083 < 6094)."""
    ws, _ = _build_wb_a4_con_exentos({
        "6081": 100.0,
        "6083": 300.0,
        "6094": 800.0,
    })
    valores_b = [ws.cell(r, 2).value for r in range(16, 26)]
    presentes = [v for v in valores_b if v]
    assert "6081" in presentes, f"Cas 6081 debe aparecer. B16:B25 = {presentes}"
    assert "6083" in presentes, f"Cas 6083 debe aparecer. B16:B25 = {presentes}"
    assert "6094" in presentes, f"Cas 6094 debe aparecer. B16:B25 = {presentes}"
    # Orden creciente por número de cas
    nums = [int(v) for v in presentes if str(v).isdigit()]
    assert nums == sorted(nums), (
        f"Cas deben ir en orden creciente. Encontrado: {nums}"
    )


def test_a4_cas_exentos_con_valor_cero_no_aparecen():
    """Cas exentos con valor 0 en F-101 no deben trasladarse."""
    ws, _ = _build_wb_a4_con_exentos({
        "6081": 0.0,    # No debe aparecer
        "6094": 800.0,  # Sí debe aparecer
        "6150": 0.0,    # No debe aparecer
    })
    valores_b = [ws.cell(r, 2).value for r in range(16, 26) if ws.cell(r, 2).value]
    assert "6094" in valores_b, f"Cas 6094 con valor debe estar. {valores_b}"
    assert "6081" not in valores_b, f"Cas 6081 con valor 0 NO debe estar. {valores_b}"
    assert "6150" not in valores_b, f"Cas 6150 con valor 0 NO debe estar. {valores_b}"


def test_a4_no_traslada_6150_informativo():
    """El casillero 6150 ('INGRESOS NO OBJETO DE IMPUESTO A LA RENTA') es
    informativo (no está en la librería del cliente) → NO debe trasladarse
    al Cuadro 1, aunque el F-101 lo declare con valor."""
    ws, _ = _build_wb_a4_con_exentos({"6150": 50_000.0, "6094": 800.0})
    valores_b = [ws.cell(r, 2).value for r in range(16, 26) if ws.cell(r, 2).value]
    assert "6094" in valores_b, f"6094 (exento real) debe estar. {valores_b}"
    assert "6150" not in valores_b, f"6150 (informativo) NO debe estar. {valores_b}"


def test_a4_sin_cas_exentos_no_modifica_B16():
    """Si el F-101 no declara cas exentos, B16 queda como antes (vacio
    si no hay balance_mapeado a cas 804/805/812/1112)."""
    ws, _ = _build_wb_a4_con_exentos({"6001": 5000.0})  # cas no-exento
    valores_b = [ws.cell(r, 2).value for r in range(16, 26) if ws.cell(r, 2).value]
    # Ningún cas exento debe estar
    for cas in CAS_EXENTOS_NO_OBJETO:
        assert cas not in valores_b, (
            f"Cas {cas} no debe aparecer si F-101 no lo declara. {valores_b}"
        )


# ── Autocompletado descripción (E) y normativa (F) desde librería CMIE ───────

def test_a4_autocompleta_descripcion_E_y_normativa_F():
    """Al trasladar un casillero exento, el A4 escribe descripción en col E
    y normativa en col F desde el catálogo de ingresos exentos (CMIE)."""
    from backend.app.ict.catalogo_ingresos_exentos import ie_descripcion, ie_normativa
    ws, _ = _build_wb_a4_con_exentos({"6094": 800.0})
    assert ws["B16"].value == "6094"
    assert ws["E16"].value == ie_descripcion("6094")
    assert ws["F16"].value == ie_normativa("6094")
    assert ws["E16"].value  # no vacío
    assert ws["F16"].value  # no vacío


def test_ie_entradas_tienen_descripcion_y_normativa():
    """Toda entrada de la librería de ingresos exentos debe traer descripción
    y normativa no vacías."""
    from backend.app.ict.catalogo_ingresos_exentos import IE_CASILLERO_INFO
    assert len(IE_CASILLERO_INFO) > 0
    for cas, (desc, norm) in IE_CASILLERO_INFO.items():
        assert desc, f"IE {cas}: descripción vacía"
        assert norm, f"IE {cas}: normativa vacía"


# ── Inserción dinámica de filas cuando hay > 10 ingresos exentos ─────────────

def test_a4_inserta_filas_para_12_exentos():
    """12 casilleros exentos → se insertan 2 filas y entran TODOS (no trunca)."""
    cas = _cas_exentos_disponibles(12)
    assert len(cas) == 12, f"Se necesitan 12 cas exentos, hay {len(cas)}"
    f101 = {c: float(i + 1) * 100 for i, c in enumerate(cas)}
    ws, _ = _build_wb_a4_con_exentos(f101)
    presentes = [ws.cell(r, 2).value for r in range(16, 30) if ws.cell(r, 2).value]
    for c in cas:
        assert c in presentes, f"falta {c}; presentes={presentes}"
    assert len(presentes) == 12


def test_a4_total_G_cubre_filas_tras_insercion():
    """12 exentos → +2 filas → datos 16-27, total en fila 28 = SUM(G16:G27)."""
    cas = _cas_exentos_disponibles(12)
    f101 = {c: 100.0 for c in cas}
    ws, _ = _build_wb_a4_con_exentos(f101)
    total = ws.cell(28, 7).value  # G28
    assert isinstance(total, str) and total.startswith("=SUM(G16:G27)"), f"G28={total!r}"


def test_a4_cuadro2_se_desplaza_tras_insercion():
    """El Cuadro 2 (804/805/812/1112) se desplaza +offset sin romperse."""
    cas = _cas_exentos_disponibles(12)  # +2 filas
    f101 = {c: 100.0 for c in cas}
    f101["804"] = 5_000.0
    ws, _ = _build_wb_a4_con_exentos(f101)
    # cas 804 estaba en G32 → ahora G34
    g34 = ws.cell(32 + 2, 7).value
    assert isinstance(g34, str) and "DATOS F-101" in g34, f"G34={g34!r}"


def test_a4_pocos_exentos_no_inserta_cuadro2_intacto():
    """Con ≤10 exentos no se insertan filas: Cuadro 2 queda en G32."""
    ws, _ = _build_wb_a4_con_exentos({"6094": 800.0, "804": 5_000.0})
    g32 = ws.cell(32, 7).value
    assert isinstance(g32, str) and "DATOS F-101" in g32, f"G32={g32!r}"
