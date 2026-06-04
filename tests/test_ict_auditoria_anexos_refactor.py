"""Tests for the refactored auditoria_anexos.py with fill_auditoria_anexos."""
from datetime import datetime
from decimal import Decimal

import openpyxl
import pytest

from backend.app.ict.audit.schemas import (
    AnexoInterpretation, AnexoStatus, AnexosMetrics, Status,
)


@pytest.fixture
def fake_anexos_metrics():
    statuses = []
    statuses.append(AnexoStatus(codigo="A1", nombre="Mapeo",
                                status=Status.OK, observacion_corta="Cuadra"))
    statuses.append(AnexoStatus(codigo="A2", nombre="Ingresos",
                                status=Status.OK, observacion_corta="$4.2M"))
    statuses.append(AnexoStatus(codigo="A3", nombre="Costos",
                                status=Status.REVISAR,
                                observacion_corta="Δ $1.2K"))
    for c in ["A4", "A5", "A7", "A9"]:
        statuses.append(AnexoStatus(codigo=c, nombre=c,
                                    status=Status.OK, observacion_corta="OK"))
    statuses.append(AnexoStatus(codigo="A6", nombre="Beneficios",
                                status=Status.CRITICO,
                                observacion_corta="Falta"))
    statuses.append(AnexoStatus(codigo="A8", nombre="Com. exterior",
                                status=Status.REVISAR,
                                observacion_corta="Revisar"))
    statuses_sorted = sorted(statuses, key=lambda s: s.codigo)
    return AnexosMetrics(
        anexos=statuses_sorted,
        resumen_global={Status.OK: 6, Status.REVISAR: 2,
                        Status.CRITICO: 1, Status.NA: 0},
    )


@pytest.fixture
def fake_interpretations():
    return {
        c: AnexoInterpretation(
            anexo_codigo=c, anexo_nombre=c,
            resumen_ejecutivo=f"Resumen {c}",
            findings=[], confianza_modelo="alta",
            requiere_revision_humana=False,
            timestamp_analisis=datetime(2026, 6, 4),
            modelo_usado="claude-sonnet-4-7-20260101",
            tokens_consumidos=1000,
        )
        for c in ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9"]
    }


def test_fill_auditoria_anexos_writes_banner_and_grid(
    fake_anexos_metrics, fake_interpretations,
):
    from backend.app.ict.fillers.auditoria_anexos import fill_auditoria_anexos
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("AUDITORÍA DE ANEXOS")
    contexto = {"razon_social": "PROPHAR S.A.",
                "ruc": "1791859596001", "periodo": "2025"}
    fill_auditoria_anexos(
        ws,
        metrics=fake_anexos_metrics,
        interpretations=fake_interpretations,
        contexto=contexto,
    )

    found_banner = False
    found_codes = set()
    found_interpretation_section = False
    found_disclaimer = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            sv = str(v) if v is not None else ""
            if "AUDITBRAIN" in sv:
                found_banner = True
            for code in [f"A{i}" for i in range(1, 10)]:
                if sv == code:
                    found_codes.add(code)
            if "INTERPRETACIÓN" in sv.upper():
                found_interpretation_section = True
            if "validada por el auditor" in sv.lower():
                found_disclaimer = True

    assert found_banner, "Banner ejecutivo no encontrado"
    assert found_codes == {f"A{i}" for i in range(1, 10)}, (
        f"Faltan códigos de anexo en el grid: {found_codes}"
    )
    assert found_interpretation_section, "Sección INTERPRETACIÓN POR ANEXO faltante"
    assert found_disclaimer, "Disclaimer IA obligatorio no presente"
