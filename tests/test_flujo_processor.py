"""Tests del processor de la Herramienta Flujo de Efectivo.

Cubre el pipeline interno (leer inputs por slot -> parsear -> generar Excel)
de forma pragmática, sin montar toda la DB. El test de humo verifica que:
  - el processor existe y es invocable,
  - la lógica de leer los DOS slots (balanza_anterior/balanza_actual) desde
    file_storage, parsear y generar produce un .xlsx válido de 11 hojas,
  - el summary refleja las cuadraturas.
"""
import io

from openpyxl import Workbook, load_workbook

from backend.app.client_portal.flujo import processor


def _balanza_xlsx_bytes(rows):
    wb = Workbook()
    ws = wb.active
    ws.append(["Cuenta", "CODIGO SUPER CIAS", "Codigos SRI", "Saldos 31 DIC"])
    for r in rows:
        ws.append(list(r))
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_processor_existe_y_es_invocable():
    assert callable(processor.flujo_efectivo_processor)


def test_pipeline_interno_genera_excel_de_9_hojas(tmp_path):
    # Slots separados en subcarpetas, como los organiza file_storage.
    from backend.app.aud.obligaciones_fiscales import file_storage

    job_dir = tmp_path / "inputs_job"
    (job_dir / file_storage.INPUTS_DIR).mkdir(parents=True)

    ant = _balanza_xlsx_bytes([
        ("Caja", "1010101", "311", 400.0),
        ("Proveedores", "2010301", "413", -400.0),
    ])
    act = _balanza_xlsx_bytes([
        ("Caja", "1010101", "311", 500.0),
        ("Proveedores", "2010301", "413", -500.0),
    ])
    file_storage.save_input(job_dir, "balanza_anterior", "ant.xlsx", ant)
    file_storage.save_input(job_dir, "balanza_actual", "act.xlsx", act)

    # El processor expone un helper interno testeable que hace parse+generar
    # sobre un job_dir concreto y devuelve (out_path, summary).
    out_path, summary = processor._procesar_job_dir(job_dir)

    assert out_path.exists()
    wb = load_workbook(out_path)
    assert len(wb.sheetnames) == 11  # + Notas + Balance resumido
    assert "Notas" in wb.sheetnames
    assert summary["filas_anterior"] == 2
    assert summary["filas_actual"] == 2
    assert "cuadre_esf" in summary
    assert "cuadre_af" in summary


def test_recalcular_desde_balanzas_regenera_previews_y_artefactos(tmp_path, monkeypatch):
    # El recálculo (editor de balanzas) reusa los motores del server: dado un
    # job existente, con balanzas editadas regenera previews + artefactos.
    from backend.app.aud.obligaciones_fiscales import file_storage

    job_dir = tmp_path / "job_recalc"
    (job_dir / file_storage.INPUTS_DIR).mkdir(parents=True)
    # el helper resuelve el job_dir por id: lo apuntamos al tmp
    monkeypatch.setattr(file_storage, "job_dir", lambda _id: job_dir)

    bal_ant = [
        {"cuenta": "Caja", "super_cias": "1010101", "sri": "311", "saldo": 400.0},
        {"cuenta": "Proveedores", "super_cias": "2010301", "sri": "413", "saldo": -400.0},
    ]
    bal_act = [
        {"cuenta": "Caja", "super_cias": "1010101", "sri": "311", "saldo": 900.0},
        {"cuenta": "Proveedores", "super_cias": "2010301", "sri": "413", "saldo": -900.0},
    ]

    prev = processor.recalcular_desde_balanzas(1, bal_ant, bal_act)

    # devuelve previews frescos con las secciones clave
    assert "FLU_95" in prev and "MAP" in prev and "MAP_ANT" in prev
    assert len(prev["MAP"]["rows"]) == 2 and len(prev["MAP_ANT"]["rows"]) == 2
    # el saldo editado se refleja (saldo en índice 4 tras agregar la col Nombre)
    assert prev["MAP"]["rows"][0][4] == 900.0
    # regeneró los artefactos descargables (Excel + TXT del flujo + previews.json)
    art = job_dir / processor.ARTIFACTS_DIR
    assert (art / processor.ARCH_FLU).exists()
    assert (art / "previews.json").exists()
    assert file_storage.output_path(job_dir).exists()


def test_recalcular_rechaza_balanza_vacia(tmp_path, monkeypatch):
    from backend.app.aud.obligaciones_fiscales import file_storage
    monkeypatch.setattr(file_storage, "job_dir", lambda _id: tmp_path / "j")
    import pytest
    with pytest.raises(ValueError):
        processor.recalcular_desde_balanzas(1, [], [{"cuenta": "x", "super_cias": "1010101", "sri": "311", "saldo": 1.0}])
