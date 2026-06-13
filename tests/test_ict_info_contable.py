"""Tests de los helpers de Información contable (A4/A5/A9) y su uso en fillers."""

from backend.app.ict.fillers.referential_helpers import (
    balance_codigo_ref,
    libros_sumif_reactivo_formula,
)


def test_balance_codigo_ref_una_cuenta():
    assert balance_codigo_ref([5]) == "='DATOS BALANCE'!B5"


def test_balance_codigo_ref_varias_cuentas_separa_con_barra():
    f = balance_codigo_ref([5, 7])
    assert f == '=TEXTJOIN(" / ",TRUE,\'DATOS BALANCE\'!B5,\'DATOS BALANCE\'!B7)'


def test_balance_codigo_ref_sin_cuentas_devuelve_none():
    assert balance_codigo_ref([]) is None


def test_libros_sumif_reactivo_con_abs():
    f = libros_sumif_reactivo_formula("$B17")
    assert f == ('=IF($B17="","",ABS(SUMIF(\'DATOS BALANCE\'!$A:$A,$B17,'
                 '\'DATOS BALANCE\'!$D:$D)))')


def test_libros_sumif_reactivo_sin_abs():
    f = libros_sumif_reactivo_formula("$B17", take_abs=False)
    assert f == ('=IF($B17="","",SUMIF(\'DATOS BALANCE\'!$A:$A,$B17,'
                 '\'DATOS BALANCE\'!$D:$D))')


from backend.app.ict.fillers.base import load_template
from backend.app.ict.fillers.a9_inventarios import A9Filler


def _a9_session():
    return {"razon_social": "X", "ruc": "1", "ejercicio_fiscal": "2025",
            "numero_adhesivo": ""}


def test_a9_costo_total_inventario_final_usa_abs():
    """7022 (inv. final) tiene saldo negativo en balance → Costo Total con ABS."""
    wb = load_template()
    anexo_data = {
        "balance_mapeado": [
            {"casillero_sri": "7022", "codigo": "5PYG.53602.017",
             "descripcion": "(-) Inventario final de materia prima",
             "saldo": -930768.56},
        ],
        "_balance_lookup": [5],
    }
    A9Filler().fill(wb, _a9_session(), anexo_data)
    ws = wb["INVENTARIOS A9"]
    assert ws["G21"].value == "=ABS('DATOS BALANCE'!D5)"


def test_a9_costo_total_ajustes_7037_respeta_signo():
    """7037 (ajustes) NO usa ABS — mantiene el signo del balance."""
    wb = load_template()
    anexo_data = {
        "balance_mapeado": [
            {"casillero_sri": "7037", "codigo": "5PYG.53602.017",
             "descripcion": "(+/-) Ajustes", "saldo": -223636.86},
        ],
        "_balance_lookup": [9],
    }
    A9Filler().fill(wb, _a9_session(), anexo_data)
    ws = wb["INVENTARIOS A9"]
    assert ws["G26"].value == "='DATOS BALANCE'!D9"


def test_a9_codigo_cuenta_col_d_es_referencia():
    wb = load_template()
    anexo_data = {
        "balance_mapeado": [
            {"casillero_sri": "7013", "codigo": "5PYG.53602.017",
             "descripcion": "Inventario inicial de materia prima",
             "saldo": 1018613.72},
        ],
        "_balance_lookup": [4],
    }
    A9Filler().fill(wb, _a9_session(), anexo_data)
    ws = wb["INVENTARIOS A9"]
    assert ws["D20"].value == "='DATOS BALANCE'!B4"
    assert ws["H20"].value == "=G20-C20"


def test_a9_diferencia_h22_corregida_a_formula():
    """H22 (cas 7025) estaba hardcodeada en 0 en la plantilla → debe ser =G22-C22."""
    wb = load_template()
    A9Filler().fill(wb, _a9_session(), {"balance_mapeado": [], "_balance_lookup": []})
    ws = wb["INVENTARIOS A9"]
    assert ws["H22"].value == "=G22-C22"
    assert ws["H18"].value == "=G18-C18"
    assert ws["H19"].value == "=G19-C19"


from backend.app.ict.fillers.a4_conciliacion_ingresos import A4Filler


def test_a4_cuadro1_filas_vacias_tienen_formula_reactiva():
    """Sin cuentas exentas pre-llenadas, cada fila del Cuadro 1 lleva una
    fórmula reactiva al casillero (col B) en la col G (valor en libros)."""
    wb = load_template()
    session = {"razon_social": "X", "ruc": "1", "ejercicio_fiscal": "2025", "numero_adhesivo": ""}
    A4Filler().fill(wb, session, {"mayor_exentos": [], "balance_mapeado": []})
    ws = wb["CONCILIACIÓN INGRESOS A4"]
    assert ws["G16"].value == ('=IF($B16="","",ABS(SUMIF(\'DATOS BALANCE\'!$A:$A,'
                               '$B16,\'DATOS BALANCE\'!$D:$D)))')



from backend.app.ict.fillers.a5_conciliacion_costos import A5Filler


def test_a5_cuadro_a_filas_vacias_tienen_formula_reactiva():
    wb = load_template()
    session = {"razon_social": "X", "ruc": "1", "ejercicio_fiscal": "2025", "numero_adhesivo": ""}
    A5Filler().fill(wb, session, {"mayor_no_deducibles": [], "balance_mapeado": []})
    ws = wb["CONCILIACIÓN COSTOS Y GASTOS A5"]
    assert ws["K17"].value == ('=IF($B17="","",ABS(SUMIF(\'DATOS BALANCE\'!$A:$A,'
                               '$B17,\'DATOS BALANCE\'!$D:$D)))')


def test_a9_cuadre_prophar_cero_diferencias():
    """En PROPHAR, el inventario contable (col G, ABS) cuadra con lo declarado:
    todas las diferencias = 0 una vez resueltas las fórmulas."""
    import os
    import pytest
    from openpyxl import load_workbook

    artifact = "audit_artifacts/ict15_papel_trabajo.xlsx"
    if not os.path.exists(artifact):
        pytest.skip("requiere ICT15 PROPHAR regenerado")

    wb = load_workbook(artifact, data_only=False)
    db = wb["DATOS BALANCE"]
    suma: dict[str, float] = {}
    for r in range(4, db.max_row + 1):
        c = db.cell(r, 1).value
        if c is None:
            continue
        c = str(c).strip()
        s = db.cell(r, 4).value or 0
        suma[c] = suma.get(c, 0) + (s if isinstance(s, (int, float)) else 0)
    f1 = wb["DATOS F-101"]
    # Limitar al rango DETALLE — la sección "🔍 CUADRE POR CASILLERO" al
    # final reutiliza col A=cas y col C=fórmula, sobreescribiría los valores.
    f1_end = f1.max_row
    for r in range(1, f1.max_row + 1):
        a = f1.cell(r, 1).value
        if a and "CUADRE" in str(a).upper():
            f1_end = r - 1
            break
    dec: dict[str, float] = {}
    for r in range(1, f1_end + 1):
        v = f1.cell(r, 1).value
        if v:
            dec[str(v).strip()] = f1.cell(r, 3).value
    for cas in ["7013", "7022", "7025", "7028", "7031", "7034", "7037"]:
        d = dec.get(cas) or 0
        g = suma.get(cas, 0)
        g = g if cas == "7037" else abs(g)
        d = d if isinstance(d, (int, float)) else 0
        assert abs(g - d) < 0.01, f"cas {cas}: costo {g} != declarado {d}"
