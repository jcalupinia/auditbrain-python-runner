"""Ensamblado del libro DM."""

import datetime
import re
from io import BytesIO

from openpyxl import load_workbook

from backend.app.aud.obligaciones_fiscales.libro.ensamblador import armar_libro
from backend.app.aud.obligaciones_fiscales.mayor.tipos import Movimiento


class _Fila:
    def __init__(self, codigo, nombre, categoria, por_mes):
        self.codigo_cuenta = codigo
        self.nombre_cuenta = nombre
        self.categoria_final = categoria
        self.por_mes_json = por_mes
        self.n_movimientos = 1
        self.debe = 0.0
        self.haber = 0.0


CLASIFICACION = [
    _Fila("1.1.5.1.1", "IVA sobre Compras", "IVA_COMPRAS", {"01": 659.57}),
    _Fila("4.1.1.1", "Venta de mercadería", "VENTAS", {"01": 5000.0}),
    _Fila("2.1.7.4.1", "IVA sobre Ventas", "IVA_VENTAS", {"01": 750.0}),
    _Fila("2.1.7.3.1", "Ret. 30% Bienes", "RET_IVA", {"01": 12.0}),
    _Fila("2.1.7.2.1", "Ret. 1% Bienes", "RET_RENTA", {"01": 20.0}),
]
MOVS = [Movimiento(codigo="1.1.5.1.1", cuenta="IVA sobre Compras",
                   fecha=datetime.date(2025, 1, 5), asiento="COM 1", debe=659.57)]


def _libro(**kw):
    datos = dict(
        clasificacion=CLASIFICACION, movimientos=MOVS,
        f104_monthly={"2025-01": {"casilleros": {"429": 4341.16}}},
        f103_monthly={"2025-01": {"casilleros": {"499": 915.70}}},
        cliente="MI CLIENTE S.A.", periodo="2025",
    )
    datos.update(kw)
    return load_workbook(BytesIO(armar_libro(**datos)))


def test_el_libro_trae_las_cuatro_hojas_de_datos():
    wb = _libro()
    assert {"Mayores homologados", "Detalle mayor", "DATOS F-104", "DATOS F-103"} <= set(wb.sheetnames)


def test_no_queda_la_hoja_vacia_por_defecto_de_openpyxl():
    assert "Sheet" not in _libro().sheetnames


def test_el_resumen_va_antes_que_el_detalle():
    nombres = _libro().sheetnames
    assert nombres.index("Mayores homologados") < nombres.index("Detalle mayor")


def test_el_libro_se_abre_sin_reparaciones():
    """Regla del proyecto: el Excel no puede pedir reparación al abrirse."""
    wb = _libro()
    for hoja in wb.sheetnames:
        for fila in wb[hoja].iter_rows():
            for celda in fila:
                if isinstance(celda.value, str) and celda.value.startswith("="):
                    assert celda.value.count("(") == celda.value.count(")")


def test_sin_declaraciones_el_libro_igual_se_genera():
    wb = _libro(f104_monthly={}, f103_monthly={})
    assert "DATOS F-104" in wb.sheetnames


def test_el_libro_trae_las_nueve_hojas():
    wb = _libro()
    esperadas = {
        "Mayores homologados", "Detalle mayor", "DM3 Revisión de saldos",
        "DM4 Compras", "DM5 Ventas", "DM6 IVA", "DM7 Retenciones x pagar",
        "DATOS F-104", "DATOS F-103",
    }
    assert esperadas <= set(wb.sheetnames)
    assert len(wb.sheetnames) == 9


def test_el_orden_de_las_hojas_es_el_obligatorio():
    wb = _libro()
    assert wb.sheetnames == [
        "Mayores homologados", "Detalle mayor", "DM3 Revisión de saldos",
        "DM4 Compras", "DM5 Ventas", "DM6 IVA", "DM7 Retenciones x pagar",
        "DATOS F-104", "DATOS F-103",
    ]


def test_ninguna_formula_referencia_una_hoja_inexistente():
    wb = _libro()
    patron = re.compile(r"'([^']+)'!")
    for hoja in wb.sheetnames:
        for fila in wb[hoja].iter_rows():
            for celda in fila:
                valor = celda.value
                if isinstance(valor, str) and valor.startswith("="):
                    for nombre_hoja in patron.findall(valor):
                        assert nombre_hoja in wb.sheetnames, (
                            f"{hoja}!{celda.coordinate} referencia "
                            f"'{nombre_hoja}' que no existe: {valor}"
                        )
