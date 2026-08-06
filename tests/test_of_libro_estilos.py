"""Encabezado común de cédula y marcas de auditoría."""

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.estilos import (
    MARCAS,
    escribir_encabezado_cedula,
    escribir_leyenda_marcas,
)


def _hoja():
    wb = Workbook()
    return wb, wb.active


def test_escribe_el_titulo_de_la_firma_y_de_la_cedula():
    wb, ws = _hoja()
    escribir_encabezado_cedula(ws, titulo="IVA", referencia="DM6",
                               cliente="MI CLIENTE S.A.", periodo="2025")
    assert ws["A1"].value == "OBLIGACIONES FISCALES"
    assert ws["A3"].value == "IVA"


def test_escribe_los_datos_del_encargo():
    wb, ws = _hoja()
    escribir_encabezado_cedula(ws, titulo="IVA", referencia="DM6",
                               cliente="MI CLIENTE S.A.", periodo="2025",
                               preparado_por="JT", revisado_por="V")
    valores = [ws.cell(r, c).value for r in range(1, 11) for c in range(1, 6)]
    assert "MI CLIENTE S.A." in valores
    assert "2025" in valores
    assert "JT" in valores
    assert "V" in valores


def test_la_referencia_de_la_cedula_queda_visible():
    wb, ws = _hoja()
    escribir_encabezado_cedula(ws, titulo="Compras", referencia="DM4",
                               cliente="C", periodo="2025")
    valores = [ws.cell(r, c).value for r in range(1, 11) for c in range(1, 6)]
    assert "DM4" in valores


def test_no_arrastra_datos_de_encargos_anteriores():
    """El modelo del auditor conservaba 'Elaborado por: JT, 2024-09-26'."""
    wb, ws = _hoja()
    escribir_encabezado_cedula(ws, titulo="IVA", referencia="DM6",
                               cliente="C", periodo="2025")
    valores = [str(ws.cell(r, c).value) for r in range(1, 11) for c in range(1, 6)]
    assert not any("2024" in v for v in valores)


def test_las_cuatro_marcas_de_auditoria_usan_la_misma_fuente():
    """El modelo mezclaba Arial y Wingdings para la misma marca."""
    assert set(MARCAS) == {"verificado", "declarado", "diferencia", "sumado"}
    wb, ws = _hoja()
    escribir_leyenda_marcas(ws, fila=20)
    fuentes = {ws.cell(f, 1).font.name for f in range(20, 24)}
    assert len(fuentes) == 1


def test_la_leyenda_explica_cada_marca():
    wb, ws = _hoja()
    escribir_leyenda_marcas(ws, fila=20)
    textos = [str(ws.cell(f, 2).value or "").lower() for f in range(20, 24)]
    assert any("libros" in t for t in textos)
    assert any("diferencia" in t for t in textos)
