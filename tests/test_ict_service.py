"""Tests for ICT service layer."""
import uuid

import pytest

from backend.app.auth.service import create_user
from backend.app.db.session import SessionLocal
from backend.app.ict import service as ict_service


def _unique(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex[:8]}"


@pytest.fixture()
def db_session():
    db = SessionLocal()
    yield db
    db.close()


@pytest.fixture()
def client_user(db_session):
    suffix = _unique("ict")
    email = f"ictuser-{suffix}@x.com"
    user = create_user(db_session, email=email, password="testpass123")
    return user


def test_create_session_sets_expires_at_90_days(db_session, client_user):
    s = ict_service.create_session(
        db_session,
        user=client_user,
        ejercicio_fiscal="2025",
        ruc="1234567890001",
        razon_social="Test S.A.",
        numero_adhesivo=None,
    )
    assert s.status == "in_progress"
    delta = s.expires_at - s.created_at
    assert 89 <= delta.days <= 91


def test_create_session_bootstraps_10_anexos(db_session, client_user):
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="X", numero_adhesivo=None,
    )
    assert len(s.anexos) == 10
    codes = {a.anexo_code for a in s.anexos}
    assert codes == {"INDICE", "A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9"}
    for a in s.anexos:
        assert a.status == "empty"


def test_create_session_returns_existing_if_in_progress(db_session, client_user):
    s1 = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="X", numero_adhesivo=None,
    )
    s2 = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="9999999999999", razon_social="Different", numero_adhesivo=None,
    )
    assert s1.id == s2.id
    assert s2.ruc == "1234567890001"  # original kept


def test_get_active_session_returns_none_if_no_session(db_session, client_user):
    assert ict_service.get_active_session(db_session, user=client_user) is None


def test_get_active_session_returns_in_progress(db_session, client_user):
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="X", numero_adhesivo=None,
    )
    found = ict_service.get_active_session(db_session, user=client_user)
    assert found is not None
    assert found.id == s.id


def test_get_session_raises_for_other_user(db_session, client_user):
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="X", numero_adhesivo=None,
    )
    # Create a different user
    suffix = _unique("other")
    other = create_user(db_session, email=f"other-{suffix}@x.com", password="testpass123")

    with pytest.raises(PermissionError):
        ict_service.get_session(db_session, session_id=s.id, user=other)


def test_update_session_updates_fields_and_touches(db_session, client_user):
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="Old", numero_adhesivo=None,
    )
    original_activity = s.last_activity_at
    import time
    time.sleep(0.01)

    updated = ict_service.update_session(
        db_session, session=s, razon_social="New", numero_adhesivo="ADH-1"
    )
    assert updated.razon_social == "New"
    assert updated.numero_adhesivo == "ADH-1"
    assert updated.last_activity_at > original_activity


def test_expire_session_marks_status(db_session, client_user):
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="X", numero_adhesivo=None,
    )
    ict_service.expire_session(db_session, session=s)
    db_session.refresh(s)
    assert s.status == "expired"


def test_update_anexo_data_persists(db_session, client_user):
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="X", numero_adhesivo=None,
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A1",
        extracted_data={"f101": {"311": 5000}, "balance": {"1.1.1.01.01": {"nombre": "Caja", "saldo": 5000}}},
        warnings=["Casillero 312 no detectado"],
        uploaded_file_meta={"slot": "f101", "filename": "test.pdf", "size": 1024},
        new_status="partial",
    )
    db_session.refresh(s)
    a1 = next(a for a in s.anexos if a.anexo_code == "A1")
    assert a1.status == "partial"
    assert a1.extracted_data["f101"]["311"] == 5000
    assert "Casillero 312 no detectado" in a1.warnings
    assert a1.uploaded_files["f101"]["filename"] == "test.pdf"


def test_indice_recomputed_after_anexo_updated(db_session, client_user):
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="X", numero_adhesivo="ABC-1",
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A1",
        extracted_data={"f101": {"311": 100}},
        warnings=[], uploaded_file_meta={"slot": "f101"},
        new_status="ready",
    )
    ict_service.recompute_indice(db_session, session=s)
    db_session.refresh(s)
    indice = next(a for a in s.anexos if a.anexo_code == "INDICE")
    assert indice.status == "ready"
    assert indice.extracted_data["aplica"]["A1"] == "SI"
    assert indice.extracted_data["aplica"]["A2"] == "NO"


def test_generate_excel_returns_bytes_with_all_sheets(db_session, client_user):
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="Test S.A.", numero_adhesivo="ABC-1",
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A9",
        extracted_data={
            "f101": {"7001": 1000.0, "7010": 2000.0, "7013": 0.0, "7022": 0.0,
                     "7025": 0.0, "7028": 0.0, "7031": 0.0, "7034": 0.0, "7037": 0.0},
            "kardex_items": [],
        },
        warnings=[], uploaded_file_meta={"slot": "f101"},
        new_status="ready",
    )
    ict_service.recompute_indice(db_session, session=s)
    db_session.refresh(s)

    excel_bytes = ict_service.generate_excel(db_session, session=s)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 1000

    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
    assert "INVENTARIOS A9" in wb.sheetnames
    ws = wb["INVENTARIOS A9"]
    # Refactor referencial: C18 contiene fórmula ='DATOS F-101'!C<row>
    # (no valor literal). El valor 1000.0 vive en DATOS F-101.
    c18 = ws["C18"].value
    assert isinstance(c18, str) and c18.startswith("='DATOS F-101'!"), \
        f"C18 debe ser fórmula referencial, obtenido {c18!r}"
    assert "DATOS F-101" in wb.sheetnames
    datos = wb["DATOS F-101"]
    found = False
    for row in datos.iter_rows(min_row=4, values_only=True):
        if row and str(row[0]) == "7001" and row[2] == 1000.0:
            found = True
            break
    assert found, "DATOS F-101 debe tener cas 7001 con valor 1000.0"


def test_save_uploaded_file_writes_to_tmp(tmp_path, monkeypatch):
    # Monkey-patch the OF tmp root to use a clean tmp_path
    from backend.app.aud.obligaciones_fiscales import file_storage as fs
    monkeypatch.setattr(fs, "_root", lambda: tmp_path)

    saved = ict_service.save_uploaded_file(
        session_id=999, anexo_code="A1", slot_name="f101",
        filename="test/with/slashes.pdf", data=b"%PDF fake",
    )
    assert saved.exists()
    assert saved.read_bytes() == b"%PDF fake"
    # Ensure path safe: no slashes, etc.
    assert "/" not in saved.name and "\\" not in saved.name
