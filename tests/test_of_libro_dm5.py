"""DM5 Ventas: ventas gravadas, ventas 0% e IVA en ventas, libros vs declaraciones."""

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm5_ventas import (
    CASILLEROS_IVA_VENTAS, CASILLEROS_VENTAS, CASILLEROS_VENTAS_0, SHEET_DM5, build_dm5,
)

PERIODOS = [f"2025-{m:02d}" for m in range(1, 13)]

DIR_MAYORES = {
    ("cuenta:4.1.1.1.1", "01"): "'Mayores homologados'!D4",
    ("cuenta:4.1.1.1.2", "01"): "'Mayores homologados'!D5",
    ("orden:VENTAS", "cuentas"): ["4.1.1.1.1", "4.1.1.1.2"],
    ("cuenta:2.1.3.1.1", "01"): "'Mayores homologados'!D8",
    ("orden:IVA_VENTAS", "cuentas"): ["2.1.3.1.1"],
}

_TODOS_CASILLEROS = list(dict.fromkeys(
    CASILLEROS_VENTAS + CASILLEROS_VENTAS_0 + CASILLEROS_IVA_VENTAS
))
DIR_F104 = {("2025-01", cas): f"'DATOS F-104'!C{i}"
            for i, cas in enumerate(_TODOS_CASILLEROS, start=20)}

NOMBRES = {
    "4.1.1.1.1": "Ventas Tarifa 15%",
    "4.1.1.1.2": "Ventas Tarifa 0%",
    "2.1.3.1.1": "IVA en Ventas",
}


def _cedula(**kw):
    wb = Workbook()
    datos = dict(dir_mayores=DIR_MAYORES, dir_f104=DIR_F104, periodos=PERIODOS,
                 nombres_cuenta=NOMBRES, cliente="C", periodo="2025")
    datos.update(kw)
    build_dm5(wb, **datos)
    return wb[SHEET_DM5]


def _etiquetas(ws):
    return [ws.cell(r, 2).value for r in range(1, ws.max_row + 1)]


def test_las_cuentas_de_ventas_aparecen_en_los_dos_bloques_de_ventas():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    assert etiquetas.count("Ventas Tarifa 15%") == 2
    assert etiquetas.count("Ventas Tarifa 0%") == 2


def test_el_bloque_de_iva_en_ventas_lista_su_cuenta():
    ws = _cedula()
    assert "IVA en Ventas" in _etiquetas(ws)


def test_el_valor_de_una_cuenta_es_una_formula_al_resumen():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(r, 2).value == "IVA en Ventas")
    assert ws.cell(fila, 3).value == "='Mayores homologados'!D8"


def test_los_casilleros_son_formulas_a_la_hoja_de_datos_no_valores():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if str(ws.cell(r, 2).value or "").startswith("Casillero 411"))
    valor = ws.cell(fila, 3).value
    assert isinstance(valor, str) and valor.startswith("='DATOS F-104'!")


def test_estan_los_casilleros_de_los_tres_bloques():
    ws = _cedula()
    etiquetas = [str(ws.cell(r, 2).value or "") for r in range(1, ws.max_row + 1)]
    for cas in CASILLEROS_VENTAS + CASILLEROS_VENTAS_0 + CASILLEROS_IVA_VENTAS:
        assert any(e.startswith(f"Casillero {cas}") for e in etiquetas), cas


def test_hay_una_diferencia_por_cada_uno_de_los_tres_bloques():
    ws = _cedula()
    assert _etiquetas(ws).count("Diferencia") == 3


def test_la_diferencia_resta_libros_menos_declarado_y_redondea():
    ws = _cedula()
    filas = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "Diferencia"]
    for fila in filas:
        assert ws.cell(fila, 3).value.startswith("=ROUND(")


def test_hay_una_fila_total_ventas_declaradas():
    ws = _cedula()
    assert "Total ventas declaradas" in _etiquetas(ws)


def test_lleva_el_encabezado_de_cedula_con_su_referencia():
    ws = _cedula()
    valores = [ws.cell(r, c).value for r in range(1, 11) for c in range(1, 6)]
    assert "OBLIGACIONES FISCALES" in valores
    assert "DM5" in valores


def test_publica_las_direcciones_que_dm6_consume():
    wb = Workbook()
    lookup = build_dm5(wb, dir_mayores=DIR_MAYORES, dir_f104=DIR_F104, periodos=PERIODOS,
                        nombres_cuenta=NOMBRES, cliente="C", periodo="2025")
    for clave in ("ventas_libros", "ventas_0_libros", "iva_ventas_libros", "total_declarado"):
        assert (clave, "01") in lookup, clave
        assert lookup[(clave, "01")].startswith(f"'{SHEET_DM5}'!")


def test_la_direccion_publicada_de_ventas_libros_apunta_a_la_fila_segun_libros():
    wb = Workbook()
    lookup = build_dm5(wb, dir_mayores=DIR_MAYORES, dir_f104=DIR_F104, periodos=PERIODOS,
                        nombres_cuenta=NOMBRES, cliente="C", periodo="2025")
    ws = wb[SHEET_DM5]
    addr = lookup[("ventas_libros", "01")]
    fila = int(addr.split("!")[1].lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
    assert ws.cell(fila, 2).value == "Según libros"


def test_el_bloque_crece_con_mas_cuentas_de_venta_y_el_subtotal_se_desplaza():
    """El numero de cuentas de venta varia por cliente: el subtotal debe
    sumar exactamente el rango de cuentas listadas, no una posicion fija."""
    dir_mayores_grande = {
        **{(f"cuenta:4.1.1.1.{i}", "01"): f"'Mayores homologados'!D{i}" for i in range(1, 13)},
        ("orden:VENTAS", "cuentas"): [f"4.1.1.1.{i}" for i in range(1, 13)],
    }
    nombres_grande = {f"4.1.1.1.{i}": f"Venta {i}" for i in range(1, 13)}
    wb = Workbook()
    build_dm5(wb, dir_mayores=dir_mayores_grande, dir_f104=DIR_F104, periodos=PERIODOS,
              nombres_cuenta=nombres_grande, cliente="C", periodo="2025")
    ws = wb[SHEET_DM5]
    primera_fila_libros = next(r for r in range(1, ws.max_row + 1)
                                if ws.cell(r, 2).value == "Según libros")
    formula = ws.cell(primera_fila_libros, 3).value
    assert formula == f"=SUM(C14:C{primera_fila_libros - 1})"
