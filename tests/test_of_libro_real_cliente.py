"""El libro generado debe reproducir las cifras reales del cliente.

Requiere AUD_OF_FIXTURES_DIR (datos de cliente, fuera del repo público).
"""

from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.aud.obligaciones_fiscales.cedulas.f104_extractor import extract_all_f104
from backend.app.aud.obligaciones_fiscales.libro.ensamblador import armar_libro
from backend.app.aud.obligaciones_fiscales.libro.fuentes import a_periodos_anuales
from backend.app.aud.obligaciones_fiscales.mayor.clasificador import clasificar
from backend.app.aud.obligaciones_fiscales.mayor.cuentas import monto_segun_libros, perfilar
from backend.app.aud.obligaciones_fiscales.mayor.reader import leer_mayor

pytestmark = pytest.mark.skipif(
    not os.getenv("AUD_OF_FIXTURES_DIR"),
    reason="Requiere AUD_OF_FIXTURES_DIR con los archivos reales del cliente",
)


class _FilaClasif:
    """Doble de ``MayorClasificacionJob``: reproduce lo que hace
    ``clasificacion_service.guardar_clasificacion`` para que este test
    ejercite la misma lógica que corre en producción."""

    def __init__(self, r, p):
        self.codigo_cuenta = r.codigo
        self.nombre_cuenta = r.nombre
        self.categoria_final = r.categoria
        self.por_mes_json = monto_segun_libros(p, r.categoria) if p else {}
        self.n_movimientos = p.n_movimientos if p else 0
        self.debe = p.debe if p else 0.0
        self.haber = p.haber if p else 0.0


@pytest.fixture(scope="module")
def libro():
    base = Path(os.environ["AUD_OF_FIXTURES_DIR"])
    lectura = leer_mayor((base / "MAYOR DE IMPUESTOS.xlsx").read_bytes())
    perfiles = perfilar(lectura.movimientos)
    resultados = clasificar(perfiles)
    clasif = [_FilaClasif(r, perfiles.get(r.codigo)) for r in resultados]
    f104_mes, _ = extract_all_f104(sorted((base / "104").glob("*.pdf")))
    data = armar_libro(
        clasificacion=clasif, movimientos=lectura.movimientos,
        f104_monthly=a_periodos_anuales(f104_mes), f103_monthly={},
        cliente="CLIENTE DE PRUEBA", periodo="2025",
    )
    return load_workbook(BytesIO(data))


def test_el_detalle_trae_todos_los_movimientos_del_mayor(libro):
    ws = libro["Detalle mayor"]
    assert ws.max_row - 3 == 4680


def test_ninguna_fila_del_detalle_queda_sin_categoria(libro):
    ws = libro["Detalle mayor"]
    sin = [r for r in range(4, ws.max_row + 1) if ws.cell(r, 1).value == "SIN_CLASIFICAR"]
    assert not sin


def test_el_resumen_tiene_las_veintiocho_cuentas(libro):
    ws = libro["Mayores homologados"]
    codigos = {
        ws.cell(r, 2).value for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value
    }
    assert len(codigos) == 28


def test_los_casilleros_declarados_llegan_con_sus_valores(libro):
    """cas 429 de enero: el IVA generado en ventas del mes."""
    ws = libro["DATOS F-104"]
    fila = next(r for r in range(4, ws.max_row + 1) if str(ws.cell(r, 1).value) == "429")
    valores = [ws.cell(fila, c).value for c in range(3, 15)]
    assert any(v for v in valores), "el casillero 429 no debería estar todo en cero"


def test_el_bloque_de_credito_tributario_esta_presente(libro):
    """605-608 y 615-619: los que faltaban en el catálogo hasta este trabajo."""
    ws = libro["DATOS F-104"]
    presentes = {str(ws.cell(r, 1).value) for r in range(4, ws.max_row + 1)}
    assert {"605", "606", "615", "617"} <= presentes
