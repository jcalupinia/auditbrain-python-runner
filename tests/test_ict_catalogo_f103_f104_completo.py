"""REGLAS OBLIGATORIAS — DATOS F-103 y DATOS F-104 completos.

Pedido del usuario:
    "verificar que se traslada TODOS los casilleros con sus códigos,
    nombres, valores — no importa que estén cero, con la finalidad de
    que no haya saldos de líneas [vacías]"

Esta suite garantiza la regla aplicada a F-103 y F-104 (igual que
catalogo_f101 para F-101).
"""

import openpyxl

from backend.app.aud.obligaciones_fiscales.cedulas.f104_extractor import ALL_CASILLEROS as F104_PARSER_CAS
from backend.app.ict.catalogo_f103 import F103_CASILLERO_NAMES, get_casillero_name as get_f103_name
from backend.app.ict.catalogo_f104 import F104_CASILLERO_NAMES, get_casillero_name as get_f104_name
from backend.app.ict.parsers.f103_pdf import ALL_CASILLEROS as F103_PARSER_CAS


# ═══════════════════════════════════════════════════════════════════════════
# CAPA A — REGLAS ESTÁTICAS F-103
# ═══════════════════════════════════════════════════════════════════════════

def test_regla_f103_todos_los_casilleros_parseables_tienen_nombre():
    """ALL_CASILLEROS del parser F-103 debe ⊆ F103_CASILLERO_NAMES."""
    faltantes = [c for c in F103_PARSER_CAS if c not in F103_CASILLERO_NAMES]
    assert not faltantes, (
        f"❌ {len(faltantes)} casilleros F-103 parseables SIN nombre canónico:\n"
        f"    {faltantes}\n"
        f"Acción: agregar cada uno a backend/app/ict/catalogo_f103.py"
    )


def test_regla_f103_nombres_no_son_vacios():
    vacios = [c for c, n in F103_CASILLERO_NAMES.items() if not n or not n.strip()]
    assert not vacios, f"❌ F-103 cas con nombre vacío: {vacios}"


def test_regla_f103_casilleros_criticos_presentes():
    """Casilleros críticos del F-103 que SIEMPRE deben estar."""
    criticos = [
        "302",  # Relación dependencia (siempre hay sueldos)
        "303",  # Honorarios profesionales
        "312",  # Bienes muebles
        "349",  # Subtotal país
        "399",  # Subtotal país retenido
        "499",  # Total retención IR
    ]
    for cas in criticos:
        assert cas in F103_CASILLERO_NAMES, f"F-103 cas crítico {cas} sin nombre"


def test_regla_f103_pagos_exterior_completos():
    """Los 3 bloques de pagos al exterior (CDI, sin CDI, paraísos) deben estar."""
    # Con CDI: 402-412
    for cas in range(402, 413):
        assert str(cas) in F103_CASILLERO_NAMES, f"Cas exterior CON CDI {cas} faltante"
    # Sin CDI: 413-422
    for cas in range(413, 423):
        assert str(cas) in F103_CASILLERO_NAMES, f"Cas exterior SIN CDI {cas} faltante"
    # Paraísos: 424-433
    for cas in range(424, 434):
        assert str(cas) in F103_CASILLERO_NAMES, f"Cas paraísos {cas} faltante"


def test_regla_f103_helper_get_name_funciona():
    assert get_f103_name("302") == "En relación de dependencia que supera o no la base desgravada"
    assert get_f103_name("99999") == ""
    assert get_f103_name("99999", fallback="?") == "?"


# ═══════════════════════════════════════════════════════════════════════════
# CAPA A — REGLAS ESTÁTICAS F-104
# ═══════════════════════════════════════════════════════════════════════════

def test_regla_f104_todos_los_casilleros_parseables_tienen_nombre():
    """ALL_CASILLEROS del parser F-104 debe ⊆ F104_CASILLERO_NAMES."""
    faltantes = [c for c in F104_PARSER_CAS if c not in F104_CASILLERO_NAMES]
    assert not faltantes, (
        f"❌ {len(faltantes)} casilleros F-104 parseables SIN nombre canónico:\n"
        f"    {faltantes}\n"
        f"Acción: agregar cada uno a backend/app/ict/catalogo_f104.py"
    )


def test_regla_f104_nombres_no_son_vacios():
    vacios = [c for c, n in F104_CASILLERO_NAMES.items() if not n or not n.strip()]
    assert not vacios, f"❌ F-104 cas con nombre vacío: {vacios}"


def test_regla_f104_casilleros_criticos_presentes():
    """Casilleros críticos del F-104 que SIEMPRE deben estar."""
    criticos = [
        "411",  # Ventas locales tarifa diferente 0%
        "413",  # Ventas locales tarifa 0% sin derecho
        "419",  # Transferencias no objeto
        "480",  # TOTAL VENTAS
        "499",  # Impuesto causado
        "529",  # Total a pagar percepción
        "799",  # Total retención IVA
    ]
    for cas in criticos:
        assert cas in F104_CASILLERO_NAMES, f"F-104 cas crítico {cas} sin nombre"


def test_regla_f104_retenciones_iva_completas():
    """Casilleros de retención IVA 721-799 — todos deben tener nombre."""
    for cas in ("721", "723", "725", "727", "729", "731", "799"):
        assert cas in F104_CASILLERO_NAMES, f"F-104 retención IVA {cas} faltante"


# ═══════════════════════════════════════════════════════════════════════════
# CAPA B — REGLAS E2E (generando Excel real)
# ═══════════════════════════════════════════════════════════════════════════

def _build_f103_test(f103_monthly: dict) -> openpyxl.Workbook:
    from backend.app.ict.fillers.source_data_sheets import build_f103_sheet
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    build_f103_sheet(wb, f103_monthly)
    return wb


def _build_f104_test(f104_monthly: dict) -> openpyxl.Workbook:
    from backend.app.ict.fillers.source_data_sheets import build_f104_sheet
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    build_f104_sheet(wb, f104_monthly)
    return wb


def test_e2e_f103_tiene_todos_los_canonicos_incluso_sin_datos():
    """Aún con F-103 vacío (cliente no subió), DATOS F-103 debe tener
    una fila por cada cas canónico con su nombre."""
    wb = _build_f103_test({})
    assert "DATOS F-103" in wb.sheetnames
    ws = wb["DATOS F-103"]

    cas_en_hoja = set()
    for r in range(4, ws.max_row + 1):
        cas = ws.cell(r, 1).value
        if cas:
            cas_en_hoja.add(str(cas).strip())

    faltantes = set(F103_CASILLERO_NAMES.keys()) - cas_en_hoja
    assert not faltantes, f"DATOS F-103 sin {len(faltantes)} canónicos: {sorted(faltantes)[:10]}"


def test_e2e_f104_tiene_todos_los_canonicos_incluso_sin_datos():
    wb = _build_f104_test({})
    assert "DATOS F-104" in wb.sheetnames
    ws = wb["DATOS F-104"]

    cas_en_hoja = set()
    for r in range(4, ws.max_row + 1):
        cas = ws.cell(r, 1).value
        if cas:
            cas_en_hoja.add(str(cas).strip())

    faltantes = set(F104_CASILLERO_NAMES.keys()) - cas_en_hoja
    assert not faltantes, f"DATOS F-104 sin {len(faltantes)} canónicos: {sorted(faltantes)[:10]}"


def test_e2e_f103_ninguna_fila_sin_nombre():
    """Recorrer DATOS F-103 y verificar que ninguna fila tenga columna B vacía."""
    f103_monthly = {
        "2025-01": {"casilleros": {"302": 178259.63, "303": 18042.55, "499": 25000.0}},
    }
    wb = _build_f103_test(f103_monthly)
    ws = wb["DATOS F-103"]

    sin_nombre = []
    for r in range(4, ws.max_row + 1):
        cas = ws.cell(r, 1).value
        if not cas:
            continue
        nombre = ws.cell(r, 2).value
        if not nombre or not str(nombre).strip():
            sin_nombre.append((r, cas))

    assert not sin_nombre, f"❌ Filas F-103 SIN nombre: {sin_nombre[:10]}"


def test_e2e_f104_ninguna_fila_sin_nombre():
    f104_monthly = {
        "2025-01": {"casilleros": {"411": 49641.66, "499": 7479.28}},
    }
    wb = _build_f104_test(f104_monthly)
    ws = wb["DATOS F-104"]

    sin_nombre = []
    for r in range(4, ws.max_row + 1):
        cas = ws.cell(r, 1).value
        if not cas:
            continue
        nombre = ws.cell(r, 2).value
        if not nombre or not str(nombre).strip():
            sin_nombre.append((r, cas))

    assert not sin_nombre, f"❌ Filas F-104 SIN nombre: {sin_nombre[:10]}"


def test_e2e_f103_casilleros_no_declarados_aparecen_con_cero():
    """Cas no declarado en ningún mes → debe aparecer con valor 0, no omitirse."""
    f103_monthly = {"2025-01": {"casilleros": {"302": 100.0}}}  # solo cas 302
    wb = _build_f103_test(f103_monthly)
    ws = wb["DATOS F-103"]

    # Cas 499 (no declarado) debe estar con 0
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, 1).value) == "499":
            val_mes = ws.cell(r, 3).value  # primer mes = col C
            assert val_mes == 0, f"Cas 499 (no declarado) debe ser 0, es {val_mes!r}"
            nombre = ws.cell(r, 2).value
            assert "RETENCIÓN DEL IMPUESTO A LA RENTA" in nombre
            return
    assert False, "Cas 499 (canónico) NO aparece en DATOS F-103"


def test_e2e_f104_valores_declarados_se_preservan():
    """Lo que el cliente declaró en F-104 sí debe pasar al Excel."""
    f104_monthly = {
        "2025-01": {"casilleros": {"411": 49641.66, "480": 49641.66}},
    }
    wb = _build_f104_test(f104_monthly)
    ws = wb["DATOS F-104"]

    encontrados = {}
    for r in range(4, ws.max_row + 1):
        cas = str(ws.cell(r, 1).value or "").strip()
        if cas in ("411", "480"):
            encontrados[cas] = ws.cell(r, 3).value  # col C = primer mes

    assert encontrados.get("411") == 49641.66
    assert encontrados.get("480") == 49641.66


def test_e2e_f103_tiene_columna_NOMBRE_y_TOTAL_ANUAL():
    """La hoja debe tener: A=Casillero, B=Nombre, C..=meses, última=TOTAL ANUAL"""
    wb = _build_f103_test({})
    ws = wb["DATOS F-103"]

    # Fila 3 = headers
    assert ws.cell(3, 1).value == "Casillero"
    assert ws.cell(3, 2).value == "Nombre del Casillero"
    # Última columna = TOTAL ANUAL
    assert ws.cell(3, ws.max_column).value == "TOTAL ANUAL"


def test_e2e_f104_tiene_columna_NOMBRE_y_TOTAL_ANUAL():
    wb = _build_f104_test({})
    ws = wb["DATOS F-104"]
    assert ws.cell(3, 1).value == "Casillero"
    assert ws.cell(3, 2).value == "Nombre del Casillero"
    assert ws.cell(3, ws.max_column).value == "TOTAL ANUAL"


def test_e2e_f103_total_anual_es_formula_SUM():
    """La columna TOTAL ANUAL debe ser fórmula =SUM(...) sobre las celdas mensuales."""
    wb = _build_f103_test({})
    ws = wb["DATOS F-103"]
    # Fila 4 (primer casillero) — última columna debe ser fórmula
    total_col = ws.max_column
    total_val = ws.cell(4, total_col).value
    assert isinstance(total_val, str) and total_val.startswith("=SUM("), \
        f"TOTAL ANUAL debe ser fórmula =SUM(...), es {total_val!r}"
