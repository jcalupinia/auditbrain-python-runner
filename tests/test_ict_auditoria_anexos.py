"""Tests para la hoja AUDITORÍA DE ANEXOS del Excel ICT 2025.

Valida que:
  1. La hoja existe después de generate_excel().
  2. Tiene los 10 bloques esperados (INDICE + A1..A9) con su título.
  3. Cada bloque tiene Estado, Metodología, Métricas y Análisis.
  4. Las cuadraturas detectadas en A1 aparecen como diferencias.
  5. Los warnings de los fillers se integran al análisis del anexo correspondiente.
"""

import io
import uuid

import openpyxl
import pytest

from backend.app.client_portal.service import create_portal_user
from backend.app.context.models import Client, Organization
from backend.app.db.session import SessionLocal
from backend.app.ict import service as ict_service
from backend.app.ict.fillers.auditoria_anexos import (
    ANEXOS_META,
    SHEET_NAME,
    build_auditoria_anexos_sheet,
    _analyze_anexo,
)


# ─────────────────────────────────────────────────────────────────────────────
# Tests unitarios del analizador (no requieren DB)
# ─────────────────────────────────────────────────────────────────────────────

def test_analyzer_a1_detecta_cuadratura_perfecta():
    """A1 con activo = pasivo + patrimonio en F-101 Y balance que cuadra
    cuenta-por-cuenta → estado ok."""
    meta = next(m for m in ANEXOS_META if m["code"] == "A1")
    ctx = {
        "f101": {"311": 100000, "499": 100000, "511": 60000, "601": 40000,
                 "550": 60000, "698": 40000, "699": 100000},
        "balance_mapeado": [
            # Activo
            {"casillero_sri": "311", "codigo": "1", "descripcion": "Caja", "saldo": 100000},
            # Pasivo
            {"casillero_sri": "511", "codigo": "2", "descripcion": "Prov", "saldo": 60000},
            # Patrimonio
            {"casillero_sri": "601", "codigo": "3", "descripcion": "Cap", "saldo": 40000},
        ],
        "anexo_warnings": {},
    }
    result = _analyze_anexo(meta, ctx)
    assert result["color"] == "ok", \
        f"Esperaba ok, dio {result['color']}. Estado: {result['estado']}"


def test_analyzer_a1_detecta_descuadre():
    """A1 con A != P+Pa en F-101 → estado bad."""
    meta = next(m for m in ANEXOS_META if m["code"] == "A1")
    ctx = {
        "f101": {"499": 100000, "699": 80000},  # desfase 20000
        "balance_mapeado": [],
        "anexo_warnings": {},
    }
    result = _analyze_anexo(meta, ctx)
    # No cuadran las cuentas TOTAL
    assert result["color"] in ("warn", "bad")


def test_analyzer_a2_advierte_f104_no_cargado():
    """A2 sin F-104 mensual → warning."""
    meta = next(m for m in ANEXOS_META if m["code"] == "A2"),
    meta = meta[0]
    ctx = {"f101": {"6001": 50000}, "f104_monthly": {}, "anexo_warnings": {}}
    result = _analyze_anexo(meta, ctx)
    assert result["color"] == "warn"
    assert "f-104" in result["analysis"].lower()


def test_analyzer_a8_advierte_falta_ats_si_hay_pagos_103():
    """Si F-103 declara pagos exterior (cas 402) pero no hay ATS XML → bad."""
    meta = next(m for m in ANEXOS_META if m["code"] == "A8")
    ctx = {
        "f101": {},
        "f103_monthly": {"2025-01": {"casilleros": {"402": 5000.0}}},
        "ats_pagos_exterior": [],
        "anexo_warnings": {},
    }
    result = _analyze_anexo(meta, ctx)
    assert result["color"] == "bad"
    assert "ats" in result["analysis"].lower()


def test_analyzer_a9_advierte_falta_kardex_si_hay_inventarios():
    """A9 con inventarios en F-101 pero sin Kardex → bad."""
    meta = next(m for m in ANEXOS_META if m["code"] == "A9")
    ctx = {
        "f101": {"7001": 50000, "7010": 30000},
        "kardex_items": [],
        "anexo_warnings": {},
    }
    result = _analyze_anexo(meta, ctx)
    assert result["color"] == "bad"
    assert "kardex" in result["analysis"].lower()


def test_analyzer_integra_warnings_del_filler():
    """Si el filler reporta warnings, deben aparecer en el análisis."""
    meta = next(m for m in ANEXOS_META if m["code"] == "A1")
    ctx = {
        "f101": {"499": 100000, "699": 100000},
        "balance_mapeado": [],
        "anexo_warnings": {"A1": ["Casillero 311 sin contraparte balance"]},
    }
    result = _analyze_anexo(meta, ctx)
    assert "warning" in result["analysis"].lower() or "Casillero 311" in result["analysis"]


# ─────────────────────────────────────────────────────────────────────────────
# Tests del builder con un workbook real
# ─────────────────────────────────────────────────────────────────────────────

def test_builder_crea_hoja_auditoria():
    """build_auditoria_anexos_sheet añade la hoja al workbook."""
    wb = openpyxl.Workbook()
    build_auditoria_anexos_sheet(
        wb,
        session_data={"razon_social": "TEST S.A.", "ruc": "123",
                      "ejercicio_fiscal": "2025"},
        f101={"499": 100000, "699": 100000},
        balance_mapeado=[],
    )
    assert SHEET_NAME in wb.sheetnames


def test_builder_pinta_los_10_bloques():
    """La hoja debe tener una sección por cada anexo (INDICE + A1..A9)."""
    wb = openpyxl.Workbook()
    build_auditoria_anexos_sheet(
        wb,
        session_data={"razon_social": "X", "ruc": "1", "ejercicio_fiscal": "2025"},
        f101={"499": 100000, "699": 100000},
    )
    ws = wb[SHEET_NAME]
    # Reunir todos los valores de col A
    all_a_values = " ".join(str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1))
    for meta in ANEXOS_META:
        assert meta["code"] in all_a_values, f"Falta sección del anexo {meta['code']}"
        # Y el nombre humano del anexo
        assert meta["nombre"][:15] in all_a_values, f"Falta nombre del anexo {meta['code']}"


def test_builder_incluye_metodologia_y_fuentes():
    """Cada sección debe documentar qué pide el SRI y de qué fuente viene."""
    wb = openpyxl.Workbook()
    build_auditoria_anexos_sheet(
        wb,
        session_data={"razon_social": "X", "ruc": "1", "ejercicio_fiscal": "2025"},
        f101={},
    )
    ws = wb[SHEET_NAME]
    all_text = " ".join(str(ws.cell(r, c).value or "")
                        for r in range(1, ws.max_row + 1)
                        for c in range(1, 9))
    # Frases que prueban que la metodología quedó escrita
    assert "Qué pide el SRI" in all_text or "pide el SRI" in all_text
    assert "Fuentes esperadas" in all_text or "Fuentes" in all_text
    assert "F-101" in all_text
    assert "Balance Mapeado" in all_text


def test_builder_kpi_anexos_completos_calculado():
    """KPI debe contar anexos en estado ok del total ANEXOS_META."""
    wb = openpyxl.Workbook()
    build_auditoria_anexos_sheet(
        wb,
        session_data={"razon_social": "X", "ruc": "1", "ejercicio_fiscal": "2025"},
        f101={"499": 100000, "699": 100000},
    )
    ws = wb[SHEET_NAME]
    all_text = " ".join(str(ws.cell(r, c).value or "")
                        for r in range(1, 12) for c in range(1, 9))
    # Debe tener el KPI ANEXOS GENERADOS visible
    assert "ANEXOS GENERADOS" in all_text or "GENERADOS" in all_text


# ─────────────────────────────────────────────────────────────────────────────
# Test E2E vía generate_excel (requiere DB)
# ─────────────────────────────────────────────────────────────────────────────

def _unique(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex[:8]}"


@pytest.fixture()
def db_session():
    db = SessionLocal()
    yield db
    db.close()


@pytest.fixture()
def client_user(db_session):
    suffix = _unique("audit")
    org = Organization(name=f"O-{suffix}", slug=f"o-{suffix}", is_active=True)
    db_session.add(org)
    db_session.commit()
    db_session.refresh(org)
    cli = Client(organization_id=org.id, name=f"C-{suffix}", is_active=True)
    db_session.add(cli)
    db_session.commit()
    db_session.refresh(cli)
    user, _ = create_portal_user(db_session, client_id=cli.id, email=f"audit-{suffix}@x.com")
    return user


@pytest.mark.skip(
    reason="2026-06-17: hoja AUDITORÍA DE ANEXOS eliminada del flujo por "
           "decisión del cliente. La info clave vive ahora en VERIFICACIÓN A1 "
           "y en CUADRE POR CASILLERO de DATOS BALANCE/F-101. El test se "
           "mantiene como referencia por si se reactiva la hoja."
)
def test_e2e_generate_excel_incluye_hoja_auditoria(db_session, client_user):
    """generate_excel() debe producir un Excel con la hoja AUDITORÍA."""
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="9999999999001", razon_social="TEST AUDITORÍA",
        numero_adhesivo=None,
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A1",
        extracted_data={
            "f101": {"311": 5000.0, "499": 5000.0, "699": 5000.0},
            "balance_mapeado": [
                {"casillero_sri": "311", "codigo": "X", "descripcion": "Caja", "saldo": 5000.0},
            ],
        },
        warnings=[], uploaded_file_meta={"slot": "balance_mapeado"},
        new_status="ready",
    )
    ict_service.recompute_indice(db_session, session=s)
    db_session.refresh(s)

    _sri, excel_bytes = ict_service.generate_excel(db_session, session=s)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)

    assert SHEET_NAME in wb.sheetnames, \
        f"Falta hoja '{SHEET_NAME}'. Hojas presentes: {wb.sheetnames}"

    ws = wb[SHEET_NAME]
    all_text = " ".join(str(ws.cell(r, c).value or "")
                        for r in range(1, ws.max_row + 1)
                        for c in range(1, 9))
    # Verifica que el nombre del cliente aparece en la hoja
    assert "TEST AUDITORÍA" in all_text
    # Verifica que están los 10 anexos
    for meta in ANEXOS_META:
        assert meta["code"] in all_text, f"Falta análisis del anexo {meta['code']}"
