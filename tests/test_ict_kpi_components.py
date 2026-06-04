"""Tests for backend.app.ict.fillers.kpi_components — Excel visual helpers."""
from decimal import Decimal

import openpyxl

from backend.app.ict.audit.schemas import (
    AnexoFinding,
    AnexoStatus,
    Status,
)
from backend.app.ict.fillers.kpi_components import (
    STATUS_COLORS,
    build_executive_banner,
    build_finding_box,
    build_kpi_card,
    build_traffic_light_grid,
)


def test_status_colors_has_all_4_statuses():
    assert Status.OK in STATUS_COLORS
    assert Status.REVISAR in STATUS_COLORS
    assert Status.CRITICO in STATUS_COLORS
    assert Status.NA in STATUS_COLORS


def test_build_kpi_card_writes_title_and_value():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_kpi_card(
        ws, anchor="B2", title="ACTIVO TOTAL",
        value="$ 21,671,880.68", status=Status.OK,
        subtitle="F-101 cas 499", width_cols=3, height_rows=4,
    )
    assert ws["B2"].value == "ACTIVO TOTAL"
    # Valor en la fila central del card (anchor row + 2)
    assert "21" in str(ws["B4"].value)


def test_build_traffic_light_grid_creates_9_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    statuses = [
        AnexoStatus(codigo=f"A{i}", nombre=f"Anexo {i}", status=Status.OK,
                    observacion_corta="OK")
        for i in range(1, 10)
    ]
    build_traffic_light_grid(ws, anchor="B2", anexos_status=statuses)
    # Verifica que las 9 celdas tienen su código
    found_codes = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell in [f"A{i}" for i in range(1, 10)]:
                found_codes.add(cell)
    assert found_codes == {f"A{i}" for i in range(1, 10)}


def test_build_finding_box_writes_all_fields():
    wb = openpyxl.Workbook()
    ws = wb.active
    f = AnexoFinding(
        severity="critico", categoria="subdeclaracion_ventas",
        titulo="Subdeclaración Q4",
        descripcion_tecnica="cas 6999: $4.2M vs $5.4M",
        implicacion_tributaria="Riesgo glosa",
        recomendacion="Conciliar Q4",
        monto_disputa=Decimal("1200000.00"),
        casilleros_afectados=["6999"],
    )
    end_row = build_finding_box(ws, anchor_row=2, anchor_col=2, finding=f)
    # The finding box should have written content rows; collect non-empty values
    contents = []
    for row in ws.iter_rows(min_row=2, max_row=end_row, values_only=True):
        for v in row:
            if v:
                contents.append(str(v))
    joined = " ".join(contents)
    assert "Subdeclaración Q4" in joined
    assert "Riesgo glosa" in joined
    assert "Conciliar Q4" in joined


def test_build_executive_banner_writes_title():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_executive_banner(
        ws, anchor="A1",
        title_main="AUDITBRAIN · PAPEL DE TRABAJO",
        title_sub="VERIFICACIÓN ANEXO A1",
        meta="PROPHAR S.A. · RUC 1791859596001",
    )
    assert "AUDITBRAIN" in str(ws["A1"].value)
