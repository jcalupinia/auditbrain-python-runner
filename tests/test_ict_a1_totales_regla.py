"""REGLA OBLIGATORIA para el ICT 2025 — A1 debe tener TOTAL por cada grupo.

Esta regla protege contra el bug que reportó el usuario:
    "en la columna de totales de F-101 declarado no está sacando totales
    de varias cuentas"

El bug ocurría porque algunos casilleros TOTAL (cas 550, 1005, 1045,
7991, etc.) NO estaban en alguna de estas 4 estructuras:
    1. A1_CASILLEROS_ORDERED — para que el filler los procese
    2. ALL_F101_CASILLEROS — para que el parser los extraiga del PDF
    3. A1Filler.TOTAL_CASILLEROS — para que el filler los marque como TOTAL
    4. A1Filler.PRIMARY_TOTAL_BLOCKS o COMPOSITE_TOTALS — para que tengan
       fórmula SUM en la columna F (saldo contable)

Si una sola estructura está incompleta, la columna 'F-101 declarado' o
'Balance contable' aparece vacía en VERIFICACIÓN A1, dando la falsa
impresión de que el sistema está roto.

ESTAS PRUEBAS son la regla que evita esa regresión:
- Si alguien borra un TOTAL del parser → CI falla.
- Si alguien agrega un TOTAL nuevo y olvida registrarlo en una de las
  4 estructuras → CI falla con un mensaje explícito de qué falta.

Para AGREGAR un nuevo TOTAL al sistema, actualizar TOTALES_ESPERADOS
abajo + las 4 estructuras mencionadas. El test que se rompa te dirá
cuál falta.
"""

import io

import openpyxl
import pytest

from backend.app.ict.cell_maps.a1 import (
    A1_CASILLEROS_ORDERED,
    A1_FIRST_DATA_ROW,
    A1_SHEET,
)
from backend.app.ict.fillers.a1_mapeo import A1Filler
from backend.app.ict.fillers.base import load_template, reset_trace
from backend.app.ict.parsers.f101_pdf import ALL_F101_CASILLEROS


# ─────────────────────────────────────────────────────────────────────────────
# REGLA: estos son los TOTALES canónicos que SIEMPRE deben aparecer en el A1
# del ICT 2025. Si el SRI cambia el formulario y agrega o renombra alguno,
# actualizar este diccionario es lo PRIMERO que hay que hacer.
# ─────────────────────────────────────────────────────────────────────────────
TOTALES_ESPERADOS: dict[str, str] = {
    # Estado de Situación Financiera — Balance
    "361":  "TOTAL ACTIVOS CORRIENTES",
    "449":  "TOTAL ACTIVOS NO CORRIENTES",
    "499":  "TOTAL DEL ACTIVO",
    "550":  "TOTAL PASIVOS CORRIENTES",
    "589":  "TOTAL PASIVOS NO CORRIENTES",
    "599":  "TOTAL DEL PASIVO",
    "698":  "TOTAL DEL PATRIMONIO",
    "699":  "TOTAL PASIVO + PATRIMONIO",
    # Estado de Resultados — Ingresos
    "1005": "TOTAL INGRESOS DE ACTIVIDADES ORDINARIAS",
    "1045": "TOTAL INGRESOS NO OPERACIONALES",
    "6999": "TOTAL INGRESOS",
    # Estado de Resultados — Costos y Gastos
    "7991": "TOTAL COSTOS OPERACIONALES",
    "7992": "TOTAL GASTOS",
    "7999": "TOTAL COSTOS Y GASTOS",
}


# ─────────────────────────────────────────────────────────────────────────────
# Reglas estructurales: las 4 estructuras del código DEBEN contener cada total
# ─────────────────────────────────────────────────────────────────────────────

def test_regla_todos_los_totales_estan_en_a1_casilleros_ordered():
    """Si un TOTAL no está en A1_CASILLEROS_ORDERED, el filler nunca lo
    procesa → la fila correspondiente queda en blanco en el A1 y la
    VERIFICACIÓN dice 'F-101 NO DECLARÓ'."""
    a1_set = {c for c, _ in A1_CASILLEROS_ORDERED}
    faltan = [cas for cas in TOTALES_ESPERADOS if cas not in a1_set]
    assert not faltan, (
        f"Casilleros TOTAL faltantes en A1_CASILLEROS_ORDERED: {faltan}. "
        f"Agregarlos a backend/app/ict/cell_maps/a1.py para que el filler "
        f"los procese."
    )


def test_regla_todos_los_totales_estan_en_parser_f101():
    """Si un TOTAL no está en ALL_F101_CASILLEROS, el parser nunca extrae
    su valor del PDF → la columna 'F-101 declarado' sale vacía."""
    faltan = [cas for cas in TOTALES_ESPERADOS if cas not in ALL_F101_CASILLEROS]
    assert not faltan, (
        f"Casilleros TOTAL faltantes en ALL_F101_CASILLEROS: {faltan}. "
        f"Agregarlos a backend/app/ict/parsers/f101_pdf.py."
    )


def test_regla_todos_los_totales_estan_marcados_como_total_en_filler():
    """Si un TOTAL no está en TOTAL_CASILLEROS, el filler no le aplica
    formato de TOTAL ni le genera la fórmula F=SUM(...)."""
    faltan = [cas for cas in TOTALES_ESPERADOS if cas not in A1Filler.TOTAL_CASILLEROS]
    assert not faltan, (
        f"Casilleros TOTAL faltantes en A1Filler.TOTAL_CASILLEROS: {faltan}. "
        f"Agregarlos a backend/app/ict/fillers/a1_mapeo.py."
    )


def test_regla_todos_los_totales_tienen_formula_F_definida():
    """Cada TOTAL debe estar en PRIMARY_TOTAL_BLOCKS (suma directa de cuentas
    del bloque) o en COMPOSITE_TOTALS (suma de dos sub-totales). Si no
    está en ninguno, su F<row> queda sin fórmula → 'Balance contable'
    sale vacío en VERIFICACIÓN."""
    primary = set(A1Filler.PRIMARY_TOTAL_BLOCKS.keys())
    composite = set(A1Filler.COMPOSITE_TOTALS.keys())
    cubiertos = primary | composite
    faltan = [cas for cas in TOTALES_ESPERADOS if cas not in cubiertos]
    assert not faltan, (
        f"Casilleros TOTAL sin fórmula SUM en F: {faltan}. "
        f"Agregar cada uno a A1Filler.PRIMARY_TOTAL_BLOCKS (si es primario, "
        f"suma cuentas) o a A1Filler.COMPOSITE_TOTALS (si suma dos "
        f"sub-totales). Sin esto la columna 'Balance contable' en "
        f"VERIFICACIÓN sale vacía."
    )


def test_regla_composite_totals_solo_referencian_subtotales_existentes():
    """Cada COMPOSITE_TOTALS apunta a 2 sub-totales — ambos deben existir
    en TOTAL_CASILLEROS, o la fórmula =F<sub1>+F<sub2> apuntará a celdas
    sin valor."""
    for cas_total, (sub1, sub2) in A1Filler.COMPOSITE_TOTALS.items():
        assert sub1 in A1Filler.TOTAL_CASILLEROS, (
            f"COMPOSITE_TOTALS[{cas_total}] referencia sub1={sub1} pero "
            f"ese sub-total no está en TOTAL_CASILLEROS"
        )
        assert sub2 in A1Filler.TOTAL_CASILLEROS, (
            f"COMPOSITE_TOTALS[{cas_total}] referencia sub2={sub2} pero "
            f"ese sub-total no está en TOTAL_CASILLEROS"
        )


def test_regla_primary_totals_tienen_block_first_cas():
    """Cada PRIMARY_TOTAL_BLOCKS asigna su TOTAL a un bloque (ACT_CORR,
    PAS_CORR, etc.); ese bloque DEBE tener un BLOCK_FIRST_CAS para que
    el filler sepa dónde empieza el rango SUM(F<inicio>:F<row-1>)."""
    for cas_total, bloque_id in A1Filler.PRIMARY_TOTAL_BLOCKS.items():
        assert bloque_id in A1Filler.BLOCK_FIRST_CAS, (
            f"PRIMARY_TOTAL_BLOCKS[{cas_total}]={bloque_id} pero "
            f"BLOCK_FIRST_CAS no tiene ese bloque definido. "
            f"Sin esto el filler no sabe desde qué fila iniciar SUM(F:F)."
        )


# ─────────────────────────────────────────────────────────────────────────────
# REGLA E2E: al generar Excel, cada TOTAL DEBE tener fórmula en C y en F
# ─────────────────────────────────────────────────────────────────────────────

def _build_min_f101_with_totals() -> dict:
    """F-101 mínimo con TODOS los totales esperados, para que el filler
    encuentre algo que escribir y podamos verificar la salida."""
    return {cas: 1000.0 + i * 100 for i, cas in enumerate(TOTALES_ESPERADOS)}


def test_regla_e2e_a1_filler_escribe_formula_C_en_cada_total():
    """Al correr el filler con F-101 que declara cada TOTAL, la columna C
    debe quedar con fórmula referencial ='DATOS F-101'!Cxx o con valor."""
    wb = load_template()
    reset_trace()
    sess = {"razon_social": "X", "ruc": "1", "ejercicio_fiscal": "2025", "numero_adhesivo": ""}
    f101 = _build_min_f101_with_totals()
    A1Filler().fill(wb, sess, {"f101": f101, "balance_mapeado": []})

    ws = wb[A1_SHEET]
    # Recorrer A1 buscando cada cas TOTAL y verificar que C tenga contenido
    faltantes = []
    for r in range(A1_FIRST_DATA_ROW, ws.max_row + 1):
        cas_val = ws[f"A{r}"].value
        if cas_val is None:
            continue
        cas = str(cas_val).strip()
        if cas in TOTALES_ESPERADOS:
            c_val = ws[f"C{r}"].value
            if c_val is None or c_val == "":
                faltantes.append((cas, r, "C vacía"))

    assert not faltantes, (
        f"TOTAL sin valor en columna C (valor declarado F-101): {faltantes}. "
        f"Esto causa 'F-101 declarado' vacío en VERIFICACIÓN A1."
    )


def test_regla_e2e_a1_filler_escribe_formula_F_en_cada_total():
    """Al correr el filler, la columna F (saldo contable) de cada TOTAL
    debe tener una fórmula SUM(F<a>:F<b>) o =F<sub1>+F<sub2>. Si queda
    vacía, la columna 'Balance contable' en VERIFICACIÓN se ve mal."""
    wb = load_template()
    reset_trace()
    sess = {"razon_social": "X", "ruc": "1", "ejercicio_fiscal": "2025", "numero_adhesivo": ""}
    f101 = _build_min_f101_with_totals()
    A1Filler().fill(wb, sess, {"f101": f101, "balance_mapeado": []})

    ws = wb[A1_SHEET]
    sin_formula_f = []
    for r in range(A1_FIRST_DATA_ROW, ws.max_row + 1):
        cas_val = ws[f"A{r}"].value
        if cas_val is None:
            continue
        cas = str(cas_val).strip()
        if cas not in TOTALES_ESPERADOS:
            continue
        f_val = ws[f"F{r}"].value
        # Debe ser una fórmula =SUM(...) o =Fx+Fy
        if not (isinstance(f_val, str) and f_val.startswith("=")):
            sin_formula_f.append((cas, r, f"F={f_val!r}"))

    assert not sin_formula_f, (
        f"TOTAL sin fórmula en columna F (saldo contable): {sin_formula_f}. "
        f"Cada TOTAL debe tener fórmula SUM o suma de sub-totales en F<row>."
    )
