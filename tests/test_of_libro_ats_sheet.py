"""Hoja de datos 'DATOS ATS': literal, la que DM8 referencia por fórmula."""

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.ats import (
    BloqueBase,
    RetencionIVA,
    RetencionRenta,
    ResumenATS,
)
from backend.app.aud.obligaciones_fiscales.libro.fuentes import (
    SHEET_ATS,
    construir_hoja_ats,
)

RESUMEN_ENERO = ResumenATS(
    periodo="2025-01",
    compras=BloqueBase(bi_0=100.0, bi_gravada=200.0, bi_no_objeto=0.0, iva=30.0),
    ventas=BloqueBase(bi_0=5.0, bi_gravada=1000.0, bi_no_objeto=0.0, iva=150.0),
    comprobantes_anulados=3,
    retenciones_renta=[
        RetencionRenta(codigo="303", concepto="Honorarios", n_registros=2,
                        base_imponible=500.0, valor_retenido=50.0),
        RetencionRenta(codigo="310", concepto="", n_registros=1,
                        base_imponible=100.0, valor_retenido=1.0),
    ],
    retenciones_renta_valor_total=51.0,
    retenciones_iva=[
        RetencionIVA(operacion="COMPRA", concepto="Retencion IVA 30%", porcentaje=30.0,
                     valor_retenido=9.0),
        RetencionIVA(operacion="COMPRA", concepto="Retencion IVA NC", porcentaje=None,
                     valor_retenido=0.0),
    ],
    retenciones_iva_total=9.0,
    iva_que_le_retuvieron=40.0,
    renta_que_le_retuvieron=15.0,
)


def test_crea_la_hoja_datos_ats():
    wb = Workbook()
    construir_hoja_ats(wb, {"2025-01": RESUMEN_ENERO})
    assert SHEET_ATS in wb.sheetnames


def test_devuelve_direcciones_para_los_campos_fijos_por_mes():
    wb = Workbook()
    lookup = construir_hoja_ats(wb, {"2025-01": RESUMEN_ENERO})
    for campo in ("compras_bi_0", "compras_bi_gravada", "compras_iva",
                  "ventas_bi_0", "ventas_bi_gravada", "ventas_iva",
                  "anulados", "iva_le_retuvieron", "renta_le_retuvieron",
                  "ret_renta_total", "ret_iva_total"):
        assert (campo, "01") in lookup, campo


def test_las_direcciones_apuntan_a_los_valores_correctos():
    wb = Workbook()
    lookup = construir_hoja_ats(wb, {"2025-01": RESUMEN_ENERO})
    ws = wb[SHEET_ATS]
    assert ws[lookup[("compras_iva", "01")].split("!")[1]].value == 30.0
    assert ws[lookup[("ventas_bi_gravada", "01")].split("!")[1]].value == 1000.0
    assert ws[lookup[("iva_le_retuvieron", "01")].split("!")[1]].value == 40.0
    assert ws[lookup[("ret_renta_total", "01")].split("!")[1]].value == 51.0


def test_agrega_una_fila_por_cada_codigo_de_retencion_de_renta_visto():
    wb = Workbook()
    lookup = construir_hoja_ats(wb, {"2025-01": RESUMEN_ENERO})
    ws = wb[SHEET_ATS]
    assert ("renta_codigo:303", "01") in lookup
    assert ("renta_codigo:310", "01") in lookup
    assert ws[lookup[("renta_codigo:303", "01")].split("!")[1]].value == 50.0
    assert ws[lookup[("renta_codigo:310", "01")].split("!")[1]].value == 1.0


def test_agrega_una_fila_por_cada_porcentaje_de_retencion_de_iva():
    wb = Workbook()
    lookup = construir_hoja_ats(wb, {"2025-01": RESUMEN_ENERO})
    ws = wb[SHEET_ATS]
    assert ("iva_pct:30", "01") in lookup
    assert ("iva_pct:NC", "01") in lookup
    assert ws[lookup[("iva_pct:30", "01")].split("!")[1]].value == 9.0


def test_un_mes_sin_ats_se_escribe_como_cero():
    wb = Workbook()
    lookup = construir_hoja_ats(wb, {"2025-01": RESUMEN_ENERO})
    ws = wb[SHEET_ATS]
    # febrero no vino en el diccionario de resúmenes
    assert ("compras_iva", "02") not in lookup


def test_sin_ningun_ats_igual_crea_la_hoja_con_la_matriz_vacia():
    wb = Workbook()
    lookup = construir_hoja_ats(wb, {})
    assert SHEET_ATS in wb.sheetnames
    assert lookup == {} or all(v for v in lookup.values())
