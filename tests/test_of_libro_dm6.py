"""DM6 IVA: la conciliacion grande, 29 columnas {1}..{29} + AF, una fila por mes."""

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.bloques import MESES, NOMBRES_MES
from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm6_iva import (
    SHEET_DM6, build_dm6, letra_columna,
)

PERIODOS = [f"2025-{m:02d}" for m in range(1, 13)]

# --- direcciones ficticias (todas de fuentes ficticias, repo publico) ---
DIR_DM5 = {}
DIR_DM4 = {}
DIR_MAYORES = {}
for i, mes in enumerate(MESES, start=1):
    DIR_DM5[("ventas_libros", mes)] = f"'DM5 Ventas'!B{i}"
    DIR_DM5[("ventas_0_libros", mes)] = f"'DM5 Ventas'!C{i}"
    DIR_DM5[("total_declarado", mes)] = f"'DM5 Ventas'!D{i}"
    DIR_DM4[("base", mes)] = f"'DM4 Compras'!E{i}"
    DIR_MAYORES[("IVA_RETENIDO", mes)] = f"'Mayores homologados'!F{i}"

CASILLEROS = ["480", "424", "483", "605", "606", "609", "615", "617", "699", "413", "417", "418"]
DIR_F104 = {
    (f"2025-{m:02d}", cas): f"'DATOS F-104'!C{m}{cas}"
    for m in range(1, 13) for cas in CASILLEROS
}


def _cedula(**kw):
    wb = Workbook()
    datos = dict(dir_dm5=DIR_DM5, dir_dm4=DIR_DM4, dir_mayores=DIR_MAYORES,
                 dir_f104=DIR_F104, periodos=PERIODOS, cliente="C", periodo="2025")
    datos.update(kw)
    build_dm6(wb, **datos)
    return wb[SHEET_DM6]


def _fila_de(ws, nombre_mes: str) -> int:
    return next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == nombre_mes)


def test_lleva_el_encabezado_de_cedula_con_su_referencia():
    ws = _cedula()
    valores = [ws.cell(r, c).value for r in range(1, 11) for c in range(1, 6)]
    assert "OBLIGACIONES FISCALES" in valores
    assert "DM6" in valores


def test_tiene_una_fila_por_cada_uno_de_los_doce_meses():
    ws = _cedula()
    for nombre in NOMBRES_MES:
        _fila_de(ws, nombre)  # no lanza StopIteration


def test_no_lleva_filas_de_cuentas_solo_columnas_por_mes():
    ws = _cedula()
    etiquetas_col2 = {ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
    # nunca aparece "Según libros" ni "Cuenta": eso es de DM4/DM5, no de DM6.
    assert "Según libros" not in etiquetas_col2
    assert "Cuenta" not in etiquetas_col2


# --- columnas que vienen de fuera ---

def test_columna_b_viene_de_dm5_ventas_libros():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    b = letra_columna(1)
    assert ws[f"{b}{fila}"].value == "='DM5 Ventas'!B1"


def test_columna_c_viene_de_dm5_ventas_0_libros():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    c = letra_columna(2)
    assert ws[f"{c}{fila}"].value == "='DM5 Ventas'!C1"


def test_columna_n_viene_de_dm4_base():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    n = letra_columna(13)
    assert ws[f"{n}{fila}"].value == "='DM4 Compras'!E1"


def test_columna_y_viene_de_dm5_total_declarado():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    y = letra_columna(24)
    assert ws[f"{y}{fila}"].value == "='DM5 Ventas'!D1"


def test_columna_f_viene_del_casillero_480():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    f = letra_columna(5)
    assert ws[f"{f}{fila}"].value == "='DATOS F-104'!C1480"


def test_columna_i_viene_del_casillero_424():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    i_ = letra_columna(8)
    assert ws[f"{i_}{fila}"].value == "='DATOS F-104'!C1424"


def test_columna_d_viene_del_casillero_413():
    """{3} D: ventas netas gravadas con tarifa 0% que NO dan derecho a
    crédito tributario. Sale del casillero 413 del F-104, no es un cero
    sin fuente definida: si queda en 0 literal, la columna AB {27} =
    Y-B-C-D-E da un descuadre falso."""
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    d = letra_columna(3)
    assert ws[f"{d}{fila}"].value == "='DATOS F-104'!C1413"


def test_columna_e_viene_de_los_casilleros_417_mas_418():
    """{4} E: exportaciones de bienes y servicios. Sale de 417 + 418 del
    F-104, no de un cero sin fuente definida."""
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    e = letra_columna(4)
    assert ws[f"{e}{fila}"].value == "='DATOS F-104'!C1417+'DATOS F-104'!C1418"


def test_columna_u_viene_de_los_libros_de_iva_retenido():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    u = letra_columna(20)
    assert ws[f"{u}{fila}"].value == "='Mayores homologados'!F1"


def test_columna_z_es_la_suma_de_los_casilleros_615_y_617():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    z = letra_columna(25)
    valor = ws[f"{z}{fila}"].value
    assert valor == "='DATOS F-104'!C1615+'DATOS F-104'!C1617"


def test_columna_aa_viene_del_casillero_699():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    aa = letra_columna(26)
    assert ws[f"{aa}{fila}"].value == "='DATOS F-104'!C1699"


def test_columna_af_viene_del_casillero_609():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    af = "AF"
    assert ws[f"{af}{fila}"].value == "='DATOS F-104'!C1609"


def test_la_tarifa_por_defecto_es_quince_por_ciento():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    g = letra_columna(6)
    assert ws[f"{g}{fila}"].value == 0.15


def test_una_tarifa_distinta_en_un_mes_se_respeta():
    ws = _cedula(tarifas={"01": 0.12})
    fila = _fila_de(ws, "Enero")
    g = letra_columna(6)
    assert ws[f"{g}{fila}"].value == 0.12


# --- formulas derivadas ---

def test_columna_h_es_b_por_g():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    b, g, h = letra_columna(1), letra_columna(6), letra_columna(7)
    assert ws[f"{h}{fila}"].value == f"={b}{fila}*{g}{fila}"


def test_columna_k_es_f_por_g_mas_i():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    f, g, i_, k = letra_columna(5), letra_columna(6), letra_columna(8), letra_columna(10)
    assert ws[f"{k}{fila}"].value == f"={f}{fila}*{g}{fila}+{i_}{fila}"


def test_columna_l_es_h_mas_i_menos_k():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    h, i_, k, l = letra_columna(7), letra_columna(8), letra_columna(10), letra_columna(11)
    assert ws[f"{l}{fila}"].value == f"=({h}{fila}+{i_}{fila})-{k}{fila}"


def test_columna_m_es_j_mas_k():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    j, k, m = letra_columna(9), letra_columna(10), letra_columna(12)
    assert ws[f"{m}{fila}"].value == f"={j}{fila}+{k}{fila}"


def test_columna_o_es_n_por_g():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    n, g, o = letra_columna(13), letra_columna(6), letra_columna(14)
    assert ws[f"{o}{fila}"].value == f"={n}{fila}*{g}{fila}"


def test_columna_r_es_o_mas_p_por_q():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    o, p, q, r = letra_columna(14), letra_columna(15), letra_columna(16), letra_columna(17)
    assert ws[f"{r}{fila}"].value == f"=({o}{fila}+{p}{fila})*{q}{fila}"


def test_columna_ab_es_y_menos_b_c_d_e():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    y, b, c, d, e, ab = (letra_columna(24), letra_columna(1), letra_columna(2),
                         letra_columna(3), letra_columna(4), letra_columna(27))
    assert ws[f"{ab}{fila}"].value == f"={y}{fila}-{b}{fila}-{c}{fila}-{d}{fila}-{e}{fila}"


def test_columna_ac_es_z_menos_w():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    z, w, ac = letra_columna(25), letra_columna(22), letra_columna(28)
    assert ws[f"{ac}{fila}"].value == f"={z}{fila}-{w}{fila}"


def test_columna_ad_es_aa_menos_x():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    aa, x, ad = letra_columna(26), letra_columna(23), letra_columna(29)
    assert ws[f"{ad}{fila}"].value == f"={aa}{fila}-{x}{fila}"


def test_columna_w_es_el_valor_absoluto_del_negativo_de_m_r_s_t_u_mas_v():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    m, r, s, t, u, v, w = (letra_columna(12), letra_columna(17), letra_columna(18),
                           letra_columna(19), letra_columna(20), letra_columna(21),
                           letra_columna(22))
    expr = f"({m}{fila}-{r}{fila}-{s}{fila}-{t}{fila}-{u}{fila}+{v}{fila})"
    assert ws[f"{w}{fila}"].value == f"=ABS(IF({expr}<0,{expr},0))"


def test_columna_x_es_el_positivo_de_m_r_s_t_u_mas_v():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    m, r, s, t, u, v, x = (letra_columna(12), letra_columna(17), letra_columna(18),
                           letra_columna(19), letra_columna(20), letra_columna(21),
                           letra_columna(23))
    expr = f"({m}{fila}-{r}{fila}-{s}{fila}-{t}{fila}-{u}{fila}+{v}{fila})"
    assert ws[f"{x}{fila}"].value == f"=IF({expr}>0,{expr},0)"


def test_los_parentesis_de_las_formulas_derivadas_estan_balanceados():
    ws = _cedula()
    fila = _fila_de(ws, "Enero")
    for n in range(1, 30):
        col = letra_columna(n)
        valor = ws[f"{col}{fila}"].value
        if isinstance(valor, str) and valor.startswith("="):
            assert valor.count("(") == valor.count(")"), (col, valor)


# --- las tres pruebas irrenunciables ---

def test_a_enero_usa_los_casilleros_483_y_605_606_sin_referencia_circular():
    ws = _cedula()
    fila_enero = _fila_de(ws, "Enero")
    j, t, l, w = (letra_columna(9), letra_columna(19), letra_columna(11), letra_columna(22))

    valor_j = ws[f"{j}{fila_enero}"].value
    valor_t = ws[f"{t}{fila_enero}"].value

    assert valor_j == "='DATOS F-104'!C1483"
    assert valor_t == "='DATOS F-104'!C1605+'DATOS F-104'!C1606"
    # nunca debe encadenar contra su propia fila (eso crearia un ciclo: L
    # depende de K, y J encadenado contra L de la MISMA fila dependeria de
    # si mismo a traves de M).
    assert f"{l}{fila_enero}" not in str(valor_j)
    assert f"{w}{fila_enero}" not in str(valor_t)


def test_b_febrero_encadena_contra_la_fila_de_enero():
    ws = _cedula()
    fila_enero = _fila_de(ws, "Enero")
    fila_febrero = _fila_de(ws, "Febrero")
    assert fila_febrero == fila_enero + 1

    j, t, l, w = (letra_columna(9), letra_columna(19), letra_columna(11), letra_columna(22))
    assert ws[f"{j}{fila_febrero}"].value == f"={l}{fila_enero}"
    assert ws[f"{t}{fila_febrero}"].value == f"={w}{fila_enero}"


def test_c_marzo_tambien_encadena_contra_febrero_no_contra_enero():
    ws = _cedula()
    fila_febrero = _fila_de(ws, "Febrero")
    fila_marzo = _fila_de(ws, "Marzo")
    assert fila_marzo == fila_febrero + 1

    j, l = letra_columna(9), letra_columna(11)
    assert ws[f"{j}{fila_marzo}"].value == f"={l}{fila_febrero}"


def test_q_no_divide_por_cero_cuando_el_mes_no_tiene_ventas():
    """DIR_DM5 no trae datos para 'Abril': B, C, D y E quedan en 0. La
    formula debe seguir siendo el guard IF(SUM(...)=0, 0, ...), nunca una
    division directa que reventaria con /0."""
    dir_dm5_sin_abril = {k: v for k, v in DIR_DM5.items() if k[1] != "04"}
    ws = _cedula(dir_dm5=dir_dm5_sin_abril)
    fila_abril = _fila_de(ws, "Abril")
    b, c, d, e, q = (letra_columna(1), letra_columna(2), letra_columna(3),
                     letra_columna(4), letra_columna(16))
    assert ws[f"{b}{fila_abril}"].value == 0
    assert ws[f"{c}{fila_abril}"].value == 0
    valor_q = ws[f"{q}{fila_abril}"].value
    esperado = (f"=IF(SUM({b}{fila_abril}:{e}{fila_abril})=0,0,"
                f"({b}{fila_abril}+{c}{fila_abril}+{e}{fila_abril})/"
                f"({b}{fila_abril}+{c}{fila_abril}+{d}{fila_abril}+{e}{fila_abril}))")
    assert valor_q == esperado
