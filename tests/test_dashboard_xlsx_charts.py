"""Tests del dashboard .xlsx con GRÁFICOS NATIVOS de Excel (openpyxl.chart).

Verifica que build_dashboard_workbook:
  - produce un libro que openpyxl puede RECARGAR (no da "reparar" en Excel),
  - contiene la hoja "Datos",
  - genera al menos 3 gráficos NATIVOS,
  - escribe los valores correctos en celdas de datos conocidas.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from backend.app.tax.planificacion_utilidades import exporter


def _data_sintetico() -> dict:
    """Modelo D con 2 períodos (2024, 2025). Costos/pasivos positivos: la
    función debe tomar valor absoluto donde corresponde."""
    return {
        "ventas": [1000.0, 1200.0],
        "otrosIng": [50.0, 60.0],
        "otrosIngFin": [10.0, 12.0],
        "costo": [600.0, 700.0],       # positivo -> abs
        "gAdmin": [150.0, 160.0],
        "gFin": [20.0, 25.0],
        "partTrab": [30.0, 35.0],
        "irCausado": [40.0, 45.0],
        "impDif": [5.0, 6.0],
        "efectivo": [300.0, 350.0],
        "cxc": [200.0, 250.0],
        "inventario": [400.0, 420.0],
        "ppe": [800.0, 850.0],
        "capital": [500.0, 500.0],
        "resAcum": [700.0, 900.0],
    }


def _reload(raw: bytes):
    return load_workbook(io.BytesIO(raw))


def test_no_repara_y_tiene_hoja_datos():
    raw = exporter.build_dashboard_workbook(
        _data_sintetico(), ["2024", "2025"], [12, 12], "Test SA", "combo"
    )
    wb = _reload(raw)  # si el libro estuviera corrupto, esto lanzaría
    assert "Datos" in wb.sheetnames
    assert "Dashboard" in wb.sheetnames


def test_al_menos_3_graficos_nativos():
    raw = exporter.build_dashboard_workbook(
        _data_sintetico(), ["2024", "2025"], [12, 12], "Test SA", "combo"
    )
    wb = _reload(raw)
    total = len(wb["Dashboard"]._charts) + len(wb["Datos"]._charts)
    assert total >= 3, f"esperaba >= 3 gráficos, obtuve {total}"


def test_valor_de_celda_conocida_ingresos_2025():
    """Ingresos ordinarios 2025 (columna del 2º período) = ventas[1] = 1200."""
    raw = exporter.build_dashboard_workbook(
        _data_sintetico(), ["2024", "2025"], [12, 12], "Test SA", "combo"
    )
    wb = _reload(raw)
    ws = wb["Datos"]
    # Buscar la fila "Ingresos ordinarios" en la columna A.
    fila = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Ingresos ordinarios":
            fila = r
            break
    assert fila is not None, "no se encontró la fila 'Ingresos ordinarios'"
    # Col 2 = período 2024, col 3 = período 2025.
    assert ws.cell(row=fila, column=3).value == 1200.0


def test_utilidad_neta_calculada():
    """Utilidad neta 2024 = 1000 - 600 - 150 - 20 + 50 + 10 - 30 - 40 - 5 = 215."""
    raw = exporter.build_dashboard_workbook(
        _data_sintetico(), ["2024", "2025"], [12, 12], "Test SA", "combo"
    )
    wb = _reload(raw)
    ws = wb["Datos"]
    fila = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Utilidad neta":
            fila = r
            break
    assert fila is not None
    assert ws.cell(row=fila, column=2).value == pytest.approx(215.0)


@pytest.mark.parametrize("estilo", ["barras", "lineas", "area", "combo"])
def test_todos_los_estilos_recargan(estilo):
    raw = exporter.build_dashboard_workbook(
        _data_sintetico(), ["2024", "2025"], [12, 12], "Test SA", estilo
    )
    wb = _reload(raw)
    total = len(wb["Dashboard"]._charts)
    assert total >= 3, f"estilo {estilo}: {total} gráficos"


# ============ Arquitectura de datos (2ª pasada Big Data) ============

def test_zip_integro_sin_reparar():
    """El .xlsx debe ser un zip íntegro (Excel no debe pedir 'reparar')."""
    import zipfile
    raw = exporter.build_dashboard_workbook(
        _data_sintetico(), ["2024", "2025"], [12, 12], "Test SA", "combo"
    )
    assert zipfile.ZipFile(io.BytesIO(raw)).testzip() is None


def test_existe_hoja_datos_powerbi_formato_largo():
    """Hoja 'Datos_PowerBI' en formato largo: Estado|Concepto|Periodo|Valor."""
    raw = exporter.build_dashboard_workbook(
        _data_sintetico(), ["2024", "2025"], [12, 12], "Test SA", "combo"
    )
    wb = _reload(raw)
    assert "Datos_PowerBI" in wb.sheetnames
    wp = wb["Datos_PowerBI"]
    assert [wp.cell(1, c).value for c in range(1, 5)] == [
        "Estado", "Concepto", "Periodo", "Valor"]
    # 12 conceptos (5 ER + 7 Balance) x 2 períodos = 24 filas de datos.
    filas = sum(1 for r in range(2, wp.max_row + 1) if wp.cell(r, 1).value)
    assert filas >= 24, f"esperaba >= 24 filas largas, obtuve {filas}"


def test_existe_al_menos_una_tabla_estructurada():
    """Debe existir al menos 1 Table (Ctrl+T). tblDatosLargo es obligatoria."""
    raw = exporter.build_dashboard_workbook(
        _data_sintetico(), ["2024", "2025"], [12, 12], "Test SA", "combo"
    )
    wb = _reload(raw)
    nombres = set()
    for sh in wb.sheetnames:
        nombres |= set(wb[sh].tables.keys())
    assert len(nombres) >= 1, f"no hay Tables; encontradas: {nombres}"
    assert "tblDatosLargo" in nombres, f"falta tblDatosLargo; hay {nombres}"


def test_celda_auditoria_cuadre_presente():
    """La hoja Dashboard debe tener la fórmula de cuadre contable."""
    raw = exporter.build_dashboard_workbook(
        _data_sintetico(), ["2024", "2025"], [12, 12], "Test SA", "combo"
    )
    wb = _reload(raw)
    d = wb["Dashboard"]
    encontrada = False
    for row in d.iter_rows():
        for c in row:
            if (isinstance(c.value, str) and c.value.startswith("=IF")
                    and "Cuadra contable" in c.value):
                encontrada = True
    assert encontrada, "no se encontró la celda de auditoría de cuadre"
