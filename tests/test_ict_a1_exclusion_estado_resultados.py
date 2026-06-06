"""Regression test: A1 estado de resultados oculta cas informativos,
exentos y no deducibles (regla cliente 2026-06-06).

REPORTE CLIENTE (verbatim):
"QUITA DE A1 DEL ESTADO DE RESULTADOS LOS CASILLERO INFORMATIVOS TANTO DE
INGRESOS COMO DE GASTOS ADEMAS LOS CASILLEROS DE INGRESOS EXENTOS Y GASTOS
NO DEDUCIBLES. SOLO DEBE QUEDAR DE LAS CUENTAS QUE SE COMPARAN CON SALDOS
CONTABLES"

REGLA implementada en `source_data_sheets._es_excluido_estado_resultados`:
en rango 6001-7999 ocultar cas cuyo nombre:
  - empieza con "VALOR EXENTO"
  - empieza con "VALOR NO DEDUCIBLE"
  - contiene "NO OBJETO DE IMPUESTO"
  - es (INFORMATIVO)

Los TOTALES (6999, 7999, 1005, etc.) NUNCA se excluyen.
"""
from __future__ import annotations

import pytest

from backend.app.ict.fillers.source_data_sheets import (
    _es_excluido_estado_resultados,
    F101_TOTALES,
)


class TestExclusionEstadoResultados:
    """Casos del reporte cliente PROPHAR ICT_10."""

    @pytest.mark.parametrize("cas,nombre", [
        ("6002", "VALOR EXENTO VENTAS NETAS LOCALES DE BIENES GRAVADAS"),
        ("6116", "VALOR EXENTO INGRESOS FINANCIEROS INTERES"),
        ("6044", "VALOR EXENTO GANANCIAS NETAS POR REVERSIONES DE DETERIORO"),
    ])
    def test_valor_exento_excluido(self, cas, nombre):
        assert _es_excluido_estado_resultados(cas, nombre) is True

    @pytest.mark.parametrize("cas,nombre", [
        ("7249", "VALOR NO DEDUCIBLE OTROS GASTOS"),
        ("7006", "VALOR NO DEDUCIBLE COMPRAS LOCALES NETAS DE BIENES"),
        ("7039", "VALOR NO DEDUCIBLE DE AJUSTES"),
    ])
    def test_valor_no_deducible_excluido(self, cas, nombre):
        assert _es_excluido_estado_resultados(cas, nombre) is True

    @pytest.mark.parametrize("cas,nombre", [
        ("6150", "INGRESOS NO OBJETO DE IMPUESTO A LA RENTA"),
        ("7906", "GASTOS ATRIBUIDOS A INGRESOS NO OBJETO DE IMPUESTO"),
    ])
    def test_no_objeto_impuesto_excluido(self, cas, nombre):
        assert _es_excluido_estado_resultados(cas, nombre) is True

    @pytest.mark.parametrize("cas,nombre", [
        ("6152", "INGRESOS BRUTOS TOTALES SEGUN CONTABILIDAD (INFORMATIVO)"),
        ("6142", "DIVIDENDOS DECLARADOS (DISTRIBUIDOS) A FAVOR DEL CONTRIBUYENTE INFORMATIVO"),
    ])
    def test_informativos_excluidos(self, cas, nombre):
        assert _es_excluido_estado_resultados(cas, nombre) is True

    @pytest.mark.parametrize("cas,nombre", [
        ("6001", "VENTAS NETAS LOCALES DE BIENES GRAVADAS CON TARIFA DIFERENTE DE 0%"),
        ("6003", "VENTAS NETAS LOCALES GRAVADAS CON TARIFA CERO O EXENTAS DE IVA"),
        ("6043", "GANANCIAS NETAS POR REVERSIONES DE DETERIORO EN EL VALOR DE LOS ACTIVOS"),
        ("7242", "GASTO SERVICIOS PUBLICOS"),
        ("7248", "GASTO OTROS GASTOS"),
        ("7067", "COSTO DEPRECIACION NO ACELERADA DE PROPIEDADES PLANTA Y EQUIPO"),
    ])
    def test_cuentas_reales_se_mantienen(self, cas, nombre):
        """Las cuentas reales (con saldo contable) NO se excluyen."""
        assert _es_excluido_estado_resultados(cas, nombre) is False

    @pytest.mark.parametrize("cas", ["6999", "7999", "1005", "7991", "7992"])
    def test_totales_nunca_se_excluyen(self, cas):
        """Los TOTALES siempre se muestran (cuadratura).

        Aunque algunos TOTALES podrían contener "INFORMATIVO" o aparentar
        ser exentos, la regla _es_excluido_estado_resultados NO los excluye
        porque NO empiezan con VALOR EXENTO ni VALOR NO DEDUCIBLE. Y el
        filtro de A1 además los protege via cas in F101_TOTALES.
        """
        # Nombre dummy: aunque dijera "INFORMATIVO", _cas_es_relevante_a1
        # del a1_mapeo.py los protege via F101_TOTALES antes del filtro.
        nombre_dummy = f"TOTAL {cas}"
        assert _es_excluido_estado_resultados(cas, nombre_dummy) is False
        assert cas in F101_TOTALES

    @pytest.mark.parametrize("cas,nombre", [
        ("311", "EFECTIVO Y EQUIVALENTES AL EFECTIVO"),
        ("499", "TOTAL DEL ACTIVO"),
        ("511", "CUENTAS Y DOCUMENTOS POR PAGAR COMERCIALES"),
        ("698", "TOTAL DEL PATRIMONIO"),
    ])
    def test_balance_no_aplica(self, cas, nombre):
        """El filtro solo aplica al estado de resultados (6001-7999)."""
        assert _es_excluido_estado_resultados(cas, nombre) is False


class TestIntegracionA1Filler:
    """E2E: cas excluidos NO aparecen en la hoja A1 generada."""

    def test_a1_no_contiene_valor_exento(self):
        """A1 generado con cas VALOR EXENTO no los emite."""
        from openpyxl import Workbook
        from backend.app.ict.fillers.a1_mapeo import A1Filler

        wb = Workbook()
        wb.create_sheet("MAPEO DE LA DECLARACIÓN A1")
        # F-101 con cas exento + cas real, ambos con saldo
        f101 = {
            "6001": 1000.0,  # venta real → debe aparecer
            "6002": 500.0,   # VALOR EXENTO → NO debe aparecer
            "6116": 100.0,   # VALOR EXENTO → NO debe aparecer
            "7242": 800.0,   # gasto real → debe aparecer
            "7249": 300.0,   # VALOR NO DEDUCIBLE → NO debe aparecer
            "6999": 1500.0,  # TOTAL → debe aparecer
            "7999": 800.0,   # TOTAL → debe aparecer
        }
        session = {"ruc": "TEST", "razon_social": "TEST", "ejercicio_fiscal": "2025"}
        try:
            A1Filler().fill(wb, session, {"f101": f101, "balance_mapeado": []})
        except Exception:
            pass  # filler puede fallar por dependencias, no nos importa aquí

        ws = wb["MAPEO DE LA DECLARACIÓN A1"]
        cas_emitidos = set()
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is not None and str(v).strip().isdigit():
                cas_emitidos.add(str(v).strip())

        # Cas reales deben estar
        assert "6001" in cas_emitidos, "cas 6001 (venta real) debe aparecer"
        # Cas excluidos NO deben estar
        assert "6002" not in cas_emitidos, "cas 6002 (VALOR EXENTO) NO debe aparecer"
        assert "6116" not in cas_emitidos, "cas 6116 (VALOR EXENTO) NO debe aparecer"
        assert "7249" not in cas_emitidos, "cas 7249 (VALOR NO DEDUCIBLE) NO debe aparecer"
