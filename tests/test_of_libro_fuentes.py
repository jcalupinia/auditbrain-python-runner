"""Adaptación de los extractores SRI al formato de los builders de hojas."""

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales import file_storage
from backend.app.aud.obligaciones_fiscales.libro.fuentes import (
    a_periodos_anuales,
    construir_hojas_de_casilleros,
    leer_ats,
)


def test_convierte_los_meses_del_extractor_f104_a_periodos_anuales():
    entrada = {
        "01": {"periodo": "01/2025", "casilleros": {"429": 10.0}},
        "12": {"periodo": "12/2025", "casilleros": {"429": 20.0}},
    }
    assert a_periodos_anuales(entrada) == {
        "2025-01": {"casilleros": {"429": 10.0}},
        "2025-12": {"casilleros": {"429": 20.0}},
    }


def test_un_mes_sin_periodo_detectado_se_descarta():
    entrada = {"01": {"periodo": None, "casilleros": {"429": 10.0}}}
    assert a_periodos_anuales(entrada) == {}


def test_construye_las_dos_hojas_de_casilleros_y_devuelve_sus_direcciones():
    wb = Workbook()
    lookups = construir_hojas_de_casilleros(
        wb,
        f104_monthly={"2025-01": {"casilleros": {"429": 4341.16}}},
        f103_monthly={"2025-01": {"casilleros": {"499": 915.70}}},
    )
    assert "DATOS F-104" in wb.sheetnames
    assert "DATOS F-103" in wb.sheetnames
    assert ("2025-01", "429") in lookups["f104"]
    assert ("2025-01", "499") in lookups["f103"]


def test_la_direccion_devuelta_apunta_al_valor_correcto():
    wb = Workbook()
    lookups = construir_hojas_de_casilleros(
        wb,
        f104_monthly={"2025-01": {"casilleros": {"429": 4341.16}}},
        f103_monthly={},
    )
    addr = lookups["f104"][("2025-01", "429")]
    assert wb["DATOS F-104"][addr].value == 4341.16


def test_sin_pdfs_igual_se_crean_las_hojas_con_la_matriz_vacia():
    """El auditor debe ver qué casilleros se esperaban, aunque no haya datos."""
    wb = Workbook()
    lookups = construir_hojas_de_casilleros(wb, f104_monthly={}, f103_monthly={})
    assert "DATOS F-104" in wb.sheetnames
    assert lookups["f104"]


def test_leer_ats_sin_archivos_devuelve_diccionario_vacio(tmp_path, monkeypatch):
    monkeypatch.setattr(file_storage, "_root", lambda: tmp_path)
    job_dir = file_storage.create_job_dir(999)
    assert leer_ats(job_dir) == {}


def test_leer_ats_agrupa_por_periodo_los_archivos_subidos(tmp_path, monkeypatch):
    monkeypatch.setattr(file_storage, "_root", lambda: tmp_path)
    job_dir = file_storage.create_job_dir(998)
    texto_pdf_invalido = b"no es un pdf real"
    file_storage.save_input(job_dir, "ats", "anexo.xml", b"<xml/>")
    # No es un PDF real (pdfplumber no puede abrirlo): se registra pero no
    # aporta período, así que leer_ats() no debe reventar.
    file_storage.save_input(job_dir, "ats", "anexo.pdf", texto_pdf_invalido)
    resultado = leer_ats(job_dir)
    assert resultado == {}
