"""Constructor de bloques de cédula: la anatomía común de DM3..DM7."""

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.bloques import (
    MESES,
    escribir_encabezado_meses,
    fila_diferencia,
    fila_referencias,
    fila_suma_rango,
)


def _ws():
    return Workbook().active


def test_el_encabezado_pone_los_doce_meses_y_el_total():
    ws = _ws()
    escribir_encabezado_meses(ws, fila=13, titulo="IVA EN COMPRAS")
    assert ws.cell(13, 1).value == "IVA EN COMPRAS"
    assert ws.cell(13, 3).value == "Enero"
    assert ws.cell(13, 14).value == "Diciembre"
    assert ws.cell(13, 15).value == "Total"


def test_una_fila_de_referencias_escribe_una_formula_por_mes():
    ws = _ws()
    direcciones = {m: f"'Mayores homologados'!D{i}" for i, m in enumerate(MESES, 4)}
    fila_referencias(ws, fila=16, etiqueta="IVA EN COMPRAS", direcciones=direcciones)
    assert ws.cell(16, 2).value == "IVA EN COMPRAS"
    assert ws.cell(16, 3).value == "='Mayores homologados'!D4"
    assert ws.cell(16, 14).value == "='Mayores homologados'!D15"


def test_un_mes_sin_direccion_queda_en_cero_no_vacio():
    ws = _ws()
    fila_referencias(ws, fila=16, etiqueta="x", direcciones={"01": "'H'!A1"})
    assert ws.cell(16, 4).value == 0


def test_la_fila_de_suma_totaliza_el_rango_indicado():
    ws = _ws()
    fila_suma_rango(ws, fila=20, etiqueta="Según libros", desde=16, hasta=19)
    assert ws.cell(20, 3).value == "=SUM(C16:C19)"
    assert ws.cell(20, 15).value == "=SUM(O16:O19)"


def test_la_fila_de_diferencia_resta_dos_filas():
    ws = _ws()
    fila_diferencia(ws, fila=33, etiqueta="Diferencia", fila_libros=20, fila_declarado=31)
    assert ws.cell(33, 3).value == "=ROUND(C20-C31,2)"


def test_los_importes_llevan_formato_contable():
    ws = _ws()
    fila_suma_rango(ws, fila=20, etiqueta="Según libros", desde=16, hasta=19)
    assert ws.cell(20, 3).number_format == "#,##0.00"
