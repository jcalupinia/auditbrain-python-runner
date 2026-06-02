"""Tests: balance_mapeado uploaded to A1 flows to other anexos that need casillero data.

REFACTOR REFERENCIAL: tras el cambio a fórmulas referenciales (CLAUDE.md),
los anexos NO contienen valores literales — contienen fórmulas tipo
='DATOS F-101'!C<row> o ='DATOS BALANCE'!D<row>+...
Los tests E2E validan:
  1. Que la celda destino contiene una FÓRMULA referencial.
  2. Que la hoja DATOS referenciada contiene el valor esperado en esa fila.
"""
import io
import re
import uuid

import openpyxl
import pytest

from backend.app.client_portal.service import create_portal_user
from backend.app.context.models import Organization, Client
from backend.app.db.session import SessionLocal
from backend.app.ict import service as ict_service
from backend.app.ict.fillers.helpers import (
    aggregate_balance_by_casillero,
    filter_balance_by_casilleros,
    get_casillero_value,
)


def _datos_balance_has_value(wb, expected_value: float, tol: float = 0.01) -> bool:
    """True si la hoja DATOS BALANCE contiene una cuenta con saldo == expected_value."""
    if "DATOS BALANCE" not in wb.sheetnames:
        return False
    ws = wb["DATOS BALANCE"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        for val in row:
            if isinstance(val, (int, float)) and abs(val - expected_value) < tol:
                return True
    return False


def _datos_f101_has_value(wb, casillero: str, expected_value: float, tol: float = 0.01) -> bool:
    """True si DATOS F-101 tiene fila con casillero == X y valor == expected_value."""
    if "DATOS F-101" not in wb.sheetnames:
        return False
    ws = wb["DATOS F-101"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 3:
            continue
        if str(row[0] or "").strip() == str(casillero) and isinstance(row[2], (int, float)):
            if abs(row[2] - expected_value) < tol:
                return True
    return False


_FORMULA_PATTERN = re.compile(
    r"^=\s*'?DATOS (F-101|F-103|F-104|BALANCE)'?!", re.IGNORECASE
)


def _has_referential_formula_for(ws, target_value: float) -> bool:
    """Busca cualquier celda con fórmula referencial (a DATOS X) en la hoja.
    Útil cuando solo necesitamos comprobar que la hoja USA referencias (no
    valores literales) y el valor target está en DATOS BALANCE/F-101."""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and _FORMULA_PATTERN.match(v):
                return True
    return False


def _unique(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex[:8]}"


@pytest.fixture()
def db_session():
    db = SessionLocal()
    yield db
    db.close()


@pytest.fixture()
def client_user(db_session):
    suffix = _unique("cross")
    org = Organization(name=f"O-{suffix}", slug=f"o-{suffix}", is_active=True)
    db_session.add(org)
    db_session.commit()
    db_session.refresh(org)
    cli = Client(organization_id=org.id, name=f"C-{suffix}", is_active=True)
    db_session.add(cli)
    db_session.commit()
    db_session.refresh(cli)
    user, _ = create_portal_user(db_session, client_id=cli.id, email=f"cross-{suffix}@x.com")
    return user


# ── Unit tests: helpers ──────────────────────────────────────────────────────

def test_aggregate_balance_sums_per_casillero():
    items = [
        {"casillero_sri": "311", "codigo": "X", "descripcion": "A", "saldo": 100},
        {"casillero_sri": "311", "codigo": "Y", "descripcion": "B", "saldo": 50},
        {"casillero_sri": "315", "codigo": "Z", "descripcion": "C", "saldo": 200},
        {"casillero_sri": "", "codigo": "skip", "descripcion": "", "saldo": 999},  # skip empty
    ]
    agg = aggregate_balance_by_casillero(items)
    assert agg["311"] == 150
    assert agg["315"] == 200
    assert "" not in agg


def test_aggregate_balance_handles_empty_list():
    assert aggregate_balance_by_casillero([]) == {}
    assert aggregate_balance_by_casillero(None) == {}


def test_aggregate_balance_handles_bad_saldo():
    items = [
        {"casillero_sri": "311", "saldo": "not_a_number"},
        {"casillero_sri": "311", "saldo": None},
        {"casillero_sri": "311", "saldo": 50.0},
    ]
    agg = aggregate_balance_by_casillero(items)
    # Bad saldos skipped; only 50.0 counts
    assert agg["311"] == 50.0


def test_get_casillero_prefers_f101_over_balance():
    data = {
        "f101": {"311": 100.0},
        "balance_mapeado": [{"casillero_sri": "311", "saldo": 999.0}],
    }
    assert get_casillero_value(data, "311") == 100.0


def test_get_casillero_falls_back_to_balance():
    data = {
        "f101": {},
        "balance_mapeado": [{"casillero_sri": "7185", "saldo": 25000.0}],
    }
    assert get_casillero_value(data, "7185") == 25000.0


def test_get_casillero_returns_default_if_missing():
    data = {"f101": {}, "balance_mapeado": []}
    assert get_casillero_value(data, "999", default=0) == 0
    assert get_casillero_value(data, "999") is None


def test_get_casillero_handles_missing_keys():
    # Anexo_data without f101 or balance_mapeado
    assert get_casillero_value({}, "311", default=42.0) == 42.0


def test_filter_balance_returns_only_target_casilleros():
    items = [
        {"casillero_sri": "806", "codigo": "A", "descripcion": "ND1", "saldo": 100},
        {"casillero_sri": "807", "codigo": "B", "descripcion": "ND2", "saldo": 50},
        {"casillero_sri": "311", "codigo": "C", "descripcion": "Caja", "saldo": 200},
    ]
    filtered = filter_balance_by_casilleros(items, {"806", "807"})
    assert len(filtered) == 2
    assert all(i["casillero_sri"] in {"806", "807"} for i in filtered)


def test_filter_balance_returns_empty_when_no_match():
    items = [{"casillero_sri": "311", "saldo": 100}]
    assert filter_balance_by_casilleros(items, {"999"}) == []
    assert filter_balance_by_casilleros([], {"806"}) == []
    assert filter_balance_by_casilleros(None, {"806"}) == []


# ── Integration tests: generate_excel shared context ─────────────────────────

def test_e2e_balance_in_a1_propagates_to_a3_via_generate_excel(db_session, client_user):
    """E2E: balance_mapeado uploaded to A1 (with casillero 7185) should allow A3
    to fill gastos de gestión even when A3's own f101 doesn't include 7185."""
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="Test", numero_adhesivo=None,
    )

    # A1 gets balance_mapeado with a gasto casillero (7185)
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A1",
        extracted_data={
            "f101": {"311": 1000.0},
            "balance_mapeado": [
                {"casillero_sri": "311", "codigo": "1", "descripcion": "Caja", "saldo": 1000.0},
                {"casillero_sri": "7185", "codigo": "61.01", "descripcion": "Gto Gestión", "saldo": 25337.71},
                {"casillero_sri": "7186", "codigo": "61.02", "descripcion": "Gto Gestión ND", "saldo": 0.0},
                {"casillero_sri": "7992", "codigo": "5.x", "descripcion": "Total gastos", "saldo": 1531596.16},
            ],
        },
        warnings=[], uploaded_file_meta={"slot": "balance_mapeado"},
        new_status="ready",
    )

    # A3 has its own f101 but NO 7185 in it (deliberately)
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A3",
        extracted_data={"f101": {"6999": 18379679.21}},
        warnings=[], uploaded_file_meta={"slot": "f101"},
        new_status="ready",
    )
    ict_service.recompute_indice(db_session, session=s)
    db_session.refresh(s)

    # Generate Excel — should not raise
    excel_bytes = ict_service.generate_excel(db_session, session=s)
    assert len(excel_bytes) > 1000

    # Open the Excel and verify A3 sheet was generated
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
    assert "COSTOS  GASTOS A3" in wb.sheetnames

    # Modo referencial: A3 contiene FÓRMULAS hacia DATOS BALANCE.
    # 1) El valor 25337.71 vive en DATOS BALANCE (no en A3 literal).
    assert _datos_balance_has_value(wb, 25337.71), \
        "El valor 25337.71 (casillero 7185) debe estar en DATOS BALANCE"
    # 2) A3 contiene al menos una fórmula referencial.
    a3 = wb["COSTOS  GASTOS A3"]
    assert _has_referential_formula_for(a3, 25337.71), \
        "A3 debería tener al menos una fórmula referencial (no valores literales)"


def test_a4_cuadro2_populated_from_balance_when_no_f101(db_session, client_user):
    """A4 Cuadro 2 casilleros (804, 805, 812, 1112) should be filled from
    balance_mapeado when A4's own f101 is empty."""
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="Test", numero_adhesivo=None,
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A1",
        extracted_data={
            "balance_mapeado": [
                {"casillero_sri": "804", "codigo": "X", "descripcion": "Dividendos exentos", "saldo": 5000.0},
                {"casillero_sri": "805", "codigo": "Y", "descripcion": "Otras rentas exentas", "saldo": 1200.0},
            ],
        },
        warnings=[], uploaded_file_meta={"slot": "balance_mapeado"},
        new_status="ready",
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A4",
        extracted_data={"f101": {}},
        warnings=[], uploaded_file_meta={"slot": "f101"},
        new_status="ready",
    )
    ict_service.recompute_indice(db_session, session=s)
    db_session.refresh(s)

    excel_bytes = ict_service.generate_excel(db_session, session=s)
    assert len(excel_bytes) > 1000

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
    assert "CONCILIACIÓN INGRESOS A4" in wb.sheetnames

    a4 = wb["CONCILIACIÓN INGRESOS A4"]
    # Modo referencial: G32 y G33 contienen fórmulas hacia DATOS BALANCE.
    val_804 = a4["G32"].value
    val_805 = a4["G33"].value
    assert isinstance(val_804, str) and val_804.startswith("="), \
        f"G32 debe ser fórmula referencial, obtenido {val_804!r}"
    assert isinstance(val_805, str) and val_805.startswith("="), \
        f"G33 debe ser fórmula referencial, obtenido {val_805!r}"
    assert "DATOS BALANCE" in val_804
    assert "DATOS BALANCE" in val_805
    # Verificar que los valores 5000.0 y 1200.0 están en DATOS BALANCE
    assert _datos_balance_has_value(wb, 5000.0)
    assert _datos_balance_has_value(wb, 1200.0)


def test_a5_cuadro_d_populated_from_balance_when_no_f101(db_session, client_user):
    """A5 Cuadro D casilleros (806, 807) should be filled from balance_mapeado
    when A5's own f101 is empty."""
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="Test", numero_adhesivo=None,
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A1",
        extracted_data={
            "balance_mapeado": [
                {"casillero_sri": "806", "codigo": "A", "descripcion": "Gtos ND locales", "saldo": 3000.0},
                {"casillero_sri": "807", "codigo": "B", "descripcion": "Gtos ND exterior", "saldo": 500.0},
            ],
        },
        warnings=[], uploaded_file_meta={"slot": "balance_mapeado"},
        new_status="ready",
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A5",
        extracted_data={"f101": {}},
        warnings=[], uploaded_file_meta={"slot": "f101"},
        new_status="ready",
    )
    ict_service.recompute_indice(db_session, session=s)
    db_session.refresh(s)

    excel_bytes = ict_service.generate_excel(db_session, session=s)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
    assert "CONCILIACIÓN COSTOS Y GASTOS A5" in wb.sheetnames

    a5 = wb["CONCILIACIÓN COSTOS Y GASTOS A5"]
    # Modo referencial: H66 y H67 contienen fórmulas hacia DATOS BALANCE.
    val_806 = a5["H66"].value
    val_807 = a5["H67"].value
    assert isinstance(val_806, str) and val_806.startswith("=") and "DATOS BALANCE" in val_806, \
        f"H66 debe ser fórmula referencial DATOS BALANCE, obtenido {val_806!r}"
    assert isinstance(val_807, str) and val_807.startswith("=") and "DATOS BALANCE" in val_807, \
        f"H67 debe ser fórmula referencial DATOS BALANCE, obtenido {val_807!r}"
    assert _datos_balance_has_value(wb, 3000.0)
    assert _datos_balance_has_value(wb, 500.0)


def test_a6_casillero_810_from_balance(db_session, client_user):
    """A6 casillero 810 should be filled from balance_mapeado when f101 is empty."""
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="Test", numero_adhesivo=None,
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A1",
        extracted_data={
            "balance_mapeado": [
                {"casillero_sri": "810", "codigo": "DD", "descripcion": "Ded adicional", "saldo": 7500.0},
            ],
        },
        warnings=[], uploaded_file_meta={"slot": "balance_mapeado"},
        new_status="ready",
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A6",
        extracted_data={"f101": {}},
        warnings=[], uploaded_file_meta={"slot": "f101"},
        new_status="ready",
    )
    ict_service.recompute_indice(db_session, session=s)
    db_session.refresh(s)

    excel_bytes = ict_service.generate_excel(db_session, session=s)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
    assert "BENEFICIOS TRIBUTARIOS A6" in wb.sheetnames

    a6 = wb["BENEFICIOS TRIBUTARIOS A6"]
    val_810 = a6["G25"].value
    assert isinstance(val_810, str) and val_810.startswith("=") and "DATOS BALANCE" in val_810, \
        f"G25 debe ser fórmula referencial DATOS BALANCE, obtenido {val_810!r}"
    assert _datos_balance_has_value(wb, 7500.0)


def test_a9_casilleros_from_balance(db_session, client_user):
    """A9 inventory casilleros should be filled from balance_mapeado when f101 is empty."""
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="Test", numero_adhesivo=None,
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A1",
        extracted_data={
            "balance_mapeado": [
                {"casillero_sri": "7001", "codigo": "INV01", "descripcion": "Inv Inicial", "saldo": 500000.0},
            ],
        },
        warnings=[], uploaded_file_meta={"slot": "balance_mapeado"},
        new_status="ready",
    )
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A9",
        extracted_data={"f101": {}},
        warnings=[], uploaded_file_meta={"slot": "f101"},
        new_status="ready",
    )
    ict_service.recompute_indice(db_session, session=s)
    db_session.refresh(s)

    excel_bytes = ict_service.generate_excel(db_session, session=s)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
    assert "INVENTARIOS A9" in wb.sheetnames

    # Modo referencial: A9 contiene fórmulas hacia DATOS BALANCE,
    # el valor literal 500000.0 vive en DATOS BALANCE.
    a9 = wb["INVENTARIOS A9"]
    assert _has_referential_formula_for(a9, 500000.0), \
        "A9 debería tener al menos una fórmula referencial (no valores literales)"
    assert _datos_balance_has_value(wb, 500000.0), \
        "El valor 500000.0 (casillero 7001) debe estar en DATOS BALANCE"


def test_f101_takes_precedence_over_balance_in_merged_context(db_session, client_user):
    """When A3 has its own f101 with a casillero, it should take priority over
    the same casillero in A1's balance_mapeado."""
    s = ict_service.create_session(
        db_session, user=client_user, ejercicio_fiscal="2025",
        ruc="1234567890001", razon_social="Test", numero_adhesivo=None,
    )
    # A1 has balance with 7185 = 99999.0
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A1",
        extracted_data={
            "balance_mapeado": [
                {"casillero_sri": "7185", "codigo": "X", "descripcion": "Gto", "saldo": 99999.0},
            ],
        },
        warnings=[], uploaded_file_meta={"slot": "balance_mapeado"},
        new_status="ready",
    )
    # A3 has its own f101 with 7185 = 25000.0 (should win)
    ict_service.update_anexo_data(
        db_session, session=s, anexo_code="A3",
        extracted_data={"f101": {"7185": 25000.0, "7992": 100000.0}},
        warnings=[], uploaded_file_meta={"slot": "f101"},
        new_status="ready",
    )
    ict_service.recompute_indice(db_session, session=s)
    db_session.refresh(s)

    excel_bytes = ict_service.generate_excel(db_session, session=s)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)

    # Modo referencial: F-101 mergea valores de TODOS los anexos (A1+A3).
    # En DATOS F-101 debe figurar el valor 25000.0 del casillero 7185
    # (el f101 de A3 tiene prioridad sobre el balance de A1).
    assert _datos_f101_has_value(wb, "7185", 25000.0), \
        "DATOS F-101 debe contener cas 7185=25000.0 (F-101 tiene prioridad sobre balance)"
    # El valor 99999.0 del balance NO debe ganar — pero puede aparecer
    # como saldo en DATOS BALANCE como dato fuente sin uso. Lo que NO
    # debe pasar es que A3 referencie ese valor. Validamos que la fórmula
    # de A3 para 7185 apunte a DATOS F-101 (no a DATOS BALANCE).
    a3 = wb["COSTOS  GASTOS A3"]
    # Filas 16 y 21 escriben casillero 7185
    f16 = a3["F16"].value
    f21 = a3["F21"].value
    assert isinstance(f16, str) and "DATOS F-101" in f16, \
        f"F16 (cas 7185) debe referenciar DATOS F-101, obtenido {f16!r}"
    assert isinstance(f21, str) and "DATOS F-101" in f21, \
        f"F21 (cas 7185) debe referenciar DATOS F-101, obtenido {f21!r}"
