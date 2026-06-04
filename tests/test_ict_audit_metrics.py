"""Tests for backend.app.ict.audit.metrics."""
from decimal import Decimal

import openpyxl
import pytest

from backend.app.ict.audit.metrics import (
    compute_a1_metrics,
    compute_anexos_metrics,
)
from backend.app.ict.audit.schemas import A1Metrics, AnexosMetrics, Status


def _build_minimal_a1_workbook():
    """Build a minimal workbook simulating sheets needed by compute_a1_metrics."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    f101 = wb.create_sheet("DATOS F-101")
    f101["A1"] = "Cas"
    f101["B1"] = "Nombre"
    f101["C1"] = "Valor"
    f101["A2"] = "499"
    f101["B2"] = "TOTAL ACTIVOS"
    f101["C2"] = 21671880.68
    f101["A3"] = "699"
    f101["B3"] = "TOTAL PASIVO Y PATRIMONIO"
    f101["C3"] = 21671880.68

    a1 = wb.create_sheet("A1")
    a1["A1"] = "Cas"
    a1["B1"] = "Nombre"
    a1["C1"] = "Declarado"
    a1["F1"] = "Contable"
    for i, cas in enumerate(["302", "303", "304"], start=2):
        a1[f"A{i}"] = cas
        a1[f"C{i}"] = 1000.0 * i
        a1[f"F{i}"] = 1000.0 * i
    return wb


def test_compute_a1_metrics_cuadra_perfecto():
    wb = _build_minimal_a1_workbook()
    m = compute_a1_metrics(wb)
    assert isinstance(m, A1Metrics)
    assert m.activo_total == Decimal("21671880.68")
    assert m.pasivo_patrimonio_total == Decimal("21671880.68")
    assert m.diferencia == Decimal("0.00")
    assert m.status_cuadre == Status.OK


def test_compute_a1_metrics_cobertura_pct():
    wb = _build_minimal_a1_workbook()
    m = compute_a1_metrics(wb)
    # 3 cas con contrapartida contable (F columna no vacía), 3 total → 100%
    assert m.cobertura_mapeo_pct == 100.0
    assert m.cas_mapeados == 3
    assert m.cas_total == 3
    assert m.cas_sin_contrapartida == []


def test_compute_a1_metrics_cas_sin_contrapartida():
    wb = _build_minimal_a1_workbook()
    # cas 305 declarado pero sin contable
    a1 = wb["A1"]
    a1["A5"] = "305"
    a1["C5"] = 500.0
    a1["F5"] = None
    m = compute_a1_metrics(wb)
    assert "305" in m.cas_sin_contrapartida
    assert m.cas_mapeados == 3
    assert m.cas_total == 4
    assert m.cobertura_mapeo_pct == 75.0


def test_compute_anexos_metrics_returns_9_anexos():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for code in ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9"]:
        wb.create_sheet(code)
    am = compute_anexos_metrics(wb)
    assert isinstance(am, AnexosMetrics)
    assert len(am.anexos) == 9
    codes = [a.codigo for a in am.anexos]
    assert codes == ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9"]
