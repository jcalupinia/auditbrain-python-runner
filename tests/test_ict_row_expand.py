"""Tests para el helper de expansión de bloques tabulares (row_expand).

Cuando un anexo tiene más datos que filas disponibles (ej. A5 Cuadro A con
15 casilleros no deducibles vs 5 filas), se insertan filas y se reajustan
las fórmulas del template que quedaron desplazadas.

openpyxl.insert_rows NO reajusta las referencias de las fórmulas; este helper
(`shift_formula_rows`) lo hace: suma `amount` a cada referencia de fila >=
`threshold`, respetando refs absolutas ($), rangos, y SIN tocar literales
de texto entre comillas ni referencias a otras hojas.
"""
from __future__ import annotations

from backend.app.ict.fillers.row_expand import shift_formula_rows, expand_tabular_block


def test_shift_simple_sum_range():
    # =SUM(G28:G32) con inserción de 3 filas en pos 22 → ambos refs >= 22 +3
    assert shift_formula_rows("=SUM(G28:G32)", threshold=22, amount=3) == "=SUM(G31:G35)"


def test_shift_no_cambia_refs_por_encima_del_umbral():
    # K17 y K21 están por encima del umbral (17,21 < 22) → no cambian
    assert shift_formula_rows("=SUM(K17:K21)", threshold=22, amount=3) == "=SUM(K17:K21)"


def test_shift_refs_mixtas():
    # =+K22+G51+H61-H72 con +3: todas >= 22 → +3
    assert shift_formula_rows("=+K22+G51+H61-H72", threshold=22, amount=3) == "=+K25+G54+H64-H75"


def test_shift_no_toca_porcentajes_ni_numeros_sueltos():
    # 15% NO es una referencia de celda (no tiene letra de columna delante)
    f = "=IF(((H59-H60)>=0),(H58*15%)+((H59-H60)*15%),(H58*15%))"
    out = shift_formula_rows(f, threshold=22, amount=3)
    assert out == "=IF(((H62-H63)>=0),(H61*15%)+((H62-H63)*15%),(H61*15%))"


def test_shift_respeta_literales_de_texto_entre_comillas():
    # "0,00%" es un literal — no debe modificarse
    f = '=IF(G33=0,"0,00%",G33/G41)'
    out = shift_formula_rows(f, threshold=22, amount=3)
    assert out == '=IF(G36=0,"0,00%",G36/G44)'


def test_shift_referencias_absolutas():
    # $B$17 con umbral 22: 17 < 22 → no cambia; $B$30 → $B$33
    assert shift_formula_rows("=$B$30+$B$17", threshold=22, amount=3) == "=$B$33+$B$17"


def test_shift_no_toca_referencias_a_otra_hoja():
    # 'DATOS F-101'!C500 es una ref externa — la fila NO debe desplazarse
    f = "='DATOS F-101'!C500"
    assert shift_formula_rows(f, threshold=22, amount=3) == "='DATOS F-101'!C500"


def test_shift_amount_cero_es_identidad():
    f = "=+K22+G51+H61-H72"
    assert shift_formula_rows(f, threshold=22, amount=0) == f


# ── expand_tabular_block: preservación de formato (bordes + merges) ───────────

def _thin():
    from openpyxl.styles import Side
    return Side(style="thin")


def test_expand_preserva_bordes_de_filas_desplazadas():
    """Tras insertar filas, las celdas desplazadas (con y sin valor) deben
    conservar sus bordes — incluso celdas vacías con borde."""
    from openpyxl import Workbook
    from openpyxl.styles import Border
    wb = Workbook()
    ws = wb.active
    b = Border(top=_thin(), bottom=_thin(), left=_thin(), right=_thin())
    # Fila 30: A con valor + borde, C vacía pero CON borde (caso que rompe openpyxl)
    ws.cell(30, 1).value = "dato"
    ws.cell(30, 1).border = b
    ws.cell(30, 3).border = b  # vacía pero con borde
    # Insertar 4 filas en pos 25 → fila 30 baja a 34
    expand_tabular_block(ws, insert_at=25, amount=4, style_row=24, last_col=5)
    assert ws.cell(34, 1).value == "dato"
    assert ws.cell(34, 1).border.top.style == "thin", "A: borde top perdido"
    assert ws.cell(34, 1).border.bottom.style == "thin", "A: borde bottom perdido"
    assert ws.cell(34, 3).border.left.style == "thin", "C vacía: borde perdido"


def test_expand_preserva_merges_de_filas_desplazadas():
    """Los merges en filas desplazadas deben moverse +amount, ninguno se pierde."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.cell(30, 1).value = "encabezado"
    ws.merge_cells("A30:E30")
    ws.cell(31, 1).value = "fila2"
    ws.merge_cells("A31:C31")
    expand_tabular_block(ws, insert_at=25, amount=4, style_row=24, last_col=5)
    merges = {str(m) for m in ws.merged_cells.ranges}
    assert "A34:E34" in merges, f"merge A30→A34 perdido. {merges}"
    assert "A35:C35" in merges, f"merge A31→A35 perdido. {merges}"
    # No deben quedar merges fantasma en la posición vieja
    assert "A30:E30" not in merges, f"merge fantasma en pos vieja. {merges}"
