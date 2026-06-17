"""Regression test: cuentas balance mapeadas al cas 888 deben trasladarse
al cas 850 cuando la regla muestra cas 850 (cliente ICT_26, 2026-06-17).

Escenario reportado por cliente (FOSFORERA ICT_26):
  - Balance: cuenta "610102.02 Gasto imp renta causado" mapeada al cas 888
  - F-101: NO declara cas 888 (valor 0)
  - F-101: declara cas 850 = 990,162.31

Antes del fix: A1 mostraba cas 850 (correcto segun regla 888/850 condicional)
PERO la cuenta del balance "610102.02" quedaba "huerfana" — aparecia en col
D/E con codigo pintado en rojo pero col F vacia (sin saldo).

Despues del fix: cuando la regla decide mostrar cas 850, las cuentas
balance mapeadas al cas 888 se trasladan al cas 850 → aparecen como
cuentas del 850 con su codigo, descripcion y saldo en col D/E/F.

Razon: ambos cas representan conceptualmente el mismo impuesto. Si la
empresa lo mapeo a 888 en el balance pero a 850 en F-101, la cuenta NO
debe perderse.
"""
from __future__ import annotations

import pytest


def _find_cas_row(ws, cas_buscado: str) -> int | None:
    for r in range(13, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).strip() == cas_buscado:
            return r
    return None


def _build_wb_escenario_ict26():
    """Genera ICT con cuenta balance mapeada al cas 888 pero F-101 declara 850."""
    from openpyxl import Workbook
    from backend.app.ict.cell_maps.a1 import A1_SHEET
    from backend.app.ict.fillers.a1_mapeo import A1Filler
    from backend.app.ict.fillers.source_data_sheets import build_f101_sheet, build_balance_sheet

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.create_sheet(A1_SHEET)
    f101_data = {
        "6001": 5000.0, "7001": 2000.0, "801": 4376307.96, "803": 656446.19,
        "850": 990162.31,  # F-101 declara IR causado bajo 850
        "888": 0,          # F-101 NO declara cas 888
    }
    f101_lookup = build_f101_sheet(wb, f101_data, {})
    # Balance: cuenta mapeada al cas 888 (como el caso del cliente)
    balance = [
        {
            "casillero_sri": "888", "codigo": "610102.02",
            "descripcion": "Gasto imp renta causado", "saldo": 990162.31,
        },
    ]
    balance_lookup = build_balance_sheet(wb, balance)

    filler = A1Filler()
    session_data = {"razon_social": "TEST", "ruc": "T",
                    "ejercicio_fiscal": "2025", "numero_adhesivo": ""}
    anexo_data = {
        "f101": f101_data,
        "balance_mapeado": balance,
        "_f101_lookup": f101_lookup,
        "_balance_lookup": balance_lookup,
    }
    filler.fill(wb, session_data, anexo_data)
    return wb


def test_cas_850_tiene_la_cuenta_balance_del_888():
    """Cuando A1 muestra cas 850 (porque F-101 cas 888=0), las cuentas
    balance mapeadas al cas 888 deben aparecer como cuentas del cas 850
    (codigo, descripcion y saldo no se pierden)."""
    from backend.app.ict.cell_maps.a1 import A1_SHEET
    wb = _build_wb_escenario_ict26()
    ws = wb[A1_SHEET]

    row_850 = _find_cas_row(ws, "850")
    assert row_850 is not None, "cas 850 debe aparecer (fallback)"

    # La cuenta balance debe estar como detalle del cas 850
    codigo = ws.cell(row_850, 4).value  # col D
    descripcion = ws.cell(row_850, 5).value  # col E

    assert codigo == "610102.02", (
        f"col D del cas 850 debe tener codigo '610102.02' (cuenta mapeada "
        f"originalmente al cas 888). Encontrado: {codigo!r}"
    )
    assert descripcion == "Gasto imp renta causado", (
        f"col E del cas 850 debe tener descripcion 'Gasto imp renta causado'. "
        f"Encontrado: {descripcion!r}"
    )

    # Col F debe tener fórmula (no estar vacía)
    f_val = ws.cell(row_850, 6).value
    assert f_val is not None and str(f_val) != "", (
        f"col F del cas 850 NO debe estar vacia (debe tener saldo o formula). "
        f"Encontrado: {f_val!r}"
    )
    assert isinstance(f_val, str) and f_val.startswith("="), (
        f"col F del cas 850 debe ser una formula referencial al balance. "
        f"Encontrado: {f_val!r}"
    )


def test_cas_888_no_aparece_porque_f101_no_declara():
    """Verifica que cas 888 NO aparezca cuando F-101 lo declara en 0."""
    from backend.app.ict.cell_maps.a1 import A1_SHEET
    wb = _build_wb_escenario_ict26()
    ws = wb[A1_SHEET]
    row_888 = _find_cas_row(ws, "888")
    assert row_888 is None, (
        "cas 888 NO debe aparecer cuando F-101 lo declara en 0 "
        "(regla 888/850 condicional)"
    )
