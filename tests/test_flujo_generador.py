import io

from openpyxl import load_workbook

from backend.app.client_portal.flujo import generador

# Balanza sintética mínima con activo, pasivo, patrimonio, efectivo, ingresos y
# gastos. Los códigos Super Cías siguen el plan Superintendencia (Activo "1",
# Pasivo "2", Patrimonio "3", ingresos/gastos del ERI "4"/"5").
HOJAS_ESPERADAS = [
    "RESUMEN", "Homologación", "ESF", "ERI", "Flujo de Efectivo",
    "Evolución del Patrimonio", "Movimiento no Efectivo", "Balance resumido",
    "Formulario 101", "Notas", "Indicadores",
]


def _balanza_anterior():
    return [
        {"cuenta": "Caja", "super_cias": "1010101", "sri": "311", "saldo": 400.0},
        {"cuenta": "Proveedores", "super_cias": "2010301", "sri": "413", "saldo": -300.0},
        {"cuenta": "Capital", "super_cias": "30101", "sri": "601", "saldo": -100.0},
    ]


def _balanza_actual():
    return [
        {"cuenta": "Caja", "super_cias": "1010101", "sri": "311", "saldo": 500.0},
        {"cuenta": "Proveedores", "super_cias": "2010301", "sri": "413", "saldo": -250.0},
        {"cuenta": "Capital", "super_cias": "30101", "sri": "601", "saldo": -100.0},
        {"cuenta": "Ventas", "super_cias": "4010101", "sri": "6011", "saldo": -800.0},
        {"cuenta": "Sueldos", "super_cias": "5020101", "sri": "7011", "saldo": 650.0},
    ]


def test_generar_excel_devuelve_bytes_con_todas_las_hojas():
    data = generador.generar_excel(_balanza_anterior(), _balanza_actual())
    assert isinstance(data, (bytes, bytearray))
    assert len(data) > 0
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == HOJAS_ESPERADAS


def test_resumen_muestra_cuadraturas():
    data = generador.generar_excel(_balanza_anterior(), _balanza_actual())
    wb = load_workbook(io.BytesIO(data))
    ws = wb["RESUMEN"]
    textos = " ".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    ).upper()
    assert "AUDITCONSULTING" in textos
    assert "AUDIT-IA" in textos
    # las tres cuadraturas semáforo aparecen nombradas
    assert "ESF" in textos
    assert "AF" in textos or "FLUJO" in textos
    assert "NETA" in textos


def test_no_lanza_excepcion_con_balanza_minima():
    # smoke: solo activo=pasivo, sin ERI
    bal = [
        {"cuenta": "Caja", "super_cias": "1010101", "sri": "311", "saldo": 100.0},
        {"cuenta": "Proveedores", "super_cias": "2010301", "sri": "413", "saldo": -100.0},
    ]
    data = generador.generar_excel(bal, bal)
    wb = load_workbook(io.BytesIO(data))
    assert len(wb.sheetnames) == 11  # + Notas + Balance resumido
    assert "Notas" in wb.sheetnames


def test_ninguna_celda_texto_parece_formula():
    # Regla del proyecto: ninguna celda de texto debe empezar con = + - @
    # (Excel levantaría el cuadro "Reparaciones").
    data = generador.generar_excel(_balanza_anterior(), _balanza_actual())
    wb = load_workbook(io.BytesIO(data))
    problemas = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value[:1] in ("=", "+", "-", "@"):
                    problemas.append((ws.title, c.coordinate, c.value))
    assert problemas == [], f"Celdas de texto tipo fórmula: {problemas}"
