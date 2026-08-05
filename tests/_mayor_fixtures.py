"""Constructores de mayores sintéticos para los tests del motor.

Los importes son inventados: NUNCA se commitean cifras de clientes a este
repositorio público.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

# Encabezado exacto del ERP del cliente de referencia.
ENCABEZADO_REAL = (
    "Código", "Cuenta", "Fecha", "Asiento", "Documento", "Identificación",
    "Persona", "Persona Cruce Cuenta", "Descripción", "Debe", "Haber", "Saldo",
)


def mayor_xlsx(
    filas: list[list],
    *,
    encabezado: tuple[str, ...] = ENCABEZADO_REAL,
    fila_encabezado: int = 1,
    hoja: str = "Hoja1",
    hojas_previas: tuple[str, ...] = (),
) -> bytes:
    """Devuelve los bytes de un .xlsx con el encabezado y las filas dadas."""
    wb = Workbook()
    ws_primera = wb.active
    for i, nombre in enumerate(hojas_previas):
        (ws_primera if i == 0 else wb.create_sheet()).title = nombre
    ws = ws_primera if not hojas_previas else wb.create_sheet()
    ws.title = hoja

    for col, valor in enumerate(encabezado, start=1):
        ws.cell(fila_encabezado, col, valor)
    for j, fila in enumerate(filas, start=fila_encabezado + 1):
        for col, valor in enumerate(fila, start=1):
            ws.cell(j, col, valor)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def mayor_xlsx_multihoja(
    hojas: dict[str, list[list]],
    *,
    encabezado: tuple[str, ...] = ENCABEZADO_REAL,
    fila_encabezado: int = 1,
) -> bytes:
    """Como mayor_xlsx, pero escribe VARIAS hojas con datos, cada una con su
    propio encabezado (idéntico) y sus propias filas. Sirve para simular un
    mayor repartido en varias hojas (una por mes o por tipo de comprobante).
    """
    wb = Workbook()
    nombres = list(hojas.keys())
    hojas_ws = {nombres[0]: wb.active}
    hojas_ws[nombres[0]].title = nombres[0]
    for nombre in nombres[1:]:
        hojas_ws[nombre] = wb.create_sheet(title=nombre)

    for nombre, filas in hojas.items():
        ws = hojas_ws[nombre]
        for col, valor in enumerate(encabezado, start=1):
            ws.cell(fila_encabezado, col, valor)
        for j, fila in enumerate(filas, start=fila_encabezado + 1):
            for col, valor in enumerate(fila, start=1):
                ws.cell(j, col, valor)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
