"""Verificación numérica de las cédulas DM contra el archivo modelo del auditor.

`openpyxl` no evalúa fórmulas: al leer el libro recién generado, una celda con
fórmula devuelve la cadena "=..." en vez de su resultado. Por eso esta
verificación no abre el libro que genera el sistema: **calcula en Python**,
directamente a partir del mayor real y de los PDFs F-104 reales del cliente,
lo que cada fila de cada cédula debería dar, y lo contrasta contra los
**valores cacheados** del archivo modelo del auditor (abierto con
``data_only=True``).

Requiere dos insumos, ambos fuera del repo público:

- ``AUD_OF_FIXTURES_DIR``: carpeta con ``MAYOR DE IMPUESTOS.xlsx`` y los PDFs
  F-104 en ``104/`` (datos reales del cliente).
- El archivo modelo del auditor en
  ``%USERPROFILE%\\Downloads\\DM - Obligaciones Fiscales FINAL (1).xlsx``.

Si cualquiera de los dos falta, toda la suite se salta (``skip``).
"""

from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.aud.obligaciones_fiscales.cedulas.f104_extractor import extract_all_f104
from backend.app.aud.obligaciones_fiscales.mayor.clasificador import clasificar
from backend.app.aud.obligaciones_fiscales.mayor.cuentas import monto_segun_libros, perfilar
from backend.app.aud.obligaciones_fiscales.mayor.reader import leer_mayor

TOLERANCIA = 0.01
MESES = [f"{m:02d}" for m in range(1, 13)]

_MODELO_PATH = (
    Path(os.environ.get("USERPROFILE", ""))
    / "Downloads"
    / "DM - Obligaciones Fiscales FINAL (1).xlsx"
)

pytestmark = pytest.mark.skipif(
    not os.getenv("AUD_OF_FIXTURES_DIR") or not _MODELO_PATH.exists(),
    reason="Requiere AUD_OF_FIXTURES_DIR y el archivo modelo del auditor en Downloads",
)


@pytest.fixture(scope="module")
def datos_reales():
    """Lee el mayor y los F-104 reales; los clasifica sin pasar por Excel."""
    base = Path(os.environ["AUD_OF_FIXTURES_DIR"])
    lectura = leer_mayor((base / "MAYOR DE IMPUESTOS.xlsx").read_bytes())
    perfiles = perfilar(lectura.movimientos)
    resultados = clasificar(perfiles)
    f104_mes, errores = extract_all_f104(sorted((base / "104").glob("*.pdf")))
    assert not errores, f"errores extrayendo los F-104 reales: {errores}"
    return {"resultados": resultados, "perfiles": perfiles, "f104": f104_mes}


@pytest.fixture(scope="module")
def modelo():
    """El archivo del auditor con los valores cacheados (no fórmulas)."""
    return load_workbook(_MODELO_PATH, data_only=True)


def _iva_compras_libros(datos_reales: dict, mes: str) -> float:
    """Suma del mayor de las cuentas IVA_COMPRAS en el mes dado.

    Usa ``monto_segun_libros`` (débito bruto, no el neto) porque IVA en
    compras es una cuenta de activo: la misma lógica que aplica
    ``clasificacion_service.guardar_clasificacion`` en producción.
    """
    total = 0.0
    for r in datos_reales["resultados"]:
        if r.categoria == "IVA_COMPRAS":
            perfil = datos_reales["perfiles"].get(r.codigo)
            if perfil:
                total += monto_segun_libros(perfil, r.categoria).get(mes, 0.0)
    return total


def _cas(datos_reales: dict, mes: str, cas: str) -> float:
    """Valor de un casillero del F-104 real de ese mes; 0 si no está."""
    return datos_reales["f104"].get(mes, {}).get("casilleros", {}).get(cas) or 0.0


def _fila_dm6_por_mes(modelo, mes: str) -> int:
    """Fila del modelo DM6 correspondiente al mes "01".."12"."""
    ws = modelo["DM6 IVA"]
    nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
               "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    nombre = nombres[int(mes) - 1]
    return next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == nombre)


# --- Caso 1: DM4 IVA en compras según libros, enero = 9.911,57 ---

def test_dm4_iva_compras_segun_libros_enero(datos_reales):
    obtenido = _iva_compras_libros(datos_reales, "01")
    assert obtenido == pytest.approx(9911.57, abs=TOLERANCIA)


# --- Caso 2: DM4 base imponible, enero = 66.077,13 (IVA / 0,15) ---

def test_dm4_base_imponible_enero(datos_reales):
    base = _iva_compras_libros(datos_reales, "01") / 0.15
    assert base == pytest.approx(66077.13, abs=TOLERANCIA)


# --- Caso 3: DM6 columna F = casillero 480, los 12 meses ---

def test_dm6_columna_f_casillero_480_los_doce_meses(datos_reales, modelo):
    for mes in MESES:
        fila = _fila_dm6_por_mes(modelo, mes)
        esperado = modelo["DM6 IVA"].cell(fila, 6).value  # columna F
        obtenido = _cas(datos_reales, mes, "480")
        assert obtenido == pytest.approx(esperado, abs=TOLERANCIA), f"mes {mes}"


# --- Caso 4: DM6 columna Z = 615+617, enero = 9.541,74 ---

def test_dm6_columna_z_615_mas_617_enero(datos_reales):
    obtenido = _cas(datos_reales, "01", "615") + _cas(datos_reales, "01", "617")
    assert obtenido == pytest.approx(9541.74, abs=TOLERANCIA)


# --- Caso 5: DM6 columna T de enero = 605+606 = 3.770,72 ---

def test_dm6_columna_t_605_mas_606_enero(datos_reales):
    obtenido = _cas(datos_reales, "01", "605") + _cas(datos_reales, "01", "606")
    assert obtenido == pytest.approx(3770.72, abs=TOLERANCIA)


# --- Caso 6: DM3 crédito tributario = 615+617 de diciembre = 68,07 ---

def test_dm3_credito_tributario_615_mas_617_diciembre(datos_reales):
    obtenido = _cas(datos_reales, "12", "615") + _cas(datos_reales, "12", "617")
    assert obtenido == pytest.approx(68.07, abs=TOLERANCIA)


# --- Caso 7: DM3 IVA diferido = casillero 485 de diciembre = 14.139,28 ---

def test_dm3_iva_diferido_485_diciembre(datos_reales):
    obtenido = _cas(datos_reales, "12", "485")
    assert obtenido == pytest.approx(14139.28, abs=TOLERANCIA)
