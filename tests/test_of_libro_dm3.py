"""DM3 Revisión de saldos: tres bloques de una cifra anual, libros vs F-104.

El "según libros" es el movimiento ACUMULADO DEL AÑO (columna Total del
resumen), no el saldo al cierre. Si el cliente no tiene la cuenta, el bloque
se escribe igual con 0 y una nota.
"""

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm3_saldos import (
    SHEET_DM3,
    build_dm3,
)

PERIODOS = [f"2025-{m:02d}" for m in range(1, 13)]

DIR_MAYORES = {
    ("cuenta:1.1.5.1.2", "TOTAL"): "'Mayores homologados'!P10",
    ("cuenta:2.1.7.4.2", "TOTAL"): "'Mayores homologados'!P20",
    # Nótese: NO hay ("cuenta:2.1.7.5.6", "TOTAL") — simula un cliente sin
    # esa cuenta en su mayor, el caso que la nota debe cubrir.
}

DIR_F104 = {
    ("2025-12", "615"): "'DATOS F-104'!N100",
    ("2025-12", "617"): "'DATOS F-104'!N101",
    ("2025-12", "485"): "'DATOS F-104'!N102",
    ("2025-12", "859"): "'DATOS F-104'!N103",
}

DIR_DM7 = {
    ("ret_renta_declarado", "12"): "'DM7 Retenciones x pagar'!O50",
}


def _cedula(**kw):
    wb = Workbook()
    datos = dict(dir_mayores=DIR_MAYORES, dir_f104=DIR_F104, dir_dm7=DIR_DM7,
                 periodos=PERIODOS, cliente="C", periodo="2025")
    datos.update(kw)
    build_dm3(wb, **datos)
    return wb[SHEET_DM3]


def _todas_las_celdas(ws):
    return [ws.cell(r, c).value for r in range(1, ws.max_row + 1) for c in range(1, 6)]


def test_lleva_el_encabezado_de_cedula_con_su_referencia():
    ws = _cedula()
    valores = [ws.cell(r, c).value for r in range(1, 11) for c in range(1, 6)]
    assert "OBLIGACIONES FISCALES" in valores
    assert "DM3" in valores


def test_el_bloque_de_credito_tributario_esta_presente():
    ws = _cedula()
    valores = _todas_las_celdas(ws)
    assert any("CREDITO TRIBUTARIO" in str(v).upper() for v in valores if v)


def test_el_credito_tributario_segun_libros_es_formula_a_la_cuenta_total():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(r, 2).value == "Según libros" and
                str(ws.cell(r - 2, 1).value or "") == "1.1.5.1.2")
    assert ws.cell(fila, 3).value == "='Mayores homologados'!P10"


def test_el_credito_tributario_segun_declaracion_suma_615_y_617_de_diciembre():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if str(ws.cell(r, 2).value or "").startswith("Según F-104"))
    assert ws.cell(fila, 3).value == "='DATOS F-104'!N100+'DATOS F-104'!N101"


def test_el_iva_diferido_segun_declaracion_es_el_casillero_485_de_diciembre():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if "485" in str(ws.cell(r, 2).value or ""))
    assert ws.cell(fila, 3).value == "='DATOS F-104'!N102"


def test_el_sri_por_pagar_segun_declaracion_suma_859_y_retencion_renta_de_diciembre():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if "859" in str(ws.cell(r, 2).value or ""))
    valor = ws.cell(fila, 3).value
    assert valor == "='DATOS F-104'!N103+'DM7 Retenciones x pagar'!O50"


def test_cuenta_faltante_se_escribe_en_cero_con_una_nota():
    """2.1.7.5.6 (SRI por pagar) no está en DIR_MAYORES: el bloque no debe
    reventar, debe quedar en 0 y dejar una nota visible para el auditor."""
    ws = _cedula()
    fila_libros_sri = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(r, 2).value == "Según libros" and
        str(ws.cell(r - 2, 1).value or "") == "2.1.7.5.6"
    )
    assert ws.cell(fila_libros_sri, 3).value == 0
    notas = [str(ws.cell(r, c).value or "") for r in range(1, ws.max_row + 1) for c in range(1, 6)]
    assert any("2.1.7.5.6" in n and ("no" in n.lower()) for n in notas)


def test_las_tres_diferencias_restan_libros_menos_declarado_y_redondean():
    ws = _cedula()
    filas = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "Diferencia"]
    assert len(filas) == 3
    for fila in filas:
        assert ws.cell(fila, 3).value.startswith("=ROUND(")


def test_las_cifras_de_dm3_no_llevan_columnas_de_meses():
    """A diferencia de DM4/DM5/DM7: DM3 es una cifra anual, no una tabla de
    12 meses. La columna D (mes de febrero en las otras cedulas) debe
    quedar vacia en las filas de 'Según libros'."""
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "Según libros")
    assert ws.cell(fila, 4).value is None


def test_se_puede_configurar_otra_cuenta_para_cada_bloque():
    ws = _cedula(
        cuenta_credito_tributario="9.9.9.1",
        cuenta_iva_diferido="9.9.9.2",
        cuenta_sri_por_pagar="9.9.9.3",
    )
    valores = _todas_las_celdas(ws)
    assert "9.9.9.1" in valores
    assert "9.9.9.2" in valores
    assert "9.9.9.3" in valores
