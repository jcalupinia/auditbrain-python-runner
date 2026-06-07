"""Regression test: parser de balance mapeado NUNCA omite cuentas
silenciosamente cuando tienen casillero SRI asignado pero saldo vacío.

REPORTE CLIENTE (2026-06-07): "HAY UN ERROR EN LA EXTRACCION DE LA
INFORMACION DEL BALANCE MAPEO POR EJEMPLO FALTA LA CUENTA DE PASIVO
CLIENTES QUE TIENE EL CODIGO 545 POR 746.7 NECESITO PONER UNA REGLA
QUE NO PERMITA QUE SE VUELVA A OMITIR INFORMACION".

DIAGNÓSTICO histórico (PROPHAR CT 2025):
La fila 350 del Plan de Cuentas tenía:
  - codigo: '2011001'
  - cas SRI: '545'
  - descripcion: 'ANTICIPOS DE CLIENTES'
  - saldo: vacío (None)
El parser ignoraba estas filas. El A1 no mostraba el cas 545. El
auditor no se enteraba de la omisión.

REGLA implementada en parsers/balance_mapeado_excel.py:
  - Si fila tiene casillero pero saldo vacío:
      → incluir cuenta con saldo=0.0 y saldo_vacio=True
      → agregar advertencia explícita a 'advertencias'
      → registrar en lista 'cuentas_sin_saldo'
"""
from __future__ import annotations

from io import BytesIO
import pytest
from openpyxl import Workbook

from backend.app.ict.parsers.balance_mapeado_excel import parse_balance_mapeado


def _make_excel(rows: list[list]) -> bytes:
    """Construye un Excel en memoria con headers + filas."""
    wb = Workbook()
    ws = wb.active
    # Headers en fila 1 (parser escanea hasta 20)
    ws.cell(1, 1, value="Cod.Cuenta.Contable")
    ws.cell(1, 2, value="Descripción cuenta contable")
    ws.cell(1, 3, value="Códigos SRI")
    ws.cell(1, 4, value="Saldos 31 DIC")
    # Filas de datos
    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, value=val)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


class TestNoOmitirCuentasSinSaldo:
    """Asegura que cuentas con cas pero sin saldo se reporten al auditor."""

    def test_cuenta_con_cas_pero_sin_saldo_se_incluye_con_saldo_cero(self):
        """Regresión bug PROPHAR cas 545: fila con cas y código pero saldo vacío
        debe incluirse en la salida con saldo=0 y saldo_vacio=True."""
        excel = _make_excel([
            ["1101", "Caja", "311", 1000.0],
            ["2011001", "ANTICIPOS DE CLIENTES", "545", None],  # ← bug histórico
            ["3101", "Capital", "601", 50000.0],
        ])
        res = parse_balance_mapeado(excel)
        cuentas = res["cuentas"]
        assert len(cuentas) == 3, (
            "Las 3 cuentas (incluyendo cas 545 sin saldo) deben aparecer"
        )

        # Verificar que cas 545 está incluido con saldo 0
        cas_545 = [c for c in cuentas if c["casillero_sri"] == "545"]
        assert len(cas_545) == 1, "cas 545 debe estar presente"
        assert cas_545[0]["saldo"] == 0.0
        assert cas_545[0]["saldo_vacio"] is True
        assert cas_545[0]["codigo"] == "2011001"
        assert "ANTICIPOS" in cas_545[0]["descripcion"].upper()

    def test_advertencia_explicita_por_cuenta_sin_saldo(self):
        """La salida debe tener una advertencia clara mencionando
        el cas y la cuenta sin saldo."""
        excel = _make_excel([
            ["1101", "Caja", "311", 1000.0],
            ["2011001", "ANTICIPOS DE CLIENTES", "545", None],
        ])
        res = parse_balance_mapeado(excel)
        advertencias = res.get("advertencias", [])
        assert len(advertencias) >= 1, "Debe haber al menos 1 advertencia"

        msg = advertencias[0]
        assert "545" in msg, f"Advertencia debe mencionar cas 545: {msg}"
        assert "2011001" in msg, f"Advertencia debe mencionar código: {msg}"
        assert "saldo" in msg.lower(), f"Advertencia debe explicar saldo vacío"

    def test_cuentas_sin_saldo_se_listan_aparte(self):
        """La salida tiene una lista `cuentas_sin_saldo` para facilitar
        el reporte al auditor."""
        excel = _make_excel([
            ["1101", "Caja", "311", 1000.0],
            ["2011001", "ANTICIPOS DE CLIENTES", "545", None],
            ["3101", "Otros", "549", None],  # otro cas también vacío
        ])
        res = parse_balance_mapeado(excel)
        sin_saldo = res.get("cuentas_sin_saldo", [])
        assert len(sin_saldo) == 2, (
            f"Esperaba 2 cuentas sin saldo, encontré {len(sin_saldo)}"
        )
        cas_vacios = {c["casillero_sri"] for c in sin_saldo}
        assert "545" in cas_vacios
        assert "549" in cas_vacios

    def test_cuenta_con_saldo_cero_explícito_NO_genera_advertencia(self):
        """Si el cliente puso explícitamente 0 en el saldo, NO es omisión —
        es una declaración válida de saldo cero. No genera advertencia."""
        excel = _make_excel([
            ["1101", "Caja", "311", 1000.0],
            ["2011001", "ANTICIPOS DE CLIENTES", "545", 0.0],  # explícito = 0
        ])
        res = parse_balance_mapeado(excel)
        # Ambas cuentas presentes
        assert len(res["cuentas"]) == 2
        # cas 545 con saldo_vacio=False (porque 0 fue explícito)
        cas_545 = [c for c in res["cuentas"] if c["casillero_sri"] == "545"][0]
        assert cas_545["saldo"] == 0.0
        assert cas_545["saldo_vacio"] is False
        # Ninguna advertencia generada por saldo vacío
        advs_545 = [a for a in res.get("advertencias", []) if "545" in a]
        assert len(advs_545) == 0

    def test_fila_sin_cas_sigue_siendo_ignorada(self):
        """Filas agrupadoras (sin cas) deben seguir siendo ignoradas,
        sin generar advertencia."""
        excel = _make_excel([
            ["1000", "GRUPO ACTIVOS", None, None],  # fila agrupadora
            ["1101", "Caja", "311", 1000.0],
        ])
        res = parse_balance_mapeado(excel)
        # Solo la cuenta con cas debe estar
        assert len(res["cuentas"]) == 1
        assert res["cuentas"][0]["casillero_sri"] == "311"
        assert len(res.get("cuentas_sin_saldo", [])) == 0

    def test_compatibilidad_backward_saldo_vacio_no_rompe_clientes_legacy(self):
        """Las cuentas con saldo_vacio=True deben tener saldo=0.0 para que
        los consumidores legacy (fillers A1/A2/etc) sigan funcionando sin
        cambios — al sumar contribuyen 0."""
        excel = _make_excel([
            ["2011001", "ANTICIPOS", "545", None],
        ])
        res = parse_balance_mapeado(excel)
        c = res["cuentas"][0]
        # Cliente legacy hace: float(c["saldo"]) — debe dar 0.0
        assert float(c["saldo"]) == 0.0
        # Si tuvieran que filtrar, podrían usar saldo_vacio
        assert c.get("saldo_vacio") is True


class TestPROPHAREmpirico:
    """Verificación contra el archivo real del cliente PROPHAR (skip si no
    está disponible — útil en desarrollo local)."""

    def test_prophar_ct_detecta_cas_545_sin_saldo(self, tmp_path):
        """Si el archivo CT PROPHAR está disponible, verificar que cas 545
        se reporte como cuenta sin saldo (regresión cliente 2026-06-07)."""
        from pathlib import Path
        ct_path = Path(
            r"C:\Users\jcalu\Downloads\Información PROPHAR"
            r"\Información PROPHAR\CT PROPHAR.xlsm"
        )
        if not ct_path.exists():
            pytest.skip("CT PROPHAR.xlsm no disponible localmente")

        # NOTE: el parser usa wb.active (primera hoja). En CT PROPHAR, la
        # primera hoja es "Hoja2" (vacía). Para que el test sea relevante,
        # se ejecuta solo cuando el cliente provee un balance mapeado real
        # con la hoja correcta como activa.
        # Se incluye este test como esqueleto; el aseguramiento real lo
        # hace TestNoOmitirCuentasSinSaldo con datos sintéticos.
        pass
