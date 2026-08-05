"""El constructor de fixtures debe producir un xlsx legible por openpyxl."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from tests._mayor_fixtures import ENCABEZADO_REAL, mayor_xlsx, mayor_xlsx_multihoja
from backend.app.aud.obligaciones_fiscales.mayor.reader import leer_mayor


def test_construye_un_xlsx_con_el_encabezado_en_la_fila_indicada():
    data = mayor_xlsx(
        [["1.1.5.1.1", "IVA sobre Compras", None, "COM 1", "", "", "", "", "", 10, 0, 10]],
        fila_encabezado=3,
    )
    ws = load_workbook(BytesIO(data)).active
    assert [c.value for c in ws[3]] == list(ENCABEZADO_REAL)
    assert ws.cell(4, 1).value == "1.1.5.1.1"


FILA = ["1.1.5.1.1", "IVA sobre Compras", "2025-01-05", "COM 202501000001",
        "FAC 001-001-000000001", "9999999999001", "PROVEEDOR DEMO S.A.", "",
        "COMPRA DE PRUEBA", 2.39, None, 2.39]


def test_detecta_las_doce_columnas_del_erp_real():
    lectura = leer_mayor(mayor_xlsx([FILA]))
    assert lectura.mapeo_suficiente
    assert lectura.columnas_detectadas["codigo"] == 0
    assert lectura.columnas_detectadas["cuenta"] == 1
    assert lectura.columnas_detectadas["debe"] == 9
    assert lectura.columnas_detectadas["haber"] == 10
    assert lectura.columnas_detectadas["saldo"] == 11


def test_no_confunde_persona_cruce_cuenta_con_la_columna_cuenta():
    """La columna 8 se llama 'Persona Cruce Cuenta' y contiene 'cuenta'."""
    lectura = leer_mayor(mayor_xlsx([FILA]))
    assert lectura.columnas_detectadas["cuenta"] == 1


def test_encuentra_el_encabezado_aunque_no_este_en_la_primera_fila():
    lectura = leer_mayor(mayor_xlsx([FILA], fila_encabezado=6))
    assert lectura.fila_encabezado == 6
    assert len(lectura.movimientos) == 1


def test_reporta_las_columnas_que_no_pudo_mapear():
    lectura = leer_mayor(
        mayor_xlsx([["1.1.5.1.1", 10]], encabezado=("Cta", "Valor"))
    )
    assert lectura.mapeo_suficiente is False
    assert "debe" in lectura.columnas_faltantes


def test_elige_la_hoja_con_mas_columnas_reconocidas():
    data = mayor_xlsx([FILA], hoja="MAYOR", hojas_previas=("Portada",))
    lectura = leer_mayor(data)
    assert lectura.hoja == "MAYOR"


def test_acepta_sinonimos_de_otros_erp():
    lectura = leer_mayor(
        mayor_xlsx(
            [["1.1.5.1.1", "IVA", "2025-01-05", "A1", 10, 0]],
            encabezado=("Cuenta Contable", "Nombre", "Fecha", "Comprobante",
                        "Débito", "Crédito"),
        )
    )
    assert lectura.mapeo_suficiente
    assert lectura.columnas_detectadas["codigo"] == 0
    assert lectura.columnas_detectadas["cuenta"] == 1
    assert lectura.columnas_detectadas["debe"] == 4


@pytest.mark.parametrize(
    "texto,esperado",
    [
        ("178.259,63", 178259.63),   # europeo
        ("178,259.63", 178259.63),   # US
        ("183724.10", 183724.10),    # plano
        ("-150,00", -150.0),         # negativo con coma decimal
        ("0,00", 0.0),
    ],
)
def test_importes_en_cualquier_formato_regional(texto, esperado):
    lectura = leer_mayor(
        mayor_xlsx([["1.1.5.1.1", "IVA", "2025-01-05", "A1", "", "", "", "",
                     "", texto, None, texto]])
    )
    assert lectura.movimientos[0].debe == esperado


@pytest.mark.parametrize(
    "texto,esperado",
    [
        ("(150.00)", -150.0),        # negativo contable entre parentesis
        ("$ 1,234.56", 1234.56),     # simbolo de moneda
        ("1.234,56 USD", 1234.56),   # sufijo de moneda, formato europeo
        ("-", 0.0),                  # guion como cero
    ],
)
def test_defecto_3_formatos_habituales_de_erp_que_antes_quedaban_en_cero(texto, esperado):
    lectura = leer_mayor(
        mayor_xlsx([["1.1.5.1.1", "IVA", "2025-01-05", "A1", "", "", "", "",
                     "", texto, None, texto]])
    )
    assert lectura.movimientos[0].debe == esperado
    assert lectura.importes_no_parseables == 0


def test_importe_no_parseable_se_cuenta_y_deja_rastro_en_errores():
    lectura = leer_mayor(
        mayor_xlsx([["1.1.5.1.1", "IVA", "2025-01-05", "A1", "", "", "", "",
                     "", "N/D", None, "N/D"]])
    )
    assert lectura.movimientos[0].debe == 0.0
    # "haber" viene vacio (None) y no cuenta como error; "debe" y "saldo"
    # traen "N/D" y ambos si.
    assert lectura.importes_no_parseables == 2
    assert any("N/D" in e for e in lectura.errores)
    assert any("2" in e for e in lectura.errores)  # numero de fila del dato


def test_importes_no_parseables_se_limitan_a_diez_entradas_en_errores():
    filas = [
        ["1.1.5.1.1", "IVA", "2025-01-05", f"A{i}", "", "", "", "", "",
         "N/D", 0, 0]
        for i in range(15)
    ]
    lectura = leer_mayor(mayor_xlsx(filas))
    assert lectura.importes_no_parseables == 15
    assert len(lectura.errores) == 10


def test_descarta_filas_de_total_sin_codigo_de_cuenta():
    filas = [
        ["1.1.5.1.1", "IVA sobre Compras", "2025-01-05", "COM 1", "", "", "",
         "", "", 10, 0, 10],
        [None, "TOTAL GENERAL", None, None, None, None, None, None, None,
         999, 0, 999],
    ]
    lectura = leer_mayor(mayor_xlsx(filas))
    assert len(lectura.movimientos) == 1
    assert lectura.filas_descartadas == 1


def test_descarta_un_encabezado_repetido_a_mitad_del_listado():
    filas = [
        ["1.1.5.1.1", "IVA sobre Compras", "2025-01-05", "COM 1", "", "", "",
         "", "", 10, 0, 10],
        list(ENCABEZADO_REAL),
        ["1.1.5.1.3", "IVA en Importaciones", "2025-01-06", "COM 2", "", "",
         "", "", "", 20, 0, 20],
    ]
    lectura = leer_mayor(mayor_xlsx(filas))
    assert [m.codigo for m in lectura.movimientos] == ["1.1.5.1.1", "1.1.5.1.3"]


def test_descarta_filas_de_total_por_cuenta_que_duplican_el_saldo():
    """Defecto 2: el ERP emite al cierre de cada cuenta una fila con codigo
    pero sin fecha/asiento y los acumulados del periodo; hoy entra como un
    movimiento mas y duplica el debe/haber de la cuenta."""
    filas = [
        ["1.1.5.1.1", "IVA sobre Compras", "2025-01-05", "COM 1", "", "", "",
         "", "", 10, 0, 10],
        ["1.1.5.1.1", "TOTAL CUENTA", None, None, "", "", "", "", "",
         21167.49, 21167.49, 0],
    ]
    lectura = leer_mayor(mayor_xlsx(filas))
    assert len(lectura.movimientos) == 1
    assert lectura.movimientos[0].asiento == "COM 1"
    assert lectura.filas_descartadas == 1


def test_descarta_filas_de_saldo_anterior():
    filas = [
        ["1.1.5.1.1", "SALDO ANTERIOR", None, None, "", "", "", "", "",
         100.0, 0, 100.0],
        ["1.1.5.1.1", "IVA sobre Compras", "2025-01-05", "COM 1", "", "", "",
         "", "", 10, 0, 10],
    ]
    lectura = leer_mayor(mayor_xlsx(filas))
    assert len(lectura.movimientos) == 1
    assert lectura.filas_descartadas == 1


def test_descarta_fila_cuando_la_descripcion_indica_subtotal():
    """La palabra clave puede venir en la glosa/descripcion, no en el nombre."""
    filas = [
        ["1.1.5.1.1", "", None, None, "", "", "", "", "Subtotal cuenta",
         50.0, 0, 50.0],
    ]
    lectura = leer_mayor(mayor_xlsx(filas))
    assert len(lectura.movimientos) == 0
    assert lectura.filas_descartadas == 1


def test_no_descarta_una_fila_legitima_sin_fecha_ni_asiento():
    """Ojo: no toda fila sin fecha es un acumulado. Solo se descarta si el
    nombre/descripcion delata que es TOTAL/SUMA/SUBTOTAL/SALDO."""
    filas = [
        ["1.1.5.1.1", "IVA sobre Compras", None, None, "", "", "", "",
         "AJUSTE MANUAL SIN FECHA", 10, 0, 10],
    ]
    lectura = leer_mayor(mayor_xlsx(filas))
    assert len(lectura.movimientos) == 1
    assert lectura.filas_descartadas == 0


def test_lee_todas_las_hojas_cuando_el_mayor_esta_repartido():
    """Defecto 1: un ERP que exporta una hoja por mes no debe perder las
    demás hojas en silencio."""
    fila_enero = ["1.1.5.1.1", "IVA sobre Compras", "2025-01-05", "COM 1",
                  "", "", "", "", "", 10, 0, 10]
    fila_febrero = ["1.1.5.1.3", "IVA en Importaciones", "2025-02-05", "COM 2",
                     "", "", "", "", "", 20, 0, 20]
    data = mayor_xlsx_multihoja({"ENERO": [fila_enero], "FEBRERO": [fila_febrero]})

    lectura = leer_mayor(data)

    assert len(lectura.movimientos) == 2
    assert lectura.hojas_leidas == ["ENERO", "FEBRERO"]
    assert lectura.hoja == "ENERO"  # la primera hoja leída, por compatibilidad
    assert lectura.filas_descartadas == 0


def test_celda_vacia_de_haber_cuenta_como_cero():
    lectura = leer_mayor(
        mayor_xlsx([["1.1.5.1.1", "IVA", "2025-01-05", "A1", "", "", "", "",
                     "", 2.39, None, 2.39]])
    )
    assert lectura.movimientos[0].haber == 0.0
    assert lectura.movimientos[0].neto == 2.39
