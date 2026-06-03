"""REGLA OBLIGATORIA del proyecto — DATOS F-101 completo y trazable.

Pedido del usuario:
    "se verifique que se traslada TODOS los casilleros con sus códigos,
    nombres, valores — no importa que estén cero, con la finalidad de
    que no haya saldos de líneas [vacías]"

Esta suite de tests garantiza la regla a nivel CÓDIGO (no solo del Excel
descargado). Si alguno falla, el deploy se bloquea.

CAPA A — REGLAS ESTÁTICAS (sin generar Excel):
  1. ALL_F101_CASILLEROS ⊆ F101_CASILLERO_NAMES
     Todo casillero parseado por el regex tiene nombre canónico.
  2. Todos los nombres están NO vacíos.
  3. Rangos críticos de Estado de Situación cubiertos (311-449, 511-699).
  4. Rangos de Resultados cubiertos (6001-7999).

CAPA B — REGLAS E2E (generando Excel):
  5. build_f101_sheet escribe TODOS los casilleros canónicos en DATOS F-101.
  6. NINGUNA fila de DATOS F-101 queda con nombre vacío.
  7. Casilleros NO declarados por el F-101 del cliente igual aparecen
     con valor 0.00 (no se omiten).
  8. Casilleros declarados conservan su valor.
"""

import openpyxl
import pytest

from backend.app.ict.catalogo_f101 import F101_CASILLERO_NAMES, get_casillero_name
from backend.app.ict.parsers.f101_pdf import ALL_F101_CASILLEROS


# ─────────────────────────────────────────────────────────────────────────────
# CAPA A — REGLAS ESTÁTICAS
# ─────────────────────────────────────────────────────────────────────────────

def test_regla_todos_los_casilleros_parseables_tienen_nombre():
    """Si el parser ALL_F101_CASILLEROS contiene un casillero que NO está
    en F101_CASILLERO_NAMES, va a salir SIN nombre en DATOS F-101.
    Esto es el bug que reportó el usuario."""
    faltantes = [c for c in ALL_F101_CASILLEROS if c not in F101_CASILLERO_NAMES]
    assert not faltantes, (
        f"❌ {len(faltantes)} casilleros parseables SIN nombre canónico:\n"
        f"    {faltantes}\n"
        f"Acción: agregar cada uno a backend/app/ict/catalogo_f101.py "
        f"(F101_CASILLERO_NAMES) con su nombre oficial SRI."
    )


def test_regla_nombres_canonicos_no_son_vacios():
    """Aunque un cas esté en el catálogo, su nombre no puede ser ''."""
    vacios = [c for c, n in F101_CASILLERO_NAMES.items() if not n or not n.strip()]
    assert not vacios, (
        f"❌ {len(vacios)} casilleros con nombre vacío en catálogo: {vacios}. "
        f"Revisar backend/app/ict/catalogo_f101.py."
    )


def test_regla_estado_situacion_financiera_completo():
    """Casilleros críticos del Estado de Situación Financiera deben estar.
    Estos son los que el cliente PROPHAR usaba y que faltaban en la imagen
    que envió (437-439, 490-491, 515-518)."""
    criticos = ["437", "438", "439", "490", "491", "515", "516", "517", "518"]
    faltantes = [c for c in criticos if c not in F101_CASILLERO_NAMES]
    assert not faltantes, (
        f"Casilleros del EEFF reportados por el usuario sin nombre: {faltantes}"
    )


def test_regla_totales_cubiertos():
    """Los totales del F-101 (361, 449, 499, 550, 589, 599, 698, 699,
    1005, 1045, 6999, 7991, 7992, 7999) deben tener nombre."""
    totales = ["361", "449", "499", "550", "589", "599", "698", "699",
               "1005", "1045", "6999", "7991", "7992", "7999"]
    for cas in totales:
        assert cas in F101_CASILLERO_NAMES, f"Total {cas} sin nombre"
        nombre = F101_CASILLERO_NAMES[cas]
        assert "TOTAL" in nombre.upper(), \
            f"Cas {cas} debería decir 'TOTAL' en su nombre, dice: {nombre!r}"


def test_get_casillero_name_devuelve_nombre_o_fallback():
    """API helper get_casillero_name no debe lanzar excepción nunca."""
    assert get_casillero_name("311") == "EFECTIVO Y EQUIVALENTES AL EFECTIVO"
    assert get_casillero_name("574") == "DESAHUCIO"
    # Casillero inexistente devuelve fallback
    assert get_casillero_name("99999") == ""
    assert get_casillero_name("99999", fallback="?") == "?"


def test_regla_cobertura_minima_del_catalogo():
    """Defensa: el catálogo debe tener al menos 300 casilleros (el F-101
    real tiene unos 360+). Si baja drásticamente es una regresión."""
    assert len(F101_CASILLERO_NAMES) >= 300, (
        f"Catálogo F-101 demasiado chico: {len(F101_CASILLERO_NAMES)} entradas. "
        f"Probablemente alguien borró entradas por error."
    )


# ─────────────────────────────────────────────────────────────────────────────
# CAPA B — REGLAS E2E (generando Excel)
# ─────────────────────────────────────────────────────────────────────────────

def _build_test_wb_with_datos_f101(f101_dict: dict) -> openpyxl.Workbook:
    from backend.app.ict.fillers.source_data_sheets import build_f101_sheet
    wb = openpyxl.Workbook()
    # Eliminar la hoja por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    # Crear hoja DATOS F-101
    build_f101_sheet(wb, f101_dict, {})
    return wb


def test_e2e_datos_f101_tiene_todos_los_casilleros_canonicos():
    """Aún con F-101 vacío, DATOS F-101 debe tener una fila por cada
    casillero canónico — esto es el control de saldos de líneas vacías."""
    wb = _build_test_wb_with_datos_f101({})
    assert "DATOS F-101" in wb.sheetnames
    ws = wb["DATOS F-101"]

    casilleros_en_hoja = set()
    for r in range(4, ws.max_row + 1):
        cas = ws.cell(r, 1).value
        if cas is not None and str(cas).strip():
            casilleros_en_hoja.add(str(cas).strip())

    faltantes = set(F101_CASILLERO_NAMES.keys()) - casilleros_en_hoja
    assert not faltantes, (
        f"DATOS F-101 NO tiene {len(faltantes)} casilleros del catálogo: "
        f"{sorted(faltantes)[:20]}"
    )


def test_e2e_ninguna_fila_DATOS_F101_sin_nombre():
    """Después de build_f101_sheet, recorrer la hoja y verificar que
    NINGUNA fila tenga columna B vacía. Esto es la REGLA del usuario."""
    # Probamos con f101 con valores variados
    f101 = {"311": 5000.0, "574": 341311.19, "490": 94089.36}
    wb = _build_test_wb_with_datos_f101(f101)
    ws = wb["DATOS F-101"]

    filas_sin_nombre = []
    for r in range(4, ws.max_row + 1):
        cas = ws.cell(r, 1).value
        nombre = ws.cell(r, 2).value
        if cas is None or str(cas).strip() == "":
            continue  # fila vacía
        if nombre is None or str(nombre).strip() == "":
            filas_sin_nombre.append((r, cas))

    assert not filas_sin_nombre, (
        f"❌ {len(filas_sin_nombre)} filas en DATOS F-101 SIN NOMBRE: "
        f"{filas_sin_nombre[:10]}"
    )


def test_e2e_casilleros_no_declarados_aparecen_con_valor_cero():
    """Casillero que el F-101 NO declara igual debe aparecer en DATOS
    F-101 con valor 0.00 (no se omite). Esto evita 'saldos de líneas'
    vacías que reportó el usuario."""
    f101 = {"311": 1000.0}  # solo 1 casillero
    wb = _build_test_wb_with_datos_f101(f101)
    ws = wb["DATOS F-101"]

    # Buscar cas 574 (que no estaba en f101) — debe estar con 0
    encontrado_574 = False
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == "574":
            encontrado_574 = True
            val = ws.cell(r, 3).value
            assert val == 0 or val == 0.0, \
                f"Cas 574 no declarado debe tener valor 0, tiene {val!r}"
            nombre = ws.cell(r, 2).value
            assert nombre == "DESAHUCIO", f"Cas 574 nombre incorrecto: {nombre!r}"
            break
    assert encontrado_574, "Cas 574 (canónico) NO aparece en DATOS F-101"


def test_e2e_casilleros_declarados_conservan_valor():
    """Lo que el F-101 declara sí debe pasar al Excel."""
    f101 = {"574": 341311.19, "311": 8500.0}
    wb = _build_test_wb_with_datos_f101(f101)
    ws = wb["DATOS F-101"]

    valores_esperados = {"574": 341311.19, "311": 8500.0}
    encontrados = {}
    for r in range(4, ws.max_row + 1):
        cas = str(ws.cell(r, 1).value or "").strip()
        if cas in valores_esperados:
            encontrados[cas] = ws.cell(r, 3).value

    for cas, expected in valores_esperados.items():
        assert encontrados.get(cas) == expected, (
            f"Cas {cas}: esperaba {expected}, encontrado {encontrados.get(cas)!r}"
        )


def test_e2e_casilleros_extras_del_pdf_no_catalogados_aparecen_al_final():
    """Caso edge: si el PDF trae un casillero que no está en el catálogo
    canónico (ej. SRI publica casillero nuevo y aún no actualizamos),
    debe aparecer al final con observación '⚠ no catalogado' — no se pierde."""
    f101 = {"311": 5000, "99999": 12345.67}  # 99999 no existe en catálogo
    wb = _build_test_wb_with_datos_f101(f101)
    ws = wb["DATOS F-101"]

    encontrado_extra = False
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == "99999":
            encontrado_extra = True
            assert ws.cell(r, 3).value == 12345.67
            obs = str(ws.cell(r, 4).value or "")
            assert "no catalogado" in obs.lower(), \
                f"Observación debe avisar que no está catalogado: {obs!r}"
            break
    assert encontrado_extra, "Cas extra 99999 del PDF no aparece en hoja"
