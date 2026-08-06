"""DM8 ATS: cruza el Anexo Transaccional contra lo declarado en F-104/F-103."""

import re

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.cedulas.dm8_ats import (
    CASILLEROS_IMPORTACIONES,
    CASILLERO_COMPRAS_TOTAL,
    CASILLERO_IVA_COMPRAS,
    CASILLERO_IVA_RETENIDO_RECIBIDO,
    CASILLERO_IVA_VENTAS,
    CASILLERO_VENTAS_GRAVADAS,
    CASILLEROS_VENTAS_0,
    SHEET_DM8,
    build_dm8,
)

PERIODOS = [f"2025-{m:02d}" for m in range(1, 13)]

CAMPOS_ATS_FIJOS = [
    "ventas_bi_0", "ventas_bi_gravada", "ventas_iva",
    "compras_bi_0", "compras_bi_gravada", "compras_iva",
    "anulados", "iva_le_retuvieron", "renta_le_retuvieron",
    "ret_renta_total", "ret_iva_total",
]

DIR_ATS = {
    **{(campo, "01"): f"'DATOS ATS'!B{i}" for i, campo in enumerate(CAMPOS_ATS_FIJOS, start=4)},
    ("renta_codigo:303", "01"): "'DATOS ATS'!B90",
    ("renta_codigo:310", "01"): "'DATOS ATS'!B91",
    ("iva_pct:10", "01"): "'DATOS ATS'!B92",
    ("iva_pct:20", "01"): "'DATOS ATS'!B93",
    ("iva_pct:30", "01"): "'DATOS ATS'!B94",
    ("iva_pct:50", "01"): "'DATOS ATS'!B95",
    ("iva_pct:70", "01"): "'DATOS ATS'!B96",
    ("iva_pct:100", "01"): "'DATOS ATS'!B97",
    ("iva_pct:NC", "01"): "'DATOS ATS'!B98",
}

_CAS_F104 = (
    CASILLEROS_VENTAS_0
    + [CASILLERO_VENTAS_GRAVADAS, CASILLERO_IVA_VENTAS, CASILLERO_COMPRAS_TOTAL,
       CASILLERO_IVA_COMPRAS, CASILLERO_IVA_RETENIDO_RECIBIDO]
    + CASILLEROS_IMPORTACIONES
    + ["721", "723", "725", "727", "729", "731", "799"]
)
DIR_F104 = {("2025-01", cas): f"'DATOS F-104'!C{i}" for i, cas in enumerate(_CAS_F104, start=20)}
DIR_F103 = {("2025-01", "499"): "'DATOS F-103'!C40"}


def _cedula(**kw):
    wb = Workbook()
    datos = dict(dir_ats=DIR_ATS, dir_f104=DIR_F104, dir_f103=DIR_F103,
                 periodos=PERIODOS, cliente="C", periodo="2025")
    datos.update(kw)
    build_dm8(wb, **datos)
    return wb[SHEET_DM8]


def _etiquetas(ws):
    return [str(ws.cell(r, 2).value or "") for r in range(1, ws.max_row + 1)]


def test_lleva_el_encabezado_de_cedula_con_su_referencia():
    ws = _cedula()
    valores = [ws.cell(r, c).value for r in range(1, 11) for c in range(1, 6)]
    assert "OBLIGACIONES FISCALES" in valores
    assert "DM8" in valores


def test_ventas_0_segun_ats_es_formula_a_datos_ats():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(r, 2).value == "Ventas 0% según ATS")
    valor = ws.cell(fila, 3).value
    assert isinstance(valor, str) and valor.startswith("='DATOS ATS'!")


def test_ventas_gravadas_se_compara_contra_casillero_411():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    assert any(e.startswith(f"Casillero {CASILLERO_VENTAS_GRAVADAS}") for e in etiquetas)


def test_iva_en_ventas_se_compara_contra_casillero_421():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    assert any(e.startswith(f"Casillero {CASILLERO_IVA_VENTAS}") for e in etiquetas)


def test_compras_gravadas_se_compara_contra_casillero_519():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    assert any(e.startswith(f"Casillero {CASILLERO_COMPRAS_TOTAL}") for e in etiquetas)


def test_iva_en_compras_se_compara_contra_casillero_520():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    assert any(e.startswith(f"Casillero {CASILLERO_IVA_COMPRAS}") for e in etiquetas)


def test_compras_0_resta_importaciones_del_casillero_519():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    for cas in CASILLEROS_IMPORTACIONES:
        assert any(e.startswith(f"Casillero {cas}") for e in etiquetas), cas
    fila_import_total = next(r for r in range(1, ws.max_row + 1)
                              if ws.cell(r, 2).value == "Total importaciones")
    assert ws.cell(fila_import_total, 3).value.startswith("=SUM(")


def test_comprobantes_anulados_es_informativo_sin_casillero():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if str(ws.cell(r, 2).value or "").startswith("Anulados según ATS"))
    valor = ws.cell(fila, 3).value
    assert isinstance(valor, str) and valor.startswith("='DATOS ATS'!")


def test_retencion_de_iva_tiene_una_fila_por_porcentaje_con_su_casillero():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    for pct, cas in (("10", "721"), ("20", "723"), ("30", "725"),
                     ("50", "727"), ("70", "729"), ("100", "731")):
        assert any(e.startswith(f"Ret. IVA {pct}% según ATS") for e in etiquetas), pct
        assert any(e.startswith(f"Casillero {cas}") for e in etiquetas), cas


def test_retencion_de_iva_nc_es_informativa_sin_casillero():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    assert any("NC" in e and "informativo" in e for e in etiquetas)


def test_retencion_de_iva_total_control_es_el_casillero_799():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    assert any(e.startswith("Casillero 799") for e in etiquetas)


def test_retencion_de_renta_tiene_una_fila_por_codigo_visto_en_los_ats():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    assert any("303" in e for e in etiquetas)
    assert any("310" in e for e in etiquetas)


def test_retencion_de_renta_se_compara_contra_casillero_499_del_f103():
    ws = _cedula()
    fila = next(r for r in range(1, ws.max_row + 1)
                if "Casillero 499" in str(ws.cell(r, 2).value or ""))
    valor = ws.cell(fila, 3).value
    assert valor == "='DATOS F-103'!C40"


def test_le_efectuaron_iva_se_compara_contra_casillero_609():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    assert any(e.startswith(f"Casillero {CASILLERO_IVA_RETENIDO_RECIBIDO}") for e in etiquetas)


def test_le_efectuaron_renta_es_informativo():
    ws = _cedula()
    etiquetas = _etiquetas(ws)
    assert any("Renta que le retuvieron" in e and "informativo" in e for e in etiquetas)


def test_todas_las_diferencias_restan_y_redondean():
    ws = _cedula()
    # Solo las filas de datos: "Diferencia" o "Diferencia 70%". Se excluye
    # la leyenda del pie ("Diferencia determinada"), que describe la marca
    # de auditoría y por definición no lleva fórmula.
    es_fila_diferencia = re.compile(r"^Diferencia( \d+(?:[.,]\d+)?%)?$").match
    filas_diff = [r for r in range(1, ws.max_row + 1)
                  if es_fila_diferencia(str(ws.cell(r, 2).value or ""))]
    assert len(filas_diff) >= 8
    for fila in filas_diff:
        valor = ws.cell(fila, 3).value
        assert isinstance(valor, str) and valor.startswith("=ROUND(")


def test_ningun_dato_se_escribe_como_valor_literal_en_la_cedula():
    """Regla innegociable: todo importe que viene de otra hoja va como
    fórmula, nunca como valor pegado."""
    ws = _cedula()
    for fila in range(13, ws.max_row + 1):
        for col in range(3, 16):
            valor = ws.cell(fila, col).value
            if isinstance(valor, (int, float)):
                assert valor == 0, (fila, col, valor)
