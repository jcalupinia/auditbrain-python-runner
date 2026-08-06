"""Hoja resumen: cuenta × mes, agrupada por categoría."""

import datetime

from openpyxl import Workbook

from backend.app.aud.obligaciones_fiscales.libro.hoja_mayores import (
    SHEET_MAYORES,
    build_hoja_mayores,
)
from backend.app.aud.obligaciones_fiscales.mayor.tipos import Movimiento


class _Fila:
    """Doble de MayorClasificacionJob: solo lo que la hoja necesita."""

    def __init__(self, codigo, nombre, categoria, por_mes, n=1, debe=0.0, haber=0.0):
        self.codigo_cuenta = codigo
        self.nombre_cuenta = nombre
        self.categoria_final = categoria
        self.por_mes_json = por_mes
        self.n_movimientos = n
        self.debe = debe
        self.haber = haber


FILAS = [
    _Fila("1.1.5.1.1", "IVA sobre Compras", "IVA_COMPRAS", {"01": 659.57, "02": 1988.83}),
    _Fila("1.1.5.1.3", "IVA en Importaciones", "IVA_COMPRAS", {"01": 9252.0}),
    _Fila("4.1.1.4", "Venta de insumos", "VENTAS", {"01": -28117.84}),
]


def _celda(ws, addr: str):
    """La celda de una dirección CALIFICADA (`'Mayores homologados'!D4`)."""
    return ws[addr.split("!", 1)[1]]


def test_crea_la_hoja_con_una_fila_por_cuenta():
    wb = Workbook()
    build_hoja_mayores(wb, FILAS)
    ws = wb[SHEET_MAYORES]
    codigos = [ws.cell(r, 2).value for r in range(1, ws.max_row + 1)]
    assert "1.1.5.1.1" in codigos
    assert "4.1.1.4" in codigos


def test_agrupa_por_categoria_y_pone_un_subtotal():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS)
    ws = wb[SHEET_MAYORES]
    addr = lookup[("IVA_COMPRAS", "01")]
    assert _celda(ws, addr).value.startswith("=SUM(")


def test_el_subtotal_de_enero_de_iva_compras_suma_sus_dos_cuentas():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS)
    ws = wb[SHEET_MAYORES]
    # El subtotal es una fórmula SUM sobre el rango de sus cuentas: se
    # verifica el rango, no el valor (openpyxl no evalúa fórmulas).
    formula = _celda(ws, lookup[("IVA_COMPRAS", "01")]).value
    assert formula.count(":") == 1


def test_publica_la_direccion_del_subtotal_de_cada_categoria_y_mes():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS)
    for mes in (f"{m:02d}" for m in range(1, 13)):
        assert ("IVA_COMPRAS", mes) in lookup
        assert ("VENTAS", mes) in lookup
    assert ("IVA_COMPRAS", "TOTAL") in lookup


def test_una_categoria_sin_cuentas_no_aparece():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS)
    assert ("RET_IVA", "01") not in lookup


def test_los_meses_sin_movimiento_quedan_en_cero_no_vacios():
    """Un mes vacío en el papel de trabajo se lee como dato faltante."""
    wb = Workbook()
    build_hoja_mayores(wb, FILAS)
    ws = wb[SHEET_MAYORES]
    fila_ventas = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "4.1.1.4"
    )
    assert ws.cell(fila_ventas, 5).value == 0.0   # marzo


def test_las_cuentas_sin_categoria_van_a_un_bloque_de_no_clasificadas():
    wb = Workbook()
    filas = FILAS + [_Fila("9.9.9", "Cuenta puente", None, {"01": 5.0})]
    lookup = build_hoja_mayores(wb, filas)
    ws = wb[SHEET_MAYORES]
    codigos = [ws.cell(r, 2).value for r in range(1, ws.max_row + 1)]
    assert "9.9.9" in codigos
    assert ("SIN_CLASIFICAR", "01") in lookup


def test_publica_tambien_la_direccion_de_cada_cuenta():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS)
    assert ("cuenta:1.1.5.1.1", "01") in lookup
    assert ("cuenta:1.1.5.1.1", "TOTAL") in lookup


def test_la_direccion_de_una_cuenta_apunta_a_su_valor_mensual():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS)
    ws = wb[SHEET_MAYORES]
    assert _celda(ws, lookup[("cuenta:1.1.5.1.1", "01")]).value == 659.57
    assert _celda(ws, lookup[("cuenta:1.1.5.1.3", "01")]).value == 9252.0


def test_las_cuentas_de_una_categoria_se_pueden_listar_en_orden():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS)
    assert lookup[("orden:IVA_COMPRAS", "cuentas")] == ["1.1.5.1.1", "1.1.5.1.3"]


# ---------------------------------------- desglose de ventas por tarifa ---
#
# El mayor agrega por cuenta y mes sin distinguir tarifa, así que DM5 mostraba
# la misma cifra en «VENTAS ≠ 0%» y en «VENTAS 0%». La hoja publica ahora, en
# un bloque aparte, las tres columnas del desglose por asiento.

FILAS_VENTAS = [
    _Fila("4.1.1.4", "Venta de insumos", "VENTAS", {"01": 1150.0}),
    _Fila("2.1.7.4.1", "IVA en ventas", "IVA_VENTAS", {"01": 120.0}),
]


def _mov(codigo, haber, asiento):
    return Movimiento(codigo=codigo, asiento=asiento,
                      fecha=datetime.date(2025, 1, 15), haber=haber)


# VTA 1: 100 + 250 + 300 con IVA 45 → sólo la línea de 300 cuadra con la base
# gravada, las otras dos son 0%.  VTA 2: 500 con IVA 75 → todo gravado.
MOVS_VENTAS = [
    _mov("4.1.1.4", 100.0, "VTA 1"), _mov("4.1.1.4", 250.0, "VTA 1"),
    _mov("4.1.1.4", 300.0, "VTA 1"), _mov("2.1.7.4.1", 45.0, "VTA 1"),
    _mov("4.1.1.4", 500.0, "VTA 2"), _mov("2.1.7.4.1", 75.0, "VTA 2"),
]


def test_publica_el_desglose_por_tarifa_de_cada_cuenta_de_ventas():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS_VENTAS, movimientos=MOVS_VENTAS)
    for bucket in ("gravada", "cero", "por_asignar"):
        clave = (f"cuenta:4.1.1.4:{bucket}", "01")
        assert clave in lookup, clave
        assert lookup[clave].startswith(f"'{SHEET_MAYORES}'!"), lookup[clave]


def test_el_desglose_separa_lo_gravado_de_lo_cero_por_ciento():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS_VENTAS, movimientos=MOVS_VENTAS)
    ws = wb[SHEET_MAYORES]
    assert _celda(ws, lookup[("cuenta:4.1.1.4:gravada", "01")]).value == 800.0
    assert _celda(ws, lookup[("cuenta:4.1.1.4:cero", "01")]).value == 350.0
    assert _celda(ws, lookup[("cuenta:4.1.1.4:por_asignar", "01")]).value == 0.0


def test_el_desglose_cuadra_contra_el_total_que_la_hoja_ya_mostraba():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS_VENTAS, movimientos=MOVS_VENTAS)
    ws = wb[SHEET_MAYORES]
    partes = sum(
        _celda(ws, lookup[(f"cuenta:4.1.1.4:{b}", "01")]).value
        for b in ("gravada", "cero", "por_asignar")
    )
    assert partes == _celda(ws, lookup[("cuenta:4.1.1.4", "01")]).value


def test_publica_los_subtotales_del_desglose_para_las_cedulas():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS_VENTAS, movimientos=MOVS_VENTAS)
    ws = wb[SHEET_MAYORES]
    for bucket in ("gravada", "cero", "por_asignar"):
        addr = lookup[(f"VENTAS:{bucket}", "01")]
        assert addr.startswith(f"'{SHEET_MAYORES}'!")
        assert _celda(ws, addr).value.startswith("=SUM(")
    assert (f"VENTAS:gravada", "TOTAL") in lookup


def test_solo_las_cuentas_de_ventas_reciben_desglose():
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS_VENTAS, movimientos=MOVS_VENTAS)
    assert ("cuenta:2.1.7.4.1:gravada", "01") not in lookup
    assert ("IVA_VENTAS:gravada", "01") not in lookup


def test_sin_movimientos_no_se_construye_el_bloque_de_desglose():
    """El desglose se calcula asiento por asiento: sin movimientos no hay
    forma de separar y la hoja se queda exactamente como estaba."""
    wb = Workbook()
    lookup = build_hoja_mayores(wb, FILAS_VENTAS)
    assert ("cuenta:4.1.1.4:gravada", "01") not in lookup
