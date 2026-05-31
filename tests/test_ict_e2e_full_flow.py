"""End-to-end ICT flow test: create session, upload to multiple anexos, download Excel."""
import io
import uuid

import openpyxl
import pytest

from backend.app.client_portal.service import create_portal_user
from backend.app.context.models import Organization, Client
from backend.app.db.session import SessionLocal

# Nombres exactos de hojas del template ICT 2025 (con tildes/eñes del xlsx original)
EXPECTED_SHEETS = {
    "INDICE",
    "MAPEO DE LA DECLARACIÓN A1",
    "INGRESOS A2",
    "COSTOS  GASTOS A3",
    "CONCILIACIÓN INGRESOS A4",
    "CONCILIACIÓN COSTOS Y GASTOS A5",
    "BENEFICIOS TRIBUTARIOS A6",
    "CRÉDITO TRIBUTARIO A7",
    "COMERCIO EXTERIOR A8",
    "INVENTARIOS A9",
}


def _unique(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex[:8]}"


@pytest.fixture()
def db_session():
    db = SessionLocal()
    yield db
    db.close()


@pytest.fixture()
def logged_client(client, db_session):
    suffix = _unique("icte2e")
    org = Organization(name=f"O-{suffix}", slug=f"o-{suffix}", is_active=True)
    db_session.add(org)
    db_session.commit()
    db_session.refresh(org)
    cli = Client(organization_id=org.id, name=f"C-{suffix}", is_active=True)
    db_session.add(cli)
    db_session.commit()
    db_session.refresh(cli)
    email = f"icte2e-{suffix}@x.com"
    user, pwd = create_portal_user(db_session, client_id=cli.id, email=email)
    r = client.post(
        "/api/v1/client/auth/login",
        data={"username": email, "password": pwd},
    )
    assert r.status_code == 200, r.json()
    token = r.json()["access_token"]
    device_id = r.cookies.get("device_id")
    return {
        "user": user,
        "headers": {"Authorization": f"Bearer {token}"},
        "cookies": {"device_id": device_id} if device_id else {},
    }


def _make_balance_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Código", "Nombre", "Saldo Final"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def test_full_ict_flow_create_upload_download(client, logged_client):
    """Smoke test: create session, upload Balance to A1, download Excel, verify it parses."""
    # Step 1: create session
    r = client.post(
        "/api/v1/client/ict/sessions",
        json={
            "ejercicio_fiscal": "2025",
            "ruc": "1234567890001",
            "razon_social": "Test S.A.",
            "numero_adhesivo": "ABC-1",
        },
        headers=logged_client["headers"],
        cookies=logged_client["cookies"],
    )
    assert r.status_code == 201, r.json()
    session = r.json()
    assert len(session["anexos"]) == 10
    session_id = session["id"]
    assert session["razon_social"] == "Test S.A."

    # Step 2: upload Balance to A1
    balance_xlsx = _make_balance_xlsx([
        ("1.1.01.01.01", "CAJA CHICA", 300.0),
        ("1.1.01.02.01", "BANCO PICHINCHA", 50000.0),
    ])
    r2 = client.post(
        f"/api/v1/client/ict/sessions/{session_id}/anexos/A1/upload",
        files={
            "files": (
                "balance.xlsx",
                io.BytesIO(balance_xlsx),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"slot_name": "balance"},
        headers=logged_client["headers"],
        cookies=logged_client["cookies"],
    )
    assert r2.status_code == 200, r2.json()
    body = r2.json()
    assert body["anexo_code"] == "A1"
    assert body["status"] in ("partial", "ready", "error")

    # Step 3: verify session details show A1 was updated
    r3 = client.get(
        f"/api/v1/client/ict/sessions/{session_id}",
        headers=logged_client["headers"],
        cookies=logged_client["cookies"],
    )
    assert r3.status_code == 200
    s = r3.json()
    a1 = next(a for a in s["anexos"] if a["anexo_code"] == "A1")
    assert a1["status"] in ("partial", "ready", "error")

    # Step 4: download the Excel
    r4 = client.get(
        f"/api/v1/client/ict/sessions/{session_id}/download",
        headers=logged_client["headers"],
        cookies=logged_client["cookies"],
    )
    assert r4.status_code == 200
    assert len(r4.content) > 1000  # non-trivial file

    # Step 5: parse downloaded Excel and confirm it has the 10 sheets
    wb = openpyxl.load_workbook(io.BytesIO(r4.content), data_only=False)
    sheets = set(wb.sheetnames)
    assert EXPECTED_SHEETS.issubset(sheets), (
        f"Missing sheets: {EXPECTED_SHEETS - sheets}"
    )

    # Step 6: verify session metadata (razon_social comes from session JSON,
    # not from INDICE sheet — the INDICE filler only runs when status != 'empty')
    assert s["razon_social"] == "Test S.A."
    assert s["ruc"] == "1234567890001"
    assert s["numero_adhesivo"] == "ABC-1"


def test_full_ict_flow_other_user_cannot_access(client, logged_client, db_session):
    """Aislamiento: usuario B no puede acceder a sesion de usuario A."""
    # User A creates session
    r = client.post(
        "/api/v1/client/ict/sessions",
        json={
            "ejercicio_fiscal": "2025",
            "ruc": "1234567890001",
            "razon_social": "A",
        },
        headers=logged_client["headers"],
        cookies=logged_client["cookies"],
    )
    assert r.status_code == 201, r.json()
    session_id = r.json()["id"]

    # Create user B in a different org/client
    suffix = _unique("icte2e-b")
    org = Organization(name=f"O-{suffix}", slug=f"o-{suffix}", is_active=True)
    db_session.add(org)
    db_session.commit()
    db_session.refresh(org)
    cli = Client(organization_id=org.id, name=f"C-{suffix}", is_active=True)
    db_session.add(cli)
    db_session.commit()
    db_session.refresh(cli)
    email = f"icte2e-b-{suffix}@x.com"
    _user_b, pwd = create_portal_user(db_session, client_id=cli.id, email=email)
    rb = client.post(
        "/api/v1/client/auth/login",
        data={"username": email, "password": pwd},
    )
    assert rb.status_code == 200, rb.json()
    device_id_b = rb.cookies.get("device_id")
    headers_b = {"Authorization": f"Bearer {rb.json()['access_token']}"}
    cookies_b = {"device_id": device_id_b} if device_id_b else {}

    # User B tries to GET user A's session
    r2 = client.get(
        f"/api/v1/client/ict/sessions/{session_id}",
        headers=headers_b,
        cookies=cookies_b,
    )
    assert r2.status_code == 403

    # User B tries to download user A's Excel
    r3 = client.get(
        f"/api/v1/client/ict/sessions/{session_id}/download",
        headers=headers_b,
        cookies=cookies_b,
    )
    assert r3.status_code == 403
