"""Genera el XML de la Declaracion Patrimonial del SRI a partir del libro Excel
producido por declaracion_patrimonial_excel.py.

Lee las hojas de los 9 modulos, toma los valores de la columna del anio
indicado y reconstruye el XML respetando las etiquetas y el orden del esquema
del SRI (etiquetas confirmadas con XML reales del SRI).

Uso:
    python tools/generar_xml_sri.py LIBRO.xlsx [SALIDA.xml] [--anio=AAAA]
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from declaracion_patrimonial_excel import (  # noqa: E402
    MODULOS, ORDEN_XML, desc_cols, valor_tag)

import re  # noqa: E402

ENCODING_XML = "ISO-8859-1"
MOD = {m["key"]: m for m in MODULOS}

# Etiquetas XML que deben limpiarse (SRI exige solo letras/digitos/espacios).
_NAME_TAGS = {"nombre", "nombreCony", "nombreDeudor", "nombreAcreedor",
              "nombreIfiExterior", "nombempresa", "descripcion"}
_RX_NOM = re.compile(r"[^A-Za-z\xf1\xd10-9\s]+")


def _saneo_nombre(s: str) -> str:
    s = s.replace("&", " Y ")
    s = _RX_NOM.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:60]   # SRI: maximo 60 caracteres


def _code(v):
    if v is None:
        return ""
    s = str(v).strip()
    return s.split(" - ", 1)[0].strip() if " - " in s else s


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _esc(t):
    return (str(t).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;"))


def _columnas_valor(ws):
    """Devuelve {anio(str): indice_columna} leyendo la fila 5."""
    cols = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=5, column=c).value
        if h and str(h).startswith("Valor "):
            cols[str(h).replace("Valor", "").strip()] = c
    return cols


def _leer_modulo(wb, modulo, anio):
    """Lee las filas de detalle de la hoja de un modulo.

    Devuelve lista de (desc_valores, valor) en el orden de desc_cols(modulo).
    """
    ws = wb[modulo["hoja"]]
    nd = len(desc_cols(modulo))
    colv = _columnas_valor(ws)
    col_val = colv.get(str(anio)) or min(colv.values())
    filas = []
    r = 6
    while r <= ws.max_row:
        if ws.cell(row=r, column=1).value == "TOTAL":
            break
        desc = [ws.cell(row=r, column=c).value for c in range(1, nd + 1)]
        valor = _num(ws.cell(row=r, column=col_val).value)
        filas.append((desc, valor))
        r += 1
    return filas


def _datos_generales(wb):
    ws = wb["Datos del XML"]
    datos = {}
    for r in range(1, ws.max_row + 1):
        etiqueta = ws.cell(row=r, column=1).value
        if etiqueta:
            datos[str(etiqueta).strip()] = ws.cell(row=r, column=2).value
    return datos


def _justificacion(wb):
    """Codigos de justificacion marcados con 'Si' en la hoja Justificacion."""
    if "Justificacion" not in wb.sheetnames:
        return []
    ws = wb["Justificacion"]
    codigos = []
    for r in range(1, ws.max_row + 1):
        etiqueta = ws.cell(row=r, column=2).value
        marca = ws.cell(row=r, column=5).value
        if (etiqueta and " - " in str(etiqueta) and
                str(marca).strip().lower() in ("si", "sí", "x", "true")):
            codigos.append(_code(etiqueta))
    return codigos


def _detalle_xml(modulo, desc, valor):
    """Construye un bloque <detalle...> respetando el orden del esquema."""
    lineas = [f"<{modulo['detalle']}>"]
    di = 0
    for _h, _w, _cat, xml_tag, tipo in modulo["cols"]:
        if tipo == "val":
            lineas.append(f"    <{xml_tag}>{valor:.2f}</{xml_tag}>")
            continue
        crudo = desc[di]
        di += 1
        if xml_tag == "_ifi":
            if not crudo or str(crudo).strip() == "":
                continue
            txt = str(crudo).strip()
            if " - " in txt and _code(txt).isdigit():
                lineas.append(f"    <ifiEcuador>{_code(txt)}</ifiEcuador>")
            else:
                lineas.append(f"    <nombreIfiExterior>{_esc(txt)}"
                               f"</nombreIfiExterior>")
            continue
        if crudo is None or str(crudo).strip() == "":
            continue
        # Numeros largos (claveCat, identificaciones) llegan como float en
        # notacion cientifica; convertirlos a entero plano.
        if tipo == "txt" and isinstance(crudo, float) and crudo.is_integer():
            crudo = int(crudo)
        val = _code(crudo) if tipo == "cod" else str(crudo).strip()
        if xml_tag in _NAME_TAGS:
            val = _saneo_nombre(val)
        lineas.append(f"    <{xml_tag}>{_esc(val)}</{xml_tag}>")
    lineas.append(f"</{modulo['detalle']}>")
    return "\n".join(lineas)


def generar_xml(libro: Path, anio: int) -> str:
    wb = load_workbook(libro, data_only=True)
    dg = _datos_generales(wb)

    tipo_dec = _code(dg.get("Tipo de declaracion"))
    tipo_ident = _code(dg.get("Tipo de identificacion"))
    num_ident = dg.get("Numero de identificacion") or ""
    nombre = _saneo_nombre(str(dg.get("Nombre del declarante") or ""))
    tipo_ident_cony = _code(dg.get("Identificacion del conyuge"))
    num_ident_cony = dg.get("Numero ident. conyuge") or ""
    nombre_cony = _saneo_nombre(str(dg.get("Nombre del conyuge") or ""))
    regularizacion = _code(dg.get("Regularizacion de activos"))

    # filas de cada modulo
    filas = {k: _leer_modulo(wb, MOD[k], anio) for k in MOD}

    # Solo se suman los valores positivos: el SRI no acepta saldos <= 0 en los
    # detalles, y la suma debe cuadrar con los detalles emitidos en el XML.
    def total(key):
        return round(sum(v for _d, v in filas[key] if v > 0), 2)

    total_creditos = total("cxc")
    total_derechos = total("derechos")
    total_activos = sum(total(m["key"]) for m in MODULOS
                        if m["clase"] == "activo")
    total_pasivos = sum(total(m["key"]) for m in MODULOS
                        if m["clase"] == "pasivo")
    total_declarado = round(total_activos - total_pasivos, 2)

    # patrimonio del anio anterior
    ws_dinero = wb["Dinero"]
    anios_libro = {int(a) for a in _columnas_valor(ws_dinero) if a.isdigit()}
    if (anio - 1) in anios_libro:
        prev = {k: _leer_modulo(wb, MOD[k], anio - 1) for k in MOD}
        act_prev = sum(round(sum(v for _d, v in prev[m["key"]] if v > 0), 2)
                       for m in MODULOS if m["clase"] == "activo")
        pas_prev = sum(round(sum(v for _d, v in prev[m["key"]] if v > 0), 2)
                       for m in MODULOS if m["clase"] == "pasivo")
        anio_anterior = round(act_prev - pas_prev, 2)
    else:
        anio_anterior = _num(dg.get("Patrimonio del anio anterior"))
    diff = round(total_declarado - anio_anterior, 2)

    # reparto del patrimonio (informativo): se respeta el del XML si cuadra
    atr = _num(dg.get("Atribuible a hijos no emancipados"))
    soc = _num(dg.get("Patrimonio en la sociedad conyugal"))
    ind = _num(dg.get("Patrimonio individual del declarante"))
    if abs(atr + soc + ind - total_declarado) > 0.01:
        atr = 0.0
        if tipo_dec == "SOC":
            soc, ind = total_declarado, 0.0
        else:
            soc, ind = 0.0, total_declarado

    out = ['<?xml version="1.0" encoding="ISO-8859-1" standalone="yes"?>',
           "<decPat>"]
    out.append(f"    <anio>{anio}</anio>")
    out.append(f"    <tipoDec>{_esc(tipo_dec)}</tipoDec>")
    out.append(f"    <tipoIdent>{_esc(tipo_ident)}</tipoIdent>")
    out.append(f"    <numIdent>{_esc(num_ident)}</numIdent>")
    out.append(f"    <nombre>{_esc(nombre)}</nombre>")
    if num_ident_cony:
        out.append(f"    <tipoIdentCony>{_esc(tipo_ident_cony)}"
                   f"</tipoIdentCony>")
        out.append(f"    <numIdentCony>{_esc(num_ident_cony)}</numIdentCony>")
        out.append(f"    <nombreCony>{_esc(nombre_cony)}</nombreCony>")
    if regularizacion:
        out.append(f"    <regularizacionActivos>{_esc(regularizacion)}"
                   f"</regularizacionActivos>")
    out.append(f"    <totalCreditos>{total_creditos:.2f}</totalCreditos>")
    if any(v > 0 for _d, v in filas["derechos"]):
        out.append(f"    <totalDerechos>{total_derechos:.2f}</totalDerechos>")

    out.append("    <patrimonio>")
    out.append(f"        <totalDeclarado>{total_declarado:.2f}"
               f"</totalDeclarado>")
    out.append(f"        <atribuibleHijos>{atr:.2f}</atribuibleHijos>")
    out.append(f"        <sociedadConyugal>{soc:.2f}</sociedadConyugal>")
    out.append(f"        <individual>{ind:.2f}</individual>")
    out.append(f"        <anioAnterior>{anio_anterior:.2f}</anioAnterior>")
    if diff >= 0:
        out.append(f"        <crecimientoPat>{diff:.2f}</crecimientoPat>")
    else:
        out.append(f"        <decrecimientoPat>{abs(diff):.2f}"
                    f"</decrecimientoPat>")
    out.append("    <justificacion>")
    # La justificacion solo aplica cuando hay CRECIMIENTO patrimonial (SRI).
    if diff > 0.005:
        for cod in _justificacion(wb):
            out.append("<detalleJustificacion>")
            out.append(f"    <justificVariacion>{_esc(cod)}"
                        "</justificVariacion>")
            out.append("</detalleJustificacion>")
    out.append("</justificacion></patrimonio>")

    for key in ORDEN_XML:
        modulo = MOD[key]
        registros = [(d, v) for d, v in filas[key] if v > 0]
        if not registros:
            continue
        out.append(f"<{modulo['contenedor']}>")
        for desc, valor in registros:
            out.append(_detalle_xml(modulo, desc, valor))
        out.append(f"</{modulo['contenedor']}>")

    out.append("</decPat>")
    return "\n".join(out)


def main(argv):
    args = [a for a in argv[1:] if not a.startswith("--")]
    opts = {a.split("=")[0]: a.split("=")[1]
            for a in argv[1:] if a.startswith("--") and "=" in a}
    if not args:
        print(__doc__)
        return 1
    libro = Path(args[0])
    if not libro.exists():
        print(f"No existe el libro: {libro}")
        return 1

    wb = load_workbook(libro, data_only=True, read_only=True)
    cols = _columnas_valor(wb["Dinero"])
    wb.close()
    anios = sorted(int(a) for a in cols if a.isdigit())
    anio = int(opts.get("--anio", anios[-1] if anios else 2025))

    salida = Path(args[1]) if len(args) > 1 else libro.with_name(
        f"DeclaracionPatrimonial_{anio}.xml")
    xml = generar_xml(libro, anio)
    salida.write_bytes(xml.encode(ENCODING_XML, errors="replace"))
    print(f"XML generado: {salida}  (anio {anio})")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
