"""Generador de Excel ejecutivo a partir de una Declaracion Patrimonial (XML del SRI).

Lee un archivo XML de Declaracion Patrimonial y produce un libro .xlsx con
formato premium que incluye:
  - Una hoja por cada modulo de activo del formulario del SRI (8 modulos) mas
    el modulo de pasivos. Los modulos que no constan en el XML se generan como
    plantilla en blanco, lista para llenar.
  - Columnas editables con variacion calculada por formula.
  - Listas desplegables con los catalogos oficiales del SRI.
  - Una hoja "Dashboard" ejecutiva con tarjetas KPI y el resumen separado en
    bloques de ACTIVOS, PASIVOS y PATRIMONIO NETO, con graficos.
  - Una hoja "Justificacion" con dashboard de la variacion patrimonial.

Las etiquetas XML de los 9 modulos estan tomadas de XML reales del SRI.

Uso:
    python tools/declaracion_patrimonial_excel.py ENTRADA.xml [SALIDA.xlsx]
"""

from __future__ import annotations

import sys
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent))
import sri_catalogos as cat  # noqa: E402

# --- Paleta ejecutiva --------------------------------------------------------

NAVY = "10243E"
NAVY_SOFT = "1F3B5C"
GOLD = "C8A24B"
INK = "1A1A1A"
GRIS_TXT = "7A7A7A"

ACT = "1E6B52"
ACT_L = "E6F0EB"
PAS = "A23B3B"
PAS_L = "F4E6E6"
PAT = "20507D"
PAT_L = "E2EAF2"

EDIT = "FCEFC9"
GRIS = "F3F4F6"

FONT_NORMAL = Font(name="Calibri", size=11, color=INK)
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=INK)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

_THIN = Side(style="thin", color="D0D3D8")
BORDE = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

MONEDA = '"$"#,##0.00'
PORCENT = '+0.0%;-0.0%;0.0%'

N_BLANK = 8  # filas en blanco para modulos que no constan en el XML

CATALOGOS = {
    "DINERO_EN": cat.DINERO_EN,
    "TIPO_INVERSION": cat.TIPO_INVERSION,
    "UBICACION": cat.UBICACION,
    "PAIS": cat.PAIS,
    "INSTITUCION_FINANCIERA": cat.INSTITUCION_FINANCIERA,
    "PARTES_RELACIONADAS": cat.PARTES_RELACIONADAS,
    "TIPO_DRC": cat.TIPO_DRC,
    "TIPO_PERSONA": cat.TIPO_PERSONA,
    "TIPO_BIEN_MUEBLE": cat.TIPO_BIEN_MUEBLE,
    "TIPO_VEHICULO": cat.TIPO_VEHICULO,
    "TIPO_DERECHO": cat.TIPO_DERECHO,
    "TIPO_INMUEBLE": cat.TIPO_INMUEBLE,
    "PROVINCIA": cat.PROVINCIA,
    "CANTON": cat.CANTON,
    "TIPO_ACREEDOR": cat.TIPO_ACREEDOR,
    "TIPO_IDENTIFICACION": cat.TIPO_IDENTIFICACION,
    "REGULARIZACION": cat.REGULARIZACION_ACTIVOS,
    "JUSTIF_INCREMENTO": cat.JUSTIFICACION_INCREMENTO,
    "JUSTIF_DECREMENTO": cat.JUSTIFICACION_DECREMENTO,
}


def _txt(node, tag, default=""):
    if node is None:
        return default
    el = node.find(tag)
    return el.text.strip() if el is not None and el.text else default


def _num(node, tag, default=0.0):
    try:
        return float(_txt(node, tag, "0") or "0")
    except ValueError:
        return default


def _fmt(catalogo, code):
    code = (code or "").strip()
    if not code:
        return ""
    desc = catalogo.get(code)
    return f"{code} - {desc}" if desc else code


def _entries(catalogo):
    def _key(k):
        try:
            return (0, int(k))
        except ValueError:
            return (1, k)
    return [f"{k} - {v}" for k, v in sorted(catalogo.items(),
                                            key=lambda kv: _key(kv[0]))]


# --- Definicion de los modulos del formulario --------------------------------
# Cada columna: (encabezado, ancho, nombre_catalogo|None, etiqueta_xml, tipo)
#   tipo: "cod" = columna codificada (catalogo) | "txt" = texto libre
#         "val" = columna monetaria (va a las columnas Valor del Excel)
# La columna "_ifi" (institucion financiera) se resuelve de forma especial.

_REG = ("Regularizacion activos", 16, "REGULARIZACION",
        "regularizacionActivos", "cod")
_ADQ = ("Anio fiscal adquisicion", 15, None, "anioFiscalAdquisicion", "txt")

MODULOS = [
    {
        "key": "dinero", "hoja": "Dinero", "seccion": "4.1.1",
        "titulo": "DINERO E INVERSIONES EN INSTITUCIONES FINANCIERAS",
        "clase": "activo", "dash": "Dinero e inversiones",
        "contenedor": "dinero", "detalle": "detalleDinero",
        "cols": [
            ("Dinero en", 18, "DINERO_EN", "dineroEn", "cod"),
            ("Tipo de inversion", 26, "TIPO_INVERSION", "tipoInversion", "cod"),
            ("Ubicacion", 15, "UBICACION", "ubicacion", "cod"),
            ("Pais", 18, "PAIS", "pais", "cod"),
            ("Institucion financiera", 28, "INSTITUCION_FINANCIERA", "_ifi",
             "cod"),
            ("Tipo de moneda", 13, None, "tipoMoneda", "txt"),
            (None, 16, None, "saldo", "val"),
            ("Partes relacionadas", 15, "PARTES_RELACIONADAS",
             "partesRelacionadas", "cod"),
            _REG, _ADQ,
        ],
    },
    {
        "key": "drc", "hoja": "Derechos de Capital", "seccion": "4.1.2",
        "titulo": "INVERSIONES EN DERECHOS REPRESENTATIVOS DE CAPITAL",
        "clase": "activo", "dash": "Derechos representativos de capital",
        "contenedor": "inversiones", "detalle": "detalleInversiones",
        "cols": [
            ("Tipo de inversion", 30, "TIPO_DRC", "tipoInversion", "cod"),
            ("Ubicacion", 15, "UBICACION", "ubicacion", "cod"),
            (None, 16, None, "valor", "val"),
            ("Pais", 18, "PAIS", "codpais", "cod"),
            ("Empresa / Administradora / Fideicomiso", 30, None, "nombempresa",
             "txt"),
            ("% de participacion", 14, None, "porcentpart", "txt"),
            ("N. de acciones / participaciones", 16, None, "numAcc", "txt"),
            ("Partes relacionadas", 15, "PARTES_RELACIONADAS", "relacionadas",
             "cod"),
            _REG, _ADQ,
        ],
    },
    {
        "key": "cxc", "hoja": "Cuentas por Cobrar", "seccion": "4.1.3",
        "titulo": "CREDITOS, DOCUMENTOS Y CUENTAS POR COBRAR",
        "clase": "activo", "dash": "Cuentas por cobrar",
        "contenedor": "ctasXCobrar", "detalle": "detalleCtasXCobrar",
        "cols": [
            ("Nombre del deudor", 28, None, "nombreDeudor", "txt"),
            ("Tipo de deudor", 16, "TIPO_PERSONA", "tipoDeudor", "cod"),
            ("Tipo de identificacion", 18, "TIPO_IDENTIFICACION",
             "tipoIdentificacion", "cod"),
            ("N. de identificacion", 20, None, "numeroIdentificacion", "txt"),
            ("Ubicacion", 15, "UBICACION", "ubicacion", "cod"),
            ("Pais", 18, "PAIS", "pais", "cod"),
            ("Partes relacionadas", 15, "PARTES_RELACIONADAS",
             "partesRelacionadas", "cod"),
            (None, 16, None, "saldo", "val"),
            _REG, _ADQ,
        ],
    },
    {
        "key": "muebles", "hoja": "Bienes Muebles", "seccion": "4.1.4",
        "titulo": "BIENES MUEBLES",
        "clase": "activo", "dash": "Bienes muebles",
        "contenedor": "otrosBienes", "detalle": "detalleOtrosBienes",
        "cols": [
            ("Tipo de bien", 36, "TIPO_BIEN_MUEBLE", "tipoBien", "cod"),
            ("Ubicacion", 15, "UBICACION", "ubicacion", "cod"),
            ("Pais", 18, "PAIS", "pais", "cod"),
            (None, 16, None, "valor", "val"),
            _REG, _ADQ,
        ],
    },
    {
        "key": "vehiculos", "hoja": "Vehiculos", "seccion": "4.1.5",
        "titulo": "VEHICULOS MOTORIZADOS TERRESTRES, NAVES Y AERONAVES",
        "clase": "activo", "dash": "Vehiculos, naves y aeronaves",
        "contenedor": "vehiculos", "detalle": "detalleVehiculos",
        "cols": [
            ("Tipo de vehiculo", 26, "TIPO_VEHICULO", "tipoVehiculo", "cod"),
            ("Registro / placa / chasis", 20, None, "placa", "txt"),
            ("Ubicacion", 15, "UBICACION", "ubicacion", "cod"),
            ("Pais", 18, "PAIS", "pais", "cod"),
            (None, 16, None, "valor", "val"),
            _REG, _ADQ,
        ],
    },
    {
        "key": "derechos", "hoja": "Derechos", "seccion": "4.1.6",
        "titulo": "DERECHOS (USUFRUCTO, USO, HABITACION, MARCAS, ETC.)",
        "clase": "activo", "dash": "Derechos (usufructo, uso, etc.)",
        "contenedor": "derechos", "detalle": "detalleDerechos",
        "cols": [
            ("Tipo de derecho", 28, "TIPO_DERECHO", "tipoDerecho", "cod"),
            ("Ubicacion", 15, "UBICACION", "ubicacion", "cod"),
            ("Pais", 18, "PAIS", "pais", "cod"),
            (None, 16, None, "valor", "val"),
            _REG, _ADQ,
        ],
    },
    {
        "key": "inmuebles", "hoja": "Bienes Inmuebles", "seccion": "4.1.7",
        "titulo": "BIENES INMUEBLES",
        "clase": "activo", "dash": "Bienes inmuebles",
        "contenedor": "bienesInmuebles", "detalle": "detalleBienesInmuebles",
        "cols": [
            ("Tipo de inmueble", 20, "TIPO_INMUEBLE", "tipoInmueble", "cod"),
            ("Ubicacion", 15, "UBICACION", "ubicacion", "cod"),
            ("Pais", 18, "PAIS", "codPais", "cod"),
            ("Provincia", 22, "PROVINCIA", "provincia", "cod"),
            ("Canton", 22, "CANTON", "canton", "cod"),
            ("Fecha de inscripcion", 16, None, "fechaInscripcion", "txt"),
            ("Clave catastral", 22, None, "claveCat", "txt"),
            (None, 16, None, "valor", "val"),
            _REG, _ADQ,
        ],
    },
    {
        "key": "otros", "hoja": "Otros Activos", "seccion": "4.1.8",
        "titulo": "OTROS ACTIVOS (CREDITOS TRIBUTARIOS Y OTROS)",
        "clase": "activo", "dash": "Otros activos",
        "contenedor": "otrosActivos", "detalle": "detalleOtrosActivos",
        "cols": [
            ("Ubicacion", 16, "UBICACION", "ubicacion", "cod"),
            ("Pais", 20, "PAIS", "pais", "cod"),
            ("Descripcion", 32, None, "descripcion", "txt"),
            (None, 16, None, "valorTotal", "val"),
            _REG, _ADQ,
        ],
    },
    {
        "key": "pasivo", "hoja": "Pasivos", "seccion": "4.2.1",
        "titulo": "DEUDAS CONTRAIDAS / PASIVOS",
        "clase": "pasivo", "dash": "Deudas y obligaciones",
        "contenedor": "pasivo", "detalle": "detallePasivo",
        "cols": [
            ("Tipo de acreedor", 22, "TIPO_ACREEDOR", "tipoAcreedor", "cod"),
            ("Domicilio del acreedor", 18, "UBICACION", "domicilioAcreedor",
             "cod"),
            (None, 16, None, "valorDeuda", "val"),
            ("Pais", 18, "PAIS", "paisAcreedor", "cod"),
            ("Nombre del acreedor", 26, None, "nombreAcreedor", "txt"),
            ("Tipo de identificacion", 18, "TIPO_IDENTIFICACION",
             "tipoIdentificacionAcreedor", "cod"),
            ("N. de identificacion", 20, None, "numeroIdentificacionAcreedor",
             "txt"),
            ("N. registro Banco Central", 18, None,
             "numeroRegistroBancoCentral", "txt"),
            ("Partes relacionadas", 15, "PARTES_RELACIONADAS",
             "partesRelacionadas", "cod"),
        ],
    },
]

# Orden en que los contenedores aparecen en el XML del SRI
ORDEN_XML = ["dinero", "drc", "cxc", "vehiculos", "muebles", "inmuebles",
             "pasivo", "otros", "derechos"]


def desc_cols(modulo):
    """Columnas descriptivas (todas menos la monetaria)."""
    return [c for c in modulo["cols"] if c[4] != "val"]


def valor_tag(modulo):
    return next(c[3] for c in modulo["cols"] if c[4] == "val")


def _parse_modulo(root, modulo):
    """Lee las filas de detalle de un modulo desde el XML."""
    rows = []
    vtag = valor_tag(modulo)
    for d in root.iter(modulo["detalle"]):
        desc = []
        for _h, _w, catname, xml_tag, tipo in modulo["cols"]:
            if tipo == "val":
                continue
            if xml_tag == "_ifi":
                ext = _txt(d, "nombreIfiExterior")
                ec = _txt(d, "ifiEcuador")
                desc.append(ext or (_fmt(cat.INSTITUCION_FINANCIERA, ec)
                                    if ec in cat.INSTITUCION_FINANCIERA
                                    else ""))
            elif tipo == "cod" and catname:
                desc.append(_fmt(CATALOGOS[catname], _txt(d, xml_tag)))
            else:
                desc.append(_txt(d, xml_tag))
        rows.append((desc, _num(d, vtag)))
    return rows


def parse_xml(path: Path) -> dict:
    root = ET.parse(path).getroot()
    anio = int(_txt(root, "anio", "0") or 0)

    encabezado = {
        "anio": anio,
        "tipoDec": _fmt(cat.TIPO_DECLARACION, _txt(root, "tipoDec")),
        "tipoIdent": _fmt(cat.TIPO_IDENTIFICACION, _txt(root, "tipoIdent")),
        "numIdent": _txt(root, "numIdent"),
        "nombre": _txt(root, "nombre"),
        "tipoIdentCony": _fmt(cat.TIPO_IDENTIFICACION,
                              _txt(root, "tipoIdentCony")),
        "numIdentCony": _txt(root, "numIdentCony"),
        "nombreCony": _txt(root, "nombreCony"),
        "regularizacion": _fmt(cat.REGULARIZACION_ACTIVOS,
                               _txt(root, "regularizacionActivos")),
        "totalCreditos": _num(root, "totalCreditos"),
        "totalDerechos": (_num(root, "totalDerechos")
                          if root.find("totalDerechos") is not None else None),
    }

    pat = root.find("patrimonio")
    patrimonio = {
        "totalDeclarado": _num(pat, "totalDeclarado"),
        "atribuibleHijos": _num(pat, "atribuibleHijos"),
        "sociedadConyugal": _num(pat, "sociedadConyugal"),
        "individual": _num(pat, "individual"),
        "anioAnterior": _num(pat, "anioAnterior"),
        "crecimientoPat": _num(pat, "crecimientoPat"),
    }
    es_incremento = patrimonio["crecimientoPat"] >= 0
    # La justificacion de la variacion patrimonial usa siempre la Tabla 14.
    justif = [(_txt(d, "justificVariacion"),
               _fmt(cat.JUSTIFICACION_INCREMENTO, _txt(d, "justificVariacion")))
              for d in root.iter("detalleJustificacion")]

    modulos = {m["key"]: _parse_modulo(root, m) for m in MODULOS}

    return {
        "encabezado": encabezado,
        "patrimonio": patrimonio,
        "justificacion": justif,
        "es_incremento": es_incremento,
        "modulos": modulos,
    }


# --- Helpers de formato ------------------------------------------------------

def _banner(ws, ncols, titulo, subtitulo=""):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    if subtitulo:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        s = ws.cell(row=2, column=1, value=subtitulo)
        s.font = Font(size=10, italic=True, color=GRIS_TXT)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    ws.cell(row=3, column=1).fill = PatternFill("solid", fgColor=GOLD)
    ws.row_dimensions[3].height = 5


def _build_catalogos(wb):
    """Hoja oculta con los catalogos del SRI; devuelve los rangos por nombre."""
    ws = wb.create_sheet("Catalogos")
    ws.sheet_state = "hidden"
    rangos = {}
    for ci, (nombre, dic) in enumerate(CATALOGOS.items(), start=1):
        L = get_column_letter(ci)
        ws.cell(row=1, column=ci, value=nombre).font = FONT_BOLD
        entradas = _entries(dic)
        for ri, e in enumerate(entradas, start=2):
            ws.cell(row=ri, column=ci, value=e)
        rangos[nombre] = f"Catalogos!${L}$2:${L}${1 + len(entradas)}"
    return rangos


# --- Hoja de modulo ----------------------------------------------------------

def _grupo_sheet(wb, modulo, anio, rows, rangos):
    ws = wb.create_sheet(modulo["hoja"])
    ws.sheet_view.showGridLines = False
    cols = desc_cols(modulo)
    nd = len(cols)
    ncols = nd + 4
    plantilla = not rows
    sub = f"Declaracion {anio}  -  columna {anio + 1} editable"
    if plantilla:
        sub += "   |   PLANTILLA EN BLANCO (este modulo no consta en el XML)"
    _banner(ws, ncols, f"{modulo['seccion']}   {modulo['titulo']}", sub)

    hdr = 5
    headers = [c[0] for c in cols] + [
        f"Valor {anio}", f"Valor {anio + 1}", "Variacion", "Variacion %"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr, column=j, value=h)
        c.font = FONT_HEADER
        c.fill = PatternFill("solid", fgColor=NAVY_SOFT)
        c.alignment = CENTER
        c.border = BORDE
    ws.row_dimensions[hdr].height = 32

    col_25, col_26, col_var, col_pct = nd + 1, nd + 2, nd + 3, nd + 4
    L25, L26 = get_column_letter(col_25), get_column_letter(col_26)

    accent = ACT if modulo["clase"] == "activo" else PAS

    filas = rows if rows else [([""] * nd, 0.0) for _ in range(N_BLANK)]
    first = hdr + 1
    r = first
    for idx, (descs, valor) in enumerate(filas):
        zebra = PatternFill("solid", fgColor=GRIS) if idx % 2 else None
        for j, d in enumerate(descs, start=1):
            c = ws.cell(row=r, column=j, value=d)
            c.font = FONT_NORMAL
            c.alignment = LEFT
            c.border = BORDE
            if zebra:
                c.fill = zebra
        c25 = ws.cell(row=r, column=col_25, value=valor)
        c26 = ws.cell(row=r, column=col_26, value=valor)
        cv = ws.cell(row=r, column=col_var, value=f"={L26}{r}-{L25}{r}")
        cp = ws.cell(row=r, column=col_pct,
                     value=f'=IF({L25}{r}=0,"",({L26}{r}-{L25}{r})/{L25}{r})')
        for c, fmt in ((c25, MONEDA), (c26, MONEDA), (cv, MONEDA), (cp, PORCENT)):
            c.number_format = fmt
            c.font = FONT_NORMAL
            c.border = BORDE
            c.alignment = RIGHT
            if zebra:
                c.fill = zebra
        if plantilla:
            c25.fill = PatternFill("solid", fgColor=EDIT)
        c26.fill = PatternFill("solid", fgColor=EDIT)
        r += 1
    last = r - 1

    for j in range(1, nd + 1):
        cc = ws.cell(row=r, column=j, value="TOTAL" if j == 1 else None)
        cc.font = FONT_HEADER
        cc.fill = PatternFill("solid", fgColor=accent)
        cc.border = BORDE
        cc.alignment = LEFT
    t25 = ws.cell(row=r, column=col_25, value=f"=SUM({L25}{first}:{L25}{last})")
    t26 = ws.cell(row=r, column=col_26, value=f"=SUM({L26}{first}:{L26}{last})")
    tv = ws.cell(row=r, column=col_var, value=f"={L26}{r}-{L25}{r}")
    tp = ws.cell(row=r, column=col_pct,
                 value=f'=IF({L25}{r}=0,"",({L26}{r}-{L25}{r})/{L25}{r})')
    for c, fmt in ((t25, MONEDA), (t26, MONEDA), (tv, MONEDA), (tp, PORCENT)):
        c.number_format = fmt
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=accent)
        c.border = BORDE
        c.alignment = RIGHT
    ws.row_dimensions[r].height = 22
    total_row = r

    for j, (_, _, catname, _, _) in enumerate(cols, start=1):
        if not catname:
            continue
        dv = DataValidation(type="list", formula1=rangos[catname],
                            allow_blank=True, showErrorMessage=False)
        ws.add_data_validation(dv)
        L = get_column_letter(j)
        dv.add(f"{L}{first}:{L}{last}")

    for j, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = c[1]
    for j in range(nd + 1, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16

    ws.freeze_panes = ws.cell(row=first, column=1)
    return {"hoja": modulo["hoja"], "clase": modulo["clase"],
            "dash": modulo["dash"], "total_row": total_row,
            "L25": L25, "L26": L26}


# --- Dashboard ---------------------------------------------------------------

def _kpi_card(ws, r, c1, c2, label, value, accent, light,
              sub_label="", sub_value=None, sub_fmt=PORCENT):
    L1, L2 = get_column_letter(c1), get_column_letter(c2)
    ws.merge_cells(f"{L1}{r}:{L2}{r}")
    ws.cell(row=r, column=c1).fill = PatternFill("solid", fgColor=accent)
    ws.row_dimensions[r].height = 6

    ws.merge_cells(f"{L1}{r + 1}:{L2}{r + 1}")
    lc = ws.cell(row=r + 1, column=c1, value=label)
    lc.font = Font(bold=True, size=10, color=accent)
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells(f"{L1}{r + 2}:{L2}{r + 2}")
    vc = ws.cell(row=r + 2, column=c1, value=value)
    vc.font = Font(bold=True, size=17, color=INK)
    vc.number_format = MONEDA
    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r + 2].height = 28

    cap = ws.cell(row=r + 3, column=c1, value=sub_label)
    cap.font = Font(size=8, color=GRIS_TXT)
    cap.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if sub_value is not None:
        pc = ws.cell(row=r + 3, column=c2, value=sub_value)
        pc.font = Font(bold=True, size=10, color=accent)
        pc.number_format = sub_fmt
        pc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for rr in range(r + 1, r + 4):
        for c in (c1, c2):
            ws.cell(row=rr, column=c).fill = PatternFill("solid", fgColor=light)
    for rr in range(r, r + 4):
        for c in (c1, c2):
            ws.cell(row=rr, column=c).border = BORDE


def _sec_header(ws, row, texto, accent):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    c = ws.cell(row=row, column=2, value=texto)
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=accent)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def _col_headers(ws, row, anio):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    textos = ["Concepto", None, f"Total {anio}", f"Total {anio + 1}",
              "Variacion", "Variacion %"]
    for j, t in enumerate(textos, start=2):
        c = ws.cell(row=row, column=j, value=t)
        c.font = FONT_HEADER
        c.fill = PatternFill("solid", fgColor=NAVY_SOFT)
        c.alignment = CENTER if j > 3 else Alignment(
            horizontal="left", vertical="center", indent=1)
        c.border = BORDE
    ws.row_dimensions[row].height = 24


def _data_row(ws, row, concepto, v2025, v2026, light, *, bold=False,
              total=False, accent=None):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    cc = ws.cell(row=row, column=2, value=concepto)
    txtcolor = "FFFFFF" if total else INK
    fill = None
    if total:
        fill = PatternFill("solid", fgColor=accent)
    elif light:
        fill = PatternFill("solid", fgColor=light)
    cc.font = Font(bold=bold or total, size=11, color=txtcolor)
    cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    c25 = ws.cell(row=row, column=4, value=v2025)
    c26 = ws.cell(row=row, column=5, value=v2026)
    cv = ws.cell(row=row, column=6, value=f"=E{row}-D{row}")
    cp = ws.cell(row=row, column=7,
                 value=f'=IF(D{row}=0,"",(E{row}-D{row})/D{row})')
    for c, fmt in ((c25, MONEDA), (c26, MONEDA), (cv, MONEDA), (cp, PORCENT)):
        c.number_format = fmt
        c.font = Font(bold=bold or total, size=11, color=txtcolor)
        c.alignment = RIGHT
    for c in (cc, c25, c26, cv, cp):
        c.border = BORDE
        if fill:
            c.fill = fill
    ws.row_dimensions[row].height = 22 if total else 20


def _build_dashboard(wb, data, infos, anio):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    enc, pat = data["encabezado"], data["patrimonio"]
    _banner(ws, 8, "DECLARACION PATRIMONIAL  -  TABLERO EJECUTIVO",
            f"{enc['nombre']}   |   {enc['tipoIdent']}: {enc['numIdent']}   "
            f"|   {enc['tipoDec']}   |   Comparativo {anio} vs {anio + 1}")

    activos = [i for i in infos if i["clase"] == "activo"]
    pasivos = [i for i in infos if i["clase"] == "pasivo"]

    def link(info, L):
        return f"='{info['hoja']}'!{info[L]}{info['total_row']}"

    r = 10
    _sec_header(ws, r, "ACTIVOS", ACT)
    r += 1
    _col_headers(ws, r, anio)
    r += 1
    act_first = r
    for info in activos:
        _data_row(ws, r, info["dash"], link(info, "L25"), link(info, "L26"),
                  ACT_L)
        r += 1
    act_last = r - 1
    _data_row(ws, r, "TOTAL ACTIVOS", f"=SUM(D{act_first}:D{act_last})",
              f"=SUM(E{act_first}:E{act_last})", None, total=True, accent=ACT)
    tot_act = r
    r += 2

    _sec_header(ws, r, "PASIVOS", PAS)
    r += 1
    _col_headers(ws, r, anio)
    r += 1
    pas_first = r
    for info in pasivos:
        _data_row(ws, r, info["dash"], link(info, "L25"), link(info, "L26"),
                  PAS_L)
        r += 1
    pas_last = r - 1
    _data_row(ws, r, "TOTAL PASIVOS", f"=SUM(D{pas_first}:D{pas_last})",
              f"=SUM(E{pas_first}:E{pas_last})", None, total=True, accent=PAS)
    tot_pas = r
    r += 2

    _sec_header(ws, r, "PATRIMONIO NETO", PAT)
    r += 1
    _col_headers(ws, r, anio)
    r += 1
    _data_row(ws, r, "(+) Total activos", f"=D{tot_act}", f"=E{tot_act}", PAT_L)
    ra = r
    r += 1
    _data_row(ws, r, "(-) Total pasivos", f"=D{tot_pas}", f"=E{tot_pas}", PAT_L)
    rp = r
    r += 1
    _data_row(ws, r, "(=) PATRIMONIO NETO", f"=D{ra}-D{rp}", f"=E{ra}-E{rp}",
              None, total=True, accent=PAT)
    pneto = r
    r += 1

    for txt, val in [
        (f"Patrimonio declarado en el XML ({anio})", pat["totalDeclarado"]),
        ("Patrimonio neto del anio anterior", pat["anioAnterior"]),
        ("Crecimiento / decremento patrimonial (XML)", pat["crecimientoPat"]),
    ]:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        cc = ws.cell(row=r, column=2, value=txt)
        cc.font = Font(size=9, italic=True, color=GRIS_TXT)
        cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cv = ws.cell(row=r, column=4, value=val)
        cv.number_format = MONEDA
        cv.font = Font(size=9, italic=True, color=GRIS_TXT)
        cv.alignment = RIGHT
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = BORDE
        ws.row_dimensions[r].height = 16
        r += 1
    r += 1

    _sec_header(ws, r, "CONSOLIDADO", NAVY_SOFT)
    r += 1
    cons_hdr = r
    _col_headers(ws, r, anio)
    r += 1
    cons_first = r
    _data_row(ws, r, "Activos", f"=D{tot_act}", f"=E{tot_act}", ACT_L, bold=True)
    r += 1
    _data_row(ws, r, "Pasivos", f"=D{tot_pas}", f"=E{tot_pas}", PAS_L, bold=True)
    r += 1
    _data_row(ws, r, "Patrimonio neto", f"=D{pneto}", f"=E{pneto}", PAT_L,
              bold=True)
    cons_last = r

    _kpi_card(ws, 5, 2, 3, "TOTAL ACTIVOS", f"=D{tot_act}", ACT, ACT_L,
              "Var. proyectada", f"=G{tot_act}")
    _kpi_card(ws, 5, 4, 5, "TOTAL PASIVOS", f"=D{tot_pas}", PAS, PAS_L,
              "Var. proyectada", f"=G{tot_pas}")
    _kpi_card(ws, 5, 6, 7, "PATRIMONIO NETO", f"=D{pneto}", PAT, PAT_L,
              "Var. proyectada", f"=G{pneto}")

    # --- grafico de barras: concentracion de los 8 tipos de activo ---
    cats_act = f"'Dashboard'!$B${act_first}:$B${act_last}"
    barr = BarChart()
    barr.type = "col"
    barr.title = f"Concentracion de activos {anio} (por tipo)"
    barr.style = 10
    barr.add_data(Reference(ws, min_col=4, min_row=act_first - 1,
                            max_row=act_last), titles_from_data=True)
    for s in barr.series:
        s.cat = AxDataSource(strRef=StrRef(f=cats_act))
    barr.dataLabels = DataLabelList()
    barr.dataLabels.showVal = True
    barr.legend = None
    barr.height, barr.width = 10, 21
    ws.add_chart(barr, "I3")

    # --- grafico circular: composicion activos / pasivos / patrimonio ---
    cats_cons = f"'Dashboard'!$B${cons_first}:$B${cons_last}"
    comp = PieChart()
    comp.title = "Composicion: activos, pasivos y patrimonio"
    comp.add_data(Reference(ws, min_col=4, min_row=cons_first,
                            max_row=cons_last), titles_from_data=False)
    for s in comp.series:
        s.cat = AxDataSource(strRef=StrRef(f=cats_cons))
    comp.dataLabels = DataLabelList()
    comp.dataLabels.showCatName = True
    comp.dataLabels.showPercent = True
    comp.dataLabels.showVal = False
    comp.dataLabels.showSerName = False
    comp.dataLabels.showLegendKey = False
    comp.height, comp.width = 10, 15
    ws.add_chart(comp, "I24")

    nota = ws.cell(row=cons_last + 2, column=2,
                   value="Edite las columnas amarillas en las hojas de cada "
                         "modulo; las tarjetas, secciones y graficos se "
                         "recalculan solos.")
    nota.font = Font(italic=True, size=9, color=GRIS_TXT)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    for col in ("D", "E", "F", "G"):
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["H"].width = 3
    return {"pneto_row": pneto}


def _build_justificacion(wb, data, anio):
    """Hoja Justificacion: lista de seleccion igual al formulario del SRI."""
    ws = wb.create_sheet("Justificacion")
    ws.sheet_view.showGridLines = False
    pat = data["patrimonio"]
    crec = pat["crecimientoPat"]
    es_inc = data["es_incremento"]
    palabra = "Incremento" if es_inc else "Decremento"
    presentes = {code for code, _lbl in data["justificacion"]}

    _banner(ws, 5, "JUSTIFICACION DE LA VARIACION PATRIMONIAL",
            "Marque 'Si' los conceptos que justifican la variacion "
            f"patrimonial  -  ejercicio {anio}")

    # linea informativa con el crecimiento / decremento
    ws.merge_cells("B5:D5")
    ci = ws.cell(row=5, column=2,
                 value=f"{palabra} patrimonial declarado en el XML")
    ci.font = FONT_BOLD
    ci.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cv = ws.cell(row=5, column=5, value=crec)
    cv.font = FONT_BOLD
    cv.number_format = MONEDA
    cv.alignment = RIGHT
    for c in range(2, 6):
        ws.cell(row=5, column=c).fill = PatternFill("solid", fgColor=GRIS)
        ws.cell(row=5, column=c).border = BORDE
    ws.row_dimensions[5].height = 20

    # cabecera de la tabla
    hdr = 7
    ws.merge_cells(start_row=hdr, start_column=2, end_row=hdr, end_column=4)
    for col, txt in ((2, "Concepto (Tabla 14 del SRI)"),
                     (5, "Justifica la variacion")):
        c = ws.cell(row=hdr, column=col, value=txt)
        c.font = FONT_HEADER
        c.fill = PatternFill("solid", fgColor=NAVY_SOFT)
        c.alignment = CENTER
        c.border = BORDE
    ws.row_dimensions[hdr].height = 26

    first = hdr + 1
    r = first
    for code in ("1", "2", "3", "4", "5", "6"):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cc = ws.cell(row=r, column=2,
                     value=_fmt(cat.JUSTIFICACION_INCREMENTO, code))
        cc.font = FONT_NORMAL
        cc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cm = ws.cell(row=r, column=5,
                     value="Si" if code in presentes else "No")
        cm.font = FONT_BOLD
        cm.alignment = CENTER
        cm.fill = PatternFill("solid", fgColor=EDIT)
        for col in range(2, 6):
            ws.cell(row=r, column=col).border = BORDE
        ws.row_dimensions[r].height = 22
        r += 1
    last = r - 1

    dv = DataValidation(type="list", formula1='"Si,No"', allow_blank=False,
                        showErrorMessage=False)
    ws.add_data_validation(dv)
    dv.add(f"E{first}:E{last}")

    nota = ws.cell(
        row=last + 2, column=2,
        value="Esta hoja reproduce la seccion Justificacion del formulario "
              "del SRI. Marque 'Si' en los conceptos que apliquen; el "
              "generador de XML los toma de aqui.")
    nota.font = Font(italic=True, size=9, color=GRIS_TXT)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 22


def _build_info(wb, data, anio, pneto_row):
    ws = wb.create_sheet("Datos del XML")
    ws.sheet_view.showGridLines = False
    _banner(ws, 2, "DATOS GENERALES DE LA DECLARACION")
    enc, pat = data["encabezado"], data["patrimonio"]

    identificacion = [
        ("Anio de la declaracion", enc["anio"], None),
        ("Tipo de declaracion", enc["tipoDec"], None),
        ("Tipo de identificacion", enc["tipoIdent"], None),
        ("Numero de identificacion", enc["numIdent"], None),
        ("Nombre del declarante", enc["nombre"], None),
    ]
    if enc["numIdentCony"] or enc["nombreCony"]:
        identificacion += [
            ("Identificacion del conyuge", enc["tipoIdentCony"], None),
            ("Numero ident. conyuge", enc["numIdentCony"], None),
            ("Nombre del conyuge", enc["nombreCony"], None),
        ]
    if enc["regularizacion"]:
        identificacion.append(
            ("Regularizacion de activos", enc["regularizacion"], None))
    identificacion.append(("Total creditos / cuentas por cobrar",
                           enc["totalCreditos"], MONEDA))
    if enc["totalDerechos"] is not None:
        identificacion.append(("Total derechos", enc["totalDerechos"], MONEDA))

    secciones = [
        ("IDENTIFICACION", identificacion),
        ("PATRIMONIO NETO (enlazado a la hoja Dashboard)", [
            (f"Patrimonio neto {anio}", f"=Dashboard!D{pneto_row}", MONEDA),
            (f"Patrimonio neto {anio + 1} (proyectado)",
             f"=Dashboard!E{pneto_row}", MONEDA),
            ("Variacion proyectada", f"=Dashboard!F{pneto_row}", MONEDA),
            ("Variacion % proyectada", f"=Dashboard!G{pneto_row}", PORCENT),
        ]),
        ("PATRIMONIO DECLARADO EN EL XML", [
            ("Patrimonio total declarado", pat["totalDeclarado"], MONEDA),
            ("Atribuible a hijos no emancipados", pat["atribuibleHijos"],
             MONEDA),
            ("Patrimonio en la sociedad conyugal", pat["sociedadConyugal"],
             MONEDA),
            ("Patrimonio individual del declarante", pat["individual"], MONEDA),
            ("Patrimonio del anio anterior", pat["anioAnterior"], MONEDA),
            ("Crecimiento / decremento patrimonial", pat["crecimientoPat"],
             MONEDA),
        ]),
    ]
    r = 5
    for titulo, filas in secciones:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        h = ws.cell(row=r, column=1, value=titulo)
        h.font = Font(bold=True, size=11, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor=NAVY_SOFT)
        h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20
        r += 1
        for nombre, valor, fmt in filas:
            cc = ws.cell(row=r, column=1, value=nombre)
            cc.font = FONT_BOLD
            cc.alignment = LEFT
            c = ws.cell(row=r, column=2, value=valor)
            c.font = FONT_NORMAL
            c.alignment = RIGHT
            if fmt:
                c.number_format = fmt
            for col in (1, 2):
                ws.cell(row=r, column=col).border = BORDE
            r += 1
        r += 1

    h = ws.cell(row=r, column=1,
                value="Justificacion de la variacion (ver hoja Justificacion)")
    h.font = Font(bold=True, size=11, color=NAVY)
    r += 1
    for _, etiqueta in data["justificacion"]:
        ws.cell(row=r, column=1, value="- " + etiqueta).font = FONT_NORMAL
        r += 1
    r += 1
    ws.cell(row=r, column=1,
            value="Catalogos del SRI integrados (Tablas 1-19 del archivo "
                  "CATALOGO.xls). El XML se genera con generar_xml_sri.py."
            ).font = Font(italic=True, size=9, color=GRIS_TXT)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 34


def _build_mapa(wb):
    """Hoja oculta con el mapa de etiquetas XML; la usa la macro VBA."""
    ws = wb.create_sheet("_Mapa")
    ws.sheet_state = "hidden"
    ws.append(["key", "hoja", "contenedor", "detalle", "clase", "ordenXml",
               "tipo", "xmlTag", "descIndex"])
    orden = {k: i for i, k in enumerate(ORDEN_XML)}
    for m in MODULOS:
        di = 0
        for _h, _w, _cat, xml_tag, tipo in m["cols"]:
            if tipo == "val":
                idx = 0
            else:
                di += 1
                idx = di
            ws.append([m["key"], m["hoja"], m["contenedor"], m["detalle"],
                       m["clase"], orden[m["key"]], tipo, xml_tag, idx])


def _build_instrucciones(wb):
    """Hoja visible con el boton para generar el XML y su configuracion."""
    ws = wb.create_sheet("Generar XML")
    ws.sheet_view.showGridLines = False
    _banner(ws, 6, "GENERAR XML PARA EL SRI",
            "Convierte este libro en el archivo XML de la Declaracion "
            "Patrimonial, listo para subir al portal del SRI")

    ws.merge_cells("B5:D6")
    b = ws.cell(row=5, column=2, value="►   GENERAR XML")
    b.font = Font(bold=True, size=14, color="FFFFFF")
    b.fill = PatternFill("solid", fgColor=ACT)
    b.alignment = CENTER
    for r in (5, 6):
        for c in range(2, 5):
            ws.cell(row=r, column=c).border = BORDE
    ws.cell(row=7, column=2,
            value="En el archivo .xlsm el boton verde ya esta vinculado a la "
                  "macro GenerarXmlSRI.").font = Font(italic=True, size=8,
                                                      color=GRIS_TXT)

    pasos = [
        ("COMO USAR ESTE ARCHIVO", True),
        ("1. Al abrir el archivo .xlsm, pulse 'Habilitar contenido' en la "
         "barra de seguridad de Excel.", False),
        ("2. Llene las columnas amarillas de cada modulo (Dinero, Vehiculos, "
         "etc.) y la hoja Justificacion.", False),
        ("3. Pulse el boton verde GENERAR XML, indique el anio y elija donde "
         "guardar el archivo.", False),
        ("4. El archivo .xml queda listo para subir al portal del SRI.",
         False),
        ("", False),
        ("SI EXCEL BLOQUEA LAS MACROS", True),
        ("- Cierre Excel. En el Explorador, clic derecho sobre el archivo > "
         "Propiedades.", False),
        ("- Al final de la pestana General marque 'Desbloquear' y pulse "
         "Aceptar.", False),
        ("- Vuelva a abrirlo y pulse 'Habilitar contenido'. Conserve siempre "
         "la extension .xlsm.", False),
        ("", False),
        ("Alternativa sin macros:  ejecutar  "
         "python tools/generar_xml_sri.py  desde la terminal.", False),
    ]
    r = 9
    for texto, encab in pasos:
        c = ws.cell(row=r, column=2, value=texto)
        if encab:
            c.font = Font(bold=True, size=11, color=NAVY)
        else:
            c.font = FONT_NORMAL
        r += 1

    ws.column_dimensions["A"].width = 3
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 22


def build_workbook(data: dict, salida: Path):
    anio = data["encabezado"]["anio"] or 2025
    wb = Workbook()
    wb.remove(wb.active)

    rangos = _build_catalogos(wb)

    infos = []
    for modulo in MODULOS:
        rows = data["modulos"][modulo["key"]]
        infos.append(_grupo_sheet(wb, modulo, anio, rows, rangos))

    dash = _build_dashboard(wb, data, infos, anio)
    _build_justificacion(wb, data, anio)
    _build_instrucciones(wb)
    _build_info(wb, data, anio, dash["pneto_row"])
    _build_mapa(wb)

    orden = ["Dashboard"] + [m["hoja"] for m in MODULOS] + \
            ["Justificacion", "Generar XML", "Datos del XML", "_Mapa",
             "Catalogos"]
    wb._sheets.sort(key=lambda ws: orden.index(ws.title))
    wb.active = 0

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)


def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 1
    entrada = Path(argv[1])
    if not entrada.exists():
        print(f"No existe el archivo: {entrada}")
        return 1
    salida = Path(argv[2]) if len(argv) > 2 else entrada.with_suffix(".xlsx")
    data = parse_xml(entrada)
    build_workbook(data, salida)
    print(f"Excel generado: {salida}")

    # Version .xlsm con la macro VBA y el boton embebidos
    try:
        from _xlsm_builder import construir_xlsm
        bas = (Path(__file__).resolve().parent / "GenerarXmlSRI.bas")
        if bas.exists():
            xlsm = salida.with_suffix(".xlsm")
            construir_xlsm(salida, xlsm, bas.read_text(encoding="utf-8"))
            print(f"Excel con macro generado: {xlsm}")
    except Exception as exc:  # noqa: BLE001
        print(f"Aviso: no se pudo generar el .xlsm ({exc})")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
